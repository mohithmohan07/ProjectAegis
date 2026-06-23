import os
import re
import json
import time
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# USER CONFIG (edit these if needed)
# ============================================================
BASE_DIR = Path(r"C:\Users\FCI\Documents\CM\data")
FOLDER_PREFIX = "ICSE_SE_G10_CH"   # matches ICSE_SE_G10_CH01, CH02, ...
OUTPUT_XLSX = Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G10_Concepts.xlsx")

BOARD = "ICSE"
BOOK = "Selina"
GRADE = 10
SUBJECT = "Mathematics"

# Override when processing a folder via CLI (e.g. mmds_ICSE_MS_G10_Geography)
MMDS_FOLDER = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_MS_G10_Geography")
MMDS_FOLDER_CONFIG = {"BOOK": "MS", "SUBJECT": "Geography", "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_MS_G10_Geography_Concepts.xlsx")}

# CBSE: run all subjects from root folder (MMDs\CBSE) with one command
MMDS_CBSE_ROOT = Path(r"C:\Users\FCI\OneDrive\Documents\CM\MMDs\CBSE")
CBSE_ALL_OUTPUT_DIR = Path(r"G:\My Drive\Concept Mapping Version 01")

# CBSE RD Mathematics: one Excel per grade (G09 and G10)
MMDS_CBSE_RD_MATHEMATICS = Path(r"C:\Users\FCI\Documents\CM\mmds_CBSE_RD_Mathematics")
# Alternative source: MMDs\CBSE\Class 10\CBSE_RD_G10_Mathematics, Class 9\CBSE_RD_G09_Mathematics
MMDS_CBSE_RD_FOLDER_NAMES = ("mmds_CBSE_RD_Mathematics", "CBSE_RD_G10_Mathematics", "CBSE_RD_G09_Mathematics")
MMDS_CBSE_RD_CONFIG = {
    "BOARD": "CBSE",
    "BOOK": "RD",
    "SUBJECT": "Mathematics",
    "OUTPUT_DIR": Path(r"G:\My Drive\Concept Mapping Version 01"),
    "SPLIT_BY_GRADE": True,
    "USE_GRADE_SUBFOLDERS": True,  # Output to Grade 09/ and Grade 10/ subfolders
}

# ICSE SE Mathematics: one Excel per grade (G09 and G10). Optional 2nd arg = 9 for G09 only, or 10 for G10 only.
MMDS_ICSE_SE_MATHEMATICS = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_Mathematics")
MMDS_ICSE_SE_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Mathematics",
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "SPLIT_BY_GRADE": True,
}

# ICSE SE G09 Biology (subfolder of Biology_Chemistry): single output file
MMDS_ICSE_SE_G09_BIOLOGY = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_Biology_Chemistry\Biology")
MMDS_ICSE_SE_G09_BIOLOGY_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Biology",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_Biology_Concepts.xlsx"),
}

# ICSE SE G09 Chemistry (subfolder of Biology_Chemistry): single output file
MMDS_ICSE_SE_G09_CHEMISTRY = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_Biology_Chemistry\Chemistry")
MMDS_ICSE_SE_G09_CHEMISTRY_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Chemistry",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_Chemistry_Concepts.xlsx"),
}

# ICSE SE G09 Geography: single output file
MMDS_ICSE_SE_G09_GEOGRAPHY = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_Geography")
MMDS_ICSE_SE_G09_GEOGRAPHY_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Geography",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_Geography_Concepts.xlsx"),
}

# ICSE SE G09 Physics: single output file
MMDS_ICSE_SE_G09_PHYSICS = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_Physics")
MMDS_ICSE_SE_G09_PHYSICS_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Physics",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_Physics_Concepts.xlsx"),
}

# ICSE SE G09 History: single output file
MMDS_ICSE_SE_G09_HISTORY = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_History")
MMDS_ICSE_SE_G09_HISTORY_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "History",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_History_Concepts.xlsx"),
}

# ICSE SE G09 Civics: single output file
MMDS_ICSE_SE_G09_CIVICS = Path(r"C:\Users\FCI\Documents\CM\mmds_ICSE_SE_G09_Civics")
MMDS_ICSE_SE_G09_CIVICS_CONFIG = {
    "BOARD": "ICSE",
    "BOOK": "Selina",
    "SUBJECT": "Civics",
    "GRADE": 9,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\Documents\CM\ICSE_SE_G09_Civics_Concepts.xlsx"),
}

# ICSE MS G10 History & Civics (MMDs\ICSE\ICSE_SE_G10_History, ICSE_SE_G10_Civics - folder names kept for path compatibility)
MMDS_ICSE_SE_G10_HISTORY_NAME = "ICSE_SE_G10_History"
MMDS_ICSE_SE_G10_CIVICS_NAME = "ICSE_SE_G10_Civics"

# Temporary: MSBSHSE Symbiosis (MMDs\CBSE\Symbiosis) - Board=MSBSHSE, Book=M-State, Grade=11, Subject from filename
MMDS_SYMBIOSIS = Path(r"C:\Users\FCI\OneDrive\Documents\CM\MMDs\CBSE\Symbiosis")
MMDS_SYMBIOSIS_CONFIG = {
    "BOARD": "MSBSHSE",
    "BOOK": "M-State",
    "GRADE": 11,
    "SUBJECT_FROM_FILENAME": True,
    "OUTPUT_DIR": Path(r"C:\Users\FCI\OneDrive\Documents\CM"),
    "OUTPUT": Path(r"C:\Users\FCI\OneDrive\Documents\CM\MSBSHSE_M-State_G11_Symbiosis_Concepts.xlsx"),
}

MODEL = "gpt-5.4-mini-2026-03-17"  # change if you want
MAX_OUTPUT_TOKENS = int(os.getenv("AEGIS_OPENAI_MAX_OUTPUT_TOKENS", "128000"))
CONCEPTS_PER_CHAPTER_MIN = 40

# Temporary: models to compare for Chapter 01 (use 2nd arg "compare")
COMPARE_MODELS = [
    "gpt-4.1-mini-2025-04-14",
    "gpt-5-mini-2025-08-07",
    "gpt-5-nano-2025-08-07",
]
CONCEPTS_PER_CHAPTER_MAX = 60

# If a .mmd is huge, we safely trim it to avoid token blowups
MAX_MMD_CHARS = 220_000

# API retry behavior
MAX_RETRIES = 3
RETRY_SLEEP_SECONDS = 2


# ============================================================
# OPENAI CLIENT
# ============================================================
client = OpenAI()  # uses OPENAI_API_KEY environment variable


# ============================================================
# STRICT JSON SCHEMA (Structured Outputs)
# ============================================================
CONCEPT_SCHEMA: Dict[str, Any] = {
    "name": "concept_rows",
    "strict": True,
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "rows": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "topic": {"type": "string"},
                        "parent_concept": {"type": "string"},
                        "concept": {"type": "string"},
                        "concept_description": {"type": "string"}
                    },
                    "required": ["topic", "parent_concept", "concept", "concept_description"]
                }
            }
        },
        "required": ["rows"]
    }
}


# ============================================================
# HELPERS
# ============================================================
def natural_sort_key(s: str):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def make_unique_output_path(output_path: Path) -> Path:
    """
    Avoid overwriting an existing output file by adding a timestamp suffix.
    Example: file.xlsx -> file_20260304_173005.xlsx
    """
    if not output_path.exists():
        return output_path
    ts = time.strftime("%Y%m%d_%H%M%S")
    return output_path.with_name(f"{output_path.stem}_{ts}{output_path.suffix}")


def infer_metadata_from_mmd_folder_name(folder_name: str) -> Optional[Dict[str, Any]]:
    """
    Infer board/book/grade/subject tokens from folder names like:
      mmds_ICSE_SE_G10_Mathematics
      mmds_CBSE_RD_G09_NUMBER_SYSTEM
    """
    m = re.match(r"^mmds_([A-Za-z]+)_([A-Za-z]+)_G(\d{2})_(.+)$", folder_name)
    if not m:
        return None
    board_code, book_code, grade_str, subject_token = m.groups()
    return {
        "board_code": board_code.upper(),
        "book_code": book_code.upper(),
        "grade": int(grade_str),
        "subject_token": subject_token,
        "subject_display": subject_token.replace("_", " "),
    }


def remove_vowels(s: str) -> str:
    """Remove vowels and collapse to alphanumeric + underscore for ChptrNm / TpcNm codes. Uppercase."""
    if not s or not isinstance(s, str):
        return ""
    s = re.sub(r"[aeiouAEIOU]", "", s)
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s.upper() if s else ""


def normalize_for_code(s: str) -> str:
    """Keep full name, replace spaces/special with underscore. E.g. 'Rational numbers – definition' -> 'Rational_numbers_definition'."""
    if not s or not isinstance(s, str):
        return ""
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def get_chapter_name_for_code(mmd_filename: str) -> str:
    """Extract chapter name stem from MMD filename for ChptrNm. E.g. CBSE_RD_G09_CH01_NUMBER_SYSTEM.mmd -> NUMBER_SYSTEM."""
    m = re.search(r"_CH\d+_(.+)\.mmd$", mmd_filename, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    stem = Path(mmd_filename).stem
    return stem


def chapter_name_stem_to_title(stem: str) -> str:
    """Convert chapter name from filename to Title Case. E.g. NUMBER_SYSTEM -> Number System."""
    if not stem or not isinstance(stem, str):
        return ""
    return stem.replace("_", " ").strip().title()


def board_code_from_board(board: str) -> str:
    """CBSE -> CB, ICSE -> IC, MSBSHSE -> MS."""
    if not board:
        return "CB"
    u = (board or "").upper()
    if "ICSE" in u:
        return "IC"
    if "MSBSHSE" in u or "MAHARASHTRA" in u:
        return "MS"
    return "CB"


def infer_subject_from_chapter_stem(stem: str) -> str:
    """Infer school subject from chapter/PDF filename stem (MSBSHSE G11 batch)."""
    s = (stem or "").lower()
    if "psychology" in s:
        return "Psychology"
    if "chemistry" in s:
        return "Chemistry"
    if "systematics" in s or "organism" in s:
        return "Biology"
    if "angle" in s or "sets" in s or "relation" in s:
        return "Mathematics"
    if "unit" in s and "measurement" in s:
        return "Physics"
    return "General"


# PDFs often live here after Mathpix extraction from Desktop (optional path for PDF column)
MSBSHSE_DEFAULT_PDF_DIR = Path(r"C:\Users\FCI\OneDrive\Desktop\My Details")


def strip_grade_suffix_from_chapter_stem(stem: str) -> str:
    """Strip trailing '(Grade 11)' / '(G11)' style suffixes from an MMD/PDF stem."""
    s = (stem or "").strip()
    s = re.sub(r"\s*\(\s*Grade\s*\d+\s*\)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(\s*G\s*\d+\s*\)\s*$", "", s, flags=re.IGNORECASE)
    return s.strip()


def parse_grade_from_chapter_stem(stem: str) -> Optional[int]:
    m = re.search(r"\(\s*Grade\s*(\d+)\s*\)", stem, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"\(\s*G\s*(\d+)\s*\)", stem, flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def ncert_plant_or_animal_physiology_parent(stem: str) -> str:
    """
    Map NCERT Biology chapter filename to required Parent Concept bucket.
    Uses the stem with or without '(Grade NN)' suffix.
    """
    s_plain = strip_grade_suffix_from_chapter_stem(stem).lower()
    s_full = (stem or "").lower()

    def hit(needle: str) -> bool:
        return needle in s_plain or needle in s_full

    plant_needles = (
        "photosynthesis in higher plants",
        "plant growth and development",
        "respiration in plants",
        "life processes in plants",
    )
    animal_needles = (
        "body fluids and circulation",
        "breathing and exchange of gases",
        "excretory products and their elimination",
        "neural control and coordination",
        "locomotion and movement",
        "chemical coordination and integration",
        "life processes in animals",
    )
    for needle in plant_needles:
        if hit(needle):
            return "Plant Physiology"
    for needle in animal_needles:
        if hit(needle):
            return "Animal Physiology"
    if "life processes" in s_plain and "plant" not in s_plain and "animal" not in s_plain:
        return "Animal Physiology"
    if "control" in s_plain and "coordination" in s_plain:
        return "Animal Physiology"
    return "Animal Physiology"


def resolve_pdf_beside_mmd(mmd_path: Path, extra_dirs: Optional[List[Path]] = None) -> Optional[Path]:
    """Find matching <stem>.pdf next to the .mmd or under known folders."""
    stem = mmd_path.stem
    candidates = [mmd_path.parent / f"{stem}.pdf"]
    if extra_dirs:
        for d in extra_dirs:
            candidates.append(d / f"{stem}.pdf")
    for p in candidates:
        if p.exists():
            return p
    for pdf in mmd_path.parent.glob("*.pdf"):
        return pdf
    return None


def format_topic_display(grade: int, board_code: str, chptr_code: str, topic_num: int, topic_name: str) -> str:
    """Topic 01: TopicName (09CBMA_ChapterName_PL). Full names, no vowel removal."""
    prefix = f"{grade:02d}{board_code}MA_{chptr_code}_PL"
    return f"Topic {topic_num:02d}: {topic_name} ({prefix})"


def format_concept_display(grade: int, board_code: str, chptr_code: str, tpc_code: str, concept_name: str) -> str:
    """ConceptName (09CBMA_ChapterName_PL_TopicName). Full names, no vowel removal."""
    prefix = f"{grade:02d}{board_code}MA_{chptr_code}_PL_{tpc_code}"
    return f"{concept_name} ({prefix})"


def find_chapter_dirs(base_dir: Path, prefix: str) -> List[Path]:
    dirs = [p for p in base_dir.iterdir() if p.is_dir() and p.name.startswith(prefix)]
    return sorted(dirs, key=lambda p: natural_sort_key(p.name))


def chapter_no_from_dirname(dirname: str) -> Optional[int]:
    m = re.search(r"_CH(\d+)$", dirname)
    return int(m.group(1)) if m else None


def extract_subject_from_msbshse_filename(filename: str) -> str:
    """
    Extract subject from MSBSHSE filename. E.g.:
    MSBSHSE_M-State_Grade11_Biology_Chapter - Systematics... -> Biology
    MSBSHSE_M-State_Grade11_Mathematics (Commerce)_Chapter - Sets.mmd -> Mathematics (Commerce)
    """
    m = re.search(r"Grade11_(.+?)_Chapter", filename)
    return m.group(1).strip() if m else "General"


def extract_chapter_info_from_filename(filename: str) -> Tuple[Optional[int], str]:
    """
    Extract chapter number and code from filename.
    Examples:
    - "icse-class-9-biology-chapter-2-cell--the-unit-of-life.mmd" -> (2, "Chapter 2")
    - "858489527-ICSE-Class-10-Selina-Biology-Chapter-07-the-Circulatory-System.mmd" -> (7, "Chapter 7")
    - "ICSE_MS_G10_CH01_INTERPRETATION_OF_TOPOGRAPHICAL_MAPS.mmd" -> (1, "Chapter 1")
    """
    # Try pattern: CH01, CH02, _CH1_ (ICSE_MS_G10 style)
    m = re.search(r"_CH0?(\d+)_", filename, re.IGNORECASE)
    if m:
        ch_no = int(m.group(1))
        ch_code = f"Chapter {ch_no}"
        return ch_no, ch_code

    # Try pattern: Chapter-07 or chapter-2
    m = re.search(r"[Cc]hapter[-_]?0?(\d+)", filename, re.IGNORECASE)
    if m:
        ch_no = int(m.group(1))
        ch_code = f"Chapter {ch_no}"
        return ch_no, ch_code

    # Fallback: try to extract any number sequence
    m = re.search(r"(\d+)", filename)
    if m:
        ch_no = int(m.group(1))
        ch_code = f"Chapter {ch_no}"
        return ch_no, ch_code

    return None, Path(filename).stem


def pick_first_file(ch_dir: Path, ext: str) -> Optional[Path]:
    files = sorted(ch_dir.glob(f"*{ext}"))
    return files[0] if files else None


def safe_read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def trim_mmd_text(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text

    # Keep the beginning (usually chapter heading + structure) and the end (often exercises/summary)
    head = text[: int(max_chars * 0.7)]
    tail = text[-int(max_chars * 0.3):]
    return head + "\n\n[...TRIMMED FOR LENGTH...]\n\n" + tail


def infer_chapter_title(mmd_text: str, fallback: str) -> str:
    # Try first markdown header
    m = re.search(r"^\s*#\s+(.+?)\s*$", mmd_text, flags=re.MULTILINE)
    if m:
        return m.group(1).strip()
    return fallback


def make_concept_id(chapter_code: str, idx: int) -> str:
    # stable, sortable IDs
    return f"{chapter_code}-C{idx:03d}"


def sanitize_mmd_references(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Replace 'MMD' references in output with natural language. Model sometimes outputs
    'MMD problem', 'from the MMD', etc. Replace with 'problem', 'from the chapter', etc.
    """
    replacements = [
        (r"\bMMD\s+problem\b", "problem"),
        (r"\bMMD\s+problems\b", "problems"),
        (r"\bMMDs?\b", "chapter"),  # MMD or MMDs -> chapter
        (r"in the MMD", "in the chapter"),
        (r"from the MMD", "from the chapter"),
        (r"the MMD\b", "the chapter"),
    ]
    for row in rows:
        for key in ("topic", "parent_concept", "concept", "concept_description"):
            if key in row and row[key]:
                s = row[key]
                for pat, repl in replacements:
                    s = re.sub(pat, repl, s, flags=re.IGNORECASE)
                row[key] = s
    return rows


def enforce_description_format(desc: str, subject: str) -> str:
    """
    Convert // separators to line breaks for structured display in Excel.
    The model outputs // as separators, which we convert to newlines for Excel cells.
    Excel will display these as structured multi-line content within each cell.
    """
    # First, normalize any existing newline patterns to // for consistency
    desc = re.sub(r'\n\s*(?=(?:Definition|Description|Usage|Misconception|Types|Examples):)', ' // ', desc)
    
    # Replace pipe separators with //
    if " | " in desc:
        desc = desc.replace(" | ", " // ")
    
    # Replace multiple newlines with //
    desc = re.sub(r'\n\s*\n+', ' // ', desc)
    
    # Now convert // separators to line breaks (newlines) for Excel
    # This creates structured multi-line content within the Excel cell
    # Excel cells with wrap_text enabled will display each section on a new line
    desc = re.sub(r'\s*//\s*', '\n', desc)
    desc = desc.strip()
    
    # Ensure it starts properly (remove leading newline if any)
    if desc.startswith('\n'):
        desc = desc[1:].strip()
    
    return desc


def sanitize_for_excel(value: str) -> str:
    """
    Remove illegal characters that Excel cannot handle.
    Excel doesn't allow control characters (0x00-0x1F) except tab, newline, and carriage return.
    """
    if not isinstance(value, str):
        return str(value) if value is not None else ""
    
    # Remove control characters except tab (\t), newline (\n), and carriage return (\r)
    cleaned = "".join(
        char if ord(char) >= 32 or char in "\t\n\r" else ""
        for char in value
    )
    return cleaned


def is_culmination_concept(concept_name: str) -> bool:
    """Return True if a concept row is marked as culmination."""
    if not concept_name or not isinstance(concept_name, str):
        return False
    name = concept_name.strip().lower()
    return name.startswith("culmination")


def ensure_culmination_per_topic(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Ensure each topic has exactly one culmination concept as the last row.

    Behavior:
    - Keeps topic order by first appearance.
    - Keeps non-culmination rows in original order.
    - If multiple culmination rows exist for a topic, keeps the first one.
    - If no culmination exists, appends a generated culmination row.
    """
    if not rows:
        return rows

    topic_order: List[str] = []
    topic_to_rows: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        topic = (row.get("topic") or "").strip() or "General"
        if topic not in topic_to_rows:
            topic_to_rows[topic] = []
            topic_order.append(topic)
        topic_to_rows[topic].append(row)

    normalized_rows: List[Dict[str, str]] = []
    for topic in topic_order:
        topic_rows = topic_to_rows[topic]
        non_culmination = [r for r in topic_rows if not is_culmination_concept(r.get("concept", ""))]
        culmination_rows = [r for r in topic_rows if is_culmination_concept(r.get("concept", ""))]

        # Keep all regular concepts first (original order preserved).
        normalized_rows.extend(non_culmination)

        if culmination_rows:
            culmination = culmination_rows[0]
            culmination["topic"] = topic
            normalized_rows.append(culmination)
        else:
            # Synthesize one fallback culmination row if model omitted it.
            parent = non_culmination[-1].get("parent_concept", "").strip() if non_culmination else "Integrated understanding"
            fallback = {
                "topic": topic,
                "parent_concept": parent if parent else "Integrated understanding",
                "concept": f"Culmination - {topic}",
                "concept_description": (
                    "Description: Consolidates the key ideas in this topic into one integrated understanding.\n"
                    "Types: Type 01: Mixed application Case 01: Solve questions combining multiple concepts from this topic.\n"
                    "Misconception: Treating each sub-concept in isolation instead of connecting them during problem solving."
                ),
            }
            normalized_rows.append(fallback)

    return normalized_rows


def get_system_prompt(subject: str) -> str:
    """
    Get subject-specific system prompt for concept extraction.
    """
    subject_lower = subject.lower()
    
    if subject_lower in ["mathematics", "math", "physics"]:
        # Math & Physics: Description (with examples), Types, Misconception
        return (
            "You are a STRICT concept mapping engine for school " + subject + " (board-level rigor).\n"
            "Return ONLY JSON matching the provided schema. No markdown. No commentary.\n\n"

            "OUTPUT CONTRACT (MUST FOLLOW EXACTLY):\n"
            "Each row has: topic, parent_concept, concept, concept_description.\n"
            "concept_description must be ONE string with exactly THREE sections, separated by // in this order:\n"
            "Description: <Detailed understanding: definition, explanation, key points, step-by-step reasoning, and worked examples integrated into the description> // Types: <Type 01: Name Case 01: ... Case 02: ... Type 02: Name Case 01: ... (organised classification of all numerical/question types)> // Misconception: <Common misconceptions students have about this concept>\n\n"
            "IMPORTANT: Use // (double slash) as the separator between each section. Do NOT use newlines.\n\n"

            "STYLE REQUIREMENTS:\n"
            "1) Concept naming grammar must be academic and patterned.\n"
            "   Prefer these templates:\n"
            "   - Properties and Applications of <X>\n"
            "   - Proof and Derivation of <rule/law>\n"
            "   - Conditions for Applying <rule/law>\n"
            "   - Representation of <X>\n"
            "   - Conceptual Meaning of <X>\n"
            "   - Methods of <procedure>\n"
            "   - Laws and Applications of <X>\n"
            "   - Converting Between <A> and <B>\n"
            "   - Simplifying Using <rule/law>\n"
            "   If a row is a synthesis, use:\n"
            "   Culmination – <Concept A> & <Concept B>\n\n"

            "2) TOPIC SEGREGATION AND NAMING (CRITICAL): A topic groups MULTIPLE related concepts together. Do NOT create one concept per topic.\n"
            "   - Each topic = one distinct section or logical block in the MMD, containing typically 5–15 (or more) concepts under it.\n"
            "   - Group related concepts under the same topic. For example, 'Rational numbers – definition and properties' should have several concepts: definition, properties, representation, etc. The LAST concept under that topic is the culmination.\n"
            "   - Follow the document structure. Match topic boundaries to natural content divisions (headings, sections). Do NOT create a new topic for each concept.\n"
            "   - DO NOT create a separate topic for exercises. Exercise sections (Exercise 1.1, Ex 2.1, etc.) at the end of the chapter must NOT become their own topic. Instead, distribute exercise problems into the content topics they test: add their question types as additional Types/Cases under the relevant concepts in the preceding topics.\n"
            "   - TOPIC NAMES MUST BE SOLID AND CONTENT-BASED, not 'Exercise 1.1', 'Ex 2.1'. Use descriptive names (e.g. 'Rational numbers – definition and properties', 'Operations on rational numbers', 'Laws of exponents and simplification').\n"
            "   - TOPIC names describe SUBJECT MATTER (e.g. 'Refraction through a lens', 'Image formation by convex lens'), NOT exercise categories like 'Ray-diagram Problem Solving and Exercises', 'Diagram completion and identification', or 'Exercise Problem Taxonomy'.\n\n"

            "3) CULMINATION PER TOPIC: The LAST concept under each topic must be a culmination row that synthesizes the concepts of that topic.\n"
            "   - Concept name: 'Culmination – <brief topic summary or Concept A & Concept B & ...>'.\n"
            "   - Place it as the last row for that topic before the next topic starts.\n"
            "   - concept_description: Description: <synthesis of the topic's concepts> // Types: <multi-concept problem types if any> (Misconception optional).\n\n"

            "4) Description must be DETAILED:\n"
            "   - Include complete definition and explanation, key properties, when/how to use.\n"
            "   - Integrate worked examples and step-by-step reasoning within the description.\n"
            "   - Include real-world applications where relevant.\n\n"

            "5) Types – CLASSIFY ALL numerical/questions under this concept. Use this EXACT structure:\n"
            "   Types: Type 01: <Name of this question/numerical type> Case 01: <brief description or example prompt> Case 02: ... Case 03: ... "
            "Type 02: <Name of next type> Case 01: ... Case 02: ... "
            "Type 03: ... and so on. Use zero-padded numbers (Type 01, Type 02, Case 01, Case 02).\n"
            "   - Extract EVERY numerical and question variety from the concept from the MMD (including from exercise sections); do not leave any out. Exercise problems belong under the concepts they test.\n"
            "   - Each Type is a distinct category of problem/question (e.g. 'Evaluating numerical expressions', 'Word problems involving rate').\n"
            "   - Under each Type, list Cases (Case 01, Case 02, ...) for specific sub-variants or example prompts (e.g. 'Evaluate:', 'Find:', 'Prove:').\n"
            "   - As many Type 01, Type 02, ... as needed to cover all numerical/question types; under each type as many Case 01, Case 02, ... as needed.\n\n"

            "6) Misconception: include only when it meaningfully applies (common errors, confusion with similar concepts); omit the Misconception section if not applicable.\n\n"

            "7) Parent concept should be meaningful and reusable across concepts. Avoid vague terms like 'Exercise Problem Taxonomy', 'Completion Tasks', 'Problem Solving and Exercises'. Use the actual concept (e.g. 'Ray diagrams for convex lens', 'Image formation in lenses').\n\n"

            "8) NEVER use 'MMD' or 'MMDs' in topic, parent_concept, concept, or concept_description. Write in normal academic language. Use 'chapter', 'text', 'problem', 'example', 'worked example' instead. E.g. 'Worked example: problem where C''=294 J K^-1 found by solving...' not 'MMD problem'.\n\n"

            "FORMAT EXAMPLE (Types section):\n"
            "Types: Type 01: Evaluating numerical exponential expressions Case 01: Evaluate 2^3 × 2^2 Case 02: Evaluate (3^2)^4 Case 03: Simplify and find value "
            "Type 02: Simplifying using laws of indices Case 01: Simplify a^m × a^n Case 02: Express as single power "
            "Type 03: ... // Misconception: ...\n\n"

            "QUALITY RULES:\n"
            "- Produce 40–60 concepts (excluding culmination rows).\n"
            "- No duplicates or near-duplicates: if the same concept appears in both theory and exercises, output it ONCE with the more comprehensive definition.\n"
            "- No vague filler like 'Introduction', 'Misc', 'Basics'.\n"
            "- Prefer small, testable, taggable concepts. Segregate topics by chapter flow; one culmination at the end of each topic.\n"
            "- CONCEPT names must be specific and content-based. Avoid 'Representation of Common Ray-Diagram Exercises and Completion Tasks'. Use clear subject matter (e.g. 'Ray diagrams for image formation', 'Lens formula applications').\n"
        )
    else:
        # Biology, Chemistry, History&Civics, Geography: Description (with examples), Types, Misconception
        return (
            "You are a STRICT concept mapping engine for school " + subject + " (board-level rigor).\n"
            "Return ONLY JSON matching the provided schema. No markdown. No commentary.\n\n"

            "OUTPUT CONTRACT (MUST FOLLOW EXACTLY):\n"
            "Each row has: topic, parent_concept, concept, concept_description.\n"
            "concept_description must be ONE string with exactly THREE sections, separated by // in this order:\n"
            "Description: <DETAILED understanding: complete definition and explanation, key characteristics and features, processes or relationships, relevant details from the chapter. Integrate concrete examples within the description to illustrate points.> // Types: <Type 01: Name Case 01: ... Case 02: ... Type 02: Name Case 01: ... (organised classification of all numerical/question types or problem varieties from the concept)> // Misconception: <Common misconceptions students have about this concept>\n\n"
            "IMPORTANT: Use // (double slash) as the separator between each section. Do NOT use newlines.\n\n"

            "STYLE REQUIREMENTS:\n"
            "1) Concept naming grammar must be academic and patterned.\n"
            "   Prefer these templates:\n"
            "   - Structure and Function of <X>\n"
            "   - Process of <X>\n"
            "   - Types and Classification of <X>\n"
            "   - Characteristics of <X>\n"
            "   - Relationship between <A> and <B>\n"
            "   - Causes and Effects of <X>\n"
            "   - Importance and Significance of <X>\n"
            "   - Comparison of <A> and <B>\n\n"

            "2) TOPIC SEGREGATION AND NAMING (CRITICAL): A topic groups MULTIPLE related concepts together. Do NOT create one concept per topic.\n"
            "   - Each topic = one distinct section or logical block, containing typically 5–15 (or more) concepts under it.\n"
            "   - Group related concepts under the same topic. The LAST concept under each topic is the culmination.\n"
            "   - Do NOT create a new topic for each concept. Topics are broader; each topic has many concepts within it.\n"
            "   - DO NOT create a separate topic for exercises. Exercise sections (Exercise 1.1, Ex 2.1, etc.) at the end of the chapter must NOT become their own topic. Instead, distribute exercise problems into the content topics they test: add their question types as additional Types/Cases under the relevant concepts in the preceding topics.\n"
            "   - Use solid, content-based topic names (e.g. 'Structure and function of X', 'Process of Y', 'Application to real scenarios').\n"
            "   - TOPIC names describe SUBJECT MATTER, NOT exercise categories like 'Problem Solving and Exercises', 'Exercise Problem Taxonomy'.\n\n"

            "3) CULMINATION PER TOPIC: The LAST concept under each topic must be a culmination row that synthesizes the concepts of that topic.\n"
            "   - Concept name: 'Culmination – <brief topic summary or Concept A & Concept B & ...>'.\n"
            "   - Place it as the last row for that topic before the next topic starts. Supports questions spanning multiple concepts within that topic.\n"
            "   - concept_description: Description: <synthesis> // Types: <multi-concept problem types if any> (Misconception optional).\n\n"

            "4) Description must be COMPREHENSIVE:\n"
            "   - Include ALL relevant information from the chapter; do not omit important details.\n"
            "   - Explain processes step-by-step where applicable; include relationships and dependencies.\n"
            "   - Use examples naturally within the description to clarify points (Biology/Chemistry: biological/chemical examples; History/Civics: events, dates, contexts; Geography: locations, phenomena, data).\n\n"

            "5) Types – CLASSIFY ALL numerical questions, problem types, or exercise varieties under this concept. Use this EXACT structure:\n"
            "   Types: Type 01: <Name of this question/problem type> Case 01: <brief description or example> Case 02: ... Case 03: ... "
            "Type 02: <Name of next type> Case 01: ... Case 02: ... "
            "Type 03: ... and so on. Use zero-padded numbers (Type 01, Type 02, Case 01, Case 02).\n"
            "   - Extract EVERY question/numerical/problem variety from the concept in the MMD (including from exercise sections); do not leave any out. Exercise problems belong under the concepts they test.\n"
            "   - Each Type is a distinct category (e.g. 'Calculation-based', 'Application to real scenario', 'Comparison/analysis').\n"
            "   - Under each Type, list Cases (Case 01, Case 02, ...) for specific sub-variants or example prompts.\n"
            "   - As many Type 01, Type 02, ... as needed; under each type as many Case 01, Case 02, ... as needed.\n\n"

            "6) Misconception: include only when it meaningfully applies; omit the Misconception section if not applicable.\n\n"

            "7) Parent concept should be meaningful and reusable across concepts. Avoid vague terms like 'Exercise Problem Taxonomy', 'Completion Tasks'. Use the actual concept.\n\n"

            "8) NEVER use 'MMD' or 'MMDs' in topic, parent_concept, concept, or concept_description. Write in normal academic language. Use 'chapter', 'problem', 'example', 'worked example' instead.\n\n"

            "FORMAT EXAMPLE (Types section):\n"
            "Types: Type 01: Direct calculation Case 01: Find value when given... Case 02: Compare two values "
            "Type 02: Word problems Case 01: Given X find Y Case 02: Application to scenario "
            "Type 03: ... // Misconception: ...\n\n"

            "QUALITY RULES:\n"
            "- Produce 40–60 concepts (excluding culmination rows).\n"
            "- No duplicates or near-duplicates: same concept in theory and exercises → output ONCE.\n"
            "- No vague filler like 'Introduction', 'Misc', 'Basics'.\n"
            "- Prefer small, testable, taggable concepts. Segregate topics by chapter flow; one culmination at the end of each topic.\n"
            "- Description must be complete and comprehensive - do not leave out important information.\n"
            "- CONCEPT names must be specific and content-based. Avoid vague exercise-style names like 'Representation of Common Exercises and Completion Tasks'.\n"
        )


def gpt_extract_concepts(chapter_label: str, mmd_text: str, subject: str, model: Optional[str] = None) -> List[Dict[str, str]]:
    system = get_system_prompt(subject)

    user = (
        f"CHAPTER: {chapter_label}\n\n"
        f"Extract {CONCEPTS_PER_CHAPTER_MIN}–{CONCEPTS_PER_CHAPTER_MAX} high-quality concepts.\n"
        "Return rows with topic, parent_concept, concept, concept_description.\n"
        "Group concepts under topics: each topic has 5–15+ concepts; the last concept per topic is the culmination. "
        "Do NOT create one concept per topic. Segregate by MMD flow (section order, headings). "
        "Do NOT create a separate topic for exercises—distribute exercise problems into the content topics they test. "
        "At the end of each topic, add exactly one culmination row: 'Culmination – <summary or Concept A & Concept B & ...>'.\n"
        "Never use 'MMD' in your output—use 'chapter', 'problem', 'example' instead.\n\n"
        "CHAPTER CONTENT:\n"
        f"{mmd_text}"
    )

    use_model = model if model else MODEL
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.responses.create(
                model=use_model,
                max_output_tokens=MAX_OUTPUT_TOKENS,
                input=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                text={
                    "format": {
                        "type": "json_schema",
                        "name": CONCEPT_SCHEMA["name"],
                        "schema": CONCEPT_SCHEMA["schema"]
                    }
                }
            )

            # Official SDK convenience: aggregated text output
            out_text = resp.output_text
            data = json.loads(out_text)

            rows = data.get("rows", [])
            # normalize description to // separator format (best-effort safety)
            for r in rows:
                r["concept_description"] = enforce_description_format(r["concept_description"], subject)
            return rows

        except Exception as e:
            last_err = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_SLEEP_SECONDS)
            else:
                raise RuntimeError(f"GPT extraction failed after {MAX_RETRIES} retries: {repr(last_err)}")


def autosize_columns(ws):
    # Simple autosize based on max length per column (bounded)
    max_width = 60
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(12, max_len * 0.9), max_width)


# ============================================================
# MAIN
# ============================================================
def process_mmd_file(
    mmd_path: Path,
    ws,
    subject: str,
    output_dir: Optional[Path] = None,
    *,
    board: str = None,
    book: str = None,
    grade: int = None,
    model: Optional[str] = None,
    chapter_no_override: Optional[int] = None,
    chapter_stem_override: Optional[str] = None,
    pdf_path_override: Optional[Path] = None,
    write_display_name_column: bool = False,
    parent_concept_override: Optional[str] = None,
    chapter_label_for_extraction: Optional[str] = None,
) -> int:
    """
    Process a single mmd file and add rows to the worksheet.
    Returns the number of rows added.
    """
    if not mmd_path.exists():
        print(f"[SKIP] MMD file not found: {mmd_path}")
        return 0
    
    # Resolve PDF path
    if pdf_path_override is not None and pdf_path_override.exists():
        pdf_path = pdf_path_override
    else:
        pdf_path = mmd_path.parent / mmd_path.name.replace(".mmd", ".pdf")
        if not pdf_path.exists():
            pdf_path = resolve_pdf_beside_mmd(mmd_path, [MSBSHSE_DEFAULT_PDF_DIR])
        if pdf_path is None or not pdf_path.exists():
            pdf_candidates = list(mmd_path.parent.glob("*.pdf"))
            pdf_path = pdf_candidates[0] if pdf_candidates else None

    # Extract chapter info from filename; chapter title from filename stem in Title Case (e.g. NUMBER_SYSTEM -> Number System)
    ch_no, chapter_code = extract_chapter_info_from_filename(mmd_path.name)
    chptr_stem_from_file = get_chapter_name_for_code(mmd_path.name)
    chapter_title = chapter_name_stem_to_title(chptr_stem_from_file) if chptr_stem_from_file else None
    if not chapter_title:
        mmd_text_for_title = safe_read_text(mmd_path)
        chapter_title = infer_chapter_title(mmd_text_for_title, fallback=chapter_code)

    if chapter_stem_override:
        stem = chapter_stem_override.strip()
        chapter_title = stem
        chapter_code = stem
        chptr_stem_from_file = stem
        if chapter_no_override is not None:
            ch_no = chapter_no_override
    
    mmd_text = safe_read_text(mmd_path)
    mmd_text = trim_mmd_text(mmd_text, MAX_MMD_CHARS)
    
    chapter_label = f"{chapter_code} | {chapter_title}" if chapter_code != chapter_title else chapter_title
    extract_label = chapter_label_for_extraction if chapter_label_for_extraction else chapter_label
    print(f"[Extracting] {extract_label} (Subject: {subject})" + (f" [Model: {model}]" if model else ""))
    
    concepts = gpt_extract_concepts(extract_label, mmd_text, subject, model=model)
    concepts = ensure_culmination_per_topic(concepts)
    concepts = sanitize_mmd_references(concepts)
    
    _board = board if board is not None else BOARD
    _book = book if book is not None else BOOK
    _grade = grade if grade is not None else GRADE
    _board_code = board_code_from_board(_board)
    chptr_stem = chptr_stem_from_file or chapter_title or chapter_code
    chptr_code = normalize_for_code(chptr_stem) if isinstance(chptr_stem, str) and chptr_stem else normalize_for_code(str(chapter_code))
    if not chptr_code:
        chptr_code = "Chapter"

    topic_to_num: Dict[str, int] = {}
    next_topic_num = 1
    rows_added = 0
    for i, r in enumerate(concepts, start=1):
        if chapter_stem_override and ch_no is not None:
            concept_id = make_concept_id(f"CH{ch_no:02d}", i)
        else:
            concept_id = make_concept_id(chapter_code, i)
        topic_name = r["topic"].strip()
        concept_name = r["concept"].strip()
        if topic_name not in topic_to_num:
            topic_to_num[topic_name] = next_topic_num
            next_topic_num += 1
        topic_num = topic_to_num[topic_name]
        tpc_code = normalize_for_code(topic_name) or f"Topic_{topic_num:02d}"
        topic_display = format_topic_display(_grade, _board_code, chptr_code, topic_num, topic_name)
        concept_display = format_concept_display(_grade, _board_code, chptr_code, tpc_code, concept_name)
        chapter_display = chapter_title if chptr_stem_from_file else chapter_code
        parent_cell = (
            parent_concept_override.strip()
            if parent_concept_override
            else r["parent_concept"].strip()
        )
        row_cells: List[Any] = [
            _board, _book, _grade, subject,
            ch_no, sanitize_for_excel(chapter_display), sanitize_for_excel(chapter_title),
            sanitize_for_excel(topic_display),
            sanitize_for_excel(parent_cell),
            sanitize_for_excel(concept_display),
        ]
        if write_display_name_column:
            row_cells.append(sanitize_for_excel(concept_name))
        row_cells.extend([
            sanitize_for_excel(r["concept_description"].strip()),
            concept_id,
            str(mmd_path),
            str(pdf_path) if pdf_path else ""
        ])
        ws.append(row_cells)
        rows_added += 1
    
    # Wrap text for Concept Description column for the rows just added
    desc_col = 12 if write_display_name_column else 11
    for row_idx in range(ws.max_row - len(concepts) + 1, ws.max_row + 1):
        ws.cell(row=row_idx, column=desc_col).alignment = Alignment(wrap_text=True, vertical="top")
    
    return rows_added


def _run_cbse_rd_by_grade(
    mmd_folder: Path,
    mmd_files: List[Path],
    chapter_limit: Optional[int] = None,
    grade_filter: Optional[int] = None,
) -> None:
    """Process mmds_CBSE_RD_Mathematics: one Excel per grade. grade_filter=9 or 10 runs only that grade. chapter_limit limits G09 chapters (for testing)."""
    cfg = MMDS_CBSE_RD_CONFIG
    board = cfg.get("BOARD", "CBSE")
    book = cfg.get("BOOK", "RD")
    subject = cfg.get("SUBJECT", "Mathematics")
    output_dir = cfg.get("OUTPUT_DIR", mmd_folder.parent)

    g09 = sorted([f for f in mmd_files if "_G09_" in f.name], key=lambda p: natural_sort_key(p.name))
    g10 = sorted([f for f in mmd_files if "_G10_" in f.name], key=lambda p: natural_sort_key(p.name))

    if grade_filter == 9:
        g10 = []
        print(f"[FILTER] Processing only Grade 09. Output: CBSE_RD_G09_Mathematics_Concepts.xlsx\n")
    elif grade_filter == 10:
        g09 = []
        print(f"[FILTER] Processing only Grade 10. Output: CBSE_RD_G10_Mathematics_Concepts.xlsx\n")
    elif chapter_limit is not None:
        g09 = g09[:chapter_limit]
        g10 = []
        print(f"[LIMIT] Processing only Grade 09, first {chapter_limit} chapter(s). Output: CBSE_RD_G09_Mathematics_Concepts.xlsx\n")

    use_grade_subfolders = cfg.get("USE_GRADE_SUBFOLDERS", False)
    for grade, files in [(9, g09), (10, g10)]:
        if not files:
            print(f"[SKIP] No G{grade:02d} .mmd files in {mmd_folder.name}")
            continue
        grade_dir = output_dir / f"Grade {grade:02d}" if use_grade_subfolders else output_dir
        output_xlsx = grade_dir / f"CBSE_RD_G{grade:02d}_Mathematics_Concepts.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Concepts"
        ws.freeze_panes = "A2"
        headers = [
            "Board", "Book", "Grade", "Subject",
            "Chapter No", "Chapter Code", "Chapter Title",
            "Topic", "Parent Concept", "Concept", "Concept Description",
            "Concept ID", "MMD Path", "PDF Path"
        ]
        ws.append(headers)
        header_font = Font(bold=True)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        total_rows = 0
        print(f"\n--- Grade {grade} ({len(files)} chapter(s)) -> {output_xlsx.name} ---")
        for ch_idx, mmd_path in enumerate(files, start=1):
            print(f"  [Chapter {ch_idx}/{len(files)}] {mmd_path.name}")
            rows = process_mmd_file(
                mmd_path, ws, subject,
                board=board, book=book, grade=grade
            )
            total_rows += rows
            grade_dir.mkdir(parents=True, exist_ok=True)
            wb.save(output_xlsx)
            print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
        autosize_columns(ws)
        grade_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_xlsx)
        print(f"[DONE] Grade {grade}: {total_rows} rows -> {output_xlsx}")


def _infer_cbse_folder_config(folder_path: Path) -> Optional[Dict[str, Any]]:
    """Infer board, book, subject, grade from CBSE folder path (e.g. Class 10\\CBSE_RD_G10_Mathematics)."""
    path_str = str(folder_path)
    folder_name = folder_path.name
    grade = 10 if "Class 10" in path_str or "Class10" in path_str else (9 if "Class 9" in path_str or "Class9" in path_str else None)
    if grade is None:
        m = re.search(r"_G0?(\d)_", folder_name, re.IGNORECASE)
        grade = int(m.group(1)) if m else 10
    if "CBSE_RD" in folder_name and ("Mathematics" in folder_name or "MATH" in folder_name.upper()):
        return {"board": "CBSE", "book": "RD", "subject": "Mathematics", "grade": grade}
    if folder_name == "Physics":
        return {"board": "CBSE", "book": "Schand", "subject": "Physics", "grade": grade}
    if folder_name == "Chemistry":
        return {"board": "CBSE", "book": "Schand", "subject": "Chemistry", "grade": grade}
    if folder_name == "Biology":
        return {"board": "CBSE", "book": "Schand", "subject": "Biology", "grade": grade}
    if "SocialScience" in folder_name or "Social_Science" in folder_name:
        return {"board": "CBSE", "book": "NCERT", "subject": "Social Science", "grade": grade}
    return None


def _run_all_cbse_subjects(root: Path) -> None:
    """Process all CBSE subject folders under root in one run. Output to Grade 09/ and Grade 10/ subfolders."""
    output_dir = CBSE_ALL_OUTPUT_DIR
    all_mmds = list(root.rglob("*.mmd"))
    if not all_mmds:
        raise RuntimeError(f"No .mmd files found under {root}")
    folder_to_files: Dict[Path, List[Path]] = {}
    for p in all_mmds:
        parent = p.parent
        if parent not in folder_to_files:
            folder_to_files[parent] = []
        folder_to_files[parent].append(p)
    subject_folders = sorted(folder_to_files.keys(), key=lambda p: (natural_sort_key(str(p)), p.name))
    print(f"[CBSE ALL] Found {len(subject_folders)} subject folder(s) under {root}\n")
    for folder in subject_folders:
        mmd_files = sorted(folder_to_files[folder], key=lambda p: natural_sort_key(p.name))
        cfg = _infer_cbse_folder_config(folder)
        if not cfg:
            print(f"[SKIP] Unknown config for {folder.name} - skipping")
            continue
        board = cfg["board"]
        book = cfg["book"]
        subject = cfg["subject"]
        grade = cfg["grade"]
        grade_dir = output_dir / f"Grade {grade:02d}"
        subject_token = re.sub(r"[^A-Za-z0-9]+", "_", subject).strip("_")
        output_xlsx = grade_dir / f"CBSE_{book}_G{grade:02d}_{subject_token}_Concepts.xlsx"
        output_xlsx = make_unique_output_path(output_xlsx)
        wb = Workbook()
        ws = wb.active
        ws.title = "Concepts"
        ws.freeze_panes = "A2"
        headers = [
            "Board", "Book", "Grade", "Subject",
            "Chapter No", "Chapter Code", "Chapter Title",
            "Topic", "Parent Concept", "Concept", "Concept Description",
            "Concept ID", "MMD Path", "PDF Path"
        ]
        ws.append(headers)
        header_font = Font(bold=True)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        total_rows = 0
        print(f"\n--- {subject} Grade {grade} ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
        for ch_idx, mmd_path in enumerate(mmd_files, start=1):
            print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
            rows = process_mmd_file(mmd_path, ws, subject, board=board, book=book, grade=grade)
            total_rows += rows
            grade_dir.mkdir(parents=True, exist_ok=True)
            wb.save(output_xlsx)
            print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
        for row_idx in range(ws.max_row - total_rows + 1, ws.max_row + 1):
            ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical="top")
        autosize_columns(ws)
        grade_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_xlsx)
        print(f"[DONE] {subject} Grade {grade}: {total_rows} rows -> {output_xlsx}")


def _run_icse_se_by_grade(mmd_folder: Path, mmd_files: List[Path], grade_filter: Optional[int] = None) -> None:
    """Process mmds_ICSE_SE_Mathematics: one Excel per grade. grade_filter=9 runs only G09, grade_filter=10 only G10, None runs both."""
    cfg = MMDS_ICSE_SE_CONFIG
    board = cfg.get("BOARD", "ICSE")
    book = cfg.get("BOOK", "Selina")
    subject = cfg.get("SUBJECT", "Mathematics")
    output_dir = cfg.get("OUTPUT_DIR", mmd_folder.parent)

    g09 = sorted([f for f in mmd_files if "_G09_" in f.name], key=lambda p: natural_sort_key(p.name))
    g10 = sorted([f for f in mmd_files if "_G10_" in f.name], key=lambda p: natural_sort_key(p.name))

    if grade_filter == 9:
        g10 = []
        print(f"[FILTER] Processing only Grade 09. Output: ICSE_SE_G09_Mathematics_Concepts.xlsx\n")
    elif grade_filter == 10:
        g09 = []
        print(f"[FILTER] Processing only Grade 10. Output: ICSE_SE_G10_Mathematics_Concepts.xlsx\n")

    for grade, files in [(9, g09), (10, g10)]:
        if not files:
            print(f"[SKIP] No G{grade:02d} .mmd files in {mmd_folder.name}")
            continue
        output_xlsx = output_dir / f"ICSE_SE_G{grade:02d}_Mathematics_Concepts.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Concepts"
        ws.freeze_panes = "A2"
        headers = [
            "Board", "Book", "Grade", "Subject",
            "Chapter No", "Chapter Code", "Chapter Title",
            "Topic", "Parent Concept", "Concept", "Concept Description",
            "Concept ID", "MMD Path", "PDF Path"
        ]
        ws.append(headers)
        header_font = Font(bold=True)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        total_rows = 0
        print(f"\n--- Grade {grade} ({len(files)} chapter(s)) -> {output_xlsx.name} ---")
        for ch_idx, mmd_path in enumerate(files, start=1):
            print(f"  [Chapter {ch_idx}/{len(files)}] {mmd_path.name}")
            rows = process_mmd_file(
                mmd_path, ws, subject,
                board=board, book=book, grade=grade
            )
            total_rows += rows
            output_dir.mkdir(parents=True, exist_ok=True)
            wb.save(output_xlsx)
            print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
        autosize_columns(ws)
        wb.save(output_xlsx)
        print(f"[DONE] Grade {grade}: {total_rows} rows -> {output_xlsx}")


def _run_compare_models(mmd_folder: Path, mmd_files: List[Path]) -> None:
    """Temporary: run Chapter 01 only with each model in COMPARE_MODELS, output separate Excels."""
    # Pick first CH01 file (prefer G09)
    ch01_files = [f for f in mmd_files if "_CH01_" in f.name.upper()]
    if not ch01_files:
        ch01_files = sorted([f for f in mmd_files if "_G09_" in f.name], key=lambda p: natural_sort_key(p.name))[:1]
    if not ch01_files:
        ch01_files = mmd_files[:1]
    mmd_path = ch01_files[0]

    # Infer config from folder name
    if mmd_folder.name in MMDS_CBSE_RD_FOLDER_NAMES:
        cfg = MMDS_CBSE_RD_CONFIG
    elif mmd_folder.name == MMDS_ICSE_SE_MATHEMATICS.name:
        cfg = MMDS_ICSE_SE_CONFIG
    else:
        cfg = {"BOARD": "CBSE", "BOOK": "RD", "SUBJECT": "Mathematics"}
    board = cfg.get("BOARD", "CBSE")
    book = cfg.get("BOOK", "RD")
    subject = cfg.get("SUBJECT", "Mathematics")
    output_dir = cfg.get("OUTPUT_DIR", mmd_folder.parent)
    grade = 9 if "_G09_" in mmd_path.name else 10

    print(f"[COMPARE] Chapter 01: {mmd_path.name}")
    print(f"  Models: {', '.join(COMPARE_MODELS)}\n")

    for model in COMPARE_MODELS:
        safe_name = model.replace(".", "-")
        output_xlsx = output_dir / f"Chapter01_{safe_name}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Concepts"
        ws.freeze_panes = "A2"
        headers = [
            "Board", "Book", "Grade", "Subject",
            "Chapter No", "Chapter Code", "Chapter Title",
            "Topic", "Parent Concept", "Concept", "Concept Description",
            "Concept ID", "MMD Path", "PDF Path"
        ]
        ws.append(headers)
        header_font = Font(bold=True)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        print(f"  [{model}] Extracting...")
        rows = process_mmd_file(mmd_path, ws, subject, board=board, book=book, grade=grade, model=model)
        for row_idx in range(ws.max_row - rows + 1, ws.max_row + 1):
            ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical="top")
        autosize_columns(ws)
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(output_xlsx)
        print(f"  [{model}] Done -> {output_xlsx.name} ({rows} rows)\n")
    print("[DONE] Compare mode complete. Outputs in", output_dir)


def main():
    mmd_files = None
    output_xlsx = OUTPUT_XLSX
    subject = SUBJECT
    book = BOOK
    board = BOARD
    grade = GRADE

    # Check if mmd files or folder are provided as command-line arguments
    if len(sys.argv) > 1:
        if sys.argv[1].lower() in ("msbshse11", "msbshse_g11"):
            paths = [Path(p).resolve() for p in sys.argv[2:]]
            if not paths:
                raise RuntimeError(
                    "Usage: python mmd_to_concepts_excel.py msbshse11 "
                    "<folder_with_mmd> | <file.mmd> [...]"
                )
            mmd_batch: List[Path] = []
            for p in paths:
                if p.is_dir():
                    mmd_batch.extend(sorted(p.glob("*.mmd"), key=lambda x: natural_sort_key(x.name)))
                elif p.suffix.lower() == ".mmd":
                    mmd_batch.append(p)
            mmd_batch = sorted(set(mmd_batch), key=lambda p: natural_sort_key(p.name))
            if not mmd_batch:
                raise RuntimeError("msbshse11: no .mmd files found")
            out_parent = paths[0] if paths[0].is_dir() else paths[0].parent
            if len(mmd_batch) == 1:
                stem_slug = re.sub(r"[^A-Za-z0-9]+", "_", mmd_batch[0].stem).strip("_") or "Chapter"
                output_xlsx = make_unique_output_path(
                    out_parent / f"MSBSHSE_G11_{stem_slug}_Concepts.xlsx"
                )
            else:
                output_xlsx = make_unique_output_path(out_parent / "MSBSHSE_G11_Concepts.xlsx")
            board, book, grade = "MSBSHSE", "M-State", 11
            wb = Workbook()
            ws = wb.active
            ws.title = "Concepts"
            ws.freeze_panes = "A2"
            headers = [
                "Board", "Book", "Grade", "Subject",
                "Chapter No", "Chapter Code", "Chapter Title",
                "Topic", "Parent Concept", "Concept", "Display Name", "Concept Description",
                "Concept ID", "MMD Path", "PDF Path"
            ]
            ws.append(headers)
            header_font = Font(bold=True)
            for c in range(1, len(headers) + 1):
                ws.cell(row=1, column=c).font = header_font
                ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            total_rows = 0
            print(f"\n--- MSBSHSE G11 ({len(mmd_batch)} chapter(s)) -> {output_xlsx.name} ---")
            for ch_idx, mmd_path in enumerate(mmd_batch, start=1):
                subj = infer_subject_from_chapter_stem(mmd_path.stem)
                print(f"  [Chapter {ch_idx}/{len(mmd_batch)}] {mmd_path.name} (Subject: {subj})")
                rows = process_mmd_file(
                    mmd_path,
                    ws,
                    subj,
                    board=board,
                    book=book,
                    grade=grade,
                    chapter_no_override=ch_idx,
                    chapter_stem_override=mmd_path.stem,
                    write_display_name_column=True,
                )
                total_rows += rows
                output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                wb.save(output_xlsx)
                print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
            autosize_columns(ws)
            output_xlsx.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_xlsx)
            print(f"[DONE] {total_rows} rows -> {output_xlsx}")
            return

        if sys.argv[1].lower() in ("cbse_ncert_bio", "cbse_ncert_biology"):
            paths = [Path(p).resolve() for p in sys.argv[2:]]
            if not paths:
                raise RuntimeError(
                    "Usage: python mmd_to_concepts_excel.py cbse_ncert_bio "
                    "<folder_with_mmd> | <file.mmd> [...]"
                )
            mmd_batch: List[Path] = []
            for p in paths:
                if p.is_dir():
                    mmd_batch.extend(sorted(p.glob("*.mmd"), key=lambda x: natural_sort_key(x.name)))
                elif p.suffix.lower() == ".mmd":
                    mmd_batch.append(p)
            mmd_batch = sorted(set(mmd_batch), key=lambda p: natural_sort_key(p.name))
            if not mmd_batch:
                raise RuntimeError("cbse_ncert_bio: no .mmd files found")
            out_parent = paths[0] if paths[0].is_dir() else paths[0].parent
            output_xlsx = make_unique_output_path(out_parent / "CBSE_NCERT_Biology_Concepts.xlsx")
            board, book, subject = "CBSE", "NCERT", "Biology"
            wb = Workbook()
            ws = wb.active
            ws.title = "Concepts"
            ws.freeze_panes = "A2"
            headers = [
                "Board", "Book", "Grade", "Subject",
                "Chapter No", "Chapter Code", "Chapter Title",
                "Topic", "Parent Concept", "Concept", "Display Name", "Concept Description",
                "Concept ID", "MMD Path", "PDF Path"
            ]
            ws.append(headers)
            header_font = Font(bold=True)
            for c in range(1, len(headers) + 1):
                ws.cell(row=1, column=c).font = header_font
                ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
            total_rows = 0
            print(f"\n--- CBSE NCERT Biology ({len(mmd_batch)} chapter(s)) -> {output_xlsx.name} ---")
            for ch_idx, mmd_path in enumerate(mmd_batch, start=1):
                stem_raw = mmd_path.stem
                cleaned_title = strip_grade_suffix_from_chapter_stem(stem_raw)
                g = parse_grade_from_chapter_stem(stem_raw)
                if g is None:
                    print(f"  [WARN] No (Grade N) in filename, defaulting to 10: {mmd_path.name}")
                    g = 10
                parent_phys = ncert_plant_or_animal_physiology_parent(stem_raw)
                extract_lbl = f"{cleaned_title} | Grade {g} | CBSE NCERT Biology"
                print(
                    f"  [Chapter {ch_idx}/{len(mmd_batch)}] {mmd_path.name} "
                    f"(Grade {g}, Parent bucket: {parent_phys})"
                )
                rows = process_mmd_file(
                    mmd_path,
                    ws,
                    subject,
                    board=board,
                    book=book,
                    grade=g,
                    chapter_no_override=ch_idx,
                    chapter_stem_override=cleaned_title,
                    parent_concept_override=parent_phys,
                    chapter_label_for_extraction=extract_lbl,
                    write_display_name_column=True,
                )
                total_rows += rows
                output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                wb.save(output_xlsx)
                print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
            autosize_columns(ws)
            output_xlsx.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_xlsx)
            print(f"[DONE] {total_rows} rows -> {output_xlsx}")
            return

        first_arg = Path(sys.argv[1]).resolve()
        if first_arg.is_dir():
            # CBSE root: process ALL subjects in one run
            if first_arg.name == "CBSE" or first_arg == MMDS_CBSE_ROOT.resolve():
                _run_all_cbse_subjects(first_arg)
                return
            # First arg is a folder: process all .mmd files in it
            mmd_files = sorted(first_arg.glob("*.mmd"), key=lambda p: natural_sort_key(p.name))
            if not mmd_files:
                raise RuntimeError(f"No .mmd files found in {first_arg}")
            # Temporary compare mode: run Chapter 01 with COMPARE_MODELS, output one Excel per model.
            if len(sys.argv) > 2 and sys.argv[2].lower() == "compare":
                _run_compare_models(first_arg, mmd_files)
                return
            # MSBSHSE Symbiosis: Board=MSBSHSE, Book=M-State, Grade=11, Subject from filename
            if MMDS_SYMBIOSIS_CONFIG and (first_arg.name == "Symbiosis" or first_arg.resolve() == MMDS_SYMBIOSIS.resolve()):
                cfg = MMDS_SYMBIOSIS_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- MSBSHSE M-State G11 Symbiosis ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    subject = extract_subject_from_msbshse_filename(mmd_path.name)
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name} (Subject: {subject})")
                    rows = process_mmd_file(
                        mmd_path, ws, subject,
                        board=cfg.get("BOARD", "MSBSHSE"),
                        book=cfg.get("BOOK", "M-State"),
                        grade=cfg.get("GRADE", 11),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # CBSE RD Mathematics: one output per grade. 2nd arg: 9 (G09 only), 10 (G10 only), or small number (chapter limit for G09 testing).
            if MMDS_CBSE_RD_CONFIG and first_arg.name in MMDS_CBSE_RD_FOLDER_NAMES:
                chapter_limit = None
                grade_filter = None
                if len(sys.argv) > 2 and sys.argv[2].isdigit():
                    n = int(sys.argv[2])
                    if n == 9:
                        grade_filter = 9
                    elif n == 10:
                        grade_filter = 10
                    else:
                        chapter_limit = n
                _run_cbse_rd_by_grade(first_arg, mmd_files, chapter_limit=chapter_limit, grade_filter=grade_filter)
                return
            # ICSE SE G09 Biology: single output file
            if MMDS_ICSE_SE_G09_BIOLOGY_CONFIG and first_arg.resolve() == MMDS_ICSE_SE_G09_BIOLOGY.resolve():
                cfg = MMDS_ICSE_SE_G09_BIOLOGY_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 Biology ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "Biology"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G09 Chemistry: single output file
            if MMDS_ICSE_SE_G09_CHEMISTRY_CONFIG and first_arg.resolve() == MMDS_ICSE_SE_G09_CHEMISTRY.resolve():
                cfg = MMDS_ICSE_SE_G09_CHEMISTRY_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 Chemistry ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "Chemistry"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G09 Geography: single output file
            if MMDS_ICSE_SE_G09_GEOGRAPHY_CONFIG and first_arg.name == MMDS_ICSE_SE_G09_GEOGRAPHY.name:
                cfg = MMDS_ICSE_SE_G09_GEOGRAPHY_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 Geography ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "Geography"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G09 Physics: single output file
            if MMDS_ICSE_SE_G09_PHYSICS_CONFIG and first_arg.name == MMDS_ICSE_SE_G09_PHYSICS.name:
                cfg = MMDS_ICSE_SE_G09_PHYSICS_CONFIG
                output_xlsx = cfg.get("OUTPUT", first_arg.parent / "ICSE_SE_G09_Physics_Concepts.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 Physics ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "Physics"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G10 History
            if first_arg.name == MMDS_ICSE_SE_G10_HISTORY_NAME:
                subject = "History"
                output_xlsx = make_unique_output_path(first_arg.parent / f"{first_arg.name}_Concepts.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE MS G10 History ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, subject,
                        board="ICSE",
                        book="MS",
                        grade=10,
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                for row_idx in range(ws.max_row - total_rows + 1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical="top")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G10 Civics
            if first_arg.name == MMDS_ICSE_SE_G10_CIVICS_NAME:
                subject = "Civics"
                output_xlsx = make_unique_output_path(first_arg.parent / f"{first_arg.name}_Concepts.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE MS G10 Civics ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, subject,
                        board="ICSE",
                        book="MS",
                        grade=10,
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                for row_idx in range(ws.max_row - total_rows + 1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical="top")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G09 History: single output file
            if MMDS_ICSE_SE_G09_HISTORY_CONFIG and first_arg.name == MMDS_ICSE_SE_G09_HISTORY.name:
                cfg = MMDS_ICSE_SE_G09_HISTORY_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 History ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "History"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE G09 Civics: single output file
            if MMDS_ICSE_SE_G09_CIVICS_CONFIG and first_arg.name == MMDS_ICSE_SE_G09_CIVICS.name:
                cfg = MMDS_ICSE_SE_G09_CIVICS_CONFIG
                output_xlsx = cfg.get("OUTPUT")
                wb = Workbook()
                ws = wb.active
                ws.title = "Concepts"
                ws.freeze_panes = "A2"
                headers = [
                    "Board", "Book", "Grade", "Subject",
                    "Chapter No", "Chapter Code", "Chapter Title",
                    "Topic", "Parent Concept", "Concept", "Concept Description",
                    "Concept ID", "MMD Path", "PDF Path"
                ]
                ws.append(headers)
                header_font = Font(bold=True)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=1, column=c).font = header_font
                    ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
                total_rows = 0
                print(f"\n--- ICSE SE G09 Civics ({len(mmd_files)} chapter(s)) -> {output_xlsx.name} ---")
                for ch_idx, mmd_path in enumerate(mmd_files, start=1):
                    print(f"  [Chapter {ch_idx}/{len(mmd_files)}] {mmd_path.name}")
                    rows = process_mmd_file(
                        mmd_path, ws, cfg.get("SUBJECT", "Civics"),
                        board=cfg.get("BOARD", "ICSE"),
                        book=cfg.get("BOOK", "Selina"),
                        grade=cfg.get("GRADE", 9),
                    )
                    total_rows += rows
                    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                    wb.save(output_xlsx)
                    print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
                autosize_columns(ws)
                wb.save(output_xlsx)
                print(f"[DONE] {total_rows} rows -> {output_xlsx}")
                return
            # ICSE SE Mathematics: one output per grade. Optional 2nd arg = 9 (G09 only) or 10 (G10 only).
            if MMDS_ICSE_SE_CONFIG and first_arg.name == MMDS_ICSE_SE_MATHEMATICS.name:
                grade_filter = None
                if len(sys.argv) > 2 and sys.argv[2].isdigit():
                    g = int(sys.argv[2])
                    if g in (9, 10):
                        grade_filter = g
                _run_icse_se_by_grade(first_arg, mmd_files, grade_filter=grade_filter)
                return
            # Use folder-specific config if defined
            if MMDS_FOLDER_CONFIG and first_arg.name == "mmds_ICSE_MS_G10_Geography":
                book = MMDS_FOLDER_CONFIG.get("BOOK", BOOK)
                subject = MMDS_FOLDER_CONFIG.get("SUBJECT", SUBJECT)
                output_xlsx = MMDS_FOLDER_CONFIG.get("OUTPUT", output_xlsx)
            else:
                # Generic folder mode: infer a stable filename and save in CM root.
                inferred = infer_metadata_from_mmd_folder_name(first_arg.name)
                if inferred:
                    board = inferred["board_code"]
                    grade = inferred["grade"]
                    # Keep runtime subject aligned with folder token unless explicitly specialized above.
                    subject = inferred["subject_display"]
                    filename = f"{inferred['board_code']}_{inferred['book_code']}_G{grade:02d}_{inferred['subject_token']}_Concepts.xlsx"
                else:
                    subject_token = re.sub(r"[^A-Za-z0-9]+", "_", subject).strip("_") or "Subject"
                    filename = f"{board}_{book}_G{int(grade):02d}_{subject_token}_Concepts.xlsx"
                output_root = Path(r"C:\Users\FCI\OneDrive\Documents\CM")
                output_xlsx = output_root / filename
        else:
            # Process specific mmd files from command line
            mmd_files = [Path(f) for f in sys.argv[1:]]
            first_mmd = mmd_files[0].resolve()
            # Single file: output same name as input (.xlsx), same folder
            if len(mmd_files) == 1:
                output_xlsx = first_mmd.parent / f"{first_mmd.stem}.xlsx"
            else:
                output_xlsx = first_mmd.parent / "Concepts.xlsx"
            # Use Geography config when file is from mmds_ICSE_MS_G10_Geography or mmds_ICSE_SE_G09_Geography
            if "mmds_ICSE_MS_G10_Geography" in str(first_mmd):
                book = MMDS_FOLDER_CONFIG.get("BOOK", BOOK)
                subject = MMDS_FOLDER_CONFIG.get("SUBJECT", SUBJECT)
            elif "mmds_ICSE_SE_G09_Geography" in str(first_mmd):
                book = "Selina"
                subject = "Geography"
                board = "ICSE"
                grade = 9
            # ICSE SE G10 Biology: separate output in CM folder
            elif "ICSE_SE_G10_Biology" in str(first_mmd):
                book = "Selina"
                subject = "Biology"
                board = "ICSE"
                grade = 10
                output_xlsx = Path(r"C:\Users\FCI\OneDrive\Documents\CM") / f"{first_mmd.stem}_Concepts.xlsx"

    # Resolve output path once so all saves (checkpoint + final) go to the same file
    if output_xlsx:
        output_xlsx = make_unique_output_path(output_xlsx)
    
    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Concepts"
    ws.freeze_panes = "A2"

    headers = [
        "Board", "Book", "Grade", "Subject",
        "Chapter No", "Chapter Code", "Chapter Title",
        "Topic", "Parent Concept", "Concept", "Concept Description",
        "Concept ID", "MMD Path", "PDF Path"
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    total_rows = 0

    if mmd_files:
        # Process specific mmd files (or all from folder); save after each chapter when using folder
        save_after_each = isinstance(sys.argv[1] if len(sys.argv) > 1 else "", str) and Path(sys.argv[1]).is_dir()
        for mmd_path in mmd_files:
            rows = process_mmd_file(mmd_path, ws, subject, board=board, book=book, grade=grade)
            total_rows += rows
            if save_after_each and output_xlsx:
                output_xlsx.parent.mkdir(parents=True, exist_ok=True)
                wb.save(output_xlsx)
                print(f"  [Checkpoint] Saved {total_rows} rows to {output_xlsx.name}")
    else:
        # Original directory-based processing
        chapter_dirs = find_chapter_dirs(BASE_DIR, FOLDER_PREFIX)
        if not chapter_dirs:
            raise RuntimeError(f"No chapter folders found in {BASE_DIR} with prefix {FOLDER_PREFIX}")

        for ch_dir in chapter_dirs:
            chapter_code = ch_dir.name  # folder name becomes chapter_code
            ch_no = chapter_no_from_dirname(ch_dir.name)

            mmd_path = pick_first_file(ch_dir, ".mmd")
            pdf_path = pick_first_file(ch_dir, ".pdf")

            if not mmd_path:
                print(f"[SKIP] {chapter_code}: .mmd not found")
                continue

            mmd_text = safe_read_text(mmd_path)
            mmd_text = trim_mmd_text(mmd_text, MAX_MMD_CHARS)
            chptr_stem_from_file = get_chapter_name_for_code(mmd_path.name)
            chapter_title = chapter_name_stem_to_title(chptr_stem_from_file) if chptr_stem_from_file else infer_chapter_title(mmd_text, fallback=chapter_code)

            chapter_label = f"{chapter_code} | {chapter_title}"
            print(f"[Extracting] {chapter_label} (Subject: {SUBJECT})")

            concepts = gpt_extract_concepts(chapter_label, mmd_text, SUBJECT)
            concepts = ensure_culmination_per_topic(concepts)
            concepts = sanitize_mmd_references(concepts)
            _board_code = board_code_from_board(BOARD)
            chptr_stem = chptr_stem_from_file or chapter_title
            chptr_code = normalize_for_code(chptr_stem) or "Chapter"
            topic_to_num = {}
            next_topic_num = 1
            for i, r in enumerate(concepts, start=1):
                concept_id = make_concept_id(chapter_code, i)
                topic_name = r["topic"].strip()
                concept_name = r["concept"].strip()
                if topic_name not in topic_to_num:
                    topic_to_num[topic_name] = next_topic_num
                    next_topic_num += 1
                topic_num = topic_to_num[topic_name]
                tpc_code = normalize_for_code(topic_name) or f"Topic_{topic_num:02d}"
                topic_display = format_topic_display(GRADE, _board_code, chptr_code, topic_num, topic_name)
                concept_display = format_concept_display(GRADE, _board_code, chptr_code, tpc_code, concept_name)
                chapter_display = chapter_title if chptr_stem_from_file else chapter_code
                ws.append([
                    BOARD, BOOK, GRADE, SUBJECT,
                    ch_no, sanitize_for_excel(chapter_display), sanitize_for_excel(chapter_title),
                    sanitize_for_excel(topic_display),
                    sanitize_for_excel(r["parent_concept"].strip()),
                    sanitize_for_excel(concept_display),
                    sanitize_for_excel(r["concept_description"].strip()),
                    concept_id,
                    str(mmd_path),
                    str(pdf_path) if pdf_path else ""
                ])
                total_rows += 1

            # Wrap text for Concept Description column for the rows just added
            # (Column 11 is "Concept Description")
            for row_idx in range(ws.max_row - len(concepts) + 1, ws.max_row + 1):
                ws.cell(row=row_idx, column=11).alignment = Alignment(wrap_text=True, vertical="top")

    autosize_columns(ws)
    
    # Create output directory if it doesn't exist
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"[DONE] Rows written: {total_rows}")
    print(f"[DONE] Saved Excel: {output_xlsx}")


if __name__ == "__main__":
    """
    Requirements:
      pip install openai openpyxl

    Environment:
      setx OPENAI_API_KEY "your_key_here"

    Run:
      # Process specific mmd files:
      python mmd_to_concepts_excel.py "path/to/file1.mmd" "path/to/file2.mmd"
      
      # Or use original directory-based approach:
      python mmd_to_concepts_excel.py
    """
    main()
