"""
Concept Mapping to Pre-Learning Tool

Reads Concept Mapping Excel files (from mmd_to_concepts_excel.py) and creates
minimal pre-learning Topics, Concepts, and Concept Descriptions - chapter wise,
subject wise.

Pre-Learning Guidelines:
- Dependency-based prerequisites: 4-6 topics per chapter; 5-7 concepts per topic.
- Concept Description matches mmd_to_concepts_excel style: Description // Types // Misconception.
- Cognitive tag per concept: FL | NU | VC | RS | GR.
- Topic / Parent Concept / Concept ID format as specified in the curriculum prompt.
- Topic naming: mmd_to_concepts_excel format (Topic 01: Name (10ICBI_CHAPTER_PrL))
- Concept naming: Concept Name (Subject_PrL)

Post-processing: Highlights concepts/descriptions that are 95%+ similar.
Yellow highlighting in Excel = rows to review for consolidation.

When a sheet has multiple **Subject** values in the same chapter (e.g. Mathematics (General)
vs Mathematics (Commerce)), the tool runs **one pre-learning generation per stream** so each
stream’s mapping rows are sent to the model in full (blank Subject rows are included in every
stream’s batch).

Pipeline: **[1] Pre-learning generator** (dependency prompt with syllabus filter rules) →
**[2] Syllabus boundary filter** (second API pass; prior-grade only; optional `--syllabus-outline`).

Usage:
  python concept_mapping_to_prelearning.py
  python concept_mapping_to_prelearning.py --chapter 1 --subject Biology
  python concept_mapping_to_prelearning.py --all-chapters
  python concept_mapping_to_prelearning.py --board CBSE --input-dir "...\\Grade 10\\CBSE" --all-chapters
  python concept_mapping_to_prelearning.py --grade 9 --board CBSE --input-dir "...\\Grade 09\\CBSE" --all-chapters
  python concept_mapping_to_prelearning.py --input-file "path\\to\\Concepts.xlsx" --all-chapters
  python concept_mapping_to_prelearning.py --input-file "path\\to\\Concepts.xlsx" --sheet Mathematics --sheet Physics

  For workbooks where Chapter No is not 1, use --all-chapters (default --chapter 1 only selects chapter 1).

Requirements: pip install pandas openpyxl openai
"""

import argparse
import difflib
import json
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openai import OpenAI
    from openai import __version__ as openai_version
except ImportError:
    OpenAI = None

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = Workbook = None


# ============================================================
# CONFIG
# ============================================================
DEFAULT_INPUT_DIR = Path(r"C:\Users\FCI\OneDrive\Documents\CM\Concept Mapping Version 01\Grade 10\ICSE")
DEFAULT_OUTPUT_DIR = Path(
    r"G:\My Drive\Concept Mapping Version 01\Concept Mapping Pre"
)

# Expected Concept Mapping workbook filenames per grade + board (inside --input-dir).
SUBJECT_FILES_BY_GRADE_AND_BOARD: Dict[int, Dict[str, Dict[str, str]]] = {
    10: {
        "ICSE": {
            "Biology": "ICSE_SE_G10_Biology_Concepts.xlsx",
            "Chemistry": "ICSE_SE_G10_Chemistry_Concepts.xlsx",
            "Mathematics": "ICSE_SE_G10_Mathematics_Concepts.xlsx",
            "Physics": "ICSE_SE_G10_Physics_Concepts.xlsx",
        },
        "CBSE": {
            "Biology": "CBSE_Schand_G10_Biology_Concepts.xlsx",
            "Chemistry": "CBSE_Schand_G10_Chemistry_Concepts.xlsx",
            "Mathematics": "CBSE_RD_G10_Mathematics_Concepts.xlsx",
            "Physics": "CBSE_Schand_G10_Physics_Concepts.xlsx",
        },
    },
    9: {
        "ICSE": {
            "Biology": "ICSE_SE_G09_Biology_Concepts.xlsx",
            "Chemistry": "ICSE_SE_G09_Chemistry_Concepts.xlsx",
            "Mathematics": "ICSE_SE_G09_Mathematics_Concepts.xlsx",
            "Physics": "ICSE_SE_G09_Physics_Concepts.xlsx",
        },
        "CBSE": {
            "Biology": "CBSE_Schand_G09_Biology_Concepts.xlsx",
            "Chemistry": "CBSE_Schand_G09_Chemistry_Concepts.xlsx",
            "Mathematics": "CBSE_RD_G09_Mathematics_Concepts.xlsx",
            "Physics": "CBSE_Schand_G09_Physics_Concepts.xlsx",
        },
    },
}

# Backward-compatible alias (grade 10 ICSE)
SUBJECT_FILES = SUBJECT_FILES_BY_GRADE_AND_BOARD[10]["ICSE"]

# Alias for board-only lookups at default grade (CLI uses grade + board)
SUBJECT_FILES_BY_BOARD = SUBJECT_FILES_BY_GRADE_AND_BOARD[10]
MODEL = "gpt-5.4-mini-2026-03-17"


def board_sequencing_guidance(board: str) -> str:
    """
    Tell the model how to interpret 'previous grades', 'later chapters', and prior exposure.
    CBSE: anchor to NCERT/CBSE official progression (not ICSE or arbitrary publisher order).
    """
    b = (board or "").strip().upper()
    if "CBSE" in b:
        return (
            "## BOARD-SPECIFIC CURRICULUM (MANDATORY)\n"
            "This run is **CBSE-aligned**. When you research or judge **previous years’ portions**, "
            "**what is new in the current grade**, or **what belongs in later chapters**, use "
            "**official CBSE progression with the NCERT textbook line (Classes 6–10) for this subject** "
            "as the primary reference.\n"
            "- Use **NCERT chapter and grade sequencing** (not ICSE, not another country’s curriculum).\n"
            "- If the publisher workbook’s chapter numbers or order **differ from NCERT**, still decide "
            "prerequisites using **where CBSE/NCERT places ideas across grades and chapters**.\n"
            "- \"Later in the same course\" means **later in the CBSE/NCERT progression** for this subject "
            "at this grade, not merely a higher chapter index in a non-NCERT book."
        )
    if "ICSE" in b:
        return (
            "## BOARD-SPECIFIC CURRICULUM (MANDATORY)\n"
            "This run is **ICSE-aligned**. When judging **prior grades** vs **current grade** and "
            "**chapter order**, use **typical official ICSE syllabus progression** for this subject and grade.\n"
            "- Do **not** replace ICSE sequencing with NCERT/CBSE chapter order.\n"
        )
    if "MSBSHSE" in b or "MAHARASHTRA" in b:
        return (
            "## BOARD-SPECIFIC CURRICULUM (MANDATORY)\n"
            "This run is **Maharashtra State Board (MSBSHSE)**-aligned. When judging **prior grades**, "
            "**what is new in the current grade**, and **chapter order**, use **official MSBSHSE / state "
            "textbook progression** for this subject (not NCERT-only sequencing unless the book follows it).\n"
        )
    return (
        "## BOARD-SPECIFIC CURRICULUM\n"
        f"Board: {board!r}. Use the **official progression for that board** when judging prior vs "
        "current-grade content and chapter order."
    )


def prelearning_output_prefix(
    input_fname: str, subject: str, board_preset: str, grade: int = 10
) -> str:
    """
    Base name for output workbooks (before _PreLearning_*.xlsx).
    ICSE: ICSE_SE_G{NN}_{Subject}. CBSE: workbook stem without _Concepts.
    """
    if board_preset == "ICSE":
        return f"ICSE_SE_G{grade:02d}_{subject}"
    stem = Path(input_fname).stem
    if stem.endswith("_Concepts"):
        stem = stem[: -len("_Concepts")]
    return stem
# Pre-learning structure (dependency architecture prompt)
MIN_TOPICS = 4
MAX_TOPICS = 6
MIN_CONCEPTS_PER_TOPIC = 5
MAX_CONCEPTS_PER_TOPIC = 7
ALLOWED_TAGS = frozenset({"FL", "NU", "VC", "RS", "GR"})
DESC_CHARS_PER_CONCEPT = 350  # full chapter mapping; truncate each row for token budget
SIMILARITY_THRESHOLD = 0.95
MAX_RETRIES = 3
RETRY_SLEEP = 2

# Highlight color for 95%+ similar pairs
HIGHLIGHT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def detect_board_book_grade(
    df: Any,
    default_board: str = "ICSE",
    default_book: str = "Selina",
    default_grade: int = 10,
) -> Tuple[str, str, int]:
    """Read Board, Book, Grade from the Concept Mapping Excel (row 1) when columns exist."""
    board, book, grade = default_board, default_book, int(default_grade)
    if df is None or not len(df):
        return board, book, grade
    if "Board" in df.columns:
        v = str(df["Board"].iloc[0]).strip()
        if v:
            board = v
    if "Book" in df.columns:
        v = str(df["Book"].iloc[0]).strip()
        if v:
            book = v
    if "Grade" in df.columns:
        try:
            g = df["Grade"].iloc[0]
            if g is not None and str(g).strip() != "":
                grade = int(float(g))
        except (TypeError, ValueError):
            pass
    return board, book, grade


# ============================================================
# HELPERS
# ============================================================
def normalize_for_code(s: str) -> str:
    if not s or not isinstance(s, str):
        return ""
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(s).strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def subject_short_name(subject: str) -> str:
    """Full subject name for Concept display: Biology_PrL, Chemistry_PrL, etc."""
    m = {
        "Biology": "Biology",
        "Chemistry": "Chemistry",
        "Mathematics": "Mathematics",
        "Physics": "Physics",
        "Psychology": "Psychology",
    }
    s = str(subject or "").strip()
    if "Mathematics" in s:
        return "Mathematics"
    return m.get(s, s)


def subject_code(subject: str) -> str:
    """Short code for Topic display: BI, CH, MA, PH (mmd_to_concepts_excel format)."""
    s = str(subject or "").strip()
    m = {
        "Biology": "BI",
        "Chemistry": "CH",
        "Mathematics": "MA",
        "Physics": "PH",
        "Psychology": "PY",
    }
    if "Mathematics" in s:
        return "MA"
    return m.get(s, "XX")


def board_code_from_board(board: str) -> str:
    b = (board or "").upper()
    if "ICSE" in b:
        return "IC"
    if "CBSE" in b:
        return "CB"
    if "MSBSHSE" in b or "MAHARASHTRA" in b:
        return "MS"
    return "CB"


def format_concept_prl(concept_name: str, subject: str) -> str:
    """Format: Concept Name (Subject_PrL) - e.g. Cell Definition (Biology_PrL)"""
    subj = subject_short_name(subject)
    return f"{concept_name} ({subj}_PrL)"


def format_topic_prl(
    topic_name: str,
    subject: str,
    grade: int,
    board: str,
    chapter_code: str,
    topic_num: int,
) -> str:
    """Topic format from mmd_to_concepts_excel: Topic 01: TopicName (10ICBI_CELL_PrL)."""
    board_code = board_code_from_board(board)
    subj_code = subject_code(subject)
    prefix = f"{grade:02d}{board_code}{subj_code}_{chapter_code}_PrL"
    return f"Topic {topic_num:02d}: {topic_name} ({prefix})"


def sanitize_for_excel(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    return "".join(c if ord(c) >= 32 or c in "\t\n\r" else "" for c in s)


def enforce_description_format(desc: str, subject: str = "") -> str:
    """
    Align with mmd_to_concepts_excel / excel_to_concepts_prelearning:
    Description // Types // Misconception as newline-separated sections in Excel.
    """
    desc = str(desc or "")
    desc = re.sub(
        r"\n\s*(?=(?:Definition|Description|Usage|Misconception|Types|Examples):)",
        " // ",
        desc,
    )
    if " | " in desc:
        desc = desc.replace(" | ", " // ")
    desc = re.sub(r"\n\s*\n+", " // ", desc)
    desc = re.sub(r"\s*//\s*", "\n", desc)
    desc = desc.strip()
    if desc.startswith("\n"):
        desc = desc[1:].strip()
    return desc


def chapter_code_for_ids(chapter_title: str) -> str:
    """Short code for Concept IDs, e.g. CELL, QUAD (uppercase, alphanumeric)."""
    raw = normalize_for_code(chapter_title)
    code = re.sub(r"[^A-Za-z0-9]", "", raw).upper()
    if not code:
        code = "CH"
    return code[:12]


def assign_structured_concept_ids(chapter_title: str, items: List[Dict[str, Any]]) -> None:
    """Set concept_id on each item: [CHAPTER_CODE]-PL-PC#-C##."""
    base = chapter_code_for_ids(chapter_title)
    parents_order: List[str] = []
    seen = set()
    for it in items:
        p = str(it.get("parent_concept") or "Prerequisites").strip() or "Prerequisites"
        if p not in seen:
            seen.add(p)
            parents_order.append(p)
    pc_map = {p: i + 1 for i, p in enumerate(parents_order)}
    counters: Dict[str, int] = defaultdict(int)
    for it in items:
        p = str(it.get("parent_concept") or "Prerequisites").strip() or "Prerequisites"
        pc = pc_map.get(p, 1)
        counters[p] += 1
        cnum = counters[p]
        it["concept_id"] = f"{base}-PL-PC{pc}-C{cnum:02d}"


# ============================================================
# READ CONCEPT MAPPING
# ============================================================
def read_concept_mapping_excel(path: Path, sheet_name: str = "Concepts") -> pd.DataFrame:
    if pd is None:
        raise ImportError("pandas is required. pip install pandas openpyxl")
    df = pd.read_excel(path, sheet_name=sheet_name)
    return df


def _excel_cell_str(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def _mask_chapter_no(df: pd.DataFrame, chapter_no: int) -> Any:
    ch_no = int(chapter_no)
    return df["Chapter No"].apply(lambda x: int(x) == ch_no if pd.notna(x) else False)


def _filtered_chapter_df(
    df: pd.DataFrame,
    chapter_no: int,
    subject_stream: Optional[str] = None,
) -> Any:
    """Rows for one chapter; if subject_stream is set, keep that Subject plus blank Subject rows."""
    ch_df = df[_mask_chapter_no(df, int(chapter_no))]
    if (
        subject_stream is not None
        and str(subject_stream).strip() != ""
        and "Subject" in df.columns
    ):
        stream = str(subject_stream).strip()

        def ok_row(r: Any) -> bool:
            v = r["Subject"]
            if pd.isna(v) or str(v).strip() == "":
                return True
            return str(v).strip() == stream

        ch_df = ch_df[ch_df.apply(ok_row, axis=1)]
    return ch_df


def distinct_subject_streams_for_chapter(df: pd.DataFrame, chapter_no: int) -> List[Optional[str]]:
    """
    One run per stream when the same chapter lists multiple Subject values
    (e.g. Mathematics (Commerce) vs Mathematics (General)). Rows with blank
    Subject are included in every stream's batch when multiple streams exist.
    """
    if "Subject" not in df.columns:
        return [None]
    base = df[_mask_chapter_no(df, chapter_no)]
    seen: List[str] = []
    for v in base["Subject"]:
        if pd.isna(v) or str(v).strip() == "":
            continue
        s = str(v).strip()
        if s not in seen:
            seen.append(s)
    if len(seen) <= 1:
        return [seen[0]] if seen else [None]
    return seen


def get_chapter_concepts(
    df: pd.DataFrame,
    chapter_no: int,
    subject_stream: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Get concepts for a chapter; optionally only rows matching a Subject stream (+ blank Subject rows)."""
    ch_df = _filtered_chapter_df(df, chapter_no, subject_stream)
    rows = []
    for _, r in ch_df.iterrows():
        parent = _excel_cell_str(r.get("Parent Concept", ""))
        if not parent:
            parent = _excel_cell_str(r.get("Display Name", r.get("Unnamed: 9", "")))
        rows.append({
            "topic": r.get("Topic", ""),
            "parent_concept": parent,
            "concept": r.get("Concept", ""),
            "concept_description": r.get("Concept Description", ""),
        })
    return rows


# ============================================================
# AI: DERIVE PRE-LEARNING
# ============================================================
DEPENDENCY_PRELEARNING_PROMPT = """
You are an expert curriculum designer specializing in dependency-based learning architecture aligned with formal school syllabi (**ICSE/CBSE** and equivalent Indian boards).

Your task is to generate **Pre-Learning Concepts** for a given chapter.

## OBJECTIVE
Generate ONLY those concepts that:

1. Are **strict prerequisites** for understanding the chapter
2. Belong to **previous grade levels OR foundational knowledge already expected** before this grade
3. Can **reasonably** be assumed to have been **taught or encountered before** this chapter (not first introduced in this grade’s syllabus for this subject)

Pre-learning concepts are still:
- Foundational knowledge, prior skills, cognitive building blocks

They are NOT:
- Chapter content, simplified re-teaching of the chapter, topic introductions

## CRITICAL SYLLABUS FILTER (MANDATORY) — apply when generating
Before **including** any concept, validate:

❓ Is this concept **explicitly taught as new** in the **CURRENT grade** syllabus (same subject, typical ICSE/CBSE progression)?
- If **YES** → ❌ **REJECT** (do not output as pre-learning mastery)

❓ Is this concept **typically introduced in this chapter or later chapters** of the same course?
- If **YES** → ❌ **REJECT**
- For **CBSE**, interpret “same course” / chapter order using **NCERT/CBSE** progression for the subject (see BOARD-SPECIFIC CURRICULUM in the run context).

**Only include** concepts that are:
- From **previous grades**, OR
- **Foundational** (e.g. basic arithmetic, basic algebra, general science literacy, reading graphs at an earlier level)

## INPUT CONTEXT (IMPORTANT)
You will receive in the user message:
- **Subject**, **Grade**, **Chapter** name/number
- **Topics** — full Concept Mapping rows for that chapter (use to infer dependencies and to **exclude** in-chapter teaching)
- **Optional** — short **syllabus outline** for the **current grade** (same subject); use it to tighten what counts as “new in this grade”

Use this for **strict boundary filtering**.

## STRICT EXCLUSIONS
DO NOT include:
- "Introduction to…", "Definition of…", "Overview of…", "Examples of…"
- Any concept that is directly taught inside the chapter
If a concept belongs to the chapter itself → REMOVE IT

## INCLUSION RULE (MANDATORY TEST)
Each concept must pass: "If a student does NOT know this, will they struggle to understand the chapter even after teaching?"
- If YES → include
- If NO → exclude

## CONCEPT DESIGN RULES
1. Each concept must be **atomic but meaningful** (not an entire topic; not a trivial single-definition fragment unless it is the meaningful unit).
2. Each concept must represent a **skill**, OR a **relationship**, OR a **reasoning structure**.
3. Avoid fragmentation: do NOT split into definition, formula, example separately; combine into a meaningful unit.

## NAMING RULES (VERY IMPORTANT)
Each concept_name must follow one of these patterns when possible:
- "Relationship Between ___ and ___"
- "Application of ___ in ___ Contexts"
- "Interpretation of ___ in Mathematical/Scientific Situations"
- "Quantitative Handling of ___"
- "Structural Understanding of ___"
- "Transformation and Manipulation of ___"
DO NOT use dull names like: Types of ___, Definition of ___, Basics of ___

## COGNITIVE TAGGING (MANDATORY)
Assign ONE primary tag per concept: FL | NU | VC | RS | GR
- FL → Foundational Logic
- NU → Numerical Handling
- VC → Vocabulary Concept
- RS → Real-world Sense
- GR → Graphical Reasoning

## TOPIC AND CONCEPT COUNTS (STRICT)
- Topics per chapter: **{MIN_T} to {MAX_T}** (inclusive).
- Concepts per topic: **{MIN_CT} to {MAX_CT}** (inclusive), every topic must satisfy this.

## CONCEPT DESCRIPTION FORMAT (MANDATORY — matches mmd_to_concepts_excel)
Each concept_description must be a **single string** with **exactly three sections**, in order, separated by ` // ` (space-slash-slash-space):

**Description:** What the student should already know before the chapter (2–4 short lines). Must align with **concept_name** and must NOT teach the chapter itself.

**Types:** Must be **specific** and mirror the main Concept Mapping style. Use **at least two** numbered types and **concrete examples/cases** grounded in **this** concept’s Description (not generic):
- Start with **Type 01:** then a short type title (what kind of prerequisite check / skill variant).
  - Under it use **Case 01:** … **Case 02:** … (each case = a brief example or instance that illustrates that type—tied to the idea/description above).
- Then **Type 02:** with another type title.
  - Again **Case 01:** … **Case 02:** … (examples specific to Type 02).
- If the prerequisite naturally supports a third variant, add **Type 03:** with **Case 01:** / **Case 02:** as needed.
- Use zero-padded labels exactly: `Type 01:`, `Type 02:`, `Case 01:`, `Case 02:` (no skipping numbers within a type).
- Do **not** write vague Types like "Type 01: General" without cases; every type must include named cases with examples.

**Misconception:** Typical precursor gaps or wrong prior ideas students have (or **N/A** if not applicable).

Use the labels exactly: "Description:", "Types:", "Misconception:" before each section.

## OUTPUT (STRICT — JSON ONLY)
Return a JSON object with one key: **"topics"** (array).

Each topic object must have:
- **topic_name**: short pre-learning topic label (groups dependencies; NOT an "Introduction to…" to the chapter)
- **concepts**: array of **{MIN_CT}–{MAX_CT}** objects, each with:
  - **parent_concept**: umbrella prerequisite category for this row
  - **concept_name**: follows naming rules above
  - **concept_description**: single string with Description: ... // Types: ... // Misconception: ...
  - **tag**: one of FL, NU, VC, RS, GR

There must be **{MIN_T}–{MAX_T}** topics in the array. Order topics and concepts by dependency (earlier prerequisites first). No duplicate or overlapping ideas.

## FINAL VALIDATION STEP (VERY IMPORTANT)
Before outputting, for **EACH** concept ask:

**“Was this already expected knowledge *before* this grade (or clearly foundational)?”**

- If **unsure** → **REMOVE or replace** with a safer prior-grade prerequisite
- If **borderline** → **REMOVE or replace**

Only keep concepts with **HIGH certainty** of **prior exposure** (earlier grades or universal foundations), and that pass the **CRITICAL SYLLABUS FILTER** above.

## FINAL CHECK BEFORE OUTPUT
Remove weak concepts; merge fragmented ones; ensure every concept is truly a prerequisite, **not** chapter content, and **not** a current-grade-only learning objective.

Return ONLY valid JSON. No markdown. No table. No explanation.
"""


def get_prelearning_system_prompt(subject: str, grade: int = 10, board: str = "ICSE") -> str:
    base = DEPENDENCY_PRELEARNING_PROMPT.format(
        MIN_T=MIN_TOPICS,
        MAX_T=MAX_TOPICS,
        MIN_CT=MIN_CONCEPTS_PER_TOPIC,
        MAX_CT=MAX_CONCEPTS_PER_TOPIC,
    )
    return (
        base
        + f"\n\n## RUN CONTEXT\nSubject: {subject}\nGrade: {grade}\nBoard: {board}\n\n"
        + board_sequencing_guidance(board)
        + "\n\n"
        + "You will receive board, chapter, Topics (Concept Mapping), and optionally a **current-grade syllabus outline**. "
        + "Use them to infer dependencies, exclude in-chapter teaching, and enforce **CRITICAL SYLLABUS FILTER**."
    )


SYLLABUS_BOUNDARY_FILTER_SYSTEM = """You are a strict curriculum auditor for **ICSE/CBSE**-aligned pre-learning.

## YOUR JOB (STAGE 2 — after pre-learning generation)
You receive **draft** pre-learning JSON (`topics` with nested `concepts`) plus chapter context.

**Remove or REPLACE** any concept that:
1. Is **explicitly taught as new** in the **current grade** for this subject (typical official progression for the board) → not acceptable as “already known” pre-learning.
2. Is **typically introduced in this chapter or later** in the same course → reject.
   - **CBSE:** “later” follows **NCERT/CBSE** chapter and grade sequencing for this subject, not ICSE ordering.
3. Fails: *“Was this already expected knowledge **before** this grade (or clearly foundational)?”* — if **unsure** or **borderline** → **REPLACE**.

**ALLOW:** previous-grade ideas, or **foundational** skills (basic arithmetic, basic algebra, early science literacy, etc.).

## STRUCTURE (MANDATORY)
- Output **exactly the same number of topics** as the draft.
- Each topic must keep **exactly the same number of concepts** as in the draft for that topic.
- **Do not** delete slots: **substitute** rejected concepts with **new** rows that pass the filter and keep:
  - `parent_concept`, `concept_name` (pattern names), `concept_description` (Description: // Types: // Misconception: with Type 01/02 and ≥2× Case 01/02 each), `tag` (FL|NU|VC|RS|GR).

Return **ONLY** JSON with one key `"topics"` (array) — same schema as input. No markdown, no commentary.
"""


def _chapter_topic_lines_summary(chapter_concepts: List[Dict[str, Any]], max_topics: int = 60) -> str:
    """Deduped Topic column values from Concept Mapping (for syllabus boundary context)."""
    seen: List[str] = []
    for c in chapter_concepts:
        t = str(c.get("topic", "")).strip()
        if t and t not in seen:
            seen.append(t)
    return "\n".join(f"- {s}" for s in seen[:max_topics])


def syllabus_boundary_filter(
    client: Any,
    topics: List[Dict[str, Any]],
    chapter_title: str,
    subject: str,
    grade: int,
    board: str,
    chapter_no: int,
    chapter_topics_summary: str,
    syllabus_outline: Optional[str],
) -> List[Dict[str, Any]]:
    """Stage 2: enforce prior-grade / syllabus boundary; preserve topic and concept counts."""
    try:
        draft = json.dumps({"topics": topics}, ensure_ascii=False)
    except (TypeError, ValueError):
        return topics
    if len(draft) > 100_000:
        draft = draft[:100_000] + "\n... [truncated for length]"

    outline = (
        syllabus_outline.strip()[:16_000]
        if syllabus_outline and syllabus_outline.strip()
        else (
            "(none — for CBSE use NCERT/CBSE grade+chapter progression; for ICSE use typical ICSE progression)"
        )
    )
    user_msg = f"""Board: {board}
{board_sequencing_guidance(board)}

Grade: {grade}
Subject: {subject}
Chapter number: {chapter_no}
Chapter name: {chapter_title}

## Chapter topics (Concept Mapping — boundary reference only)
{chapter_topics_summary[:12_000]}

## Optional current-grade syllabus outline (same subject)
{outline}

## DRAFT pre-learning JSON (return cleaned "topics" only)
{draft}
"""
    resp = client.chat.completions.create(
        model=MODEL,
        messages=[
            {
                "role": "system",
                "content": SYLLABUS_BOUNDARY_FILTER_SYSTEM + "\n\n" + board_sequencing_guidance(board),
            },
            {"role": "user", "content": user_msg},
        ],
        response_format={"type": "json_object"},
    )
    text = resp.choices[0].message.content or "{}"
    data = json.loads(text)
    out = data.get("topics", data.get("Topics", []))
    if not isinstance(out, list) or len(out) != len(topics):
        raise ValueError("filter returned invalid topics list length")
    # verify each topic concept count matches
    for ti, (a, b) in enumerate(zip(topics, out)):
        if not isinstance(b, dict):
            raise ValueError(f"topic {ti} not an object")
        ca = a.get("concepts", []) if isinstance(a, dict) else []
        cb = b.get("concepts", [])
        if not isinstance(cb, list) or len(cb) != len(ca):
            raise ValueError(f"topic {ti} concept count mismatch after filter")
    return out


def _build_full_chapter_context(chapter_concepts: List[Dict[str, Any]]) -> str:
    """Serialize entire chapter mapping for the model (truncate descriptions per row for token limits)."""
    def _safe(s: str, max_len: int) -> str:
        s = str(s or "")[:max_len]
        return s.encode("ascii", "replace").decode("ascii")

    lines = []
    for idx, c in enumerate(chapter_concepts, start=1):
        lines.append(
            f"[{idx}] Topic: {_safe(c.get('topic', ''), 200)} | "
            f"Parent: {_safe(c.get('parent_concept', ''), 120)} | "
            f"Concept: {_safe(c.get('concept', ''), 200)} | "
            f"Description: {_safe(c.get('concept_description', ''), DESC_CHARS_PER_CONCEPT)}"
        )
    return "\n".join(lines)


def _normalize_nested_concept_rows(topics: List[Dict[str, Any]], subject: str) -> None:
    """Normalize tag and concept_description on each nested concept."""
    for t in topics:
        if not isinstance(t, dict):
            continue
        for c in t.get("concepts", []) or []:
            if not isinstance(c, dict):
                continue
            if c.get("tag") is not None:
                c["tag"] = str(c["tag"]).strip().upper()
            if c.get("concept_description"):
                c["concept_description"] = enforce_description_format(
                    str(c["concept_description"]), subject
                )


def _extract_types_block(desc: str) -> str:
    """Portion of concept_description after Types: and before Misconception:."""
    if "Types:" not in desc:
        return ""
    after = desc.split("Types:", 1)[1]
    if "Misconception:" in after:
        after = after.split("Misconception:", 1)[0]
    return after


def _validate_types_structure(types_block: str, topic_idx: int, concept_idx: int) -> Tuple[bool, str]:
    """
    Require structured Types: Type 01:, Type 02:, with Case 01:/Case 02: examples (mmd-style).
    """
    t = types_block.strip()
    if len(t) < 40:
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: Types section too short or empty",
        )
    if not re.search(r"Type\s*01\s*:", t, re.IGNORECASE):
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: Types must include Type 01: ...",
        )
    if not re.search(r"Type\s*02\s*:", t, re.IGNORECASE):
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: Types must include Type 02: ...",
        )
    # At least two Case 01 / two Case 02 so each of Type 01 and Type 02 can carry concrete examples
    n_c1 = len(re.findall(r"Case\s*01\s*:", t, re.IGNORECASE))
    n_c2 = len(re.findall(r"Case\s*02\s*:", t, re.IGNORECASE))
    if n_c1 < 2:
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: Types need Case 01: under "
            f"Type 01 and Type 02 (at least 2× Case 01:)",
        )
    if n_c2 < 2:
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: Types need Case 02: under "
            f"Type 01 and Type 02 (at least 2× Case 02:)",
        )
    return True, ""


def _validate_concept_row(c: Dict[str, Any], topic_idx: int, concept_idx: int) -> Tuple[bool, str]:
    tag = str(c.get("tag", "")).strip().upper()
    if tag not in ALLOWED_TAGS:
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: tag must be one of {sorted(ALLOWED_TAGS)}, "
            f"got {c.get('tag')!r}",
        )
    desc = str(c.get("concept_description", ""))
    if "Description:" not in desc or "Types:" not in desc or "Misconception:" not in desc:
        return (
            False,
            f"topic {topic_idx} concept {concept_idx}: concept_description must contain "
            f"Description:, Types:, and Misconception:",
        )
    types_block = _extract_types_block(desc)
    ok_t, reason_t = _validate_types_structure(types_block, topic_idx, concept_idx)
    if not ok_t:
        return False, reason_t
    if not str(c.get("concept_name", "")).strip():
        return False, f"topic {topic_idx} concept {concept_idx}: missing concept_name"
    if not str(c.get("parent_concept", "")).strip():
        return False, f"topic {topic_idx} concept {concept_idx}: missing parent_concept"
    return True, ""


def _validate_prelearning_topics(topics: List[Dict[str, Any]]) -> Tuple[bool, str]:
    if not isinstance(topics, list):
        return False, "topics is not a list"
    n = len(topics)
    if n < MIN_TOPICS or n > MAX_TOPICS:
        return False, f"topic count {n} must be {MIN_TOPICS}-{MAX_TOPICS}"
    for ti, t in enumerate(topics):
        if not isinstance(t, dict):
            return False, f"topic {ti} is not an object"
        if not str(t.get("topic_name", "")).strip():
            return False, f"topic {ti}: missing topic_name"
        concepts = t.get("concepts", [])
        if not isinstance(concepts, list):
            return False, f"topic {ti}: concepts is not a list"
        m = len(concepts)
        if m < MIN_CONCEPTS_PER_TOPIC or m > MAX_CONCEPTS_PER_TOPIC:
            return (
                False,
                f"topic {ti} ({t.get('topic_name', '')!r}) has {m} concepts; "
                f"need {MIN_CONCEPTS_PER_TOPIC}-{MAX_CONCEPTS_PER_TOPIC}",
            )
        for ci, c in enumerate(concepts):
            if not isinstance(c, dict):
                return False, f"topic {ti} concept {ci} is not an object"
            ok, reason = _validate_concept_row(c, ti, ci)
            if not ok:
                return False, reason
    return True, ""


def derive_prelearning_from_chapter(
    chapter_title: str,
    chapter_concepts: List[Dict[str, Any]],
    subject: str,
    grade: int = 10,
    board: str = "ICSE",
    chapter_no: int = 1,
    syllabus_outline: Optional[str] = None,
    skip_boundary_filter: bool = False,
) -> List[Dict[str, Any]]:
    """Stage [1] generate pre-learning; Stage [2] syllabus boundary filter (unless disabled)."""
    if OpenAI is None:
        raise ImportError("openai is required. pip install openai")

    client = OpenAI()
    concepts_text = _build_full_chapter_context(chapter_concepts)
    n_rows = len(chapter_concepts)
    tags = sorted(ALLOWED_TAGS)
    outline_section = ""
    if syllabus_outline and syllabus_outline.strip():
        outline_section = (
            f"\n## Optional syllabus outline — **current grade** ({grade}, {subject})\n"
            f"Use only to judge what is typically **new in this grade** vs **prior grades**. "
            f"Do not copy outline verbatim as pre-learning.\n"
            f"{syllabus_outline.strip()[:12_000]}\n"
        )

    base_user = f"""Board: {board}
{board_sequencing_guidance(board)}

Chapter number: {chapter_no}
Chapter name: {chapter_title}
Subject (curriculum stream for this batch): {subject}
Grade: {grade}
{outline_section}
Topics within the chapter (Concept Mapping — {n_rows} rows in THIS batch only, analyse ALL of them; use only to infer prerequisites and to EXCLUDE in-chapter teaching). If multiple streams exist in the workbook (e.g. General vs Commerce), this list is filtered to ONE stream — prerequisites must still reflect every row below.

{concepts_text}

Return JSON with key "topics" only (array).
- Exactly {MIN_TOPICS}-{MAX_TOPICS} topics.
- Each topic has topic_name and concepts array with {MIN_CONCEPTS_PER_TOPIC}-{MAX_CONCEPTS_PER_TOPIC} items.
- Each concept: parent_concept, concept_name, tag (one of {", ".join(tags)}),
  concept_description = Description: ... // Types: ... // Misconception: ...
  In Types:, include Type 01: (title) with Case 01:/Case 02: examples, then Type 02: (title) with Case 01:/Case 02:, all specific to that concept.
Allowed tags: {", ".join(tags)}."""

    reminder = (
        f"CRITICAL: Fix your JSON. Requirements: "
        f"{MIN_TOPICS}-{MAX_TOPICS} topics; each topic {MIN_CONCEPTS_PER_TOPIC}-{MAX_CONCEPTS_PER_TOPIC} concepts; "
        f"each concept: parent_concept, concept_name, tag (one of {tags}); "
        f"Types must have Type 01: and Type 02:, each with Case 01: and Case 02: "
        f"(≥2× Case 01: and ≥2× Case 02: total, examples tied to that concept); "
        f"full string: Description: ... // Types: ... // Misconception: ..."
    )

    messages = [
        {"role": "system", "content": get_prelearning_system_prompt(subject, grade, board)},
        {"role": "user", "content": base_user},
    ]

    last_err = None
    out: List[Dict[str, Any]] = []

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=MODEL,
                messages=messages,
                response_format={"type": "json_object"},
            )
            content = resp.choices[0].message.content
            data = json.loads(content) if isinstance(content, str) else json.loads(content or "{}")
            topics = data.get("topics", data.get("Topics", []))
            if not isinstance(topics, list):
                topics = []
            _normalize_nested_concept_rows(topics, subject)
            ok, reason = _validate_prelearning_topics(topics)
            if ok:
                if skip_boundary_filter:
                    return topics
                summary = _chapter_topic_lines_summary(chapter_concepts)
                try:
                    print("    [Pipeline] Stage [2] Syllabus boundary filter…")
                    filtered = syllabus_boundary_filter(
                        client,
                        topics,
                        chapter_title,
                        subject,
                        grade,
                        board,
                        chapter_no,
                        summary,
                        syllabus_outline,
                    )
                    _normalize_nested_concept_rows(filtered, subject)
                    ok_f, reason_f = _validate_prelearning_topics(filtered)
                    if ok_f:
                        return filtered
                    print(
                        f"    [WARN] Boundary filter failed validation ({reason_f}); "
                        f"using stage [1] draft."
                    )
                except Exception as ex:
                    print(f"    [WARN] Boundary filter skipped: {ex}; using stage [1] draft.")
                return topics
            messages.append({"role": "assistant", "content": content or "{}"})
            messages.append({"role": "user", "content": reminder + " Reason: " + reason})
            out = topics
        except Exception as e:
            last_err = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_SLEEP)

    if out:
        total_c = sum(len(t.get("concepts", []) or []) for t in out if isinstance(t, dict))
        print(
            f"    [WARN] Pre-learning validation failed after retries; "
            f"using last response ({len(out)} topics, {total_c} concepts). Check output."
        )
        return out
    raise RuntimeError(f"AI failed after {MAX_RETRIES} retries: {repr(last_err)}")


# ============================================================
# SIMILARITY DETECTION
# ============================================================
def text_similarity(a: str, b: str) -> float:
    """Return similarity ratio 0-1 using SequenceMatcher."""
    a = (a or "").strip().lower()
    b = (b or "").strip().lower()
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def find_similar_pairs(
    items: List[Tuple[int, str, str]],
    threshold: float = SIMILARITY_THRESHOLD,
) -> List[Tuple[int, int, float, float, float]]:
    """
    items: [(row_idx, concept_name, concept_description), ...]
    Returns: [(row_i, row_j, sim_name, sim_desc, max_sim), ...] for pairs >= threshold
    """
    pairs = []
    n = len(items)
    for i in range(n):
        for j in range(i + 1, n):
            _, name_i, desc_i = items[i]
            _, name_j, desc_j = items[j]
            sim_name = text_similarity(name_i, name_j)
            sim_desc = text_similarity(desc_i, desc_j)
            sim = max(sim_name, sim_desc)
            if sim >= threshold:
                pairs.append((i, j, sim_name, sim_desc, sim))
    return pairs


# ============================================================
# WRITE OUTPUT EXCEL
# ============================================================
def write_similarity_report(
    all_rows: List[Dict[str, Any]],
    report_path: Path,
) -> int:
    """Write a text report of 95%+ similar concept pairs. Returns count of pairs."""
    items = [
        (i, r.get("concept", ""), r.get("concept_description", ""))
        for i, r in enumerate(all_rows)
    ]
    pairs = find_similar_pairs(items, SIMILARITY_THRESHOLD)
    if not pairs:
        return 0
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write("Pre-Learning Similarity Report (95%+ similar - review for consolidation)\n")
        f.write("=" * 70 + "\n\n")
        for idx, (i, j, sim_name, sim_desc, sim) in enumerate(pairs, 1):
            row_i, row_j = all_rows[i], all_rows[j]
            f.write(f"Pair {idx}: Similarity {sim:.1%}\n")
            f.write(f"  Row {i+2}: {row_i.get('concept','')[:60]}...\n")
            f.write(f"  Row {j+2}: {row_j.get('concept','')[:60]}...\n")
            f.write(f"  (Name sim: {sim_name:.1%}, Desc sim: {sim_desc:.1%})\n\n")
    return len(pairs)


def write_prelearning_excel(
    all_rows: List[Dict[str, Any]],
    output_path: Path,
    subject: str,
    board: str = "ICSE",
    book: str = "Selina",
    grade: int = 10,
) -> None:
    if Workbook is None:
        raise ImportError("openpyxl is required. pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Learning Concepts"
    ws.freeze_panes = "A2"

    headers = [
        "Board", "Book", "Grade", "Subject",
        "Chapter No", "Chapter Title",
        "Topic", "Parent Concept", "Concept", "Concept Description",
        "Concept ID", "Tag",
    ]
    ws.append(headers)
    desc_col = headers.index("Concept Description") + 1
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    for r in all_rows:
        ws.append([
            sanitize_for_excel(r.get("board", board)),
            sanitize_for_excel(r.get("book", book)),
            r.get("grade", grade),
            sanitize_for_excel(r.get("subject", subject)),
            r.get("chapter_no"),
            sanitize_for_excel(r.get("chapter_title")),
            sanitize_for_excel(r.get("topic")),
            sanitize_for_excel(r.get("parent_concept")),
            sanitize_for_excel(r.get("concept")),
            sanitize_for_excel(r.get("concept_description")),
            sanitize_for_excel(r.get("concept_id")),
            sanitize_for_excel(r.get("tag")),
        ])
        ws.cell(row=ws.max_row, column=desc_col).alignment = Alignment(wrap_text=True, vertical="top")

    # Similarity highlighting
    items = [
        (i + 2, r.get("concept", ""), r.get("concept_description", ""))
        for i, r in enumerate(all_rows)
    ]
    pairs = find_similar_pairs(items, SIMILARITY_THRESHOLD)
    rows_to_highlight = set()
    for i, j, *_ in pairs:
        # i,j are 0-based indices into all_rows; Excel data rows are 2 + index (row 1 = header)
        rows_to_highlight.add(i + 2)
        rows_to_highlight.add(j + 2)

    for row_idx in rows_to_highlight:
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).fill = HIGHLIGHT_FILL

    # Autosize
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            (len(str(c.value or "")) for c in ws[col_letter]),
            default=12,
        )
        ws.column_dimensions[col_letter].width = min(max(12, int(max_len * 0.9)), 60)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================
def process_subject(
    input_path: Path,
    output_path: Path,
    subject: str,
    chapter_no: Optional[int] = None,
    chapter_nos: Optional[List[int]] = None,
    use_ai: bool = True,
    grade: int = 10,
    syllabus_outline: Optional[str] = None,
    skip_boundary_filter: bool = False,
    default_board: str = "ICSE",
    sheet_name: str = "Concepts",
) -> int:
    """Process one subject. Use chapter_nos for several chapters, chapter_no for one, or neither for all."""
    df = read_concept_mapping_excel(input_path, sheet_name=sheet_name)
    board, book, grade_eff = detect_board_book_grade(df, default_board, "Selina", grade)
    chapters_raw = list(df["Chapter No"].dropna().unique())
    if chapter_nos is not None and len(chapter_nos) > 0:
        wanted = {int(x) for x in chapter_nos}
        chapters = [c for c in chapters_raw if int(c) in wanted]
    elif chapter_no is not None:
        chapters = [c for c in chapters_raw if int(c) == int(chapter_no)]
    else:
        chapters = chapters_raw
    chapters = sorted(chapters, key=lambda c: int(c))

    all_rows = []

    for ch_no in chapters:
        streams = distinct_subject_streams_for_chapter(df, int(ch_no))
        for subject_stream in streams:
            subj_label = (
                str(subject_stream).strip()
                if subject_stream is not None and str(subject_stream).strip() != ""
                else subject
            )
            ch_concepts = get_chapter_concepts(df, int(ch_no), subject_stream)
            if not ch_concepts:
                print(
                    f"    [SKIP] Chapter {ch_no} stream {subj_label!r}: no mapping rows"
                )
                continue

            ch_title_df = _filtered_chapter_df(df, int(ch_no), subject_stream)
            ch_title = str(ch_title_df["Chapter Title"].iloc[0]).strip()
            ch_code = normalize_for_code(ch_title) or "Chapter"

            if use_ai:
                topics_result = derive_prelearning_from_chapter(
                    ch_title,
                    ch_concepts,
                    subj_label,
                    grade_eff,
                    board=board,
                    chapter_no=int(ch_no),
                    syllabus_outline=syllabus_outline,
                    skip_boundary_filter=skip_boundary_filter,
                )
            else:
                topics_result = []

            flat_for_ids: List[Dict[str, Any]] = []
            for t in topics_result:
                if not isinstance(t, dict):
                    continue
                for c in t.get("concepts", []) or []:
                    if isinstance(c, dict):
                        flat_for_ids.append(c)

            id_base = f"{ch_title}_{subj_label}"
            assign_structured_concept_ids(id_base, flat_for_ids)

            if topics_result:
                n_conc = len(flat_for_ids)
                stream_note = f" [{subj_label}]" if len(streams) > 1 else ""
                print(
                    f"    Chapter {ch_no}{stream_note}: {len(topics_result)} topics, {n_conc} pre-learning concepts"
                )

            for topic_num, t in enumerate(topics_result, start=1):
                if not isinstance(t, dict):
                    continue
                topic_name = str(t.get("topic_name", "")).strip() or "Prerequisites"
                topic_display = format_topic_prl(
                    topic_name, subj_label, grade_eff, board, ch_code, topic_num
                )
                for it in t.get("concepts", []) or []:
                    if not isinstance(it, dict):
                        continue
                    concept_name = str(it.get("concept_name", "")).strip()
                    if not concept_name:
                        continue
                    concept_display = format_concept_prl(concept_name, subj_label)
                    all_rows.append({
                        "board": board,
                        "book": book,
                        "grade": grade_eff,
                        "subject": subj_label,
                        "chapter_no": int(ch_no),
                        "chapter_title": ch_title,
                        "topic": topic_display,
                        "parent_concept": str(it.get("parent_concept", "")).strip(),
                        "concept": concept_display,
                        "concept_description": str(it.get("concept_description", "")),
                        "concept_id": str(it.get("concept_id", "")),
                        "tag": str(it.get("tag", "")).strip().upper(),
                    })

    if all_rows:
        write_prelearning_excel(all_rows, output_path, subject, grade=grade_eff)
        report_path = output_path.with_suffix(".similarity_report.txt")
        n_pairs = write_similarity_report(all_rows, report_path)
        if n_pairs:
            print(f"    Similarity: {n_pairs} pairs 95%+ similar -> {report_path.name}")

    return len(all_rows)


def main():
    parser = argparse.ArgumentParser(
        description="Create pre-learning from Concept Mapping Excel files"
    )
    parser.add_argument(
        "--input-file",
        type=str,
        default=None,
        help="Path to a single Concept Mapping .xlsx (use with --sheet or omit to process all sheets).",
    )
    parser.add_argument(
        "--input-dir",
        default=str(DEFAULT_INPUT_DIR),
        help="Directory containing Concept Mapping Excel files",
    )
    parser.add_argument(
        "--sheet",
        action="append",
        dest="sheets",
        metavar="NAME",
        help="Sheet name to read (repeatable). Default with --input-file: all sheets in the workbook.",
    )
    parser.add_argument(
        "--grade",
        type=int,
        default=10,
        help="Default grade hint when not present in Excel (directory mode: must be 9 or 10).",
    )
    parser.add_argument(
        "--board",
        choices=list(SUBJECT_FILES_BY_GRADE_AND_BOARD[10].keys()),
        default="ICSE",
        help="Which workbook naming set to use in --input-dir (default: ICSE). CBSE = RD Math + Schand Sci.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory for pre-learning Excel files (default: Google Drive path; with --input-file: same folder as the workbook).",
    )
    parser.add_argument(
        "--subject",
        type=str,
        default=None,
        metavar="SUBJECT",
        help="Process only this subject (directory mode: Biology, Chemistry, Mathematics, Physics). Ignored when using --input-file unless set (overrides per-sheet Subject column).",
    )
    parser.add_argument(
        "--chapter",
        type=int,
        default=1,
        help="Process only this chapter number. Ignored if --chapters is set. Default: 1",
    )
    parser.add_argument(
        "--chapters",
        type=int,
        nargs="+",
        metavar="N",
        help="Process only these chapter numbers, e.g. --chapters 5 6 17 (overrides --chapter)",
    )
    parser.add_argument(
        "--all-chapters",
        action="store_true",
        help="Process all chapters (overrides --chapter; not used with --chapters)",
    )
    parser.add_argument(
        "--no-ai",
        action="store_true",
        help="Skip AI (for testing structure only)",
    )
    parser.add_argument(
        "--syllabus-outline",
        type=str,
        default=None,
        help="Optional path to .txt: current-grade syllabus excerpt (tightens stage [2] filter)",
    )
    parser.add_argument(
        "--no-boundary-filter",
        action="store_true",
        help="Skip stage [2] syllabus boundary filter (draft only)",
    )
    args = parser.parse_args()

    if args.chapters:
        chapter_nos_list: Optional[List[int]] = list(args.chapters)
        chapter_no_single: Optional[int] = None
    elif args.all_chapters:
        chapter_nos_list = None
        chapter_no_single = None
    else:
        chapter_nos_list = None
        chapter_no_single = args.chapter

    input_file_path: Optional[Path] = Path(args.input_file).resolve() if args.input_file else None
    if args.output_dir:
        output_dir = Path(args.output_dir)
    elif input_file_path is not None:
        output_dir = input_file_path.parent
    else:
        output_dir = Path(DEFAULT_OUTPUT_DIR)

    syllabus_outline_text: Optional[str] = None
    if args.syllabus_outline:
        op = Path(args.syllabus_outline)
        if op.exists():
            syllabus_outline_text = op.read_text(encoding="utf-8", errors="replace")
        else:
            print(f"[WARN] --syllabus-outline not found: {op}")

    # --- Single workbook mode (--input-file) ---
    if input_file_path is not None:
        if pd is None:
            print("[ERROR] pandas is required. pip install pandas openpyxl")
            sys.exit(1)
        if not input_file_path.exists():
            print(f"[ERROR] Input file not found: {input_file_path}")
            sys.exit(1)

        xl = pd.ExcelFile(input_file_path)
        if args.sheets:
            sheet_names = [s for s in args.sheets if s in xl.sheet_names]
            missing = set(args.sheets) - set(sheet_names)
            if missing:
                print(f"[WARN] Unknown sheet(s) skipped: {sorted(missing)}")
        else:
            sheet_names = list(xl.sheet_names)

        if not sheet_names:
            print("[ERROR] No sheets to process.")
            sys.exit(1)

        for sheet in sheet_names:
            df_probe = read_concept_mapping_excel(input_file_path, sheet_name=sheet)
            if "Chapter No" not in df_probe.columns or "Concept" not in df_probe.columns:
                print(f"[SKIP] Sheet {sheet!r}: missing Chapter No / Concept columns")
                continue

            subject_eff = args.subject
            if not subject_eff and "Subject" in df_probe.columns and len(df_probe):
                subject_eff = str(df_probe["Subject"].iloc[0]).strip()
            if not subject_eff:
                subject_eff = sheet

            stem = input_file_path.stem
            safe_sheet = normalize_for_code(sheet) or "Sheet"
            if chapter_nos_list:
                ch_label = "_".join(str(n) for n in sorted(set(chapter_nos_list)))
                out_name = f"{stem}_{safe_sheet}_PreLearning_Chapters_{ch_label}.xlsx"
            elif chapter_no_single is None:
                out_name = f"{stem}_{safe_sheet}_PreLearning_All.xlsx"
            else:
                out_name = f"{stem}_{safe_sheet}_PreLearning_Chapter{chapter_no_single}.xlsx"
            output_path = output_dir / out_name

            ch_log = (
                f"chapters {', '.join(str(n) for n in sorted(set(chapter_nos_list)))}"
                if chapter_nos_list
                else ("ALL" if chapter_no_single is None else str(chapter_no_single))
            )
            print(f"\n--- {input_file_path.name} :: sheet {sheet!r} ({ch_log}) ---")
            try:
                count = process_subject(
                    input_file_path,
                    output_path,
                    subject_eff,
                    chapter_no=chapter_no_single,
                    chapter_nos=chapter_nos_list,
                    use_ai=not args.no_ai,
                    grade=args.grade,
                    syllabus_outline=syllabus_outline_text,
                    skip_boundary_filter=args.no_boundary_filter,
                    default_board=args.board,
                    sheet_name=sheet,
                )
                print(f"[DONE] {count} pre-learning concepts -> {output_path}")
            except Exception as e:
                print(f"[ERROR] {e}")
                raise

        print("\n[DONE] Pre-learning generation complete.")
        print("Yellow highlighting = concepts/descriptions 95%+ similar (review for consolidation).")
        return

    # --- Directory mode (original) ---
    input_dir = Path(args.input_dir)
    if not input_dir.exists():
        print(f"[ERROR] Input directory not found: {input_dir}")
        sys.exit(1)

    if args.grade not in SUBJECT_FILES_BY_GRADE_AND_BOARD:
        print(f"[ERROR] Unsupported --grade {args.grade} (directory mode supports 9 and 10).")
        sys.exit(1)

    if args.subject and args.subject not in SUBJECT_FILES:
        print(f"[ERROR] Unknown --subject {args.subject!r}. Expected one of: {sorted(SUBJECT_FILES.keys())}")
        sys.exit(1)

    grade_map = SUBJECT_FILES_BY_GRADE_AND_BOARD[args.grade]
    subjects = [args.subject] if args.subject else list(SUBJECT_FILES.keys())
    subject_files = grade_map[args.board]

    for subject in subjects:
        fname = subject_files.get(subject)
        if not fname:
            continue
        input_path = input_dir / fname
        if not input_path.exists():
            print(f"[SKIP] Not found: {input_path}")
            continue

        out_base = prelearning_output_prefix(fname, subject, args.board, args.grade)
        if chapter_nos_list:
            ch_label = "_".join(str(n) for n in sorted(set(chapter_nos_list)))
            output_path = output_dir / f"{out_base}_PreLearning_Chapters_{ch_label}.xlsx"
            ch_log = f"chapters {', '.join(str(n) for n in sorted(set(chapter_nos_list)))}"
        elif chapter_no_single is None:
            output_path = output_dir / f"{out_base}_PreLearning_All.xlsx"
            ch_log = "ALL"
        else:
            output_path = output_dir / f"{out_base}_PreLearning_Chapter{chapter_no_single}.xlsx"
            ch_log = str(chapter_no_single)
        print(f"\n--- Grade {args.grade} {args.board} {subject} ({ch_log}) ---")
        try:
            count = process_subject(
                input_path,
                output_path,
                subject,
                chapter_no=chapter_no_single,
                chapter_nos=chapter_nos_list,
                use_ai=not args.no_ai,
                grade=args.grade,
                syllabus_outline=syllabus_outline_text,
                skip_boundary_filter=args.no_boundary_filter,
                default_board=args.board,
            )
            print(f"[DONE] {count} pre-learning concepts -> {output_path}")
        except Exception as e:
            print(f"[ERROR] {e}")
            raise

    print("\n[DONE] Pre-learning generation complete.")
    print("Yellow highlighting = concepts/descriptions 95%+ similar (review for consolidation).")


if __name__ == "__main__":
    main()
