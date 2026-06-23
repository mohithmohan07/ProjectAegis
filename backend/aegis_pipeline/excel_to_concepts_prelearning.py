"""
Excel to Concepts - Pre-Learning (Grade IX ICSE)

Reads an Excel file with Math and Physics sheets. Columns: Chapter, Pre Topic, Concept,
Concept Description, Types. Uses gpt-5-mini to enhance each concept into the standard
output format (Description // Types // Misconception). Outputs subject-wise Excel files
matching mmd_to_concepts_excel.py format.

Pre-Learning context: These are fundamentals required before the actual chapter.
"""

import argparse
import os
import re
import json
import time
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openai import OpenAI
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# USER CONFIG
# ============================================================
INPUT_XLSX = Path(r"C:\Users\FCI\OneDrive\Desktop\ICSE\Grade 9 Pre concepts.xlsx")
OUTPUT_DIR = Path(r"C:\Users\FCI\OneDrive\Documents\CM")

BOARD = "ICSE"
BOOK = "Selina"
GRADE = 9
MODEL = "gpt-5.4-mini-2026-03-17"
MAX_OUTPUT_TOKENS = int(os.getenv("AEGIS_OPENAI_MAX_OUTPUT_TOKENS", "128000"))

# Subject code for topic/concept display: MA=Math, PH=Physics
SUBJECT_CODES = {"Math": "MA", "Mathematics": "MA", "Physics": "PH"}

MAX_RETRIES = 3
RETRY_SLEEP_SECONDS = 2


# ============================================================
# STRICT JSON SCHEMA (single concept enhancement)
# ============================================================
CONCEPT_ENHANCE_SCHEMA: Dict[str, Any] = {
    "name": "concept_enhance",
    "strict": True,
    "schema": {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "parent_concept": {"type": "string"},
            "concept_description": {"type": "string"}
        },
        "required": ["parent_concept", "concept_description"]
    }
}


# ============================================================
# OPENAI CLIENT
# ============================================================
client = OpenAI()


# ============================================================
# HELPERS (from mmd_to_concepts_excel.py)
# ============================================================
def normalize_for_code(s: str) -> str:
    """Keep full name, replace spaces/special with underscore."""
    if not s or not isinstance(s, str):
        return ""
    s = re.sub(r"[^A-Za-z0-9]+", "_", s.strip())
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def board_code_from_board(board: str) -> str:
    return "IC" if board and "ICSE" in (board or "").upper() else "CB"


def enforce_description_format(desc: str, subject: str) -> str:
    """Convert // separators to line breaks for Excel display."""
    desc = re.sub(r'\n\s*(?=(?:Definition|Description|Usage|Misconception|Types|Examples):)', ' // ', desc)
    if " | " in desc:
        desc = desc.replace(" | ", " // ")
    desc = re.sub(r'\n\s*\n+', ' // ', desc)
    desc = re.sub(r'\s*//\s*', '\n', desc)
    desc = desc.strip()
    if desc.startswith('\n'):
        desc = desc[1:].strip()
    return desc


def sanitize_for_excel(value: str) -> str:
    if not isinstance(value, str):
        return str(value) if value is not None else ""
    return "".join(
        char if ord(char) >= 32 or char in "\t\n\r" else ""
        for char in value
    )


def extract_chapter_no(chapter_str: str) -> Tuple[Optional[int], str]:
    """Extract chapter number and title from Chapter column."""
    if not chapter_str or not isinstance(chapter_str, str):
        return None, str(chapter_str or "")
    s = str(chapter_str).strip()
    # Try "Chapter 1", "Ch 1", "1. Rational Numbers", etc.
    m = re.search(r"(?:Chapter|Ch\.?)\s*0?(\d+)[:\s]*(.*)", s, re.IGNORECASE)
    if m:
        return int(m.group(1)), (m.group(2) or s).strip()
    m = re.search(r"^0?(\d+)[\.\:\s]+(.+)", s)
    if m:
        return int(m.group(1)), m.group(2).strip()
    m = re.search(r"\b0?(\d+)\b", s)
    if m:
        return int(m.group(1)), s
    return None, s


def format_topic_display(grade: int, board_code: str, subj_code: str, chptr_code: str, topic_num: int, topic_name: str) -> str:
    """Topic 01: TopicName (09ICMA_ChapterName_PL)."""
    prefix = f"{grade:02d}{board_code}{subj_code}_{chptr_code}_PL"
    return f"Topic {topic_num:02d}: {topic_name} ({prefix})"


def format_concept_display(grade: int, board_code: str, subj_code: str, chptr_code: str, tpc_code: str, concept_name: str) -> str:
    """ConceptName (09ICMA_ChapterName_PL_TopicName)."""
    prefix = f"{grade:02d}{board_code}{subj_code}_{chptr_code}_PL_{tpc_code}"
    return f"{concept_name} ({prefix})"


def make_concept_id(chapter_code: str, idx: int) -> str:
    return f"{chapter_code}-C{idx:03d}"


def autosize_columns(ws):
    max_width = 60
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(12, max_len * 0.9), max_width)


# ============================================================
# AI ENHANCEMENT
# ============================================================
def get_prelearning_system_prompt(subject: str) -> str:
    """System prompt for enhancing pre-learning concepts (Grade IX ICSE)."""
    return (
        "You are a STRICT concept mapping assistant for ICSE Grade IX " + subject + " PRE-LEARNING.\n"
        "Pre-Learning = fundamentals students need BEFORE starting the actual chapter.\n"
        "Return ONLY JSON matching the provided schema. No markdown. No commentary.\n\n"

        "TASK: Enhance the given concept into the standard format.\n"
        "Output: parent_concept, concept_description.\n\n"

        "concept_description MUST be ONE string with exactly THREE sections, separated by // in this order:\n"
        "Description: <DETAILED understanding: complete definition, explanation, key points, step-by-step reasoning. "
        "Integrate worked examples within the description. Consider Grade IX ICSE level.> "
        "// Types: <Type 01: Name Case 01: ... Case 02: ... Type 02: Name Case 01: ... "
        "(organised classification of all question/numerical types from the concept)> "
        "// Misconception: <Common misconceptions students have about this concept, or 'N/A' if not applicable>\n\n"

        "IMPORTANT: Use // (double slash) as the separator between sections. Do NOT use newlines.\n\n"

        "RULES:\n"
        "1) parent_concept: A meaningful, reusable parent category for this concept.\n"
        "2) Description: Expand and enrich the given concept description. Add clarity, examples, and rigor.\n"
        "3) Types: Use the provided Types if given; otherwise infer from the concept. Use structure: "
        "Type 01: <Name> Case 01: ... Case 02: ... Type 02: ... Use zero-padded numbers.\n"
        "4) Misconception: Add common student errors if relevant; otherwise write 'N/A'.\n"
        "5) Keep the concept name and topic as provided - do not change them.\n"
        "6) All content must be appropriate for Grade IX ICSE pre-learning (foundational, prerequisite knowledge).\n"
    )


def gpt_enhance_concept(
    chapter: str,
    pre_topic: str,
    concept: str,
    concept_description: str,
    types: str,
    subject: str,
) -> Tuple[str, str]:
    """
    Use gpt-5-mini to enhance a single concept into the standard format.
    Returns (parent_concept, concept_description).
    """
    user = (
        f"CHAPTER (Grade IX ICSE): {chapter}\n"
        f"Pre Topic: {pre_topic}\n"
        f"Concept: {concept}\n"
        f"Concept Description (existing): {concept_description or '(none)'}\n"
        f"Types (existing): {types or '(none)'}\n\n"
        "Enhance into the standard format. Return parent_concept and concept_description "
        "(Description // Types // Misconception)."
    )

    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.responses.create(
                model=MODEL,
                max_output_tokens=MAX_OUTPUT_TOKENS,
                input=[
                    {"role": "system", "content": get_prelearning_system_prompt(subject)},
                    {"role": "user", "content": user},
                ],
                text={
                    "format": {
                        "type": "json_schema",
                        "name": CONCEPT_ENHANCE_SCHEMA["name"],
                        "schema": CONCEPT_ENHANCE_SCHEMA["schema"]
                    }
                }
            )
            data = json.loads(resp.output_text)
            desc = data.get("concept_description", "")
            desc = enforce_description_format(desc, subject)
            return data.get("parent_concept", "").strip(), desc
        except Exception as e:
            last_err = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_SLEEP_SECONDS)
            else:
                raise RuntimeError(f"GPT enhancement failed after {MAX_RETRIES} retries: {repr(last_err)}")


# ============================================================
# MAIN PROCESSING
# ============================================================
def read_input_sheet(wb, sheet_name: str, limit: Optional[int] = None) -> List[Dict[str, Any]]:
    """Read rows from sheet. Expected columns: Chapter, Pre Topic, Concept, Concept Description, Types."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    # Normalize column names (case-insensitive, allow variations)
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower()
        if "chapter" in h_lower:
            col_map["chapter"] = i
        elif "pre" in h_lower and "topic" in h_lower:
            col_map["pre_topic"] = i
        elif h_lower == "concept":
            col_map["concept"] = i
        elif "description" in h_lower or "detail" in h_lower:
            col_map["concept_description"] = i
        elif "types" in h_lower:
            col_map["types"] = i

    if "chapter" not in col_map or "concept" not in col_map:
        print(f"  [WARN] Sheet '{sheet_name}': missing Chapter or Concept column. Found: {headers}")
        return []

    def safe_cell(row, key: str, default_idx: int = 0):
        """Safely get cell value; return empty string if index out of range."""
        if key not in col_map:
            return ""
        idx = col_map[key]
        return row[idx] if idx < len(row) else ""

    result = []
    for row in rows[1:]:
        if not row:
            continue
        chapter = safe_cell(row, "chapter", 0)
        pre_topic = safe_cell(row, "pre_topic", 1)
        concept = safe_cell(row, "concept", 2)
        desc = safe_cell(row, "concept_description", 3)
        types_val = safe_cell(row, "types", 4)
        if not concept and not chapter:
            continue
        result.append({
            "chapter": str(chapter or "").strip(),
            "pre_topic": str(pre_topic or "").strip(),
            "concept": str(concept or "").strip(),
            "concept_description": str(desc or "").strip(),
            "types": str(types_val or "").strip(),
        })
        if limit and len(result) >= limit:
            break
    return result


def process_subject(
    rows: List[Dict[str, Any]],
    subject: str,
    output_path: Path,
) -> int:
    """Process all rows for a subject, enhance with AI, write to Excel."""
    if not rows:
        print(f"  [SKIP] No rows for {subject}")
        return 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Concepts"
    ws.freeze_panes = "A2"
    headers = [
        "Board", "Book", "Grade", "Subject",
        "Chapter No", "Chapter Code", "Chapter Title",
        "Topic", "Parent Concept", "Concept", "Concept Description",
        "Concept ID", "MMD Path", "PDF Path"
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    board_code = board_code_from_board(BOARD)
    subj_code = SUBJECT_CODES.get(subject, "MA" if "math" in subject.lower() else "PH")

    # Group by chapter for numbering
    chapter_to_info: Dict[str, Tuple[int, str]] = {}
    chapter_order: List[str] = []
    for r in rows:
        ch = r["chapter"] or "Unknown"
        if ch not in chapter_to_info:
            ch_no, ch_title = extract_chapter_no(ch)
            ch_code = normalize_for_code(ch_title or ch) or "Chapter"
            chapter_to_info[ch] = (ch_no or len(chapter_order) + 1, ch_title or ch)
            chapter_order.append(ch)

    topic_to_num: Dict[str, int] = {}
    next_topic_num = 1
    concept_idx = 0
    total_rows = 0

    for i, r in enumerate(rows):
        chapter = r["chapter"] or "Unknown"
        pre_topic = r["pre_topic"] or "Pre-Learning"
        concept = r["concept"] or ""
        if not concept:
            continue

        ch_no, ch_title = chapter_to_info.get(chapter, (i + 1, chapter))
        ch_code = normalize_for_code(ch_title or chapter) or "Chapter"

        print(f"  [{i+1}/{len(rows)}] Enhancing: {concept[:50]}...")
        try:
            parent_concept, concept_description = gpt_enhance_concept(
                chapter=chapter,
                pre_topic=pre_topic,
                concept=concept,
                concept_description=r["concept_description"],
                types=r["types"],
                subject=subject,
            )
        except Exception as e:
            print(f"    [ERROR] {e}")
            parent_concept = pre_topic or "Pre-Learning"
            concept_description = (
                f"Description: {r['concept_description'] or concept}\n"
                f"Types: {r['types'] or 'N/A'}\n"
                "Misconception: N/A"
            )
            concept_description = enforce_description_format(concept_description, subject)

        topic_key = f"{chapter}||{pre_topic}"
        if topic_key not in topic_to_num:
            topic_to_num[topic_key] = next_topic_num
            next_topic_num += 1
        topic_num = topic_to_num[topic_key]
        tpc_code = normalize_for_code(pre_topic) or f"Topic_{topic_num:02d}"
        topic_display = format_topic_display(GRADE, board_code, subj_code, ch_code, topic_num, pre_topic)
        concept_display = format_concept_display(GRADE, board_code, subj_code, ch_code, tpc_code, concept)

        concept_idx += 1
        concept_id = make_concept_id(ch_code, concept_idx)

        ws.append([
            BOARD, BOOK, GRADE, subject,
            ch_no, sanitize_for_excel(ch_title or chapter), sanitize_for_excel(ch_title or chapter),
            sanitize_for_excel(topic_display),
            sanitize_for_excel(parent_concept),
            sanitize_for_excel(concept_display),
            sanitize_for_excel(concept_description),
            concept_id,
            "",  # MMD Path
            ""   # PDF Path
        ])
        total_rows += 1

        # Wrap Concept Description
        ws.cell(row=ws.max_row, column=11).alignment = Alignment(wrap_text=True, vertical="top")

    autosize_columns(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return total_rows


def main():
    parser = argparse.ArgumentParser(description="Excel to Concepts - Pre-Learning (Grade IX ICSE)")
    parser.add_argument("input", nargs="?", default=str(INPUT_XLSX), help="Input Excel path")
    parser.add_argument("output_dir", nargs="?", default=str(OUTPUT_DIR), help="Output directory")
    parser.add_argument("--limit", "-n", type=int, help="Limit rows per sheet (e.g. 10 for testing)")
    parser.add_argument("--sheet", "-s", choices=["Math", "Mathematics", "Physics"], help="Process only this sheet")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)

    if not input_path.exists():
        print(f"[ERROR] Input Excel not found: {input_path}")
        sys.exit(1)

    print(f"Input: {input_path}")
    print(f"Output dir: {output_dir}")
    print(f"Model: {MODEL}")
    if args.limit:
        print(f"Limit: {args.limit} rows per sheet")
    if args.sheet:
        print(f"Sheet: {args.sheet} only")
    print()

    wb = load_workbook(input_path, read_only=True, data_only=True)

    sheets = [args.sheet] if args.sheet else ["Math", "Mathematics", "Physics"]
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        subject = "Mathematics" if sheet_name in ["Math", "Mathematics"] else "Physics"
        rows = read_input_sheet(wb, sheet_name, limit=args.limit)
        if not rows:
            continue
        output_name = f"ICSE_SE_G09_{subject}_PreLearning_Concepts.xlsx"
        output_path = output_dir / output_name
        print(f"\n--- {subject} ({len(rows)} rows) -> {output_path.name} ---")
        total = process_subject(rows, subject, output_path)
        print(f"[DONE] {subject}: {total} rows -> {output_path}")

    wb.close()
    print("\n[DONE] All subjects processed.")


if __name__ == "__main__":
    """
    Requirements:
      pip install openai openpyxl

    Environment:
      setx OPENAI_API_KEY "your_key_here"

    Run:
      python excel_to_concepts_prelearning.py
      python excel_to_concepts_prelearning.py --limit 10 --sheet Math
      python excel_to_concepts_prelearning.py "path/to/input.xlsx" "path/to/output_dir"
    """
    main()
