#!/usr/bin/env python3
"""Hand-authored bulk-import workbook for Balbharati Std 6 CH01.

Written manually against the source MMD, not produced by Aegis. Every term used
appears in the book; the audit that motivated this found apex, lateral face,
curved solid, measurement, diameter and polygonal base all absent from it.
"""
import argparse
import pathlib
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

CODE = "06MSMA_ThreeDimensi"
CHAPTER = "Three-Dimensional Shapes (06_Mathematics_Maharashtra)"
CH_DISPLAY = "Three-Dimensional Shapes"
CH_DESC = ("Maharashtra Board Std 6 Mathematics: Three-Dimensional Shapes. "
           "Dimensions, elements of shapes, polygons, prism, cylinder, "
           "pyramids, cone, sphere, and the difference between a prism and "
           "a pyramid.")

# --------------------------------------------------------------------------
# Topics — the book's own printed sections, in printed order.
# "Difference between Prism and Pyramid" is a printed section and gets its own
# topic; Aegis had filed its concepts under Sphere.
# --------------------------------------------------------------------------
TOPICS = [
    ("T01", "Dimensions"),
    ("T02", "Elements of shapes"),
    ("T03", "Two Dimentional Shapes"),      # book's printed spelling, kept
    ("T04", "Three Dimentional Shapes"),    # book's printed spelling, kept
    ("T05", "Pyramid"),
    ("T06", "Cone"),
    ("T07", "Sphere"),
    ("T08", "Difference between Prism and Pyramid"),
]
TOPIC_DESC = {
    "T01": "One dimensional, two dimensional and three dimensional shapes.",
    "T02": "The vertex, side, edge and face of two and three dimensional shapes.",
    "T03": "Polygons and the number of their sides and vertices.",
    "T04": "The prism and the cylinder, formed by stacking equal shapes.",
    "T05": "Pyramids, whose vertical faces are triangular and meet at the top.",
    "T06": "The cone and its elements.",
    "T07": "The sphere, and showing the height of three dimensional shapes.",
    "T08": "How the faces, vertices and edges of a prism differ from a pyramid.",
}

# --------------------------------------------------------------------------
# Concepts — short grade-appropriate noun phrases. Descriptions explain first
# and give the book's example after, in Std 6 language.
# --------------------------------------------------------------------------
C = {}


def concept(key, topic, title, camel, details, mastery, keywords):
    # concept_details is composed later, once the questions exist, into the
    # canonical Aegis string (generation.py "OUTPUT CONTRACT for
    # concept_description" + concept_refiner.py):
    #   Description: ...\nAchieving Mastery: ... // Activity/Info Hub: ...
    #   // Types: Type NN: ... Case NN: ... Example NN: ...
    #   // Misconception/ Error Analysis: Misconceptions: ...; Error Analysis: ...
    C[key] = {
        "topic": topic, "title": title, "camel": camel,
        "description": details, "mastery": mastery, "keywords": keywords,
    }


concept(
    "dim", "T01", "One, Two and Three Dimensional Shapes", "OneTwoThreeDim",
    "A shape is one dimensional when it has only length. A shape is two "
    "dimensional when it occupies space on a flat surface. A shape is three "
    "dimensional when it occupies place in space. To decide which a shape is, "
    "look at whether it stays on a flat surface or takes up space. The book "
    "fills a matchbox with wet soil to make a brick. The brick that comes out "
    "is a cuboid, and it takes up space, so it is three dimensional.",
    "The learner can look at an object and say whether it is one, two or "
    "three dimensional, and give the reason.",
    "one dimensional, two dimensional, three dimensional, length, width, "
    "height, flat surface, space, cuboid",
)
concept(
    "elem", "T02", "Vertex, Side, Edge and Face", "VertexSideEdgeFace",
    "Every shape is described using the same few parts. In a two dimensional "
    "shape, the common point of two sides is a vertex, and a line segment "
    "joining two adjacent vertices is a side. A two dimensional shape has "
    "only one surface. In a three dimensional shape, a common point of three "
    "or more edges is a vertex, and the common side of two surfaces is an "
    "edge. A three dimensional shape is formed by the combination of many "
    "surfaces. The book shows a rectangle with its vertex and side, and a "
    "cuboid with its vertex, edge and face.",
    "The learner can point to the vertex, side, edge and face of a given "
    "shape and say what each one is.",
    "vertex, side, arm, edge, face, surface, rectangle, cuboid, "
    "two dimensional shape, three dimensional shape",
)
concept(
    "poly", "T03", "Polygons", "Polygons",
    "A polygon is a closed figure with three or more sides. To check whether "
    "a figure is a polygon, count its sides and see that the sides join all "
    "the way round so the figure closes. A figure that does not close is not "
    "a polygon. The book's table starts with a triangle, which has 3 vertices "
    "and 3 sides, and goes on to the quadrilateral, pentagon, hexagon and "
    "heptagon.",
    "The learner can decide whether a figure is a polygon and can count its "
    "sides and vertices.",
    "polygon, closed figure, side, vertex, triangle, quadrilateral, pentagon, "
    "hexagon, heptagon",
)
concept(
    "prism", "T04", "Prism", "Prism",
    "A prism is made by stacking equal copies of the same flat shape on top "
    "of each other. The shape that is stacked is the bottom face, and the "
    "prism is named after it. When 10 to 15 equal triangles are stacked, the "
    "shape formed is a triangular prism. It has 6 vertices, 2 triangular "
    "faces, 3 rectangular faces and 9 edges. When equal rectangles are "
    "stacked the shape formed is a cuboid, which is also called a rectangular "
    "prism. Squares, pentagons and hexagons can be stacked in the same way.",
    "The learner can name a prism from the shape that is stacked, and can "
    "count its vertices, faces and edges.",
    "prism, triangular prism, rectangular prism, cuboid, bottom face, "
    "vertical face, vertex, edge, face",
)
concept(
    "cyl", "T04", "Cylinder", "Cylinder",
    "A cylinder is formed by stacking equal circles on top of each other, so "
    "it is also called a circular prism. Because circles are stacked, the "
    "shape has circular faces at the two ends and one curved face going "
    "round. The boundary of each circular face is a circular edge. The book "
    "forms it by cutting 10 to 15 circles of equal radius from thick "
    "cardboard and placing them one on top of the other.",
    "The learner can form a cylinder by stacking circles and can name its "
    "circular faces, curved face and circular edges.",
    "cylinder, circular prism, circular face, curved face, circular edge, "
    "radius, stacking",
)
concept(
    "pyr", "T05", "Pyramids", "Pyramids",
    "A pyramid is a shape whose vertical faces are triangular and meet at the "
    "top. To name a pyramid, look at its bottom face: a triangular bottom "
    "face gives a triangular pyramid and a square bottom face gives a square "
    "pyramid. The book shows a pyramid and a tent as examples. Its triangular "
    "pyramid has 4 vertices, 4 triangular faces and 6 edges.",
    "The learner can name a pyramid from its bottom face and can count its "
    "vertices, triangular faces and edges.",
    "pyramid, triangular pyramid, square pyramid, bottom face, vertical face, "
    "triangular face, vertex, edge, tent",
)
concept(
    "cone", "T06", "Cone", "Cone",
    "A cone is the shape formed when a semicircle of paper is rolled so that "
    "its two radii are joined, and a smaller circle is stuck to the round "
    "boundary. The book makes one by drawing a circle of radius 6 cm, cutting "
    "two semicircles from it, gluing the two radii of a semicircle together, "
    "then cutting a circle of radius 3 cm and sticking it on the "
    "circumference of the shape made. A cone has a vertex, a circular edge, a "
    "circular face and a curved face.",
    "The learner can make a cone from paper and can name its vertex, circular "
    "edge, circular face and curved face.",
    "cone, vertex, circular edge, circular face, curved face, semicircle, "
    "radius, circumference",
)
concept(
    "sph", "T07", "Sphere", "Sphere",
    "A sphere is a shape with a curved surface and no edges and no vertices. "
    "To decide whether a shape is a sphere, look for the curved surface and "
    "check that there is no edge and no vertex anywhere on it. The book gives "
    "the ball and the ladoo as shapes we see in daily life, and says the "
    "shape of a ball is spherical.",
    "The learner can identify a spherical shape by its curved surface and by "
    "the absence of edges and vertices.",
    "sphere, spherical, ball, ladoo, curved surface, no edge, no vertex",
)
concept(
    "hgt", "T07", "Height of Three Dimensional Shapes", "HeightOfThreeDimShapes",
    "Height is shown differently for different three dimensional shapes, so "
    "the shape has "
    "to be identified first. The book asks how the height of a prism, a "
    "pyramid, a cylinder and a cone can be shown, and then asks separately "
    "whether the height of a sphere can be shown. The sphere is asked about "
    "on its own because it has no flat face and no vertex.",
    "The learner can say how height is shown for a prism, pyramid, cylinder "
    "and cone, and can treat the sphere as a separate case.",
    "height, prism, pyramid, cylinder, cone, sphere",
)
concept(
    "diff", "T08", "Prism and Pyramid Compared", "PrismPyramidCompared",
    "A prism and a pyramid are told apart by their faces, vertices and edges. "
    "In a prism the bottom face and the top face are triangular, square or "
    "polygonal and the vertical faces are rectangular; its vertices are on "
    "the bottom and top surfaces; and its edges are three times the number of "
    "sides of the top or bottom surface. In a pyramid the bottom face is "
    "triangular, square or polygonal and the vertical faces are triangular; "
    "it has one vertex at the top and the other vertices at the bottom; and "
    "its edges are twice the number of sides of the bottom surface.",
    "The learner can decide whether a given shape is a prism or a pyramid by "
    "checking its faces, vertices and edges against the book's table.",
    "prism, pyramid, bottom face, top face, vertical face, rectangular face, "
    "triangular face, vertex, edge",
)

# --------------------------------------------------------------------------
# Questions — every task in the book, placed with the concept that teaches it.
# Tables are written out as readable lines, not flattened into pipe soup.
# needs_image records a task the book prints with a picture it depends on.
# --------------------------------------------------------------------------
POLY_TABLE = (
    "Complete the table below.\n"
    "Name | Figure | Number of Vertices | Number of Sides\n"
    "Triangle | | 3 | 3\n"
    "Quadrilateral | | | \n"
    "Pentagon | | | \n"
    "Hexagon | | | \n"
    "Heptagon | | | \n"
    ".......... | | | "
)
PRISM_PYRAMID_TABLE = (
    "Complete the table below.\n"
    "Columns: Triangular Prism | Triangular Pyramid | Square Prism | "
    "Square Pyramid | Pentagonal Prism | Pentagonal Pyramid\n"
    "Picture: (to be drawn for each)\n"
    "Bottom face: Triangular | Triangular | | | | \n"
    "No. of vertices: 6 | 4 | | | | \n"
    "No. of edges: 9 | 6 | | | | \n"
    "No. of vertical faces: 3 | 3 | | | | \n"
    "No. of bottom/top faces: 2 | 1 | | | | "
)
SURROUND_TABLE = (
    "Complete the table by writing the names of the shapes that appear in "
    "the surrounding area.\n"
    "Row 1: Cone | | | | "
)

Q = [
    # (concept, sheet, category, cognitive, difficulty, marks, text, answer, needs_image)
    ("dim", "D", "Long Answer", "Understanding", "Moderate", 3,
     "Identify the shapes of the following objects and classify them as "
     "two-dimensional and three-dimensional shapes.\n"
     "Objects shown: Drum; Surface of a table; Reed pipe / Mouthpiece; "
     "Carrom-board surface; Harmonium; Road; Playground; Spinning Top\n"
     "Sr.No. | Two dimensional shape | Three dimensional shape\n"
     "1 | Surface of a table | \n2 | | \n3 | | \n4 | | ",
     "Two dimensional: surface of a table, carrom-board surface, road, "
     "playground. Three dimensional: drum, reed pipe/mouthpiece, harmonium, "
     "spinning top.", True),
    ("dim", "D", "Long Answer", "Applying", "Moderate", 2,
     "Write down the names of some other shapes that you know and classify "
     "them.",
     "For example: two dimensional - triangle, rectangle, circle; three "
     "dimensional - cuboid, cylinder, sphere.", False),
    ("dim", "D", "Long Answer", "Applying", "Less", 1,
     "Observe two-dimensional and three-dimensional shapes through the "
     "Internet.",
     "Learner records the two and three dimensional shapes observed.", False),
    ("dim", "D", "Long Answer", "Analysing", "Moderate", 3,
     "Discuss the two-dimensional and three-dimensional shapes in the picture "
     "based on the following points.\n"
     "1) A tree and its shadow\n2) Faces of cupboard and cupboard\n"
     "3) Faces of trunk and trunk",
     "The shadow, the faces of the cupboard and the faces of the trunk are "
     "two dimensional. The tree, the cupboard and the trunk are three "
     "dimensional.", True),
    ("dim", "D", "Long Answer", "Applying", "Moderate", 3,
     "I will fill a matchbox with wet soil and make a brick out of it. How "
     "can you create three-dimensional shapes? Is the resulting shape "
     "(cuboid, brick shape) three-dimensional? Check this.",
     "Filling the matchbox with wet soil makes a shape that takes up space. "
     "The brick formed is a cuboid and it is three dimensional.", True),
    ("poly", "D", "Long Answer", "Understanding", "Moderate", 4,
     POLY_TABLE,
     "Quadrilateral 4 vertices 4 sides; Pentagon 5 and 5; Hexagon 6 and 6; "
     "Heptagon 7 and 7.", True),
    ("poly", "S", "Short Answer", "Analysing", "Moderate", 2,
     "If we keep on increasing the number of sides of a polygon, what closed "
     "figure will be formed?",
     "As the number of sides keeps increasing, the polygon looks more and "
     "more like a circle.", False),
    ("poly", "D", "Long Answer", "Applying", "Less", 2,
     "Create various polygons using straws/sticks.",
     "Learner joins straws or sticks end to end so the figure closes, and "
     "makes polygons with different numbers of sides.", False),
    ("prism", "D", "Long Answer", "Applying", "Moderate", 4,
     "Activity 1:\n(1) Cut 10 to 15 triangles of equal size from thick "
     "cardboard.\n(2) Place the cut triangles on top of each other.\n"
     "(3) Write the elements of the three-dimensional shape formed.",
     "The shape formed is a triangular prism. It has 6 vertices, 2 triangular "
     "faces, 3 rectangular faces and 9 edges.", True),
    ("prism", "D", "Long Answer", "Applying", "Moderate", 3,
     "Activity 2:\n(1) Cut 10 to 15 rectangles of equal size from thick "
     "cardboard.\n(2) Place the cut rectangles on top of each other.\n"
     "(3) Write the elements of the three-dimensional shape formed.",
     "The shape formed is a cuboid.", True),
    ("prism", "S", "Fill in the blank", "Remembering", "Less", 3,
     "Cuboid has ▯ vertices, ▯ rectangular faces, ▯ edges.",
     "8 vertices, 6 rectangular faces, 12 edges.", False),
    ("prism", "D", "Long Answer", "Applying", "Moderate", 4,
     "As in activity No. 1 and 2 above, make a pile of different shapes like "
     "squares, pentagons, hexagons and stack them on top of each other. "
     "Record their names, vertices, faces and edges.",
     "A prism from squares, a prism from pentagons and a prism from hexagons, "
     "each with its names, vertices, faces and edges recorded.", False),
    ("cyl", "D", "Long Answer", "Applying", "Moderate", 3,
     "Activity 3:\n(1) Cut 10 to 15 circles of equal radius from thick "
     "cardboard.\n(2) Place the cut circles on top of each other.",
     "The shape formed is a cylinder, which is also called a circular prism.",
     False),
    ("cyl", "D", "Long Answer", "Understanding", "Moderate", 3,
     "Write the elements of the three-dimensional shape formed.",
     "The cylinder has 2 circular faces, 1 curved face and 2 circular edges.",
     True),
    ("cyl", "S", "Fill in the blank", "Remembering", "Less", 3,
     "The three-dimensional figure formed is a cylinder. A cylinder has ▯ "
     "circular faces, ▯ curved face, ▯ circular edges.",
     "2 circular faces, 1 curved face, 2 circular edges.", False),
    ("pyr", "S", "Fill in the blank", "Remembering", "Less", 4,
     "The adjacent three-dimensional shape is a square pyramid. It has ▯ "
     "vertices, ▯ triangular faces, ▯ edges, ▯ square face.",
     "5 vertices, 4 triangular faces, 8 edges, 1 square face.", True),
    ("cone", "D", "Long Answer", "Applying", "High", 5,
     "Activity:\n(1) Draw a circle with a radius of 6 cm on a piece of "
     "paper.\n(2) Cut the circle with scissors. Cut out two semicircles from "
     "it.\n(3) Glue the two radii of a semicircle together.\n(4) Draw another "
     "circle with radius 3 cm on the paper. Cut it out and stick it on the "
     "circumference of the shape you have made.\n(5) The three-dimensional "
     "shape formed is a cone.\n(6) Write the elements of a cone.",
     "The elements of a cone are the vertex, the circular edge, the circular "
     "face and the curved face.", True),
    ("hgt", "D", "Long Answer", "Understanding", "Moderate", 3,
     "How can the height of prism, pyramid, cylinder and cone be shown?",
     "For a prism and a cylinder the height is shown between the two end "
     "faces. For a pyramid and a cone it is shown from the top vertex "
     "straight down to the bottom face.", False),
    ("hgt", "S", "Short Answer", "Analysing", "Moderate", 2,
     "Can the height of a three-dimensional shape sphere be shown?",
     "A sphere has no flat face and no vertex, so its height cannot be shown "
     "in the same way as the other shapes.", False),
    ("diff", "D", "Long Answer", "Understanding", "High", 5,
     PRISM_PYRAMID_TABLE,
     "Square prism: 8 vertices, 12 edges, 4 vertical faces, 2 bottom/top "
     "faces. Square pyramid: 5 vertices, 8 edges, 4 vertical faces, 1 bottom "
     "face. Pentagonal prism: 10 vertices, 15 edges, 5 vertical faces, 2 "
     "bottom/top faces. Pentagonal pyramid: 6 vertices, 10 edges, 5 vertical "
     "faces, 1 bottom face.", True),
    ("prism", "D", "Long Answer", "Applying", "Moderate", 3,
     SURROUND_TABLE,
     "Names of shapes seen around, for example cone, cylinder, cuboid, "
     "sphere.", True),
    ("diff", "S", "Short Answer", "Analysing", "Moderate", 2,
     "Is cube a Prism or a Pyramid?",
     "A cube is a prism, because its bottom and top faces are equal squares "
     "and its vertical faces are rectangular.", False),
    ("diff", "S", "Short Answer", "Analysing", "Moderate", 2,
     "In which shape are all the faces triangular in shape?",
     "In a triangular pyramid all four faces are triangular.", False),
    ("diff", "S", "Short Answer", "Analysing", "Moderate", 2,
     "If we keep on increasing the number of sides of base polygon then which "
     "3D shape will be formed?",
     "As the number of sides of the bottom face keeps increasing, a prism "
     "comes to look like a cylinder and a pyramid comes to look like a cone.",
     False),
    ("prism", "O", "Multiple Choice Question", "Understanding", "Less", 1,
     "What shape will be formed if carrom pieces are placed one on top of "
     "the other?", "B) Cylinder", False),
    ("cone", "O", "Multiple Choice Question", "Analysing", "Moderate", 1,
     "By which 2 similar shapes is hourglass formed?", "D) Cone", False),
    ("dim", "O", "Multiple Choice Question", "Remembering", "Less", 1,
     "Which of the following shapes is not three-dimensional?",
     "A) square", False),
    ("pyr", "D", "Long Answer", "Applying", "Moderate", 3,
     "Draw the three-dimensional figure from the numbers given below and "
     "write its name.\nThere are 5 faces, 5 vertices, and 8 edges. One face "
     "is square and the remaining faces are triangular.",
     "The shape is a square pyramid.", False),
    ("sph", "D", "Long Answer", "Applying", "Moderate", 3,
     "Draw the three-dimensional figure from the numbers given below and "
     "write its name.\nOne face is curved. It has no vertices and no edges.",
     "The shape is a sphere.", False),
    ("prism", "D", "Long Answer", "Applying", "Moderate", 3,
     "Draw the three-dimensional figure from the numbers given below and "
     "write its name.\nThere are 6 faces, 12 edges and 8 vertices.",
     "The shape is a cuboid.", False),
    ("cyl", "D", "Long Answer", "Applying", "Moderate", 3,
     "Draw the three-dimensional figure from the numbers given below and "
     "write its name.\nThere is 1 curved face, 2 circular edges, and 2 flat "
     "faces.", "The shape is a cylinder.", False),
]

MCQ_OPTIONS = {
    "What shape will be formed if carrom pieces are placed one on top of "
    "the other?": [("Cuboid", "No"), ("Cylinder", "Yes"), ("Cone", "No"),
                   ("Cube", "No")],
    "By which 2 similar shapes is hourglass formed?":
        [("Cube", "No"), ("Triangular prism", "No"), ("Sphere", "No"),
         ("Cone", "Yes")],
    "Which of the following shapes is not three-dimensional?":
        [("square", "Yes"), ("Sphere", "No"), ("Cylinder", "No"),
         ("Cone", "No")],
}

# --------------------------------------------------------------------------
# Activity/Info Hub — the book's activities, experiments and discussion cases.
# The contract puts this material here and explicitly NOT into Cases
# ("Textbook activities, experiments, and discussion cases belong in
# Activity/Info Hub, not as Cases").
# --------------------------------------------------------------------------
HUB = {
    "dim":
        "The chapter fills a matchbox with wet soil to make a brick and asks "
        "whether the resulting cuboid is three dimensional. It also asks "
        "learners to observe two dimensional and three dimensional shapes "
        "through the Internet, and to discuss the shapes in the picture of "
        "the playground and the musical instruments.",
    "poly":
        "Learners create various polygons using straws or sticks, which shows "
        "that a figure is a polygon only when the sides close it.",
    "prism":
        "Activity 1 cuts 10 to 15 triangles of equal size from thick cardboard "
        "and places them on top of each other; Activity 2 does the same with "
        "rectangles. Learners then make a pile of squares, pentagons and "
        "hexagons and record the names, vertices, faces and edges of each "
        "shape formed.",
    "cyl":
        "Activity 3 cuts 10 to 15 circles of equal radius from thick cardboard "
        "and places them one on top of the other to form the shape.",
    "cone":
        "The chapter draws a circle of radius 6 cm, cuts two semicircles from "
        "it, glues the two radii of one semicircle together, then cuts a "
        "circle of radius 3 cm and sticks it on to close the shape.",
}

# --------------------------------------------------------------------------
# Types / Cases. A Type is one reusable assessable pattern; a Case is a
# conceptual sub-type named by what is given, what is asked and under what
# constraint — never a raw question and never an activity title. Every source
# question that is not Hub material sits on its own Example line, in full.
# Type numbers are allocated continuously across the chapter in topic order
# (concept_refiner rules 1-2); Case restarts per Type, Example per Case.
# --------------------------------------------------------------------------
TYPES = {
    "dim": [
        ("Deciding the dimensions of a shape", [
            ("Given everyday objects or named shapes, sort them into two "
             "dimensional and three dimensional", [0, 1]),
            ("Given a list of named shapes, pick out the one that is not "
             "three dimensional", [26]),
        ]),
    ],
    "poly": [
        ("Counting the sides and vertices of a polygon", [
            ("Given the name of a polygon, state how many vertices and sides "
             "it has", [5]),
            ("Given a polygon whose number of sides keeps increasing, name "
             "the closed figure that is approached", [6]),
        ]),
    ],
    "prism": [
        ("Naming a prism and counting its elements", [
            ("Given a prism, state its vertices, faces and edges", [10]),
            ("Given equal objects stacked one on top of another, name the "
             "shape that is formed", [24]),
            ("Given the counts of vertices, faces and edges, draw the shape "
             "and name it", [29]),
            ("Given the shapes seen in the surrounding area, record their "
             "names against the shape they match", [20]),
        ]),
    ],
    "cyl": [
        ("Naming the elements of a cylinder", [
            ("Given a cylinder formed by stacking circles, write its "
             "elements", [13]),
            ("Given a cylinder, state how many circular faces, curved faces "
             "and circular edges it has", [14]),
            ("Given the counts of vertices, faces and edges, draw the shape "
             "and name it", [30]),
        ]),
    ],
    "pyr": [
        ("Naming a pyramid and counting its elements", [
            ("Given a square pyramid, state its vertices, triangular faces "
             "and edges", [15]),
            ("Given the counts of vertices, faces and edges, draw the shape "
             "and name it", [27]),
        ]),
    ],
    "cone": [
        ("Recognising a cone from the shapes that form it", [
            ("Given an object made of two similar shapes, name the shape it "
             "is built from", [25]),
        ]),
    ],
    "sph": [
        ("Recognising a sphere from its elements", [
            ("Given the counts of vertices, faces and edges, draw the shape "
             "and name it", [28]),
        ]),
    ],
    "hgt": [
        ("Showing the height of a three dimensional shape", [
            ("Given shapes with two end faces or with a top vertex, describe "
             "where the height is shown", [17]),
            ("Given a sphere, decide whether its height can be shown and say "
             "why", [18]),
        ]),
    ],
    "diff": [
        ("Telling a prism from a pyramid by its elements", [
            ("Given a prism and a pyramid on the same bottom face, complete their "
             "vertices, edges and faces", [19]),
            ("Given a named shape, decide whether it is a prism or a "
             "pyramid", [21]),
            ("Given a property of the faces, name the shape that has it",
             [22]),
            ("Given a base polygon whose number of sides keeps increasing, "
             "name the three dimensional shape that is approached", [23]),
        ]),
    ],
}

# --------------------------------------------------------------------------
# Learner analysis. Every normal concept carries exactly one combined section
# with both labelled parts: a commonly held incorrect belief, and a distinct
# plausible application/reasoning mistake.
# --------------------------------------------------------------------------
ANALYSIS = {
    "dim": ("Learners believe that any shape drawn on paper is two "
            "dimensional and any object they can hold is three dimensional, "
            "so they call a picture of a ball two dimensional and a flat "
            "cardboard square three dimensional.",
            "Asked to classify, the learner sorts by the material of the "
            "object instead of by whether it takes up space, and puts a "
            "sheet of paper with the cuboids because it can be held."),
    "elem": ("Learners believe a vertex and a side mean the same thing in "
             "two and three dimensional shapes, so they call the edge of a "
             "cuboid a side and count the faces as sides.",
             "Counting the elements of a cuboid, the learner counts the "
             "edges of one face only and reports 4 edges instead of 12, "
             "because they read the drawing as a flat rectangle."),
    "poly": ("Learners believe any figure with straight sides is a polygon, "
             "so they accept a figure whose sides do not meet.",
             "Counting the sides of a pentagon from a drawing, the learner "
             "counts the vertices instead and reports the wrong number, or "
             "counts one side twice where two sides meet."),
    "prism": ("Learners believe every prism is a cuboid, because the cuboid "
              "is the prism they meet most often.",
              "Counting the edges of a triangular prism, the learner counts "
              "only the 3 edges of the top face and the 3 of the bottom face "
              "and reports 6, leaving out the 3 vertical edges."),
    "cyl": ("Learners believe a cylinder has no edges because it is round, "
            "so they say it has only faces.",
            "Asked for the elements of a cylinder, the learner counts the "
            "curved face as two faces because it can be seen from both "
            "sides, and reports 3 faces instead of 2 circular faces and 1 "
            "curved face."),
    "pyr": ("Learners believe every pyramid is the square pyramid they have "
            "seen in pictures, so they expect 5 vertices for every pyramid.",
            "Naming a pyramid, the learner names it from a vertical face "
            "instead of the bottom face, and calls a square pyramid a "
            "triangular pyramid because its vertical faces are triangular."),
    "cone": ("Learners believe a cone and a pyramid are the same shape "
             "because both come to a point at the top.",
             "Asked what a cone is made from, the learner says a full circle "
             "rather than a semicircle, and cannot explain why the paper "
             "takes a curved shape when the two radii are joined."),
    "sph": ("Learners believe a sphere has one curved edge where it appears "
            "to end, because its outline is drawn as a circle.",
            "Asked for the elements of a sphere, the learner reports 1 edge "
            "and 1 vertex by reading the drawn outline and the shading dot "
            "as an edge and a vertex."),
    "hgt": ("Learners believe height is always the distance from the bottom "
            "face to the top face, so they look for two flat faces on every "
            "shape.",
            "Showing the height of a cone, the learner follows the side from the "
            "top vertex down to the edge of the bottom face, instead of going "
            "straight down to the bottom face."),
    "diff": ("Learners believe a shape is a pyramid whenever its bottom face "
             "is a polygon, without checking the vertical faces.",
             "Counting the edges of a square pyramid, the learner applies the "
             "prism rule and reports three times the base sides (12) instead "
             "of twice the base sides (8)."),
}

TOPIC_TITLE = {k: v for k, v in TOPICS}
TOPIC_NO = {k: i + 1 for i, (k, _) in enumerate(TOPICS)}


def _one_line(text):
    """An Example line carries the whole question but no newline.

    The contract forbids newlines inside concept_details except the Achieving
    Mastery line, while also requiring each source question copied in full. The
    book's multi-line tables and activities are therefore joined, not cut.
    """
    return " ".join(part.strip() for part in str(text).splitlines()
                    if part.strip())


def compose_details():
    """Build every concept_details in Aegis's canonical section order."""
    type_no = 0
    placed = {}
    for ck, c in C.items():
        sections = [f"Description: {c['description']}\n"
                    f"Achieving Mastery: {c['mastery']}"]
        if HUB.get(ck):
            sections.append(f"Activity/Info Hub: {HUB[ck]}")

        blocks = []
        for type_title, cases in TYPES.get(ck) or []:
            type_no += 1
            parts = [f"Type {type_no:02d}: {type_title}"]
            for case_no, (case_title, qids) in enumerate(cases, 1):
                parts.append(f"Case {case_no:02d}: {case_title}")
                for ex_no, qid in enumerate(qids, 1):
                    if qid in placed:
                        raise SystemExit(
                            f"question {qid} placed twice: "
                            f"{placed[qid]} and {ck}")
                    placed[qid] = ck
                    if Q[qid][0] != ck:
                        raise SystemExit(
                            f"question {qid} belongs to {Q[qid][0]}, "
                            f"listed under {ck}")
                    parts.append(
                        f"Example {ex_no:02d}: {_one_line(Q[qid][6])}")
            blocks.append(" ".join(parts))
        # Rule 3: a purely theoretical concept carries no Types section.
        if blocks:
            sections.append("Types: " + " ".join(blocks))

        miscon, error = ANALYSIS[ck]
        sections.append(f"Misconception/ Error Analysis: Misconceptions: "
                        f"{miscon}; Error Analysis: {error}")
        c["details"] = " // ".join(sections)

    # Fail closed: a question that reached neither a Case nor the Hub would be
    # silently invisible in the concept column.
    for qid, q in enumerate(Q):
        if qid not in placed and not HUB.get(q[0]):
            raise SystemExit(f"question {qid} has no Case and no Hub: {q[6][:60]}")
    return placed


PLACED = compose_details()


def topic_label(tk):
    return f"Topic {TOPIC_NO[tk]:02d}: {TOPIC_TITLE[tk]} ({CODE}_PL_{tk})"


def concept_label(ck):
    c = C[ck]
    return f"{c['title']} ({CODE}_PL_{c['topic']}_{c['camel']})"


# Concept -> Group -> Question is the publishable hierarchy: the Group is the
# Type/Case a question is hosted under, and it always hangs off a concept.
# Aegis gives every concept all three shells (build_concepts.py _add_concept),
# so a question can be filed, moved or added later without inventing one. The
# same three are written here whether or not they currently hold a question.
TIERS = ("Basic", "Intermediate", "Advanced")
TIER_OF = {"Less": "Basic", "Moderate": "Intermediate", "High": "Advanced"}


def group_of(ck, tier):
    """(display name, stable code) of one concept's Type/Case shell."""
    c = C[ck]
    return (f"{c['title']} — {tier}",
            f"{CODE}_PL_{c['topic']}_{c['camel']}_{tier[0]}")


def tier_columns(ck):
    """basic_groups / intermediate_groups / advanced_groups for a concept."""
    return {f"{t.lower()}_groups": group_of(ck, t)[0] for t in TIERS}


# --------------------------------------------------------------------------
# Column layout. The importer addresses columns POSITIONALLY (see the module
# docstring in app/bulk_import/reader.py) and the canonical layout repeats
# field names across bands — Descriptive carries question_label in both the
# Group band and the Question band. So the header is derived from the
# importer's own constants rather than hand-listed, and every value is written
# by (band, field) so a repeated name cannot land in the wrong column.
# --------------------------------------------------------------------------
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[2]))
from app.bulk_import import (  # noqa: E402
    CHAPTER_FIELDS, TOPIC_FIELDS, CONCEPT_FIELDS, OBJECTIVE_GROUP_FIELDS,
    DESCRIPTIVE_GROUP_FIELDS, FIELDS_BY_KIND, SHEET_BY_KIND,
)

KIND = {"O": "objective", "S": "subjective", "D": "descriptive"}
SHEETS = {sk: SHEET_BY_KIND[kind] for sk, kind in KIND.items()}


def _bands(sk):
    """(header, {band: (offset, [field, ...])}) for one sheet."""
    gf = (DESCRIPTIVE_GROUP_FIELDS if sk == "D" else OBJECTIVE_GROUP_FIELDS)
    layout = [("chapter", CHAPTER_FIELDS), ("topic", TOPIC_FIELDS),
              ("concept", CONCEPT_FIELDS), ("group", gf),
              ("question", FIELDS_BY_KIND[KIND[sk]][
                  len(CHAPTER_FIELDS) + len(TOPIC_FIELDS)
                  + len(CONCEPT_FIELDS) + len(gf):])]
    bands, header, offset = {}, [], 0
    for band, fields in layout:
        bands[band] = (offset, list(fields))
        header.extend(fields)
        offset += len(fields)
    return header, bands


class Row:
    """A row addressed by (band, field), flushed to a positional list."""

    def __init__(self, header, bands):
        self.cells = [""] * len(header)
        self.bands = bands

    def set(self, band, **values):
        offset, fields = self.bands[band]
        for field, value in values.items():
            self.cells[offset + fields.index(field)] = value
        return self


BANNER = {"chapter": "Chapter", "topic": "Topic", "concept": "Concept",
          "group": "Group", "question": "Question"}

filled = set()   # (concept, tier) shells a question already creates

wb = Workbook()
wb.remove(wb.active)
counters = {}

for sk, sname in SHEETS.items():
    ws = wb.create_sheet(sname)
    hdr, bands = _bands(sk)
    banner = [""] * len(hdr)
    for band, (offset, _fields) in bands.items():
        banner[offset] = BANNER[band]
    ws.append(banner)
    ws.append(hdr)

    for ck, qsheet, cat, cog, diff, marks, qtext, ans, needs_img in Q:
        if qsheet != sk:
            continue
        c = C[ck]
        tk = c["topic"]
        counters[ck] = counters.get(ck, 0) + 1
        tier = TIER_OF[diff]
        gdisp, gname = group_of(ck, tier)
        filled.add((ck, tier))
        qlabel = f"{gname}{counters[ck]:02d}"
        row = Row(hdr, bands)
        row.set("chapter",
                chapter_title=CHAPTER, chapter_display_name=CH_DISPLAY,
                chapter_duration="3", chapter_description=CH_DESC)
        row.set("topic",
                topic_title=topic_label(tk),
                topic_display_name=TOPIC_TITLE[tk],
                pre_post_learning="Post",
                topic_concept_labels=concept_label(ck),
                topic_description=TOPIC_DESC[tk])
        row.set("concept",
                concept_title=concept_label(ck),
                concept_display_name=c["title"],
                concept_details=c["details"],
                concept_source="Balbharati Std 6 Mathematics, "
                               "Chapter 1 Three-Dimensional Shapes")
        row.set("concept", **tier_columns(ck))
        row.set("group",
                concept_question_labels=qlabel,
                group_name=gname, group_display_name=gdisp,
                group_description=f"{tier} group for {c['title']}.",
                group_status="Active", group_type=tier,
                group_question_labels=qlabel)
        row.set("question",
                question_label=qlabel, question_category=cat,
                cognitive_skills=cog,
                question_source="Balbharati Std 6 Mathematics Ch.1",
                question_duration="1",
                question_appears_in="Pre-test, Post-test, Worksheet, Test",
                level_of_difficulty=diff,
                question=qtext, marks=str(marks),
                answer_explanation=ans,
                question_disclaimer=(
                    "This question is printed with a picture in the book; "
                    "the picture must be attached for the question to be "
                    "answerable." if needs_img else ""))
        if sk == "O":
            opts = MCQ_OPTIONS.get(qtext.strip(), [])
            for i, (opt, correct) in enumerate(opts[:6], 1):
                row.set("question", **{
                    f"answer_type_{i}": "Words",
                    f"answer_content_{i}": opt,
                    f"correct_answer_{i}": correct,
                    f"answer_weightage_{i}": "1" if correct == "Yes" else "0",
                })
        elif sk == "S":
            row.set("question", math_keyboard="No", answer_type_1="Words",
                    answer_1=ans, answer_display_1="Yes",
                    weightage_1=str(marks))
        else:
            row.set("question", math_keyboard="No", display_answer="Yes",
                    answer_type_1="Words", answer_weightage_1=str(marks),
                    answer_content_1=ans)
        ws.append(row.cells)

    # Rows are emitted per question, so a Type/Case that currently holds no
    # question — and a concept the book sets no question on at all — would
    # never reach the sheet. The importer builds the hierarchy from a row whose
    # Question band is empty, so every remaining shell gets one context-only
    # row. That keeps the rule intact in both directions: a concept may have
    # no question, but no question exists without a Type/Case above it.
    if sk == "D":
        for ck, c in C.items():
            tk = c["topic"]
            for tier in TIERS:
                if (ck, tier) in filled:
                    continue
                gdisp, gname = group_of(ck, tier)
                row = Row(hdr, bands)
                row.set("chapter",
                        chapter_title=CHAPTER,
                        chapter_display_name=CH_DISPLAY,
                        chapter_duration="3", chapter_description=CH_DESC)
                row.set("topic",
                        topic_title=topic_label(tk),
                        topic_display_name=TOPIC_TITLE[tk],
                        pre_post_learning="Post",
                        topic_concept_labels=concept_label(ck),
                        topic_description=TOPIC_DESC[tk])
                row.set("concept",
                        concept_title=concept_label(ck),
                        concept_display_name=c["title"],
                        concept_details=c["details"],
                        concept_source="Balbharati Std 6 Mathematics, "
                                       "Chapter 1 Three-Dimensional Shapes")
                row.set("concept", **tier_columns(ck))
                row.set("group",
                        group_name=gname, group_display_name=gdisp,
                        group_description=f"{tier} group for {c['title']}.",
                        group_status="Active", group_type=tier)
                ws.append(row.cells)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[2]:
        cell.font = Font(bold=True, size=9)
    for col in ws.iter_cols(min_row=1, max_row=1):
        ws.column_dimensions[col[0].column_letter].width = 30
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

_ap = argparse.ArgumentParser()
_ap.add_argument("--out", default="CH01_Three_Dimensional_BULK_IMPORT.xlsx")
out = _ap.parse_args().out
pathlib.Path(out).parent.mkdir(parents=True, exist_ok=True)
wb.save(out)

n = {s: sum(1 for q in Q if q[1] == s) for s in SHEETS}
print(f"wrote {out}")
print(f"topics={len(TOPICS)} concepts={len(C)} questions={len(Q)} "
      f"(objective={n['O']} subjective={n['S']} descriptive={n['D']})")
print("questions needing the book's picture:",
      sum(1 for q in Q if q[8]))
for tk, tt in TOPICS:
    cs = [C[k]['title'] for k in C if C[k]['topic'] == tk]
    qs = sum(1 for q in Q if C[q[0]]['topic'] == tk)
    print(f"  {tk} {tt:38} concepts={len(cs)} questions={qs}")
