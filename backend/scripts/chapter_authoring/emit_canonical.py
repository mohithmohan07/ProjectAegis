#!/usr/bin/env python3
"""Emit the publishable workbook through Aegis's own writer.

The content is hand-authored (build_bulk_import.py). This step only applies
Aegis's formatting contract to it, so no rule is hand-copied and none can drift:

* chapter_duration comes from ``chapter_durations.lookup_duration_minutes`` —
  the Clarius chapter-time table — stored as "<n> minutes" exactly as
  ``build_concepts._sync_chapter_topic_summary`` stores it.
* pre_topics / post_topics, topic_description and chapter_description are
  written by that same function.
* every tag, label and group column is composed by ``bulk_import.writer``
  (chapter_tag / topic_tag / concept_tag, "Topic NN: <title> (<tag>)",
  "<concept> — <tier>", canonical question labels).

No model call and no Aegis generation phase runs here.
"""
from __future__ import annotations

import os
import pathlib
import sys

SCRATCH = pathlib.Path(__file__).resolve().parent
OUT = pathlib.Path(os.environ.get("AEGIS_AUTHORING_OUT", "out")).resolve()
OUT.mkdir(parents=True, exist_ok=True)
BACKEND = SCRATCH.parents[1]

DB = OUT / "canonical.db"
os.environ["AEGIS_DB_URL"] = f"sqlite:///{DB}"
os.environ["AEGIS_DATA_DIR"] = str(OUT / "data")
os.environ.setdefault("AEGIS_AUTH_MODE", "none")

sys.path.insert(0, str(BACKEND))
os.chdir(BACKEND)

from app.db import SessionLocal, init_db  # noqa: E402
from app import models  # noqa: E402
from app.bulk_import import writer  # noqa: E402
from app.bulk_import.reader import import_workbook  # noqa: E402
from app.services import build_concepts, chapter_durations  # noqa: E402
from app.services import generation  # noqa: E402

TIER_ORDER = {"Basic": 0, "Intermediate": 1, "Advanced": 2}

AUTHORED = OUT / "CH01_Three_Dimensional_BULK_IMPORT.xlsx"
FINAL = OUT / "CH01_Three_Dimensional_Shapes_BULK_IMPORT.xlsx"


def authored_order():
    """The book's printed topic order and concept order, from the source file."""
    import contextlib
    import importlib.util
    import io

    spec = importlib.util.spec_from_file_location(
        "authored", SCRATCH / "build_bulk_import.py")
    module = importlib.util.module_from_spec(spec)
    with contextlib.redirect_stdout(io.StringIO()):
        spec.loader.exec_module(module)
    topics = {title: n for n, (_, title) in enumerate(module.TOPICS, 1)}
    concepts = {c["title"]: n for n, c in enumerate(module.C.values(), 1)}
    return topics, concepts


def name_the_topics(path: pathlib.Path) -> int:
    """Put the topic (and concept) name into the tag codes.

    ``directory.topic_tag`` takes only chapter identity, so the writer stamps
    ONE code — ``06MSMA_<chapter>_PL`` — onto all eight topics, and the concept
    code stops at the topic, leaving sibling concepts sharing a code. Nothing
    in the sheet can then address a single topic or concept, which is what the
    codes are for when tagging.

    The names are appended, so every code still starts with exactly the tag
    Aegis composes and ``strip_title_tag`` / ``strip_topic_title`` recover the
    same clean values. Applied to every column that repeats a code, so the
    cross-references the importer links on stay identical.
    """
    import openpyxl

    from app import bulk_import as bi
    from app.services import directory as d

    # Codes are rebuilt from the names, never edited in place: a concept whose
    # name equals its topic's ("Cone" under topic Cone) would otherwise be
    # skipped by any "already ends with the slug" test and keep a code
    # identical to its topic's.
    base = None

    def topic_code(name: str) -> str:
        return f"{base}_{d._underscore_slug(name)}"

    def concept_code(topic_name: str, concept_name: str) -> str:
        return f"{topic_code(topic_name)}_{d._underscore_slug(concept_name)}"

    def retag_topic(cell: str) -> str:
        """'Topic 01: Name (tag)' -> '... (base_Name)'."""
        name = bi.strip_topic_title(cell)
        if not name or not cell.rstrip().endswith(")"):
            return cell
        head = cell.rstrip()[:-1].rpartition(" (")[0]
        return f"{head} ({topic_code(name)})"

    def retag_concept(cell: str, topic_name: str) -> str:
        """'Concept Name (tag)' -> '... (base_Topic_Concept)'."""
        name = bi.strip_title_tag(cell)
        if not name or not cell.rstrip().endswith(")") or not topic_name:
            return cell
        head = cell.rstrip()[:-1].rpartition(" (")[0]
        return f"{head} ({concept_code(topic_name, name)})"

    wb = openpyxl.load_workbook(path)
    changed = 0
    for kind, sheet in bi.SHEET_BY_KIND.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [str(c.value or "").strip() for c in ws[2]]
        col = {name: i for i, name in enumerate(header)}
        for row in ws.iter_rows(min_row=3):
            values = [c.value for c in row]
            if not any(values):
                continue
            if base is None:
                # The tag the writer composed, before any name is appended.
                sample = str(values[col["topic_title"]] or "")
                base = sample.rstrip()[:-1].rpartition(" (")[2]
            topic_name = str(values[col["topic_display_name"]] or "").strip()
            for cell in row:
                field = header[cell.column - 1] if cell.column <= len(header) else ""
                if not isinstance(cell.value, str) or not cell.value:
                    continue
                if field in {"topic_title", "pre_topics", "post_topics"}:
                    new = _map_tagged(cell.value, retag_topic)
                elif field in {"concept_title", "topic_concept_labels"}:
                    new = _map_tagged(
                        cell.value, lambda v: retag_concept(v, topic_name))
                else:
                    continue
                if new != cell.value:
                    cell.value = new
                    changed += 1
    wb.save(path)
    return changed


def _map_tagged(value: str, fn):
    """Apply ``fn`` to each 'Name (tag)' entry of a comma-separated cell.

    Split on ", " alone would cut names that contain a comma
    ("One, Two and Three Dimensional Shapes"), so entries are only closed at
    the ")" that ends a tag.
    """
    parts, buf = [], ""
    for chunk in value.split(", "):
        buf = f"{buf}, {chunk}" if buf else chunk
        if buf.rstrip().endswith(")"):
            parts.append(fn(buf))
            buf = ""
    if buf:
        parts.append(fn(buf))
    return ", ".join(parts)


def main() -> int:
    if DB.exists():
        DB.unlink()
    init_db()
    db = SessionLocal()
    try:
        counts = import_workbook(db, AUTHORED)
        issues = counts.pop("issues", [])
        print(f"loaded authored content: {counts}")
        for issue in issues:
            print(f"  !! {issue}")
        if issues:
            return 1

        chapter = db.query(models.Chapter).one()

        # Source order drives the writer's "Topic NN:" numbering and the order
        # concepts are listed in. Rows arrive in question order, so it has to
        # be set from the book's printed sequence, not from row/insert order.
        topic_order, concept_order = authored_order()
        for topic in chapter.topics:
            if topic.topic_title not in topic_order:
                print(f"!! topic not in the authored order: {topic.topic_title}")
                return 1
            topic.source_order = topic_order[topic.topic_title]
            for concept in topic.concepts:
                concept.source_order = concept_order[concept.concept_title]
        db.flush()

        # Book token for the chapter tag: primary_book_source() takes the
        # leading source, so keep it the book name alone.
        for concept in db.query(models.Concept):
            concept.sources = "Balbharati"

        # generation._topic_index orders a chapter's topics by row id. In a
        # real run topics are created in textbook order, so id order and
        # source order agree; here rows arrive in question order, which would
        # stamp the wrong T-number into every label. Index by source_order —
        # the same key writer._topic_number uses — so the label matches the
        # topic number the writer prints beside it.
        generation._topic_index = lambda concept: sorted(
            concept.topic.chapter.topics,
            key=lambda t: (t.source_order or 10**9, t.id),
        ).index(concept.topic) + 1

        # Group naming and question labels follow Aegis's own generators
        # rather than the codes the authoring sheet carried in:
        # "<concept> — <tier>" (build_concepts._add_concept /
        # build_assessments._group_for) and generation.question_label().
        for concept in db.query(models.Concept):
            counter = 0
            for group in sorted(concept.groups, key=lambda g: TIER_ORDER[g.group_type]):
                name = f"{concept.concept_title} — {group.group_type}"
                group.group_name = name
                group.group_display_name = name
                for question in sorted(group.questions, key=lambda q: q.id):
                    counter += 1
                    question.question_label = generation.question_label(
                        concept, counter)
        db.flush()

        # THE duration rule: the Clarius table, not an estimate.
        minutes = chapter_durations.lookup_duration_minutes(
            board=chapter.board, grade=chapter.grade,
            subject=chapter.subject, chapter_title=chapter.chapter_title,
        )
        if minutes is None:
            print("!! no duration row for this chapter; refusing to guess")
            return 1
        chapter.chapter_duration = f"{minutes} minutes"

        # Let Aegis write the chapter/topic summary fields it owns.
        chapter.chapter_description = ""
        for topic in chapter.topics:
            topic.topic_description = ""
        build_concepts._sync_chapter_topic_summary(chapter)
        db.commit()

        print(f"\nchapter_duration    : {chapter.chapter_duration}")
        print(f"chapter_description : {chapter.chapter_description}")
        print(f"post_topics         : {chapter.post_topics[:110]}...")

        writer.write_workbook(db, FINAL)
        added = name_the_topics(FINAL)
        print(f"\nwrote {FINAL} ({FINAL.stat().st_size:,} bytes); "
              f"named tags rewritten in {added} cells")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
