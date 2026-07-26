"""Concept-mapping output format: tags in title columns, label columns,
required fields, and tag-stripping round-trip."""
import io

import pytest

from app import bulk_import as bi
from app import models
from app.bulk_import import writer
from tests.conftest import convert_concept_upload, stream_result


def test_strip_helpers():
    assert bi.strip_title_tag("Number System (09_Mathematics_CBSE_RS)") \
        == "Number System"
    assert bi.strip_title_tag("Understanding Social Science (09_SocialScience_CBSE_NCERT)") \
        == "Understanding Social Science"
    assert bi.strip_title_tag("Understanding Social Science (09_Social_Science_CBSE_NCERT)") \
        == "Understanding Social Science"
    assert bi.strip_title_tag("What is Social Science (09CBSS_Ch_PL_Topic)") \
        == "What is Social Science"
    # A real parenthetical without underscores is preserved.
    assert bi.strip_title_tag("Photosynthesis (C3)") == "Photosynthesis (C3)"
    assert bi.strip_topic_title("Topic 01: Meaning of Social Science (09CBSS_Ch_PL)") \
        == "Meaning of Social Science"


def test_writer_composes_tags_labels_and_clean_display(db):
    concept = (
        db.query(models.Concept)
        .join(models.Topic).join(models.Chapter)
        .filter(models.Chapter.board != "").first()
    )
    assert concept is not None
    chapter = concept.topic.chapter
    row = writer._concept_to_row(concept, "objective")

    # Chapter: tag in title (col 0), clean display (col 1).
    assert bi.strip_title_tag(row[0]) == chapter.chapter_title
    assert "_" in row[0] and row[0].endswith(")")
    assert row[1] == chapter.chapter_title

    # Topic title col (6): "Topic NN: <title> (<tag>)"; display col (7):
    # clean topic title only, with no number/code/tag.
    clean_topic = bi.strip_topic_title(concept.topic.topic_title)
    assert row[6].startswith("Topic ") and row[6].endswith(")")
    assert bi.strip_topic_title(row[6]) == clean_topic
    assert row[7] == clean_topic
    assert "_" not in row[7]  # display name carries no code/tag

    # topic_concept_labels (col 9) lists the topic's concept titles in the
    # tagged "Name (tag)" form of the concept_title column (review: labels
    # must be concept title, not the clean display name).
    assert bi.OBJECTIVE_FIELDS[9] == "topic_concept_labels"
    assert concept.concept_title in row[9]
    # Each label matches the tagged concept_title column (col 12) exactly.
    assert row[12] in row[9]

    # Concept: tag in title (col 12), clean display (col 13).
    assert bi.strip_title_tag(row[12]) == concept.concept_title
    assert row[13] == concept.concept_title


def test_deposit_applies_numbering_recap_titlecase_and_topic_columns(db):
    """End-to-end deposit: continuous Type NN, Miscellaneous for culmination,
    Recap description, Title Case, and tagged pre/post topic columns."""
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="07CBMA_FmtRules", board="CBSE", grade="07",
        subject="Mathematics", unit="Mathematics Unit",
        chapter_title="Format Rules Chapter",
        chapter_display_name="Format Rules Chapter",
    )
    db.add(chapter)
    db.commit()

    records = [
        {"topic": "operations on numbers", "concept_title": "addition of integers",
         "parent_concept": "integer operations",
         "concept_details": (
             "Description: a // Types: Type 01: Direct "
             "Case 01: Direct addition of two positive integers. "
             "Example 01: Add 2 and 3 using integer addition rules. "
             "Case 02: Direct addition with a two-digit total. "
             "Example 01: Add 5 and 9 using integer addition rules. // "
             "Misconception: Students may believe adding integers always increases the value."), "keywords": ""},
        {"topic": "operations on numbers", "concept_title": "Culmination - Operations On Numbers",
         "parent_concept": "Culmination",
         "concept_details": ("Description: a long synthesis paragraph // "
                             "Types: Type 01: Mixed "
                             "Case 01: Mixed-operation integer review. "
                             "Example 01: Combine addition and subtraction steps in one integer problem. // "
                             "Misconception: Students may believe mixed review tasks use only one operation."), "keywords": ""},
        {"topic": "powers and roots", "concept_title": "squares of numbers",
         "parent_concept": "powers",
         "concept_details": (
             "Description: b // Types: Type 01: Compute "
             "Case 01: Computing a square by repeated multiplication. "
             "Example 01: Calculate 4 squared and explain the multiplication. // "
             "Misconception: Students may think squaring a number means doubling it."), "keywords": ""},
        {"topic": "powers and roots", "concept_title": "Culmination - Powers and Roots",
         "parent_concept": "Culmination",
         "concept_details": (
             "Description: recap // Types: Type 01: Mixed "
             "Case 01: Mixed review of powers and inverse roots. "
             "Example 01: Combine square and square-root facts in one review problem. // "
             "Misconception: Students may believe roots and powers cannot appear together."), "keywords": ""},
    ]
    build_concepts._deposit_concepts(db, chapter, records, "Post", "")
    build_concepts._sync_chapter_topic_summary(chapter)
    db.commit()

    by_title = {c.concept_title: c for t in chapter.topics for c in t.concepts}

    # Title Case on topics and concepts.
    topic_titles = {t.topic_title for t in chapter.topics}
    assert {"Operations on Numbers", "Powers and Roots"} <= topic_titles
    assert "Addition of Integers" in by_title

    # Continuous Type numbering across the chapter (culmination excluded).
    assert "Type 01: Direct" in by_title["Addition of Integers"].concept_details
    assert "Type 02: Compute" in by_title["Squares of Numbers"].concept_details

    # Culmination: Miscellaneous Type sequence + Description collapses to "Recap".
    culm = by_title["Culmination - Operations on Numbers"].concept_details
    assert "Miscellaneous Type 01: Mixed" in culm
    assert "Description: Recap" in culm
    assert "long synthesis" not in culm
    assert "Misconceptions: Students may believe mixed review tasks use only one operation." in culm

    # Column E (post_topics) lists tagged Topic Titles (with the code).
    assert "Topic 01:" in chapter.post_topics
    assert "07CBMA" in chapter.post_topics  # the code is present

    # Topic display column shows only the title, without "Topic NN:" or code.
    row = writer._concept_to_row(by_title["Addition of Integers"], "objective")
    assert row[4] == chapter.post_topics  # chapter band col E
    assert row[7] == "Operations on Numbers"
    assert "_" not in row[7]


def test_post_deposit_keeps_math_recap_rich_text_canonical(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBMA_MathRecap",
        board="CBSE",
        grade="10",
        subject="Mathematics",
        unit="Mathematics Unit",
        chapter_title="Math Recap",
        chapter_display_name="Math Recap",
    )
    db.add(chapter)
    db.commit()
    raw_title = r"Use the Finite-sum Formula S_n = \frac{n}{2}(a+l)"
    records = [
        {
            "topic": "Sequences",
            "concept_title": raw_title,
            "parent_concept": "Finite Sums",
            "concept_details": (
                "Description: The finite-sum formula combines the first and "
                "last terms of an arithmetic progression. // Misconception/ "
                "Error Analysis: Misconceptions: Students may believe the "
                "factor n can be omitted.; Error Analysis: Students may "
                "substitute the common difference for the last term."
            ),
            "keywords": "",
        },
        {
            "topic": "Sequences",
            "concept_title": f"Culmination - {raw_title}",
            "parent_concept": "Culmination",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]

    created, _merged = build_concepts._deposit_concepts(
        db, chapter, records, "Post", "")
    db.flush()

    culmination = next(
        concept for concept in (db.get(models.Concept, cid) for cid in created)
        if concept.parent_concept == "Culmination"
    )
    assert "[Katex]" in culmination.concept_details
    assert not build_concepts.generation.kr.rich_text_issues(
        culmination.concept_details)


def test_post_deposit_rejects_cases_without_numbered_examples(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="07CBMA_StrictCases", board="CBSE", grade="07",
        subject="Mathematics", unit="Mathematics Unit",
        chapter_title="Strict Case Contract",
        chapter_display_name="Strict Case Contract",
    )
    db.add(chapter)
    db.commit()
    records = [
        {
            "topic": "Integer Operations",
            "concept_title": "Adding Integers",
            "parent_concept": "Integer Operations",
            "concept_details": (
                "Description: Addition combines signed integer quantities. // "
                "Types: Type 01: Direct calculation "
                "Case 01: Direct addition of two positive integers. // "
                "Misconceptions: Students may believe signed addition always "
                "increases the magnitude."
            ),
            "keywords": "",
        },
        {
            "topic": "Integer Operations",
            "concept_title": "Culmination - Integer Operations",
            "parent_concept": "Culmination",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]

    with pytest.raises(ValueError, match="case_without_example"):
        build_concepts._deposit_concepts(
            db, chapter, records, "Post", "")


def test_post_deposit_treats_explicit_empty_inventory_as_closed_world(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBMA_EmptyInventory",
        board="CBSE",
        grade="10",
        subject="Mathematics",
        unit="Algebra",
        chapter_title="Explicit Empty Inventory",
        chapter_display_name="Explicit Empty Inventory",
    )
    db.add(chapter)
    db.commit()
    records = [
        {
            "topic": "Linear Equations",
            "parent_concept": "Equation Solving",
            "concept_title": "Balancing Linear Equations",
            "concept_details": (
                "Description: A linear equation remains balanced when the same "
                "valid operation is applied to both sides. // "
                "Types: Type 01: Solving linear balance equations "
                "Case 01: Isolate the unknown using inverse operations. "
                "Example 01: Solve 3x + 5 = 20 and explain each balancing step. "
                "// Misconception/ Error Analysis: Misconceptions: Students "
                "may believe an operation applies to only one side.; Error "
                "Analysis: Students may change a sign without performing the "
                "same inverse operation on both sides."
            ),
            "keywords": "",
        },
        {
            "topic": "Linear Equations",
            "parent_concept": "Culmination",
            "concept_title": "Culmination - Linear Equations",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]

    with pytest.raises(
        ValueError,
        match=r"source_inventory_semantics; unexpected_examples=1",
    ):
        build_concepts._deposit_concepts(
            db,
            chapter,
            records,
            "Post",
            "",
            inventory={"items": [], "stats": {}},
        )

    assert not chapter.topics


def test_post_deposit_without_inventory_uses_generation_fatal_policy(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBSS_NoInventoryFatal",
        board="CBSE",
        grade="10",
        subject="Social Science",
        unit="History",
        chapter_title="No Inventory Fatal Validation",
        chapter_display_name="No Inventory Fatal Validation",
    )
    db.add(chapter)
    db.commit()
    copied_description = (
        "National allegories transform abstract political communities into "
        "recognizable female figures whose attributes help citizens identify "
        "shared history ideals and collective belonging across generations."
    )
    records = [
        {
            "topic": "National Allegories",
            "parent_concept": "Visual Nationalism",
            "concept_title": "Representing the Nation",
            "concept_details": (
                f"Description: {copied_description} // "
                "Misconception/ Error Analysis: Misconceptions: Students may "
                "believe every allegory represents the same nation.; Error "
                "Analysis: Students may name a figure without connecting its "
                "attributes to the represented nation."
            ),
            "keywords": "",
        },
        {
            "topic": "National Allegories",
            "parent_concept": "Culmination",
            "concept_title": "Culmination - National Allegories",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]

    with pytest.raises(
        ValueError,
        match=r"concept validation failed before deposit: "
              r"verbatim_source_description",
    ):
        build_concepts._deposit_concepts(
            db,
            chapter,
            records,
            "Post",
            "",
            inventory=None,
            source_text=copied_description,
        )

    assert not chapter.topics


def test_post_deposit_preserves_inventory_example_and_image_in_export(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBPH_Circuits", board="CBSE", grade="10",
        subject="Physics", unit="Electricity",
        chapter_title="Circuit Interpretation",
        chapter_display_name="Circuit Interpretation",
    )
    db.add(chapter)
    db.commit()
    item = {
        "qid": "QINV-0001",
        "source_kind": "exercise",
        "topic_hint": "Circuit Diagrams",
        "source_label": "Fig. 11.1",
        "raw_task": (
            "Compare the circuits in Fig. 11.1 and state which lamp glows."
        ),
        "image_urls": ["https://example.test/figure-11-1.png"],
    }
    prompt = build_concepts.generation._inventory_task_text(item)
    records = [
        {
            "topic": "Circuit Diagrams",
            "concept_title": "Reading Circuit Diagrams",
            "parent_concept": "Electric Circuits",
            "concept_details": (
                "Description: Circuit diagrams show connected paths for current. // "
                "Types: Type 01: Diagram interpretation "
                "Case 01: Comparisons based on a visible circuit configuration. "
                f"Example 01: {prompt} // "
                "Misconceptions: Students may believe an open switch increases "
                "current. // Error Analysis: A learner may trace current through "
                "the broken branch and reach the wrong result."
            ),
            "keywords": "",
        },
        {
            "topic": "Circuit Diagrams",
            "concept_title": "Culmination - Circuit Diagrams",
            "parent_concept": "Culmination",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]
    inventory = {"items": [item]}

    created, merged = build_concepts._deposit_concepts(
        db,
        chapter,
        records,
        "Post",
        "",
        inventory=inventory,
        source_text=item["raw_task"],
    )
    assert not merged
    assert len(created) == 2
    concept = next(
        concept
        for topic in chapter.topics
        for concept in topic.concepts
        if concept.concept_title == "Reading Circuit Diagrams"
    )
    assert "Case 01:" in concept.concept_details
    assert "Example 01:" in concept.concept_details
    assert (
        '[img src="https://example.test/figure-11-1.png" '
        'alt="Fig. 11.1"]'
    ) in concept.concept_details

    row = writer._concept_to_row(concept, "objective")
    exported = row[bi.OBJECTIVE_FIELDS.index("concept_details")]
    assert exported == concept.concept_details
    assert "Example 01:" in exported
    assert "https://example.test/figure-11-1.png" in exported


def test_post_deposit_refreshes_existing_concept_with_current_contract(db):
    """A title collision must not leave legacy content in the exported row."""
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBSS_Refresh", board="CBSE", grade="10",
        subject="Social Science", unit="History",
        chapter_title="Refresh Contract", chapter_display_name="Refresh Contract",
    )
    topic = models.Topic(
        topic_title="Nationalism", topic_display_name="Nationalism",
        pre_post_learning="Post",
    )
    chapter.topics.append(topic)
    legacy = models.Concept(
        concept_title="Popular Sovereignty",
        concept_display_name="Popular Sovereignty",
        parent_concept="Legacy Parent",
        concept_details=(
            "Description: Legacy text. // Misconceptions: Students may believe "
            "only rulers hold political authority. // Error Analysis: Students "
            "may omit the role of citizens."
        ),
        keywords="legacy",
    )
    topic.concepts.append(legacy)
    db.add(chapter)
    db.commit()
    legacy_id = legacy.id

    records = [
        {
            "topic": "Nationalism",
            "concept_title": "Popular Sovereignty",
            "parent_concept": "Democratic Nationhood",
            "concept_details": (
                "Description: Popular sovereignty locates political authority "
                "with citizens who participate in shaping public institutions. // "
                "Misconception/ Error Analysis: Misconceptions: Students may "
                "believe sovereignty belongs only to a monarch.; Error Analysis: "
                "Students may omit citizens' role when explaining political "
                "authority."
            ),
            "keywords": "sovereignty, citizens",
        },
        {
            "topic": "Nationalism",
            "concept_title": "Culmination - Nationalism",
            "parent_concept": "Culmination",
            "concept_details": "Description: Recap",
            "keywords": "",
        },
    ]

    created, merged = build_concepts._deposit_concepts(
        db, chapter, records, "Post", "NCERT")

    assert legacy_id in merged
    assert len(created) == 1
    db.flush()
    refreshed = db.get(models.Concept, legacy_id)
    assert refreshed is not None
    assert refreshed.parent_concept == "Democratic Nationhood"
    assert refreshed.keywords == "sovereignty, citizens"
    assert refreshed.concept_details.count("Misconception/ Error Analysis:") == 1
    assert " // Misconceptions:" not in refreshed.concept_details
    assert " // Error Analysis:" not in refreshed.concept_details


def test_post_deposit_repairs_missing_inventory_question_before_writes(db):
    from app.services import build_concepts

    chapter = models.Chapter(
        chapter_code="10CBSS_Coverage", board="CBSE", grade="10",
        subject="Social Science", unit="History",
        chapter_title="Coverage Contract",
        chapter_display_name="Coverage Contract",
    )
    db.add(chapter)
    db.commit()
    task = "Explain how popular sovereignty changed political authority."
    records = [{
        "topic": "Nationalism",
        "concept_title": "Popular Sovereignty",
        "parent_concept": "Nation States",
        "concept_details": (
            "Description: Popular sovereignty places political authority with "
            "the people. // Misconception/ Error Analysis: Misconceptions: "
            "Students may believe sovereignty belongs only to a monarch.; Error "
            "Analysis: Students may treat political authority as hereditary "
            "rather than civic."
        ),
        "keywords": "",
    }, {
        "topic": "Nationalism",
        "concept_title": "Culmination - Nationalism",
        "parent_concept": "Culmination",
        "concept_details": "Description: Recap",
        "keywords": "",
    }]
    inventory = {"items": [{
        "qid": "QINV-0007",
        "source_kind": "checkpoint_question",
        "topic_hint": "Nationalism",
        "raw_task": task,
    }]}

    created, merged = build_concepts._deposit_concepts(
        db,
        chapter,
        records,
        "Post",
        "",
        inventory=inventory,
    )

    assert len(created) == 2
    assert not merged
    normal = next(
        concept
        for topic in chapter.topics
        for concept in topic.concepts
        if concept.concept_title == "Popular Sovereignty"
    )
    assert f"Example 01: {task}" in normal.concept_details
    assert build_concepts.generation._rendered_inventory_coverage_defects(
        [{
            "concept_details": normal.concept_details,
            "concept_title": normal.concept_title,
            "topic": normal.topic.topic_title,
        }],
        inventory,
    ) == {"missing": [], "duplicate": []}


def test_group_label_columns_present_and_ordered():
    f = bi.OBJECTIVE_FIELDS
    assert f[21] == "concept_question_labels"
    gi = f.index("group_question_labels")
    assert f[gi + 1] == "related_digicards"


def test_writer_leaves_group_columns_empty_for_concept_rows(db):
    """Concept-catalog rows must not pre-fill group columns at generation time."""
    concept = (
        db.query(models.Concept)
        .join(models.Topic).join(models.Chapter)
        .filter(models.Chapter.board != "").first()
    )
    assert concept is not None
    row = writer._concept_to_row(concept, "objective")
    basic_idx = bi.OBJECTIVE_FIELDS.index("basic_groups")
    assert row[basic_idx] == ""
    assert row[basic_idx + 1] == ""
    assert row[basic_idx + 2] == ""


def test_writer_uses_canonical_parent_concept_column(db):
    """parent_concept is now a first-class canonical column (no fallback)."""
    concept = db.query(models.Concept).join(models.Topic).join(models.Chapter).first()
    concept.parent_concept = "Cell Organisation"
    row = writer._concept_to_row(concept, "objective")
    parent_idx = bi.OBJECTIVE_FIELDS.index("parent_concept")
    assert row[parent_idx] == "Cell Organisation"
    # keywords / related_concepts columns are gone from new workbooks.
    assert "keywords" not in bi.OBJECTIVE_FIELDS
    assert "related_concepts" not in bi.OBJECTIVE_FIELDS


def test_writer_falls_back_to_related_concepts_on_legacy_workbook(db):
    """Legacy workbooks (no parent_concept column) still get the parent marker
    in their related_concepts column."""
    concept = db.query(models.Concept).join(models.Topic).join(models.Chapter).first()
    concept.parent_concept = "Cell Organisation"
    row = writer._concept_to_row(
        concept, "objective", concept_fields=bi.LEGACY_CONCEPT_FIELDS)
    related_idx = (
        len(bi.CHAPTER_FIELDS) + len(bi.TOPIC_FIELDS)
        + bi.LEGACY_CONCEPT_FIELDS.index("related_concepts")
    )
    assert "parent: Cell Organisation" in row[related_idx]


def test_append_concepts_writes_parent_concept_column(db, tmp_path):
    import openpyxl

    concept = db.query(models.Concept).join(models.Topic).join(models.Chapter).first()
    concept.parent_concept = "Cell Organisation"
    db.flush()

    path = tmp_path / "parent_template.xlsx"
    writer._new_workbook().save(path)

    result = writer.append_concepts(db, path, [concept.id])
    assert result["parent_column"] is True
    out = openpyxl.load_workbook(path)[bi.SHEET_OBJECTIVE]
    headers = [c.value for c in out[2]]
    parent_idx = headers.index("parent_concept") + 1
    assert out.cell(row=3, column=parent_idx).value == "Cell Organisation"


def test_roundtrip_recovers_clean_titles(db):
    """Export concepts (tagged cells) then re-import: clean titles are recovered."""
    concept = db.query(models.Concept).join(models.Topic).join(models.Chapter).first()
    data = writer.write_concepts_workbook(db, [concept.id])
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(data))[bi.SHEET_OBJECTIVE]
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if r and r[0]]
    assert rows
    # The exported title cell carries a tag; stripping recovers the clean title.
    assert bi.strip_title_tag(str(rows[0][12])) == concept.concept_title
    assert "_" in str(rows[0][12])  # a tag is present


def test_deposit_fills_required_fields(client, db):
    # A fresh chapter with blank (NA-equivalent) required fields.
    chapter = models.Chapter(
        chapter_code="09CBSS_ReqTest", board="CBSE", grade="09",
        subject="Social Science", unit="Social Science Unit",
        chapter_title="Req Test Chapter", chapter_display_name="Req Test Chapter",
        chapter_duration="", chapter_description="", pre_topics="", post_topics="",
    )
    db.add(chapter)
    db.commit()
    chapter_id = chapter.id

    files = {"file": ("req.txt", io.BytesIO(
        b"## Meaning of Social Science\nWhat is social science studies.\n"
        b"Scope of social science explained.\n"
        b"## Importance of Social Science\nWhy social science matters today."
    ), "text/plain")}
    job = client.post("/build-concepts/post-learning/uploads", files=files).json()
    convert_concept_upload(client, job["id"])
    stream_result(client.post(
        f"/build-concepts/post-learning/uploads/{job['id']}/generate",
        json={"target_chapter_id": chapter_id}))

    db.expire_all()
    chapter = db.get(models.Chapter, chapter_id)
    # Required fields are filled (no "NA").
    assert chapter.chapter_duration.endswith("minutes")
    assert chapter.chapter_description and chapter.chapter_description.lower() != "na"
    # pre/post topic lists are comma-separated (never semicolons).
    assert ";" not in (chapter.post_topics or "")
    # Newly created post topics carry a synthesized summary.
    new = [t for t in chapter.topics if t.topic_title in
           ("Meaning of Social Science", "Importance of Social Science")]
    assert new and all(t.topic_description for t in new)
