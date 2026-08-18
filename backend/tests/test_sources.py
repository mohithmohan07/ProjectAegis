"""Multi-source tracking: concept/question dedupe across books + source merge."""
import io

import openpyxl

from app import bulk_import as bi
from app import config, models
from app.bulk_import import layouts, writer


def _use_specific_dry_learner_analysis(monkeypatch):
    # The deterministic learner-analysis fallbacks are deleted (filler is
    # never synthesized); dry rows simply carry whatever analysis their
    # fixtures author. Kept as a no-op seam so callers stay explicit.
    del monkeypatch


def test_merge_sources_dedupes_case_insensitively():
    assert bi.merge_sources("NCERT", "RD Sharma") == "NCERT, RD Sharma"
    # comma is the only supported separator; legacy "; " input normalizes
    assert bi.merge_sources("NCERT; RD Sharma", "ncert") == "NCERT, RD Sharma"
    assert bi.merge_sources("", "Arihant") == "Arihant"
    assert bi.merge_sources("S Chand", "") == "S Chand"


def test_vocab_exposes_book_sources(client):
    v = client.get("/directory/vocab").json()
    assert "NCERT" in v["book_sources"]
    assert "RD Sharma" in v["book_sources"]
    # Maharashtra State Board's own textbook imprint.
    assert "Balbharati" in v["book_sources"]


def test_legacy_workbook_without_concept_source_still_imports(client, db, tmp_path):
    """Old-layout files (no concept_source column) must not mis-align bands.

    The header rows are built from the FROZEN registry entry, never from
    ``writer._write_headers``: that writer migrates to the reference layout in
    S7 and a fixture built from it would follow the writer instead of pinning
    the legacy layout this test exists for. All three sheets carry their real
    header — the reader's layout gate identifies a workbook, not a sheet, and
    a one-column stub is not a layout.
    """
    legacy = layouts.layout("canonical-legacy-concept-band")
    legacy_fields = list(legacy.sheet("objective").fields)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for kind in ("objective", "subjective", "descriptive"):
        sheet_layout = legacy.sheet(kind)
        ws = wb.create_sheet(sheet_layout.sheet_name)
        ws.append(["Chapter"])  # band row (content irrelevant)
        ws.append(list(sheet_layout.fields))

    ws = wb[legacy.sheet("objective").sheet_name]
    row = [""] * len(legacy_fields)
    # Group band now has 8 fields (added group_question_labels) -> question
    # band starts one column later than the old layout.
    row[0] = "Legacy Chapter (10CBMA_Legacy)"
    row[6] = "Legacy Topic"
    row[12] = "Legacy Concept Unique XYZ"
    row[21] = "10CBMA_Lgcy_PL_T01_X Q99"   # concept_question_labels (group label)
    row[26] = "Basic"                       # group_type
    row[29] = "10CBMA_Lgcy_PL_T01_X Q99"   # question-band label
    row[30] = "Multiple Choice Question"
    row[31] = "Remembering"
    row[32] = "NCERT"
    row[37] = "Legacy unique question text 9871?"
    row[38] = 1
    ws.append(row)
    path = tmp_path / "legacy.xlsx"
    wb.save(path)

    files = {"file": ("legacy.xlsx", io.BytesIO(path.read_bytes()),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    counts = client.post("/data/import", files=files).json()
    assert counts["questions"] == 1

    q = db.query(models.Question).filter_by(
        question_label="10CBMA_Lgcy_PL_T01_X Q99").one()
    assert q.question == "Legacy unique question text 9871?"
    assert q.group.concept.concept_title == "Legacy Concept Unique XYZ"
    assert q.group.group_type == "Basic"


def test_concept_resused_across_books_merges_sources(
    client, db, first_chapter, monkeypatch,
):
    """Same concept from a second book: not duplicated, sources accumulate."""
    _use_specific_dry_learner_analysis(monkeypatch)
    body = (b"## Optics Basics\n"
            b"Refraction of light through glass slabs\n"
            b"Total internal reflection in prisms")

    from tests.conftest import convert_concept_upload, stream_result

    def upload_and_generate(book):
        files = {"file": (f"{book.replace(' ', '_')}.txt", io.BytesIO(body), "text/plain")}
        job = client.post(
            f"/build-concepts/post-learning/uploads?source_book={book}", files=files,
        ).json()
        assert job["source_book"] == book
        convert_concept_upload(client, job["id"])
        return stream_result(client.post(
            f"/build-concepts/post-learning/uploads/{job['id']}/generate",
            json={"target_chapter_id": first_chapter["id"]}))

    first = upload_and_generate("NCERT")
    assert first["concepts_created"] == 3
    assert first["concepts_merged"] == 0

    second = upload_and_generate("RD Sharma")
    assert second["concepts_created"] == 0
    assert second["concepts_merged"] == 3

    c = (db.query(models.Concept)
         .filter(models.Concept.concept_title.like("Refraction of light%")).one())
    assert c.sources == "NCERT, RD Sharma"


def test_duplicate_questions_across_books_merge_sources(client, db, first_chapter):
    """Same question text from another book: skipped, question_source merged."""
    body = (b"# Qs\n\n"
            b"State the law of refraction with one worked example 4417.\n\n"
            b"Define critical angle for a glass-air interface 4417.")

    from tests.conftest import convert_assessment_upload, stream_result

    def run(book):
        files = {"file": (f"q_{book.replace(' ', '_')}.txt", io.BytesIO(body), "text/plain")}
        job = client.post(
            f"/build-assessments/uploads?upload_type=questions&source_book={book}",
            files=files,
        ).json()
        convert_assessment_upload(client, job["id"])
        client.post(f"/build-assessments/uploads/{job['id']}/deposit", json={
            "scope_type": "chapter", "scope_ids": [first_chapter["id"]],
        })
        return stream_result(client.post(
            f"/build-assessments/uploads/{job['id']}/generate",
            json={"question_type": "objective"}))

    first = run("S Chand")
    assert first["created"] == 2
    assert first["duplicates_merged"] == 0

    second = run("Arihant")
    assert second["created"] == 0
    assert second["duplicates_merged"] == 2

    q = (db.query(models.Question)
         .filter(models.Question.question.like("State the law of refraction%")).one())
    assert q.question_source == "S Chand, Arihant"


def test_output_workbook_source_cells_update_in_place(db, tmp_path, client, first_chapter):
    """Re-appending an existing concept refreshes its concept_source cell."""
    detail = client.get(f"/directory/chapters/{first_chapter['id']}").json()
    concept_id = detail["topics"][0]["concepts"][0]["id"]
    concept = db.get(models.Concept, concept_id)
    concept.sources = "NCERT"
    db.commit()

    path = tmp_path / "out.xlsx"
    first = writer.append_concepts(db, path, [concept_id])
    assert first["written"] >= 1

    concept.sources = bi.merge_sources(concept.sources, "RS Aggarwal")
    db.commit()
    db.expire_all()

    second = writer.append_concepts(db, path, [concept_id])
    assert second["written"] == 0
    assert second["sources_updated"] >= 1

    wb = openpyxl.load_workbook(path)
    ws = wb[bi.SHEET_OBJECTIVE]
    src_col = bi.OBJECTIVE_FIELDS.index("concept_source")
    values = {
        writer._cell_str(row, src_col)
        for row in ws.iter_rows(min_row=3, values_only=True)
        if bi.strip_title_tag(writer._cell_str(row, 12)) == concept.concept_title
    }
    assert "NCERT, RS Aggarwal" in values
