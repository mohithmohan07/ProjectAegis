"""MES PR 5 — Concept and Master renderers from one release snapshot.

The uploaded Grade-6 MES FINAL templates are the positional authority;
``assessment_workbook_template.json`` captures them verbatim and these tests pin
both renderers to it (spec §0 item 2, §1, §9, §10, §13).
"""
from __future__ import annotations

import copy
import io

import openpyxl
import pytest

from app.bulk_import import assessment_workbook as mp
from app.services import assessment_grouping as ag
from app.services import assessment_profile
from app.services import assessment_release as rel


def _snapshot() -> dict:
    chapter = {
        "chapter_title": (
            "Three-Dimensional Shapes (06_Mathematics_MSBSHSE_Balbharati)"),
        "chapter_display_name": "Three-Dimensional Shapes",
        "pre_topics": "", "post_topics": "Topic 01",
        "chapter_description": "Solids and their properties.",
    }
    concept_a = {
        "concept_key": "C_A",
        "concept_title": "Two-dimensional shape (06MSMA_T01_TwoDim)",
        "concept_display_name": "Two-dimensional shape",
        "concept_details": "Description: flat shapes.",
        "keywords": "shape; plane",
        "related_concepts": "",
        "digicards": "",
        "concept_source": "Balbharati",
    }
    concept_b = {  # deliberately questionless
        "concept_key": "C_B",
        "concept_title": "Three-dimensional shape (06MSMA_T01_ThreeDim)",
        "concept_display_name": "Three-dimensional shape",
        "concept_details": "Description: solid shapes.",
        "keywords": "solid",
        "related_concepts": "",
        "digicards": "",
        "concept_source": "Balbharati",
    }
    groups = []
    for key, machine, name in (
        ("C_A", "06MSMA_T01_TwoDim", "Two-dimensional shape"),
        ("C_B", "06MSMA_T01_ThreeDim", "Three-dimensional shape"),
    ):
        for shell in ag.required_shells(key, machine, name):
            shell = dict(shell)
            shell["concept_key"] = key
            groups.append(shell)
    objective = {
        "candidate_id": "CAND-1",
        "question_label": "06MSMA_T01_TwoDim Q01",
        "sheet_kind": "objective",
        "question_category": "Multiple Choice Question",
        "cognitive_skill": "Remember",
        "difficulty": "Less",
        "marks": 1.0,
        "question_duration": 1.0,
        "math_keyboard": "",
        "question_appears_in": "Pre/Post-Worksheet/Test",
        "answer_restriction": "Specific",
        "question": "Which of these is a two-dimensional shape?",
        "question_text": "Which of these is a two-dimensional shape?",
        "answers": [
            {"answer_type": "Phrases", "answer_content": "Circle",
             "correct_answer": "1", "answer_weightage": "1"},
            {"answer_type": "Phrases", "answer_content": "Sphere",
             "correct_answer": "0", "answer_weightage": "0"},
        ],
        "sub_questions": [],
        "answer_explanation": "A circle is flat; a sphere is a solid.",
        "concept_key": "C_A",
        "group_key": "(06MSMA_T01_TwoDim) BG01",
        "flags": [],
    }
    descriptive = {
        "candidate_id": "CAND-2",
        "question_label": "06MSMA_T01_TwoDim Q02",
        "sheet_kind": "descriptive",
        "question_category": "Long Answer Type (4 Marks)",
        "cognitive_skill": "Understand",
        "difficulty": "Moderate",
        "marks": 4.0,
        "question_duration": 6.0,
        "math_keyboard": "No",
        "question_appears_in": "Pre/Post-Worksheet/Test",
        "answer_restriction": "Open",
        "question": "Explain how a square differs from a cube.",
        "question_text": "Explain how a square differs from a cube.",
        "display_answer": "A square is flat with two dimensions ...",
        # Multipart items keep all scoring evidence with their parts.  Main
        # answer blocks would duplicate the same four marks.
        "answers": [],
        "sub_questions": [
            {"text": "Name the number of faces of a cube.", "marks": "2",
             "keywords": [
                 {"answer_type": "Phrases", "weightage": "2",
                  "keyword": "[content]: six faces"}]},
            {"text": "State the dimensions of a square.", "marks": "2",
             "keywords": [
                 {"answer_type": "Phrases", "weightage": "2",
                  "keyword": "[content]: two dimensions"}]},
        ],
        "answer_explanation": "",
        "concept_key": "C_A",
        "group_key": "(06MSMA_T01_TwoDim) IG01",
        "flags": [],
    }
    return {
        "chapter": chapter,
        "topics": [{
            "topic_title": "Topic 01: Dimensions (06MSMA_ThreeDim_PL)",
            "topic_display_name": "Dimensions",
            "pre_post_learning": "Post",
            "topic_concept_labels": "",
            "related_topics": "",
            "topic_description": "Dimensions of shapes.",
            "concepts": [concept_a, concept_b],
        }],
        "groups": groups,
        "candidates": [objective, descriptive],
    }


def test_manifest_matches_the_uploaded_templates():
    assert mp.SHEET_ORDER == ["Objective", "Descriptive", "Subjective"]
    assert len(mp.FIELDS["Objective"]) == 67
    assert len(mp.FIELDS["Descriptive"]) == 374
    assert len(mp.FIELDS["Subjective"]) == 144
    obj = mp.FIELDS["Objective"]
    # answer_restriction sits between appears_in and difficulty; question_text
    # directly after question; answer_explanation is the last column.
    assert obj[obj.index("question_appears_in") + 1] == "answer_restriction"
    assert obj[obj.index("answer_restriction") + 1] == "level_of_difficulty"
    assert obj[obj.index("question") + 1] == "question_text"
    assert obj[-1] == "answer_explanation"
    # Concept band keeps keywords/related_concepts; no parent_concept.
    assert "keywords" in obj and "related_concepts" in obj
    assert "parent_concept" not in obj
    # Subjective carries twenty answer blocks.
    assert "placeholder_20" in mp.FIELDS["Subjective"]


def test_dual_output_shares_one_snapshot_and_validates():
    result = mp.build_dual_output(_snapshot())
    assert result["valid"], result["manifest"]["read_back"]
    manifest = result["manifest"]
    assert manifest["concept_snapshot_sha256"] == mp.snapshot_sha256(
        _snapshot())
    assert set(manifest["workbook_sha256s"]) == {
        "concepts_xlsx", "master_xlsx"}
    assert manifest["issues"]["unplaced"] == []
    assert manifest["issues"]["placed_questions"] == 2


def test_concept_file_is_a_clean_catalogue():
    parsed = mp.parse_workbook(mp.render_concept_file(_snapshot()))
    assert parsed["sheet_order"] == ["Objective", "Descriptive", "Subjective"]
    rows = parsed["sheets"]["Objective"]["rows"]
    assert [r["concept_title"] for r in rows] == [
        "Two-dimensional shape (06MSMA_T01_TwoDim)",
        "Three-dimensional shape (06MSMA_T01_ThreeDim)",
    ]
    for row in rows:
        assert row["group_name"] == "" and row["question_label"] == ""
        assert row["basic_groups"] == "" and row["concept_question_labels"] == ""
        assert row["chapter_duration"] == ""
        # Identity values survive byte-exact, spelling variants included.
        assert row["chapter_title"].endswith("(06_Mathematics_MSBSHSE_Balbharati)")
        assert row["keywords"] != "" or row["concept_title"].startswith("Three")
    assert parsed["sheets"]["Descriptive"]["rows"] == []
    assert parsed["sheets"]["Subjective"]["rows"] == []


def test_master_contains_everything_including_questionless_concepts():
    master, issues = mp.render_master_file(_snapshot())
    parsed = mp.parse_workbook(master)
    assert issues["unplaced"] == []
    objective_rows = parsed["sheets"]["Objective"]["rows"]
    descriptive_rows = parsed["sheets"]["Descriptive"]["rows"]
    assert parsed["sheets"]["Subjective"]["rows"] == []
    group_keys = {group["group_key"] for group in _snapshot()["groups"]}
    for row in [*objective_rows, *descriptive_rows]:
        assert row["group_name"] == row["group_display_name"]
        # Q16 (SOP §6.1): the visible group name IS the group ID; the
        # questionless tail row keeps its group band empty.
        if str(row.get("group_name") or "").strip():
            assert row["group_name"] in group_keys
    provenance = {
        (item["sheet"], item["row"]): item["group_key"]
        for item in issues["group_provenance"]
    }
    assert provenance[("Objective", 3)] == "(06MSMA_T01_TwoDim) BG01"
    assert provenance[("Descriptive", 3)] == "(06MSMA_T01_TwoDim) IG01"
    assert "(06MSMA_T01_TwoDim) AG01" not in provenance.values()
    assert parsed["sheets"]["Objective"]["row_numbers"] == [3, 4]
    assert parsed["sheets"]["Descriptive"]["row_numbers"] == [3]

    # Empty tier shells remain in the snapshot but do not become assessment
    # rows.  The sole catalogue row is the one OD5 tail row for the genuinely
    # questionless concept.
    question_rows = [r for r in objective_rows if r["question_label"]]
    catalogue_rows = [r for r in objective_rows if not r["question_label"]]
    assert len(question_rows) == 1
    assert len(catalogue_rows) == 1
    tail = [
        r for r in catalogue_rows
        if r["concept_title"].startswith("Three-dimensional")]
    assert len(tail) == 1
    assert tail[0] is catalogue_rows[-1]
    assert all(
        not str(tail[0].get(field) or "").strip()
        for field in (
            "group_name", "group_display_name", "group_description",
            "group_status", "group_type", "group_question_labels",
            "related_digicards", "concept_question_labels", "basic_groups",
            "intermediate_groups", "advanced_groups",
        )
    )
    assert [
        item["concept_title"] for item in issues["questionless_concepts"]
    ] == ["Three-dimensional shape (06MSMA_T01_ThreeDim)"]
    assert sorted(
        issues["questionless_concepts"][0]["shell_group_keys"]) == [
        "(06MSMA_T01_ThreeDim) AG01",
        "(06MSMA_T01_ThreeDim) BG01",
        "(06MSMA_T01_ThreeDim) IG01",
    ]
    # The shells stay in the payload; the difference is RECORDED, not silent.
    assert {g["group_key"] for g in _snapshot()["groups"]} >= set(
        issues["questionless_concepts"][0]["shell_group_keys"])
    assert sorted(issues["omitted_empty_group_shells"]) == [
        "(06MSMA_T01_ThreeDim) AG01",
        "(06MSMA_T01_ThreeDim) BG01",
        "(06MSMA_T01_ThreeDim) IG01",
        "(06MSMA_T01_TwoDim) AG01",
    ]

    q = question_rows[0]
    assert q["question_appears_in"] == "Pre/Post-Worksheet/Test"
    assert q["answer_restriction"] == "Specific"
    assert q["answer_content_1"] == "Circle"
    # Rendered in the CMS's Yes/No spelling (owner ruling, 2026-08-21);
    # the wire "1"/"0" and "Yes"/"No" are both accepted on read-back.
    assert str(q["correct_answer_1"]) == "Yes"
    assert q["group_question_labels"] == "06MSMA_T01_TwoDim Q01"
    assert q["concept_question_labels"] == (
        "06MSMA_T01_TwoDim Q01, 06MSMA_T01_TwoDim Q02")
    assert q["basic_groups"] == "(06MSMA_T01_TwoDim) BG01"
    assert q["intermediate_groups"] == "(06MSMA_T01_TwoDim) IG01"
    assert q["advanced_groups"] == ""

    d = descriptive_rows[0]
    assert d["sub_question_1"] == "Name the number of faces of a cube."
    assert d["sq1_keyword_1"] == "[content]: six faces"
    assert d["sq2_keyword_1"] == "[content]: two dimensions"
    assert d["display_answer"].startswith("A square is flat")
    assert d["answer_restriction"] == "Open"
    assert d["answer_content_1"] == ""
    assert d["basic_groups"] == "(06MSMA_T01_TwoDim) BG01"
    assert d["intermediate_groups"] == "(06MSMA_T01_TwoDim) IG01"
    assert d["advanced_groups"] == ""


def test_unresolved_placement_rides_the_manifest_never_disappears():
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["group_key"] = "(unknown) BG01"
    master, issues = mp.render_master_file(snapshot)
    assert issues["unplaced"][0]["candidate_id"] == "CAND-1"
    parsed = mp.parse_workbook(master)
    labels = [
        r["question_label"] for r in parsed["sheets"]["Objective"]["rows"]
        if r["question_label"]
    ]
    assert labels == []  # not silently placed anywhere
    errors = mp.validate_master_file(parsed, snapshot)
    assert errors == []  # structure stays valid; the ledger carries the issue


def test_master_validation_names_wire_value_and_arithmetic_defects():
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["question_appears_in"] = (
        "Pre-test, Post-test, Worksheet, Test")  # expanded = wrong for MES
    snapshot["candidates"][0]["answers"][0]["answer_weightage"] = "3"
    master, _ = mp.render_master_file(snapshot)
    errors = mp.validate_master_file(mp.parse_workbook(master), snapshot)
    assert any("not the profile wire value" in e for e in errors)
    assert any("correct weightage 3 != marks 1" in e for e in errors)


@pytest.mark.parametrize(
    ("replacement", "expected"),
    [
        pytest.param(
            "", "populated Question band requires a non-blank question_label",
            id="blank-label",
        ),
        pytest.param(
            "ROGUE Q99", "question label row set/order differs",
            id="rogue-label",
        ),
    ],
)
def test_master_readback_reconciles_question_row_identity(
    replacement: str, expected: str,
):
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    row = next(
        item for item in parsed["sheets"]["Objective"]["rows"]
        if item.get("question_label")
    )
    row["question_label"] = replacement

    errors = mp.validate_master_file(
        parsed,
        copy.deepcopy(snapshot),
        group_provenance=issues["group_provenance"],
    )

    assert any(expected in error for error in errors), errors
    assert any(
        "question label row set/order differs" in error for error in errors
    ), errors


def test_master_readback_refuses_a_labelled_blank_question():
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    row = next(
        item for item in parsed["sheets"]["Objective"]["rows"]
        if item.get("question_label")
    )
    row["question"] = ""

    errors = mp.validate_master_file(
        parsed,
        copy.deepcopy(snapshot),
        group_provenance=issues["group_provenance"],
    )

    assert any("question must not be blank" in error for error in errors)


def test_master_readback_rejects_duplicate_concept_tail_rows():
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    sheet = parsed["sheets"]["Objective"]
    tail = next(
        item for item in sheet["rows"]
        if not any(
            str(item.get(field) or "").strip()
            for field in mp.FIELDS["Objective"][
                mp._INDEX["Objective"]["question_label"]:
            ]
        )
    )
    sheet["rows"].append(copy.deepcopy(tail))
    sheet["row_numbers"].append(max(sheet["row_numbers"]) + 1)

    errors = mp.validate_master_file(
        parsed,
        copy.deepcopy(snapshot),
        group_provenance=issues["group_provenance"],
    )

    assert any(
        "concept-only tail row set/order differs" in error
        for error in errors
    ), errors


def test_master_readback_rejects_questionless_group_scaffolding():
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    tail = next(
        item for item in parsed["sheets"]["Objective"]["rows"]
        if not item.get("question_label")
    )
    group = next(
        item for item in snapshot["groups"]
        if item["concept_key"] == "C_B" and item["group_type"] == "Basic"
    )
    tail.update({
        "group_name": group["group_name"],
        "group_display_name": group["group_display_name"],
        "group_description": group["semantic_description"],
        "group_status": group["group_status"],
        "group_type": group["group_type"],
        "group_question_labels": "",
    })

    errors = mp.validate_master_file(
        parsed,
        copy.deepcopy(snapshot),
        group_provenance=issues["group_provenance"],
    )

    assert any(
        "without a Question band must also keep the Group band blank" in error
        for error in errors
    ), errors


@pytest.mark.parametrize(
    "malformed_field", ["answers", "sub_questions", "keywords"],
)
def test_master_projection_remains_total_for_non_array_containers(
    malformed_field: str,
):
    snapshot = _snapshot()
    candidate = (
        snapshot["candidates"][0]
        if malformed_field == "answers"
        else snapshot["candidates"][1]
    )
    if malformed_field == "keywords":
        candidate["sub_questions"][0]["keywords"] = 1
        ledger_field = "sub_questions[1].keywords"
    else:
        candidate[malformed_field] = 1
        ledger_field = malformed_field

    findings = rel.unresolved_question_homes(snapshot)
    master, issues = mp.render_master_file(snapshot)

    assert master[:2] == b"PK"
    assert isinstance(findings, list)
    assert any(
        entry.get("field") == ledger_field
        and entry.get("reason") == "invalid_container_shape"
        for entry in issues["truncated_rows"]
    ), issues["truncated_rows"]


def test_master_provenance_distinguishes_shared_friendly_group_names():
    snapshot = copy.deepcopy(_snapshot())
    first_group = next(
        group for group in snapshot["groups"]
        if group["group_key"] == "(06MSMA_T01_TwoDim) BG01"
    )
    first_group["group_name"] = "Two-dimensional shape — Basic"
    first_group["group_display_name"] = "Two-dimensional shape — Basic"
    second_group = copy.deepcopy(first_group)
    second_group["group_key"] = "(06MSMA_T01_TwoDim) BG02"
    second_group["semantic_description"] = "A second Basic variant family."
    snapshot["groups"].append(second_group)

    second_question = copy.deepcopy(snapshot["candidates"][0])
    second_question["candidate_id"] = "CAND-3"
    second_question["question_label"] = "06MSMA_T01_TwoDim Q03"
    second_question["question"] = "Which listed figure is flat?"
    second_question["question_text"] = "Which listed figure is flat?"
    second_question["group_key"] = "(06MSMA_T01_TwoDim) BG02"
    snapshot["candidates"].append(second_question)

    master, issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    friendly_rows = [
        (row_number, row)
        for row_number, row in zip(
            parsed["sheets"]["Objective"]["row_numbers"],
            parsed["sheets"]["Objective"]["rows"],
        )
        if row["group_name"] == "Two-dimensional shape — Basic"
    ]
    assert [row["group_display_name"] for _, row in friendly_rows] == [
        "Two-dimensional shape — Basic",
        "Two-dimensional shape — Basic",
    ]
    assert [row["group_question_labels"] for _, row in friendly_rows] == [
        "06MSMA_T01_TwoDim Q01",
        "06MSMA_T01_TwoDim Q03",
    ]

    provenance = {
        (item["sheet"], item["row"]): item["group_key"]
        for item in issues["group_provenance"]
    }
    assert [
        provenance[("Objective", row_number)]
        for row_number, _ in friendly_rows
    ] == [
        "(06MSMA_T01_TwoDim) BG01",
        "(06MSMA_T01_TwoDim) BG02",
    ]
    assert mp.validate_master_file(
        parsed, snapshot,
        group_provenance=issues["group_provenance"],
    ) == []
    assert any(
        "group provenance required" in error
        for error in mp.validate_master_file(parsed, snapshot)
    )

    result = mp.build_dual_output(snapshot)
    assert result["valid"], result["manifest"]["read_back"]


def test_equation_answer_and_keyword_cells_are_raw():
    """Q21 supersedes the older wrapped answer_content interpretation."""
    answer_snapshot = copy.deepcopy(_snapshot())
    descriptive = answer_snapshot["candidates"][1]
    descriptive["sub_questions"] = []
    descriptive["answers"] = [{
        "answer_type": "Equation", "answer_weightage": "2",
        "answer_content": "1 mark: States [Katex] 3n=12 [/Katex].",
    }, {
        "answer_type": "Phrases", "answer_weightage": "2",
        "answer_content": "[method]: Completes the comparison.",
    }]
    master, _ = mp.render_master_file(answer_snapshot)
    row = mp.parse_workbook(master)["sheets"]["Descriptive"]["rows"][0]
    assert row["answer_content_1"] == (
        r"\text{1 mark: States }3n=12\text{.}"
    )

    keyword_snapshot = copy.deepcopy(_snapshot())
    descriptive = keyword_snapshot["candidates"][1]
    descriptive["sub_questions"][0]["keywords"][0] = {
        "answer_type": "Equation", "weightage": "2",
        "keyword": "[Katex] 3n=3\\times4=12 [/Katex]",
    }
    master, _ = mp.render_master_file(keyword_snapshot)
    row = mp.parse_workbook(master)["sheets"]["Descriptive"]["rows"][0]
    assert row["sq1_keyword_1"] == "3n=3\\times4=12"


def test_equation_option_is_raw_in_cell_but_wrapped_in_lowercase_question_text():
    snapshot = copy.deepcopy(_snapshot())
    objective = snapshot["candidates"][0]
    objective["answers"][0].update({
        "answer_type": "Equation",
        "answer_content": "[Katex]x=2[/Katex]",
    })

    master, _ = mp.render_master_file(snapshot)
    row = mp.parse_workbook(master)["sheets"]["Objective"]["rows"][0]

    assert row["answer_content_1"] == "x=2"
    assert "a) [Katex] x=2 [/Katex]" in row["question_text"]
    assert "b) Sphere" in row["question_text"]


def test_master_renderer_projects_screenshot_shaped_markdown_table_cells():
    snapshot = copy.deepcopy(_snapshot())
    objective = snapshot["candidates"][0]
    source = "\n".join([
        "Study the following table:",
        "| Name of peak | Altitude (in metres) |",
        "|:---|---:|",
        "| K-2 | 8611 |",
        "| Lao Tse | 8516 |",
        "| Mount Everest (Sagarmatha) | 8849 |",
        "| Makalu | 8485 |",
        "| Kanchanjunga | 8586 |",
    ])
    objective["question"] = source
    objective["question_text"] = source

    master, _ = mp.render_master_file(snapshot)
    row = mp.parse_workbook(master)["sheets"]["Objective"]["rows"][0]

    for field in ("question", "question_text"):
        assert "Table row 1, column 1: Name of peak" in row[field]
        assert "Table row 6, column 2: 8586" in row[field]
        assert "|:---|---:|" not in row[field]
    assert "a) Circle" in row["question_text"]


def test_master_records_a_duplicate_group_key_instead_of_raising():
    """INVERTED by spec-step8 T7.5/B4.

    The raise cost every row on every sheet of all four outputs while the
    same defect was already refused at freeze by
    ``rel.duplicate_group_keys`` and again by ``validate_master_file``'s
    read-back. The workbook is now written and the defect is recorded.
    """
    snapshot = copy.deepcopy(_snapshot())
    duplicate = copy.deepcopy(snapshot["groups"][0])
    duplicate["concept_key"] = "C_B"
    duplicate["group_type"] = "Advanced"
    duplicate["group_name"] = "Three-dimensional shape — Advanced"
    duplicate["group_display_name"] = duplicate["group_name"]
    snapshot["groups"].append(duplicate)
    master, issues = mp.render_master_file(snapshot)
    assert master
    assert any(
        "duplicate group_key" in item["message"]
        for item in issues["group_defects"])
    # The staging twin, unchanged and still refusing at freeze.
    assert rel.duplicate_group_keys(snapshot["groups"]) == [
        "(06MSMA_T01_TwoDim) BG01"]
    # And the read-back twin.
    assert any(
        "duplicate group_key" in error
        for error in mp.validate_master_file(
            mp.parse_workbook(master), snapshot))


def test_a_non_sop_visible_group_name_is_a_named_defect_not_a_render_error():
    # Q16 inverted this pin: the retired friendly "Concept — Tier" name is
    # now the mismatch; the group ID itself is the only conforming value.
    snapshot = copy.deepcopy(_snapshot())
    group = snapshot["groups"][0]
    group["group_name"] = "Two-dimensional shape — Basic"
    group["group_display_name"] = "Two-dimensional shape — Basic"
    master, _ = mp.render_master_file(snapshot)
    assert master
    codes = {
        f["code"] for f in rel.unresolved_question_homes(snapshot)}
    assert rel.GROUP_VISIBLE_NAME_MISMATCH in codes

    missing_display = copy.deepcopy(_snapshot())
    missing_display["topics"][0]["concepts"][0][
        "concept_display_name"
    ] = ""
    master, _ = mp.render_master_file(missing_display)
    assert master
    assert rel.GROUP_HOME_UNNAMED in {
        f["code"] for f in rel.unresolved_question_homes(missing_display)}


def test_a_group_home_disagreement_is_a_defect_not_a_render_error():
    """INVERTED by spec-step8 T7.5/B4 — the ``:373`` raise.

    Today it cost all four workbooks; now all four are written, the
    candidate reaches no data row, and the defect is NAMED at staging so
    the database write is refused.
    """
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["concept_key"] = "C_B"
    master, issues = mp.render_master_file(snapshot)
    assert master
    assert [item["question_label"] for item in issues["unplaced"]] == [
        "06MSMA_T01_TwoDim Q01"]
    findings = rel.unresolved_question_homes(snapshot)
    assert [f["code"] for f in findings] == [rel.GROUP_HOME_DISAGREEMENT]
    assert findings[0]["question_label"] == "06MSMA_T01_TwoDim Q01"
    # And the label reaches NO data row.
    parsed = mp.parse_workbook(master)
    for sheet in parsed["sheets"].values():
        assert all(
            row.get("question_label") != "06MSMA_T01_TwoDim Q01"
            for row in sheet["rows"])


def test_master_provenance_uses_physical_sheet_row_numbers():
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    workbook = openpyxl.load_workbook(io.BytesIO(master))
    workbook["Objective"].insert_rows(3)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    parsed = mp.parse_workbook(buffer.getvalue())
    assert parsed["sheets"]["Objective"]["row_numbers"] == [4, 5]
    assert parsed["sheets"]["Descriptive"]["row_numbers"] == [3]
    shifted_provenance = copy.deepcopy(issues["group_provenance"])
    for item in shifted_provenance:
        if item["sheet"] == "Objective":
            item["row"] += 1
    assert mp.validate_master_file(
        parsed, snapshot, group_provenance=shifted_provenance,
    ) == []


def test_formula_injection_and_cell_limit_guards():
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["question"] = "=HYPERLINK evil"
    master, _ = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    q = [r for r in parsed["sheets"]["Objective"]["rows"]
         if r["question_label"]][0]
    assert q["question"] == "=HYPERLINK evil"

    # INVERTED by spec-step8 S9, and inverted deliberately rather than
    # deleted: this assertion pinned a REAL behaviour that is now wrong.
    # [measured at 76c84fb] a 40 000-character question raised
    # ``WorkbookRenderError`` out of ``render_master_file`` — costing every
    # row on every sheet of ALL FOUR outputs for one cell — while the
    # staged verdict returned ``[]``, so nothing refused the database write
    # either. Rule E says a defect blocks the write, never a download. The
    # new contract is asserted here in full, and the coverage this test had
    # is strengthened, not weakened: it now pins the workbook, the
    # truncation marker, the whole recorded value, and the staged refusal.
    oversized = copy.deepcopy(_snapshot())
    full_text = "x" * (mp.CELL_LIMIT + 1)
    oversized["candidates"][0]["question"] = full_text
    master, issues = mp.render_master_file(oversized)
    assert master, "the Master File is written, not refused"

    recorded = [
        row for row in issues["oversized_cells"]
        if row["reason"] == mp.CELL_TEXT_TOO_LONG
    ]
    assert recorded, "the oversized cell is recorded, never silently cut"
    # R4: the SILENCE is the defect. The complete value is where a
    # reviewer reads it, unabridged.
    assert recorded[0]["full_value"] == full_text
    assert recorded[0]["actual"] == len(full_text)
    assert recorded[0]["cap"] == mp.CELL_LIMIT

    # And the cell itself says the tail is elsewhere.
    written = [
        row for row in mp.parse_workbook(master)["sheets"]["Objective"]["rows"]
        if row["question_label"]
    ][0]["question"]
    assert len(written) <= mp.CELL_LIMIT
    assert "the complete value is recorded" in written.lower()

    # The database write is refused, by name, at STAGING.
    findings = [
        f for f in rel.unresolved_question_homes(oversized)
        if f["code"] == rel.RENDER_SHAPE_OVERFLOW
    ]
    assert [f["field"] for f in findings] == ["question"]
    assert findings[0]["question_label"] == "06MSMA_T01_TwoDim Q01"


def test_staged_shape_uses_the_selected_descriptive_answer_capacity():
    snapshot = copy.deepcopy(_snapshot())
    descriptive = snapshot["candidates"][1]
    descriptive["sub_questions"] = []
    descriptive["answers"] = [{
        "answer_type": "Phrases",
        "answer_content": f"[content]: criterion {number}",
        "answer_weightage": "",
    } for number in range(1, 12)]
    english_profile = assessment_profile.resolve_for_metadata(None, {
        "board": "MSBSHSE",
        "grade": "6",
        "subject": "English",
    })

    post_findings = rel.unresolved_question_homes(snapshot, english_profile)
    assert not [
        finding for finding in post_findings
        if finding.get("candidate_id") == descriptive["candidate_id"]
        and finding.get("field") == "answers"
    ]

    # The staging audit must inspect every widened cell too, not merely stop
    # counting at the historical ten-slot layout.
    descriptive["answers"][10]["answer_content"] = "x" * (mp.CELL_LIMIT + 1)
    widened_cell_findings = rel.unresolved_question_homes(
        snapshot, english_profile,
    )
    assert [
        finding["field"] for finding in widened_cell_findings
        if finding.get("candidate_id") == descriptive["candidate_id"]
        and finding.get("field") == "answer_content_11"
    ] == ["answer_content_11"]

    descriptive["answers"][10]["answer_content"] = "[content]: criterion 11"
    snapshot["topics"][0]["pre_post_learning"] = "Pre"
    pre_overflows = [
        finding for finding in rel.unresolved_question_homes(
            snapshot, english_profile,
        )
        if finding.get("candidate_id") == descriptive["candidate_id"]
        and finding.get("field") == "answers"
    ]
    assert [(finding["cap"], finding["actual"]) for finding in pre_overflows] == [
        (10, 11),
    ]


def test_a_control_character_is_repaired_not_raised():
    """The second half of the ``render_shape_overflow`` family (S9).

    [measured at 76c84fb] a single ``\\x01`` in a question reached openpyxl
    and raised ``IllegalCharacterError`` out of ``wb.save`` — an exception
    from a third-party library, with no defect name, no record, and all
    four outputs gone. XML 1.0 forbids those code points; the cell is
    repaired, the whole original is recorded, and the write is refused.
    """
    ctl = copy.deepcopy(_snapshot())
    ctl["candidates"][0]["question"] = "bad\x01char"

    master, issues = mp.render_master_file(ctl)
    assert master, "the Master File is written, not refused"

    recorded = [
        row for row in issues["oversized_cells"]
        if row["reason"] == mp.CELL_TEXT_ILLEGAL_CHARACTER
    ]
    assert recorded, "the illegal character is recorded"
    assert recorded[0]["full_value"] == "bad\x01char"
    assert recorded[0]["code_points"] == ["U+0001"]

    written = [
        row for row in mp.parse_workbook(master)["sheets"]["Objective"]["rows"]
        if row["question_label"]
    ][0]["question"]
    assert written == "badchar"

    findings = [
        f for f in rel.unresolved_question_homes(ctl)
        if f["code"] == rel.RENDER_SHAPE_OVERFLOW
    ]
    assert [f["reason"] for f in findings] == [
        mp.CELL_TEXT_ILLEGAL_CHARACTER
    ]


def test_a_cell_that_is_both_too_long_and_illegal_is_repaired_for_both():
    """The compound case — S9's own failure mode, still reachable until now.

    [measured on the first cut of S9] ``cell_text_defect`` returned the
    FIRST reason only. A value that was over the cap AND carried an
    XML-illegal code point reported ``cell_text_too_long``, ``_cell_value``
    dispatched on that one reason, truncated, and never stripped — so the
    illegal code point rode into openpyxl and

        aw.render_master_file(snap)
        -> openpyxl.utils.exceptions.IllegalCharacterError

    took all four outputs, the exact loss this slice exists to end. It was
    data-position-dependent: the same defect at the tail was truncated away
    by luck, so one chapter shipped and the next did not. Staging named the
    length only, so a reviewer reading the findings could not have known
    what actually broke the file.

    Both halves are asserted, on the cell and at staging, plus the two cap
    arithmetic facts the same repair path owns.
    """
    both = copy.deepcopy(_snapshot())
    full_text = "\x01" + "y" * (mp.CELL_LIMIT + 5000)
    both["candidates"][0]["question"] = full_text

    master, issues = mp.render_master_file(both)
    assert master, "the Master File is written, not refused"
    assert mp.build_dual_output(both)["valid"], "and it reads back clean"

    reasons = [
        row["reason"] for row in issues["oversized_cells"]
        if row["context"].endswith(":question")
    ]
    assert reasons == [mp.CELL_TEXT_TOO_LONG, mp.CELL_TEXT_ILLEGAL_CHARACTER]
    assert all(
        row["full_value"] == full_text for row in issues["oversized_cells"]
        if row["context"].endswith(":question")
    ), "both entries carry the WHOLE value; neither is a partial record"

    written = [
        row for row in mp.parse_workbook(master)["sheets"]["Objective"]["rows"]
        if row["question_label"]
    ][0]["question"]
    assert "\x01" not in written
    assert len(written) <= mp.CELL_LIMIT

    # Staging names BOTH, so the findings say what broke the file.
    findings = [
        f for f in rel.unresolved_question_homes(both)
        if f["code"] == rel.RENDER_SHAPE_OVERFLOW
    ]
    assert sorted({f["reason"] for f in findings}) == sorted({
        mp.CELL_TEXT_TOO_LONG, mp.CELL_TEXT_ILLEGAL_CHARACTER,
    })


def test_the_truncated_cell_states_a_true_count_and_stays_under_the_cap():
    """Two arithmetic defects in one repair, both measured, both fixed.

    * the note said "holds the first 32585 of 40000 characters" while
      holding 32589 — ``kept`` was computed from the unformatted TEMPLATE
      (182 chars) and the cut used the FORMATTED mark (178). A false
      sentence inside a repair whose whole justification is that the record
      stays true.
    * formula-like content must remain exact and is protected by Cell string
      typing, not a payload-mutating apostrophe.
    """
    import re

    plain = mp._cell_value("z" * 40000, context="t")
    stated = int(re.search(r"first (\d+) of (\d+)", plain).group(1))
    assert stated == plain.index("\n[Aegis:"), "the count is the true count"
    assert len(plain) <= mp.CELL_LIMIT

    guarded = mp._cell_value("=" + "x" * 40000, context="t")
    assert guarded.startswith("=")
    assert len(guarded) <= mp.CELL_LIMIT
    stated = int(re.search(r"first (\d+) of (\d+)", guarded).group(1))
    assert stated == guarded.index("\n[Aegis:")


def test_formula_like_and_negative_answers_roundtrip_as_literal_text():
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["answers"][0].update({
        "answer_type": "Equation",
        "answer_content": "-3, -2, -1",
    })
    snapshot["candidates"][0]["answers"][1].update({
        "answer_type": "Phrases",
        "answer_content": "=HYPERLINK(\"https://invalid\",\"x\")",
    })

    master, _issues = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(master)
    row = next(
        item for item in parsed["sheets"]["Objective"]["rows"]
        if item.get("question_label")
    )
    assert row["answer_content_1"] == "-3, -2, -1"
    assert row["answer_content_2"] == (
        "=HYPERLINK(\"https://invalid\",\"x\")"
    )

    workbook = openpyxl.load_workbook(io.BytesIO(master), data_only=False)
    worksheet = workbook["Objective"]
    formula_cell = worksheet.cell(
        row=3, column=mp._INDEX["Objective"]["answer_content_2"] + 1,
    )
    assert formula_cell.data_type == "s"
    assert formula_cell.value == "=HYPERLINK(\"https://invalid\",\"x\")"
    workbook.close()


def test_capacity_overflow_is_a_named_defect_with_the_label_named():
    """INVERTED by spec-step8 T7.5/B4 — ``_question_record``'s four caps.

    The cap is the LAYOUT's own column count, so a row that overflows it
    cannot be written whole. Raising cost every row on every sheet of all
    four outputs; now the row ships with the slots the layout has, the
    overflow is named ``render_shape_overflow`` at staging with the label,
    the cap and the actual count, and the database write is refused.
    """
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["answers"] *= 4  # 8 options
    master, _ = mp.render_master_file(snapshot)
    assert master
    findings = [
        f for f in rel.unresolved_question_homes(snapshot)
        if f["code"] == rel.RENDER_SHAPE_OVERFLOW
    ]
    assert len(findings) == 1
    assert findings[0]["question_label"] == "06MSMA_T01_TwoDim Q01"
    assert findings[0]["cap"] == mp.MAX_OBJECTIVE_OPTIONS
    assert findings[0]["actual"] == 8
