"""The layout registry and the reader's fail-closed header gate (spec-step8 S2).

Every fixture here is built from a FROZEN registry entry or from a committed
``.xlsx``. None is built from ``writer._new_workbook``/``_write_headers``: that
writer migrates to the reference layout in S7, and a fixture that follows it
would prove nothing about the layouts this gate exists to keep readable.
"""
import io
import json

import openpyxl
import pytest

from app import bulk_import as bi
from app import models
from app.bulk_import import layouts, reader
from app.services import identity

_XLSX = ("application/vnd.openxmlformats-officedocument."
         "spreadsheetml.sheet")


def _reference_fixture(name: str):
    return layouts.REFERENCE_WORKBOOK_PATH.parent / name


def _new_workbook_for(layout: layouts.Layout, rows_by_kind: dict) -> openpyxl.Workbook:
    """A workbook carrying ``layout``'s exact headers plus the given rows.

    ``rows_by_kind`` maps a kind to a list of ``{(block, field): value}``
    dicts, so every fixture value is addressed by name through the registry.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for kind, sheet_layout in layout.sheets.items():
        ws = wb.create_sheet(sheet_layout.sheet_name)
        ws.append([band.label for band in sheet_layout.bands])
        ws.append(list(sheet_layout.fields))
        for cells in rows_by_kind.get(kind, []):
            # Synthetic valid-question fixtures should exercise the contract
            # under a real positive duration. Tests that intentionally probe a
            # blank duration provide the field explicitly and are not changed.
            cells = dict(cells)
            if (
                cells.get(("question", "question_label"))
                and ("question", "question_duration") not in cells
            ):
                cells[("question", "question_duration")] = "1"
            row = [""] * len(sheet_layout.fields)
            for (block, field), value in cells.items():
                index = sheet_layout.column(block, field)
                assert index is not None, (kind, block, field)
                row[index] = value
            ws.append(row)
    return wb


def _save(wb, tmp_path, name):
    path = tmp_path / name
    wb.save(path)
    return path


def _post(client, path):
    return client.post(
        "/data/import",
        files={"file": (path.name, io.BytesIO(path.read_bytes()), _XLSX)},
    )


def _math_policy_context(marker: str) -> dict:
    return {
        ("chapter", "chapter_title"): (
            f"Policy {marker} "
            "(06_Mathematics_MSBSHSE_Balbharati)"
        ),
        ("topic", "topic_title"): f"Topic 01: Policy {marker}",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): f"Policy {marker}",
        ("concept", "concept_display_name"): f"Policy {marker}",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
    }


def _math_policy_descriptive_cells(
    marker: str,
    *,
    category: str = "Short Answer Type (2 Marks)",
    marks: str = "2",
    difficulty: str = "Easy",
    duration: str = "2",
) -> dict:
    cells = _math_policy_context(marker)
    cells.update({
        ("question", "question_label"):
            f"06MSMA_Policy{marker}_PL_T01_C01 Q01",
        ("question", "question_category"): category,
        ("question", "question"): "Show the method.",
        ("question", "question_text"): "Show the method.",
        ("question", "marks"): marks,
        ("question", "level_of_difficulty"): difficulty,
        ("question", "question_duration"): duration,
        ("question", "math_keyboard"): "No",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "[method]: valid method",
        ("question", "answer_weightage_1"): marks,
    })
    return cells


def _math_policy_subjective_cells(
    marker: str,
    *,
    marks: str = "2",
    duration: str = "2",
    answer_count: int = 2,
) -> dict:
    cells = _math_policy_context(marker)
    tokens = [f"$${chr(ord('a') + n)}$$" for n in range(answer_count)]
    prompt = "Complete " + " and ".join(tokens) + "."
    cells.update({
        ("question", "question_label"):
            f"06MSMA_Policy{marker}_PL_T01_C01 Q01",
        ("question", "question_category"): "Fill in the blanks",
        ("question", "question"): prompt,
        ("question", "question_text"): prompt,
        ("question", "marks"): marks,
        ("question", "level_of_difficulty"): "Medium",
        ("question", "question_duration"): duration,
        ("question", "math_keyboard"): "No",
    })
    for n in range(1, answer_count + 1):
        placeholder = chr(ord("a") + n - 1)
        cells[("question", f"answer_type_{n}")] = "Phrases"
        cells[("question", f"answer_{n}")] = f"answer {n}"
        cells[("question", f"answer_display_{n}")] = f"answer {n}"
        cells[("question", f"weightage_{n}")] = "1"
        cells[("question", f"placeholder_{n}")] = placeholder
    return cells


@pytest.fixture(autouse=True)
def _drop_chapters_this_file_imports(db):
    """Leave the session database exactly as this file found it.

    ``conftest._prepare`` seeds the fixture workbook ONCE per session and
    never resets between tests, so a file that imports workbooks adds
    chapters every other test then sees. That is not hypothetical: without
    this fixture, ``test_data_io`` fails four ways when it runs after this
    file and passes when it runs alone — an order-dependent flake that the
    full suite happens to hide.

    Deleting the chapter is enough: ``Chapter.topics`` cascades
    ``all, delete-orphan`` through Topic, Concept, Group and Question
    (models.py:40, :59, :83, :111), so nothing this file imported survives.
    """
    before = {chapter_id for (chapter_id,) in db.query(models.Chapter.id)}
    yield
    db.expire_all()
    for chapter in db.query(models.Chapter).filter(
        ~models.Chapter.id.in_(before) if before else True
    ):
        db.delete(chapter)
    db.commit()


# --------------------------------------------------------------------------- #
# The registry itself
# --------------------------------------------------------------------------- #

def test_the_committed_format_workbook_and_the_template_json_agree_field_for_field():
    """T3.1b: the JSON is a derived cache of the owner-supplied workbook.

    A hand-edit to either copy fails here instead of shipping a mis-banded
    workbook.
    """
    workbook = openpyxl.load_workbook(
        layouts.REFERENCE_WORKBOOK_PATH, data_only=True, read_only=True)
    try:
        manifest = json.loads(
            layouts.MANIFEST_PATH.read_text(encoding="utf-8"))
        assert workbook.sheetnames == manifest["sheet_order"]
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            header = ["" if value is None else str(value).strip()
                      for value in rows[1]]
            assert header == manifest["sheets"][sheet_name]["fields"], sheet_name
            data_rows = [
                row for row in rows[2:]
                if row and any(
                    value is not None and str(value).strip() for value in row)
            ]
            assert data_rows == [], f"{sheet_name} carries data rows"
    finally:
        workbook.close()


def test_the_registry_prefers_the_workbook_and_records_the_drift(tmp_path):
    """T3.1b: on disagreement THE WORKBOOK WINS, and the drift is recorded."""
    manifest = json.loads(layouts.MANIFEST_PATH.read_text(encoding="utf-8"))
    manifest["sheets"]["Objective"]["fields"][15] = "keywords_RENAMED"

    layout, defects = layouts.build_reference_layout(manifest=manifest)

    assert layout is not None
    # The entry is the workbook's, not the JSON's.
    assert layout.sheet("objective").fields[15] == "keywords"
    drift = [d for d in defects if d["code"] == "layout_manifest_drift"]
    assert len(drift) == 1, defects
    assert drift[0]["sheet"] == "Objective"
    assert drift[0]["column_index"] == 15
    assert drift[0]["workbook_value"] == "keywords"
    assert drift[0]["manifest_value"] == "keywords_RENAMED"
    # A drift never halts anything: the live registry is still built.
    assert layouts.registry_defects() == []


def test_the_canonical_layout_entries_are_frozen_literals_not_live_constants():
    """T6/C5b: the canonical entries are transcriptions, not derivations.

    The tuples below are written out HERE. S7 redefines
    ``bi.OBJECTIVE_FIELDS`` / ``SUBJECTIVE_FIELDS`` / ``DESCRIPTIVE_FIELDS``
    to the reference layout; a registry entry derived from those constants
    would silently BECOME the reference layout, un-registering the canonical
    one and turning every older workbook into a 422 while this file's fixtures
    stayed green. After S7 the canonical entries must therefore NOT equal
    ``bi.FIELDS_BY_KIND`` — the two equalities at the bottom of this test are
    the ones S7 has to flip, deliberately and visibly.
    """
    front = (
        "chapter_title", "chapter_display_name", "chapter_duration",
        "pre_topics", "post_topics", "chapter_description",
        "topic_title", "topic_display_name", "pre_post_learning",
        "topic_concept_labels", "related_topics", "topic_description",
    )
    concept = (
        "concept_title", "concept_display_name", "parent_concept",
        "concept_details", "digicards",
        "basic_groups", "intermediate_groups", "advanced_groups",
        "concept_source",
    )
    legacy_concept = (
        "concept_title", "concept_display_name", "concept_details",
        "keywords", "digicards", "related_concepts",
        "basic_groups", "intermediate_groups", "advanced_groups",
    )
    objective_group = (
        "concept_question_labels", "group_name", "group_display_name",
        "group_description", "group_status", "group_type",
        "group_question_labels", "related_digicards",
    )
    descriptive_group = (
        "concept_question_labels", "group_display_name", "group_description",
        "group_name", "group_status", "group_type",
        "question_label", "group_question_labels", "related_digicards",
    )
    objective_scalars = (
        "question_label", "question_category", "cognitive_skills",
        "question_source", "question_disclaimer", "question_duration",
        "question_appears_in", "level_of_difficulty", "question", "marks",
    )
    objective_answers = tuple(
        f"{prefix}_{n}"
        for n in range(1, 7)
        for prefix in ("answer_type", "answer_content", "correct_answer",
                       "answer_weightage")
    )

    current = layouts.layout("canonical-current")
    objective = current.sheet("objective")
    assert objective.sheet_name == "Objective "  # the trailing space is real
    assert objective.fields == (
        front + concept + objective_group + objective_scalars
        + objective_answers + ("answer_explanation", "question_text")
    )
    assert len(objective.fields) == 65
    # The 10 / 11 / 12 scalar counts the reader used to hard-code, and the
    # answer-block counts, are properties of these literals.
    assert objective.question_scalar_fields == objective_scalars
    assert objective.answer_block_numbers == tuple(range(1, 7))
    assert len(current.sheet("subjective").question_scalar_fields) == 11
    assert current.sheet("subjective").answer_block_numbers == tuple(range(1, 11))
    assert len(current.sheet("descriptive").question_scalar_fields) == 12
    assert current.sheet("descriptive").answer_block_numbers == tuple(range(1, 11))
    assert current.sheet("descriptive").sub_question_numbers == tuple(range(1, 16))
    assert current.sheet("descriptive").block_fields("group") == descriptive_group

    legacy = layouts.layout("canonical-legacy-concept-band")
    assert legacy.sheet("objective").block_fields("concept") == legacy_concept
    assert "concept_source" not in legacy.sheet("objective").fields

    no_qt = layouts.layout("canonical-no-question-text")
    assert len(no_qt.sheet("objective").fields) == 64
    assert "question_text" not in no_qt.sheet("objective").fields

    # --- the two S7 flips, now FLIPPED ------------------------------------ #
    # S7 moved bi.FIELDS_BY_KIND / bi.SHEET_BY_KIND onto the owner's layout.
    # These two inequalities ARE the pin that the frozen literals above stayed
    # frozen: had the canonical entries been written as live references to
    # those constants, they would have moved too, the canonical layout would
    # have stopped being registered, and every older workbook would 422.
    assert list(objective.fields) != bi.FIELDS_BY_KIND["objective"]
    assert objective.sheet_name != bi.SHEET_BY_KIND["objective"]
    assert bi.SHEET_BY_KIND["objective"] == "Objective"
    # And the canonical layout is still registered and still reachable.
    assert layouts.LAYOUTS["canonical-current"] is current


def test_the_canonical_registry_entries_did_not_move_with_FIELDS_BY_KIND():
    """S7's direct pin, stated once on its own so it cannot be lost in a diff.

    ``layouts.CANONICAL_CURRENT`` is a frozen literal transcription. After S7
    it must NOT equal ``bi.FIELDS_BY_KIND``; if it ever does, the entry was
    re-derived and ``canonical-current`` has silently become the reference
    layout.
    """
    canonical = layouts.CANONICAL_CURRENT.fields_by_kind()
    for kind in ("objective", "subjective", "descriptive"):
        assert list(canonical[kind]) != list(bi.FIELDS_BY_KIND[kind]), kind
    assert layouts.CANONICAL_CURRENT.sheet_name_by_kind() != bi.SHEET_BY_KIND
    # All four entries are registered, so an older workbook still identifies.
    assert set(layouts.LAYOUTS) == {
        layouts.REFERENCE_LAYOUT_ID, "canonical-current",
        "canonical-no-question-text", "canonical-legacy-concept-band",
    }


def test_the_reference_entry_carries_the_measured_reference_geometry():
    """The counts the reader used to hard-code, measured on sop-mes-1."""
    reference = layouts.layout(layouts.REFERENCE_LAYOUT_ID)
    assert reference.sheet_name_by_kind() == {
        "objective": "Objective",
        "descriptive": "Descriptive",
        "subjective": "Subjective",
    }
    assert len(reference.sheet("objective").fields) == 67
    assert len(reference.sheet("descriptive").fields) == 374
    assert len(reference.sheet("subjective").fields) == 144
    # 12 / 13 / 14 scalars, against the reader's old 10 / 11 / 12.
    assert len(reference.sheet("objective").question_scalar_fields) == 12
    assert len(reference.sheet("subjective").question_scalar_fields) == 13
    assert len(reference.sheet("descriptive").question_scalar_fields) == 14
    # 20 answer blocks on Subjective, against the reader's old range(10).
    assert reference.sheet("subjective").answer_block_numbers == tuple(range(1, 21))
    # No sheet repeats a field name — which is what makes name addressing
    # possible on this layout at all.
    for kind, sheet_layout in reference.sheets.items():
        assert len(set(sheet_layout.fields)) == len(sheet_layout.fields), kind


def test_identity_module_imports_no_bulk_import_writer():
    """T4-9(b): the import direction stays bulk_import <- services.identity."""
    import subprocess
    import sys

    code = (
        "import sys; import app.services.identity; "
        "print('app.bulk_import.writer' in sys.modules)"
    )
    out = subprocess.run(
        [sys.executable, "-c", code], capture_output=True, text=True,
        cwd=str(layouts.REFERENCE_WORKBOOK_PATH.parents[3]),
    )
    assert out.returncode == 0, out.stderr
    assert out.stdout.strip() == "False", out.stdout


# --------------------------------------------------------------------------- #
# The reader, on the accepted reference corpus
# --------------------------------------------------------------------------- #

@pytest.mark.parametrize("fixture_name", [
    "grade6_science.xlsx", "grade6_english.xlsx", "grade6_mathematics.xlsx",
])
def test_reference_workbook_imports_by_name(db, fixture_name):
    """The exact two values that were wrong before the gate existed.

    Before: the Objective sheet was dropped whole with no issue recorded
    (``SHEET_BY_KIND['objective']`` is ``'Objective '`` and the reference
    sheet is ``'Objective'``), and the Descriptive rows stored
    ``question_label = 'Very Short Answer Questions'`` — the question_category
    value, read two columns early.
    """
    counts = reader.import_workbook(db, _reference_fixture(fixture_name))
    assert counts["layout_id"] == layouts.REFERENCE_LAYOUT_ID
    # The owner's reference books predate the minted identity: their title
    # cells carry ``directory.topic_tag``/``concept_tag``, which are
    # CHAPTER-level — one tag on every topic of the chapter. Restoring those
    # as identities gave [measured] six duplicate topic machine_ids in these
    # eight topics, which is one ``question_label`` for two concepts and a
    # silently dropped question (R4). They are now recorded, one note per row,
    # and the id is minted on first use. The Science reference also retains
    # one legacy ``\mathrm`` expression in both learner text and its Equation
    # rubric. Current imports preserve that frozen content but record both
    # medium defects for review. Every OTHER issue must still be absent, so
    # the filters key on stable content, not an incidental ordering.
    identity_notes = [
        issue for issue in counts["issues"]
        if "(imported_without_machine_id)" in issue
    ]
    assert identity_notes, "the legacy chapter-level tags must be RECORDED"
    legacy_medium_notes = [
        issue for issue in counts["issues"]
        if issue.startswith(
            "06MSSC_Measur_PL_T04_LaboratoryThermometerL Q01:"
        )
        and "\\mathrm" in issue
    ]
    assert len(legacy_medium_notes) == (
        2 if fixture_name == "grade6_science.xlsx" else 0
    ), legacy_medium_notes
    legacy_multipart_notes = [
        issue for issue in counts["issues"]
        if "multipart descriptive duplicates scoring" in issue
    ]
    assert len(legacy_multipart_notes) == 1, legacy_multipart_notes
    legacy_rubric_notes = [
        issue for issue in counts["issues"]
        if "does not start with an allowed functional tag" in issue
    ]
    assert [
        issue for issue in counts["issues"]
        if issue not in identity_notes
        and issue not in legacy_medium_notes
        and issue not in legacy_multipart_notes
        and issue not in legacy_rubric_notes
    ] == []

    descriptive = [
        q for q in db.query(models.Question).all()
        if q.sheet_kind == "descriptive"
    ]
    objective = [
        q for q in db.query(models.Question).all()
        if q.sheet_kind == "objective"
    ]
    # The Objective sheet is read at all.
    assert objective, "the Objective sheet was dropped"

    workbook = openpyxl.load_workbook(
        _reference_fixture(fixture_name), data_only=True, read_only=True)
    try:
        sheet_layout = layouts.sheet(layouts.REFERENCE_LAYOUT_ID, "descriptive")
        rows = [
            row for row in workbook["Descriptive"].iter_rows(
                min_row=3, values_only=True)
            if row and any(row)
        ]
    finally:
        workbook.close()
    truth = sheet_layout.block_values(rows[0], "question")
    stored = {q.question_label: q for q in descriptive}
    assert truth["question_label"] in stored, sorted(stored)
    question = stored[truth["question_label"]]
    assert question.question_category == truth["question_category"]
    assert question.question == truth["question"]
    assert question.marks == float(truth["marks"])


def test_the_first_reference_descriptive_row_stores_the_measured_values(db):
    """The two literal values from the reproduction, pinned by hand."""
    reader.import_workbook(db, _reference_fixture("grade6_science.xlsx"))
    question = db.query(models.Question).filter_by(
        question_label="06MSSC_Measur_PL_T02_LimitationsOfHandSpanM Q01").one()
    assert question.question_category == "Very Short Answer Questions"
    assert question.sheet_kind == "descriptive"


def test_reader_answer_blocks_are_read_by_name(db, tmp_path):
    """The block COUNT comes from the layout, not from a literal ``range``.

    The fixture is synthetic: all three committed reference Subjective sheets
    carry zero data rows, and the reference sheet declares 20 answer blocks
    against the reader's old ``range(10)``.
    """
    reference = layouts.layout(layouts.REFERENCE_LAYOUT_ID)
    cells = {
        ("chapter", "chapter_title"): "Sound Waves (07_Science_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Vibrations",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Amplitude and loudness",
        ("concept", "concept_display_name"): "Amplitude and loudness",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): "07CBSC_Sound_PL_T01_C01 Q01",
        ("question", "question"): "List fifteen properties of a sound wave.",
        ("question", "marks"): "15",
    }
    for n in range(1, 16):
        cells[("question", f"answer_type_{n}")] = "Phrases"
        cells[("question", f"answer_{n}")] = f"property {n}"
        cells[("question", f"weightage_{n}")] = "1"
    path = _save(
        _new_workbook_for(reference, {"subjective": [cells]}),
        tmp_path, "reference_subjective.xlsx")

    counts = reader.import_workbook(db, path)
    # The synthetic fixture's title cells carry no machine-id tag, so the
    # importer NOTES each row as ``imported_without_machine_id`` (S4/T4-8: a
    # note, never a block — the id is minted on first use). Nothing about the
    # answer bands may be reported. The filter keys on the stable CODE token,
    # not on the prose: filtering out every issue whose text happens to
    # contain "machine id" would also hide ``machine_id_conflict``.
    assert [
        issue for issue in counts["issues"]
        if "(imported_without_machine_id)" not in issue
    ] == []
    question = db.query(models.Question).filter_by(
        question_label="07CBSC_Sound_PL_T01_C01 Q01").one()
    assert len(question.answers) == 15
    assert question.answers[14]["answer"] == "property 15"


def test_canonical_current_and_legacy_workbooks_still_import(db, tmp_path):
    """Q16: older workbooks remain readable, and are read as themselves."""
    for layout_id, marker in (
        ("canonical-current", "CanonCurrent"),
        ("canonical-legacy-concept-band", "CanonLegacy"),
        ("canonical-no-question-text", "CanonNoQT"),
    ):
        layout = layouts.layout(layout_id)
        label = f"08CBSC_{marker}_PL_T01_C01 Q01"
        cells = {
            ("chapter", "chapter_title"): f"{marker} Chapter (08_Science_CBSE_NCERT)",
            ("topic", "topic_title"): "Topic 01: Registry Topic",
            ("topic", "pre_post_learning"): "Post",
            ("concept", "concept_title"): f"{marker} Concept",
            ("concept", "concept_display_name"): f"{marker} Concept",
            ("group", "group_name"): "Basic Group 01",
            ("group", "group_type"): "Basic",
            ("question", "question_label"): label,
            ("question", "question"): f"What does {marker} pin?",
            ("question", "marks"): "1",
            ("question", "answer_type_1"): "Phrases",
            ("question", "answer_content_1"): "the layout registry",
        }
        path = _save(
            _new_workbook_for(layout, {"objective": [cells]}),
            tmp_path, f"{layout_id}.xlsx")
        counts = reader.import_workbook(db, path)
        assert counts["layout_id"] == layout_id
        question = db.query(models.Question).filter_by(
            question_label=label).one()
        assert question.question == f"What does {marker} pin?"
        assert question.group.concept.concept_title == f"{marker} Concept"


# --------------------------------------------------------------------------- #
# The gate refuses, whole
# --------------------------------------------------------------------------- #

def test_uppercase_objective_option_labels_refuse_the_whole_import(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_Uppercase_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Uppercase Option Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Uppercase Options",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Uppercase options",
        ("concept", "concept_display_name"): "Uppercase options",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Choose the first letter.",
        ("question", "question_text"):
            "Choose the first letter.\nA) Alpha\nB) Beta",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "Alpha",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_content_2"): "Beta",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path, "uppercase_objective_options.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "uppercase objective option label(s) A), B)" in detail
    assert "uppercase_objective_option_label" in detail
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_uppercase_options_in_question_fallback_refuse_the_whole_import(
    client, db, tmp_path,
):
    """Blank question_text cannot bypass the label gate via backfill."""

    layout = layouts.layout("canonical-current")
    label = "09CBPH_UpperFallback_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Uppercase Fallback Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Uppercase Fallback",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Uppercase fallback",
        ("concept", "concept_display_name"): "Uppercase fallback",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Choose.\nA) Alpha\nB) Beta",
        ("question", "question_text"): "",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "Alpha",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_content_2"): "Beta",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path,
        "uppercase_question_fallback.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert "uppercase objective option label(s) A), B)" in (
        response.json()["detail"]
    )
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


@pytest.mark.parametrize("populated_answers", [0, 1])
def test_uppercase_option_scan_uses_layout_capacity_not_populated_blocks(
    client, db, tmp_path, populated_answers,
):
    layout = layouts.layout("canonical-current")
    label = f"09CBPH_UpperCapacity{populated_answers}_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Uppercase Capacity Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Uppercase Capacity",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Uppercase capacity",
        ("concept", "concept_display_name"): "Uppercase capacity",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Choose.",
        ("question", "question_text"): "Choose.\nA) Alpha\nB) Beta",
        ("question", "marks"): "1",
    }
    if populated_answers:
        cells[("question", "answer_type_1")] = "Phrases"
        cells[("question", "answer_content_1")] = "Alpha"
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path,
        f"uppercase_capacity_{populated_answers}.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert "uppercase objective option label(s) A), B)" in (
        response.json()["detail"]
    )
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


@pytest.mark.parametrize(
    ("dialect", "table_text"),
    [
        pytest.param(
            "markdown",
            "Study the table.\n| Name | Value |\n|---|---:|\n| A | 1 |",
            id="markdown-table",
        ),
        pytest.param(
            "katex-tabular",
            (
                r"Study [Katex] \begin{tabular}{cc}"
                r"A&B\\1&2\end{tabular} [/Katex]."
            ),
            id="katex-tabular-table",
        ),
    ],
)
def test_unsupported_rich_question_tables_refuse_the_whole_import(
    client, db, tmp_path, dialect, table_text,
):
    layout = layouts.layout("canonical-current")
    marker = dialect.replace("-", "_")
    label = f"09CBPH_Table_{marker}_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            f"Unsupported Table {marker} (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): f"Topic 01: Table {marker}",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): f"Table {marker}",
        ("concept", "concept_display_name"): f"Table {marker}",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Read the supplied table.",
        ("question", "question_text"): table_text,
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "One",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path, f"unsupported_{marker}.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "question_text" in detail
    assert "unsupported" in detail.casefold()
    assert "table" in detail.casefold()
    assert "unsupported_table" in detail
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_subjective_answer_table_is_refused_before_persistence(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_SubjectiveTable_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Subjective Table Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Subjective Table",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Subjective table",
        ("concept", "concept_display_name"): "Subjective table",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "State the values.",
        ("question", "question_text"): "State the values.",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_1"):
            "| Name | Value |\n|---|---:|\n| Alpha | 1 |",
        ("question", "weightage_1"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"subjective": [cells]}),
        tmp_path,
        "subjective_answer_table.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert "unsupported_table" in response.json()["detail"]
    db.expire_all()
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_four_mark_single_rubric_refuses_before_database_mutation(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_FourMarkRubric_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Four Mark Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Four Mark",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Four mark rubric",
        ("concept", "concept_display_name"): "Four mark rubric",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Explain the method.",
        ("question", "question_text"): "Explain the method.",
        ("question", "marks"): "4",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "[content]: one combined rubric.",
        ("question", "answer_weightage_1"): "4",
        ("question", "math_keyboard"): "No",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "four_mark_single_rubric.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert "4-mark single-part descriptive requires at least two" in (
        response.json()["detail"]
    )
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_four_mark_multipart_can_score_only_subquestion_rubrics(
    client, db, tmp_path,
):
    """The two-rubric rule belongs to a single-part 4-mark question.

    A multipart item carries no main answer blocks: its subquestion marks and
    keyword rubrics own the score.
    """

    layout = layouts.layout("canonical-current")
    label = "09CBPH_FourMarkMultipart_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Four Mark Multipart Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Four Mark Multipart",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Four mark multipart",
        ("concept", "concept_display_name"): "Four mark multipart",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Answer both parts.",
        ("question", "question_text"): "Answer both parts.",
        ("question", "marks"): "4",
        ("question", "math_keyboard"): "No",
        ("question", "sub_question_1"): "State the first result.",
        ("question", "sub_question_marks_1"): "2",
        ("question", "sq1_answer_type_1"): "Phrases",
        ("question", "sq1_weightage_1"): "2",
        ("question", "sq1_keyword_1"): "[content]: first result",
        ("question", "sub_question_2"): "Explain the second result.",
        ("question", "sub_question_marks_2"): "2",
        ("question", "sq2_answer_type_1"): "Phrases",
        ("question", "sq2_weightage_1"): "2",
        ("question", "sq2_keyword_1"): "[method]: second result",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "four_mark_multipart.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 200, response.text
    assert not any(
        "4-mark descriptive requires" in issue
        for issue in response.json()["issues"]
    )
    db.expire_all()
    question = db.query(models.Question).filter_by(
        question_label=label,
    ).one()
    assert question.answers == []
    assert [part["marks"] for part in question.sub_questions] == ["2", "2"]
    assert [
        part["keywords"][0]["weightage"]
        for part in question.sub_questions
    ] == ["2", "2"]


def test_multipart_import_does_not_sum_main_weights_when_subquestions_score(
    db, tmp_path,
):
    """Legacy main rubric residue cannot replace multipart score ownership."""

    layout = layouts.layout("canonical-current")
    label = "09CBPH_MultipartScoreOwner_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Multipart Score Owner Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Multipart Score Owner",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Multipart score owner",
        ("concept", "concept_display_name"): "Multipart score owner",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Answer both parts.",
        ("question", "question_text"): "Answer both parts.",
        ("question", "marks"): "4",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "Legacy shared rubric.",
        ("question", "answer_weightage_1"): "1",
        ("question", "sub_question_1"): "Give the result.",
        ("question", "sub_question_marks_1"): "4",
        ("question", "sq1_answer_type_1"): "Phrases",
        ("question", "sq1_weightage_1"): "4",
        ("question", "sq1_keyword_1"): "[content]: result",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "multipart_score_owner.xlsx",
    )

    counts = reader.import_workbook(db, path)

    assert not any(
        "answer weightage sum" in issue
        or "4-mark descriptive requires" in issue
        for issue in counts["issues"]
    ), counts["issues"]
    assert any(
        "multipart descriptive duplicates scoring" in issue
        for issue in counts["issues"]
    )


def test_literal_newline_escape_refuses_before_database_mutation(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_LiteralNewline_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Literal Newline Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Formatting",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Option formatting",
        ("concept", "concept_display_name"): "Option formatting",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): r"Choose one.\na) Alpha\nb) Beta",
        ("question", "question_text"): r"Choose one.\na) Alpha\nb) Beta",
        ("question", "question_duration"): "1",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "Alpha",
        ("question", "correct_answer_1"): "1",
        ("question", "answer_weightage_1"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path,
        "literal_newline.xlsx",
    )

    before = db.query(models.Question).count()
    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert "literal \\n found" in response.json()["detail"]
    assert db.query(models.Question).count() == before


def test_subjective_placeholders_are_not_treated_as_raw_math(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_SubjectivePlaceholders_PL_T01_C01 Q01"
    prompt = "Complete $$a$$, then complete $$b$$."
    cells = {
        ("chapter", "chapter_title"):
            "Subjective Placeholder Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Subjective Placeholders",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Subjective placeholders",
        ("concept", "concept_display_name"): "Subjective placeholders",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): prompt,
        ("question", "question_text"): prompt,
        ("question", "marks"): "2",
        ("question", "math_keyboard"): "No",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_1"): "first",
        ("question", "answer_display_1"): "first",
        ("question", "weightage_1"): "1",
        ("question", "placeholder_1"): "a",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_2"): "second",
        ("question", "answer_display_2"): "second",
        ("question", "weightage_2"): "1",
        ("question", "placeholder_2"): "b",
    }
    path = _save(
        _new_workbook_for(layout, {"subjective": [cells]}),
        tmp_path,
        "subjective_placeholders.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 200, response.text
    assert not any(
        "raw math" in issue or "unbalanced [Katex]" in issue
        for issue in response.json()["issues"]
    ), response.json()["issues"]
    db.expire_all()
    question = db.query(models.Question).filter_by(
        question_label=label,
    ).one()
    assert question.question == prompt
    assert [answer["placeholder"] for answer in question.answers] == ["a", "b"]


def test_strict_import_accepts_all_msbshse_math_policy_sheets(
    client, db, tmp_path,
):
    """The chapter tag selects the narrow policy without local guessing."""

    layout = layouts.layout("canonical-current")
    objective = _math_policy_context("PositiveObjective")
    objective.update({
        ("question", "question_label"):
            "06MSMA_PolicyPositiveObjective_PL_T01_C01 Q01",
        ("question", "question_category"): "Multiple Choice Question",
        ("question", "question"): "Which value is even?",
        ("question", "question_text"): "Which value is even?",
        ("question", "marks"): "1",
        ("question", "level_of_difficulty"): "Hard",
        ("question", "question_duration"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "2",
        ("question", "correct_answer_1"): "Yes",
        ("question", "answer_weightage_1"): "1",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_content_2"): "3",
        ("question", "correct_answer_2"): "No",
        ("question", "answer_weightage_2"): "0",
    })
    descriptive = _math_policy_descriptive_cells("PositiveDescriptive")
    subjective = _math_policy_subjective_cells("PositiveSubjective")
    path = _save(
        _new_workbook_for(
            layout,
            {
                "objective": [objective],
                "descriptive": [descriptive],
                "subjective": [subjective],
            },
        ),
        tmp_path,
        "math_policy_positive.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 200, response.text
    assert response.json()["questions"] == 3
    db.expire_all()
    stored = {
        question.question_label: question
        for question in db.query(models.Question).filter(
            models.Question.question_label.in_({
                "06MSMA_PolicyPositiveObjective_PL_T01_C01 Q01",
                "06MSMA_PolicyPositiveDescriptive_PL_T01_C01 Q01",
                "06MSMA_PolicyPositiveSubjective_PL_T01_C01 Q01",
            })
        )
    }
    assert stored[
        "06MSMA_PolicyPositiveObjective_PL_T01_C01 Q01"
    ].level_of_difficulty == "High"
    assert stored[
        "06MSMA_PolicyPositiveDescriptive_PL_T01_C01 Q01"
    ].level_of_difficulty == "Less"
    assert stored[
        "06MSMA_PolicyPositiveSubjective_PL_T01_C01 Q01"
    ].level_of_difficulty == "Moderate"


@pytest.mark.parametrize(
    ("case", "kind", "overrides", "answer_count", "expected"),
    [
        pytest.param(
            "WrongSheet",
            "descriptive",
            {"question_category": "Multiple Choice Question"},
            0,
            "is not permitted for sheet_kind 'descriptive'",
            id="wrong-sheet-category",
        ),
        pytest.param(
            "WrongFixedMarks",
            "descriptive",
            {"marks": "3"},
            0,
            "marks 3 must be one of",
            id="wrong-fixed-marks",
        ),
        pytest.param(
            "WrongMatrixDuration",
            "descriptive",
            {"difficulty": "Hard", "duration": "2"},
            0,
            "question_duration 2 does not match policy duration 3",
            id="wrong-matrix-duration",
        ),
        pytest.param(
            "FractionalSubpoints",
            "subjective",
            {"marks": "2.5", "duration": "2"},
            2,
            "positive whole number of policy subpoints",
            id="wrong-per-subpoint-marks",
        ),
        pytest.param(
            "WrongSubpointDuration",
            "subjective",
            {"marks": "2", "duration": "1"},
            2,
            "question_duration 1 does not match policy duration 2",
            id="wrong-per-subpoint-duration",
        ),
        pytest.param(
            "WrongAnswerCount",
            "subjective",
            {"marks": "2", "duration": "2"},
            1,
            "Subjective answer count 1 does not match represented subpoint "
            "count 2",
            id="subjective-answer-count",
        ),
    ],
)
def test_strict_import_rejects_msbshse_math_policy_mismatches_before_writes(
    client, db, tmp_path, case, kind, overrides, answer_count, expected,
):
    layout = layouts.layout("canonical-current")
    if kind == "descriptive":
        cells = _math_policy_descriptive_cells(
            case,
            category=overrides.get(
                "question_category", "Short Answer Type (2 Marks)"
            ),
            marks=overrides.get("marks", "2"),
            difficulty=overrides.get("difficulty", "Easy"),
            duration=overrides.get("duration", "2"),
        )
    else:
        cells = _math_policy_subjective_cells(
            case,
            marks=overrides.get("marks", "2"),
            duration=overrides.get("duration", "2"),
            answer_count=answer_count,
        )
    label = cells[("question", "question_label")]
    path = _save(
        _new_workbook_for(layout, {kind: [cells]}),
        tmp_path,
        f"math_policy_{case}.xlsx",
    )
    chapter_count = db.query(models.Chapter).count()
    question_count = db.query(models.Question).count()

    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert expected in response.json()["detail"]
    db.expire_all()
    assert db.query(models.Chapter).count() == chapter_count
    assert db.query(models.Question).count() == question_count
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_strict_math_policy_skips_an_empty_question_band(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    cells = _math_policy_context("ConceptTail")
    for field in ("group_name", "group_type"):
        cells.pop(("group", field))
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "math_policy_concept_tail.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 200, response.text
    assert response.json()["questions"] == 0
    assert response.json()["groups"] == 0
    assert db.query(models.Concept).filter_by(
        concept_title="Policy ConceptTail"
    ).one()


def test_mixed_phrases_rubric_refuses_before_database_mutation(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_MixedPhrases_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Mixed Phrases Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Mixed Phrases",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Mixed phrases",
        ("concept", "concept_display_name"): "Mixed phrases",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Give the value.",
        ("question", "question_text"): "Give the value.",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): r"The value is \frac{1}{2}.",
        ("question", "answer_weightage_1"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "mixed_phrases_rubric.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert "Phrases content must not include LaTeX" in (
        response.json()["detail"]
    )
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


@pytest.mark.parametrize(
    ("answer_type", "content", "expected"),
    [
        pytest.param(
            "Words",
            r"The value is \frac{1}{2}.",
            "Phrases content must not include LaTeX",
            id="legacy-words-normalizes-before-medium-check",
        ),
        pytest.param(
            "",
            "Plain answer without a declared medium.",
            "unsupported or blank answer_type ''",
            id="blank-type",
        ),
        pytest.param(
            "Mystery",
            "Plain answer with an unknown medium.",
            "unsupported or blank answer_type 'Mystery'",
            id="unknown-type",
        ),
    ],
)
def test_answer_type_is_normalized_and_validated_before_database_mutation(
    client, db, tmp_path, answer_type, content, expected,
):
    layout = layouts.layout("canonical-current")
    suffix = answer_type or "Blank"
    label = f"09CBPH_Type{suffix}_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Answer Type Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Answer Type",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Answer type",
        ("concept", "concept_display_name"): "Answer type",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Give the answer.",
        ("question", "question_text"): "Give the answer.",
        ("question", "marks"): "1",
        ("question", "answer_type_1"): answer_type,
        ("question", "answer_content_1"): content,
        ("question", "answer_weightage_1"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        f"answer_type_{suffix}.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert expected in response.json()["detail"]
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_legacy_words_alias_strict_imports_as_canonical_phrases(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_TypeWordsPositive_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Words Alias Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Words Alias",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Words alias",
        ("concept", "concept_display_name"): "Words alias",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Give the answer.",
        ("question", "question_text"): "Give the answer.",
        ("question", "marks"): "1",
        ("question", "question_duration"): "1",
        ("question", "math_keyboard"): "No",
        ("question", "answer_type_1"): "Words",
        ("question", "answer_content_1"): "[content]: complete answer",
        ("question", "answer_weightage_1"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "legacy_words_positive.xlsx",
    )

    response = _post(client, path)

    assert response.status_code == 200, response.text
    db.expire_all()
    question = db.query(models.Question).filter_by(
        question_label=label,
    ).one()
    assert question.answers[0]["answer_type"] == "Phrases"


@pytest.mark.parametrize(
    ("field", "value", "expected"),
    [
        pytest.param(
            "question_label", "",
            "populated Question band requires a non-blank question_label",
            id="blank-question-label",
        ),
        pytest.param(
            "question", "", "question must not be blank",
            id="blank-question",
        ),
        pytest.param(
            "marks", "+1", "marks must be finite and positive",
            id="signed-marks-text",
        ),
        pytest.param(
            "question_duration", "+1",
            "question_duration must be finite and positive",
            id="signed-duration-text",
        ),
    ],
)
def test_strict_import_refuses_missing_identity_and_signed_numeric_text(
    client, db, tmp_path, field, value, expected,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_StrictIdentity_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Strict Identity Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Strict Identity",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Strict identity",
        ("concept", "concept_display_name"): "Strict identity",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question_category"): "Multiple Choice Question",
        ("question", "question"): "Which option is correct?",
        ("question", "question_text"): "Which option is correct?",
        ("question", "marks"): "1",
        ("question", "question_duration"): "1",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "First",
        ("question", "correct_answer_1"): "Yes",
        ("question", "answer_weightage_1"): "1",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_content_2"): "Second",
        ("question", "correct_answer_2"): "No",
        ("question", "answer_weightage_2"): "0",
    }
    cells[("question", field)] = value
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path,
        f"strict_{field}.xlsx",
    )
    before = (
        db.query(models.Chapter).count(),
        db.query(models.Question).count(),
    )

    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert expected in response.json()["detail"]
    db.expire_all()
    assert (
        db.query(models.Chapter).count(),
        db.query(models.Question).count(),
    ) == before


@pytest.mark.parametrize("multipart", [False, True])
def test_strict_import_refuses_a_tag_without_rubric_content(
    client, db, tmp_path, multipart,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_EmptyRubric_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Empty Rubric Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Empty Rubric",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Empty rubric",
        ("concept", "concept_display_name"): "Empty rubric",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "State the result.",
        ("question", "question_text"): "State the result.",
        ("question", "marks"): "1",
        ("question", "question_duration"): "1",
        ("question", "math_keyboard"): "No",
    }
    if multipart:
        cells.update({
            ("question", "sub_question_1"): "Give the result.",
            ("question", "sub_question_marks_1"): "1",
            ("question", "sq1_answer_type_1"): "Phrases",
            ("question", "sq1_weightage_1"): "1",
            ("question", "sq1_keyword_1"): "[content]:   ",
        })
    else:
        cells.update({
            ("question", "answer_type_1"): "Phrases",
            ("question", "answer_content_1"): "[content]:   ",
            ("question", "answer_weightage_1"): "1",
        })
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        f"empty_rubric_{'multipart' if multipart else 'main'}.xlsx",
    )
    before = db.query(models.Question).count()

    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert "allowed functional tag" in response.json()["detail"]
    db.expire_all()
    assert db.query(models.Question).count() == before


@pytest.mark.parametrize(
    ("kind", "positional_cells", "expected"),
    [
        pytest.param(
            "objective",
            {
                ("question", "answer_type_2"): "Phrases",
                ("question", "answer_content_2"): "Only slot two",
                ("question", "correct_answer_2"): "Yes",
                ("question", "answer_weightage_2"): "1",
            },
            "objective option blocks must be contiguous from slot 1",
            id="objective-options",
        ),
        pytest.param(
            "subjective",
            {
                ("question", "question"): "Complete $$b$$.",
                ("question", "question_text"): "Complete $$b$$.",
                ("question", "answer_type_2"): "Phrases",
                ("question", "answer_2"): "second",
                ("question", "answer_display_2"): "second",
                ("question", "weightage_2"): "1",
                ("question", "placeholder_2"): "b",
            },
            "subjective answer blocks must be contiguous from slot 1",
            id="subjective-answers",
        ),
        pytest.param(
            "descriptive",
            {
                ("question", "answer_type_2"): "Phrases",
                ("question", "answer_content_2"): "[content]: slot two",
                ("question", "answer_weightage_2"): "1",
            },
            "Descriptive main rubric blocks must be contiguous from slot 1",
            id="descriptive-main-rubrics",
        ),
        pytest.param(
            "descriptive",
            {
                ("question", "sub_question_2"): "State the second result.",
                ("question", "sub_question_marks_2"): "1",
                ("question", "sq2_answer_type_1"): "Phrases",
                ("question", "sq2_weightage_1"): "1",
                ("question", "sq2_keyword_1"): "[content]: second result",
            },
            "Descriptive subquestion blocks must be contiguous from slot 1",
            id="descriptive-subquestions",
        ),
        pytest.param(
            "descriptive",
            {
                ("question", "sub_question_1"): "State the result.",
                ("question", "sub_question_marks_1"): "1",
                ("question", "sq1_answer_type_2"): "Phrases",
                ("question", "sq1_weightage_2"): "1",
                ("question", "sq1_keyword_2"): "[content]: result",
            },
            "subquestion 1 keyword blocks must be contiguous from slot 1",
            id="descriptive-keywords",
        ),
    ],
)
def test_sparse_positional_blocks_refuse_before_database_mutation(
    client, db, tmp_path, kind, positional_cells, expected,
):
    layout = layouts.layout("canonical-current")
    marker = expected.split()[0].replace("/", "")
    label = f"09CBPH_Sparse{marker}_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            f"Sparse {marker} Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): f"Topic 01: Sparse {marker}",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): f"Sparse {marker}",
        ("concept", "concept_display_name"): f"Sparse {marker}",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Give the result.",
        ("question", "question_text"): "Give the result.",
        ("question", "marks"): "1",
        ("question", "question_duration"): "1",
        **positional_cells,
    }
    if kind != "objective":
        cells[("question", "math_keyboard")] = "No"
    path = _save(
        _new_workbook_for(layout, {kind: [cells]}),
        tmp_path,
        f"sparse_{kind}_{marker}.xlsx",
    )
    before = db.query(models.Question).count()

    response = _post(client, path)

    assert response.status_code == 422, response.text
    assert expected in response.json()["detail"]
    db.expire_all()
    assert db.query(models.Question).count() == before


def test_type_only_second_block_does_not_satisfy_four_mark_rubric_shape(
    client, db, tmp_path,
):
    layout = layouts.layout("canonical-current")
    label = "09CBPH_BlankSecondRubric_PL_T01_C01 Q01"
    cells = {
        ("chapter", "chapter_title"):
            "Blank Rubric Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Blank Rubric",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Blank second rubric",
        ("concept", "concept_display_name"): "Blank second rubric",
        ("group", "group_name"): "Basic Group 01",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): label,
        ("question", "question"): "Explain the method.",
        ("question", "question_text"): "Explain the method.",
        ("question", "marks"): "4",
        ("question", "answer_type_1"): "Phrases",
        ("question", "answer_content_1"): "[content]: first rubric point.",
        ("question", "answer_weightage_1"): "2",
        ("question", "answer_type_2"): "Phrases",
        ("question", "answer_content_2"): "",
        ("question", "answer_weightage_2"): "2",
        ("question", "math_keyboard"): "No",
    }
    path = _save(
        _new_workbook_for(layout, {"descriptive": [cells]}),
        tmp_path,
        "blank_second_rubric.xlsx",
    )

    before = db.query(models.Chapter).count()
    response = _post(client, path)

    assert response.status_code == 422
    assert "4-mark single-part descriptive requires at least two" in (
        response.json()["detail"]
    )
    db.expire_all()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label=label,
    ).first() is None


def test_an_unrecognised_header_refuses_the_whole_import(client, db, tmp_path):
    layout = layouts.layout("canonical-current")
    cells = {
        ("chapter", "chapter_title"): "Unreadable Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Unreadable Topic",
        ("concept", "concept_title"): "Unreadable Concept",
        ("concept", "concept_display_name"): "Unreadable Concept",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): "09CBPH_Unread_PL_T01_C01 Q01",
        ("question", "question"): "Does a half-read workbook import?",
        ("question", "marks"): "1",
    }
    wb = _new_workbook_for(layout, {"objective": [cells]})
    # One renamed column is enough: the geometry can no longer be established.
    wb[layout.sheet("objective").sheet_name].cell(row=2, column=16).value = "kewyords"
    path = _save(wb, tmp_path, "unreadable.xlsx")

    before = db.query(models.Chapter).count()
    response = _post(client, path)
    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "Objective " in detail
    assert "column 16" in detail
    assert "kewyords" in detail
    db.rollback()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Question).filter_by(
        question_label="09CBPH_Unread_PL_T01_C01 Q01").first() is None


def test_a_trailing_space_sheet_name_is_a_different_layout_not_a_skip(
    client, db, tmp_path,
):
    """The whole bug, as one assertion.

    ``SHEET_BY_KIND['objective']`` carries a trailing space. A sheet named
    ``'Objective'`` used to hit ``if sheet_name not in wb.sheetnames:
    continue`` and vanish with no issue recorded. It is now an unidentified
    layout, and an unidentified layout refuses the file.
    """
    layout = layouts.layout("canonical-current")
    cells = {
        ("chapter", "chapter_title"): "Trailing Space Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Trailing Space Topic",
        ("concept", "concept_title"): "Trailing Space Concept",
        ("concept", "concept_display_name"): "Trailing Space Concept",
        ("group", "group_type"): "Basic",
        ("question", "question_label"): "09CBPH_Trail_PL_T01_C01 Q01",
        ("question", "question"): "Is a renamed sheet a skip?",
        ("question", "marks"): "1",
    }
    wb = _new_workbook_for(layout, {"objective": [cells]})
    wb[layout.sheet("objective").sheet_name].title = "Objective"
    path = _save(wb, tmp_path, "trailing_space.xlsx")

    before = db.query(models.Chapter).count()
    response = _post(client, path)
    assert response.status_code == 422
    assert "Objective" in response.json()["detail"]
    db.rollback()
    assert db.query(models.Chapter).count() == before
    assert db.query(models.Concept).filter_by(
        concept_title="Trailing Space Concept").first() is None


def test_identify_sheet_matches_only_on_exact_name_and_header():
    objective = layouts.sheet("canonical-current", "objective")
    assert layouts.identify_sheet(
        objective.sheet_name, objective.fields).layout_id == "canonical-current"
    # Same header, different sheet name -> not a match.
    assert layouts.identify_sheet("Objective", objective.fields) is None
    # Same sheet name, one RENAMED column -> not a match. (Dropping the
    # trailing question_text column is a registered layout of its own, which
    # is exactly why identification is tuple equality and not a prefix test.)
    renamed = list(objective.fields)
    renamed[15] = "kewyords"
    assert layouts.identify_sheet(objective.sheet_name, renamed) is None
    reference = layouts.sheet(layouts.REFERENCE_LAYOUT_ID, "descriptive")
    canonical_descriptive = layouts.sheet("canonical-current", "descriptive")
    # Both are 374 columns on a sheet named 'Descriptive'; only the exact
    # tuple tells them apart.
    assert len(reference.fields) == len(canonical_descriptive.fields)
    assert layouts.identify_sheet(
        "Descriptive", reference.fields
    ).layout_id == layouts.REFERENCE_LAYOUT_ID
    assert layouts.identify_sheet(
        "Descriptive", canonical_descriptive.fields
    ).layout_id == "canonical-current"


# --------------------------------------------------------------------------- #
# The two silent merges in the same function
# --------------------------------------------------------------------------- #

def test_two_chapters_whose_codes_collide_are_not_merged(db, tmp_path):
    """T6.2: the derived chapter code truncates at 19 characters.

    ``'The Rise of Nationalism in Europe'`` and ``'... in Asia'`` both derive
    ``10CBSS_TheRiseOfNat``. The reader keyed chapters on that code alone and
    silently merged them, with no issue recorded. The gold's own titles do not
    collide, which is exactly why a reference-anchored suite cannot see it.
    """
    layout = layouts.layout("canonical-current")
    rows = []
    for region in ("Europe", "Asia"):
        rows.append({
            ("chapter", "chapter_title"):
                f"The Rise of Nationalism in {region} (10_History_CBSE_NCERT)",
            ("topic", "topic_title"): f"Topic 01: Nationalism in {region}",
            ("topic", "pre_post_learning"): "Post",
            ("concept", "concept_title"): f"Nationalism in {region}",
            ("concept", "concept_display_name"): f"Nationalism in {region}",
            ("group", "group_name"): "Basic Group 01",
            ("group", "group_type"): "Basic",
            ("question", "question_label"): f"10CBSS_Nat{region}_PL_T01_C01 Q01",
            ("question", "question"): f"Why did nationalism rise in {region}?",
            ("question", "marks"): "1",
        })
    path = _save(
        _new_workbook_for(layout, {"objective": rows}), tmp_path, "collide.xlsx")

    counts = reader.import_workbook(db, path)
    europe = db.query(models.Chapter).filter_by(
        chapter_title="The Rise of Nationalism in Europe").one()
    asia = db.query(models.Chapter).filter_by(
        chapter_title="The Rise of Nationalism in Asia").one()
    # Two rows, one truncated code: the collision is real and is not merged.
    assert europe.id != asia.id
    assert europe.chapter_code == asia.chapter_code
    chapters = db.query(models.Chapter).filter_by(
        chapter_code=europe.chapter_code).all()
    assert {chapter.chapter_title for chapter in chapters} == {
        "The Rise of Nationalism in Europe",
        "The Rise of Nationalism in Asia",
    }
    assert any("chapter_code_collision" in issue for issue in counts["issues"]), \
        counts["issues"]
    # Neither chapter absorbed the other's concept.
    for chapter in chapters:
        titles = {
            concept.concept_title
            for topic in chapter.topics for concept in topic.concepts
        }
        assert len(titles) == 1, titles


def test_a_pre_topic_and_a_post_topic_with_one_title_import_as_two_topics(
    db, tmp_path,
):
    """T6.4: the reader's topic identity had no lane.

    A Pre topic and a Post topic sharing a title merged into whichever row was
    created first, and the other lane's concepts were re-parented under it —
    silently. The identity also goes through ``identity.topic_identity``, the
    ONE shared normaliser, not a local copy: the two Post rows below differ in
    casing and in the ``Topic NN:`` prefix and must still be one topic.
    """
    layout = layouts.layout("canonical-current")
    chapter_cell = "Nationalism Culmination (10_History_CBSE_NCERT)"
    rows = [
        {
            ("chapter", "chapter_title"): chapter_cell,
            ("topic", "topic_title"): "Topic 04: Culmination",
            ("topic", "pre_post_learning"): "Post",
            ("concept", "concept_title"): "Post culmination concept",
            ("concept", "concept_display_name"): "Post culmination concept",
            ("group", "group_type"): "Basic",
            ("group", "group_name"): "Basic Group 01",
        },
        {
            ("chapter", "chapter_title"): chapter_cell,
            ("topic", "topic_title"): "Culmination",
            ("topic", "pre_post_learning"): "Pre",
            ("concept", "concept_title"): "Pre culmination concept",
            ("concept", "concept_display_name"): "Pre culmination concept",
            ("group", "group_type"): "Basic",
            ("group", "group_name"): "Basic Group 01",
        },
        {
            ("chapter", "chapter_title"): chapter_cell,
            ("topic", "topic_title"): "culmination",
            ("topic", "pre_post_learning"): "Post",
            ("concept", "concept_title"): "Second post culmination concept",
            ("concept", "concept_display_name"): "Second post culmination concept",
            ("group", "group_type"): "Basic",
            ("group", "group_name"): "Basic Group 01",
        },
    ]
    path = _save(
        _new_workbook_for(layout, {"objective": rows}), tmp_path, "lanes.xlsx")

    counts = reader.import_workbook(db, path)
    chapter = db.query(models.Chapter).filter_by(
        chapter_title="Nationalism Culmination").one()
    wanted = identity.topic_identity("Culmination")
    topics = [
        topic for topic in chapter.topics
        if identity.topic_identity(topic.topic_title) == wanted
    ]
    assert len(topics) == 2, [t.topic_title for t in topics]
    by_lane = {topic.pre_post_learning: topic for topic in topics}
    assert set(by_lane) == {"Pre", "Post"}
    assert {c.concept_title for c in by_lane["Pre"].concepts} == {
        "Pre culmination concept"}
    # The two Post rows converge on ONE topic through the shared normaliser,
    # and the stored title is never rewritten to the incoming casing.
    assert {c.concept_title for c in by_lane["Post"].concepts} == {
        "Post culmination concept", "Second post culmination concept"}
    assert by_lane["Post"].topic_title == "Culmination"
    assert any("topic_lane_conflict" in issue for issue in counts["issues"]), \
        counts["issues"]


def test_the_reader_and_writer_column_geometry_twins_are_gone():
    """One module decides column geometry now, and it is the registry.

    ``reader._concept_fields`` and ``writer._sheet_concept_fields`` were twins
    deciding the same geometry of the same file under the same
    ``output_workbook_lock()``; they died together.
    """
    from app.bulk_import import writer

    for module, name in (
        (reader, "_concept_fields"),
        (reader, "_concept_len"),
        (reader, "_group_slice"),
        (writer, "_sheet_concept_fields"),
        (writer, "_sheet_concept_len"),
    ):
        assert not hasattr(module, name), f"{module.__name__}.{name}"


def test_scan_workbook_records_an_unidentified_sheet_instead_of_skipping_it(
    tmp_path,
):
    """The writer's own unflagged ``continue``, made visible.

    The writer may not fail closed — it is reached mid-run with the model
    budget already spent — so an unidentified sheet is a RECORDED mismatch.
    """
    from app.bulk_import import writer

    layout = layouts.layout("canonical-current")
    wb = _new_workbook_for(layout, {})
    wb[layout.sheet("subjective").sheet_name].cell(
        row=2, column=16).value = "kewyords"
    path = _save(wb, tmp_path, "scan_mismatch.xlsx")

    index = writer.scan_workbook(path)
    assert [m["sheet"] for m in index.mismatches] == [
        layout.sheet("subjective").sheet_name]
    assert index.mismatches[0]["code"] == "sheet_layout_unidentified"
    # The sheets that DID identify are still scanned.
    assert layout.sheet("objective").sheet_name in index.sheet_meta
    assert index.sheet_meta[layout.sheet("objective").sheet_name][
        "layout_id"] == "canonical-current"


def test_the_reader_appears_in_default_does_not_come_from_the_profile(
    db, tmp_path, monkeypatch,
):
    """T6.3: sourcing the reader's default from the active profile would stamp
    one school's wire value onto every other school's imported rows."""
    from app.services import assessment_profile

    monkeypatch.setitem(
        assessment_profile.DEFAULT_PROFILE, "appears_in", "Post-test")
    layout = layouts.layout("canonical-current")
    cells = {
        ("chapter", "chapter_title"): "Appears In Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Appears In Topic",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Appears In Concept",
        ("concept", "concept_display_name"): "Appears In Concept",
        ("group", "group_type"): "Basic",
        ("group", "group_name"): "Basic Group 01",
        ("question", "question_label"): "09CBPH_Appear_PL_T01_C01 Q01",
        ("question", "question"): "Where does this question appear?",
        ("question", "marks"): "1",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path, "appears_in.xlsx")

    reader.import_workbook(db, path)
    question = db.query(models.Question).filter_by(
        question_label="09CBPH_Appear_PL_T01_C01 Q01").one()
    assert question.question_appears_in == bi.APPEARS_IN_ALL


def test_a_content_sheet_the_file_omits_is_recorded_not_silent(db, tmp_path):
    """R4: the defect this gate replaced was the SILENCE, not the drop.

    The old reader skipped an unmatched sheet with a bare ``continue`` and no
    flag, so an entire missing Objective sheet read back as a clean import.
    An absent sheet loses no row -- it has none -- but a reviewer must be able
    to tell "this file has no Subjective questions" from "this file's
    Subjective sheet went missing", and only a recorded issue carries that.
    """
    layout = layouts.layout("canonical-current")
    cells = {
        ("chapter", "chapter_title"): "Omitted Sheet Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Only Descriptive",
        ("topic", "pre_post_learning"): "Post",
        ("concept", "concept_title"): "Only Descriptive Concept",
        ("concept", "concept_display_name"): "Only Descriptive Concept",
        ("group", "group_type"): "Basic",
        ("group", "group_name"): "Basic Group 01",
    }
    workbook = _new_workbook_for(layout, {"descriptive": [cells]})
    for sheet_name in ("Objective ", "Subjective"):
        workbook.remove(workbook[sheet_name])
    path = _save(workbook, tmp_path, "only_descriptive.xlsx")

    counts = reader.import_workbook(db, path)

    assert counts["layout_id"] == "canonical-current"
    # The sheets that ARE present still import.
    assert counts["concepts"] >= 1
    # And the two that are absent are named, each exactly once.
    for sheet_name in ("Objective ", "Subjective"):
        named = [i for i in counts["issues"] if i.startswith(f"{sheet_name}:")]
        assert len(named) == 1, (sheet_name, counts["issues"])
        assert "not present" in named[0]


def test_a_declared_non_content_tab_is_recorded_rather_than_passed_over(db, tmp_path):
    """The Doc Link tab is legitimately skipped -- and says so.

    ``canonical-current`` declares it non-content, so it is not a refusal.
    Recording it is what keeps "read past" distinguishable from "read".
    """
    layout = layouts.layout("canonical-current")
    workbook = _new_workbook_for(layout, {})
    workbook.create_sheet("Doc Link <> Each fields ").append(["anything"])
    path = _save(workbook, tmp_path, "with_doc_link.xlsx")

    counts = reader.import_workbook(db, path)

    ignored = [i for i in counts["issues"] if i.startswith("Doc Link")]
    assert len(ignored) == 1, counts["issues"]
    assert "not a content sheet" in ignored[0]


def test_a_blank_lane_cell_stores_a_lane_rather_than_an_empty_string(db, tmp_path):
    """T6.4: the topic identity key needs a total lane, so blank means Post.

    Pinned because it is a stored-value change: before the lane joined the
    key, a present-but-empty ``pre_post_learning`` cell persisted ``""``. An
    unlaned Topic row is worse than a Post one -- it cannot be told apart from
    either lane on the round trip -- but nothing pinned the choice.
    """
    layout = layouts.layout("canonical-current")
    cells = {
        ("chapter", "chapter_title"): "Blank Lane Chapter (09_Physics_CBSE_NCERT)",
        ("topic", "topic_title"): "Topic 01: Blank Lane",
        ("topic", "pre_post_learning"): "",
        ("concept", "concept_title"): "Blank Lane Concept",
        ("concept", "concept_display_name"): "Blank Lane Concept",
        ("group", "group_type"): "Basic",
        ("group", "group_name"): "Basic Group 01",
    }
    path = _save(
        _new_workbook_for(layout, {"objective": [cells]}),
        tmp_path, "blank_lane.xlsx")

    reader.import_workbook(db, path)

    # The reader strips the ``Topic NN:`` ordinal prefix when it stores the
    # title, so query the stored form rather than the authored one.
    topic = db.query(models.Topic).filter_by(topic_title="Blank Lane").one()
    assert topic.pre_post_learning == "Post"
