import copy
import io

import openpyxl
import pytest

from app import bulk_import as bi
from app import config


@pytest.mark.parametrize(("path", "writer_name"), [
    ("/data/export?scope=all", "write_workbook"),
    ("/data/export/questions?ids=1", "write_workbook"),
    ("/data/export/concepts?ids=1", "write_concepts_workbook"),
    ("/data/workbook/new?subject=History", "write_subject_workbook"),
])
def test_export_cell_limit_is_an_actionable_422(
    client, monkeypatch, path, writer_name,
):
    from app.api import data as data_api

    def oversized(*_args, **_kwargs):
        raise data_api.writer.ExcelCellLimitError(
            "Excel export blocked to prevent data loss: concept_details is "
            "32,768 characters."
        )

    monkeypatch.setattr(data_api.writer, writer_name, oversized)

    response = client.get(path)

    assert response.status_code == 422
    assert response.json()["detail"].startswith(
        "Excel export blocked to prevent data loss"
    )


@pytest.mark.parametrize(
    "path", ["/data/export?scope=all", "/data/export/questions?ids=1"],
)
def test_export_capacity_defect_is_an_actionable_422(
    client, monkeypatch, path,
):
    from app.api import data as data_api

    def malformed_shape(*_args, **_kwargs):
        raise data_api.writer.WorkbookCapacityError(
            "question 'Q1' subquestion 1 must be an object for lossless "
            "workbook export"
        )

    monkeypatch.setattr(data_api.writer, "write_workbook", malformed_shape)

    response = client.get(path)

    assert response.status_code == 422
    assert "lossless workbook export" in response.json()["detail"]


def test_health(client):
    response = client.get("/health")
    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["storage"]["status"] in {"ok", "critical", "unknown"}


def test_export_all_is_canonical_workbook(client):
    r = client.get("/data/export?scope=all")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert bi.SHEET_OBJECTIVE in wb.sheetnames
    assert bi.SHEET_SUBJECTIVE in wb.sheetnames
    assert bi.SHEET_DESCRIPTIVE in wb.sheetnames
    ws = wb[bi.SHEET_OBJECTIVE]
    # Row 2 carries the canonical field names.
    field_row = [c.value for c in ws[2]]
    assert field_row[: len(bi.OBJECTIVE_FIELDS)] == bi.OBJECTIVE_FIELDS


def test_export_all_includes_questionless_concepts(client, first_chapter, db):
    """A DB holding only generated concepts (no assessments yet) must not
    export as an empty workbook — concept-catalog rows are emitted."""
    from app import models

    concept_titles = {
        c.concept_title for t in db.get(models.Chapter, first_chapter["id"]).topics
        for c in t.concepts
    }
    assert concept_titles

    r = client.get("/data/export?scope=all")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb[bi.SHEET_OBJECTIVE]
    exported = {
        str(row[13] or "") for row in ws.iter_rows(min_row=3, values_only=True)
    }  # concept_display_name column carries the clean title
    assert concept_titles <= exported


def test_generation_appends_to_output_workbook(client, first_concept):
    session = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [first_concept["id"]],
    }).json()
    client.post(f"/build-assessments/sessions/{session['id']}/batches", json={
        "cognitive_skills": ["Understanding"], "difficulty_levels": ["Moderate"],
        "categories": ["Multiple Choice Question"], "question_type": "objective",
        "num_questions": 1,
    })
    client.post(f"/build-assessments/sessions/{session['id']}/generate")

    r = client.get("/data/export?scope=output")
    assert r.status_code == 200
    assert len(r.content) > 0


def test_append_only_never_overwrites(client, first_concept, db):
    """Re-running export of the same questions must not duplicate labels."""
    from app.bulk_import import writer

    # Run one generation so an output workbook exists.
    session = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [first_concept["id"]],
    }).json()
    client.post(f"/build-assessments/sessions/{session['id']}/batches", json={
        "cognitive_skills": ["Applying"], "difficulty_levels": ["High"],
        "categories": ["Long Answer"], "question_type": "descriptive",
        "num_questions": 1,
    })
    from tests.conftest import stream_result
    gen = stream_result(client.post(f"/build-assessments/sessions/{session['id']}/generate"))
    ids = gen["pipeline"]
    # Append the same question ids again -> all skipped.
    again = writer.append_questions(
        db, config.BULK_IMPORT_OUTPUT,
        [q["id"] for q in client.get("/data/questions?origin=concept_mapping").json()],
    )
    assert again["skipped"] >= 1
    assert again["objective"] == again["subjective"] == 0


def test_export_questions_selection(client, first_concept):
    """Per-functionality export: download just the generated questions."""
    session = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [first_concept["id"]],
    }).json()
    client.post(f"/build-assessments/sessions/{session['id']}/batches", json={
        "cognitive_skills": ["Understanding"], "difficulty_levels": ["Moderate"],
        "categories": ["Multiple Choice Question"], "question_type": "objective",
        "num_questions": 2,
    })
    from tests.conftest import stream_result
    gen = stream_result(client.post(f"/build-assessments/sessions/{session['id']}/generate"))
    ids = gen["question_ids"]
    assert len(ids) == 2

    r = client.get("/data/export/questions", params={"ids": ",".join(map(str, ids))})
    assert r.status_code == 200
    assert r.headers["content-disposition"].endswith('bulk_import_questions.xlsx"')
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb[bi.SHEET_OBJECTIVE]
    # Header rows + exactly the two generated questions.
    assert ws.max_row == 2 + 2


def test_export_questions_requires_ids(client):
    assert client.get("/data/export/questions", params={"ids": ""}).status_code == 400
    assert client.get("/data/export/questions", params={"ids": "x"}).status_code == 400


def test_export_concepts_selection(client, first_chapter, db):
    """Per-functionality export for Build Concepts: download generated concepts."""
    from app import models

    concept_ids = [
        c.id for t in db.get(models.Chapter, first_chapter["id"]).topics
        for c in t.concepts
    ][:2]
    assert concept_ids

    r = client.get("/data/export/concepts", params={"ids": ",".join(map(str, concept_ids))})
    assert r.status_code == 200
    assert r.headers["content-disposition"].endswith('bulk_import_concepts.xlsx"')
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb[bi.SHEET_OBJECTIVE]
    field_row = [c.value for c in ws[2]]
    assert field_row[: len(bi.OBJECTIVE_FIELDS)] == bi.OBJECTIVE_FIELDS
    # One concept-catalog row per concept (no tags here).
    assert ws.max_row == 2 + len(concept_ids)


def test_import_workbook_roundtrip(client, db, request):
    """A valid three-sheet selection round-trips without field loss."""
    from app import models
    from app.bulk_import import layouts

    chapter = models.Chapter(
        chapter_code="09CBPH-ROUNDTRIP",
        board="CBSE", grade="9", subject="Physics", unit="Roundtrip",
        chapter_title="Roundtrip Chapter",
        chapter_display_name="Roundtrip Chapter",
        chapter_duration="45",
        chapter_description="A deliberately strict-safe roundtrip fixture.",
    )
    topic = models.Topic(
        chapter=chapter,
        topic_title="Roundtrip Topic",
        topic_display_name="Roundtrip Topic",
        pre_post_learning="Post",
        topic_description="Carries all three assessment sheets.",
    )
    concept = models.Concept(
        topic=topic,
        concept_title="Roundtrip Concept",
        concept_display_name="Roundtrip Concept",
        concept_details="Description: strict-safe export and import.",
        keywords="roundtrip, workbook",
        sources="UpSchool DB",
    )
    group = models.Group(
        concept=concept,
        group_type="Basic",
        group_name="Roundtrip Basic",
        group_display_name="Roundtrip Basic",
        group_status="Active",
    )
    common = {
        "group": group,
        "cognitive_skills": "Understand",
        "question_source": "UpSchool DB",
        "question_disclaimer": "Roundtrip disclaimer.",
        "question_appears_in": bi.APPEARS_IN_ALL,
        "level_of_difficulty": "Moderate",
        "origin": "seed",
    }
    objective = models.Question(
        **common,
        sheet_kind="objective",
        question_label="09CBPH_Roundtrip_PL_T01_C01 Q01",
        question_category="Multiple Choice Question",
        question="Which expression names pH?",
        question_text="Which expression names pH?",
        marks=1, question_duration=1, math_keyboard="",
        answer_restriction="Specific",
        answers=[
            {
                "answer_type": "Equation",
                "answer_content": r"\text{pH}",
                "correct_answer": "1",
                "answer_weightage": "1",
            },
            {
                "answer_type": "Phrases",
                "answer_content": "temperature",
                "correct_answer": "0",
                "answer_weightage": "0",
            },
        ],
        sub_questions=[],
        answer_explanation="pH is the named quantity.",
    )
    descriptive = models.Question(
        **common,
        sheet_kind="descriptive",
        question_label="09CBPH_Roundtrip_PL_T01_C01 Q02",
        question_category="Short Answer",
        question="Explain the result.",
        question_text="Explain the result.",
        marks=2, question_duration=2, math_keyboard="No",
        answer_restriction="Open",
        answers=[],
        sub_questions=[{
            "text": "State the independently scored result.",
            "marks": "2",
            "keywords": [{
                "answer_type": "Phrases",
                "weightage": "2",
                "keyword": "[content]: states and explains the result",
            }],
        }],
        display_answer="A complete explanation of the result.",
        answer_explanation="",
    )
    subjective = models.Question(
        **common,
        sheet_kind="subjective",
        question_label="09CBPH_Roundtrip_PL_T01_C01 Q03",
        question_category="Fill in the Blanks",
        question="The recorded value is $$a$$.",
        question_text="The recorded value is $$a$$.",
        marks=1, question_duration=1, math_keyboard="No",
        answer_restriction="Specific",
        answers=[{
            "answer_type": "Phrases",
            "answer_content": "one",
            "answer_display": "one",
            "answer_weightage": "1",
            "placeholder": "a",
        }],
        sub_questions=[],
        answer_explanation="The blank is one.",
    )
    db.add_all([
        chapter, topic, concept, group, objective, descriptive, subjective,
    ])
    db.commit()
    chapter_id = chapter.id

    def cleanup_roundtrip_graph():
        db.rollback()
        persisted = db.get(models.Chapter, chapter_id)
        if persisted is not None:
            db.delete(persisted)
            db.commit()

    request.addfinalizer(cleanup_roundtrip_graph)
    selected = [objective, descriptive, subjective]
    selected_ids = [question.id for question in selected]
    before_export = {
        question.id: (
            question.question,
            question.question_text,
            question.display_answer,
            question.answer_explanation,
            question.answer_restriction,
            question.question_source,
            question.question_disclaimer,
            copy.deepcopy(question.answers),
            copy.deepcopy(question.sub_questions),
        )
        for question in selected
    }
    export = client.get(
        "/data/export/questions",
        params={"ids": ",".join(str(value) for value in selected_ids)},
    )
    assert export.status_code == 200, export.text
    db.expire_all()
    after_export = {
        question.id: (
            question.question,
            question.question_text,
            question.display_answer,
            question.answer_explanation,
            question.answer_restriction,
            question.question_source,
            question.question_disclaimer,
            copy.deepcopy(question.answers),
            copy.deepcopy(question.sub_questions),
        )
        for question in db.query(models.Question).filter(
            models.Question.id.in_(selected_ids)
        )
    }
    assert after_export == before_export
    workbook = openpyxl.load_workbook(io.BytesIO(export.content))
    descriptive_layout = layouts.sheet(
        layouts.REFERENCE_LAYOUT_ID, "descriptive",
    )
    descriptive_sheet = workbook[descriptive_layout.sheet_name]
    multipart_rows = 0
    exported_text = []
    for worksheet in workbook.worksheets:
        exported_text.extend(
            str(cell.value)
            for row in worksheet.iter_rows(min_row=3)
            for cell in row
            if cell.value is not None
        )
    for row_number in range(3, descriptive_sheet.max_row + 1):
        has_subquestion = any(
            descriptive_sheet.cell(
                row=row_number,
                column=descriptive_layout.column(
                    "question", f"sub_question_{number}",
                ) + 1,
            ).value
            for number in descriptive_layout.sub_question_numbers
        )
        if not has_subquestion:
            continue
        multipart_rows += 1
        for number in descriptive_layout.answer_block_numbers:
            for field in (
                f"answer_type_{number}",
                f"answer_content_{number}",
                f"answer_weightage_{number}",
            ):
                assert descriptive_sheet.cell(
                    row=row_number,
                    column=descriptive_layout.column("question", field) + 1,
                ).value in (None, "")
    subjective_layout = layouts.sheet(
        layouts.REFERENCE_LAYOUT_ID, "subjective",
    )
    subjective_sheet = workbook[subjective_layout.sheet_name]
    subjective_row = next(
        row for row in subjective_sheet.iter_rows(min_row=3, values_only=True)
        if any(value is not None for value in row)
    )
    assert subjective_row[
        subjective_layout.column("question", "answer_1")
    ] == "one"
    assert str(subjective_row[
        subjective_layout.column("question", "weightage_1")
    ]) == "1"
    workbook.close()
    assert multipart_rows == 1
    assert not any(r"\mathrm" in value for value in exported_text)
    assert any(r"\text{pH}" in value for value in exported_text)

    response = client.post("/data/import", files={
        "file": (
            "roundtrip.xlsx",
            io.BytesIO(export.content),
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet",
        ),
    })
    assert response.status_code == 200, response.text
    assert response.json()["questions"] == 0


@pytest.mark.parametrize("append_kind", ["concepts", "questions"])
def test_append_migrates_historical_cells_before_strict_reimport(
    db, tmp_path, append_kind,
):
    """Both staged append paths repair old rows on the workbook copy.

    Contract v2.0 §17: line breaks inside a cell ship as ``<br>``, so the
    repaired option lines are read back on that marker, never on ``\\n``.
    """

    from app import models
    from app.bulk_import import layouts, reader, writer
    from app.services import katex_rules

    path = tmp_path / f"legacy-{append_kind}-output.xlsx"
    workbook = writer._new_workbook()

    def put(worksheet, sheet_layout, block, field, value, *, row=3):
        column = sheet_layout.column(block, field)
        assert column is not None, (sheet_layout.kind, block, field)
        worksheet.cell(row=row, column=column + 1).value = value

    def seed_front(worksheet, sheet_layout, suffix):
        put(worksheet, sheet_layout, "chapter", "chapter_title", suffix)
        put(worksheet, sheet_layout, "topic", "topic_title", "Topic")
        put(
            worksheet, sheet_layout, "topic", "pre_post_learning", "Post",
        )
        put(worksheet, sheet_layout, "concept", "concept_title", "Concept")
        put(worksheet, sheet_layout, "group", "group_name", "Basic Group 01")
        put(worksheet, sheet_layout, "group", "group_type", "Basic")

    objective_layout = writer._target_sheet("objective")
    objective = workbook[objective_layout.sheet_name]
    seed_front(objective, objective_layout, f"LEGACY-OBJ-{append_kind}")
    put(
        objective, objective_layout, "question", "question_label",
        f"LEGACY-OBJ-{append_kind}-Q1",
    )
    put(
        objective, objective_layout, "question", "question",
        "Answer A) stays prose.\nA) Alpha\nB) Beta",
    )
    put(
        objective, objective_layout, "question", "question_text",
        "Study the table.\n| Name | Value |\n|---|---:|\n| Alpha | 1 |"
        "\nA) Alpha\nB) Beta",
    )
    put(objective, objective_layout, "question", "marks", "1")
    put(
        objective, objective_layout, "question", "question_duration", "1",
    )
    put(
        objective, objective_layout, "question", "answer_type_1", "Phrases",
    )
    put(
        objective, objective_layout, "question", "answer_content_1",
        r"[Katex] \mathrm{x}+1 [/Katex]",
    )
    put(
        objective, objective_layout, "question", "correct_answer_1", "1",
    )
    put(
        objective, objective_layout, "question", "answer_weightage_1", "1",
    )

    descriptive_layout = writer._target_sheet("descriptive")
    descriptive = workbook[descriptive_layout.sheet_name]
    seed_front(descriptive, descriptive_layout, f"LEGACY-DES-{append_kind}")
    put(
        descriptive, descriptive_layout, "question", "question_label",
        f"LEGACY-DES-{append_kind}-Q1",
    )
    put(
        descriptive, descriptive_layout, "question", "question",
        "Explain the supplied relationship.",
    )
    put(
        descriptive, descriptive_layout, "question", "question_text",
        "Explain the supplied relationship.",
    )
    put(descriptive, descriptive_layout, "question", "marks", "1")
    put(
        descriptive, descriptive_layout, "question", "math_keyboard",
        "No",
    )
    put(
        descriptive, descriptive_layout, "question", "question_duration",
        "1",
    )
    put(
        descriptive, descriptive_layout, "question", "answer_type_1",
        "Phrases",
    )
    put(
        descriptive, descriptive_layout, "question", "answer_content_1",
        "One complete explanation.",
    )
    table = (
        r"[Katex] \begin{array}{cc}\text{Name}&\text{Value}\\A&1"
        r"\end{array} [/Katex]"
    )
    put(
        descriptive, descriptive_layout, "question", "display_answer",
        table,
    )
    put(
        descriptive, descriptive_layout, "question", "answer_weightage_1",
        "1",
    )
    put(
        descriptive, descriptive_layout, "question", "sub_question_1",
        table,
    )
    put(
        descriptive, descriptive_layout, "question", "sub_question_marks_1",
        "1",
    )
    put(
        descriptive, descriptive_layout, "question", "sq1_answer_type_1",
        "Phrases",
    )
    put(
        descriptive, descriptive_layout, "question", "sq1_weightage_1", "1",
    )
    put(
        descriptive, descriptive_layout, "question", "sq1_keyword_1",
        "[Katex] y^2 [/Katex]",
    )
    workbook.save(path)
    workbook.close()

    if append_kind == "concepts":
        concept_id = db.query(models.Concept.id).order_by(models.Concept.id).first()
        assert concept_id is not None
        result = writer.append_concepts(db, path, [concept_id[0]])
    else:
        # Even a no-new-row append is a publication touch of this durable
        # workbook and must migrate its historical rows before saving.
        result = writer.append_questions(db, path, [])
    assert result["legacy_cells_normalized"] > 0

    migrated = openpyxl.load_workbook(path, data_only=True)
    objective = migrated[objective_layout.sheet_name]
    objective_type = objective.cell(
        row=3,
        column=objective_layout.column("question", "answer_type_1") + 1,
    ).value
    objective_content = objective.cell(
        row=3,
        column=objective_layout.column("question", "answer_content_1") + 1,
    ).value
    assert objective_type == "Equation"
    assert "[Katex]" not in objective_content
    assert r"\mathrm" not in objective_content
    assert r"\text{x}" in objective_content
    assert katex_rules.answer_cell_issues(
        objective_type, objective_content,
    ) == []
    migrated_question = objective.cell(
        row=3,
        column=objective_layout.column("question", "question") + 1,
    ).value
    migrated_question_text = objective.cell(
        row=3,
        column=objective_layout.column("question", "question_text") + 1,
    ).value
    assert "Answer A) stays prose." in migrated_question
    assert "<br>a) Alpha<br>b) Beta" in migrated_question
    assert "\n" not in migrated_question
    assert "Table row 1, column 1: Name" in migrated_question_text
    assert "|---|" not in migrated_question_text
    assert "<br>a) Alpha<br>b) Beta" in migrated_question_text
    assert "\n" not in migrated_question_text

    descriptive = migrated[descriptive_layout.sheet_name]
    answer_display = descriptive.cell(
        row=3,
        column=descriptive_layout.column("question", "display_answer") + 1,
    ).value
    sub_question = descriptive.cell(
        row=3,
        column=descriptive_layout.column("question", "sub_question_1") + 1,
    ).value
    keyword_type = descriptive.cell(
        row=3,
        column=descriptive_layout.column(
            "question", "sq1_answer_type_1",
        ) + 1,
    ).value
    keyword = descriptive.cell(
        row=3,
        column=descriptive_layout.column("question", "sq1_keyword_1") + 1,
    ).value
    main_rubric = {
        field: descriptive.cell(
            row=3,
            column=descriptive_layout.column("question", field) + 1,
        ).value
        for field in (
            "answer_type_1", "answer_content_1", "answer_weightage_1",
        )
    }
    migrated.close()
    assert answer_display == table
    assert sub_question == table
    assert katex_rules.rich_text_issues(answer_display) == []
    assert katex_rules.rich_text_issues(sub_question) == []
    assert keyword_type == "Equation"
    assert "[Katex]" not in keyword
    assert katex_rules.answer_cell_issues(keyword_type, keyword) == []
    assert main_rubric == {
        "answer_type_1": None,
        "answer_content_1": None,
        "answer_weightage_1": None,
    }

    # The exact workbook the append path published now clears the same strict
    # preflight used by POST /data/import.
    reader.import_workbook(db, path, strict_content=True)


def test_fresh_export_lowercases_legacy_objective_labels_on_the_copy(
    client, db,
):
    """Fresh public export is strict-importable without rewriting storage.

    Contract v2.0 §17: the exported cell carries ``<br>`` line breaks while
    storage keeps ``\\n``; the strict re-import below reads them back.
    """

    from app import models
    from app.bulk_import import writer

    question = (
        db.query(models.Question)
        .filter(models.Question.sheet_kind == "objective")
        .order_by(models.Question.id)
        .first()
    )
    assert question is not None
    before = (question.question, question.question_text)
    legacy_question = "Answer A) stays prose.\nA) Alpha\nB) Beta"
    legacy_text = (
        "Study the table.\n| Name | Value |\n|---|---:|\n| Alpha | 1 |"
        "\nA) Alpha\nB) Beta"
    )
    question.question = legacy_question
    question.question_text = legacy_text
    db.commit()
    label = question.question_label

    try:
        exported = client.get(
            "/data/export/questions", params={"ids": str(question.id)},
        )
        assert exported.status_code == 200, exported.text
        db.expire_all()
        stored = db.get(models.Question, question.id)
        assert (stored.question, stored.question_text) == (
            legacy_question, legacy_text,
        )

        workbook = openpyxl.load_workbook(io.BytesIO(exported.content))
        sheet_layout = writer._target_sheet("objective")
        worksheet = workbook[sheet_layout.sheet_name]
        label_column = sheet_layout.column("question", "question_label")
        row_number = next(
            row
            for row in range(3, worksheet.max_row + 1)
            if worksheet.cell(row=row, column=label_column + 1).value == label
        )
        question_value = worksheet.cell(
            row=row_number,
            column=sheet_layout.column("question", "question") + 1,
        ).value
        text_value = worksheet.cell(
            row=row_number,
            column=sheet_layout.column("question", "question_text") + 1,
        ).value
        workbook.close()
        assert "Answer A) stays prose." in question_value
        assert "<br>a) Alpha<br>b) Beta" in question_value
        assert "\n" not in question_value
        assert "Table row 1, column 1: Name" in text_value
        assert "|---|" not in text_value
        assert "<br>a) Alpha<br>b) Beta" in text_value
        assert "\n" not in text_value

        response = client.post(
            "/data/import",
            files={
                "file": (
                    "fresh-strict-roundtrip.xlsx",
                    io.BytesIO(exported.content),
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet",
                ),
            },
        )
        assert response.status_code == 200, response.text
    finally:
        stored = db.get(models.Question, question.id)
        stored.question, stored.question_text = before
        db.commit()
