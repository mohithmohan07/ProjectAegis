"""S8 — one renderer: the staged home verdict, the de-raising, the profile.

Three things land here and each is atomic (spec-step8 S8):

1. ``assessment_release.unresolved_question_homes`` + its ``create_release``
   wiring + the deletion of ``_readiness``'s unplaced read. Any two without
   the third is either a double refusal or NO refusal at all.
2. All fourteen raises in ``render_master_file`` and ``_question_record``
   stop raising, each in the same change as the verdict that replaces it.
3. ``sheet_kinds`` added, ``allow_subjective_rows`` deleted, one accessor,
   and ``validate_master_file`` reading the new key.

Plus the two owner rulings: OD5 (a questionless concept is ONE tail row)
and OD3 (``keywords`` and ``related_concepts`` ship filled).
"""
from __future__ import annotations

import ast
import copy
import inspect
import io
import json
from pathlib import Path

import openpyxl
import pytest

from app.bulk_import import assessment_workbook as mp
from app.bulk_import import writer as bi_writer
from app.services import assessment_master_refiner as refiner
from app.services import assessment_profile
from app.services import assessment_release as rel

from tests.test_mes_dual_output import _snapshot


# --------------------------------------------------------------------------- #
# The test-only second profile (T12/M8). It exists to prove the plumbing
# reaches every site and must never become a second pinned school.
# --------------------------------------------------------------------------- #

REFERENCE_2 = assessment_profile.register({
    **assessment_profile.DEFAULT_PROFILE,
    "name": "reference-2",
    "sheet_kinds": ("objective", "descriptive", "subjective"),
    "forced_blank_fields": ("question_disclaimer",),
})


# --------------------------------------------------------------------------- #
# ATOMIC 2 — the fourteen raises
# --------------------------------------------------------------------------- #

def _raises_in(function) -> list[int]:
    source = inspect.getsource(function)
    tree = ast.parse(source)
    return [
        node.lineno for node in ast.walk(tree) if isinstance(node, ast.Raise)
    ]


def test_the_master_renderer_contains_no_raise():
    """The cheapest possible guard against the fifteenth raise.

    A raise inside ``render_master_file`` — called unguarded from
    ``publish_release`` — means NO WORKBOOK IS WRITTEN FOR ANY OF THE FOUR
    OUTPUTS. Every one of its conditions is decidable at staging from the
    frozen snapshot, so none of them may cost the workbooks.
    """

    assert _raises_in(mp.render_master_file) == []
    assert _raises_in(mp._question_record) == []
    assert _raises_in(mp.render_concept_file) == []


@pytest.mark.parametrize(
    ("line", "mutate", "code"),
    [
        pytest.param(
            ":315",
            lambda s: s["groups"][0].update(group_key=""),
            None, id="315-blank-group-key"),
        pytest.param(
            ":317",
            lambda s: s["groups"].append({
                **copy.deepcopy(s["groups"][1]),
                "group_key": s["groups"][0]["group_key"]}),
            None, id="317-duplicate-group-key"),
        pytest.param(
            ":324",
            lambda s: s["groups"][0].update(concept_key="C_MISSING"),
            rel.GROUP_CONCEPT_HOME_UNKNOWN, id="324-unknown-concept-home"),
        pytest.param(
            ":331",
            lambda s: s["topics"][0]["concepts"][0].update(
                concept_display_name=""),
            rel.GROUP_HOME_UNNAMED, id="331-home-unnamed"),
        pytest.param(
            ":336",
            lambda s: s["groups"][0].update(group_type="Wobbly"),
            None, id="336-invalid-group-type"),
        pytest.param(
            ":343",
            lambda s: s["groups"][0].update(group_name="not the visible name"),
            rel.GROUP_VISIBLE_NAME_MISMATCH, id="343-visible-name-mismatch"),
        pytest.param(
            ":373",
            lambda s: s["candidates"][0].update(concept_key="C_B"),
            rel.GROUP_HOME_DISAGREEMENT, id="373-group-home-disagreement"),
    ],
)
def test_every_group_edge_raise_became_a_named_defect(line, mutate, code):
    """All seven, by line, with all four workbooks written.

    ``code is None`` marks the three that already had a staging twin and
    simply went: ``validate_group``'s ``_missing`` over ``_GROUP_REQUIRED``
    (``:315``), ``duplicate_group_keys`` (``:317``) and ``validate_group``'s
    ``group_type`` enum (``:336``). Nothing was added for them, and the
    twin is asserted here rather than trusted.
    """

    snapshot = copy.deepcopy(_snapshot())
    mutate(snapshot)

    output = mp.build_dual_output(snapshot)
    assert output["concepts_xlsx"][:2] == b"PK"
    assert output["master_xlsx"][:2] == b"PK"

    findings = rel.unresolved_question_homes(snapshot)
    if code is not None:
        assert code in {f["code"] for f in findings}, line
    else:
        # The staging twin, measured rather than assumed.
        groups = [
            {**dict(g), "concept_id": 1} for g in snapshot["groups"]]
        twin = [
            error for group in groups for error in rel.validate_group(group)
        ] + [
            f"duplicate group_key {key!r}"
            for key in rel.duplicate_group_keys(groups)
        ]
        assert twin, line
        # …and the read-back twin, EXERCISED rather than short-circuited.
        # ``assert twin or validate_master_file(...)`` never evaluated this
        # branch at all — ``twin`` was asserted truthy one line above — and
        # it hid a fact worth recording: only TWO of the three have a
        # read-back twin. [measured] ``validate_master_file`` never
        # re-checks ``group_type``, so ``:336``'s single refusal is
        # ``validate_group`` at freeze, asserted above.
        read_back = mp.validate_master_file(
            mp.parse_workbook(output["master_xlsx"]), snapshot)
        if line in (":315", ":317"):
            assert read_back, line
        else:
            assert line == ":336" and read_back == [], (line, read_back)


def test_the_six_deleted_raises_each_still_have_their_staging_twin():
    """Named individually so a reviewer can check rather than trust."""

    snapshot = copy.deepcopy(_snapshot())
    group = {**dict(snapshot["groups"][0]), "concept_id": 1}
    assert "missing group_key" in rel.validate_group({**group, "group_key": ""})
    assert rel.duplicate_group_keys([group, dict(group)]) == [
        group["group_key"]]
    assert any(
        "group_type must be one of" in error
        for error in rel.validate_group({**group, "group_type": "Wobbly"}))

    candidate = {
        **dict(snapshot["candidates"][0]),
        "blueprint_cell_id": "CELL-1",
        "restriction_reason": "bounded",
        "source_atom_ids": ["QINV-1"],
    }
    assert rel.validate_candidate(candidate) == []
    assert "question_duration must be finite and positive" in (
        rel.validate_candidate({**candidate, "question_duration": 0}))
    assert "objective math_keyboard must be exactly blank" in (
        rel.validate_candidate({**candidate, "math_keyboard": "Yes"}))
    descriptive = {
        **dict(snapshot["candidates"][1]),
        "blueprint_cell_id": "CELL-2",
        "restriction_reason": "open",
        "source_atom_ids": ["QINV-2"],
    }
    assert "descriptive math_keyboard must be exactly Yes or No" in (
        rel.validate_candidate({**descriptive, "math_keyboard": ""}))


# --------------------------------------------------------------------------- #
# OWNER RULING OD5 / T16 — questionless concepts
# --------------------------------------------------------------------------- #

def _objective_rows(master: bytes) -> list[dict]:
    return mp.parse_workbook(master)["sheets"]["Objective"]["rows"]


def test_a_questionless_concept_is_one_tail_row_stopping_at_the_concept_columns():
    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    rows = _objective_rows(master)
    tail = [
        row for row in rows
        if str(row.get("concept_title") or "").startswith("Three-dimensional")
    ]
    assert len(tail) == 1
    assert tail[0] is rows[-1], "the tail row is appended LAST"
    # Filled to the concept columns…
    assert tail[0]["chapter_title"]
    assert tail[0]["topic_title"]
    assert tail[0]["concept_title"]
    assert tail[0]["concept_details"]
    # …and empty from the Group band onward.
    fields = mp.FIELDS["Objective"]
    group_band_start = fields.index("group_name")
    assert all(
        not str(tail[0].get(field) or "").strip()
        for field in fields[group_band_start:]
    )
    assert [c["concept_key"] for c in issues["questionless_concepts"]] == [
        "C_B"]


def test_the_master_validator_accepts_a_questionless_concept():
    """Fails before this slice: three "group missing from Master" errors."""

    snapshot = _snapshot()
    master, issues = mp.render_master_file(snapshot)
    assert mp.validate_master_file(
        mp.parse_workbook(master), snapshot,
        group_provenance=issues["group_provenance"],
    ) == []


def test_the_questionless_shells_stay_in_the_payload_and_the_difference_is_recorded():
    snapshot = _snapshot()
    _master, issues = mp.render_master_file(snapshot)
    shells = {
        group["group_key"] for group in snapshot["groups"]
        if group["concept_key"] == "C_B"
    }
    assert len(shells) == 3
    recorded = issues["questionless_concepts"][0]
    assert set(recorded["shell_group_keys"]) == shells
    assert recorded["concept_title"] == (
        "Three-dimensional shape (06MSMA_T01_ThreeDim)")


def test_a_concept_with_questions_still_emits_its_unrepresented_group_rows():
    """The negative control that pins where the line is."""

    snapshot = _snapshot()
    master, _ = mp.render_master_file(snapshot)
    rows = _objective_rows(master)
    catalogue = [
        row for row in rows
        if not row.get("question_label")
        and str(row.get("concept_title") or "").startswith("Two-dimensional")
    ]
    assert [row["group_type"] for row in catalogue] == ["Advanced"]


def test_the_master_refiner_readback_still_fires_for_a_missing_catalogue_row(
    monkeypatch,
):
    """The negative control for OD5's scoping of the SECOND read-back.

    Scoping ``placed_concepts`` is what stops every questionless concept
    becoming a "refiner-readback: group … has no rendered Master row"
    error. A scoping that went one step further — "a group with no
    questions" — would ALSO stop catching a genuinely missing catalogue
    row on a concept that does have questions, and the read-back would
    stay green while proving nothing. This drops one provenance entry for
    such a group and asserts the error still arrives.
    """

    from tests.test_assessment_master_refiner import _payload

    payload = copy.deepcopy(_payload())
    real_render = mp.render_master_file
    # A group the PAYLOAD declares — the refiner's read-back walks
    # ``payload["groups"]``, not the shells the snapshot builder adds.
    victim = str(payload["groups"][0]["group_key"])
    dropped: list[str] = []

    def _render_missing_one_row(snapshot, profile=None):
        master, issues = real_render(snapshot, profile)
        kept = [
            entry for entry in issues["group_provenance"]
            if entry["group_key"] != victim
        ]
        assert len(kept) < len(issues["group_provenance"])
        dropped.append(victim)
        issues["group_provenance"] = kept
        return master, issues

    monkeypatch.setattr(
        refiner.assessment_workbook, "render_master_file",
        _render_missing_one_row)
    state = refiner._validation_state(
        payload, assessment_profile.DEFAULT_PROFILE)
    assert dropped, "the probe never dropped a row"
    assert any(
        f"group {dropped[0]!r} has no rendered Master row" in error
        for error in state["errors"]
    ), state["errors"]


# --------------------------------------------------------------------------- #
# ATOMIC 3 — the profile key (B3)
# --------------------------------------------------------------------------- #

def test_allow_subjective_rows_is_deleted_not_kept_beside_sheet_kinds():
    assert "allow_subjective_rows" not in assessment_profile.DEFAULT_PROFILE
    assert assessment_profile.DEFAULT_PROFILE["sheet_kinds"] == (
        "objective", "descriptive")
    live = Path("app/services").rglob("*.py")
    for path in [*live, *Path("app/bulk_import").rglob("*.py")]:
        for number, line in enumerate(
            path.read_text(encoding="utf-8").splitlines(), start=1
        ):
            code = line.split("#")[0]
            assert "allow_subjective_rows" not in code, (path, number)


def test_the_recorded_golden_profile_still_resolves():
    """``tests/golden/rne_assessment_candidates.json`` must not move.

    It records the profile its run executed under as
    ``{name, appears_in, allow_subjective_rows}`` and feeds it straight
    into ``decide_cells``, which deliberately does not call ``resolve``.
    The accessor's fallback to ``DEFAULT_PROFILE`` is what keeps that
    dict resolving after the key it carries was deleted.
    """

    golden = json.loads(
        Path("tests/golden/rne_assessment_candidates.json").read_text(
            encoding="utf-8"))
    recorded = golden["profile"]
    assert "allow_subjective_rows" in recorded
    assert "sheet_kinds" not in recorded
    assert assessment_profile.sheet_kinds(recorded) == (
        "objective", "descriptive")

    from app.services import assessment_cells as cells
    assert cells._allowed_sheet_kinds(recorded) == (
        "objective", "descriptive")


def test_sheet_kinds_is_read_through_one_accessor():
    """No module outside ``assessment_profile`` indexes the key directly."""

    for root in ("app/services", "app/bulk_import", "app/api"):
        for path in Path(root).rglob("*.py"):
            if path.name == "assessment_profile.py":
                continue
            text = path.read_text(encoding="utf-8")
            assert '["sheet_kinds"]' not in text, path
            assert '.get("sheet_kinds")' not in text, path


def test_every_profile_key_has_a_named_reader():
    """The key-with-no-reader pin (C10/T12/M4).

    A key that no code reads is worse than no key: it reads as plumbing
    that works. This is also what enforces T12/M6b — a tree that kept both
    ``allow_subjective_rows`` and ``sheet_kinds`` fails here on the one
    with no reader.
    """

    readers = {
        # ``name`` and ``appears_in`` used to point at their own accessors
        # in ``assessment_profile.py``, which the escape clause below then
        # waved through — so the pin proved nothing for two of the seven
        # keys while the renderer indexed ``profile["name"]`` and
        # ``profile["appears_in"]`` directly, bypassing the per-key
        # DEFAULT_PROFILE fallback. Both now name a real outside reader.
        "name": (
            "app/bulk_import/assessment_workbook.py", "build_dual_output"),
        "appears_in": (
            "app/bulk_import/assessment_workbook.py", "validate_master_file"),
        "sheet_kinds": (
            "app/bulk_import/assessment_workbook.py", "validate_master_file"),
        "forced_blank_fields": (
            "app/bulk_import/assessment_workbook.py", "_row_values"),
        "question_source": (
            "app/bulk_import/assessment_workbook.py", "_question_record"),
        "group_status": (
            "app/bulk_import/assessment_workbook.py", "_group_record_fields"),
        "automatic_secondary_tags": (
            "app/services/assessment_release.py", "validate_placement"),
    }
    assert set(readers) == set(assessment_profile.DEFAULT_PROFILE)
    for key, (path, function) in readers.items():
        # No escape clause: every named reader is OUTSIDE
        # ``assessment_profile`` and goes through the accessor.
        assert not path.endswith("assessment_profile.py"), (key, path)
        source = Path(path).read_text(encoding="utf-8")
        assert f"def {function}(" in source, (key, path, function)
        assert f"assessment_profile.{key}(" in source, (key, path)

    # …and no module outside ``assessment_profile`` indexes ANY profile key
    # directly. A direct index skips the per-key DEFAULT_PROFILE fallback,
    # which is the only reason the recorded golden profile still resolves.
    for root in ("app/services", "app/bulk_import", "app/api"):
        for module in Path(root).rglob("*.py"):
            if module.name == "assessment_profile.py":
                continue
            text = module.read_text(encoding="utf-8")
            for key in assessment_profile.DEFAULT_PROFILE:
                assert f'profile["{key}"]' not in text, (module, key)
                assert f"profile['{key}']" not in text, (module, key)


def test_forced_blank_fields_is_read_at_all_three_sites():
    """Renderer, read-back gate, and both snapshot builders (T12/M4)."""

    workbook = Path(
        "app/bulk_import/assessment_workbook.py").read_text(encoding="utf-8")
    assert workbook.count("assessment_profile.forced_blank_fields(") >= 3
    for path in (
        "app/services/assessment_release_snapshot.py",
        "app/services/assessment_release_service.py",
    ):
        assert '"chapter_duration"' in Path(path).read_text(encoding="utf-8")

    # And the lever actually moves: reference-2 un-blanks chapter_duration.
    snapshot = copy.deepcopy(_snapshot())
    snapshot["chapter"]["chapter_duration"] = "45 minutes"
    default_rows = _objective_rows(mp.render_master_file(snapshot)[0])
    widened_rows = _objective_rows(
        mp.render_master_file(snapshot, REFERENCE_2)[0])
    assert all(
        not str(row.get("chapter_duration") or "") for row in default_rows)
    assert all(
        str(row.get("chapter_duration") or "") == "45 minutes"
        for row in widened_rows)


def test_renderable_sheet_kinds_is_a_module_constant_not_the_profile():
    """T7.5/B3 — a profile-derived map has no ``None`` branch to take."""

    assert mp.RENDERABLE_SHEET_KINDS == ("objective", "descriptive")
    assert mp.sheet_for_kind() == {
        "objective": "Objective", "descriptive": "Descriptive"}
    # The layout carries Subjective; the RENDERER does not.
    from app.bulk_import import layouts
    assert "subjective" in layouts.layout(
        layouts.REFERENCE_LAYOUT_ID).sheet_name_by_kind()
    assert "subjective" not in mp.sheet_for_kind()
    workbook = Path(
        "app/bulk_import/assessment_workbook.py").read_text(encoding="utf-8")
    assert "MES" not in workbook.split("_MANIFEST_PATH")[1]


def _subjective_snapshot() -> dict:
    snapshot = copy.deepcopy(_snapshot())
    snapshot["candidates"][0]["sheet_kind"] = "subjective"
    return snapshot


def test_a_subjective_candidate_a_widened_profile_allows_is_a_named_defect_not_a_render_error():
    snapshot = _subjective_snapshot()
    candidate = {
        **snapshot["candidates"][0],
        "blueprint_cell_id": "CELL-1",
        "restriction_reason": "bounded",
        "source_atom_ids": ["QINV-1"],
    }
    # (i) it PASSES validate_candidate under the widened profile — the
    #     plumbing working, and what fails today.
    assert rel.validate_candidate(candidate, REFERENCE_2) == []
    # (ii) and yields a NAMED sheet_kind_not_renderable defect.
    findings = rel.unresolved_question_homes(snapshot, REFERENCE_2)
    named = [
        f for f in findings if f["code"] == rel.SHEET_KIND_NOT_RENDERABLE]
    assert len(named) == 1
    assert named[0]["sheet_kind"] == "subjective"
    assert named[0]["profile"] == "reference-2"
    assert named[0]["renderable_sheet_kinds"] == ["objective", "descriptive"]
    # (iii)/(iv) both workbooks are written, and no Subjective data row.
    output = mp.build_dual_output(snapshot, REFERENCE_2)
    assert output["concepts_xlsx"][:2] == b"PK"
    assert output["master_xlsx"][:2] == b"PK"
    parsed = mp.parse_workbook(output["master_xlsx"])
    assert parsed["sheets"]["Subjective"]["rows"] == []
    # (v) no reason string anywhere names a school.
    for item in output["manifest"]["issues"]["unplaced"]:
        assert "MES" not in item["reason"]
    for finding in findings:
        assert "MES" not in finding["message"]


def test_a_widened_profiles_subjective_candidate_blocks_the_upload_and_still_downloads(
    db, client,
):
    """Spec S8 clauses (iii)/(iv), end to end rather than in the renderer.

    The byte-string half above proves the workbooks exist. This proves the
    consequence: ``_readiness`` BLOCKED, the database write refused, and
    every artifact still served 200 — the fact that distinguishes a named
    defect from the raise it replaced, which cost all four outputs.
    """

    from app.services import assessment_release_service as svc
    from tests.test_mes_release_lifecycle import _fresh_release, OWNER

    def widen(payload):
        payload["candidates"][0]["sheet_kind"] = "subjective"
        for cell in payload.get("blueprint_cells") or []:
            cell["sheet_kind"] = "subjective"

    release, _payload, label = _fresh_release(
        db, mutate=widen,
        provider_identity={"profile": "reference-2"})
    assert any(
        error.startswith(rel.SHEET_KIND_NOT_RENDERABLE)
        for error in release.diagnostics["payload_errors"]
    ), release.diagnostics["payload_errors"]

    published = svc.publish_release(db, release)
    assert published.diagnostics["readiness"] == svc.BLOCKED
    with pytest.raises(svc.UploadRefused, match="blocked"):
        svc.upload_master_to_database(db, published, owner_sub=OWNER)
    for filename in ("master.xlsx", "concepts.xlsx"):
        response = client.get(
            f"/build-assessments/releases/{published.id}/{filename}")
        assert response.status_code == 200, filename
        assert response.content[:2] == b"PK"
    # And the label reaches no data row of either workbook.
    master = client.get(
        f"/build-assessments/releases/{published.id}/master.xlsx").content
    for sheet in mp.parse_workbook(master)["sheets"].values():
        assert all(
            str(row.get("question_label") or "") != label
            for row in sheet["rows"])


def test_the_reference_profile_still_refuses_a_subjective_candidate_at_freeze():
    """The negative control: the two codes must not double-report."""

    snapshot = _subjective_snapshot()
    candidate = {
        **snapshot["candidates"][0],
        "blueprint_cell_id": "CELL-1",
        "restriction_reason": "bounded",
        "source_atom_ids": ["QINV-1"],
    }
    errors = rel.validate_candidate(candidate)
    assert any("sheet_kind must be one of" in error for error in errors)
    assert rel.SHEET_KIND_NOT_RENDERABLE not in {
        f["code"] for f in rel.unresolved_question_homes(snapshot)}


# --------------------------------------------------------------------------- #
# OD3 / T3.3 / T3.7 — the restored columns
# --------------------------------------------------------------------------- #

def test_keywords_ship_filled_on_all_four_outputs():
    """OD3: the gold leaving them blank is fill practice, not a rule.

    OD3 names BOTH columns, so both are asserted: a change that blanked
    ``related_concepts`` "for parity with the gold" used to pass here.
    The two renderers are the two projections every one of the four
    outputs is rendered by — Outputs 01/02 are the Pre lane's pair and
    03/04 the Post lane's, same functions, different snapshot.
    """

    snapshot = copy.deepcopy(_snapshot())
    # The fixture is a POST snapshot, whose ``related_concepts`` is blank
    # until a relations pass exists (T3.3); what OD3 forbids is the
    # renderer or the profile BLANKING an authored value, so an authored
    # value is what is asserted here.
    snapshot["topics"][0]["concepts"][0]["related_concepts"] = (
        "Three-dimensional shape (06MSMA_T01_ThreeDim)")
    for data in (
        mp.render_concept_file(snapshot),
        mp.render_master_file(snapshot)[0],
    ):
        rows = mp.parse_workbook(data)["sheets"]["Objective"]["rows"]
        assert any(str(row.get("keywords") or "").strip() for row in rows)
        assert any(
            str(row.get("related_concepts") or "").strip() for row in rows)
    for field in ("keywords", "related_concepts"):
        assert field not in assessment_profile.forced_blank_fields()


def test_topic_concept_labels_is_populated_in_every_projection():
    snapshot = copy.deepcopy(_snapshot())
    snapshot["topics"][0]["topic_concept_labels"] = (
        "Two-dimensional shape (06MSMA_T01_TwoDim), "
        "Three-dimensional shape (06MSMA_T01_ThreeDim)")
    for data in (
        mp.render_concept_file(snapshot),
        mp.render_master_file(snapshot)[0],
    ):
        rows = mp.parse_workbook(data)["sheets"]["Objective"]["rows"]
        assert all(
            str(row.get("topic_concept_labels") or "").strip()
            for row in rows)
    # And the two snapshot builders compute it rather than hard-coding "".
    for path in (
        "app/services/assessment_release_snapshot.py",
        "app/services/assessment_release_service.py",
    ):
        source = Path(path).read_text(encoding="utf-8")
        assert '"topic_concept_labels": ""' not in source, path
        assert "identity.titled(" in source, path


def test_the_writer_still_composes_the_identity_pair():
    """C9/T14 — "one renderer" did not strip the OTHER one.

    ``identity.titled`` and ``identity.topic_concept_roster`` are SHARED
    with ``writer._front_bands`` and ``writer._concept_field_value``, not
    moved out of them. Stripping them from the writer would delete the
    identity pair from Outputs 01/03 and from the append-only bulk-import
    database of record.
    """

    source = Path("app/bulk_import/writer.py").read_text(encoding="utf-8")
    assert "identity.titled(" in source
    assert "identity.topic_concept_roster(" in source
    assert "def write_concepts_workbook(" in source
    assert hasattr(bi_writer, "write_concepts_workbook")


# --------------------------------------------------------------------------- #
# The Rule-1 purge that is atomic with this slice
# --------------------------------------------------------------------------- #

def test_no_volume_derived_chapter_duration_or_description_survives():
    """Un-blanking ``chapter_duration`` before this went would ship a
    volume-derived number to every school (CLAUDE.md:17-18)."""

    code = "\n".join(
        line.split("#")[0]
        for line in Path("app/services/build_concepts.py")
        .read_text(encoding="utf-8").splitlines()
    )
    assert "max(40, n_concepts * 12)" not in code
    assert "This chapter develops" not in code
    assert "concept(s) across" not in code
    assert "n_concepts" not in code


# --------------------------------------------------------------------------- #
# The read-back gate that the manifest union closes (T3.6)
# --------------------------------------------------------------------------- #

def test_the_identity_trio_never_reaches_a_workbook_record():
    entry = {
        "chapter": {"chapter_title": "C"},
        # ``topic_machine_id`` joined the strip list in the B4 repair round;
        # the fixture carries it so a leak of the key into the record fails
        # here (the audit measured the earlier fixture could not detect it).
        "topic": {"topic_title": "T", "topic_machine_id": "X_PL_T01"},
        "concept": {
            "concept_key": "C_A",
            "concept_machine_id": "M1",
            "release_row_identity": "R1",
            "concept_title": "Title",
        },
    }
    record = mp._bands_record(entry)
    assert set(record) == {"chapter_title", "topic_title", "concept_title"}


def test_the_master_renderer_records_a_group_it_cannot_place():
    """R4: the renderer stopped raising; it did not start losing."""

    snapshot = copy.deepcopy(_snapshot())
    snapshot["groups"][0]["concept_key"] = "C_MISSING"
    master, issues = mp.render_master_file(snapshot)
    assert master[:2] == b"PK"
    assert [item["code"] for item in issues["group_defects"]] == [
        rel.GROUP_CONCEPT_HOME_UNKNOWN]


# --------------------------------------------------------------------------- #
# T3.3 / T3.3b — the Pre lane's resolved ``related_concepts``
# --------------------------------------------------------------------------- #

def test_pre_related_concepts_carries_the_resolved_post_machine_ids(db):
    from app.services import build_concepts_release as release
    from app.services import build_concepts_release_files as release_files
    from tests.test_assessment_release_run import _chapter_with_concepts
    from tests.test_pre_release_lane_wiring import _both_lanes_job

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    post = release.release_payload(job, lane=release.LANE_POST)
    _c, post_concepts, _r, _d = release_files.transient_release_hierarchy(
        db, job, payload=post)
    expected = post_concepts[0].machine_id
    assert expected

    pre = release.release_payload(job, lane=release.LANE_PRE)
    row = pre["records"][0]
    assert row[release.PRE_ROW_RELATED_CONCEPTS_FIELD] == expected
    assert row[release.PRE_ROW_RELATED_UNRESOLVED_FIELD] == []

    # And it reaches the rendered Pre Concept File.
    _c, pre_concepts, _r, _d = release_files.transient_release_hierarchy(
        db, job, payload=pre)
    assert pre_concepts[0].related_concepts == expected
    # No opaque phase-3 token survives anywhere in the rendered bytes.
    data = release_files.build_release_bulk_import_workbook(
        db, job, lane=release.LANE_PRE)
    book = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    text = "\n".join(
        str(cell)
        for sheet in book.worksheets
        for row_values in sheet.iter_rows(values_only=True)
        for cell in row_values
        if cell is not None
    )
    book.close()
    for token in ("CONCEPT-", "PRC-", "PRT-"):
        assert token not in text


def test_post_related_concepts_is_blank_until_a_relations_pass_exists(db):
    """Pinned so it FAILS the day somebody authors concept↔concept links."""

    from app.services import build_concepts_release as release
    from app.services import build_concepts_release_files as release_files
    from tests.test_assessment_release_run import _chapter_with_concepts
    from tests.test_pre_release_lane_wiring import _both_lanes_job

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    post = release.release_payload(job, lane=release.LANE_POST)
    _c, concepts, _r, _d = release_files.transient_release_hierarchy(
        db, job, payload=post)
    assert all(not (c.related_concepts or "") for c in concepts)


def test_an_unresolvable_needed_for_link_is_a_review_flag_not_a_blank(db):
    from app.services import build_concepts_release as release

    resolved, unresolved = release._resolve_needed_for(
        {"_aegis_needed_for": [
            {"post_concept_id": "CONCEPT-0009",
             "post_concept_title": "Nothing staged this"},
        ]},
        {"CONCEPT-0001": ("10CBMA_Ch_PL_T01_C01", "What makes a solid")},
    )
    assert resolved == []
    assert unresolved and unresolved[0]["post_concept_id"] == "CONCEPT-0009"
    assert "no Post concept" in unresolved[0]["reason"]

    # A position that resolves to a DIFFERENT concept is also a flag, never
    # a silent stamp: the id is the join, the title only CONFIRMS it.
    resolved, unresolved = release._resolve_needed_for(
        {"_aegis_needed_for": [
            {"post_concept_id": "CONCEPT-0001",
             "post_concept_title": "A different concept entirely"},
        ]},
        {"CONCEPT-0001": ("10CBMA_Ch_PL_T01_C01", "What makes a solid")},
    )
    assert resolved == []
    assert "staged at this position" in unresolved[0]["reason"]


def test_the_publication_lifts_the_resolved_marker_before_the_strip():
    """T3.3b — the column must not empty when the reviewer clicks Upload."""

    from app.services import build_concepts_release as release

    row = {
        "concept_title": "Counting to ten",
        release.PRE_ROW_RELATED_CONCEPTS_FIELD: "10CBMA_Ch_PL_T01_C01",
    }
    lifted = release._lift_resolved_related_concepts(row)
    assert lifted["related_concepts"] == "10CBMA_Ch_PL_T01_C01"
    stripped = release._strip_release_fields(lifted)
    assert stripped["related_concepts"] == "10CBMA_Ch_PL_T01_C01"
    assert release.PRE_ROW_RELATED_CONCEPTS_FIELD not in stripped
    # …and without the lift it would be gone, which is the whole point.
    assert "related_concepts" not in release._strip_release_fields(row)


def test_add_concept_persists_related_concepts_and_digicards():
    source = Path("app/services/build_concepts.py").read_text(encoding="utf-8")
    assert "related_concepts=rec.get(\"related_concepts\", \"\")" in source
    assert "digicards=rec.get(\"digicards\", \"\")" in source
    merge = Path(
        "app/services/build_concepts_release_publication.py"
    ).read_text(encoding="utf-8")
    assert 'existing.related_concepts = rec.get("related_concepts", "")' in (
        merge)
    assert 'existing.digicards = rec.get("digicards", "")' in merge


# --------------------------------------------------------------------------- #
# The artifact pins (OD4/D9-Q22, OD6/T17/B1)
# --------------------------------------------------------------------------- #

def test_output_03_carries_the_machine_id_pair_in_every_title_cell(db):
    """The POST Concept File — renamed from ``…output_01…`` by OD4/D9-Q22.

    The artifact and the assertion are unchanged; only the number moves.
    """

    import re

    from app.services import build_concepts_release as release
    from app.services import build_concepts_release_files as release_files
    from tests.test_assessment_release_run import _chapter_with_concepts
    from tests.test_pre_release_lane_wiring import _both_lanes_job

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    data = release_files.build_release_bulk_import_workbook(
        db, job, lane=release.LANE_POST)
    book = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    sheet = book.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(h or "") for h in rows[1]]
    index = header.index("concept_title")
    titles = [
        str(row[index] or "") for row in rows[2:]
        if any(cell not in (None, "") for cell in row)
    ]
    book.close()
    assert titles
    for title in titles:
        assert re.search(r"\([^()]+\)\s*$", title), title


def test_the_sealed_concept_projection_keeps_the_same_shape():
    """OD6/T17/B1 — ``aegis_concepts.xlsx`` is NOT one of the four outputs.

    It is the sealed internal Concept-file projection whose sha256
    ``assessment_release_service`` seals at publication and re-verifies at
    upload. A projection whose shape drifts from the artifact it projects
    is a hash that proves nothing, so its sheet set is pinned against the
    Master's here.
    """

    snapshot = copy.deepcopy(_snapshot())
    snapshot["topics"][0]["topic_concept_labels"] = "A (T01_C01)"
    concepts = mp.render_concept_file(snapshot)
    master, _ = mp.render_master_file(snapshot)
    parsed = mp.parse_workbook(concepts)
    assert parsed["sheet_order"] == mp.parse_workbook(master)["sheet_order"]
    fields = mp.FIELDS["Objective"]
    group_band_start = fields.index("group_name")
    for row in parsed["sheets"]["Objective"]["rows"]:
        assert str(row.get("topic_concept_labels") or "").strip()
        assert all(
            not str(row.get(field) or "").strip()
            for field in fields[group_band_start:]
        )


def test_the_reference_profile_cells_payload_is_byte_identical_before_and_after_threading():
    """T12/M5's cost check: zero decisions re-key.

    ``kernel.decision_key`` is computed over the request payload, and the
    profile evidence is part of it. For the DEFAULT profile the derived
    tuple is identical to the constant it replaced, so the payload stays
    byte-identical and no recorded decision moves.
    """

    from app.services import assessment_cells as cells

    assert cells._allowed_sheet_kinds(
        assessment_profile.DEFAULT_PROFILE) == ("objective", "descriptive")
    assert cells._profile_payload(assessment_profile.DEFAULT_PROFILE) == {
        "name": "reference-1",
        "allowed_sheet_kinds": ["objective", "descriptive"],
        "appears_in": "Pre/Post-Worksheet/Test",
    }


def test_the_group_home_verdict_is_the_same_four_inputs_the_renderer_reads():
    """``unresolved_question_homes`` is PURE over the snapshot."""

    snapshot = copy.deepcopy(_snapshot())
    before = json.dumps(snapshot, sort_keys=True, default=str)
    rel.unresolved_question_homes(snapshot)
    assert json.dumps(snapshot, sort_keys=True, default=str) == before
    assert rel.unresolved_question_homes(_snapshot()) == []


def test_the_master_refiner_readback_accepts_a_real_questionless_group(db=None):
    """The SECOND read-back, exercised rather than grepped (T16).

    ``assessment_master_refiner``'s ``_validation_state`` rebuilds group rows
    from ``group_provenance`` and errors with "refiner-readback: group … has
    no rendered Master row" for every payload group with no rendered row. A
    real payload group whose concept ends with ZERO placed question rows now
    legitimately has none — the concept gets one tail row instead — so
    leaving this read-back unscoped turns every affected concept into an
    error. It takes the SAME scoping as ``validate_master_file``, and it is
    NOT "a group with no questions", which would stop catching a genuinely
    missing catalogue row on a concept that does have questions.
    """

    from tests.test_assessment_master_refiner import (
        _payload, CONCEPT_KEY, MACHINE_ID,
    )

    payload = copy.deepcopy(_payload())
    topic = payload["concept_snapshot"]["topics"][0]
    quiet_key = "release:test:0002"
    quiet_machine = "06MSMA_T01_Quiet"
    topic["concepts"].append({
        **copy.deepcopy(topic["concepts"][0]),
        "concept_key": quiet_key,
        "concept_machine_id": quiet_machine,
        "concept_title": f"Quiet concept ({quiet_machine})",
        "concept_display_name": "Quiet concept",
    })
    payload["concept_snapshot"]["concept_provenance"].append({
        "concept_key": quiet_key,
        "release_row_identity": {"position": 2},
    })
    template = copy.deepcopy(payload["groups"][0])
    payload["groups"].append({
        **template,
        "group_key": f"({quiet_machine}) BG01",
        "concept_key": quiet_key,
        "concept_id": quiet_key,
        "group_type": "Basic",
        "group_sequence": 1,
        "group_name": "Quiet concept — Basic",
        "group_display_name": "Quiet concept — Basic",
        "member_candidate_ids": [],
    })

    state = refiner._validation_state(
        payload, assessment_profile.DEFAULT_PROFILE)
    readback = [
        error for error in state["errors"] if "refiner-readback" in error]
    assert readback == [], readback
    assert state["errors"] == [], state["errors"]
    # …and the concept still reaches the Master, as its ONE tail row with
    # every Group-band cell empty.
    from app.services import assessment_release_service as release_service

    snapshot = release_service.snapshot_from_staged_release(payload)
    master, issues = mp.render_master_file(snapshot)
    rows = _objective_rows(master)
    tail = [
        row for row in rows
        if str(row.get("concept_title") or "").startswith("Quiet concept")]
    assert len(tail) == 1
    assert tail[0] is rows[-1]
    assert not str(tail[0].get("group_name") or "").strip()
    assert [c["concept_key"] for c in issues["questionless_concepts"]] == [
        quiet_key]
    assert CONCEPT_KEY and MACHINE_ID  # the fixture's own concept is intact


# --------------------------------------------------------------------------- #
# ONE DECISION, ONE OWNER — the repairs the three S8 audits found
# --------------------------------------------------------------------------- #

def test_the_staged_verdict_and_the_renderer_take_the_same_placement_decision(
    monkeypatch,
):
    """The R4 hole one layer up: two sources for "can this row render".

    ``unresolved_question_homes`` compared against ``RENDERABLE_SHEET_KINDS``
    while the renderer compared against ``sheet_for_kind()`` — that tuple
    INTERSECTED with the layout's sheet-name map. [measured] under a layout
    carrying no Descriptive sheet the renderer dropped every descriptive
    candidate into the unplaced ledger while the staged verdict returned
    nothing at all: readiness read ``ready`` and the database write would
    have proceeded with those questions in no data row.

    They are now ONE function. This narrows the renderable map and asserts
    the two still agree — the divergence cannot come back silently.
    """

    monkeypatch.setattr(
        mp, "sheet_for_kind", lambda: {"objective": "Objective"})
    snapshot = copy.deepcopy(_snapshot())
    _master, issues = mp.render_master_file(snapshot)
    dropped = {item["question_label"] for item in issues["unplaced"]}
    staged = {
        finding["question_label"]
        for finding in rel.unresolved_question_homes(snapshot)
        if finding["code"] in {
            rel.UNRESOLVED_QUESTION_HOME,
            rel.GROUP_HOME_DISAGREEMENT,
            rel.SHEET_KIND_NOT_RENDERABLE,
        }
    }
    assert dropped, "the narrowed layout dropped nothing to compare"
    assert dropped == staged


def test_one_function_decides_whether_a_candidate_reaches_a_data_row():
    """T1: the renderer does not re-discover the three branches."""

    renderer = inspect.getsource(mp.render_master_file)
    assert "rel.candidate_home_finding(" in renderer
    refiner_source = inspect.getsource(refiner._validation_state)
    assert "rel.candidate_home_finding(" in refiner_source
    # And the staged pass calls the same function on the same inputs.
    assert "candidate_home_finding(" in inspect.getsource(
        rel.unresolved_question_homes)


def test_a_group_identity_defect_is_not_recorded_as_a_question_home_defect():
    """The renderer's ledger used a candidate-level code on a group."""

    blank = copy.deepcopy(_snapshot())
    blank["groups"][0]["group_key"] = ""
    assert [item["code"] for item in mp.render_master_file(blank)[1][
        "group_defects"]] == [rel.GROUP_KEY_BLANK]

    duplicate = copy.deepcopy(_snapshot())
    duplicate["groups"].append(copy.deepcopy(duplicate["groups"][0]))
    assert [item["code"] for item in mp.render_master_file(duplicate)[1][
        "group_defects"]] == [rel.GROUP_KEY_DUPLICATE]

    for code in (rel.GROUP_KEY_BLANK, rel.GROUP_KEY_DUPLICATE):
        assert code not in rel.QUESTION_HOME_CODES


def test_a_truncated_row_is_recorded_in_the_manifest_not_only_at_staging():
    """R4: the row ships truncated, so the loss must not be silent.

    ``render_shape_overflow`` names it on the release and refuses the
    database write — but the issues manifest that lands on disk as
    ``manifest.json`` carried nothing, so a reviewer reading only the
    manifest saw a complete-looking six-option row.
    """

    snapshot = copy.deepcopy(_snapshot())
    candidate = snapshot["candidates"][0]
    candidate["answers"] = [
        copy.deepcopy(candidate["answers"][0])
        for _ in range(mp.MAX_OBJECTIVE_OPTIONS + 4)
    ]
    _master, issues = mp.render_master_file(snapshot)
    assert issues["truncated_rows"] == [{
        "candidate_id": candidate["candidate_id"],
        "question_label": candidate["question_label"],
        "field": "answers",
        "cap": mp.MAX_OBJECTIVE_OPTIONS,
        "actual": mp.MAX_OBJECTIVE_OPTIONS + 4,
    }]
    assert rel.RENDER_SHAPE_OVERFLOW in {
        finding["code"]
        for finding in rel.unresolved_question_homes(snapshot)}
    # A clean snapshot records nothing.
    assert mp.render_master_file(_snapshot())[1]["truncated_rows"] == []


def test_the_visible_group_name_has_exactly_one_owner():
    """A second copy of the composition would block every release.

    ``unresolved_question_homes`` spelled ``f"{concept_name} — {tier}"``
    inline while ``assessment_grouping.friendly_group_name`` composed the
    value actually written. A change to the separator in one of them flags
    ``group_visible_name_mismatch`` on every group of every release.
    """

    from app.services import assessment_grouping as grouping

    source = inspect.getsource(rel.unresolved_question_homes)
    assert "friendly_group_name(" in source
    code = "\n".join(
        line.split("#")[0] for line in source.splitlines())
    assert "— {tier}" not in code

    snapshot = copy.deepcopy(_snapshot())
    group = snapshot["groups"][0]
    concept = snapshot["topics"][0]["concepts"][0]
    expected = grouping.friendly_group_name(
        concept["concept_display_name"], group["group_type"])
    assert group["group_name"] == expected
    assert rel.unresolved_question_homes(snapshot) == []
    group["group_name"] = expected + "!"
    assert rel.GROUP_VISIBLE_NAME_MISMATCH in {
        finding["code"]
        for finding in rel.unresolved_question_homes(snapshot)}


def test_a_run_context_carrying_a_profile_dict_still_ships_four_outputs():
    """``str()`` on a Mapping turned the profile into ``"{'name': …}"``.

    ``get_profile`` then raised ``KeyError: unknown assessment profile``
    out of ``freeze_payload`` — no release row and no workbook at all, on
    a slice whose whole theme is that nothing costs the four outputs.
    """

    from app.services import assessment_release_service as svc

    as_dict = svc._run_profile(
        {"assessment_profile": dict(assessment_profile.DEFAULT_PROFILE)})
    assert isinstance(as_dict, dict)
    assert assessment_profile.sheet_kinds(as_dict) == (
        "objective", "descriptive")
    assert svc._run_profile({"profile": "reference-1"}) == "reference-1"
    assert svc._run_profile({"profile": ""}) is None
    assert svc._run_profile(None) is None
    # A PARTIAL Mapping resolves too, through the per-key fallback.
    snapshot = _snapshot()
    output = mp.build_dual_output(snapshot, {"name": "partial"})
    assert output["master_xlsx"][:2] == b"PK"
    assert output["manifest"]["profile"] == "partial"


def test_the_parent_machinery_is_gone_from_the_bulk_import_lane():
    """The spec's "machinery is gone" pin, measured rather than assumed."""

    for path in Path("app/bulk_import").rglob("*.py"):
        assert '"parent: "' not in path.read_text(encoding="utf-8"), path
