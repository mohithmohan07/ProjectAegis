"""S7 — the layout move, and the migration that carries older files across.

Every test here fails at HEAD~1, and each fails for its own reason:

* the committed fixture is on the canonical layout until the writer moves;
* ``writer._new_workbook`` emits four sheets in the wrong order until OD6's
  sheet half lands;
* ``migrate_workbook_layout`` does not exist until this slice;
* the row-width repair is silent until T3.8 makes it a recorded decision.

The migration FIXTURES are workbooks the real writer produced — its own
``_question_to_row`` / ``_concept_to_row``, emitting through the frozen
``canonical-current`` registry entry — with a non-empty ``source_book``, so
``concept_source`` is non-blank on Descriptive exactly as a deployed volume's
``bulk_import_output.xlsx`` has it. Never ``bulk_import_database.xlsx``: per
OWNER RULING OD2 that file is the CI-regenerated database fixture, nothing
appends to it, it is not the file that migrates, and its rows are
``scripts/generate_dummy_data.py``'s rather than the pipeline's.
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from app import bulk_import as bi
from app import config
from app import models
from app.bulk_import import layouts, writer

CANONICAL = "canonical-current"


# --------------------------------------------------------------------------- #
# Fixtures built through the REAL writer, at the canonical layout
# --------------------------------------------------------------------------- #

def _graph(db, suffix: str):
    chapter = models.Chapter(
        chapter_code=f"10CBSC_Migrate_{suffix}",
        board="CBSE", grade="10", subject="Science",
        unit="Layout Migration",
        chapter_title=f"Layout Migration {suffix}",
        chapter_display_name=f"Layout Migration {suffix}",
    )
    topic = models.Topic(
        chapter=chapter, topic_title="Migrating Topic",
        topic_display_name="Migrating Topic", pre_post_learning="Post",
    )
    concept = models.Concept(
        topic=topic,
        concept_title=f"Migrating Concept {suffix}",
        concept_display_name=f"Migrating Concept {suffix}",
        concept_details="A concept written by a real run.",
        keywords="migration, layout",
        related_concepts="A genuine relation",
        # A real run sets ``concept.sources`` from the job's source_book and
        # ``writer._concept_field_value`` returns it unconditionally, so a
        # deployed workbook HAS a non-blank concept_source on every sheet.
        sources="NCERT",
        parent_concept="",
    )
    group = models.Group(
        concept=concept, group_name=f"{suffix} Basic",
        group_display_name=f"{suffix} Basic", group_type="Basic",
        group_status="Active", group_description="",
    )
    question = models.Question(
        group=group, sheet_kind="descriptive",
        question_label=f"10CBSC_Mig_PL_T01_{suffix} Q01",
        question="Explain the migrating method.",
        question_text="Explain the migrating method.",
        question_source="NCERT", marks="3", cognitive_skills="Understand",
        level_of_difficulty="Moderate", question_appears_in=bi.APPEARS_IN_ALL,
        answers=[{"answer_type": "Phrases", "answer_weightage": "1",
                  "answer_content": "Because the layout moved."}],
        sub_questions=[],
    )
    db.add_all([chapter, topic, concept, group, question])
    db.commit()
    return concept, topic, group, question


def _real_writer_canonical_workbook(db, path: Path, suffix: str):
    """A workbook the REAL writer produced, on the ``canonical-current`` layout.

    ``_question_to_row`` / ``_concept_to_row`` compose through whichever
    registered layout they are handed, so this is the writer's own output —
    not a hand-transcribed fixture — at the layout a deployed volume carries.
    """
    concept, topic, group, question = _graph(db, suffix)
    entry = layouts.layout(CANONICAL)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_layout in entry.sheets.values():
        worksheet = workbook.create_sheet(sheet_layout.sheet_name)
        for band in sheet_layout.bands:
            worksheet.cell(row=1, column=band.start, value=band.label)
        for column, name in enumerate(sheet_layout.fields, start=1):
            worksheet.cell(row=2, column=column, value=name)
    doc = workbook.create_sheet("Doc Link <> Each fields ")
    doc.cell(row=1, column=1, value="Screenshot Doc")

    objective = entry.sheet("objective")
    worksheet = workbook[objective.sheet_name]
    for column, value in enumerate(
        writer._concept_to_row(
            concept, "objective", topic, sheet_layout=objective),
        start=1,
    ):
        worksheet.cell(row=3, column=column, value=value)

    descriptive = entry.sheet("descriptive")
    worksheet = workbook[descriptive.sheet_name]
    for column, value in enumerate(
        writer._question_to_row(
            question, "descriptive", group, sheet_layout=descriptive),
        start=1,
    ):
        worksheet.cell(row=3, column=column, value=value)

    workbook.save(path)
    return concept, question


# --------------------------------------------------------------------------- #
# The migration
# --------------------------------------------------------------------------- #

def test_appending_to_an_older_layout_migrates_once_and_loses_nothing(
    db, tmp_path,
):
    path = tmp_path / "output.xlsx"
    concept, question = _real_writer_canonical_workbook(db, path, "Once")
    labels_before = _all_question_labels(path)
    assert labels_before

    first = writer.append_concepts(db, path, [concept.id])
    receipt = first["layout_migration"]
    assert receipt["source_layout_id"] == CANONICAL
    assert receipt["target_layout_id"] == layouts.REFERENCE_LAYOUT_ID
    assert receipt["sha256_before"] != receipt["sha256_after"]

    # The pre-migration file is retained beside the target, per Q16.
    sibling = Path(receipt["sibling_path"])
    assert sibling.name == f"output.pre-{CANONICAL}.xlsx"
    assert sibling.exists()

    # The workbook is now on the target layout, sheets and all.
    workbook = openpyxl.load_workbook(path)
    assert workbook.sheetnames == bi.SHEET_ORDER
    for kind, sheet_name in bi.SHEET_BY_KIND.items():
        header = [c.value for c in workbook[sheet_name][2]]
        assert header == bi.FIELDS_BY_KIND[kind], kind

    # Every question label that was in the file is still in the file.
    assert labels_before <= _all_question_labels(path)
    # ...and every row survived, counted per sheet by the receipt itself.
    assert receipt["rows_by_sheet_after"] == receipt["rows_by_sheet_before"]

    second = writer.append_concepts(db, path, [concept.id])
    assert "layout_migration" not in second


def test_a_migration_that_would_drop_a_non_blank_value_is_recorded_and_the_run_completes(
    db, tmp_path,
):
    """The canonical Descriptive ``concept_source`` has no target column.

    Cluster D specified a ``WorkbookLayoutMigrationRefused`` raise here. That
    is the halt Q13 retired: ``append_concepts`` is reached from
    ``build_concepts`` inside a run that has already spent its entire model
    budget, so a refusal destroys the whole run's work to protect one cell.
    """
    path = tmp_path / "output.xlsx"
    concept, _question = _real_writer_canonical_workbook(db, path, "Drop")
    canonical_descriptive = layouts.layout(CANONICAL).sheet("descriptive")
    assert canonical_descriptive.column("concept", "concept_source") is not None
    assert "concept_source" not in bi.concept_fields("descriptive")

    # NO RAISE.
    result = writer.append_concepts(db, path, [concept.id])

    dropped = [
        entry for entry in result["layout_migration"]["unmappable"]
        if entry["field"] == "concept_source"
    ]
    assert dropped, result["layout_migration"]["unmappable"]
    assert dropped[0]["value"] == "NCERT"
    assert dropped[0]["kind"] == "descriptive"

    # A release issue names it...
    codes = {issue["code"] for issue in result["issues"]}
    assert layouts.MIGRATION_UNMAPPABLE_VALUE in codes
    named = next(
        issue for issue in result["issues"]
        if issue["code"] == layouts.MIGRATION_UNMAPPABLE_VALUE
    )
    assert "concept_source" in named["detail"]
    assert named["sheet"] == "Descriptive"

    # ...and a recorded, flagged, content-addressed Fixer decision carries it.
    decision = next(
        entry for entry in result["fixer_decisions"]
        if entry["code"] == layouts.MIGRATION_UNMAPPABLE_VALUE
    )
    assert decision["review_flag"] is True
    assert len(decision["decision_sha256"]) == 64
    assert decision["policy_version"]

    # The run completed: append_concepts RETURNED, and the workbook it
    # returned is on the target layout with its rows intact.
    assert result["layout_migration"]["rows_by_sheet_after"] == (
        result["layout_migration"]["rows_by_sheet_before"])
    assert openpyxl.load_workbook(path).sheetnames == bi.SHEET_ORDER
    # And the value is still readable in the retained sibling.
    assert Path(result["layout_migration"]["sibling_path"]).exists()


def test_the_duplicate_question_label_columns_collapse_by_recorded_alias(
    db, tmp_path,
):
    """[measured] canonical Descriptive carries ``question_label`` TWICE.

    Index 27 in the Group band and index 30 in the Question band, both
    non-blank with identical values on all 56 rows of this repo's own
    artifact. The target carries it once. The collapse is provably safe, and
    it is an EXPLICIT registry entry rather than an inference.
    """
    canonical = layouts.layout(CANONICAL).sheet("descriptive")
    assert canonical.column("group", "question_label") == 27
    assert canonical.column("question", "question_label") == 30
    target = layouts.sheet(layouts.REFERENCE_LAYOUT_ID, "descriptive")
    assert target.column("group", "question_label") is None

    alias = next(
        entry for entry in layouts.aliases_for(
            CANONICAL, layouts.REFERENCE_LAYOUT_ID, "descriptive")
        if entry.source == ("group", "question_label")
    )
    assert alias.target == ("question", "question_label")
    assert alias.reason
    alias = (alias,)

    path = tmp_path / "output.xlsx"
    concept, question = _real_writer_canonical_workbook(db, path, "Alias")
    result = writer.append_concepts(db, path, [concept.id])

    receipt = result["layout_migration"]
    assert alias[0].name in receipt["alias_entries_applied"]
    # It is NOT reported unmappable, and the label survived.
    assert not [
        entry for entry in receipt["unmappable"]
        if entry["field"] == "question_label"
    ]
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[bi.SHEET_DESCRIPTIVE]
    header = [cell.value for cell in sheet[2]]
    column = header.index("question_label") + 1
    assert sheet.cell(row=3, column=column).value == question.question_label


def _canonical_descriptive_row(path: Path, **cells) -> None:
    """A canonical workbook with named Descriptive cells set by hand.

    The alias collapse is only interesting when the two ``question_label``
    copies DISAGREE, and the writer cannot produce that: it writes
    ``q.question_label`` into both. A deployed ``BULK_IMPORT_OUTPUT`` is a file
    a human can open in Excel, so divergence is unlikely, not impossible —
    and "unlikely" is exactly the case a receipt has to be honest about.
    """
    entry = layouts.layout(CANONICAL)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_layout in entry.sheets.values():
        worksheet = workbook.create_sheet(sheet_layout.sheet_name)
        for band in sheet_layout.bands:
            worksheet.cell(row=1, column=band.start, value=band.label)
        for column, name in enumerate(sheet_layout.fields, start=1):
            worksheet.cell(row=2, column=column, value=name)
    descriptive = entry.sheet("descriptive")
    worksheet = workbook[descriptive.sheet_name]
    worksheet.cell(
        row=3, column=descriptive.column("chapter", "chapter_title") + 1,
        value="A Chapter",
    )
    for key, value in cells.items():
        band, field = key.split("__", 1)
        worksheet.cell(
            row=3, column=descriptive.column(band, field) + 1, value=value)
    workbook.save(path)


def test_a_divergent_duplicate_question_label_is_recorded_not_overwritten(
    tmp_path,
):
    """The alias collapse must never CHOOSE in silence.

    The guard for this existed and was unreachable. Source bands are walked in
    layout order, so the Group-band copy was written into a still-blank target
    cell and the Question band then overwrote it unconditionally: the
    divergent value left with no ``unmappable`` entry, no Fixer decision and
    no release issue, while ``alias_entries_applied`` still asserted a
    lossless collapse. A ``question_label`` is the learner-facing identity
    CLAUDE.md names, so the receipt affirming a collapse it had not verified is
    the failure, not merely the silence.
    """
    path = tmp_path / "diverged.xlsx"
    _canonical_descriptive_row(
        path,
        group__question_label="GROUP-COPY-LABEL",
        question__question_label="QUESTION-COPY-LABEL",
    )
    receipt = layouts.migrate_workbook_layout(
        path, target_layout_id=layouts.REFERENCE_LAYOUT_ID)

    conflicts = [
        entry for entry in receipt.unmappable
        if entry["field"] == "question_label"
    ]
    assert len(conflicts) == 1, receipt.unmappable
    assert conflicts[0]["value"] == "GROUP-COPY-LABEL"
    assert conflicts[0]["conflicts_with"] == "QUESTION-COPY-LABEL"
    assert conflicts[0]["band"] == "group"
    assert conflicts[0]["alias"]
    # ...and the receipt does NOT claim the alias collapsed this row cleanly.
    assert receipt.alias_entries_applied == []

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[bi.SHEET_DESCRIPTIVE]
    header = [cell.value for cell in sheet[2]]
    column = header.index("question_label") + 1
    assert sheet.cell(row=3, column=column).value == "QUESTION-COPY-LABEL"


def test_a_blank_target_band_copy_does_not_erase_the_aliased_value(tmp_path):
    """The same collapse, from the other side.

    A whitespace-only cell in the target's own band used to win purely because
    its band was walked last, silently erasing a real aliased label. Blank
    never displaces non-blank, and this needs no record because nothing with
    content was dropped.
    """
    path = tmp_path / "blanked.xlsx"
    _canonical_descriptive_row(
        path,
        group__question_label="THE-ONLY-REAL-LABEL",
        question__question_label="   ",
    )
    receipt = layouts.migrate_workbook_layout(
        path, target_layout_id=layouts.REFERENCE_LAYOUT_ID)

    workbook = openpyxl.load_workbook(path)
    sheet = workbook[bi.SHEET_DESCRIPTIVE]
    header = [cell.value for cell in sheet[2]]
    column = header.index("question_label") + 1
    assert sheet.cell(row=3, column=column).value == "THE-ONLY-REAL-LABEL"
    assert not [
        entry for entry in receipt.unmappable
        if entry["field"] == "question_label"
    ]
    assert receipt.alias_entries_applied


def _canonical_objective_row(path: Path, **cells) -> None:
    entry = layouts.layout(CANONICAL)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_layout in entry.sheets.values():
        worksheet = workbook.create_sheet(sheet_layout.sheet_name)
        for band in sheet_layout.bands:
            worksheet.cell(row=1, column=band.start, value=band.label)
        for column, name in enumerate(sheet_layout.fields, start=1):
            worksheet.cell(row=2, column=column, value=name)
    objective = entry.sheet("objective")
    worksheet = workbook[objective.sheet_name]
    for key, value in cells.items():
        row, band, field = key.split("__", 2)
        worksheet.cell(
            row=int(row.lstrip("r")),
            column=objective.column(band, field) + 1,
            value=value,
        )
    workbook.save(path)


def test_the_row_counts_are_evidence_and_not_a_tautology(tmp_path):
    """``rows_after`` used to be incremented in the same breath as before.

    No branch stood between them, so ``rows_by_sheet_after ==
    rows_by_sheet_before`` was true by construction and two regressions
    asserted a field that could not be false. A source row whose every
    non-blank value is homeless in the target lands nothing — and the receipt
    now says so rather than writing an empty row and calling it survival.
    Every value it dropped is still an ``unmappable`` record and still in the
    sibling.
    """
    path = tmp_path / "homeless.xlsx"
    _canonical_objective_row(
        path,
        r3__concept__parent_concept="A REAL PARENT",
        r4__chapter__chapter_title="Real Chapter",
    )
    receipt = layouts.migrate_workbook_layout(
        path, target_layout_id=layouts.REFERENCE_LAYOUT_ID)

    assert receipt.rows_by_sheet_before[bi.SHEET_OBJECTIVE] == 2
    assert receipt.rows_by_sheet_after[bi.SHEET_OBJECTIVE] == 1
    assert [
        entry["value"] for entry in receipt.unmappable
        if entry["field"] == "parent_concept"
    ] == ["A REAL PARENT"]

    sheet = openpyxl.load_workbook(path)[bi.SHEET_OBJECTIVE]
    rows = [
        row for row in sheet.iter_rows(min_row=3, values_only=True)
        if row and any(str(value or "").strip() for value in row)
    ]
    # No empty row was emitted where the homeless one used to be.
    assert len(rows) == 1
    assert Path(receipt.sibling_path).exists()


def test_a_numeric_zero_is_a_value_and_not_a_blank(tmp_path):
    """``str(value or "").strip()`` calls ``0`` blank. A mark of 0 is not."""
    path = tmp_path / "zero.xlsx"
    _canonical_objective_row(
        path,
        r3__chapter__chapter_title="Ch",
        r3__chapter__chapter_duration=0,
    )
    layouts.migrate_workbook_layout(
        path, target_layout_id=layouts.REFERENCE_LAYOUT_ID)
    target = layouts.sheet(layouts.REFERENCE_LAYOUT_ID, "objective")
    sheet = openpyxl.load_workbook(path)[bi.SHEET_OBJECTIVE]
    column = target.column("chapter", "chapter_duration") + 1
    assert sheet.cell(row=3, column=column).value == 0


def test_a_dropped_sheet_that_carried_data_becomes_a_recorded_decision(
    db, tmp_path,
):
    """``dropped_sheets`` recorded a NAME; a reviewer's link is a value.

    A sheet the source layout declares ignored is in neither layout, so the
    migrated workbook is rebuilt without it — correctly, by OD6 the authority
    has three sheets. But a reviewer who pasted a link into that tab put
    content there, and a receipt carrying only the sheet's name cannot say
    whether anything was in it. One decision per sheet, naming the magnitude
    and where the values are still readable.
    """
    path = tmp_path / "output.xlsx"
    concept, _question = _real_writer_canonical_workbook(db, path, "Dropped")
    workbook = openpyxl.load_workbook(path)
    workbook["Doc Link <> Each fields "].cell(
        row=3, column=1, value="https://example.com/a-reviewer-added-this")
    workbook.save(path)

    result = writer.append_concepts(db, path, [concept.id])
    receipt = result["layout_migration"]
    assert receipt["dropped_sheets"] == ["Doc Link <> Each fields "]
    assert receipt["dropped_sheet_cells"]["Doc Link <> Each fields "] == 1

    issue = next(
        entry for entry in result["issues"]
        if entry["code"] == layouts.MIGRATION_DROPPED_SHEET
    )
    assert issue["sheet"] == "Doc Link <> Each fields "
    assert issue["cells"] == 1
    assert receipt["sibling_path"] in issue["detail"]
    decision = next(
        entry for entry in result["fixer_decisions"]
        if entry["code"] == layouts.MIGRATION_DROPPED_SHEET
    )
    assert decision["review_flag"] is True
    assert len(decision["decision_sha256"]) == 64
    # The value itself is still readable where the receipt says it is.
    sibling = openpyxl.load_workbook(Path(receipt["sibling_path"]))
    assert sibling["Doc Link <> Each fields "].cell(row=3, column=1).value == (
        "https://example.com/a-reviewer-added-this")


def test_an_empty_dropped_sheet_is_not_escalated(db, tmp_path):
    """The writer's own banner row is not a reviewer's content."""
    path = tmp_path / "output.xlsx"
    concept, _question = _real_writer_canonical_workbook(db, path, "Quiet")
    result = writer.append_concepts(db, path, [concept.id])
    assert result["layout_migration"]["dropped_sheets"] == [
        "Doc Link <> Each fields "]
    assert not [
        entry for entry in result.get("issues", [])
        if entry["code"] == layouts.MIGRATION_DROPPED_SHEET
    ]


def test_the_migration_never_leaves_a_half_written_workbook(tmp_path,
                                                           monkeypatch):
    """The subject is runtime state on a deployed volume.

    A bare ``Workbook.save(path)`` writes the live file in place, so a crash
    or a full disk mid-save truncates it. Every other write in this package
    goes through ``workbook_sync.atomic_save_workbook``; this one does too.
    """
    path = tmp_path / "output.xlsx"
    _canonical_descriptive_row(path, question__question_label="A LABEL")
    before = path.read_bytes()

    real_save = openpyxl.Workbook.save

    def exploding(self, filename):
        Path(filename).write_bytes(b"half a workbook")
        raise OSError("no space left on device")

    monkeypatch.setattr(openpyxl.Workbook, "save", exploding)
    with pytest.raises(OSError):
        layouts.migrate_workbook_layout(
            path, target_layout_id=layouts.REFERENCE_LAYOUT_ID)
    monkeypatch.setattr(openpyxl.Workbook, "save", real_save)

    assert path.read_bytes() == before
    assert openpyxl.load_workbook(path).sheetnames


def test_the_retained_sibling_is_named_for_the_published_workbook(
    db, tmp_path,
):
    """Not for ``mkstemp``'s dot-hidden random stem.

    ``append_concepts`` runs on the staged sibling the transactional outbox
    provides, so naming the pre-migration copy after the file it was handed
    produced ``.bulk_import_output-ozy09ajd.pre-canonical-current.xlsx`` — a
    hidden file, beside the target, whose name does not say which workbook it
    came from, one fresh orphan per run, and it is the path the operator is
    told to recover from.
    """
    from app.services import build_concepts

    target = tmp_path / "bulk_import_output.xlsx"
    concept, _question = _real_writer_canonical_workbook(db, target, "Sibling")
    staged, written = build_concepts._stage_concept_workbook(
        db, target, [concept.id])

    sibling = Path(written["layout_migration"]["sibling_path"])
    assert sibling.name == f"bulk_import_output.pre-{CANONICAL}.xlsx"
    assert sibling.parent == target.parent
    assert sibling.exists()
    assert staged.name not in sibling.name
    # A re-run overwrites this one name instead of accumulating orphans.
    assert [
        found.name for found in tmp_path.glob("*.pre-*.xlsx")
    ] == [sibling.name]


def test_a_canonical_workbook_still_imports_and_appends_after_the_layout_move(
    client, db, tmp_path,
):
    """C5b's actual failure mode: Q16's "readable AND appendable".

    ``data/Testing/AP.xlsx`` is a committed workbook on ``canonical-current``.
    Had ``layouts.py``'s frozen entries been written as live references to
    ``bi.FIELDS_BY_KIND``, this POST would be a 422 the moment S7 landed.
    """
    source = config.ROOT / "data" / "Testing" / "AP.xlsx"
    assert source.exists()
    payload = source.read_bytes()

    response = client.post(
        "/data/import",
        files={
            "file": (
                "AP.xlsx", io.BytesIO(payload),
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet",
            ),
        },
    )
    assert response.status_code == 200, response.text

    working = tmp_path / "AP.xlsx"
    working.write_bytes(payload)
    concept = db.query(models.Concept).first()
    result = writer.append_concepts(db, working, [concept.id])
    assert result["layout_migration"]["source_layout_id"] == CANONICAL
    assert openpyxl.load_workbook(working).sheetnames == bi.SHEET_ORDER


def _all_question_labels(path) -> set[str]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    labels: set[str] = set()
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header = next(
                worksheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
            names = [
                "" if value is None else str(value).strip() for value in header
            ]
            if "question_label" not in names:
                continue
            index = names.index("question_label")
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                if index < len(row) and str(row[index] or "").strip():
                    labels.add(str(row[index]).strip())
    finally:
        workbook.close()
    return labels


# --------------------------------------------------------------------------- #
# T3.8 — the silent row-width repair
# --------------------------------------------------------------------------- #

def test_a_row_wider_than_its_sheet_is_recorded_not_truncated(
    db, tmp_path, monkeypatch,
):
    _concept, _topic, group, question = _graph(db, "Wide")
    real = writer._band_cells

    def wider(sheet_layout, block, values):
        cells = list(real(sheet_layout, block, values))
        if block == "question":
            cells.append("A TAIL WITH NO COLUMN")
        return cells

    monkeypatch.setattr(writer, "_band_cells", wider)
    decisions: list[dict] = []
    row = writer._question_to_row(
        question, "descriptive", group, decisions=decisions)

    assert len(row) == len(bi.FIELDS_BY_KIND["descriptive"])
    assert len(decisions) == 1
    decision = decisions[0]
    assert decision["code"] == writer.ROW_WIDTH_MISMATCH
    assert decision["sheet_kind"] == "descriptive"
    assert decision["expected_width"] == len(bi.FIELDS_BY_KIND["descriptive"])
    assert decision["actual_width"] == decision["expected_width"] + 1
    assert decision["row_identity"] == question.question_label
    # The tail is not lost: it is IN the record.
    assert decision["overflow_values"] == ["A TAIL WITH NO COLUMN"]
    assert decision["review_flag"] is True
    assert len(decision["decision_sha256"]) == 64


def test_a_row_narrower_than_its_sheet_is_recorded_not_padded(
    db, tmp_path, monkeypatch,
):
    _concept, _topic, group, question = _graph(db, "Narrow")
    real = writer._front_bands

    def narrower(*args, **kwargs):
        return list(real(*args, **kwargs))[:-1]

    monkeypatch.setattr(writer, "_front_bands", narrower)
    decisions: list[dict] = []
    row = writer._question_to_row(
        question, "objective", group, decisions=decisions)

    assert len(row) == len(bi.FIELDS_BY_KIND["objective"])
    assert len(decisions) == 1
    decision = decisions[0]
    assert decision["code"] == writer.ROW_WIDTH_MISMATCH
    assert decision["sheet_kind"] == "objective"
    assert decision["actual_width"] == decision["expected_width"] - 1
    assert decision["overflow_values"] == []
    assert decision["review_flag"] is True


def test_a_concept_row_narrower_than_its_front_bands_is_recorded(
    db, tmp_path, monkeypatch,
):
    """T3.8's narrow-side record reaches the CONCEPT composer too.

    ``_concept_to_row`` used to pad to the sheet width before checking it, so
    the very pad that writes OD6's deliberate blank tail also absorbed a
    genuinely short Chapter/Topic/Concept composition — the defect and the
    intended tail were indistinguishable and the recorded decision could never
    fire. The check is on the FRONT bands; the tail is appended after it.
    """
    concept, topic, _group, _question = _graph(db, "ShortFront")
    real = writer._front_bands

    def narrower(*args, **kwargs):
        return list(real(*args, **kwargs))[:-1]

    monkeypatch.setattr(writer, "_front_bands", narrower)
    decisions: list[dict] = []
    row = writer._concept_to_row(
        concept, "objective", topic, decisions=decisions)

    # The row still ships at the sheet's width, tail and all.
    assert len(row) == len(bi.FIELDS_BY_KIND["objective"])
    assert len(decisions) == 1
    decision = decisions[0]
    assert decision["code"] == writer.ROW_WIDTH_MISMATCH
    assert decision["sheet_kind"] == "objective"
    assert decision["actual_width"] == decision["expected_width"] - 1
    assert decision["row_identity"] == concept.concept_title
    assert decision["review_flag"] is True


def test_a_correct_concept_row_records_nothing(db):
    """The OD6 blank tail is not a width defect and must not read as one."""
    concept, topic, _group, _question = _graph(db, "RightFront")
    decisions: list[dict] = []
    row = writer._concept_to_row(
        concept, "objective", topic, decisions=decisions)
    assert len(row) == len(bi.FIELDS_BY_KIND["objective"])
    assert decisions == []
    group_start = min(
        band["start"] for band in bi.SECTION_BANDS["objective"]
        if band["label"].strip() == "Group"
    )
    assert not any(str(value or "").strip() for value in row[group_start - 1:])


# --------------------------------------------------------------------------- #
# OD2 — the committed fixture and the writer move in ONE commit
# --------------------------------------------------------------------------- #

def test_the_committed_fixture_is_on_the_target_layout():
    """``backend/data/bulk_import_database.xlsx``, identified by the registry.

    Nothing appends to this file; ``.gitignore:57`` records it as the
    committed database fixture and CI regenerates it through
    ``app.bulk_import.writer`` before every pytest run, so it acquires the new
    layout automatically the moment the writer moves. The single obligation
    is to COMMIT the regenerated fixture in the same commit as the writer, and
    this is the guard on that: it fails on a tree whose committed fixture and
    whose writer disagree.
    """
    path = config.BULK_IMPORT_DB
    assert path.exists()
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        headers = {
            name: next(
                workbook[name].iter_rows(min_row=2, max_row=2,
                                         values_only=True), ())
            for name in workbook.sheetnames
        }
        assert list(workbook.sheetnames) == bi.SHEET_ORDER
    finally:
        workbook.close()
    identity = layouts.identify_workbook(headers)
    assert identity.layout_id == layouts.REFERENCE_LAYOUT_ID
    assert {found.kind for found in identity.sheets} == {
        "objective", "descriptive", "subjective"}


# --------------------------------------------------------------------------- #
# OWNER RULING OD6 / T17 — the Concept File's shape, on BOTH lanes
# --------------------------------------------------------------------------- #

def test_the_concept_files_are_filled_to_the_concept_band_and_no_further(db):
    """Outputs 01 and 03 — the two files the owner ruled on.

    "In Concepts File, it should be filled till concepts related columns
    (starting from first)." The COLUMN half was already true; the SHEET half
    was false, and it is this slice that makes it true: the builder emitted
    four sheets (including a ``Doc Link <> Each fields `` tab that is in no
    registered layout) in the wrong order.
    """
    from app.services import build_concepts_release_files as release_files
    from tests.test_assessment_release_run import _chapter_with_concepts
    from tests.test_pre_release_lane_wiring import _both_lanes_job

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    group_start = min(
        band["start"] for band in bi.SECTION_BANDS["objective"]
        if band["label"].strip() == "Group"
    )
    labels_index = bi.FIELDS_BY_KIND["objective"].index("topic_concept_labels")
    source_index = bi.FIELDS_BY_KIND["objective"].index("concept_source")

    for lane in (release_files.LANE_PRE, release_files.LANE_POST):
        data = release_files.build_release_bulk_import_workbook(
            db, job, lane=lane)
        workbook = openpyxl.load_workbook(io.BytesIO(data))

        # Exactly the format workbook's three sheets, in its order, and NO
        # Doc Link sheet.
        assert workbook.sheetnames == bi.SHEET_ORDER, lane
        assert "Doc Link <> Each fields " not in workbook.sheetnames

        rows_by_sheet = {}
        for kind, sheet_name in bi.SHEET_BY_KIND.items():
            sheet = workbook[sheet_name]
            assert [c.value for c in sheet[2]] == bi.FIELDS_BY_KIND[kind]
            rows_by_sheet[kind] = [
                row for row in sheet.iter_rows(min_row=3, values_only=True)
                if row and any(str(v or "").strip() for v in row)
            ]
        # Only Objective carries data rows.
        assert rows_by_sheet["objective"], lane
        assert rows_by_sheet["descriptive"] == []
        assert rows_by_sheet["subjective"] == []

        for row in rows_by_sheet["objective"]:
            tail = row[group_start - 1:]
            assert not any(str(v or "").strip() for v in tail), (lane, tail)
            assert str(row[labels_index] or "").strip(), lane
            assert str(row[source_index] or "").strip(), lane


@pytest.mark.parametrize("layout_id", [
    "canonical-current", "canonical-no-question-text",
    "canonical-legacy-concept-band",
])
def test_every_registered_older_layout_has_a_migration_home(layout_id):
    """A registered layout the migration cannot map is a trap, not a feature."""
    source = layouts.layout(layout_id)
    target = layouts.layout(layouts.REFERENCE_LAYOUT_ID)
    for kind, sheet in source.sheets.items():
        aliased = {
            alias.source
            for alias in layouts.aliases_for(
                layout_id, layouts.REFERENCE_LAYOUT_ID, kind)
        }
        homeless = [
            (block, name)
            for block, names in sheet.blocks
            for name in names
            if target.sheet(kind).column(block, name) is None
            and (block, name) not in aliased
        ]
        # parent_concept and concept_source (Descriptive) genuinely have no
        # home; every one of them is a RECORDED unmappable entry rather than a
        # silent drop, and nothing else is homeless.
        assert {name for _block, name in homeless} <= {
            "parent_concept", "concept_source"}, (layout_id, kind, homeless)
