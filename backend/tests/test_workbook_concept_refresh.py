"""Regression coverage for in-place Concept-band workbook refreshes."""

from collections import Counter
from pathlib import Path

import openpyxl

from app import bulk_import as bi
from app import models
from app.bulk_import import layouts, writer
from app.services import identity


def _write_layout_workbook(path, layout_id: str):
    """An empty workbook on one REGISTERED layout, built from the registry.

    Never from ``writer._new_workbook`` with the header row overwritten: that
    fixture moved with the writer, which is exactly how a canonical-layout
    regression can go green while the canonical layout has stopped existing.
    """
    entry = layouts.layout(layout_id)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for sheet_layout in entry.sheets.values():
        worksheet = workbook.create_sheet(sheet_layout.sheet_name)
        worksheet.append([band.label for band in sheet_layout.bands])
        worksheet.append(list(sheet_layout.fields))
    workbook.save(path)
    return path


def _refresh_graph(db, suffix: str, *, include_tag: bool = True):
    chapter = models.Chapter(
        chapter_code=f"10CBMA_Refresh_{suffix}",
        board="CBSE",
        grade="10",
        subject="Mathematics",
        unit="Workbook Refresh",
        chapter_title=f"Workbook Refresh {suffix}",
        chapter_display_name=f"Workbook Refresh {suffix}",
    )
    home_topic = models.Topic(
        chapter=chapter,
        topic_title="Home Topic",
        topic_display_name="Home Topic",
        pre_post_learning="Post",
    )
    tagged_topic = models.Topic(
        chapter=chapter,
        topic_title="Tagged Topic",
        topic_display_name="Tagged Topic",
        pre_post_learning="Post",
    )
    concept = models.Concept(
        topic=home_topic,
        concept_title="Ohm's Law",
        concept_display_name="Ohm's Law",
        parent_concept="Original Parent",
        concept_details="Original concept details",
        keywords="original keyword",
        related_concepts="Original Relation",
        sources="Original Book",
    )
    basic = models.Group(
        concept=concept,
        group_type="Basic",
        group_name="Original Basic",
        group_display_name="Original Basic",
        group_status="Active",
    )
    question = models.Question(
        group=basic,
        sheet_kind="objective",
        question_label=f"REFRESH-{suffix}",
        question_category="Multiple Choice Question",
        cognitive_skills="Understanding",
        question_source="Original Book",
        question="Which statement describes Ohm's law?",
        answers=[],
    )
    db.add(chapter)
    if include_tag:
        db.add(models.ConceptTag(concept=concept, topic=tagged_topic))
    db.commit()
    return concept, home_topic, tagged_topic, basic, question


def _objective_rows(path):
    ws = openpyxl.load_workbook(path)[bi.SHEET_OBJECTIVE]
    return ws, [
        (row_i, tuple(cell.value for cell in ws[row_i]))
        for row_i in range(3, ws.max_row + 1)
        if any(cell.value is not None for cell in ws[row_i])
    ]


def test_append_concepts_refreshes_complete_band_after_normalized_title_change(
    db, tmp_path,
):
    concept, home_topic, tagged_topic, basic, question = _refresh_graph(
        db, "Canonical")
    path = tmp_path / "canonical-refresh.xlsx"

    first = writer.append_concepts(db, path, [concept.id])
    assert first["written"] == 2
    assert writer.append_questions(db, path, [question.id])["objective"] == 1

    before_ws, before_rows = _objective_rows(path)
    before_max_row = before_ws.max_row
    before_question_rows = {
        row_i
        for row_i, row in before_rows
        if writer._cell_str(row, writer._q_start("objective"))
    }

    # A later generation pass corrects content and presentation without
    # changing the normalized identity of the concept.
    original_title_key = bi.normalize_question_text(concept.concept_title)
    concept.concept_title = "OHM'S LAW"
    concept.concept_display_name = "OHM'S LAW"
    assert bi.normalize_question_text(concept.concept_title) == original_title_key
    concept.parent_concept = "Electrical Relationships"
    concept.concept_details = "Corrected, complete concept details"
    concept.sources = "Current Book"
    basic.group_name = "Refreshed Basic"
    basic.group_display_name = "Refreshed Basic"
    db.add(models.Group(
        concept=concept,
        group_type="Intermediate",
        group_name="Refreshed Intermediate",
        group_display_name="Refreshed Intermediate",
        group_status="Active",
    ))
    db.commit()

    refreshed = writer.append_concepts(db, path, [concept.id])
    assert refreshed["written"] == 0
    assert refreshed["sources_updated"] == 3

    ws, rows = _objective_rows(path)
    assert ws.max_row == before_max_row
    assert {row_i for row_i, row in rows
            if writer._cell_str(row, writer._q_start("objective"))} \
        == before_question_rows
    assert len(rows) == 3

    details_col = bi.OBJECTIVE_FIELDS.index("concept_details")
    source_col = bi.OBJECTIVE_FIELDS.index("concept_source")
    basic_col = bi.OBJECTIVE_FIELDS.index("basic_groups")
    intermediate_col = bi.OBJECTIVE_FIELDS.index("intermediate_groups")
    placement_counts = Counter()
    expected_topics = {
        bi.normalize_question_text(home_topic.topic_title): home_topic,
        bi.normalize_question_text(tagged_topic.topic_title): tagged_topic,
    }
    title_cells_by_topic = {}

    for _, row in rows:
        clean_title = bi.strip_title_tag(
            writer._cell_str(row, writer._IDX_CONCEPT_TITLE))
        clean_topic = bi.strip_topic_title(
            writer._cell_str(row, writer._IDX_TOPIC_TITLE))
        topic_key = bi.normalize_question_text(clean_topic)
        topic = expected_topics[topic_key]
        placement_counts[topic_key] += 1

        expected_title = writer._concept_to_row(
            concept, "objective", topic)[writer._IDX_CONCEPT_TITLE]
        assert writer._cell_str(
            row, writer._IDX_CONCEPT_TITLE) == expected_title
        assert clean_title == concept.concept_title
        assert writer._cell_str(
            row, writer._IDX_CONCEPT_TITLE + 1) == concept.concept_title
        assert writer._cell_str(row, details_col) == concept.concept_details
        assert writer._cell_str(row, source_col) == (
            "Original Book, Current Book")
        title_cells_by_topic[topic_key] = expected_title

        is_question_row = bool(
            writer._cell_str(row, writer._q_start("objective")))
        if is_question_row:
            # Q16 (SOP §6.1): the export composes the group ID itself —
            # the stale stored "Refreshed …" names never reach the sheet.
            machine = identity.machine_id_for_concept(concept)
            assert writer._cell_str(row, basic_col) == (
                identity.compose_group_key(machine, "Basic", 1))
            assert writer._cell_str(row, intermediate_col) == (
                identity.compose_group_key(machine, "Intermediate", 1))
        else:
            assert writer._cell_str(row, basic_col) == ""
            assert writer._cell_str(row, intermediate_col) == ""

    assert placement_counts == {
        bi.normalize_question_text(home_topic.topic_title): 2,
        bi.normalize_question_text(tagged_topic.topic_title): 1,
    }
    # ONE identity, repeated under both placements — which is precisely how
    # the CMS reads a repeated row as a TAG rather than as a second concept.
    # Before S4 the tag came from ``directory.concept_tag``, which is derived
    # from the PLACEMENT topic, so the same concept exported two different
    # machine ids and the two rows read as two concepts. The id is now the
    # concept's own persisted ``machine_id`` (spec-step8 T4-3/T4-4), so it
    # does not move with the placement.
    assert len(set(title_cells_by_topic.values())) == 1
    assert set(title_cells_by_topic.values()) == {
        f"{concept.concept_title} ({concept.machine_id})"
    }


def test_append_concepts_relocates_catalog_and_question_rows_in_place(
    db, tmp_path,
):
    concept, old_topic, tagged_topic, _, question = _refresh_graph(
        db, "Relocation", include_tag=True)
    new_topic = models.Topic(
        chapter=old_topic.chapter,
        topic_title="Relocated Home Topic",
        topic_display_name="Relocated Home Topic",
        pre_post_learning="Post",
    )
    db.add(new_topic)
    db.commit()
    concept_id = concept.id
    new_topic_id = new_topic.id
    path = tmp_path / "home-topic-relocation.xlsx"

    assert writer.append_concepts(db, path, [concept.id])["written"] == 2
    assert writer.append_questions(db, path, [question.id])["objective"] == 1
    ws, before_rows = _objective_rows(path)
    assert len(before_rows) == 3
    before_max_row = ws.max_row
    q_start = writer._q_start("objective")
    before_question = next(
        (row_i, row)
        for row_i, row in before_rows
        if writer._cell_str(row, q_start) == question.question_label
    )
    before_question_band = before_question[1][q_start:]
    assert {
        bi.strip_topic_title(
            writer._cell_str(row, writer._IDX_TOPIC_TITLE))
        for _, row in before_rows
    } == {old_topic.topic_title, tagged_topic.topic_title}

    concept.topic_id = new_topic_id
    concept.sources = "Relocated Book"
    db.commit()
    db.expire_all()
    concept = db.get(models.Concept, concept_id)
    new_topic = db.get(models.Topic, new_topic_id)

    result = writer.append_concepts(db, path, [concept.id])
    assert result["written"] == 0
    assert result["sources_updated"] == 3

    ws, after_rows = _objective_rows(path)
    assert ws.max_row == before_max_row
    assert [row_i for row_i, _ in after_rows] == [
        row_i for row_i, _ in before_rows]
    assert len(after_rows) == 3
    assert Counter(
        bi.strip_topic_title(
            writer._cell_str(row, writer._IDX_TOPIC_TITLE))
        for _, row in after_rows
    ) == {
        new_topic.topic_title: 2,
        tagged_topic.topic_title: 1,
    }

    after_question = next(
        (row_i, row)
        for row_i, row in after_rows
        if writer._cell_str(row, q_start) == question.question_label
    )
    assert after_question[0] == before_question[0]
    # Placement/front-band movement must not rewrite assessment content,
    # answers, or question_source.
    assert after_question[1][q_start:] == before_question_band
    assert writer._cell_str(
        after_question[1], q_start + 3) == question.question_source
    source_col = bi.OBJECTIVE_FIELDS.index("concept_source")
    assert {
        writer._cell_str(row, source_col)
        for _, row in after_rows
    } == {"Original Book, Relocated Book"}

    index = writer.scan_workbook(path)
    assert writer.concept_placement_key(concept, new_topic) \
        in index.c_placements
    assert writer.concept_placement_key(concept, tagged_topic) \
        in index.c_placements
    assert not any(
        key[0] == bi.normalize_question_text(concept.concept_title)
        and key[1] == bi.normalize_question_text(
            old_topic.chapter.chapter_title)
        and key[2] == bi.normalize_question_text(old_topic.topic_title)
        for key in index.c_placements
    )


def test_reconciliation_preserves_tags_and_never_hijacks_other_chapters(
    db, tmp_path,
):
    foreign, foreign_home, _, _, _ = _refresh_graph(
        db, "Foreign", include_tag=False)
    concept, home, tagged, _, question = _refresh_graph(
        db, "Tagged", include_tag=True)
    path = tmp_path / "tags-and-chapters.xlsx"

    assert writer.append_concepts(db, path, [foreign.id])["written"] == 1
    assert writer.append_concepts(db, path, [concept.id])["written"] == 2
    assert writer.append_questions(db, path, [question.id])["objective"] == 1
    before_ws, before_rows = _objective_rows(path)
    before_max_row = before_ws.max_row
    foreign_chapter = foreign_home.chapter.chapter_title
    foreign_before = [
        row for _, row in before_rows
        if bi.strip_title_tag(
            writer._cell_str(row, writer._IDX_CHAPTER_TITLE)
        ) == foreign_chapter
    ]
    assert len(foreign_before) == 1

    concept.concept_details = "Updated only in the tagged concept's chapter"
    concept.sources = "Tagged Current Book"
    db.commit()

    result = writer.append_concepts(db, path, [concept.id])
    assert result["written"] == 0
    ws, rows = _objective_rows(path)
    assert ws.max_row == before_max_row
    assert len(rows) == 4

    details_col = bi.OBJECTIVE_FIELDS.index("concept_details")
    q_start = writer._q_start("objective")
    current_rows = []
    foreign_after = []
    for row_i, row in rows:
        chapter_title = bi.strip_title_tag(
            writer._cell_str(row, writer._IDX_CHAPTER_TITLE))
        if chapter_title == foreign_chapter:
            foreign_after.append(row)
        else:
            current_rows.append((row_i, row))

    assert foreign_after == foreign_before
    assert all(
        writer._cell_str(row, details_col) == concept.concept_details
        for _, row in current_rows
    )
    placement_counts = Counter(
        bi.strip_topic_title(
            writer._cell_str(row, writer._IDX_TOPIC_TITLE))
        for _, row in current_rows
    )
    assert placement_counts == {
        home.topic_title: 2,
        tagged.topic_title: 1,
    }
    assert [
        bi.strip_topic_title(
            writer._cell_str(row, writer._IDX_TOPIC_TITLE))
        for _, row in current_rows
        if writer._cell_str(row, q_start)
    ] == [home.topic_title]


def test_reconciliation_never_hijacks_same_title_pre_learning_rows(
    db, tmp_path,
):
    chapter = models.Chapter(
        chapter_code="10CBMA_PrePostIdentity",
        board="CBSE",
        grade="10",
        subject="Mathematics",
        unit="Workbook Refresh",
        chapter_title="Pre Post Identity",
        chapter_display_name="Pre Post Identity",
    )
    pre_topic = models.Topic(
        chapter=chapter,
        topic_title="Shared Topic",
        topic_display_name="Shared Topic",
        pre_post_learning="Pre",
    )
    post_topic = models.Topic(
        chapter=chapter,
        topic_title="Shared Topic",
        topic_display_name="Shared Topic",
        pre_post_learning="Post",
    )
    pre_concept = models.Concept(
        topic=pre_topic,
        concept_title="Shared Concept",
        concept_display_name="Shared Concept",
        concept_details="Pre-learning details must remain unchanged.",
        sources="Prerequisite Book",
    )
    post_concept = models.Concept(
        topic=post_topic,
        concept_title="Shared Concept",
        concept_display_name="Shared Concept",
        concept_details="Original post-learning details.",
        sources="Chapter Book",
    )
    db.add(chapter)
    db.commit()
    path = tmp_path / "pre-post-identity.xlsx"

    assert writer.append_concepts(
        db, path, [pre_concept.id])["written"] == 1
    assert writer.append_concepts(
        db, path, [post_concept.id])["written"] == 1
    _, initial_rows = _objective_rows(path)
    assert len(initial_rows) == 2
    assert {
        writer._cell_str(row, writer._IDX_TOPIC_PRE_POST)
        for _, row in initial_rows
    } == {"Pre", "Post"}

    relocated_post_topic = models.Topic(
        chapter=chapter,
        topic_title="Relocated Post Topic",
        topic_display_name="Relocated Post Topic",
        pre_post_learning="Post",
    )
    db.add(relocated_post_topic)
    db.flush()
    post_concept.topic_id = relocated_post_topic.id
    post_concept.concept_details = "Updated post-learning details."
    post_concept.sources = "Current Chapter Book"
    db.commit()

    result = writer.append_concepts(db, path, [post_concept.id])

    assert result["written"] == 0
    _, rows = _objective_rows(path)
    assert len(rows) == 2
    details_col = bi.OBJECTIVE_FIELDS.index("concept_details")
    rows_by_kind = {
        writer._cell_str(row, writer._IDX_TOPIC_PRE_POST): row
        for _, row in rows
    }
    assert bi.strip_topic_title(writer._cell_str(
        rows_by_kind["Pre"], writer._IDX_TOPIC_TITLE,
    )) == pre_topic.topic_title
    assert writer._cell_str(
        rows_by_kind["Pre"], details_col,
    ) == pre_concept.concept_details
    assert bi.strip_topic_title(writer._cell_str(
        rows_by_kind["Post"], writer._IDX_TOPIC_TITLE,
    )) == relocated_post_topic.topic_title
    assert writer._cell_str(
        rows_by_kind["Post"], details_col,
    ) == post_concept.concept_details

    index = writer.scan_workbook(path)
    assert writer.concept_placement_key(
        pre_concept, pre_topic) in index.c_placements
    assert writer.concept_placement_key(
        post_concept, relocated_post_topic) in index.c_placements
    assert writer.concept_placement_key(
        post_concept, post_topic) not in index.c_placements


def test_append_concepts_migrates_a_legacy_workbook_then_refreshes_its_fields(
    db, tmp_path,
):
    """Appending to an OLDER registered layout migrates it once, first.

    This test used to build a legacy-layout sheet and assert that
    ``append_concepts`` kept writing legacy columns in place. Since S7 the
    append MIGRATES the workbook onto the target layout before it writes
    anything (Q16 / T7.6) — writing target values at old-layout positions is
    the corruption the migration exists to prevent — so what is pinned here is
    that the migration happens once, that it is recorded, and that the fields
    still refresh by NAME afterwards.
    """
    concept, _, _, basic, question = _refresh_graph(db, "Legacy")
    path = tmp_path / "legacy-refresh.xlsx"
    _write_layout_workbook(path, "canonical-legacy-concept-band")

    first = writer.append_concepts(db, path, [concept.id])
    assert first["written"] == 2
    receipt = first["layout_migration"]
    assert receipt["source_layout_id"] == "canonical-legacy-concept-band"
    assert receipt["target_layout_id"] == layouts.REFERENCE_LAYOUT_ID
    assert Path(receipt["sibling_path"]).exists()
    assert writer.append_questions(db, path, [question.id])["objective"] == 1
    before_ws, _ = _objective_rows(path)
    before_max_row = before_ws.max_row

    concept.concept_details = "Corrected legacy concept details"
    concept.keywords = "corrected, searchable, keywords"
    concept.related_concepts = "Current Legacy Relation"
    basic.group_name = "Legacy Basic Updated"
    basic.group_display_name = "Legacy Basic Updated"
    db.commit()

    result = writer.append_concepts(db, path, [concept.id])
    assert result["written"] == 0
    # Migrated ONCE: the second append finds the workbook already on target.
    assert "layout_migration" not in result

    ws, rows = _objective_rows(path)
    assert ws.max_row == before_max_row
    assert len(rows) == 3
    headers = [cell.value for cell in ws[2]]
    assert headers == bi.FIELDS_BY_KIND["objective"]
    details_col = headers.index("concept_details")
    keywords_col = headers.index("keywords")
    related_col = headers.index("related_concepts")
    basic_col = headers.index("basic_groups")

    for _, row in rows:
        assert writer._cell_str(row, details_col) == concept.concept_details
        assert writer._cell_str(row, keywords_col) == concept.keywords
        # The legacy "parent: X" marker is gone with the column itself, so
        # related_concepts carries only genuine relations now.
        assert writer._cell_str(row, related_col) == "Current Legacy Relation"
        if writer._cell_str(row, writer._q_start("objective")):
            # Q16 (SOP §6.1): the export composes the group ID itself;
            # the stale stored "Legacy Basic Updated" never reaches it.
            assert writer._cell_str(row, basic_col) == (
                identity.compose_group_key(
                    identity.machine_id_for_concept(concept), "Basic", 1))
        else:
            assert writer._cell_str(row, basic_col) == ""
