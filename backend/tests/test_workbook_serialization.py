"""Concurrency and formula-safety regressions for Bulk Import output."""
from __future__ import annotations

import io
import threading

import openpyxl
import pytest

from app import config, models
from app.bulk_import import SHEET_OBJECTIVE, workbook_sync, writer
from app.db import SessionLocal
from app.services import build_assessments, build_concepts, generation


class _AttemptAwareRLock:
    """Real RLock that exposes when a second thread starts acquiring it."""

    def __init__(self) -> None:
        self._lock = threading.RLock()
        self._meta_lock = threading.Lock()
        self._attempts = 0
        self.second_attempted = threading.Event()

    def __enter__(self):
        with self._meta_lock:
            self._attempts += 1
            if self._attempts == 2:
                self.second_attempted.set()
        self._lock.acquire()
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self._lock.release()


def test_append_questions_serializes_the_complete_mutation(
    tmp_path, monkeypatch,
):
    observed_lock = _AttemptAwareRLock()
    first_scan_entered = threading.Event()
    release_first_scan = threading.Event()
    second_scan_entered = threading.Event()
    calls_lock = threading.Lock()
    scan_calls = 0
    errors: list[BaseException] = []

    def controlled_scan(_path):
        nonlocal scan_calls
        with calls_lock:
            scan_calls += 1
            call_number = scan_calls
        if call_number == 1:
            first_scan_entered.set()
            if not release_first_scan.wait(timeout=3):
                raise AssertionError("timed out waiting to release first writer")
        else:
            second_scan_entered.set()
        return writer.WorkbookIndex()

    monkeypatch.setattr(
        workbook_sync,
        "_OUTPUT_WORKBOOK_LOCK",
        observed_lock,
    )
    monkeypatch.setattr(writer, "scan_workbook", controlled_scan)
    monkeypatch.setattr(writer, "_questions", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(writer, "_new_workbook", object)
    monkeypatch.setattr(
        writer,
        "_migrate_existing_question_cells",
        lambda _workbook: 0,
    )
    monkeypatch.setattr(
        workbook_sync,
        "atomic_save_workbook",
        lambda *_args, **_kwargs: None,
    )

    def append() -> None:
        try:
            writer.append_questions(
                object(),
                tmp_path / "bulk_import_output.xlsx",
                [],
            )
        except BaseException as exc:  # surfaced on the main test thread below
            errors.append(exc)

    first = threading.Thread(target=append)
    second = threading.Thread(target=append)
    first.start()
    assert first_scan_entered.wait(timeout=3)
    second.start()
    assert observed_lock.second_attempted.wait(timeout=3)

    # The second writer has reached the same process-wide lock, but cannot
    # begin even its initial workbook scan while the first mutation is active.
    assert not second_scan_entered.is_set()

    release_first_scan.set()
    first.join(timeout=3)
    second.join(timeout=3)
    assert not first.is_alive()
    assert not second.is_alive()
    assert second_scan_entered.is_set()
    assert errors == []


def test_concurrent_assessment_sessions_reserve_unique_labels_before_publish(
    db, first_concept, tmp_path, monkeypatch,
):
    """Two sessions targeting one concept must both reach the workbook."""
    output = tmp_path / "concurrent-assessments.xlsx"
    monkeypatch.setattr(config, "BULK_IMPORT_OUTPUT", output)

    session_ids: list[int] = []
    for _ in range(2):
        session = build_assessments.create_session(
            db, "concept", [first_concept["id"]]
        )
        build_assessments.add_batch(
            db,
            session.id,
            cognitive_skills=["Understand"],
            difficulty_levels=["Moderate"],
            categories=["Multiple Choice Question"],
            question_type="objective",
            num_questions=1,
        )
        session_ids.append(session.id)

    # Both model calls finish together. Before the shared finalization lock was
    # extended to label allocation, this forced both sessions to reserve the
    # same counter and the workbook silently skipped one of their rows.
    generation_barrier = threading.Barrier(2)
    original_generate = generation.generate_questions_for_concept

    def synchronized_generate(*args, **kwargs):
        generation_barrier.wait(timeout=5)
        return original_generate(*args, **kwargs)

    monkeypatch.setattr(
        generation, "generate_questions_for_concept", synchronized_generate
    )

    results: list[dict] = []
    errors: list[BaseException] = []

    def generate_session(session_id: int) -> None:
        local_db = SessionLocal()
        try:
            results.append(build_assessments.generate(local_db, session_id))
        except BaseException as exc:  # surfaced on the main test thread below
            errors.append(exc)
        finally:
            local_db.close()

    threads = [
        threading.Thread(target=generate_session, args=(session_id,))
        for session_id in session_ids
    ]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=10)

    assert all(not thread.is_alive() for thread in threads)
    assert errors == []
    assert len(results) == 2

    question_ids = [
        question_id
        for result in results
        for question_id in result["question_ids"]
    ]
    db.expire_all()
    questions = (
        db.query(models.Question)
        .filter(models.Question.id.in_(question_ids))
        .order_by(models.Question.id)
        .all()
    )
    labels = [question.question_label for question in questions]

    assert len(questions) == 2
    assert len(set(labels)) == 2
    assert all(result["pipeline"]["appended"]["skipped"] == 0
               for result in results)
    workbook_index = writer.scan_workbook(output)
    assert set(labels).issubset(workbook_index.labels)


@pytest.mark.parametrize("value", ["=2+2", "+SUM(A1:A2)", "-1+2", "@A1"])
def test_formula_leading_xlsx_values_round_trip_as_plain_text(value):
    workbook = writer._new_workbook()
    worksheet = workbook[SHEET_OBJECTIVE]
    writer._write_cell(worksheet, row=3, column=1, value=value)

    output = io.BytesIO()
    workbook.save(output)
    loaded = openpyxl.load_workbook(io.BytesIO(output.getvalue()))
    cell = loaded[SHEET_OBJECTIVE].cell(row=3, column=1)

    assert cell.value == value
    assert cell.data_type == "s"


def test_public_question_export_keeps_formula_leading_content_literal(db):
    question = db.query(models.Question).order_by(models.Question.id).first()
    assert question is not None
    original = question.question
    try:
        question.question = "=2+2"
        db.flush()
        output = writer.write_workbook(db, question_ids=[question.id])
        loaded = openpyxl.load_workbook(io.BytesIO(output))
        formula_cells = [
            cell
            for worksheet in loaded.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value == "=2+2"
        ]
        assert formula_cells
        assert all(cell.data_type == "s" for cell in formula_cells)
    finally:
        question.question = original
        db.rollback()


def test_concept_export_rejects_cell_text_above_excel_limit(db):
    concept = db.query(models.Concept).order_by(models.Concept.id).first()
    assert concept is not None
    original = concept.concept_details
    try:
        concept.concept_details = "x" * (
            writer.EXCEL_CELL_CHARACTER_LIMIT + 1
        )
        db.flush()

        with pytest.raises(writer.ExcelCellLimitError) as caught:
            writer.write_concepts_workbook(db, [concept.id])

        message = str(caught.value)
        assert "Excel export blocked to prevent data loss" in message
        assert "field 'concept_details'" in message
        assert "32,768 characters" in message
        assert "32,767-character cell limit" in message
        assert "split it across multiple cells or records" in message
    finally:
        concept.concept_details = original
        db.rollback()


def test_concept_export_allows_text_at_exact_excel_limit(db):
    concept = db.query(models.Concept).order_by(models.Concept.id).first()
    assert concept is not None
    original = concept.concept_details
    exact_value = "x" * writer.EXCEL_CELL_CHARACTER_LIMIT
    try:
        concept.concept_details = exact_value
        db.flush()

        output = writer.write_concepts_workbook(db, [concept.id])
        workbook = openpyxl.load_workbook(
            io.BytesIO(output), data_only=True, read_only=True
        )
        try:
            worksheet = workbook[SHEET_OBJECTIVE]
            headers = next(
                worksheet.iter_rows(
                    min_row=2,
                    max_row=2,
                    values_only=True,
                )
            )
            details_column = headers.index("concept_details") + 1
            assert worksheet.cell(row=3, column=details_column).value == exact_value
        finally:
            workbook.close()
    finally:
        concept.concept_details = original
        db.rollback()


@pytest.mark.parametrize("value", ["=cmd()", "+cmd()", "-cmd()", "@cmd()"])
def test_concept_inventory_csv_formula_values_are_escaped(value):
    assert build_concepts._csv_safe_cell(value) == f"'{value}"
