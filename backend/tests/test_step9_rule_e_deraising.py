"""S9 — Rule E enforced: a defect blocks the DATABASE WRITE, never a download.

Every test here failed at 76c84fb, and each names what it measured there.
The theme is one sentence from spec-step8 §6/T9: *nothing ever blocks a
download*. Before this slice a single unusable staged row raised out of
``build_concepts_release_files.transient_release_hierarchy`` and cost the
reviewer the Concept File AND both Master builds — a run's own diagnosis
taking away the evidence a reviewer needs to act on it.
"""
from __future__ import annotations

import ast
import copy
import inspect
from pathlib import Path

import pytest

from app import models
from app.services import assessment_release_run as run
from app.services import assessment_release_snapshot as snapshot_mod
from app.services import build_concepts_release as release
from app.services import build_concepts_release_files as release_files
from app.services import build_concepts_release_publication as publication
from app.bulk_import import layouts
from app.bulk_import import assessment_workbook as aw

from tests.test_assessment_release_run import (
    OWNER,
    _authorities,
    _chapter_with_concepts,
    _decision_context,
)

_SOUND_ROW = {
    "topic": "Shapes",
    "concept_title": "Solid Shapes",
    "concept_details": (
        "Solid shapes occupy space in three dimensions; cubes are examples."
    ),
    "keywords": "solid, cube, three dimensions",
}
_SOUND_SECOND_ROW = {
    "topic": "Shapes",
    "concept_title": "Plane Shapes",
    "concept_details": "Plane shapes are flat, two-dimensional figures.",
    "keywords": "plane, flat, two dimensions",
}
# Row 2 with no ``topic``: the measured case the spec names. It has no place
# in the chapter hierarchy the four outputs are projected onto, so it reaches
# no workbook row and no database row.
_TOPICLESS_ROW = dict(_SOUND_SECOND_ROW, topic="")

_INVENTORY = {
    "items": [
        {
            "qid": "QINV-0001",
            "source_kind": "exercise",
            "source_label": "Exercise 1(1)",
            "raw_task": "Which of these is a solid?",
            "options": ["Cube", "Circle"],
        },
        {
            "qid": "QINV-0002",
            "source_kind": "exercise",
            "source_label": "Exercise 1(2)",
            "raw_task": "Explain why a cube is a solid.",
        },
    ],
}

_DOWNLOADS = (
    "release-bulk-import.xlsx",
    "release.xlsx",
    "diagnostics.zip",
    "release.json",
)


def _job_with_records(db, chapter, records) -> models.UploadJob:
    job = models.UploadJob(
        owner_sub=OWNER,
        module="build_concepts",
        upload_type="textbook",
        filename="ch.mmd",
        mmd_text="# Chapter\n\nExercise 1. Which of these is a solid?",
        status="generated",
        deposit_scope_type="chapter",
        deposit_scope_ids=[chapter.id],
        question_inventory=copy.deepcopy(_INVENTORY),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=copy.deepcopy(records),
        inventory=copy.deepcopy(_INVENTORY),
        reason="S9 fixture",
    )
    db.refresh(job)
    return job


# --------------------------------------------------------------------------- #
# 1. The Rule-E pin
# --------------------------------------------------------------------------- #

def test_no_new_raise_inside_transient_release_hierarchy():
    """ONE raise, and it is the one D8.4 names.

    [measured at 76c84fb] this function held EIGHT ``raise ValueError``
    statements. Seven of them judged the release's CONTENT — a missing
    chapter row, missing directory metadata, a records value that is not an
    array, and three row-level conditions — and every one of them cost
    ``release-bulk-import.xlsx`` (404) plus both Master builds (400).

    The survivor is ``this upload has no staged release``, which is genuine
    impossibility rather than a defect: there is no artefact to project and
    no frozen directory metadata to build a chapter row from, so returning
    a hierarchy would mean inventing one. D8.4 names exactly this raise as
    the one ``api/build_concepts`` keeps mapping to 404.

    Asserted over the parsed source rather than by behaviour, because the
    thing being pinned is that a LATER slice does not quietly add a ninth.
    """
    source = inspect.getsource(release_files.transient_release_hierarchy)
    tree = ast.parse(source.lstrip())
    raises = [node for node in ast.walk(tree) if isinstance(node, ast.Raise)]
    assert len(raises) == 1, (
        "transient_release_hierarchy may hold exactly one raise; a new one "
        "takes a download away from a reviewer"
    )
    assert "this upload has no staged release" in ast.unparse(raises[0])

    # And the function hands its verdict back rather than throwing it.
    assert release_files.transient_release_hierarchy.__code__.co_argcount >= 2
    signature = inspect.signature(release_files.transient_release_hierarchy)
    assert "payload" in signature.parameters


# --------------------------------------------------------------------------- #
# 2. One malformed row costs the row, not the run
# --------------------------------------------------------------------------- #

def test_one_malformed_staged_row_does_not_kill_any_download(db, client):
    """The measured two-row payload: row 2 carries no ``topic``.

    [measured at 76c84fb] ``transient_release_hierarchy`` raised "staged
    concept row 2 has no topic"; ``release-bulk-import.xlsx`` returned 404
    while the other three downloads returned 200, and
    ``POST /build-assessments/releases/from-job/{id}`` returned 400 with
    that same string. ``structural_defects`` returned ``[]`` and
    ``release_state`` read ``ready_with_flags`` — so the defect took the
    artifacts away and appeared in the release state nowhere at all.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW, _TOPICLESS_ROW])

    payload = release.release_payload(job)
    assert len(payload["records"]) == 2

    # The defect is RECORDED, at staging, where the payload can carry it.
    recorded = payload[release.STAGED_ROW_DEFECTS_FIELD]
    assert [row["position"] for row in recorded] == [2]
    assert recorded[0]["code"] == release.STAGED_ROW_UNUSABLE

    # A reviewer reading only the issues list sees it too, by name.
    named = [
        issue for issue in payload["issues"]
        if issue["code"] == release.STAGED_ROW_UNUSABLE
    ]
    assert named, "the row-level defect is a named release issue"
    assert "row 2" in named[0]["message"]

    # It reaches the release state — the producer's consumer.
    assert release.structural_defects(payload)
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE

    # ALL FOUR downloads, including the one that 404'd.
    for name in _DOWNLOADS:
        response = client.get(f"/build-concepts/uploads/{job.id}/{name}")
        assert response.status_code == 200, name
        assert response.content, name

    # The workbook carries the OTHER row — the sound one still ships.
    _chapter, concepts, records, defects = (
        release_files.transient_release_hierarchy(db, job, payload=payload))
    assert [c.concept_title for c in concepts] == ["Solid Shapes"]
    assert len(records) == 1
    assert [row["code"] for row in defects] == [release.STAGED_ROW_UNUSABLE]

    # And the DATABASE WRITE is refused. That is the whole of Rule E:
    # the write is blocked, the evidence is not.
    with pytest.raises(ValueError, match="row 2"):
        publication.upload_release_to_database(
            db, job.id, owner_sub=OWNER, lane="post")


def test_the_bulk_import_workbook_stays_a_readable_bulk_import_workbook(
    db, client,
):
    """The issues note may not cost the file its identity.

    ``layouts.identify_workbook`` refuses a WHOLE workbook over one sheet
    no layout claims, and ``reader.import_workbook`` reads data rows from
    row 3 — so the note cannot be a sheet and cannot be a data row. It is
    written on row 1, past the last banded column. This test is the reason
    that constraint is written down rather than assumed.
    """
    import io

    import openpyxl

    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW, _TOPICLESS_ROW])

    data = release_files.build_release_bulk_import_workbook(db, job)
    workbook = openpyxl.load_workbook(io.BytesIO(data))

    # Still the three-sheet canonical format, still identifiable.
    headers = {
        name: next(
            workbook[name].iter_rows(min_row=2, max_row=2, values_only=True),
            (),
        )
        for name in workbook.sheetnames
    }
    identity = layouts.identify_workbook(headers)
    assert identity.layout_id

    # The note is on row 1 and says what happened.
    sheet = workbook[workbook.sheetnames[0]]
    row_one = " ".join(
        str(cell.value) for cell in sheet[1] if cell.value is not None)
    assert release_files.ISSUES_NOTE_LABEL in row_one
    assert "row 2" in row_one

    # Row 3 onward is still exactly the sound concept row, unpolluted.
    titles = [
        row[list(headers[sheet.title]).index("concept_title")]
        for row in sheet.iter_rows(min_row=3, values_only=True)
        if any(row)
    ]
    assert len(titles) == 1
    assert str(titles[0]).startswith("Solid Shapes")


def test_an_empty_release_still_serves_a_workbook_that_says_it_is_empty(
    db, client,
):
    """R4 one level down: an empty file that explains itself.

    A reviewer opening a zero-row Concept File must not have to guess
    whether the run failed or the chapter is genuinely empty.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [])

    response = client.get(
        f"/build-concepts/uploads/{job.id}/release-bulk-import.xlsx")
    assert response.status_code == 200

    import io

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    row_one = " ".join(
        str(cell.value)
        for cell in workbook[workbook.sheetnames[0]][1]
        if cell.value is not None
    )
    assert "no concept row to export" in row_one
    # Emptiness, and it is named as emptiness rather than as corruption.
    assert release.nothing_to_publish(release.release_payload(job)) is True


# --------------------------------------------------------------------------- #
# 3. Output 04 builds from the sound rows
# --------------------------------------------------------------------------- #

def test_output_04_builds_from_the_sound_rows(db, client):
    """The Post Master File — 200, not 400 (renamed by OD4/D9-Q22).

    [measured at 76c84fb] ``POST /build-assessments/releases/from-job/{id}``
    returned ``400 {"detail":"staged concept row 2 has no topic"}`` for this
    exact payload: one unusable row meant no Master at all, for any concept,
    for any question. Now the snapshot carries the sound row, the Master is
    built and downloadable, and the row that could not be carried is refused
    at the assessment lane's own publication act.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW, _TOPICLESS_ROW])
    authorities, _first = _authorities(db, chapter)

    released = run.run_release_for_job(
        db, job.id, owner_sub=OWNER, authorities=authorities,
        **_decision_context())

    # It BUILT. That is the whole inversion.
    assert released is not None
    directory = Path(released.publication["directory"])
    assert directory.exists()

    # From the SOUND row: the malformed one is in no concept band.
    snapshot = released.concept_snapshot
    titles = [
        concept["concept_title"]
        for topic in snapshot["topics"]
        for concept in topic["concepts"]
    ]
    assert titles == ["Solid Shapes"]

    # And the row that could not be carried is refused, by name, at the
    # assessment lane's publication act — never at the download.
    errors = (released.diagnostics or {}).get("payload_errors") or []
    assert any(release.STAGED_ROW_UNUSABLE in str(row) for row in errors)


def test_the_snapshot_bridge_reports_what_it_could_not_carry(db):
    """``SnapshotError`` is genuine impossibility only (D8.4).

    A row the transient hierarchy skipped rides ``bridge["defects"]``
    instead of raising, so Outputs 02/04 build from the sound rows.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW, _TOPICLESS_ROW])
    payload = release.release_payload(job)

    bridge = snapshot_mod.build(db, job, payload)
    assert [row["concept_title"] for row in bridge["concepts"]] == [
        "Solid Shapes"
    ]
    assert [row["code"] for row in bridge["defects"]] == [
        release.STAGED_ROW_UNUSABLE
    ]

    # The impossibility that STAYS: a release with no frozen seal.
    unsealed = dict(payload, source_document_hash="")
    with pytest.raises(snapshot_mod.SnapshotError, match="source-document"):
        snapshot_mod.build(db, job, unsealed)


# --------------------------------------------------------------------------- #
# 4. The state split
# --------------------------------------------------------------------------- #

def test_emptiness_and_corruption_are_two_questions(db):
    """D8.1, stated as the two functions it produced.

    [measured at 76c84fb] ``structural_defects`` returned "the release
    contains no concept rows to upload" for an empty release, so an
    empty-but-correct one was named *Diagnostic release* — which §4 defines
    as "structural/import integrity failed", and nothing failed.
    """
    empty = {"records": [], "summary": {}}
    assert release.nothing_to_publish(empty) is True
    assert release.structural_defects(empty) == []
    assert release.release_state(empty) == release.READY

    corrupt = {
        "records": [dict(_SOUND_ROW)],
        "summary": {},
        release.STAGED_ROW_DEFECTS_FIELD: [
            {"code": release.STAGED_ROW_UNUSABLE, "position": 2,
             "message": "staged concept row 2 has no topic"},
        ],
    }
    assert release.nothing_to_publish(corrupt) is False
    assert release.structural_defects(corrupt)
    assert release.release_state(corrupt) == release.DIAGNOSTIC_RELEASE

    # Both at once: empty AND corrupt is corrupt.
    both = dict(corrupt, records=[])
    assert release.nothing_to_publish(both) is True
    assert release.structural_defects(both)


def test_the_row_level_gate_fails_closed_on_a_payload_that_records_nothing(
    db, client,
):
    """A missing key is not a clean bill of health.

    D8.5 moved the row-level verdict to STAGING and had the projection read
    what the payload holds. The first cut of that read the key and only the
    key — so a payload with no ``staged_row_defects`` at all reported NO
    defect, and every payload staged before this key existed, restored from
    an export, or built by any caller other than the two staging functions
    has exactly that shape.

    [measured on that cut] a Post release whose ONE record had no ``topic``,
    with the key removed from ``job.question_inventory``:

        structural_defects  : []
        release_state       : ready_with_flags
        upload(lane="post") : database_uploaded True, publication published,
                              created_concept_ids []

    The row was dropped in silence, the receipt reported a success, and
    ``database_uploaded`` latched so no later attempt could find it. At
    76c84fb the identical payload was refused. R4: the silence is the
    defect.

    The recomputation is legal and cannot drift — ``staged_row_defects`` is
    a pure function of the payload calling the SAME ``row_projection_defect``
    the staging pass and the projection both call, and D8.5 forbids a
    projection-time DISCOVERY mutating a frozen payload, not a pure
    recomputation inside the gate.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_TOPICLESS_ROW])

    inventory = copy.deepcopy(dict(job.question_inventory))
    key = release.release_key_for_lane("post")
    inventory[key].pop(release.STAGED_ROW_DEFECTS_FIELD)
    job.question_inventory = inventory
    db.commit()
    db.refresh(job)

    payload = release.release_payload(job)
    assert release.STAGED_ROW_DEFECTS_FIELD not in payload
    assert payload["records"], "the unusable row is still in the payload"

    assert release.structural_defects(payload)
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE
    with pytest.raises(ValueError, match="row 1"):
        publication.upload_release_to_database(
            db, job.id, owner_sub=OWNER, lane="post")

    # Rule E is untouched by the fail-closed read: the evidence still ships.
    for name in _DOWNLOADS:
        response = client.get(f"/build-concepts/uploads/{job.id}/{name}")
        assert response.status_code == 200, name


def test_records_that_are_not_an_array_are_corruption_not_emptiness(db):
    """The one shape ``nothing_to_publish`` could not tell from empty.

    [measured] with only the emptiness question answering, a payload whose
    ``records`` was ``{"a": 1}`` returned no defect, read *Ready with
    flags*, and published as a zero-row success — because
    ``payload.get("records") or []`` iterates a dict's KEYS and none of
    them is a publishable record. ``transient_release_hierarchy`` names
    that exact shape a defect at download time (``staged_records_not_an_
    array``), so the release was calling one thing corrupt in the workbook
    and correct in the state.

    D8.1 split emptiness from corruption. This is corruption.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW])
    inventory = copy.deepcopy(dict(job.question_inventory))
    inventory[release.release_key_for_lane("post")]["records"] = {"a": 1}
    job.question_inventory = inventory
    db.commit()
    db.refresh(job)

    payload = release.release_payload(job)
    assert release.nothing_to_publish(payload) is True, "and yet not empty"
    assert release.structural_defects(payload)
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE
    with pytest.raises(ValueError, match="not an array"):
        publication.upload_release_to_database(
            db, job.id, owner_sub=OWNER, lane="post")


def test_a_zero_row_receipt_never_claims_a_verdict_nobody_made(db):
    """D8.2's success may report what happened, not what was decided.

    [measured] the receipt read "this release is empty and the run recorded
    that its emptiness is correct" on BOTH lanes — but D8.3's verdict is
    Pre-lane only (``_pre_lane_verdict_defects`` returns ``[]`` for every
    other lane), so on the Post lane the job's own record asserted exactly
    the shape-inference D8.3 spends a model call to abolish. A receipt is
    where a reviewer reads what was decided; "a verdict says so" and
    "nothing here said otherwise" are not the same sentence.
    """
    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [])

    result = publication.upload_release_to_database(
        db, job.id, owner_sub=OWNER, lane="post")
    assert result["database_uploaded"] is True
    db.refresh(job)
    assert "records no verdict" in job.detail
    assert "recorded that its emptiness is correct" not in job.detail


def test_every_new_key_has_a_live_reader(db):
    """The producer-with-no-consumer check, run as a test.

    A key written and never read is not a record. S8 closed exactly this
    hole for ``unplaced``; this slice introduces five keys and each one is
    pinned here to a reader that CHANGES ITS ANSWER when the key is
    present — a stronger check than a grep, which a dead read would pass.
    """
    base = {"records": [dict(_SOUND_ROW)], "summary": {}}

    # 1. ``staged_row_defects`` → structural_defects → release_state.
    assert release.structural_defects(base) == []
    with_rows = dict(base, **{
        release.STAGED_ROW_DEFECTS_FIELD: [
            {"code": release.STAGED_ROW_UNUSABLE, "position": 2,
             "message": "staged concept row 2 has no topic"},
        ],
    })
    assert release.structural_defects(with_rows)

    # 2. ``pre_lane_verdict`` → structural_defects, on the PRE lane only.
    empty_pre = {
        "records": [], "summary": {},
        release.RELEASE_LANE_FIELD: release.LANE_PRE,
    }
    assert release.structural_defects(empty_pre), "no verdict blocks"
    decided = dict(empty_pre, **{
        release.PRE_LANE_VERDICT_FIELD: {
            "verdict": release.PRE_LANE_ASSUMES_NOTHING,
            "rationale": "opens from nothing",
        },
    })
    assert release.structural_defects(decided) == []

    # 3. ``concept_snapshot_defects`` → create_release → payload_errors.
    from app.services import assessment_release_service as svc

    chapter = _chapter_with_concepts(db)
    job = _job_with_records(db, chapter, [_SOUND_ROW])
    staged = release.release_payload(job)
    bridge = snapshot_mod.build(db, job, staged)
    made = svc.create_release(
        db,
        chapter_id=chapter.id,
        payload={
            "source_concept_release_sha256": bridge[
                "source_concept_release_sha256"],
            "concept_snapshot": bridge["snapshot"],
            "concept_snapshot_defects": [
                {"code": "snapshot_row_unaddressable",
                 "message": "row 9 has no machine identity"},
            ],
            "source_atoms": [],
            "blueprint_cells": [],
            "candidates": [],
            "groups": [],
            "placements": [],
        },
        owner_sub=OWNER,
    )
    assert any(
        "row 9 has no machine identity" in str(row)
        for row in (made.diagnostics or {}).get("payload_errors") or []
    )

    # 4. ``defects`` on the snapshot bridge → ``assessment_release_run``
    #    stamps it onto the payload as ``concept_snapshot_defects``.
    #
    #    THIS ASSERTION WAS INVERTED. It read ``assert "defects" in
    #    bridge``, which cannot fail: ``assessment_release_snapshot.build``
    #    emits the key unconditionally, so it was exactly the grep this
    #    test's docstring disclaims — and its comment said "read by the
    #    caller above", which was wrong, because item 3 hand-builds
    #    ``concept_snapshot_defects`` and never touches ``bridge``. Nothing
    #    connected the producer to the consumer. It now names both sides of
    #    the join and changes its answer if either key is renamed.
    assert bridge["defects"] == [], "the sound-only fixture carries none"

    unusable = _job_with_records(db, chapter, [_SOUND_ROW, _TOPICLESS_ROW])
    carried = snapshot_mod.build(
        db, unusable, release.release_payload(unusable))["defects"]
    assert [row["code"] for row in carried] == [release.STAGED_ROW_UNUSABLE]

    authorities, _first = _authorities(db, chapter)
    released = run.run_release_for_job(
        db, unusable.id, owner_sub=OWNER, authorities=authorities,
        **_decision_context())
    assert released.payload["concept_snapshot_defects"] == carried

    # 5. ``oversized_cells`` in the renderer's issues manifest.
    from tests.test_mes_dual_output import _snapshot

    over = copy.deepcopy(_snapshot())
    over["candidates"][0]["question"] = "x" * (aw.CELL_LIMIT + 1)
    _master, issues = aw.render_master_file(over)
    assert issues["oversized_cells"][0]["full_value"] == (
        "x" * (aw.CELL_LIMIT + 1)
    )
