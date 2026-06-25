"""Subject-wise Create Workbook: blank template + content modes."""
import io

import openpyxl

from app import bulk_import as bi
from app import models
from app.bulk_import import writer


def _load(content: bytes):
    return openpyxl.load_workbook(io.BytesIO(content))


def _subject_of_first_chapter(db):
    ch = db.query(models.Chapter).filter(models.Chapter.subject != "").first()
    return ch.subject, ch


def test_blank_workbook_has_canonical_headers_and_no_rows(client):
    r = client.get("/data/workbook/new?subject=Mathematics&mode=blank")
    assert r.status_code == 200
    wb = _load(r.content)
    for sheet in (bi.SHEET_OBJECTIVE, bi.SHEET_SUBJECTIVE, bi.SHEET_DESCRIPTIVE):
        assert sheet in wb.sheetnames
    ws = wb[bi.SHEET_OBJECTIVE]
    assert [c.value for c in ws[2]][: len(bi.OBJECTIVE_FIELDS)] == bi.OBJECTIVE_FIELDS
    # No data rows below the two header rows.
    assert ws.max_row == 2


def test_content_workbook_contains_only_that_subject(client, db):
    subject, _ = _subject_of_first_chapter(db)
    other = (
        db.query(models.Chapter)
        .filter(models.Chapter.subject != subject, models.Chapter.subject != "")
        .first()
    )
    r = client.get(f"/data/workbook/new?subject={subject}&mode=content")
    assert r.status_code == 200
    wb = _load(r.content)

    chapter_titles = set()
    for kind, sheet in bi.SHEET_BY_KIND.items():
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if row and row[0]:
                # The title column now carries a tag; compare the clean title.
                chapter_titles.add(bi.strip_title_tag(str(row[0])))
    assert chapter_titles, "expected content rows for the seeded subject"

    in_scope = {
        c.chapter_title for c in
        db.query(models.Chapter).filter(models.Chapter.subject == subject)
    }
    assert chapter_titles <= in_scope
    if other:
        assert other.chapter_title not in chapter_titles


def test_content_workbook_includes_conceptless_question_concepts(client, db):
    """A concept without questions still appears as a concept-catalog row."""
    subject, chapter = _subject_of_first_chapter(db)
    topic = chapter.topics[0]
    bare = models.Concept(
        topic_id=topic.id, concept_title="Bare Concept Without Questions",
        concept_display_name="Bare Concept Without Questions (test)",
        concept_details="Description: x", keywords="",
    )
    db.add(bare)
    db.commit()

    r = client.get(f"/data/workbook/new?subject={subject}&mode=content")
    wb = _load(r.content)
    ws = wb[bi.SHEET_OBJECTIVE]
    concept_col = writer._IDX_CONCEPT_TITLE
    titles = {
        bi.strip_title_tag(writer._cell_str(row, concept_col))
        for row in ws.iter_rows(min_row=3, values_only=True)
    }
    assert "Bare Concept Without Questions" in titles


def test_subject_required(client):
    assert client.get("/data/workbook/new?subject=%20&mode=blank").status_code == 400


def test_filename_includes_scope(client):
    r = client.get("/data/workbook/new?subject=Mathematics&board=CBSE&grade=10&mode=blank")
    assert r.status_code == 200
    assert "bulk_import_Mathematics_CBSE_10.xlsx" in r.headers["content-disposition"]
