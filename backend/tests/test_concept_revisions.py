"""Reviewer instructions, the edits made from them, and where they are stored.

The properties under test are product requirements, not implementation detail:
rounds are unbounded, the delivered workbook is never annotated with review
history, and an instruction survives a failed attempt to apply it.
"""
from __future__ import annotations

import pytest

from app import models
from app.db import SessionLocal
from app.services import concept_revisions, generation_recovery


@pytest.fixture()
def db():
    session = SessionLocal()
    try:
        yield session
    finally:
        session.rollback()
        session.close()


@pytest.fixture()
def job(db):
    chapter = models.Chapter(
        chapter_code="10CBMA_REV01",
        chapter_title="Arithmetic Progressions",
    )
    db.add(chapter)
    db.flush()

    topics = {}
    for order, title in enumerate(
        ["Arithmetic Progressions", "nth Term", "Sum of First n Terms"]
    ):
        topic = models.Topic(
            chapter_id=chapter.id, topic_title=title, source_order=order
        )
        db.add(topic)
        db.flush()
        topics[title] = topic

    concept = models.Concept(
        topic_id=topics["Arithmetic Progressions"].id,
        concept_title="How d governs the behaviour of Sn",
        concept_details="Description: the common difference drives the sum.",
        keywords="d, sum",
        source_order=0,
    )
    db.add(concept)

    upload = models.UploadJob(
        module="build_concepts",
        learning_kind="post",
        filename="ap.mmd",
        status="generated",
        deposit_scope_type="chapter",
        deposit_scope_ids=[chapter.id],
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)
    db.refresh(concept)
    return upload, concept, topics


def _provider(changes, summary="Moved the concept to the later topic."):
    def call(**_kwargs):
        return {"change_summary": summary, "changes": changes}

    return call


def test_rounds_are_unbounded(db, job):
    """A reviewer may keep correcting an output for as long as it takes."""

    upload, _concept, _topics = job

    for expected in range(1, 26):
        revision = concept_revisions.record_instruction(
            db, upload, f"Round {expected}: move it later."
        )
        assert revision.round_number == expected

    history = concept_revisions.list_revisions(db, upload.id)
    assert [row.round_number for row in history] == list(range(1, 26))
    # Oldest first, so the instructions replay in the order they were given.
    assert history[0].instruction.startswith("Round 1:")
    assert history[-1].instruction.startswith("Round 25:")


def test_instruction_is_committed_before_any_edit_is_attempted(db, job):
    upload, _concept, _topics = job

    revision = concept_revisions.record_instruction(
        db, upload, "Put the Sn concept under Sum of First n Terms."
    )

    assert revision.id is not None
    assert revision.status == "pending"
    assert revision.completed_at is None


def test_non_resumable_job_cannot_record_or_apply_a_revision(db, job):
    upload, _concept, _topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "This ordinary round exists before the recovery verdict."
    )
    inventory = dict(upload.question_inventory or {})
    inventory[models.GENERATION_RECOVERY_INVENTORY_KEY] = {
        "resume_allowed": False,
        "recovery_action": "reconvert_new_upload",
        "recovery": "Start a new upload and conversion.",
    }
    upload.question_inventory = inventory
    db.commit()
    db.refresh(upload)

    before = len(concept_revisions.list_revisions(db, upload.id))
    with pytest.raises(generation_recovery.NonResumableRunError):
        concept_revisions.record_instruction(
            db, upload, "This round must never be recorded."
        )
    assert len(concept_revisions.list_revisions(db, upload.id)) == before

    called = False

    def forbidden_provider(**_kwargs):
        nonlocal called
        called = True
        raise AssertionError("blocked revision entered the provider")

    with pytest.raises(generation_recovery.NonResumableRunError):
        concept_revisions.apply_instruction(
            db, upload, revision, provider=forbidden_provider
        )
    assert called is False
    db.refresh(revision)
    assert revision.status == "pending"


def test_a_failed_round_keeps_the_reviewer_words(db, job):
    """The instruction is the research artifact; it outlives the attempt."""

    upload, _concept, _topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "Move it to Sum of First n Terms."
    )

    def exploding(**_kwargs):
        raise RuntimeError("provider unavailable")

    applied = concept_revisions.apply_instruction(
        db, upload, revision, provider=exploding
    )

    assert applied.status == "failed"
    assert "provider unavailable" in applied.error
    assert applied.instruction == "Move it to Sum of First n Terms."
    assert concept_revisions.list_revisions(db, upload.id)[-1].id == revision.id


def test_placement_correction_repoints_the_concept(db, job):
    upload, concept, topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "This belongs under Sum of First n Terms."
    )

    applied = concept_revisions.apply_instruction(
        db,
        upload,
        revision,
        provider=_provider([{
            "concept_id": concept.id,
            "field": "topic",
            "after": "Sum of First n Terms",
            "reason": "Reviewer: this belongs under Sum of First n Terms.",
        }]),
    )

    db.refresh(concept)
    assert applied.status == "applied"
    assert concept.topic_id == topics["Sum of First n Terms"].id
    assert applied.changes == [{
        "concept_id": concept.id,
        "field": "topic",
        "before": "Arithmetic Progressions",
        "after": "Sum of First n Terms",
        "reason": "Reviewer: this belongs under Sum of First n Terms.",
    }]


def test_a_round_that_changes_nothing_is_recorded_as_no_change(db, job):
    """The workbook is rebuilt from concepts, so this leaves it identical."""

    upload, _concept, _topics = job
    revision = concept_revisions.record_instruction(db, upload, "Looks fine.")

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_provider([], summary="Nothing needed changing."),
    )

    assert applied.status == "no_change"
    assert applied.changes == []
    assert applied.change_summary == "Nothing needed changing."


def test_an_unknown_topic_is_declined_rather_than_created(db, job):
    """A reviewer redirects among taught topics; inventing one is generation's job."""

    upload, concept, _topics = job
    original_topic_id = concept.topic_id
    revision = concept_revisions.record_instruction(
        db, upload, "Move it to a brand new topic."
    )

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_provider([{
            "concept_id": concept.id,
            "field": "topic",
            "after": "A Topic This Chapter Never Teaches",
            "reason": "Reviewer asked for a new topic.",
        }]),
    )

    db.refresh(concept)
    assert concept.topic_id == original_topic_id
    assert applied.changes == []
    assert applied.status == "no_change"


def test_edits_outside_the_allowed_fields_are_ignored(db, job):
    upload, concept, _topics = job
    revision = concept_revisions.record_instruction(db, upload, "Change things.")

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_provider([
            {
                "concept_id": concept.id,
                "field": "source_order",
                "after": "99",
                "reason": "not an editable field",
            },
            {
                "concept_id": 10_000_000,
                "field": "concept_title",
                "after": "belongs to another job",
                "reason": "not an offered concept",
            },
        ]),
    )

    assert applied.changes == []


def test_review_history_is_not_written_into_the_delivered_workbook(
    db, job, tmp_path
):
    """Review history lives in the database; the artifact stays untouched.

    The reviewer's words and Aegis's reasons must never reach the deliverable.
    Only the corrected concept content does, so two rounds can be diffed
    without the audit trail moving underneath them.
    """

    from app.bulk_import import writer

    upload, concept, _topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "Rename this to something shorter."
    )
    concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_provider([{
            "concept_id": concept.id,
            "field": "concept_title",
            "after": "How d governs Sn",
            "reason": "Reviewer asked for a shorter title.",
        }]),
    )

    destination = tmp_path / "delivered.xlsx"
    writer.write_workbook(db, destination)

    # An xlsx is a zip, so the cells must be read rather than byte-scanned;
    # a raw search would pass whether or not the history leaked.
    from openpyxl import load_workbook

    book = load_workbook(destination, read_only=True, data_only=True)
    cells = {
        str(value)
        for sheet in book.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    }
    book.close()
    blob = " \n ".join(cells)

    for marker in (
        "Rename this to something shorter",
        "Reviewer asked for a shorter title",
        "round_number",
        "concept_revisions",
    ):
        assert marker not in blob, (
            f"review history leaked into the workbook: {marker!r}"
        )
    # The corrected content itself is delivered, proving the scan can see cells.
    assert "How d governs Sn" in blob


def test_empty_and_oversized_instructions_are_refused(db, job):
    upload, _concept, _topics = job

    with pytest.raises(concept_revisions.RevisionError):
        concept_revisions.record_instruction(db, upload, "   ")
    with pytest.raises(concept_revisions.RevisionError):
        concept_revisions.record_instruction(
            db, upload, "x" * (concept_revisions.MAX_INSTRUCTION_CHARS + 1)
        )


def test_revision_summary_exposes_the_round_for_review(db, job):
    upload, concept, _topics = job
    revision = concept_revisions.record_instruction(db, upload, "Move it later.")
    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_provider([{
            "concept_id": concept.id,
            "field": "topic",
            "after": "Sum of First n Terms",
            "reason": "Reviewer: move it later.",
        }]),
    )

    summary = concept_revisions.revision_summary(applied)
    assert summary["round_number"] == 1
    assert summary["instruction"] == "Move it later."
    assert summary["status"] == "applied"
    assert summary["change_count"] == 1
    assert summary["created_at"]


# --------------------------------------------------------------------------- #
# API surface
# --------------------------------------------------------------------------- #

def test_flagged_placements_are_collected_from_the_run_log():
    """A best-judgement placement is shipped and flagged, never retired."""

    upload = models.UploadJob(generation_log=[
        {
            "type": "log",
            "level": "warning",
            "message": (
                "Phase 3.8 terminal disposition narrow 'How d governs Sn' "
                "under 'Arithmetic Progressions': kept the supported clause"
            ),
        },
        {"type": "log", "level": "info", "message": "an ordinary line"},
        {"type": "progress", "value": 0.81},
    ])

    flags = concept_revisions.flagged_placements(upload)
    assert len(flags) == 1
    assert "How d governs Sn" in flags[0]["message"]


def test_a_round_captures_the_flags_that_were_open_when_it_was_written(db, job):
    upload, _concept, _topics = job
    upload.generation_log = [{
        "type": "log",
        "level": "warning",
        "message": "Phase 3.8 terminal disposition narrow 'Sn' under 'AP'",
    }]
    db.commit()

    revision = concept_revisions.record_instruction(db, upload, "Move it later.")

    assert len(revision.flagged_placements) == 1
    assert "Sn" in revision.flagged_placements[0]["message"]


def test_revision_routes_round_trip(client, db, job, monkeypatch):
    upload, concept, topics = job

    monkeypatch.setattr(
        concept_revisions,
        "_default_provider",
        lambda **_kwargs: {
            "change_summary": "Moved to the later topic.",
            "changes": [{
                "concept_id": concept.id,
                "field": "topic",
                "after": "Sum of First n Terms",
                "reason": "Reviewer: this needs Sn.",
            }],
        },
    )

    posted = client.post(
        f"/build-concepts/uploads/{upload.id}/revisions",
        json={"instruction": "This belongs under Sum of First n Terms."},
    )
    assert posted.status_code == 200, posted.text
    body = posted.json()
    assert body["round_number"] == 1
    assert body["status"] == "applied"
    assert body["change_count"] == 1

    listed = client.get(f"/build-concepts/uploads/{upload.id}/revisions")
    assert listed.status_code == 200, listed.text
    history = listed.json()["revisions"]
    assert [row["round_number"] for row in history] == [1]
    assert history[0]["instruction"] == (
        "This belongs under Sum of First n Terms."
    )

    db.expire_all()
    assert db.get(models.Concept, concept.id).topic_id == (
        topics["Sum of First n Terms"].id
    )


def test_empty_instruction_is_rejected_by_the_route(client, job):
    upload, _concept, _topics = job

    response = client.post(
        f"/build-concepts/uploads/{upload.id}/revisions",
        json={"instruction": "   "},
    )
    assert response.status_code in {400, 422}


def test_a_provider_failure_returns_the_failed_round_not_an_error(
    client, job, monkeypatch
):
    """The reviewer's words are already committed; the round reports the failure."""

    upload, _concept, _topics = job

    def exploding(**_kwargs):
        raise RuntimeError("provider unavailable")

    monkeypatch.setattr(concept_revisions, "_default_provider", exploding)

    response = client.post(
        f"/build-concepts/uploads/{upload.id}/revisions",
        json={"instruction": "Move the Sn concept."},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "failed"
    assert "provider unavailable" in body["error"]
    assert body["instruction"] == "Move the Sn concept."


def test_non_resumable_job_refuses_revision_before_record_or_provider(
    client, db, job, monkeypatch,
):
    """A stale/direct client cannot spend on an incomplete source graph."""

    upload, _concept, _topics = job
    inventory = dict(upload.question_inventory or {})
    inventory[models.GENERATION_RECOVERY_INVENTORY_KEY] = {
        "error": "UnattendedDecisionUnavailable: Q24",
        "message": "Generation did not complete.",
        "resume_allowed": False,
        "recovery_action": "reconvert_new_upload",
        "recovery": "Start a new upload and conversion.",
    }
    upload.question_inventory = inventory
    db.commit()

    def forbidden_record(*_args, **_kwargs):
        raise AssertionError("blocked recovery recorded a revision round")

    def forbidden_apply(*_args, **_kwargs):
        raise AssertionError("blocked recovery entered the revision provider")

    monkeypatch.setattr(
        concept_revisions, "record_instruction", forbidden_record,
    )
    monkeypatch.setattr(
        concept_revisions, "apply_instruction", forbidden_apply,
    )

    response = client.post(
        f"/build-concepts/uploads/{upload.id}/revisions",
        json={"instruction": "Try to repair this output."},
    )

    assert response.status_code == 409
    assert "Start a new upload and conversion" in response.text
    assert concept_revisions.list_revisions(db, upload.id) == []
    listed = client.get(f"/build-concepts/uploads/{upload.id}/revisions")
    assert listed.status_code == 200
    assert listed.json()["revisions"] == []


def test_unknown_job_is_not_found(client):
    response = client.post(
        "/build-concepts/uploads/99999999/revisions",
        json={"instruction": "anything"},
    )
    assert response.status_code == 404


# --------------------------------------------------------------------------- #
# Reviewer-authored additions
# --------------------------------------------------------------------------- #

def _adding_provider(additions, changes=None, summary="Added what was missing."):
    def call(**_kwargs):
        return {
            "change_summary": summary,
            "changes": changes or [],
            "additions": additions,
        }

    return call


def test_reviewer_can_add_a_concept_generation_missed(db, job):
    upload, _concept, topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "You are missing the concept on recognising an AP."
    )

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_adding_provider([{
            "topic": "Arithmetic Progressions",
            "concept_title": "Recognising a real-world sequence as an AP",
            "concept_details": "Description: ladder rungs and taxi fares.",
            "keywords": "AP, recognise",
            "parent_concept": "Arithmetic Progressions",
            "reason": "Reviewer: you are missing the recognition concept.",
        }]),
    )

    assert applied.status == "applied"
    assert applied.change_count == 1
    created = applied.changes[0]
    assert created["field"] == "created"
    assert "Recognising a real-world sequence" in created["after"]

    stored = db.get(models.Concept, created["concept_id"])
    assert stored is not None
    assert stored.topic_id == topics["Arithmetic Progressions"].id
    assert stored.concept_details == "Description: ladder rungs and taxi fares."


def test_an_addition_under_an_unknown_topic_is_declined(db, job):
    """Adding a concept must not invent chapter structure."""

    upload, _concept, _topics = job
    before = db.query(models.Concept).count()
    revision = concept_revisions.record_instruction(db, upload, "Add it.")

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_adding_provider([{
            "topic": "A Topic This Chapter Never Teaches",
            "concept_title": "Orphan",
            "concept_details": "Description: nowhere to live.",
            "reason": "Reviewer asked.",
        }]),
    )

    assert applied.changes == []
    assert db.query(models.Concept).count() == before


def test_additions_and_edits_compose_in_one_round(db, job):
    upload, concept, topics = job
    revision = concept_revisions.record_instruction(
        db, upload, "Move that one later and add the missing basic concept."
    )

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_adding_provider(
            additions=[{
                "topic": "Arithmetic Progressions",
                "concept_title": "d determines whether an AP increases",
                "concept_details": "Description: sign of the common difference.",
                "reason": "Reviewer: keep the foundational idea here.",
            }],
            changes=[{
                "concept_id": concept.id,
                "field": "topic",
                "after": "Sum of First n Terms",
                "reason": "Reviewer: this needs the sum formula.",
            }],
        ),
    )

    db.refresh(concept)
    assert applied.status == "applied"
    assert applied.change_count == 2
    # Rule 3 of the placement rules: the split keeps both sides.
    assert concept.topic_id == topics["Sum of First n Terms"].id
    fields = {row["field"] for row in applied.changes}
    assert fields == {"topic", "created"}


def test_an_added_concept_reaches_the_delivered_workbook(db, job, tmp_path):
    from app.bulk_import import writer
    from openpyxl import load_workbook

    upload, _concept, _topics = job
    revision = concept_revisions.record_instruction(db, upload, "Add it.")
    concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_adding_provider([{
            "topic": "nth Term",
            "concept_title": "Deriving the nth term formula",
            "concept_details": "Description: a plus n minus one d.",
            "reason": "Reviewer: the derivation is missing.",
        }]),
    )

    destination = tmp_path / "delivered.xlsx"
    writer.write_workbook(db, destination)
    book = load_workbook(destination, read_only=True, data_only=True)
    cells = {
        str(value)
        for sheet in book.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    }
    book.close()

    assert "Deriving the nth term formula" in cells
    # The reason the reviewer gave still never reaches the file.
    assert not any("the derivation is missing" in cell for cell in cells)


def test_deletion_is_not_offered(db, job):
    """A reviewer's removal request is recorded, never silently performed."""

    upload, concept, _topics = job
    before = db.query(models.Concept).count()
    revision = concept_revisions.record_instruction(
        db, upload, "Delete the Sn concept entirely."
    )

    applied = concept_revisions.apply_instruction(
        db, upload, revision,
        provider=_adding_provider(
            additions=[],
            changes=[],
            summary="Deletion is not available; the concept was left in place.",
        ),
    )

    assert db.query(models.Concept).count() == before
    assert db.get(models.Concept, concept.id) is not None
    assert applied.status == "no_change"
    assert "Deletion is not available" in applied.change_summary
    assert "delete" not in "".join(
        concept_revisions.EDITABLE_FIELDS + concept_revisions.ADDABLE_FIELDS
    )
