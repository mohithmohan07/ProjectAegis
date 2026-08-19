"""Bulk Import workflow spec: question_text, comma-only multi-values,
cognitive-skill normalization, standard values, validation report."""
import io

import openpyxl

from app import bulk_import as bi
from app import models
from app.bulk_import import layouts
from app.db import _backfill_and_normalize
from app.services import generation


# ---------------------- multi-value parsing (comma only) ---------------------- #

def test_multi_value_comma_only():
    assert bi.split_multi("Remember, Understand") == ["Remember", "Understand"]
    assert bi.split_multi("Pre-test, Post-test, Worksheet") == [
        "Pre-test", "Post-test", "Worksheet"]
    # newline / semicolon / pipe are NOT separators — they stay inside the value.
    assert bi.split_multi("Remember\nUnderstand") == ["Remember\nUnderstand"]
    assert bi.split_multi("Remember; Understand") == ["Remember; Understand"]
    assert bi.split_multi("Remember | Understand") == ["Remember | Understand"]


def test_newlines_preserved_in_plain_text():
    s = "Line one.\nLine two with [Katex] F = ma [/Katex] kept."
    out = bi.to_plain_text(s)
    assert "\n" in out
    assert "F = ma" in out
    assert "[katex]" not in out.lower()


# ------------------------ cognitive skill normalization ----------------------- #

def test_cognitive_normalization_map():
    cases = {
        "Remembering": "Remember", "Understanding": "Understand",
        "Applying": "Apply", "Analysing": "Analyse",
        "Evaluating": "Evaluate", "Creating": "Create",
        "Remember": "Remember",
    }
    for old, new in cases.items():
        assert bi.normalize_cognitive_skills(old) == new
    # Multi-value (comma) normalizes element-wise.
    assert bi.normalize_cognitive_skills("Remembering, Understanding") == \
        "Remember, Understand"


def test_batch_normalizes_old_cognitive_values(client, first_concept):
    s = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [first_concept["id"]],
    }).json()
    batch = client.post(f"/build-assessments/sessions/{s['id']}/batches", json={
        "cognitive_skills": ["Remembering", "Understanding"],
        "difficulty_levels": ["Less"],
        "categories": ["Multiple Choice Question"],
        "question_type": "objective", "num_questions": 1,
    }).json()
    assert batch["cognitive_skills"] == ["Remember", "Understand"]


def test_other_standard_value_normalizers():
    assert bi.normalize_appears_in("Pre/Post-Worksheet/Test") == \
        "Pre-test, Post-test, Worksheet, Test"
    assert bi.normalize_appears_in("Pre-test, Worksheet") == "Pre-test, Worksheet"
    assert bi.normalize_answer_type("Words") == "Phrases"
    assert bi.normalize_answer_type("Equation") == "Equation"


# ------------------------------ question_text -------------------------------- #

def test_generated_questions_populate_question_text(client, first_concept, db):
    s = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [first_concept["id"]],
    }).json()
    categories = {
        "objective": "Multiple Choice Question",
        "subjective": "Short Answer",
        "descriptive": "Long Answer",
    }
    for q_type in ("objective", "subjective", "descriptive"):
        client.post(f"/build-assessments/sessions/{s['id']}/batches", json={
            "cognitive_skills": ["Apply"], "difficulty_levels": ["Moderate"],
            "categories": [categories[q_type]], "question_type": q_type,
            "num_questions": 1,
        })
    from tests.conftest import stream_result
    gen = stream_result(client.post(f"/build-assessments/sessions/{s['id']}/generate"))
    assert gen["created"] == 3
    ids = client.get(f"/build-assessments/sessions/{s['id']}").json()[
        "generated_question_ids"]
    for qid in ids:
        q = db.get(models.Question, qid)
        assert q.question_text, f"question_text empty for {q.sheet_kind}"
        assert "[katex]" not in q.question_text.lower()  # plain text, not markup


def test_export_workbook_carries_question_text_where_the_authority_puts_it(
    client,
):
    """The column is no longer LAST — it sits right after ``question``.

    This test asserted "question_text is the last column" on every sheet. On
    the owner's layout it is not: Objective puts it at column 41 of 67. The
    assertion is re-pointed at the authority rather than deleted, and the
    sheet set below loses the Doc Link tab with it (OWNER RULING OD6).
    """
    r = client.get("/data/export?scope=all")
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    for kind, sheet in bi.SHEET_BY_KIND.items():
        ws = wb[sheet]
        header = [c.value for c in ws[2]]
        assert header == bi.FIELDS_BY_KIND[kind]
        assert header[header.index("question") + 1] == "question_text"
    # Exactly the authority's three sheets, in the authority's order, and NO
    # 'Doc Link <> Each fields ' sheet.
    assert wb.sheetnames == bi.SHEET_ORDER


def test_legacy_import_backfills_question_text(client, db, tmp_path):
    """Template WITHOUT question_text imports safely; backfill = plain question."""
    # Built from the FROZEN registry entry for the pre-question_text template,
    # never from ``writer._write_headers`` (that writer moves in S7). Every
    # sheet carries its real header: the layout gate identifies a WORKBOOK,
    # and a one-column stub sheet is not a layout.
    legacy = layouts.layout("canonical-no-question-text")
    legacy_fields = list(legacy.sheet("objective").fields)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for kind in ("objective", "subjective", "descriptive"):
        sheet_layout = legacy.sheet(kind)
        ws = wb.create_sheet(sheet_layout.sheet_name)
        ws.append(["Chapter"])
        ws.append(list(sheet_layout.fields))
    ws = wb[legacy.sheet("objective").sheet_name]
    row = [""] * len(legacy_fields)
    row[0] = "Legacy QT Chapter (09CBPH_LegacyQT)"
    row[6] = "Legacy QT Topic"
    row[12] = "Legacy QT Concept"
    # Group band now has 8 fields (added group_question_labels) -> question
    # band starts one column later than the old layout.
    row[21] = "09CBPH_LgQT_PL_T01_X Q01"   # concept_question_labels (group label)
    row[26] = "Basic"                       # group_type
    row[29] = "09CBPH_LgQT_PL_T01_X Q01"   # question_label
    row[31] = "Remembering"                 # cognitive — old value, must normalize
    row[37] = "State [katex] v = u + at [/katex] in words."  # question
    row[38] = 1                             # marks
    ws.append(row)
    path = tmp_path / "legacy_qt.xlsx"
    wb.save(path)

    files = {"file": ("legacy_qt.xlsx", io.BytesIO(path.read_bytes()),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    counts = client.post("/data/import", files=files).json()
    assert counts["questions"] == 1

    q = db.query(models.Question).filter_by(
        question_label="09CBPH_LgQT_PL_T01_X Q01").one()
    assert q.cognitive_skills == "Remember"          # normalized on import
    assert q.question_text == "State v = u + at in words."  # backfilled, plain
    assert q.question_appears_in == "Pre-test, Post-test, Worksheet, Test"


def test_db_backfill_for_existing_questions(db, first_concept):
    """Existing DB rows without question_text are backfilled, never overwritten."""
    concept = db.get(models.Concept, first_concept["id"])
    group = concept.groups[0]
    q = models.Question(
        group_id=group.id, sheet_kind="objective",
        question_label="BACKFILL TEST Q01",
        question="What is [katex] E = mc^2 [/katex]?", question_text="",
        cognitive_skills="Evaluating",
        question_appears_in="Pre/Post-Worksheet/Test",
        answers=[{"answer_type": "Words", "answer_content": "x",
                  "correct_answer": "Yes", "answer_weightage": "1"}],
    )
    db.add(q)
    db.commit()
    qid = q.id

    _backfill_and_normalize()

    db.expire_all()
    q2 = db.get(models.Question, qid)
    assert q2.question_text == "What is E = mc^2?"
    assert q2.cognitive_skills == "Evaluate"
    assert q2.question_appears_in == "Pre-test, Post-test, Worksheet, Test"
    assert q2.answers[0]["answer_type"] == "Phrases"

    # Re-running never overwrites an existing value.
    q2.question_text = "Custom evaluator context."
    db.commit()
    _backfill_and_normalize()
    db.expire_all()
    assert db.get(models.Question, qid).question_text == "Custom evaluator context."


# --------------------------- context handling -------------------------------- #

def test_context_attached_to_question_text():
    mmd = (
        "# Source\n\n"
        "Rahul and Meera discuss how shadows form at noon and at dusk.\n\n"
        "Based on the above passage, explain why shadows are shortest at noon."
    )
    records = generation.identify_questions_from_mmd(
        mmd, upload_type="questions", question_type="subjective", live=False)
    target = next(r for r in records if "shortest at noon" in r["question"])
    assert target["question_text"].startswith("Context:")
    assert "Rahul and Meera" in target["question_text"]


# ------------------------------ validation ----------------------------------- #

def test_import_validation_reports_issues(client, tmp_path):
    # Real headers on all three sheets, from the frozen registry entry: the
    # one-column stubs this fixture used to write match no layout, and the
    # reader now refuses a workbook whose geometry it cannot establish.
    current = layouts.layout("canonical-current")
    fields = list(current.sheet("objective").fields)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for kind in ("objective", "subjective", "descriptive"):
        sheet_layout = current.sheet(kind)
        ws = wb.create_sheet(sheet_layout.sheet_name)
        ws.append(["Chapter"])
        ws.append(list(sheet_layout.fields))
    ws = wb[current.sheet("objective").sheet_name]
    sheet_layout = current.sheet("objective")
    row = [""] * len(fields)

    def put(band, name, value):
        # Addressed by band-qualified NAME through the CANONICAL entry this
        # fixture is built from, never by an offset summed from the target's
        # constants: the two layouts no longer agree on a single position.
        row[sheet_layout.column(band, name)] = value

    put("chapter", "chapter_title", "Validation Chapter (09CBPH_Validate)")
    put("topic", "topic_title", "T")
    put("concept", "concept_title", "C")
    put("group", "group_type", "Basic")
    put("group", "concept_question_labels", "09CBPH_Val_PL_T01_X Q01")
    put("question", "question_label", "09CBPH_Val_PL_T01_X Q01")
    put("question", "cognitive_skills", "Memorising")   # unknown -> flagged
    put("question", "level_of_difficulty", "Extreme")   # unknown -> flagged
    put("question", "question", "Compute $$x^2$$ quickly")  # raw $$ -> flagged
    put("question", "marks", "abc")                     # not numeric -> flagged
    ws.append(row)
    path = tmp_path / "validate.xlsx"
    wb.save(path)

    files = {"file": ("validate.xlsx", io.BytesIO(path.read_bytes()),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    counts = client.post("/data/import", files=files).json()
    issues = "\n".join(counts["issues"])
    assert "unknown cognitive skill" in issues
    assert "unknown level_of_difficulty" in issues
    assert "raw math delimiters" in issues
    assert "[Katex]" in issues
    assert "marks not numeric" in issues
