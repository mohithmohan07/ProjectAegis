from __future__ import annotations

import copy
import io
import json
import zipfile

from openpyxl import load_workbook
import pytest

from app import bulk_import as bi
from app import models
from app.services import build_concepts
from app.services import build_concepts_release as release
from app.services import build_concepts_release_contract as release_contract
from app.services import build_concepts_release_files as release_files
from app.services import build_concepts_release_manifest
from app.services import build_concepts_release_publication as publication
from app.services import uploads


def _chapter(db):
    value = db.query(models.Chapter).order_by(models.Chapter.id).first()
    assert value is not None
    return value


def _job(db, *, checkpoint=None):
    chapter = _chapter(db)
    # Contract v2.0 §32.1: a release chapter carries a frozen duration
    # (never estimated). The shared fixture chapter guarantees one so the
    # ``chapter_duration_unregistered`` gate judges only what a test stages;
    # a test that blanks it deliberately restores it afterwards.
    if bi.duration_minutes_cell(chapter.chapter_duration) == "":
        chapter.chapter_duration = "40 minutes"
        db.commit()
        db.refresh(chapter)
    job = models.UploadJob(
        module="build_concepts",
        upload_type="document",
        learning_kind="post",
        source_book="NCERT",
        filename="release-source.mmd",
        mmd_text="## Topic A\nA source paragraph.",
        status="converted",
        deposit_scope_type="chapter",
        deposit_scope_ids=[chapter.id],
        generation_checkpoint=copy.deepcopy(checkpoint or {}),
        question_inventory={"items": [], "stats": {}, "mined_types": []},
        generation_log=[{
            "type": "log",
            "level": "error",
            "message": "TOPOLOGY-CONCEPT-0001 disagreed at BLK-0007",
            "ts": 1.0,
        }],
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job, chapter


def _records():
    return [{
        "topic": "Topic A",
        "parent_concept": "Parent A",
        "concept_title": "Released Concept Alpha",
        "concept_details": "Description: Supported core with one disputed clause.",
        "keywords": "alpha",
        "_semantic_topic_id": "TOPIC-A",
        "_phase32_origin_concept_id": "TOPOLOGY-CONCEPT-0001",
        "_source_block_ids": ["BLK-0007"],
    }]


def _rendered_records():
    """One complete Type owner; other concepts may legitimately own none."""
    rows = _records()
    rows[0]["concept_details"] = (
        "Description: Supported core with one disputed clause."
        " // Types: Type 01: Evidence-based explanation "
        "Case 01: Explain the foundational relationship. "
        "Example 01: Explain the first method. "
        "Case 02: Apply the reusable method later. "
        "Example 02: Apply the later method."
    )
    rows[0][release.RELEASE_ROW_QIDS_FIELD] = [
        "QINV-0001", "QINV-0002",
    ]
    rows[0][release.RELEASE_ROW_ROUTES_FIELD] = [
        "TYPE-0001::CASE-0001",
        "TYPE-0001::CASE-0002",
    ]
    return rows


def _host_manifest(*placements):
    return {
        "placements": {
            qid: {
                "qid": qid,
                "type_id": type_id,
                "case_id": case_id,
                "host_disposition": disposition,
            }
            for qid, type_id, case_id, disposition in placements
        },
    }


def _manifest_only_records():
    records = _rendered_records()
    records[0].pop(release.RELEASE_ROW_QIDS_FIELD)
    records[0].pop(release.RELEASE_ROW_ROUTES_FIELD)
    records[0]["_type_case_qid_host_placement_manifest"] = _host_manifest(
        (
            "QINV-0001", "TYPE-0001", "CASE-0001",
            "type_case_example",
        ),
        (
            "QINV-0002", "TYPE-0001", "CASE-0002",
            "type_case_example",
        ),
    )
    return records


def _concept_without_types():
    return {
        "topic": "Topic A",
        "parent_concept": "Parent A",
        "concept_title": "Released Concept Without A Type",
        "concept_details": "Description: A valid concept with no assigned Type.",
        "keywords": "untyped",
        "_semantic_topic_id": "TOPIC-A",
        "_phase32_origin_concept_id": "TOPOLOGY-CONCEPT-0002",
        "_source_block_ids": ["BLK-0008"],
    }


def _inventory():
    return {
        "items": [
            {"qid": "QINV-0001", "raw_task": "Explain the first method."},
            {"qid": "QINV-0002", "raw_task": "Apply the later method."},
        ],
        "stats": {"items": 2},
    }


def _mined_types(*, duplicate=False):
    second_qid = "QINV-0001" if duplicate else "QINV-0002"
    return {
        "types": [{
            "type_id": "TYPE-0001",
            "type_title": "Evidence-based explanation",
            "type_definition": (
                "Explain a source relationship by connecting evidence to the "
                "required historical framework."
            ),
            "owner_topic_ids": ["TOPIC-A", "TOPIC-B"],
            "case_prompts": [
                {
                    "case_id": "CASE-0001",
                    "case_definition": "Explain the foundational relationship.",
                    "owner_topic_ids": ["TOPIC-A"],
                    "source_question_ids": ["QINV-0001"],
                    "examples": [{
                        "source_question_id": "QINV-0001",
                        "prompt": "Explain the first method.",
                    }],
                },
                {
                    "case_id": "CASE-0002",
                    "case_definition": (
                        "Apply the same reusable method after the later topic "
                        "becomes necessary."
                    ),
                    "owner_topic_ids": ["TOPIC-B"],
                    "source_question_ids": [second_qid],
                    "examples": [{
                        "source_question_id": second_qid,
                        "prompt": "Apply the later method.",
                    }],
                },
            ],
        }],
    }


def _activity_mined_types():
    mined = _mined_types()
    mined["types"][0]["case_prompts"][1]["is_activity"] = True
    return mined


def _pending():
    return {
        "decision_id": "phase32-release-test",
        "context_hash": "a" * 64,
        "kind": "phase32_concept_blueprint_semantic_conflict",
        "phase": "3.2",
        "conflict": "The concept may require the later topic framework.",
        "diagnosis": "Autonomous review could not certify one topology action.",
        "decision_question": "Move, refine or split?",
        "item": {
            "unit_id": "TOPOLOGY-CONCEPT-0001",
            "topic": "Topic A",
            "qids": ["QINV-0001"],
        },
        "candidates": [{
            "target_id": "TOPOLOGY-CONCEPT-0001",
            "title": "Released Concept Alpha",
            "topic": "Topic A",
            "source_block_ids": ["BLK-0007"],
        }],
        "evidence": [{
            "evidence_id": "BLK-0007",
            "label": "Source block",
            "text": "The relevant source relationship.",
            "page": "9",
        }],
        "options": [],
    }


def test_one_type_can_route_cases_to_different_topics_without_qid_duplication():
    rows, issues, routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )

    assert [row["row_kind"] for row in rows] == [
        "type", "case", "example", "case", "example"
    ]
    assert rows[1]["owner_topic_ids"] == ["TOPIC-A"]
    assert rows[3]["owner_topic_ids"] == ["TOPIC-B"]
    assert rows[2]["example_qid"] == "QINV-0001"
    assert rows[4]["example_qid"] == "QINV-0002"
    assert routes["QINV-0001"][0]["case_id"] == "CASE-0001"
    assert routes["QINV-0002"][0]["case_id"] == "CASE-0002"
    assert not [issue for issue in issues if issue["severity"] == "error"]


def test_duplicate_qid_is_released_as_an_audit_error():
    _rows, issues, _routes = release.audit_type_cases(
        _mined_types(duplicate=True), _inventory()
    )
    codes = [issue["code"] for issue in issues]
    assert "duplicate_qid_assignment" in codes
    assert "unassigned_inventory_qid" in codes


@pytest.mark.parametrize("missing_piece", ["route_marker", "types_section"])
def test_rendered_example_identity_requires_route_marker_and_visible_types(
    missing_piece,
):
    records = _rendered_records()
    if missing_piece == "route_marker":
        records[0].pop(release.RELEASE_ROW_ROUTES_FIELD)
    else:
        records[0]["concept_details"] = "Description: The Types body is absent."
    type_case_rows, _issues, _routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )

    issues = release._rendered_type_case_route_issues(
        records, type_case_rows
    )

    assert len(issues) == 1
    assert issues[0]["code"] == "unrendered_type_case_qid"
    assert issues[0]["qids"] == ["QINV-0001", "QINV-0002"]


def test_partial_or_mismatched_routes_never_satisfy_another_qid():
    records = _rendered_records()
    records[0][release.RELEASE_ROW_ROUTES_FIELD] = [
        {"type_id": "TYPE-0001", "case_id": "CASE-0001"},
        {"type_id": "TYPE-9999", "case_id": "CASE-0002"},
    ]
    type_case_rows, _issues, routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )

    issues = release._rendered_type_case_route_issues(
        records, type_case_rows
    )
    annotated = release._annotate_records(records, issues, routes)

    assert len(issues) == 1
    assert issues[0]["qids"] == ["QINV-0002"]
    assert issues[0]["details"]["rendered_routes"] == [{
        "example_qid": "QINV-0001",
        "type_id": "TYPE-0001",
        "case_id": "CASE-0001",
        "is_activity": False,
    }]
    assert [
        route["example_qid"]
        for route in annotated[0][release.RELEASE_ROW_ROUTES_FIELD]
    ] == ["QINV-0001"]


def test_split_route_alias_preserves_the_exact_original_case_identity():
    records = _rendered_records()
    records[0][release.RELEASE_ROW_ROUTES_FIELD] = [
        "TYPE-0001::CASE-0001",
        "TYPE-0001::CASE-0099::split-of:CASE-0002",
    ]
    type_case_rows, _issues, routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )

    assert release._rendered_type_case_route_issues(
        records, type_case_rows
    ) == []
    annotated = release._annotate_records(records, [], routes)
    second = next(
        route
        for route in annotated[0][release.RELEASE_ROW_ROUTES_FIELD]
        if route["example_qid"] == "QINV-0002"
    )
    assert second["case_id"] == "CASE-0002"
    assert second["rendered_case_id"] == "CASE-0099"
    assert second["split_of_case_id"] == "CASE-0002"


@pytest.mark.parametrize(
    "proof_field",
    ["_activity_hub_qids", "_aegis_hub_placements", "manifest"],
)
def test_activity_examples_require_exact_hub_evidence_and_a_nonempty_body(
    proof_field,
):
    records = _rendered_records()
    record = records[0]
    record["concept_details"] = (
        "Description: Supported core. // Activity/Info Hub: "
        "Apply the later method. // Types: Type 01: Evidence-based "
        "explanation Case 01: Explain the foundational relationship. "
        "Example 01: Explain the first method."
    )
    record[release.RELEASE_ROW_QIDS_FIELD] = ["QINV-0001"]
    record[release.RELEASE_ROW_ROUTES_FIELD] = [
        "TYPE-0001::CASE-0001"
    ]
    if proof_field == "manifest":
        record["_type_case_qid_host_placement_manifest"] = _host_manifest(
            (
                "QINV-0002", "TYPE-0001", "CASE-0002",
                "activity_info_hub",
            ),
        )
    else:
        record[proof_field] = ["QINV-0002"]
    type_case_rows, _issues, routes = release.audit_type_cases(
        _activity_mined_types(), _inventory()
    )
    activity_row = next(
        row for row in type_case_rows
        if row.get("example_qid") == "QINV-0002"
    )

    assert activity_row["is_activity"] is True
    assert release._rendered_type_case_route_issues(
        records, type_case_rows
    ) == []
    annotated = release._annotate_records(records, [], routes)
    annotated_qids = {
        route["example_qid"]
        for route in annotated[0][release.RELEASE_ROW_ROUTES_FIELD]
    }
    assert "QINV-0001" in annotated_qids
    assert ("QINV-0002" in annotated_qids) is (proof_field == "manifest")

    no_body = copy.deepcopy(records)
    no_body[0]["concept_details"] = (
        "Description: Supported core. // Types: Type 01: Evidence-based "
        "explanation Case 01: Explain the foundational relationship. "
        "Example 01: Explain the first method."
    )
    missing = release._rendered_type_case_route_issues(
        no_body, type_case_rows
    )
    assert len(missing) == 1
    assert missing[0]["qids"] == ["QINV-0002"]


def test_a_mismatched_activity_manifest_never_claims_the_hub_qid():
    records = _rendered_records()
    record = records[0]
    record["concept_details"] = (
        "Description: Supported core. // Activity/Info Hub: "
        "Apply the later method. // Types: Type 01: Evidence-based "
        "explanation Case 01: Explain the foundational relationship. "
        "Example 01: Explain the first method."
    )
    record[release.RELEASE_ROW_QIDS_FIELD] = ["QINV-0001"]
    record[release.RELEASE_ROW_ROUTES_FIELD] = [
        "TYPE-0001::CASE-0001"
    ]
    record["_type_case_qid_host_placement_manifest"] = _host_manifest(
        (
            "QINV-0002", "TYPE-0001", "CASE-WRONG",
            "activity_info_hub",
        ),
    )
    type_case_rows, _issues, _routes = release.audit_type_cases(
        _activity_mined_types(), _inventory()
    )

    missing = release._rendered_type_case_route_issues(
        records, type_case_rows
    )

    assert len(missing) == 1
    assert missing[0]["qids"] == ["QINV-0002"]


def test_description_only_rows_with_mined_types_are_diagnostic_but_downloadable(
    db,
):
    """A durable pre-Type checkpoint may ship, but can never publish as done."""
    job, chapter = _job(db)

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        error=RuntimeError("host pass failed after the durable checkpoint"),
    )

    payload = release.release_payload(job)
    assert payload is not None
    missing = [
        issue for issue in payload["issues"]
        if issue["code"] == release.TYPE_CASE_ROUTE_SET_EMPTY
    ]
    assert len(missing) == 1
    assert missing[0]["qids"] == ["QINV-0001", "QINV-0002"]
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE

    # The publication gate recomputes the identity defect from the staged
    # material, so clearing the display ledger cannot open the write.
    payload_without_issues = copy.deepcopy(payload)
    payload_without_issues["issues"] = []
    assert any(
        release.TYPE_CASE_ROUTE_SET_EMPTY in defect
        for defect in release.structural_defects(payload_without_issues)
    )
    with pytest.raises(ValueError, match=release.TYPE_CASE_ROUTE_SET_EMPTY):
        publication.upload_release_to_database(db, job.id)

    # Rule E: Diagnostic blocks only the DB act; every evidence download
    # remains buildable for review and recovery.
    assert release_files.build_release_workbook(job)
    assert release_files.build_release_bulk_import_workbook(db, job)
    assert release_files.build_diagnostics_zip(job)


def test_a_whole_concept_file_with_no_type_question_routes_is_diagnostic(
    db,
):
    """Source QIDs plus concept rows need at least one exact route set."""

    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types={"types": []},
        reason="durable description-only checkpoint",
    )

    payload = release.release_payload(job)
    assert payload is not None
    route_set_issues = [
        issue for issue in payload["issues"]
        if issue["code"] == release.TYPE_CASE_ROUTE_SET_EMPTY
    ]
    assert len(route_set_issues) == 1
    assert route_set_issues[0]["qids"] == ["QINV-0001", "QINV-0002"]
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE

    # Publication recomputes from the staged material. Deleting the display
    # issue therefore cannot turn the same description-only file publishable.
    payload_without_issues = copy.deepcopy(payload)
    payload_without_issues["issues"] = []
    assert any(
        release.TYPE_CASE_ROUTE_SET_EMPTY in defect
        for defect in release.structural_defects(payload_without_issues)
    )
    with pytest.raises(ValueError, match=release.TYPE_CASE_ROUTE_SET_EMPTY):
        publication.upload_release_to_database(db, job.id)

    # Diagnostic is a publication verdict, never a download embargo.
    assert release_files.build_release_workbook(job)
    assert release_files.build_release_bulk_import_workbook(db, job)
    assert release_files.build_diagnostics_zip(job)


def test_questionless_concept_file_does_not_require_a_type_route(db):
    """Q14 stays intact: the gate is chapter-wide, not per concept."""

    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory={"items": [], "stats": {"items": 0}},
        mined_types={"types": []},
        reason="questionless concept file",
    )

    payload = release.release_payload(job)
    assert payload is not None
    assert all(
        issue["code"] != release.TYPE_CASE_ROUTE_SET_EMPTY
        for issue in payload["issues"]
    )
    assert all(
        release.TYPE_CASE_ROUTE_SET_EMPTY not in defect
        for defect in release.structural_defects(payload)
    )


def test_manifest_metadata_cannot_replace_visible_cases_and_questions(db):
    """Exact audit fields do not make a placeholder Types body visible."""

    job, chapter = _job(db)
    records = _manifest_only_records()
    records[0]["concept_details"] = (
        "Description: Supported core. // Types: Type 01: Placeholder only. "
        "Case 01: Placeholder only. Example 01:"
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=records,
        inventory=_inventory(),
        mined_types=_mined_types(),
        reason="manual edit retained audit metadata only",
    )

    payload = release.release_payload(job)
    assert payload is not None
    issues = [
        issue for issue in payload["issues"]
        if issue["code"] == release.TYPE_CASE_ROUTE_SET_EMPTY
    ]
    assert len(issues) == 1
    assert issues[0]["qids"] == ["QINV-0001", "QINV-0002"]
    assert issues[0]["details"]["missing_visible_shapes"] == [
        "a Case section containing an Example"
    ]
    assert release.release_state(payload) == release.DIAGNOSTIC_RELEASE

    payload_without_issues = copy.deepcopy(payload)
    payload_without_issues["issues"] = []
    assert any(
        release.TYPE_CASE_ROUTE_SET_EMPTY in defect
        for defect in release.structural_defects(payload_without_issues)
    )
    with pytest.raises(ValueError, match=release.TYPE_CASE_ROUTE_SET_EMPTY):
        publication.upload_release_to_database(db, job.id)


def test_manifest_only_routes_keep_staging_and_publication_in_parity(
    db, monkeypatch,
):
    job, chapter = _job(db)
    records = _manifest_only_records()
    type_case_rows, _issues, _routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )
    assert release._rendered_type_case_route_issues(
        records, type_case_rows
    ) == []

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=records,
        inventory=_inventory(),
        mined_types=_mined_types(),
    )

    payload = release.release_payload(job)
    assert payload is not None
    staged_routes = payload["records"][0][release.RELEASE_ROW_ROUTES_FIELD]
    assert [route["example_qid"] for route in staged_routes] == [
        "QINV-0001", "QINV-0002",
    ]
    assert all(route["route_evidence"] == "manifest" for route in staged_routes)
    assert release._rendered_type_case_route_issues(
        payload["records"], payload["type_case_rows"]
    ) == []
    assert release.structural_defects(payload) == []
    assert release.release_state(payload) == release.READY

    def commit_only(database, _path, ids):
        database.commit()
        return {
            "written": len(ids),
            "sources_updated": 0,
            "publication_status": "published",
        }

    monkeypatch.setattr(
        build_concepts,
        "_commit_and_publish_concept_workbook",
        commit_only,
    )
    published = publication.upload_release_to_database(db, job.id)
    assert published["database_uploaded"] is True
    created_ids = published["created_concept_ids"]
    assert created_ids
    db.query(models.Concept).filter(
        models.Concept.id.in_(created_ids)
    ).delete(synchronize_session=False)
    db.commit()


def test_a_concept_without_types_is_valid_when_all_examples_render_on_the_owner(
    db,
):
    """Q14 does not impose a per-concept Type minimum."""
    job, chapter = _job(db)
    records = [*_rendered_records(), _concept_without_types()]

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=records,
        inventory=_inventory(),
        mined_types=_mined_types(),
    )

    payload = release.release_payload(job)
    assert payload is not None
    assert not [
        issue for issue in payload["issues"]
        if issue["code"] == "unrendered_type_case_qid"
    ]
    assert not any(
        "unrendered_type_case_qid" in defect
        for defect in release.structural_defects(payload)
    )
    untyped = next(
        row for row in payload["records"]
        if row["concept_title"] == "Released Concept Without A Type"
    )
    assert "// Types:" not in untyped["concept_details"]


def test_stage_release_clears_manual_pause_and_highlights_the_affected_row(db):
    job, chapter = _job(db)

    result = release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        pending_decision=_pending(),
        reason="Release the newest durable rows instead of asking the user.",
    )

    db.refresh(job)
    payload = release.release_payload(job)
    assert payload is not None
    assert result["status"] == "released"
    assert result["database_uploaded"] is False
    assert job.status == "released"
    assert job.result_ids == []
    assert job.pending_decision is None
    assert payload["records"][0][release.RELEASE_ROW_STATUS_FIELD] == (
        "released_with_errors"
    )
    assert payload["records"][0][release.RELEASE_ROW_ERRORS_FIELD]
    assert payload["issues"][0]["unit_id"] == "TOPOLOGY-CONCEPT-0001"
    assert payload["issues"][0]["block_ids"] == ["BLK-0007"]


def test_stage_release_records_unowned_examples_on_its_own_ledger(db):
    # Q13/R4: a public Example with no exact inventory owner is recorded
    # at STAGING, beside the QC audit, so every exit that stages rows
    # carries the record — the log-only "closed-world validation remains
    # blocked" defect class cannot recur. Dry mode records the finding
    # unadjudicated rather than dropping it because the judge is off.
    from app.services import concept_example_ownership as ownership

    job, chapter = _job(db)
    records = _records()
    records[0]["concept_details"] += (
        " // Types: Type 01: Alpha Case 01: c "
        "Example 01: An invented Example no inventory item owns."
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=records,
        inventory=_inventory(),
        mined_types=_mined_types(),
    )

    payload = release.release_payload(job)
    assert payload is not None
    recorded = [
        issue for issue in payload["issues"]
        if issue["code"] == ownership.UNOWNED_EXAMPLES_ISSUE_CODE
    ]
    assert len(recorded) == 1
    assert recorded[0]["severity"] == "warning"
    assert recorded[0]["details"]["adjudicated"] is False
    assert recorded[0]["details"]["verdicts"][0]["example_text"].startswith(
        "an invented example"
    )
    # Chapter-level record: no qid anchor, so the row-annotation pass
    # cannot stamp released_with_errors onto rows whose Examples were
    # ruled legitimate while the unowned ones (naming no qid) go unmarked.
    assert recorded[0]["qids"] == []


def test_inventory_fallback_reaches_the_job_and_strips_release_slots(db):
    # _newest_checkpoint_material returns {} when the checkpoint carries
    # no inventory, which used to satisfy the isinstance test and leave
    # the job.question_inventory fallback unreachable — the ownership
    # scan then ran against an EMPTY inventory and flagged every rendered
    # Example as unowned. The empty snapshot now falls through to the
    # job's stored inventory, minus the lane release slots (a prior
    # staged release must never nest inside a new payload's inventory).
    job, _chapter = _job(db)
    job.question_inventory = {
        "items": [{"qid": "QINV-0001", "raw_task": "A real task."}],
        "stats": {"items": 1},
        release.RELEASE_KEY: {"records": [], "issues": []},
        release.PRE_RELEASE_KEY: {"records": [], "issues": []},
    }

    fallback = release._job_inventory_fallback(job)

    assert fallback["items"][0]["qid"] == "QINV-0001"
    assert release.RELEASE_KEY not in fallback
    assert release.PRE_RELEASE_KEY not in fallback


def test_lane_slots_are_stripped_from_any_inventory_source(db):
    # The capture fallback can hand job.question_inventory in VERBATIM as
    # the explicit inventory= argument; the lane release slots must be
    # stripped from the resolved value whatever branch supplied it, or a
    # prior staged release nests inside the new payload and the durable
    # row grows every staging cycle.
    job, chapter = _job(db)
    tainted = _inventory()
    tainted[release.RELEASE_KEY] = {"records": [], "issues": []}
    tainted[release.PRE_RELEASE_KEY] = {"records": [], "issues": []}

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=tainted,
        mined_types=_mined_types(),
    )

    payload = release.release_payload(job)
    assert payload is not None
    staged_inventory = payload["question_task_inventory"]
    assert release.RELEASE_KEY not in staged_inventory
    assert release.PRE_RELEASE_KEY not in staged_inventory
    assert staged_inventory["items"]


def test_stage_release_hands_the_judge_the_captured_checkpoint_identity(
    db, monkeypatch,
):
    # On the clean exit generate_post_learning has already cleared
    # job.generation_checkpoint, so the live checkpoint_target_identity
    # property is {} — the judge's chapter context must come from the
    # checkpoint THIS staging resolved (the wrapper hands the captured
    # one in), or the recorded decision loses its metadata on exactly
    # the most common path.
    from app.services import concept_example_ownership as ownership

    seen: list[dict] = []
    monkeypatch.setattr(
        ownership, "adjudication_issue",
        lambda _records, _inventory, *, meta, job_id, allow_live=True: (
            seen.append(dict(meta)) or None
        ),
    )
    job, chapter = _job(db)
    assert job.checkpoint_target_identity == {}

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        checkpoint={"target_identity": {
            "board": "cbse", "grade": "6", "subject": "english",
            "unit": "poem", "chapter_title": "the school bell",
            "chapter_code": "ch-2",
        }},
    )

    assert seen and seen[0]["board"] == "cbse"
    assert seen[0]["chapter_title"] == "the school bell"


def test_release_workbook_orders_type_case_example_and_marks_errors(db):
    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_rendered_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        pending_decision=_pending(),
    )

    workbook = load_workbook(io.BytesIO(release_files.build_release_workbook(job)))
    assert workbook.sheetnames == [
        "Released Concepts",
        "Type Case Routing",
        "Release Issues",
        "Release Manifest",
    ]
    concepts = workbook["Released Concepts"]
    assert concepts.cell(1, 2).value == "Release Status"
    assert concepts.cell(1, 3).value == "Errors / Warnings"
    assert concepts.cell(2, 2).value == "released_with_errors"
    assert concepts.cell(2, 1).fill.fgColor.rgb[-6:] == "F4CCCC"

    routes = workbook["Type Case Routing"]
    assert [routes.cell(row, 1).value for row in range(2, 7)] == [
        "type", "case", "example", "case", "example"
    ]
    assert routes.cell(3, 6).value == "Explain the foundational relationship."
    assert routes.cell(4, 9).value == "QINV-0001"


def test_diagnostics_archive_keeps_log_checkpoint_source_evidence_and_blks(db):
    checkpoint = {
        "checkpoints": [{
            "stage": "pre_type_assignment",
            "progress": 0.81,
            "records": _records(),
        }]
    }
    job, chapter = _job(db, checkpoint=checkpoint)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        pending_decision=_pending(),
        checkpoint=checkpoint,
    )

    with zipfile.ZipFile(io.BytesIO(release_files.build_diagnostics_zip(job))) as archive:
        names = set(archive.namelist())
        assert "release/released_concepts.xlsx" in names
        assert "release/release_payload.json" in names
        assert "context/generation_log.json" in names
        assert "context/generation_checkpoint.json" in names
        assert "context/question_inventory.json" in names
        assert "context/source_evidence.json" in names
        assert "context/blks.json" in names
        assert "source/converted_source.mmd" in names
        blocks = json.loads(archive.read("context/blks.json"))
        assert any(
            row.get("block_id") == "BLK-0007"
            for row in blocks["blocks"]
        )


def test_post_capture_failure_releases_captured_rows_not_an_empty_checkpoint(db):
    job, chapter = _job(db)

    def fails_after_capture(_db, _job_id, _chapter_id, **_kwargs):
        release_contract._RELEASE_CAPTURE.set({
            "records": _records(),
            "inventory": _inventory(),
            "mined_types": _mined_types(),
            "final_grounding_certificate": {"version": "test"},
            "checkpoint": {},
        })
        raise RuntimeError("late publication preparation failure")

    wrapped = release_contract._wrap_generation(fails_after_capture)
    result = wrapped(db, job.id, chapter.id)

    payload = release.release_payload(job)
    assert result["row_count"] == 1
    assert payload is not None
    assert payload["records"][0]["concept_title"] == "Released Concept Alpha"
    assert any(
        issue["code"] == "RuntimeError" for issue in payload["issues"]
    )


def _validated_cache_rows():
    return [
        {
            "topic": "Topic A",
            "parent_concept": "Parent A",
            "concept_title": "Validated Cached Concept",
            "concept_details": (
                "Description: The complete validated claim.\n"
                "Achieving Mastery: Explain it. // Misconception/ Error "
                "Analysis: Misconceptions: A learner merges two ideas.; "
                "Error Analysis: The learner treats them as one step."
            ),
            "keywords": "validated",
            "_semantic_topic_id": "TOPIC-A",
            "_semantic_graph_contract": "CONTRACT-1",
            "_source_block_ids": ["BLK-0007"],
            # T10-6 (S11): the tie-break reads the RECORDED Q1 allotment
            # marker, not a "misconception" substring — a validated row
            # with complete analysis carries the marker by construction.
            "_aegis_analysis_allotments": ["LA-0001"],
        },
        {
            "topic": "Topic A",
            "parent_concept": "Parent A",
            "concept_title": "Culmination - Topic A",
            "concept_details": "Description: Recap of the validated claim.",
            "keywords": "",
            "_semantic_topic_id": "TOPIC-A",
            "_semantic_graph_contract": "CONTRACT-1",
            "_source_block_ids": ["BLK-0007"],
        },
    ]


def _write_validated_cache(tmp_path, rows):
    from app.services import canonical_source_phase3 as phase3

    (tmp_path / release._FINAL_TOPOLOGY_ARTIFACT).write_text(
        json.dumps({
            "records": rows,
            "records_sha256": phase3._sha256_json(rows),
            "source_contract_hash": "CONTRACT-1",
        }),
        encoding="utf-8",
    )


def test_failure_release_prefers_validated_topology_cache(
    db, tmp_path, monkeypatch,
):
    """Job 23 released 51 stale checkpoint rows without learner analysis
    while a validated topology with complete learner analysis sat cached."""
    job, chapter = _job(db)
    stale = [{
        "topic": "Topic A",
        "parent_concept": "Parent A",
        "concept_title": "Stale Checkpoint Concept",
        "concept_details": "Description: No learner analysis yet.",
        "keywords": "stale",
        "_semantic_topic_id": "TOPIC-A",
        "_semantic_graph_contract": "CONTRACT-1",
        "_source_block_ids": ["BLK-0007"],
    }]
    monkeypatch.setattr(
        release.generation,
        "_newest_compatible_concept_checkpoint",
        lambda _envelope: {"records": copy.deepcopy(stale)},
    )
    _write_validated_cache(tmp_path, _validated_cache_rows())
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    result = release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        error=RuntimeError("host mutation contract failed closed"),
        reason="Generation failed after creating a durable checkpoint.",
    )

    payload = release.release_payload(job)
    assert result["status"] == release.RELEASE_STATUS
    titles = [row["concept_title"] for row in payload["records"]]
    assert "Validated Cached Concept" in titles
    assert "Stale Checkpoint Concept" not in titles
    assert any(
        issue["code"] == "release_rows_upgraded_from_validated_cache"
        for issue in payload["issues"]
    )


def test_failure_release_reads_the_rewritten_settled_rows_snapshot(
    db, tmp_path, monkeypatch,
):
    """Job 26: the new path's settled rows must reach a failure release."""
    from app.services import canonical_source_phase3 as phase3_core

    job, chapter = _job(db)
    monkeypatch.setattr(
        release.generation,
        "_newest_compatible_concept_checkpoint",
        lambda _envelope: {"records": []},
    )
    rows = _validated_cache_rows()
    (tmp_path / release._SETTLED_ROWS_ARTIFACT).write_text(
        json.dumps({
            "records": rows,
            "records_sha256": phase3_core._sha256_json(rows),
            "source_contract_hash": "CONTRACT-1",
        }),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        error=RuntimeError("host.units decision failed closed"),
        reason="Generation failed after creating a durable checkpoint.",
    )

    payload = release.release_payload(job)
    titles = [row["concept_title"] for row in payload["records"]]
    assert "Validated Cached Concept" in titles


def _allotted_row(title, allotments):
    return {
        "topic": "Topic A",
        "parent_concept": "Parent A",
        "concept_title": title,
        "concept_details": "Description: A row carrying recorded allotments.",
        "keywords": "allotted",
        "_semantic_topic_id": "TOPIC-A",
        "_semantic_graph_contract": "CONTRACT-1",
        "_source_block_ids": ["BLK-0007"],
        "_aegis_analysis_allotments": list(allotments),
    }


def test_a_stale_split_cache_never_beats_the_current_merged_rows(
    db, tmp_path, monkeypatch,
):
    """Audit finding 12: the selection counted ROWS carrying an allotment
    marker, so a stale split cache (two rows carrying one id each)
    out-counted the current merged topology (one row carrying both ids)
    while recording exactly the same allotment identities — row
    partitioning deciding the swap. The exact identity SETS are compared
    now: equal sets keep the current rows, no advisory needed."""
    job, chapter = _job(db)
    current = [_allotted_row("Merged Current Concept",
                             ["LA-0001", "LA-0002"])]
    monkeypatch.setattr(
        release.generation,
        "_newest_compatible_concept_checkpoint",
        lambda _envelope: {"records": copy.deepcopy(current)},
    )
    _write_validated_cache(tmp_path, [
        _allotted_row("Split Cached Concept One", ["LA-0001"]),
        _allotted_row("Split Cached Concept Two", ["LA-0002"]),
    ])
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        error=RuntimeError("failed after creating a durable checkpoint"),
        reason="Generation failed after creating a durable checkpoint.",
    )

    payload = release.release_payload(job)
    titles = [row["concept_title"] for row in payload["records"]]
    assert titles == ["Merged Current Concept"]
    assert not any(
        issue["code"] == "release_rows_upgraded_from_validated_cache"
        for issue in payload["issues"]
    )
    assert not any(
        issue["code"] == release.TOPOLOGY_ALLOTMENT_AMBIGUITY
        for issue in payload["issues"]
    )


def test_incomparable_allotment_sets_keep_current_and_record_the_ambiguity(
    db, tmp_path, monkeypatch,
):
    """Audit finding 12: when each side records allotment identities the
    other lacks, neither is a superset — the current rows ship and the
    ambiguity is a NAMED advisory issue, never a row-count preference and
    never a gate."""
    job, chapter = _job(db)
    current = [_allotted_row("Merged Current Concept",
                             ["LA-0001", "LA-0002"])]
    monkeypatch.setattr(
        release.generation,
        "_newest_compatible_concept_checkpoint",
        lambda _envelope: {"records": copy.deepcopy(current)},
    )
    _write_validated_cache(tmp_path, [
        _allotted_row("Sideways Cached Concept", ["LA-0001", "LA-0003"]),
    ])
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        error=RuntimeError("failed after creating a durable checkpoint"),
        reason="Generation failed after creating a durable checkpoint.",
    )

    payload = release.release_payload(job)
    titles = [row["concept_title"] for row in payload["records"]]
    assert titles == ["Merged Current Concept"]
    advisories = [
        issue for issue in payload["issues"]
        if issue["code"] == release.TOPOLOGY_ALLOTMENT_AMBIGUITY
    ]
    assert len(advisories) == 1
    assert advisories[0]["severity"] == "warning"
    assert "incomparable" in advisories[0]["message"]
    # Advisory, never a gate: the ambiguity does not block the write.
    assert not any(
        "topology" in defect
        for defect in release.structural_defects(payload)
    )


def test_captured_final_rows_are_never_overridden_by_cache(
    db, tmp_path, monkeypatch,
):
    job, chapter = _job(db)
    _write_validated_cache(tmp_path, _validated_cache_rows())
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        reason="Generation completed.",
    )

    payload = release.release_payload(job)
    titles = [row["concept_title"] for row in payload["records"]]
    assert titles == ["Released Concept Alpha"]


def test_cache_from_a_different_source_contract_is_ignored(
    db, tmp_path, monkeypatch,
):
    job, chapter = _job(db)
    stale = [{
        "topic": "Topic A",
        "parent_concept": "Parent A",
        "concept_title": "Stale Checkpoint Concept",
        "concept_details": "Description: No learner analysis yet.",
        "keywords": "stale",
        "_semantic_topic_id": "TOPIC-A",
        "_semantic_graph_contract": "CONTRACT-2",
        "_source_block_ids": ["BLK-0007"],
    }]
    monkeypatch.setattr(
        release.generation,
        "_newest_compatible_concept_checkpoint",
        lambda _envelope: {"records": copy.deepcopy(stale)},
    )
    _write_validated_cache(tmp_path, _validated_cache_rows())
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        error=RuntimeError("host mutation contract failed closed"),
        reason="Generation failed after creating a durable checkpoint.",
    )

    payload = release.release_payload(job)
    titles = [row["concept_title"] for row in payload["records"]]
    assert titles == ["Stale Checkpoint Concept"]


def test_release_routes_and_manual_decision_endpoint_are_unattended(client, db):
    build_concepts_release_manifest.install()
    release_contract.install()
    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        pending_decision=_pending(),
    )

    response = client.get(f"/build-concepts/uploads/{job.id}")
    assert response.status_code == 200
    body = response.json()
    assert body["awaiting_decision"] is False
    assert body["pending_decision"] is None
    kinds = {row["kind"] for row in body["source_artifacts"]["files"]}
    assert {
        "released_concepts",
        "release_diagnostics",
        "release_payload",
        "database_upload",
    } <= kinds

    assert client.get(
        f"/build-concepts/uploads/{job.id}/release.xlsx"
    ).status_code == 200
    assert client.get(
        f"/build-concepts/uploads/{job.id}/diagnostics.zip"
    ).status_code == 200
    assert client.get(
        f"/build-concepts/uploads/{job.id}/release.json"
    ).status_code == 200

    manual = client.post(
        f"/build-concepts/uploads/{job.id}/decisions/phase32-release-test",
        json={"choice": "accept_recommended"},
    )
    assert manual.status_code == 409
    assert "unattended" in manual.json()["detail"].lower()


def test_diagnostics_route_refuses_to_read_during_a_job_operation(
    client, db, monkeypatch,
):
    """A rebuild cannot be interleaved with the multi-file ZIP snapshot."""

    job, _chapter = _job(db)
    assembled: list[bool] = []

    def unexpected_assembly(*_args, **_kwargs):
        assembled.append(True)
        raise AssertionError("diagnostics were assembled while the job ran")

    monkeypatch.setattr(
        release_files, "build_diagnostics_zip", unexpected_assembly,
    )

    with uploads.exclusive_job_operation(job.id):
        response = client.get(
            f"/build-concepts/uploads/{job.id}/diagnostics.zip"
        )

    assert response.status_code == 409
    assert "already running" in response.json()["detail"]
    assert assembled == []


def test_explicit_upload_publishes_flagged_rows_and_is_idempotent(db, monkeypatch):
    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_rendered_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
        pending_decision=_pending(),
    )
    assert db.query(models.Concept).filter(
        models.Concept.concept_title == "Released Concept Alpha"
    ).count() == 0

    def commit_only(database, _path, ids):
        database.commit()
        return {
            "written": len(ids),
            "sources_updated": 0,
            "publication_status": "published",
        }

    monkeypatch.setattr(
        build_concepts,
        "_commit_and_publish_concept_workbook",
        commit_only,
    )
    first = publication.upload_release_to_database(db, job.id)
    second = publication.upload_release_to_database(db, job.id)

    assert first["database_uploaded"] is True
    assert first["created_concept_ids"]
    assert second["database_uploaded"] is True
    assert db.query(models.Concept).filter(
        models.Concept.concept_title == "Released Concept Alpha"
    ).count() == 1
    db.refresh(job)
    assert job.status == "generated"
    assert release.release_payload(job)["summary"]["database_uploaded"] is True


# --------------------------------------------------------------------------- #
# No generation outcome reaches the user as a decision
# --------------------------------------------------------------------------- #
#
# The 81%/89% selection screen was the last place a run could stop and wait.
# These pin every way generation can end badly, and prove each one produces a
# release instead of a question.

def _run_wrapped(db, job, chapter, outcome):
    """Drive the installed release wrapper around one generation outcome."""

    release_contract.install()

    def original(_db, _job_id, _target_chapter_id, **_kwargs):
        if isinstance(outcome, Exception):
            raise outcome
        return outcome

    return release_contract._run_generation_release(
        original,
        db,
        job.id,
        chapter.id,
        owner_sub=None,
    )


def test_an_unattended_decision_stop_becomes_a_release(db):
    """The stop that fired when only human-only actions remained."""
    job, chapter = _job(db)

    result = _run_wrapped(
        db, job, chapter,
        build_concepts.UnattendedDecisionUnavailable(
            "Unattended generation stopped: the saved decision offers only "
            "actions that a person must take (replace_source)."
        ),
    )

    assert result["status"] == release.RELEASE_STATUS
    db.refresh(job)
    assert job.status == release.RELEASE_STATUS
    payload = release.release_payload(job)
    # The issue is coded by the exception class, so the export names the
    # exact stop that was converted rather than a generic "it failed".
    codes = {issue["code"] for issue in payload["issues"]}
    assert "UnattendedDecisionUnavailable" in codes
    # The reason survives so the export explains why the run ended where it did.
    assert any(
        "replace_source" in str(issue.get("message", ""))
        + str(issue.get("details", ""))
        for issue in payload["issues"]
    )


def test_a_returned_pending_decision_becomes_a_release(db):
    """Generation answering with awaiting_decision must not reach the UI."""
    job, chapter = _job(db)

    result = _run_wrapped(db, job, chapter, {
        "job_id": job.id,
        "status": "awaiting_decision",
        "pending_decision": _pending(),
        "resume_required": False,
    })

    assert result["status"] == release.RELEASE_STATUS
    assert "pending_decision" not in result
    db.refresh(job)
    assert build_concepts._pending_human_decision(job.generation_checkpoint) is None
    payload = release.release_payload(job)
    # The issue is not discarded -- it travels with the release.
    assert payload["pending_decision_snapshot"]
    # The decision's own kind is the issue code, so the export says which
    # semantic question went unanswered.
    assert any(
        issue["code"] == _pending()["kind"] for issue in payload["issues"]
    )


def test_an_ordinary_generation_failure_becomes_a_release(db):
    job, chapter = _job(db)

    result = _run_wrapped(
        db, job, chapter, RuntimeError("phase 3.3 contract is unavailable"),
    )

    assert result["status"] == release.RELEASE_STATUS
    payload = release.release_payload(job)
    assert any(
        "phase 3.3 contract is unavailable" in str(issue.get("message", ""))
        for issue in payload["issues"]
    )


def test_release_bulk_import_workbook_renders_canonical_rows_without_db_upload(db):
    """Creators get the Bulk Import file at run completion: the released
    rows render through the canonical writer on transient objects, with the
    authored chapter metadata applied and no concept touching the database."""
    from app import bulk_import as _bi

    job, chapter = _job(db)
    # A finalized chapter duration wins over authored metadata by design;
    # this test exercises the authored path, so the shared fixture chapter
    # must not carry one while this release is staged. The directory
    # metadata is frozen AT staging, so the row is restored right after —
    # the shared chapter stays contract-conformant (v2.0 §32.1) for every
    # later release test.
    restored = (chapter.chapter_duration, chapter.chapter_description)
    chapter.chapter_duration = ""
    chapter.chapter_description = ""
    db.commit()
    concepts_before = db.query(models.Concept).count()
    try:
        release.stage_release(
            db,
            job,
            target_chapter_id=chapter.id,
            records=_records(),
            inventory=_inventory(),
            mined_types=_mined_types(),
        )
    finally:
        chapter.chapter_duration, chapter.chapter_description = restored
        db.commit()
    db.refresh(job)
    payload = release.release_payload(job)
    payload["chapter_meta"] = {
        "chapter_description": "An authored chapter description.",
        "chapter_duration_minutes": 120,
        "topic_descriptions": {
            _bi.normalize_question_text("Topic A"): "Topic A teaches the method.",
        },
    }
    inventory = copy.deepcopy(dict(job.question_inventory or {}))
    inventory[release.RELEASE_KEY] = payload
    job.question_inventory = inventory
    db.commit()

    data = release_files.build_release_bulk_import_workbook(db, job)
    workbook = load_workbook(io.BytesIO(data))
    from app.bulk_import import SHEET_OBJECTIVE

    sheet = workbook[SHEET_OBJECTIVE]
    rows = list(sheet.iter_rows(min_row=3, values_only=True))
    assert rows, "the canonical export must carry the released concept rows"
    first = rows[0]
    values = [str(value or "") for value in first]
    assert any("Released Concept Alpha" in value for value in values)
    assert any("An authored chapter description." in value for value in values)
    # Contract v2.0 §32: the chapter duration is a real numeric cell (120),
    # never unit-bearing text ("120 minutes").
    assert 120 in [value for value in first if isinstance(value, (int, float))]
    assert not any("120 minutes" in value for value in values)
    assert any("Topic A teaches the method." in value for value in values)
    assert db.query(models.Concept).count() == concepts_before


def _manifest_rows(job) -> dict[str, str]:
    workbook = load_workbook(io.BytesIO(release_files.build_release_workbook(job)))
    sheet = workbook["Release Manifest"]
    return {
        str(sheet.cell(row, 1).value): str(sheet.cell(row, 2).value)
        for row in range(2, sheet.max_row + 1)
    }


def test_a_chapter_the_outline_pass_never_read_says_so_in_the_release(db):
    # "The exercise questions were not picked up" has two very different
    # causes. The release has to name which one happened.
    job, chapter = _job(db)
    inventory = _inventory()
    inventory["extraction_provenance"] = {
        "chapter_outline_applied": False,
        "chapter_outline_version": "",
        "chapter_outline_topics": 0,
        "chapter_outline_partitions": 0,
        "chapter_outline_review_flags": [],
        "source_task_blocks": 9,
        "inventory_items": 4,
        "model_split_items": 0,
    }
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=inventory,
        mined_types=_mined_types(),
    )

    payload = release.release_payload(job)
    codes = {issue["code"] for issue in payload["issues"]}
    assert "chapter_outline_not_applied" in codes
    assert payload["extraction_provenance"]["source_task_blocks"] == 9

    rows = _manifest_rows(job)
    assert "NOT applied" in rows["Chapter outline"]
    assert rows["Questions extracted"] == "4"
    assert rows["Questions from a model-decided split"] == "0"


def test_a_model_read_chapter_reports_its_outline_without_raising_a_warning(db):
    job, chapter = _job(db)
    inventory = _inventory()
    inventory["extraction_provenance"] = {
        "chapter_outline_applied": True,
        "chapter_outline_version": "chapter-outline-1",
        "chapter_outline_topics": 8,
        "chapter_outline_partitions": 5,
        "chapter_outline_review_flags": [],
        "source_task_blocks": 9,
        "inventory_items": 21,
        "model_split_items": 12,
    }
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=inventory,
        mined_types=_mined_types(),
    )

    codes = {issue["code"] for issue in release.release_payload(job)["issues"]}
    assert "chapter_outline_not_applied" not in codes
    assert "chapter_outline_topics_unusable" not in codes

    rows = _manifest_rows(job)
    assert rows["Chapter outline"] == "applied (chapter-outline-1)"
    assert rows["Outline topics decided"] == "8"
    assert rows["Questions from a model-decided split"] == "12"


def test_boundaries_without_topics_are_reported_as_a_single_topic_risk(db):
    job, chapter = _job(db)
    inventory = _inventory()
    inventory["extraction_provenance"] = {
        "chapter_outline_applied": True,
        "chapter_outline_version": "chapter-outline-1",
        "chapter_outline_topics": 0,
        "chapter_outline_partitions": 3,
        "chapter_outline_review_flags": ["no usable content topic in the outline"],
        "source_task_blocks": 9,
        "inventory_items": 11,
        "model_split_items": 6,
    }
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=inventory,
        mined_types=_mined_types(),
    )

    issues = {issue["code"]: issue for issue in release.release_payload(job)["issues"]}
    assert "chapter_outline_topics_unusable" in issues
    assert issues["chapter_outline_topics_unusable"]["severity"] == "warning"
    assert issues["chapter_outline_review_flags"]["severity"] == "info"
    assert "no usable content topic" in _manifest_rows(job)["Outline normalization flags"]


def _repeated_wording_types():
    """Two distinct source questions that reach the learner with one wording."""
    repeated = (
        "Explain how a plant responds to sunlight reaching it from one side."
    )
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    cases[0]["examples"][0]["prompt"] = repeated
    cases[1]["examples"][0]["prompt"] = repeated
    return types


def test_one_question_reaching_the_learner_twice_is_flagged():
    _rows, issues, _routes = release.audit_type_cases(
        _repeated_wording_types(), _inventory()
    )

    repeats = [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]
    assert len(repeats) == 1
    assert repeats[0]["qids"] == ["QINV-0001", "QINV-0002"]
    assert repeats[0]["severity"] == "warning"


def test_distinct_questions_are_not_reported_as_repeats():
    _rows, issues, _routes = release.audit_type_cases(
        _mined_types(), _inventory()
    )

    assert not [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]


def test_a_short_shared_tail_is_reported_as_a_collision():
    """RE-AUTHORED under S11 (T10-5): the ``len(key) < 25`` threshold that
    suppressed this was a character count deciding whether two questions
    are the same question. A short repeat is REPORTED — one grouped
    warning a reviewer dismisses in one glance — because reporting a
    collision is identity accounting; suppressing it was the judgment.
    """
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    cases[0]["examples"][0]["prompt"] = "Why?"
    cases[1]["examples"][0]["prompt"] = "Why?"

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    repeats = [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]
    assert len(repeats) == 1
    assert repeats[0]["severity"] == "warning"
    assert len(repeats[0]["qids"]) == 2


def test_a_devanagari_question_repeated_is_reported():
    """[measured, T10-5] the old Latin-only noise class reduced Devanagari
    prompts to the empty string, so the audit was inert for every
    non-Latin source."""
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    prompt = "\u0926\u0939\u0940 \u092e\u0947\u0902 \u0915\u094c\u0928-\u0938\u093e \u0905\u092e\u094d\u0932 \u092a\u093e\u092f\u093e \u091c\u093e\u0924\u093e \u0939\u0948?"
    cases[0]["examples"][0]["prompt"] = prompt
    cases[1]["examples"][0]["prompt"] = prompt

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    repeats = [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]
    assert len(repeats) == 1
    assert len(repeats[0]["qids"]) == 2


def test_repeated_question_detection_is_case_insensitive():
    """[measured, T10-5] the old key deleted uppercase letters instead of
    folding them, so a recased repeat read as a different question."""
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    cases[0]["examples"][0]["prompt"] = "Name the acid present in curd."
    cases[1]["examples"][0]["prompt"] = "NAME THE ACID PRESENT IN CURD."

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    repeats = [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]
    assert len(repeats) == 1
    assert len(repeats[0]["qids"]) == 2


def test_a_combining_vowel_sign_difference_is_not_a_repeat():
    """Audit finding 18: Python's ``\\w`` excludes Unicode combining marks,
    so the old noise class deleted every Devanagari matra and two DISTINCT
    prompts (differing only by a vowel sign) collided into one false
    ``repeated_question_text`` warning. The key is lossless now."""
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    # "की कौन?" vs "क कौन?" — identical except the combining vowel sign
    # U+0940 on the first akshara.
    cases[0]["examples"][0]["prompt"] = "की कौन?"
    cases[1]["examples"][0]["prompt"] = "क कौन?"

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    assert not [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]


def test_the_repeat_key_is_lossless_case_and_whitespace_fold_nothing_else():
    """RE-AUTHORED under audit finding 18 (this test used to pin the lossy
    behaviour: punctuation stripped, a leading "(2)" discarded). The
    mechanical duplicate warning may only compare, never classify —
    punctuation and numbering are wording now, so prompts differing in
    them are distinct; case and whitespace runs, the lossless folds, still
    collide."""
    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    cases[0]["examples"][0]["prompt"] = (
        "(2) Explain how a plant responds to sunlight from one side."
    )
    cases[1]["examples"][0]["prompt"] = (
        "Explain how a plant responds to sunlight from one side!"
    )

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    assert not [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]

    types = _mined_types()
    cases = types["types"][0]["case_prompts"]
    cases[0]["examples"][0]["prompt"] = (
        "Explain  how a plant\nresponds to sunlight."
    )
    cases[1]["examples"][0]["prompt"] = (
        "EXPLAIN HOW A PLANT RESPONDS TO SUNLIGHT."
    )

    _rows, issues, _routes = release.audit_type_cases(types, _inventory())

    repeats = [
        issue for issue in issues
        if issue["code"] == "repeated_question_text"
    ]
    assert len(repeats) == 1
    assert len(repeats[0]["qids"]) == 2


def test_diagnostics_archive_carries_the_pre_learning_lane(
    db, tmp_path, monkeypatch,
):
    """Step-7 slice C2 (map §6.6): the export used to be blind to the Pre
    lane, so a run that built a Pre-Learning map read exactly like one
    that never did. The recorded map and the capture it was built from
    now reach the ledger and the run report — and a reviewer sees what
    the Pre lane produced without opening the artifact JSON."""
    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_records(),
        inventory=_inventory(),
        mined_types=_mined_types(),
    )
    (tmp_path / "source.phase3-prelearn-capture.json").write_text(
        json.dumps({"prerequisites": [
            {"prerequisite_id": "PR-0001", "text": "what sovereignty is"},
        ]}),
        encoding="utf-8",
    )
    (tmp_path / "source.phase3-prelearn-map.json").write_text(
        json.dumps({
            "rows": [{
                "topic": "Political Vocabulary",
                "concept_title": "Sovereignty",
                "concept_details": (
                    "Description: The highest law-making authority."
                    "\nAchieving Mastery: Saying who holds it."
                    " // Misconception/ Error Analysis: Misconceptions: "
                    "learners may believe a sovereign is merely powerful."
                ),
                "_pre_concept_id": "PRC-0001",
                "_source_grounding_contract": (
                    "derived-from-prerequisite-capture"
                ),
                "_aegis_pre_prerequisites": [
                    {"prerequisite_id": "PR-0001", "text": "what sovereignty is"},
                ],
                "_aegis_analysis_allotments": ["PLA-0001"],
                "_aegis_needed_for": [],
            }],
            "topics": [{
                "pre_topic_id": "PRT-0001",
                "title": "Political Vocabulary",
                "pre_concept_ids": ["PRC-0001"],
            }],
            "analysis": {
                "inventory": [{"item_id": "PLA-0001",
                               "kind": "misconception",
                               "text": "a prerequisite belief"}],
                "allotments": {"PLA-0001": "PRC-0001"},
                "rationales": {},
                "review_flags": {},
            },
            "review_flags": {},
            "decision_flags": {},
            "validation": [],
        }),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        release.uploads,
        "source_artifact_directory",
        lambda _job_id: str(tmp_path),
        raising=False,
    )

    with zipfile.ZipFile(
        io.BytesIO(release_files.build_diagnostics_zip(job))
    ) as archive:
        names = set(archive.namelist())
        assert "context/pre_learning_map.json" in names
        assert "context/pre_learning_capture.json" in names
        ledger = json.loads(archive.read("context/coverage_ledger.json"))
        report = json.loads(archive.read("context/run_report.json"))
        rendered = archive.read("RUN_REPORT.txt").decode("utf-8")

    pre = ledger["summary"]["pre_learning"]
    assert pre["rows"] == 1
    assert pre["prerequisites"] == {"total": 1, "mapped": 1, "unaccounted": 0}
    assert pre["analysis_items"] == {"total": 1, "allotted": 1,
                                     "unaccounted": 0}
    assert report["pre_learning"]["allotted_item_count"] == 1
    assert "PRE-LEARNING MAP (Phase 03)" in rendered
    assert "PRE-LEARNING (Phase 03)" in rendered
    assert "PRC-0001 Sovereignty" in rendered
