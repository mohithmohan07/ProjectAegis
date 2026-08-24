"""The Refiner (§8.3): polish within untouchable identities, never block.

Covers the spec's regressions: a refinement lands with a recorded diff;
identity drift is discarded (rows byte-stable); a refinement introducing a
terminal validator error is rolled back per row; provider failure stages
the unrefined rows with the "refiner unavailable" flag; sealed rows are
re-sealed after refinement; decide-once replays without a provider call.
"""
from __future__ import annotations

import copy
import io
import sys
import types

from openpyxl import load_workbook

from app import models
from app.services import build_concepts_release as release
from app.services import build_concepts_release_contract as release_contract
from app.services import build_concepts_release_files as release_files
from app.services import grounding_certificate as gc
from app.services import release_refiner
from app.services.phase3 import kernel


_METADATA = {
    "board": "CBSE",
    "grade": "9",
    "subject": "History",
    "unit": "India and the Contemporary World",
    "chapter_title": "The French Revolution",
    "chapter_code": "HIST-09-01",
    "pre_post": "Post",
    "source_book": "NCERT",
    "inventory": {"items": []},
    "mined_types": {"types": []},
    "source_text": "",
}

_BEFORE_DETAILS = (
    "Description: Learners see how ordinary citizens, not monarchs, came "
    "to define political belonging.\n"
    "Achieving Mastery: Explaining who constitutes a nation in "
    "revolutionary thought. // Misconception/ Error Analysis: "
    "Misconceptions: Learners believe monarchs granted nationhood as a "
    "favour.; Error Analysis: The learner attributes the 1789 "
    "constitution to royal initiative instead of popular sovereignty."
)

_AFTER_DETAILS = _BEFORE_DETAILS.replace(
    "Learners see how ordinary citizens",
    "Learners discover how ordinary citizens",
)

_TYPES_DETAILS = (
    "Description: The revolution built its own political vocabulary from "
    "everyday civic practice across France.\n"
    "Achieving Mastery: Explaining one revolutionary symbol in context. "
    "// Types: Type 01: Evidence-Based Explanation — Explain a source "
    "relationship by connecting evidence to the required framework. "
    "Case 01: Explain the foundational relationship. Example: Explain "
    "the first method. // Misconception/ Error Analysis: Misconceptions: "
    "Learners believe symbols were decorative rather than political."
)


def _rows() -> list[dict]:
    return [
        {
            "topic": "The Age Of Revolutions",
            "parent_concept": "",
            "concept_title": "The Nation As A People's Project",
            "concept_details": _BEFORE_DETAILS,
            "keywords": "nation, sovereignty",
            "_semantic_topic_id": "T-1",
            "_source_block_ids": ["BLK-0001"],
            "_aegis_release_qids": ["QINV-0001"],
        },
        {
            "topic": "The Age Of Revolutions",
            "parent_concept": "",
            "concept_title": "Symbols Of The Revolution",
            "concept_details": _TYPES_DETAILS,
            "keywords": "symbols",
            "_semantic_topic_id": "T-1",
            "_source_block_ids": ["BLK-0002"],
            "_aegis_release_qids": ["QINV-0002"],
        },
    ]


class _Provider:
    """A fake per-row provider applying a text replacement when it matches."""

    def __init__(self, replace: tuple[str, str] | None = None,
                 details_override=None, keywords_override=None):
        self.calls = 0
        self.replace = replace
        self.details_override = details_override
        self.keywords_override = keywords_override

    def __call__(self, payload: dict) -> dict:
        self.calls += 1
        row = payload["rows"][0]
        details = str(row["concept_details"])
        if self.details_override is not None:
            details = self.details_override(details)
        elif self.replace is not None:
            details = details.replace(*self.replace)
        keywords = str(row.get("keywords") or "")
        if self.keywords_override is not None:
            keywords = self.keywords_override(keywords)
        return {
            "rows": [{
                "row_ref": row["row_ref"],
                "concept_details": details,
                "keywords": keywords,
                "rationale": "polished wording for grade consistency",
            }]
        }


def test_assessment_master_output_kind_delegates_lazily(monkeypatch):
    records = [{"candidates": [{"candidate_id": "CAND-1"}], "groups": []}]
    metadata = {"subject": "Mathematics", "grade": "6"}
    instruction_set = {"slots": {"tone": "clear"}}
    provider = lambda _payload: {}  # noqa: E731 - identity sentinel
    critic = lambda _payload: {}  # noqa: E731 - identity sentinel
    fixer = lambda _payload: {}  # noqa: E731 - identity sentinel
    store = kernel.DecisionStore()
    seen = {}
    expected = (
        [{"delegated": True}],
        {
            "policy_version": "assessment-master-refiner-3",
            "output_kind": "assessment_master",
            "changes": [],
        },
        ["assessment refiner review"],
    )

    def refine_master(received, **kwargs):
        seen["records"] = received
        seen.update(kwargs)
        return expected

    module_name = "app.services.assessment_master_refiner"
    module = types.ModuleType(module_name)
    module.refine_master = refine_master
    monkeypatch.setitem(sys.modules, module_name, module)
    services_package = sys.modules["app.services"]
    monkeypatch.setattr(
        services_package, "assessment_master_refiner", module, raising=False
    )

    actual = release_refiner.refine_release(
        records,
        metadata=metadata,
        instruction_set=instruction_set,
        provider=provider,
        critic=critic,
        store=store,
        output_kind="assessment_master",
        envelope_sha256="sealed-envelope",
        fixer=fixer,
    )

    assert actual == expected
    assert seen["records"] == records
    assert seen["records"] is not records
    assert seen["records"][0] is not records[0]
    assert seen["metadata"] is metadata
    assert seen["instruction_set"] is instruction_set
    assert seen["provider"] is provider
    assert seen["critic"] is critic
    assert seen["store"] is store
    assert seen["envelope_sha256"] == "sealed-envelope"
    assert seen["fixer"] is fixer


def test_concepts_release_ignores_assessment_only_arguments_byte_for_byte():
    class RecordingProvider(_Provider):
        def __init__(self):
            super().__init__(replace=("see how", "discover how"))
            self.payloads = []

        def __call__(self, payload):
            self.payloads.append(copy.deepcopy(payload))
            return super().__call__(payload)

    class UnreadableEnvelope:
        def __str__(self):  # pragma: no cover - reading it fails the test
            raise AssertionError("concepts_release read assessment envelope")

    def forbidden_fixer(_payload):
        raise AssertionError("concepts_release called assessment Fixer")

    def recording_critic(log):
        def critic(payload):
            log.append(copy.deepcopy(payload))
            return {"verdict": "verified", "confidence": 1.0, "issues": []}

        return critic

    default_provider = RecordingProvider()
    explicit_provider = RecordingProvider()
    default_critic_payloads = []
    explicit_critic_payloads = []
    default_store = kernel.DecisionStore()
    explicit_store = kernel.DecisionStore()

    implicit = release_refiner.refine_release(
        _rows(),
        metadata=_METADATA,
        provider=default_provider,
        critic=recording_critic(default_critic_payloads),
        store=default_store,
        envelope_sha256=UnreadableEnvelope(),
        fixer=forbidden_fixer,
    )
    explicit = release_refiner.refine_release(
        _rows(),
        metadata=_METADATA,
        provider=explicit_provider,
        critic=recording_critic(explicit_critic_payloads),
        store=explicit_store,
        output_kind="concepts_release",
        envelope_sha256=UnreadableEnvelope(),
        fixer=forbidden_fixer,
    )

    assert explicit == implicit
    assert explicit_provider.payloads == default_provider.payloads
    assert explicit_critic_payloads == default_critic_payloads
    assert explicit_store.keys() == default_store.keys()
    assert len(explicit_store.keys()) == len(_rows())


def test_a_refinement_lands_with_a_recorded_diff_and_row_mark():
    provider = _Provider(replace=(
        "Learners see how ordinary citizens",
        "Learners discover how ordinary citizens",
    ))
    refined, diff, flags = release_refiner.refine_release(
        _rows(), metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert refined[0]["concept_details"] == _AFTER_DETAILS
    assert refined[0]["_aegis_release_refined"] is True
    assert diff["policy_version"] == release_refiner.REFINER_POLICY_VERSION
    changes = diff["changes"]
    assert len(changes) == 1
    assert changes[0]["field"] == "concept_details"
    assert changes[0]["before"] == _BEFORE_DETAILS
    assert changes[0]["after"] == _AFTER_DETAILS
    assert changes[0]["reason"] == "polished wording for grade consistency"
    assert "concept_key" in changes[0]
    # The untouched sibling row stays byte-identical and unmarked.
    assert refined[1] == _rows()[1]
    assert "refiner unavailable" not in " ".join(flags)


def test_critic_dissent_is_an_advisory_flag_never_a_gate():
    provider = _Provider(replace=("see how", "discover how"))

    def critic(_payload):
        return {"verdict": "rejected", "confidence": 0.5,
                "issues": ["tone drifts above grade level"]}

    refined, diff, flags = release_refiner.refine_release(
        _rows(), metadata=_METADATA, provider=provider, critic=critic,
        store=kernel.DecisionStore(),
    )
    # The refinement still ships (decide-once, Q10) with the dissent
    # recorded for review.
    assert refined[0]["concept_details"] != _BEFORE_DETAILS
    assert diff["changes"]
    assert any("critic" in flag for flag in flags)


def test_identity_drift_is_discarded_and_rows_stay_byte_stable():
    # The model edits a Case line inside the untouchable Types section.
    provider = _Provider(replace=(
        "Case 01: Explain the foundational relationship.",
        "Case 01: Explain the improved foundational relationship.",
    ))
    original = _rows()
    refined, diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert refined == original
    assert diff["changes"] == []
    assert any(
        "refiner refinement discarded: identity drift" in flag
        for flag in flags
    )


def test_qid_and_section_drift_are_both_discarded():
    provider = _Provider(details_override=lambda details: details.replace(
        "QINV", "QINV-X"
    ).replace("// Misconception/ Error Analysis:", "// Learner Analysis:"))
    original = _rows()
    refined, _diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert refined == original
    assert any("identity drift" in flag for flag in flags)


def test_a_refinement_introducing_a_validator_error_is_rolled_back():
    # Deleting the mastery sentence passes the identity check (Description
    # prose is editable) but introduces missing_mastery_statement at the
    # terminal gate — the mechanical re-check rolls that row back.
    def drop_mastery(details: str) -> str:
        return details.replace(
            "\nAchieving Mastery: Explaining who constitutes a nation in "
            "revolutionary thought.",
            "",
        )

    provider = _Provider(details_override=drop_mastery)
    original = _rows()
    refined, diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    # The rolled-back row is byte-stable and contributes no diff entry;
    # nothing else changed, so the diff is empty.
    assert refined == original
    assert diff["changes"] == []
    assert any(
        "rolled back" in flag and "missing_mastery_statement" in flag
        for flag in flags
    )


def test_provider_failure_stages_unrefined_rows_with_availability_flag():
    def broken(_payload):
        raise RuntimeError("provider down")

    original = _rows()
    refined, diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=broken,
        store=kernel.DecisionStore(),
    )
    assert refined == original
    assert diff["changes"] == []
    assert any(flag.startswith("refiner unavailable:") for flag in flags)


def test_contract_exhaustion_stages_unrefined_rows_never_blocks():
    def malformed(_payload):
        return {"nonsense": True}

    original = _rows()
    refined, _diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=malformed,
        store=kernel.DecisionStore(),
    )
    assert refined == original
    assert any(
        flag.startswith("refiner unavailable:") and "ContractError" in flag
        for flag in flags
    )


def _sealed_rows() -> list[dict]:
    rows = _rows()
    for row in rows:
        row["_source_grounding_contract"] = "phase38-independent-verified"
        row["_source_grounding_version"] = "phase3.8-certified-required-grounding-5"
    gc.seal_records(
        rows,
        source_contract_hash="C-1",
        semantic_topology_sha256="a" * 64,
        allowed_block_ids={"BLK-0001", "BLK-0002"},
    )
    return rows


def test_refined_rows_are_resealed_when_the_capture_carried_seals():
    provider = _Provider(replace=("see how", "discover how"))
    original = _sealed_rows()
    before_certificate = original[0][gc.ROW_CERTIFICATE_FIELD]
    before_identity = original[0][gc.ROW_IDENTITY_FIELD]
    refined, diff, _flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert diff["resealed_after_refinement"] is True
    assert diff["changes"]
    # The Description changed, so the sealed source claim and row
    # certificate were re-minted; the row IDENTITY (topic/title/placement)
    # is untouched by construction.
    assert refined[0][gc.ROW_CERTIFICATE_FIELD] != before_certificate
    assert refined[0][gc.ROW_IDENTITY_FIELD] == before_identity
    assert refined[0][gc.CONCEPT_ID_FIELD] == original[0][gc.CONCEPT_ID_FIELD]


def test_a_reseal_failure_falls_back_to_unrefined_rows(monkeypatch):
    provider = _Provider(replace=("see how", "discover how"))
    original = _sealed_rows()

    def failing_reseal(_rows):
        raise gc.GroundingCertificateError("cannot re-seal")

    monkeypatch.setattr(release_refiner, "_reseal", failing_reseal)
    refined, diff, flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert refined == original
    assert diff["resealed_after_refinement"] is False
    assert any(
        flag.startswith("refiner unavailable:") and "re-seal" in flag
        for flag in flags
    )


def test_unsealed_captures_are_never_resealed():
    provider = _Provider(replace=("see how", "discover how"))
    refined, diff, _flags = release_refiner.refine_release(
        _rows(), metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert diff["resealed_after_refinement"] is False
    assert gc.ROW_CERTIFICATE_FIELD not in refined[0]


def test_decide_once_replays_without_a_provider_invocation():
    provider = _Provider(replace=("see how", "discover how"))
    store = kernel.DecisionStore()
    first, first_diff, _ = release_refiner.refine_release(
        _rows(), metadata=_METADATA, provider=provider, store=store,
    )
    calls_after_first = provider.calls
    assert calls_after_first == 2  # one decision per row
    second, second_diff, _ = release_refiner.refine_release(
        _rows(), metadata=_METADATA, provider=provider, store=store,
    )
    assert provider.calls == calls_after_first
    assert second == first
    assert second_diff["changes"] == first_diff["changes"]


def test_the_refiner_never_edits_identity_fields_mechanically():
    # Even a provider that also returns renamed fields cannot move them:
    # only the whitelist is ever applied.
    class RenamingProvider(_Provider):
        def __call__(self, payload):
            response = super().__call__(payload)
            response["rows"][0]["concept_title"] = "A Brand New Name"
            response["rows"][0]["topic"] = "A Brand New Topic"
            return response

    provider = RenamingProvider(replace=("see how", "discover how"))
    original = _rows()
    refined, _diff, _flags = release_refiner.refine_release(
        original, metadata=_METADATA, provider=provider,
        store=kernel.DecisionStore(),
    )
    assert refined[0]["concept_title"] == original[0]["concept_title"]
    assert refined[0]["topic"] == original[0]["topic"]
    assert refined[0]["_aegis_release_qids"] == original[0]["_aegis_release_qids"]


# --------------------------------------------------------------------------- #
# The release-contract seam: refined records are staged, the payload carries
# the recorded diff, and the workbook exposes it.
# --------------------------------------------------------------------------- #


def _chapter(db):
    value = db.query(models.Chapter).order_by(models.Chapter.id).first()
    assert value is not None
    return value


def _job(db):
    chapter = _chapter(db)
    job = models.UploadJob(
        module="build_concepts",
        upload_type="document",
        learning_kind="post",
        source_book="NCERT",
        filename="refiner-source.mmd",
        mmd_text="## Topic\nA source paragraph.",
        status="converted",
        deposit_scope_type="chapter",
        deposit_scope_ids=[chapter.id],
        generation_checkpoint={},
        question_inventory={"items": [], "stats": {}, "mined_types": []},
        generation_log=[],
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job, chapter


def test_release_after_result_stages_refined_records_with_diff(db, monkeypatch):
    job, chapter = _job(db)
    rows = _rows()
    provider = _Provider(replace=(
        "Learners see how ordinary citizens",
        "Learners discover how ordinary citizens",
    ))
    monkeypatch.setattr(
        release_refiner, "_default_provider", lambda: provider
    )
    monkeypatch.setattr(release_refiner, "_default_critic", lambda: None)

    def original(_db, _job_id, _chapter_id, **_kwargs):
        release_contract._RELEASE_CAPTURE.set({
            "records": copy.deepcopy(rows),
            "inventory": {"items": [], "stats": {}},
            "mined_types": {"types": []},
            "final_grounding_certificate": {},
            "checkpoint": {},
        })
        return {"job_id": job.id, "status": "generated"}

    result = release_contract._run_generation_release(
        original, db, job.id, chapter.id, owner_sub=None,
    )
    assert result["status"] == release.RELEASE_STATUS
    payload = release.release_payload(job)
    assert payload is not None
    refinements = payload["refinements"]
    assert refinements["policy_version"] == (
        release_refiner.REFINER_POLICY_VERSION
    )
    assert refinements["changes"]
    assert refinements["changes"][0]["field"] == "concept_details"
    staged = payload["records"]
    assert "Learners discover how ordinary citizens" in (
        staged[0]["concept_details"]
    )
    assert staged[0][release.RELEASE_ROW_REFINED_FIELD] is True
    # The audit mark never reaches the database upload payload.
    assert release.RELEASE_ROW_REFINED_FIELD not in (
        release._strip_release_fields(staged[0])
    )

    workbook = load_workbook(
        io.BytesIO(release_files.build_release_workbook(job))
    )
    assert "Refinements" in workbook.sheetnames
    sheet = workbook["Refinements"]
    assert sheet.cell(1, 1).value == "Kind"
    kinds = {
        str(sheet.cell(row, 1).value) for row in range(2, sheet.max_row + 1)
    }
    assert "refinement" in kinds


def test_release_without_a_live_refiner_ships_unrefined_with_flag(db):
    job, chapter = _job(db)
    rows = _rows()

    def original(_db, _job_id, _chapter_id, **_kwargs):
        release_contract._RELEASE_CAPTURE.set({
            "records": copy.deepcopy(rows),
            "inventory": {"items": [], "stats": {}},
            "mined_types": {"types": []},
            "final_grounding_certificate": {},
            "checkpoint": {},
        })
        return {"job_id": job.id, "status": "generated"}

    result = release_contract._run_generation_release(
        original, db, job.id, chapter.id, owner_sub=None,
    )
    assert result["status"] == release.RELEASE_STATUS
    payload = release.release_payload(job)
    refinements = payload["refinements"]
    assert refinements["changes"] == []
    assert any(
        str(flag).startswith("refiner unavailable:")
        for flag in refinements["review_flags"]
    )
    # The staged rows are the captured rows, unrefined.
    assert payload["records"][0]["concept_details"] == _BEFORE_DETAILS


def test_direct_stage_release_callers_carry_no_refinements_key(db):
    job, chapter = _job(db)
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=_rows(),
        inventory={"items": []},
        mined_types={"types": []},
    )
    payload = release.release_payload(job)
    assert "refinements" not in payload
    workbook = load_workbook(
        io.BytesIO(release_files.build_release_workbook(job))
    )
    assert "Refinements" not in workbook.sheetnames
