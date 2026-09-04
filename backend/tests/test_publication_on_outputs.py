"""The upload page's Source book is the one source every output names.

Owner ruling, 2026-09-04 (register Q27; contract v2.0 §18): ``concept_source``
on Outputs 01/03 and ``concept_source`` / ``question_source`` on Outputs
02/04 all carry the run's publication — the Source book typed on the Build
Concepts page — never a concept's accumulated provenance list and never a
filename. The same ruling makes a Post Master question the source task's
wording verbatim; the prompt pins below hold that rule in place.
"""
from __future__ import annotations

import io

import openpyxl

from app import models
from app.bulk_import import assessment_workbook as aw
from app.bulk_import import layouts
from app.services import (
    assessment_item_review,
    assessment_materialization,
    assessment_release_snapshot as release_snapshot,
    build_concepts_release as release,
    build_concepts_release_files as release_files,
)
from tests.test_assessment_release_run import _chapter_with_concepts
from tests.test_bulk_import_layout_migration import (
    _both_lanes_job_with_publication,
)


def _rows(data: bytes, sheet: str = "Objective") -> tuple[list[str], list[list]]:
    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    try:
        ws = workbook[sheet]
        header = [str(c or "") for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
        rows = [
            list(row) for row in ws.iter_rows(min_row=3, values_only=True)
            if row and any(str(v or "").strip() for v in row)
        ]
        return header, rows
    finally:
        workbook.close()


def test_the_concept_file_names_the_upload_pages_source_on_every_row(db):
    """Output 03 before the database write (the transient hierarchy)."""
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job_with_publication(db, chapter, source_book="Balbharati")

    data = release_files.build_release_bulk_import_workbook(
        db, job, lane=release_files.LANE_POST,
    )
    header, rows = _rows(data)
    source = header.index("concept_source")
    title = header.index("chapter_title")
    assert rows, "the Concept File carries the released rows"
    assert {str(row[source]) for row in rows} == {"Balbharati"}
    # The chapter's human tag names the same publication.
    assert all("Balbharati" in str(row[title]) for row in rows)


def test_a_persisted_concepts_provenance_list_never_reaches_the_concept_source_cell(db):
    """Output 03 after the database write (the by-id writer)."""
    from app.bulk_import import writer

    chapter = _chapter_with_concepts(db)
    concepts = [c for t in chapter.topics for c in t.concepts]
    assert concepts
    for concept in concepts:
        concept.sources = "NCERT | Fullmarks"
    db.commit()

    data = writer.write_concepts_workbook(
        db, [c.id for c in concepts],
        layout_id=writer.CONCEPT_FILE_LAYOUT_ID, publication="Balbharati",
    )
    header, rows = _rows(data)
    source = header.index("concept_source")
    assert rows
    assert {str(row[source]) for row in rows} == {"Balbharati"}

    # Without a run there is no publication to name: the accumulator export
    # keeps the persisted provenance, exactly as before.
    legacy = writer.write_concepts_workbook(
        db, [c.id for c in concepts], layout_id=writer.CONCEPT_FILE_LAYOUT_ID,
    )
    header, rows = _rows(legacy)
    assert {str(row[header.index("concept_source")]) for row in rows} == {
        "NCERT | Fullmarks",
    }


def test_the_master_snapshot_carries_the_publication_as_every_concepts_source(db):
    """Outputs 02/04 read ``concept_source`` off the snapshot rows."""
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job_with_publication(db, chapter, source_book="Balbharati")
    for topic in chapter.topics:
        for concept in topic.concepts:
            concept.sources = "NCERT | Fullmarks"
    db.commit()

    bridge = release_snapshot.build(db, job, release.release_payload(job))
    assert bridge["source_book"] == "Balbharati"
    concept_rows = [c for t in bridge["snapshot"]["topics"] for c in t["concepts"]]
    assert concept_rows
    assert {row["concept_source"] for row in concept_rows} == {"Balbharati"}
    assert {row["concept_source"] for row in bridge["concepts"]} == {"Balbharati"}
    assert bridge["snapshot"]["source_book"] == "Balbharati"


def test_the_universal_layout_is_what_the_concept_file_renders():
    """Register Q27: 72/440/149 on Outputs 01/03 too."""
    schema = aw.output_schema("concept", None, None)
    assert [len(schema["fields"][s]) for s in aw.SHEET_ORDER] == [72, 440, 149]
    assert schema["descriptive_answer_slots"] == layouts.UNIVERSAL_DESCRIPTIVE_ANSWER_SLOTS


def test_post_questions_are_the_sources_wording_verbatim_by_rule():
    """The materialization author and both reviewers hold the ruling."""
    rules = assessment_materialization.MATERIALIZE_SYSTEM
    assert "VERBATIM" in rules
    assert "never rephrase, simplify, expand, modernise, correct or 'polish'" in rules
    assert "never create an item the source does not contain" in rules
    # The retired latitude is gone from the author's rules.
    assert "write a clear, complete, self-contained item" not in rules
    assert assessment_materialization.MATERIALIZE_POLICY_VERSION == (
        "assessment-materialize-14"
    )
    critic = assessment_materialization.MATERIALIZE_CRITIC_SYSTEM
    assert "source wording" in critic
    review = assessment_item_review.ITEM_REVIEW_SYSTEM
    assert "VERBATIM" in review
    assert "any rephrasing, simplification" in review
    assert assessment_item_review.ITEM_REVIEW_POLICY_VERSION == (
        "assessment-item-review-2"
    )
