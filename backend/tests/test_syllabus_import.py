"""Tests for syllabus structure import."""
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import models
from app.services import syllabus_import as svc


def _write_xlsx(path: Path, rows: list[list], sheet_name: str = "Sheet1") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


@pytest.fixture()
def cbse_xlsx(tmp_path):
    path = tmp_path / "cbse.xlsx"
    _write_xlsx(path, [
        ["Grade", "Subject", "Unit", "Chapter"],
        ["6", "Mathematics", "Number Systems", "Knowing Our Numbers"],
        ["6", "Mathematics", "Number Systems", "Whole Numbers"],
        ["6", "Science", "Materials", "Sorting Materials Into Groups"],
    ])
    return path


@pytest.fixture()
def english_xlsx(tmp_path):
    path = tmp_path / "english.xlsx"
    _write_xlsx(path, [
        ["Grade", "Unit", "Chapter"],
        ["8", "Reading Skills", "Comprehension Passages"],
        ["8", "Writing Skills", "Formal Letters"],
    ])
    return path


def test_parse_workbook_cbse(cbse_xlsx):
    rows = svc.parse_workbook(cbse_xlsx, default_board="CBSE")
    assert len(rows) == 3
    assert rows[0].board == "CBSE"
    assert rows[0].grade == "06"
    assert rows[0].subject == "Mathematics"
    assert rows[0].unit == "Number Systems"
    assert rows[0].chapter == "Knowing Our Numbers"


def test_english_universal_across_boards(english_xlsx):
    rows = svc.parse_workbook(
        english_xlsx,
        default_subject="English Language",
        universal_boards=svc.ALL_SYLLABUS_BOARDS,
    )
    boards = {r.board for r in rows}
    assert boards == set(svc.ALL_SYLLABUS_BOARDS)
    assert all(r.subject == "English Language" for r in rows)
    assert len(rows) == len(svc.ALL_SYLLABUS_BOARDS) * 2


def test_upsert_chapters_creates_shells(db, cbse_xlsx):
    before = db.query(models.Chapter).count()
    rows = svc.parse_workbook(cbse_xlsx, default_board="CBSE")
    result = svc.upsert_chapters(db, rows)
    assert result["created"] == 3
    assert db.query(models.Chapter).count() == before + 3
    ch = db.query(models.Chapter).filter_by(
        board="CBSE", chapter_title="Knowing Our Numbers",
    ).one()
    assert ch.board == "CBSE"
    assert ch.grade == "06"
    assert ch.unit == "Number Systems"
    assert ch.topics == []
    # Remove test rows so later tests keep the seeded fixture tree stable.
    db.query(models.Chapter).filter(
        models.Chapter.chapter_title.in_(
            ["Knowing Our Numbers", "Whole Numbers", "Sorting Materials Into Groups"],
        ),
    ).delete(synchronize_session=False)
    db.commit()


def test_upsert_skips_duplicates(db, cbse_xlsx):
    rows = svc.parse_workbook(cbse_xlsx, default_board="CBSE")
    svc.upsert_chapters(db, rows)
    again = svc.upsert_chapters(db, rows)
    assert again["created"] == 0
    assert again["skipped"] == 3


def test_bootstrap_syllabus_mirrors_the_workbooks_every_start(
    db, cbse_xlsx, monkeypatch,
):
    """Startup mirrors the bundled workbooks instead of seeding once.

    Loading only into an EMPTY database meant a deploy shipping re-issued
    workbooks never applied them — the app kept serving the old directory.
    Bootstrap now refreshes on every start and is idempotent.
    """
    import app.config as cfg

    db.query(models.Chapter).delete()
    db.commit()

    syllabus_dir = cbse_xlsx.parent
    monkeypatch.setattr(cfg, "BUNDLED_SYLLABUS_DIR", syllabus_dir)
    monkeypatch.setattr(cfg, "SYLLABUS_DIR", syllabus_dir)
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])

    result = svc.bootstrap_syllabus(db)
    assert result is not None
    assert result["created"] == 3

    # A second start changes nothing.
    again = svc.bootstrap_syllabus(db)
    assert again is not None
    assert again["created"] == 0
    assert again["pruned"] == 0
    assert db.query(models.Chapter).count() == 3

    db.query(models.Chapter).delete()
    db.commit()
    # Restore shared session fixture data for downstream tests.
    from tests.conftest import _load_test_fixtures
    _load_test_fixtures()


def test_imported_chapter_display_names_carry_no_codes(db, cbse_xlsx):
    """Display names are the human-readable half of each band.

    The *_title columns carry the code tag; the *_display_name columns stay
    clean, the same rule the workbook writer follows for topics and concepts.
    Chapters were storing the tagged form, so the directory API served
    "Determiners (06_English_Language_CBSE)" wherever a display name showed.
    """
    import re

    rows = svc.parse_workbook(cbse_xlsx, default_board="CBSE")
    svc.upsert_chapters(db, rows)

    chapter = db.query(models.Chapter).filter_by(
        chapter_title="Knowing Our Numbers").one()
    assert chapter.chapter_display_name == "Knowing Our Numbers"
    tagged = [
        c for c in db.query(models.Chapter).all()
        if re.search(r"\([A-Za-z0-9]+(?:_[A-Za-z0-9]+)+\)",
                     c.chapter_display_name or "")
    ]
    assert tagged == []


def test_english_workbook_routes_regardless_of_name_separators(tmp_path):
    """A separator change in the filename must not reroute a workbook.

    The English Language sheet carries only Unit + Chapter and is replicated
    across every board and grade by a dedicated parser. Routing matched the
    literal "english language", so re-exporting the file as
    "English_Language_Units_and_Chapters.xlsx" sent it to the generic parser,
    which found no board or grade and silently dropped all 580 chapters.
    """
    rows = [
        ["Unit", "Chapter Name"],
        ["Grammar", "Determiners"],
        ["Grammar", "Modals"],
    ]
    for name in (
        "English Language Units and Chapters.xlsx",
        "English_Language_Units_and_Chapters.xlsx",
        "English-Language-Units-and-Chapters.xlsx",
    ):
        path = tmp_path / name
        _write_xlsx(path, rows)
        parsed = svc.parse_workbook(path)
        assert parsed, f"{name} produced no rows"
        # Replicated across every board and grade.
        assert {r.board for r in parsed} == set(svc.ALL_SYLLABUS_BOARDS)
        assert {r.grade for r in parsed} == set(svc.ENGLISH_LANGUAGE_GRADES)
        assert {r.subject for r in parsed} == {"English Language"}
        assert {r.chapter for r in parsed} == {"Determiners", "Modals"}


def test_bundled_syllabus_filenames_all_route_to_a_real_parser(tmp_path):
    """Every shipped workbook name resolves to board-tagged rows."""
    for key, filename in svc.SYLLABUS_FILES.items():
        options = svc._infer_file_options(filename)
        if key == "english_language":
            assert options.get("universal_boards"), filename
        else:
            assert options.get("default_board"), filename


@pytest.fixture()
def isolated_db(tmp_path):
    """A private database for refresh tests.

    ``refresh_syllabus`` deletes superseded chapters, so it must never run
    against the shared fixture database other tests depend on.
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    from app import models as app_models

    engine = create_engine(f"sqlite:///{tmp_path / 'iso.db'}")
    app_models.Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    try:
        yield session
    finally:
        session.close()
        engine.dispose()


def _seed_superseded_chapter(db, **overrides):
    from app.services import directory

    fields = {
        "board": "CBSE", "grade": "06", "subject": "Civics",
        "unit": "Governance", "chapter_title": "Retired Chapter",
        **overrides,
    }
    chapter = models.Chapter(
        chapter_code=directory.make_chapter_code(
            fields["board"], fields["grade"], fields["subject"],
            fields["chapter_title"],
        ),
        chapter_display_name=fields["chapter_title"],
        **fields,
    )
    db.add(chapter)
    db.commit()
    return chapter


def test_refresh_retires_superseded_empty_chapters(
    isolated_db, cbse_xlsx, monkeypatch,
):
    """A re-issued syllabus must not leave the same chapter listed twice.

    load_all_syllabus_files only ADDS. A chapter that changes subject gets a
    new code, so the add-only path kept the old row too — on the live
    database that produced 96 duplicate board+grade+title pairs.
    """
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])
    stale_id = _seed_superseded_chapter(isolated_db).id

    result = svc.refresh_syllabus(isolated_db)

    assert result["pruned"] == 1
    assert isolated_db.get(models.Chapter, stale_id) is None
    assert isolated_db.query(models.Chapter).filter_by(
        chapter_title="Knowing Our Numbers").count() == 1


def test_refresh_never_deletes_a_chapter_with_authored_work(
    isolated_db, cbse_xlsx, monkeypatch,
):
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])
    kept = _seed_superseded_chapter(
        isolated_db, chapter_title="Chapter With Work")
    kept.topics.append(
        models.Topic(topic_title="Authored Topic", pre_post_learning="Post"))
    isolated_db.commit()
    kept_id = kept.id

    result = svc.refresh_syllabus(isolated_db)

    assert isolated_db.get(models.Chapter, kept_id) is not None
    assert result["pruned"] == 0
    assert any(
        "Chapter With Work" in entry
        for entry in result["retained_with_content"]
    )


def test_refresh_is_idempotent(isolated_db, cbse_xlsx, monkeypatch):
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])

    first = svc.refresh_syllabus(isolated_db)
    total = isolated_db.query(models.Chapter).count()
    second = svc.refresh_syllabus(isolated_db)

    assert first["created"] > 0
    assert second["created"] == 0
    assert second["pruned"] == 0
    assert isolated_db.query(models.Chapter).count() == total


def test_refresh_never_prunes_when_a_workbook_is_missing(
    isolated_db, cbse_xlsx, monkeypatch,
):
    """A partial or broken deploy must not empty the directory."""
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(
        svc, "_missing_expected_files", lambda: ["UnitChapter_List__ICSE.xlsx"])
    stale_id = _seed_superseded_chapter(isolated_db).id

    result = svc.refresh_syllabus(isolated_db)

    assert result["pruned"] == 0
    assert isolated_db.get(models.Chapter, stale_id) is not None


def test_bootstrap_applies_reissued_workbooks_to_a_populated_database(
    isolated_db, cbse_xlsx, monkeypatch,
):
    """Startup used to load only into an EMPTY database.

    A deploy shipping re-issued workbooks therefore never applied them: the
    app kept serving the previous directory (2508 chapters) and none of the
    new chapters appeared in the frontend.
    """
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [cbse_xlsx])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])
    _seed_superseded_chapter(isolated_db)

    result = svc.bootstrap_syllabus(isolated_db)

    assert result is not None
    assert result["created"] > 0
    assert isolated_db.query(models.Chapter).filter_by(
        chapter_title="Knowing Our Numbers").count() == 1


def test_karnataka_files_every_social_strand_under_social_science():
    """Karnataka teaches Social Science as one subject, like CBSE.

    History/Civics/Geography/Economics/Sociology/Business Studies were
    stored as separate Karnataka subjects, so the dropdown offered six
    divisions the board does not examine separately.
    """
    from app.services import directory

    for strand in (
        "History", "Civics", "Geography", "Economics",
        "Sociology", "Business Studies",
    ):
        assert directory.effective_subject_for_tags(
            "Karnataka", strand) == "Social Science"
    # Non-social subjects are untouched.
    assert directory.effective_subject_for_tags(
        "Karnataka", "Mathematics") == "Mathematics"


def test_icse_examines_history_and_civics_as_one_subject():
    """ICSE sets History and Civics as a single paper; Geography stands alone."""
    from app.services import directory

    assert directory.effective_subject_for_tags(
        "ICSE", "History") == "History and Civics"
    assert directory.effective_subject_for_tags(
        "ICSE", "Civics") == "History and Civics"
    assert directory.effective_subject_for_tags(
        "ICSE", "Geography") == "Geography"
    # ICSE does not consolidate into Social Science.
    assert directory.effective_subject_for_tags(
        "ICSE", "Economics") == "Economics"


def test_subject_banner_beats_chapter_numbers_in_the_subject_column(tmp_path):
    """A banner row names the block's subject; the rows hold chapter numbers.

    The Karnataka sheets carry "Grade 6 Physical Education" over rows whose
    subject column reads "Chapter 1", "Chapter 2" … so the directory offered
    eleven subjects called "Chapter N".
    """
    path = tmp_path / "kstate.xlsx"
    _write_xlsx(path, [
        ["Subject", "Unit", "Chapter"],
        ["Mathematics", "Number System (06_KSTATE)", "Patterns in Mathematics"],
        ["Grade 6 Physical Education", None, None],
        ["Chapter 1", "Theory (06_KSTATE)", "Meaning of Physical Education"],
        ["Chapter 2", "Theory (06_KSTATE)", "Kabaddi"],
    ])

    rows = svc.parse_workbook(path, default_board="Karnataka")

    subjects = {r.subject for r in rows}
    assert "Physical Education" in subjects
    assert not any(s.lower().startswith("chapter") for s in subjects)
    assert {r.chapter for r in rows} >= {"Kabaddi", "Patterns in Mathematics"}


def test_tracker_sheets_are_not_syllabus(tmp_path):
    path = tmp_path / "with_status.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Grade 06"
    for row in [
        ["Subject", "Unit", "Chapter"],
        ["Mathematics", "Number System", "Patterns in Mathematics"],
    ]:
        ws.append(row)
    tracker = wb.create_sheet("Maths Chapter wise status")
    for row in [
        ["Grade", "Chapter Name", "Concept Mapping", "Assessments"],
        [10, "Real Numbers", False, True],
    ]:
        tracker.append(row)
    wb.save(path)

    rows = svc.parse_workbook(path, default_board="Karnataka")

    assert {r.chapter for r in rows} == {"Patterns in Mathematics"}


def test_refresh_migrates_a_refiled_chapter_and_keeps_its_work(
    isolated_db, tmp_path, monkeypatch,
):
    """A chapter that changes subject must MOVE, not be stranded or dropped.

    Karnataka History becomes Social Science. Leaving the old row behind
    kept a subject the syllabus no longer has in the dropdown; deleting it
    would have destroyed authored concepts.
    """
    from app.services import directory

    path = tmp_path / "kstate.xlsx"
    _write_xlsx(path, [
        ["Subject", "Unit", "Chapter"],
        ["Social Science", "Ancient India (06_KSTATE)", "India - Our Pride"],
    ], sheet_name="Grade 06")
    monkeypatch.setattr(svc, "_discover_workbooks", lambda: [path])
    monkeypatch.setattr(svc, "_missing_expected_files", lambda: [])
    monkeypatch.setattr(
        svc, "_infer_file_options", lambda name: {"default_board": "Karnataka"})

    legacy = models.Chapter(
        chapter_code="06KAHI_India_Our_Pride",
        board="Karnataka", grade="06", subject="History",
        unit="Ancient India (06_KSTATE)", chapter_title="India - Our Pride",
        chapter_display_name="India - Our Pride",
    )
    isolated_db.add(legacy)
    isolated_db.commit()
    legacy_id = legacy.id
    isolated_db.add(models.Topic(
        topic_title="Authored Topic", pre_post_learning="Post",
        chapter_id=legacy_id))
    isolated_db.commit()

    result = svc.refresh_syllabus(isolated_db)
    isolated_db.expire_all()

    moved = isolated_db.get(models.Chapter, legacy_id)
    assert moved is not None, "the chapter must not be deleted"
    assert moved.subject == "Social Science"
    assert len(moved.topics) == 1, "authored work travels with the chapter"
    assert result["migrated"] == 1
    assert result["pruned"] == 0
    assert isolated_db.query(models.Chapter).count() == 1
