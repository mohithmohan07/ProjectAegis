"""The converged release core — spec-step8 T1/T2/T14/T15, slice S6.

Two release systems became one. ``models.AssessmentRelease`` is the
immutable row of record for all four outputs; ``build_concepts_release``
stays the staging and evidence layer; the three §4 names are the only
public vocabulary; and one Build Concepts run produces all four outputs
(owner ruling OD1).

The tests here pin the five properties that convergence is, plus the two
that keep it from costing anything:

* one row per LANE per run, four projections;
* a re-stage after a freeze mints version 2 and supersedes version 1,
  instead of silently overwriting a slot;
* the row records the lane, the profile and the layout its run executed
  under, in a column that was declared and never written;
* ``release_state`` is one vocabulary over BOTH payload shapes;
* the table name does not move, because renaming it is a migration this
  repo cannot express;

and then the two safety properties:

* an unplaced candidate still blocks the assessment-lane upload and still
  leaves every download open (the S6→S8 hand-off pin);
* an assessment-lane failure never costs the concept outputs (Q13).
"""
from __future__ import annotations

import copy
import threading
import uuid
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy.orm.attributes import flag_modified

from app import models
from app.api import build_assessments as build_assessments_api
from app.bulk_import import assessment_workbook, layouts
from app.services import assessment_release as rel
from app.services import assessment_release_run as run
from app.services import assessment_release_service as svc
from app.services import build_concepts_release as release
from app.services import build_concepts_release_contract as release_contract
from app.services import build_concepts_release_files as release_files
from app.services import (
    build_concepts_terminal_release_contract as terminal_release,
)
from app.services import build_concepts_release_manifest as release_manifest
from app.services import progress
from app.services import release_core
from app.services.phase3 import premap

from tests.test_assessment_release_run import (
    OWNER,
    _authorities,
    _chapter_with_concepts,
    _decision_context,
    _make_job,
)
from tests.test_assessment_pre_release_lane import (
    _cells,
    _generated_authorities,
)
from tests.test_pre_release_lane_wiring import _both_lanes_job


# --------------------------------------------------------------------------- #
# 1. One immutable row per lane per run
# --------------------------------------------------------------------------- #

def _run_both_lanes(db, job, chapter, *, questions=2):
    """Both lanes of one job, exactly as the run now builds them."""

    authorities, _ = _authorities(db, chapter)
    post = run.run_release_for_job(
        db, job.id, owner_sub=OWNER, authorities=authorities,
        **_decision_context())
    generated, _ = _generated_authorities()
    concept_key = next(iter(
        run.release_snapshot.build(
            db, job, release.release_payload(job, lane=release.LANE_PRE)
        )["concept_records_by_key"]
    ))
    pre = run.run_pre_release_for_job(
        db,
        job.id,
        owner_sub=OWNER,
        blueprint_cells=_cells(questions, concept_key),
        authorities=generated,
        **_decision_context(),
    )
    return pre, post


def _published_layout_id(release_row, filename: str) -> str:
    path = Path(release_row.publication["directory"]) / filename
    workbook = openpyxl.load_workbook(
        path, data_only=True, read_only=True,
    )
    try:
        headers = {
            name: next(
                workbook[name].iter_rows(
                    min_row=2, max_row=2, values_only=True,
                ),
                (),
            )
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()
    return layouts.identify_workbook(headers).layout_id


def _audit_post_authorities(db, chapter, *, mathematics: bool):
    authorities, _ = _authorities(db, chapter)
    cell_author, cell_critic = authorities["cells"]

    def audit_cell_author(payload):
        result = cell_author(payload)
        if result["sheet_kind"] == "descriptive":
            result["question_category"] = "Short Answer Type (3 Marks)"
        return result

    authorities["cells"] = (audit_cell_author, cell_critic)
    marking_author, marking_critic = authorities["marking"]
    if mathematics:

        def audit_marking_author(payload):
            result = marking_author(payload)
            if payload["candidate"]["sheet_kind"] == "objective":
                result["question_duration"] = 1
            return result

    else:
        # Since the owner's Question Duration Matrix (2026-08-29) the
        # Grade-6 English policy prescribes exact minutes too, so the stub
        # author reads the prescribed value off the profile contract.
        from app.services import assessment_profile

        english_meta = {
            "board": "MSBSHSE", "grade": "6", "subject": "English",
        }

        def audit_marking_author(payload):
            result = marking_author(payload)
            cell = payload["blueprint_evidence"]["explicit_blueprint_cell"]
            prescribed = assessment_profile.question_duration_minutes(
                None,
                english_meta,
                sheet_kind=str(cell.get("sheet_kind") or ""),
                question_category=str(cell.get("question_category") or ""),
                difficulty=str(cell.get("difficulty") or ""),
                marks=cell.get("marks"),
            )
            if prescribed is not None:
                result["question_duration"] = prescribed
                result["duration_basis_count"] = None
            return result

    authorities["marking"] = (audit_marking_author, marking_critic)
    return authorities


def test_one_run_mints_one_release_per_lane_with_four_projections(db):
    """Two rows, four projections (T2) — and the lane is on the row.

    Before this slice the two lanes' releases were indistinguishable in
    the database: nothing on the row said which lane it was the record
    of, so the manifest could not address either Master File and the
    supersede chain could not tell a Pre re-run from a Post one.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    pre, post = _run_both_lanes(db, job, chapter)

    rows = (
        db.query(models.AssessmentRelease)
        .filter(models.AssessmentRelease.job_id == job.id)
        .all()
    )
    assert len(rows) == 2
    assert {row.lane for row in rows} == {release.LANE_PRE, release.LANE_POST}
    assert pre.lane == release.LANE_PRE
    assert post.lane == release.LANE_POST
    assert pre.release_uid != post.release_uid, (
        "two lanes are two publications (Rule G), never two versions of one"
    )
    # Four projections: each row serves a Concept file and a Master file.
    for row in (pre, post):
        assert svc.published_artifact(row, svc.CONCEPTS_FILENAME)
        assert svc.published_artifact(row, svc.MASTER_FILENAME)
    # And the lane resolver finds exactly one live row per lane.
    assert release_core.latest_release_for_lane(
        db, job.id, release.LANE_PRE).id == pre.id
    assert release_core.latest_release_for_lane(
        db, job.id, release.LANE_POST).id == post.id


def test_restaging_after_freeze_mints_version_2_and_leaves_version_1_superseded(
    db,
):
    """The concrete hazard T2 names, closed on both sides.

    Before: ``stage_release`` OVERWROTE the job's slot
    (``durable_inventory[RELEASE_KEY] = payload``) and every run minted an
    unrelated ``release_uid``. So a reviewer could publish the Master
    File, then ``force_release`` — which refuses only when
    ``job.status == "generated"`` — and the staged payload would move
    underneath the frozen row with nothing on either side recording it.

    After: the staging slot MINTS a version, and the second run of the
    same lane on the same job supersedes the first instead of orphaning
    it. Both halves are asserted, because either alone leaves the hazard
    half-closed.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    authorities, _ = _authorities(db, chapter)

    assert release.release_payload(job)[release.STAGED_VERSION_FIELD] == 1
    first = run.run_release_for_job(
        db, job.id, owner_sub=OWNER, authorities=authorities,
        **_decision_context())
    assert first.version == 1
    assert first.provider_identity["staged_release_version"] == 1

    # The re-stage. The slot mints v2 rather than replacing v1 unseen.
    release.stage_release(
        db,
        job,
        target_chapter_id=chapter.id,
        records=release.release_payload(job)["records"],
        inventory=copy.deepcopy(job.question_inventory),
        reason="the reviewer explicitly re-released",
    )
    db.refresh(job)
    assert release.release_payload(job)[release.STAGED_VERSION_FIELD] == 2

    second = run.run_release_for_job(
        db, job.id, owner_sub=OWNER, authorities=authorities,
        **_decision_context())
    db.refresh(first)
    assert second.version == 2
    assert second.release_uid == first.release_uid
    assert first.state == "superseded"
    assert first.superseded_at is not None
    assert second.provider_identity["staged_release_version"] == 2
    # The superseded row keeps its artifacts — append-only, never edited —
    # but the reviewer's manifest points at the live one.
    assert release_core.latest_release_for_lane(
        db, job.id, release.LANE_POST).id == second.id


def test_the_release_row_carries_its_lane_the_profile_and_the_layout_the_run_executed_under(
    db,
):
    """§3 guard 2's run-context pinning, in a column nothing ever wrote.

    [measured] before this slice ``grep -rn provider_identity app/
    --include=*.py`` returned exactly one hit — the declaration at
    ``models.py`` — and ``grep -c profile
    app/services/assessment_release_service.py`` returned 0. A release
    that does not record the profile and layout it executed under cannot
    be re-read safely under a different one, and the provider spend OD1
    commits every chapter to could not be measured from the row.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    pre, post = _run_both_lanes(db, job, chapter)

    for row, lane in ((pre, release.LANE_PRE), (post, release.LANE_POST)):
        assert row.lane == lane
        assert row.layout_id == release_core.layout_id() == "sop-mes-1"
        identity = row.provider_identity
        assert identity["lane"] == lane
        assert identity["layout_id"] == "sop-mes-1"
        assert identity["workbook_outputs"] == {
            "concepts_xlsx": {
                "layout_id": "sop-mes-1",
                "contract_id": "concept-reference-1",
            },
            "master_xlsx": {
                "layout_id": "sop-mes-1",
                "contract_id": "reference-master-1",
            },
        }
        assert identity["profile"], "the run's profile name is recorded"
        assert identity["job_id"] == job.id
        assert identity["staged_release_version"] >= 1


@pytest.mark.parametrize(
    ("subject", "expected_layout_id"),
    [
        pytest.param(
            "Mathematics",
            layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
            id="mathematics",
        ),
        pytest.param(
            "English",
            layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
            id="english-post",
        ),
    ],
)
def test_release_persists_the_rendered_dynamic_master_layout(
    db, subject, expected_layout_id,
):
    chapter = _chapter_with_concepts(db)
    original_metadata = (chapter.board, chapter.grade, chapter.subject)
    try:
        chapter.board = "MSBSHSE"
        chapter.grade = "06"
        chapter.subject = subject
        db.commit()
        job = _make_job(db, chapter)
        authorities = _audit_post_authorities(
            db, chapter, mathematics=subject == "Mathematics",
        )

        released = run.run_release_for_job(
            db,
            job.id,
            owner_sub=OWNER,
            authorities=authorities,
            **_decision_context(),
        )

        rendered_layout_id = _published_layout_id(
            released, svc.MASTER_FILENAME,
        )
        assert rendered_layout_id == expected_layout_id
        assert released.layout_id == rendered_layout_id
        assert released.provider_identity["layout_id"] == rendered_layout_id
        assert released.provider_identity["workbook_outputs"] == {
            "concepts_xlsx": {
                "layout_id": layouts.REFERENCE_LAYOUT_ID,
                "contract_id": "concept-reference-1",
            },
            "master_xlsx": {
                "layout_id": rendered_layout_id,
                "contract_id": expected_layout_id,
            },
        }
        assert _published_layout_id(
            released, svc.CONCEPTS_FILENAME,
        ) == layouts.REFERENCE_LAYOUT_ID
    finally:
        chapter.board, chapter.grade, chapter.subject = original_metadata
        db.commit()


def test_the_table_name_is_unchanged():
    """A pin against a well-meant later rename.

    The table is called ``assessment_releases`` and now holds the release
    of record for the CONCEPT lanes too, so the name is a misnomer. It
    stays anyway: ``db._ensure_columns`` emits only ``ALTER TABLE … ADD
    COLUMN``, so a rename is not expressible — ``create_all`` would mint a
    new empty table and orphan every published release and its receipt.
    Adding COLUMNS is expressible, and that is what S6 did.
    """

    assert models.AssessmentRelease.__tablename__ == "assessment_releases"
    columns = set(models.AssessmentRelease.__table__.columns.keys())
    assert {"lane", "layout_id", "provider_identity", "job_id"} <= columns


# --------------------------------------------------------------------------- #
# 2. ONE public vocabulary over BOTH payload shapes
# --------------------------------------------------------------------------- #

def test_release_state_is_the_same_vocabulary_for_both_payload_shapes(db):
    """The defect that would have broken the core on arrival.

    [measured] before this slice
    ``release_state({"source_atoms": …, "candidates": …, "groups": …})``
    returned ``diagnostic_release``, because ``structural_defects``
    counted A's ``records`` and B's payload has none. Landed as it was,
    every healthy assessment release was labelled structurally corrupt the
    moment the shared core started asking.
    """

    concept_payload = {
        "records": [{"topic": "Solids", "concept_title": "What is a solid"}],
        "summary": {"issue_count": 0, "affected_row_count": 0},
    }
    assessment_payload = {
        "source_atoms": [{"atom_id": "QINV-0001"}],
        "candidates": [{
            "candidate_id": "CAND-1",
            "concept_key": "db:1",
            "group_key": "(CON-1) BG01",
        }],
        "groups": [{"group_key": "(CON-1) BG01", "concept_key": "db:1"}],
    }

    assert release.release_state(concept_payload) == release.READY
    assert release.release_state(assessment_payload) == release.READY
    assert release.structural_defects(assessment_payload) == []
    # An assessment payload with no depositable candidate IS diagnostic,
    # for the same reason a concept payload with no depositable row is.
    empty = dict(assessment_payload, candidates=[])
    assert release.release_state(empty) == release.DIAGNOSTIC_RELEASE
    homeless = dict(
        assessment_payload,
        candidates=[{"candidate_id": "CAND-1", "concept_key": "", "group_key": ""}],
    )
    assert release.release_state(homeless) == release.DIAGNOSTIC_RELEASE
    # And the three names are the whole vocabulary, for both systems.
    assert set(release_core.RELEASE_STATES) == {
        release.READY, release.READY_WITH_FLAGS, release.DIAGNOSTIC_RELEASE,
    }


def test_the_release_summary_states_the_release_and_keeps_the_diagnostics(db):
    """``release_state`` becomes public; ``state``/``readiness`` stay beside it.

    T1's resolution is a demotion, not a deletion: B's 7-state lifecycle
    and its 3-value readiness keep doing exactly what they did — readiness
    is still the only thing refusing the assessment-lane database write —
    they just stop being a second PUBLIC name for the same run.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _pre, post = _run_both_lanes(db, job, chapter)

    summary = build_assessments_api._release_summary(post)
    assert summary["release_state"] in release_core.RELEASE_STATES
    assert summary["release_state"] == release_core.release_state(post)
    assert summary["state"] == post.state
    assert summary["readiness"] == post.diagnostics["readiness"]
    assert summary["lane"] == release.LANE_POST


def test_the_readiness_verdict_is_an_input_to_the_projection_not_a_rival(db):
    """B's readiness names the state; it does not name it SEPARATELY."""

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _pre, post = _run_both_lanes(db, job, chapter)

    post.diagnostics = dict(post.diagnostics, readiness=svc.BLOCKED)
    db.commit()
    assert release_core.release_state(post) == release.DIAGNOSTIC_RELEASE

    post.diagnostics = dict(
        post.diagnostics, readiness=svc.RELEASED_WITH_WARNINGS)
    db.commit()
    assert release_core.release_state(post) == release.READY_WITH_FLAGS

    post.diagnostics = dict(post.diagnostics, readiness=svc.READY)
    db.commit()
    assert release_core.release_state(post) == release.READY
    # The constants this module reads are the service's own.
    assert release_core._BLOCKED == svc.BLOCKED
    assert release_core._RELEASED_WITH_WARNINGS == svc.RELEASED_WITH_WARNINGS


# --------------------------------------------------------------------------- #
# 3. The S6 → S8 hand-off: the consequence of an unplaced candidate
# --------------------------------------------------------------------------- #

def test_an_unplaced_candidate_still_blocks_the_upload_through_readiness(
    db, client,
):
    """UNCHANGED behaviour, pinned so the hand-off to S8 is safe.

    ``_readiness``'s read of ``manifest["issues"]["unplaced"]`` is the ONLY
    consumer of that list in the tree, and ``if readiness == BLOCKED`` is
    the ONLY thing refusing the assessment-lane database write. Its
    replacement verdict does not land until S8, in the SAME commit that
    deletes this read. Deleting it here would leave a window in which a
    candidate appearing in no data row of any of the four outputs
    publishes to the database.

    So this test is written now and stays green unchanged through S8,
    where the REASON changes — from the renderer's ``unplaced`` list to
    the staged ``unresolved_question_home`` verdict — and the CONSEQUENCE
    does not: database write refused, every download available.
    """

    from tests.test_mes_release_lifecycle import _fresh_release

    def orphan(payload):
        payload["candidates"][0]["group_key"] = "(unknown) BG01"

    release_row, _payload, _label = _fresh_release(db, mutate=orphan)
    published = svc.publish_release(db, release_row)

    assert published.diagnostics["readiness"] == svc.BLOCKED
    with pytest.raises(svc.UploadRefused, match="blocked"):
        svc.upload_master_to_database(db, published, owner_sub=OWNER)
    for filename in ("master.xlsx", "concepts.xlsx"):
        response = client.get(
            f"/build-assessments/releases/{published.id}/{filename}")
        assert response.status_code == 200, filename


def test_no_code_path_decides_on_unplaced_and_the_refusal_survives(
    db, client,
):
    """S8's replacement for S6's ``…_is_still_there`` pin (T7.5 item 5).

    Two halves, deliberately in ONE test so neither can land alone.

    Half one: no site outside ``assessment_workbook``'s own producers
    READS ``unplaced``. The key is still WRITTEN and still durable —
    ``publish_release`` puts the whole issues manifest on disk as
    ``manifest.json`` — so what is gone is the DECIDER, not the record.
    "Has no consumer" would overstate it and contradict T7.5 item 4.

    Half two: the same candidate is STILL refused, now by the staged
    ``unresolved_question_home`` verdict, and every download still ships.
    """

    import subprocess
    from pathlib import Path

    service = Path(
        "app/services/assessment_release_service.py"
    ).read_text(encoding="utf-8")
    assert '(manifest.get("issues") or {}).get("unplaced") or []' not in (
        service
    )
    assert "if readiness == BLOCKED" in service

    def _grep(pattern: str) -> list[str]:
        return subprocess.run(
            ["grep", "-rn", pattern, "app/", "--include=*.py"],
            capture_output=True, text=True, check=False,
        ).stdout.splitlines()

    # No READ of the key anywhere in the tree, in either indexing form.
    assert _grep(r'get("unplaced")\|\["unplaced"\]') == []
    # The PRODUCER is still there — the evidence channel must not be
    # deleted along with the decider.
    renderer = Path(
        "app/bulk_import/assessment_workbook.py"
    ).read_text(encoding="utf-8")
    assert "unplaced.append(" in renderer
    assert '"unplaced": unplaced' in renderer

    from tests.test_mes_release_lifecycle import _fresh_release

    def orphan(payload):
        payload["candidates"][0]["group_key"] = "(unknown) BG01"

    release_row, _payload, label = _fresh_release(db, mutate=orphan)
    assert any(
        error.startswith(rel.UNRESOLVED_QUESTION_HOME)
        for error in release_row.diagnostics["payload_errors"])
    published = svc.publish_release(db, release_row)
    assert published.diagnostics["readiness"] == svc.BLOCKED
    with pytest.raises(svc.UploadRefused, match="blocked"):
        svc.upload_master_to_database(db, published, owner_sub=OWNER)
    for filename in ("master.xlsx", "concepts.xlsx"):
        response = client.get(
            f"/build-assessments/releases/{published.id}/{filename}")
        assert response.status_code == 200, filename
    # The evidence still ships.
    assert [
        item["question_label"]
        for item in published.diagnostics["issues"]["unplaced"]
    ] == [label]


# --------------------------------------------------------------------------- #
# 4. The manifest enumerates all four outputs, in the owner's order (T14/P16)
# --------------------------------------------------------------------------- #

_OUTPUT_ORDER = [
    "pre_release_bulk_import",   # 01 Pre-Learning Concept File
    "pre_release_master",        # 02 Pre-Learning Master File
    "release_bulk_import",       # 03 Post-Learning Concept File
    "release_master",            # 04 Post-Learning Master File
]


def _manifest_blocks(job):
    """All four manifest blocks — the eager pair and the lazy pair.

    The lazy module MONKEYPATCHES the eager one, so an entry added to one
    alone is a silent no-op. The eager half must be reached through
    ``eager_release_artifact_entries``: ``install()`` rebinds the module
    attribute for the rest of the process.
    """

    return {
        "eager": release_files.eager_release_artifact_entries(job),
        "lazy": release_manifest.release_artifact_entries(job),
    }


def test_the_manifest_enumerates_all_four_outputs_in_both_lanes(db):
    """P16 closed at its root: four outputs, one place, the owner's order.

    S3 added Outputs 01 and 03 and left 02/04 as present-and-disabled
    placeholders precisely so the ORDER was fixed from that slice on. S6
    resolves them through the release row — they must land at positions 02
    and 04, not appended at the end.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    for name, entries in _manifest_blocks(job).items():
        kinds = [row["kind"] for row in entries]
        assert kinds[:4] == _OUTPUT_ORDER, name
        by_kind = {row["kind"]: row for row in entries}
        # Present and disabled with a STATED reason before the lane has a
        # release row — never absent, because an absent entry is the
        # defect this manifest exists to close.
        for kind in ("pre_release_master", "release_master"):
            assert by_kind[kind]["disabled"] is True, (name, kind)
            assert by_kind[kind]["disabled_reason"], (name, kind)

    # Now build both lanes, and both Master entries become servable.
    pre, post = _run_both_lanes(db, job, chapter)
    db.refresh(job)
    for name, entries in _manifest_blocks(job).items():
        by_kind = {row["kind"]: row for row in entries}
        assert [row["kind"] for row in entries][:4] == _OUTPUT_ORDER, name
        for kind, row_release in (
            ("pre_release_master", pre), ("release_master", post),
        ):
            entry = by_kind[kind]
            assert not entry.get("disabled"), (name, kind)
            assert entry["download_url"] == (
                f"/build-assessments/releases/{row_release.id}/master.xlsx"
            ), (name, kind)
            assert entry["release_state"] in release_core.RELEASE_STATES


def test_the_master_entries_are_served_by_the_route_they_advertise(db, client):
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _run_both_lanes(db, job, chapter)
    db.refresh(job)

    entries = release_manifest.release_artifact_entries(job)
    for row in entries:
        if row["kind"] in ("pre_release_master", "release_master"):
            response = client.get(row["download_url"])
            assert response.status_code == 200, row["kind"]
            assert response.headers["content-type"].startswith(
                "application/vnd.openxmlformats")


# --------------------------------------------------------------------------- #
# 5. OD1 — one run, four outputs — and T15-2's containment
# --------------------------------------------------------------------------- #

def test_one_build_concepts_run_produces_all_four_outputs(db, monkeypatch):
    """The owner's ruling OD1, at the call site the spec names.

    ``_run_generation_release`` stages the two concept lanes and then
    calls ``_build_master_siblings`` ONCE, on the tail all four exits
    converge on and OUTSIDE the ``_RELEASE_MODE`` / ``_RELEASE_CAPTURE``
    ``finally`` — so the deposit interceptor cannot see the assessment
    lane.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    seen = []

    def spy(name, original):
        def call(_db, job_id, **kwargs):
            seen.append((name, job_id, _release_mode_visible()))
            return original(_db, job_id, **kwargs)
        return call

    def _release_mode_visible():
        return release_contract._RELEASE_MODE.get()

    monkeypatch.setattr(
        run, "run_release_for_job",
        spy("post", _stub_release(db, chapter, release.LANE_POST)))
    monkeypatch.setattr(
        run, "run_pre_release_for_job",
        spy("pre", _stub_release(db, chapter, release.LANE_PRE)))

    def original(_db, _job_id, _chapter_id, **_kwargs):
        return {"status": "ok"}

    events: list[dict] = []
    sink_token = progress._sink.set(events.append)
    try:
        release_contract._run_generation_release(
            original, db, job.id, chapter.id, owner_sub=OWNER)
    finally:
        progress._sink.reset(sink_token)

    started_lanes = [name for name, _job, _mode in seen]
    assert len(started_lanes) == 2 and set(started_lanes) == {"pre", "post"}, (
        "both Master lanes are built exactly once; their concurrent worker "
        "start order is intentionally not a contract"
    )
    assert all(job_id == job.id for _name, job_id, _mode in seen)
    assert not any(mode for _name, _job, mode in seen), (
        "the assessment lane runs OUTSIDE _RELEASE_MODE, so the deposit "
        "interceptor cannot see it"
    )
    master_steps = [
        event.get("label") for event in events
        if event.get("type") == "step"
        and event.get("label") == "Building Master files (Outputs 02/04)"
    ]
    assert master_steps == ["Building Master files (Outputs 02/04)"], (
        "the Master fan-out needs one real stage boundary so its usage and "
        "cache evidence do not land on the preceding concept stage"
    )
    # The two concept outputs are staged and durable.
    assert release.release_payload(job) is not None
    assert release.release_payload(job, lane=release.LANE_PRE) is not None


def test_master_sibling_builds_overlap_on_separate_sessions(db, monkeypatch):
    """The two-worker Master fan-out is real concurrency, not just wiring."""

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    barrier = threading.Barrier(2)
    lock = threading.Lock()
    sessions: dict[str, object] = {}
    active = 0
    max_active = 0

    class _Stub:
        def __init__(self, lane):
            self.id = f"REL-{lane}"

    def rebuild(lane_db, job_id, lane, *, owner_sub=None, stage_progress=None):
        nonlocal active, max_active
        assert job_id == job.id
        assert owner_sub == OWNER
        with lock:
            sessions[lane] = lane_db
            active += 1
            max_active = max(max_active, active)
        try:
            barrier.wait(timeout=5)
            return _Stub(lane)
        finally:
            with lock:
                active -= 1

    monkeypatch.setattr(release_contract, "rebuild_lane_master", rebuild)

    built = release_contract._build_master_siblings(
        db, job.id, chapter.id, owner_sub=OWNER,
    )

    assert max_active == 2
    assert built == {
        release.LANE_PRE: {"release_id": "REL-pre"},
        release.LANE_POST: {"release_id": "REL-post"},
    }
    assert set(sessions) == {release.LANE_PRE, release.LANE_POST}
    assert sessions[release.LANE_PRE] is not db
    assert sessions[release.LANE_POST] is not db
    assert sessions[release.LANE_PRE] is not sessions[release.LANE_POST]


def _record_nonterminal_verdict(payload: dict) -> None:
    """Record the explicit staged verdict a non-terminal exit writes."""
    payload[terminal_release.TERMINAL_GENERATION_FIELD] = False
    summary = dict(payload.get("summary") or {})
    summary[terminal_release.TERMINAL_GENERATION_FIELD] = False
    payload["summary"] = summary


def test_nonterminal_concept_checkpoint_skips_all_master_spend(
    db, monkeypatch,
):
    """Job 81's 81% release is evidence, not Master-authoring authority."""

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    inventory = copy.deepcopy(job.question_inventory)

    post = inventory[release.release_key_for_lane(release.LANE_POST)]
    post["checkpoint_stage"] = "pre_type_assignment"
    post.setdefault("issues", []).append({
        "phase": "generation",
        "severity": "error",
        "message": "provider rejected the JSON-mode request",
    })
    # Restructure A: a run that exits non-terminally records that verdict
    # at staging; the gate reads the recorded fact, so the simulation must
    # record it too (checkpoint echoes alone no longer decide).
    _record_nonterminal_verdict(post)
    pre = inventory[release.release_key_for_lane(release.LANE_PRE)]
    pre.setdefault("snapshot_defects", []).append(
        "terminal_generation_incomplete: Post stopped at "
        "pre_type_assignment"
    )
    _record_nonterminal_verdict(pre)
    job.question_inventory = inventory
    flag_modified(job, "question_inventory")
    db.commit()

    called: list[str] = []

    def should_not_build(*_args, **_kwargs):
        called.append("provider")
        raise AssertionError("a non-terminal release reached Master generation")

    monkeypatch.setattr(release_contract, "rebuild_lane_master", should_not_build)
    events: list[dict] = []
    sink_token = progress._sink.set(events.append)
    try:
        built = release_contract._build_master_siblings(
            db, job.id, chapter.id, owner_sub=OWNER,
        )
    finally:
        progress._sink.reset(sink_token)

    assert built == {release.LANE_PRE: None, release.LANE_POST: None}
    assert called == []
    assert not any(
        event.get("label") == "Building Master files (Outputs 02/04)"
        for event in events
    )


def test_explicit_master_rebuild_rejects_nonterminal_concept_pre_spend(
    db, monkeypatch,
):
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    inventory = copy.deepcopy(job.question_inventory)
    post = inventory[release.release_key_for_lane(release.LANE_POST)]
    post["checkpoint_stage"] = "pre_type_assignment"
    _record_nonterminal_verdict(post)
    job.question_inventory = inventory
    flag_modified(job, "question_inventory")
    db.commit()

    called: list[str] = []

    def should_not_run(*_args, **_kwargs):
        called.append("provider")
        raise AssertionError("provider call escaped terminality gate")

    monkeypatch.setattr(run, "run_release_for_job", should_not_run)
    with pytest.raises(run.ReleaseRunError, match="Resume Concept generation"):
        release_contract.rebuild_lane_master(
            db,
            job.id,
            release.LANE_POST,
            owner_sub=OWNER,
        )
    assert called == []


def test_staging_records_the_verdict_and_checkpoint_drift_cannot_block(db):
    """Restructure A: the verdict is decided once, at staging.

    [measured] job 'Patterns' (2026-08-29): a fully completed run had both
    Concept files staged and healthy, yet every Master click was refused
    because eligibility re-derived terminality from checkpoint echoes.
    After staging records the verdict, a live checkpoint that later
    drifts, is cleared, or re-validates differently can never re-open
    the question.
    """
    terminal_release.install()
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    for lane in release.RELEASE_LANES:
        payload = release.release_payload(job, lane=lane)
        assert payload[terminal_release.TERMINAL_GENERATION_FIELD] is True, lane
        summary = payload.get("summary") or {}
        assert summary[terminal_release.TERMINAL_GENERATION_FIELD] is True, lane

    # Drift the live checkpoint to a mid-run stage AND clear it entirely:
    # neither may block a Master built from the staged release.
    for drifted in ({"stage": "pre_type_assignment"}, None):
        job.generation_checkpoint = drifted
        flag_modified(job, "generation_checkpoint")
        db.commit()
        for lane in release.RELEASE_LANES:
            eligible, reason = release_contract._lane_master_eligibility(
                db, job.id, lane, owner_sub=OWNER,
            )
            assert eligible is True, (lane, drifted, reason)


def test_legacy_payload_with_live_terminal_checkpoint_backfills_true(db):
    """The 'Patterns' staging-order race, retroactively unblocked.

    A payload staged before Restructure A carries no explicit verdict and
    can echo a mid-run ``checkpoint_stage`` (the final checkpoint write
    landed after the release was staged). When the LIVE checkpoint says
    ``final_content_ready``, the one-time backfill records True and the
    Masters build.
    """
    terminal_release.install()
    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    inventory = copy.deepcopy(job.question_inventory)
    for lane in release.RELEASE_LANES:
        payload = inventory[release.release_key_for_lane(lane)]
        payload.pop(terminal_release.TERMINAL_GENERATION_FIELD, None)
        summary = dict(payload.get("summary") or {})
        summary.pop(terminal_release.TERMINAL_GENERATION_FIELD, None)
        payload["summary"] = summary
        payload["checkpoint_stage"] = "pre_type_assignment"
    job.question_inventory = inventory
    flag_modified(job, "question_inventory")
    job.generation_checkpoint = {"stage": "final_content_ready"}
    flag_modified(job, "generation_checkpoint")
    db.commit()

    for lane in release.RELEASE_LANES:
        eligible, reason = release_contract._lane_master_eligibility(
            db, job.id, lane, owner_sub=OWNER,
        )
        assert eligible is True, (lane, reason)

    # The backfill stamped the fact durably: clearing the live checkpoint
    # afterwards no longer matters.
    db.refresh(job)
    for lane in release.RELEASE_LANES:
        payload = release.release_payload(job, lane=lane)
        assert payload[terminal_release.TERMINAL_GENERATION_FIELD] is True, lane
    job.generation_checkpoint = None
    flag_modified(job, "generation_checkpoint")
    db.commit()
    for lane in release.RELEASE_LANES:
        eligible, reason = release_contract._lane_master_eligibility(
            db, job.id, lane, owner_sub=OWNER,
        )
        assert eligible is True, (lane, reason)


def _stub_release(db, chapter, lane):
    """A cheap stand-in for one lane's full assessment pipeline."""

    def build(_db, job_id, **kwargs):
        row = models.AssessmentRelease(
            release_uid=f"REL-{uuid.uuid4().hex[:8]}",
            version=1,
            owner_sub=OWNER,
            job_id=job_id,
            lane=lane,
            layout_id=release_core.layout_id(),
            state="materialized",
        )
        _db.add(row)
        _db.commit()
        _db.refresh(row)
        return row

    return build


# The five exception types this call site can actually see today, each
# imported from where it is raised rather than re-declared here — a
# stand-in exception would prove nothing about the real ones.
_REAL_EXCEPTIONS = [
    pytest.param(
        lambda: run.ReleaseRunError("the staged release has no inventory"),
        id="ReleaseRunError"),
    pytest.param(
        lambda: run.SourceQuestionLeak("a chapter QID reached the Pre lane"),
        id="SourceQuestionLeak"),
    pytest.param(
        lambda: svc.UploadRefused("this release is already being published"),
        id="UploadRefused"),
    pytest.param(
        lambda: assessment_workbook.WorkbookRenderError(
            "the group has no home concept"),
        id="WorkbookRenderError"),
    pytest.param(
        lambda: premap.PreExtractionError("the capture could not be read"),
        id="PreExtractionError"),
]

# The four exits ``_run_generation_release`` has, each a
# ``stage_release`` / ``_stage_pre_sibling`` pair, each ending in
# ``return staged``. The call site is on the single tail they converge
# on, so a fifth exit added later inherits it — and this parametrisation
# is what fails if someone returns from a fifth place instead.
_EXITS = ["clean_captured", "clean_checkpoint", "raised_captured",
          "raised_checkpoint"]


def _drive_exit(db, job, chapter, exit_name):
    """Force one of the four exits and return the staged result."""

    def original(_db, _job_id, _chapter_id, **_kwargs):
        if exit_name in ("clean_captured", "raised_captured"):
            release_contract._RELEASE_CAPTURE.set({
                "records": [{
                    "topic": "Solids",
                    "concept_title": "What makes a solid",
                    "concept_details": (
                        "Description: A solid keeps its own shape.\n"
                        "Achieving Mastery: name three solids."
                    ),
                    "keywords": "solid, shape",
                }],
                "inventory": copy.deepcopy(job.question_inventory),
                "mined_types": {},
                "final_grounding_certificate": {"version": "test"},
                "checkpoint": {},
            })
        if exit_name.startswith("raised"):
            raise RuntimeError(f"generation failed at {exit_name}")
        return {"status": "ok"}

    return release_contract._run_generation_release(
        original, db, job.id, chapter.id, owner_sub=OWNER)


@pytest.mark.parametrize("exit_name", _EXITS)
@pytest.mark.parametrize("make_error", _REAL_EXCEPTIONS)
def test_an_assessment_lane_failure_does_not_cost_the_concept_outputs(
    db, client, monkeypatch, exit_name, make_error,
):
    """T15-2, and it is a Q13 requirement rather than a robustness nicety.

    By the time ``_build_master_siblings`` runs, the two CONCEPT outputs
    are finished and durable. An implementer who "simplifies" by letting
    the exception escape turns one Master-lane fault into all four outputs
    lost — a mid-run halt after the model budget is spent, which
    CLAUDE.md and Q13 forbid.

    Completed Concept exits still record every Master failure. A generation
    failure exit is different: its non-terminal Concept payload must skip the
    Master runner before the injected failure (or any provider spend) can
    occur. Both paths preserve the downloadable Concept evidence.
    """

    chapter = _chapter_with_concepts(db)
    control_job = _both_lanes_job(db, chapter)
    control = _drive_exit(db, control_job, chapter, exit_name)
    assert control["status"] == release.RELEASE_STATUS
    control_rows = _bulk_import_rows(client, control_job)
    control_states = {
        lane: release.release_state(
            release.release_payload(control_job, lane=lane))
        for lane in release.RELEASE_LANES
    }

    job = _both_lanes_job(db, chapter)

    def explode(_db, _job_id, **_kwargs):
        raise make_error()

    monkeypatch.setattr(run, "run_release_for_job", explode)
    monkeypatch.setattr(run, "run_pre_release_for_job", explode)

    # (i) nothing escapes.
    staged = _drive_exit(db, job, chapter, exit_name)
    assert staged["status"] == release.RELEASE_STATUS
    db.refresh(job)

    # (ii) both Concept Files still download, with the same rows — the
    # whole (status, rows) tuple, compared against a control run of the
    # SAME exit with the lane healthy.
    assert _bulk_import_rows(client, job) == control_rows
    if exit_name in ("clean_captured", "raised_captured"):
        assert all(
            status == 200 for status, _rows in control_rows.values()
        ), "a captured exit has rows, so both lanes must serve 200"

    # (iii) the concept releases' states are identical to the control's.
    for lane in release.RELEASE_LANES:
        payload = release.release_payload(job, lane=lane)
        assert payload is not None, lane
        assert release.release_state(payload) == control_states[lane], lane

    nonterminal = exit_name.startswith("raised")

    # (iv) completed lanes name Master failures. Non-terminal lanes never
    # enter Master generation, so they must not manufacture a second issue.
    expected = type(make_error()).__name__
    for lane in release.RELEASE_LANES:
        issue = release.assessment_lane_issue(
            release.release_payload(job, lane=lane))
        if nonterminal:
            assert issue is None, lane
            continue
        assert issue is not None, lane
        assert issue["code"] == release.ASSESSMENT_LANE_UNAVAILABLE
        assert issue["details"]["lane"] == lane
        assert issue["details"]["exception"] == expected
        assert issue["details"]["staged_version"] >= 1

    # (v) only the Master entries go disabled, carrying that reason.
    for name, entries in _manifest_blocks(job).items():
        by_kind = {row["kind"]: row for row in entries}
        assert [row["kind"] for row in entries][:4] == _OUTPUT_ORDER, name
        for kind, lane in (
            ("pre_release_master", release.LANE_PRE),
            ("release_master", release.LANE_POST),
        ):
            assert by_kind[kind]["disabled"] is True, (name, kind)
            if nonterminal:
                disabled_reason = by_kind[kind]["disabled_reason"]
                assert "not" in disabled_reason and "built" in disabled_reason, (
                    name, kind, disabled_reason,
                )
            else:
                assert expected in by_kind[kind]["disabled_reason"], (
                    name, kind,
                )
        for kind in ("pre_release_bulk_import", "release_bulk_import"):
            assert not by_kind[kind].get("disabled"), (name, kind)

    # (vi) the concept-lane database upload is still open.
    for kind, lane in (
        ("database_upload", release.LANE_POST),
        ("pre_database_upload", release.LANE_PRE),
    ):
        entry = {
            row["kind"]: row
            for row in release_manifest.release_artifact_entries(job)
        }[kind]
        assert entry["disabled"] is False, kind
        assert entry["release_state"] == control_states[lane], kind


def _bulk_import_rows(client, job) -> dict[str, list]:
    """Both lanes' Concept Files, downloaded and read back as ROWS.

    Not a byte length: the workbook carries the job id and the staging
    timestamp, so two runs of the same content differ in bytes and agree
    in rows. What the assertion is about is the CONTENT still shipping.
    """

    import io

    from openpyxl import load_workbook

    rows: dict[str, list] = {}
    for lane, query in (
        (release.LANE_POST, ""), (release.LANE_PRE, "?lane=pre"),
    ):
        response = client.get(
            f"/build-concepts/uploads/{job.id}"
            f"/release-bulk-import.xlsx{query}"
        )
        if response.status_code != 200:
            # A checkpoint-only exit stages a release with no concept row
            # at all, and that route has always refused to export an empty
            # workbook. Recorded rather than asserted away: the property
            # under test is that the assessment lane's failure changes
            # NOTHING here, and it is proved by comparing this whole tuple
            # against a control run of the same exit.
            rows[lane] = (response.status_code, None)
            continue
        book = load_workbook(io.BytesIO(response.content), read_only=True)
        rows[lane] = (200, [
            (sheet.title, [
                tuple("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
                if any(cell not in (None, "") for cell in row)
            ])
            for sheet in book.worksheets
        ])
        book.close()
    return rows


# --------------------------------------------------------------------------- #
# 6. The repairs the three S6 audits found
# --------------------------------------------------------------------------- #

def test_a_lane_that_fails_after_minting_its_row_still_names_the_reason(
    db, monkeypatch,
):
    """The realistic partial failure, and it used to lose its reason.

    ``create_release`` COMMITS the row and ``publish_release`` runs after
    it, so "a row exists and the lane still failed" is the ordinary shape
    of a Master-lane fault — not a corner. The Master entry read the
    recorded ``assessment_lane_unavailable`` issue only in the "no row at
    all" branch, so this shape showed the generic "has not finished
    publishing" sentence and threw the named exception away.

    Nothing was ever LOST (the issue is durable on the concept payload),
    but the reviewer was told the less specific of two true things, which
    is exactly what T15-2 (v) exists to prevent.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)

    def mint_then_fail(_db, job_id, **kwargs):
        row = models.AssessmentRelease(
            release_uid=f"REL-{uuid.uuid4().hex[:8]}",
            version=1,
            owner_sub=OWNER,
            job_id=job_id,
            lane=release.LANE_POST,
            layout_id=release_core.layout_id(),
            state="materialized",
        )
        _db.add(row)
        _db.commit()
        raise svc.UploadRefused("this release is already being published")

    monkeypatch.setattr(run, "run_release_for_job", mint_then_fail)
    monkeypatch.setattr(
        run, "run_pre_release_for_job",
        _stub_release(db, chapter, release.LANE_PRE))

    def original(_db, _job_id, _chapter_id, **_kwargs):
        return {"status": "ok"}

    release_contract._run_generation_release(
        original, db, job.id, chapter.id, owner_sub=OWNER)
    db.refresh(job)

    # The row IS there and unpublished — this is the branch under test.
    row = release_core.latest_release_for_lane(
        db, job.id, release.LANE_POST)
    assert row is not None
    assert not (row.publication or {}).get("manifest")

    issue = release.assessment_lane_issue(release.release_payload(job))
    assert issue is not None
    assert issue["details"]["exception"] == "UploadRefused"

    for name, entries in _manifest_blocks(job).items():
        entry = {row["kind"]: row for row in entries}["release_master"]
        assert entry["disabled"] is True, name
        assert "UploadRefused" in entry["disabled_reason"], name


def test_the_manifest_never_advertises_a_superseded_release(db):
    """``latest_release_for_lane`` now says what its docstring says.

    Its ``(live or rows or [None])[0]`` fallback returned a SUPERSEDED
    row when only superseded rows existed, and ``master_entry`` then
    published an ENABLED download pointing at it — in flat contradiction
    of the docstring's "a superseded row is not what the reviewer's
    manifest should point at". The entry is now present and disabled with
    a reason that states which of the four things happened.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _pre, post = _run_both_lanes(db, job, chapter)
    db.refresh(job)
    assert not release_files.master_entry(
        job, lane=release.LANE_POST).get("disabled")

    post.state = "superseded"
    db.commit()
    db.refresh(job)

    assert release_core.latest_release_for_lane(
        db, job.id, release.LANE_POST) is None
    for name, entries in _manifest_blocks(job).items():
        entry = {row["kind"]: row for row in entries}["release_master"]
        assert entry["disabled"] is True, name
        assert entry["download_url"] == "", name
        assert "superseded" in entry["disabled_reason"], name


def test_the_lineage_question_still_includes_a_superseded_row(db):
    """And the chain therefore never orphans its published receipts.

    The freeze seam asks ``release_chain_head``, not the manifest's
    question: a chain whose every row has been superseded is still that
    lane's chain, and appending version N+1 to it keeps every published
    receipt in one lineage where minting a fresh ``release_uid`` at
    version 1 would strand them.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _pre, post = _run_both_lanes(db, job, chapter)
    post.state = "superseded"
    db.commit()

    head = release_core.release_chain_head(db, job.id, release.LANE_POST)
    assert head is not None and head.id == post.id
    assert release_core.latest_release_for_lane(
        db, job.id, release.LANE_POST) is None


def test_a_lane_with_no_staged_concept_release_is_skipped_not_discarded(
    db, monkeypatch,
):
    """A doomed attempt whose reason cannot be recorded is not made.

    ``record_assessment_lane_unavailable`` writes onto the lane's staged
    concept payload, so a lane with no payload has nowhere to record a
    failure: the exception was caught and then DROPPED — neither built
    nor recorded. A chapter with no Phase-03 pre map is the ordinary
    shape of that, so every Post-only run paid for a doomed Pre attempt
    and discarded its reason.

    The lane's Master entry is still PRESENT and disabled, so the skip is
    not silent either.
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    inventory = dict(job.question_inventory or {})
    inventory.pop(release.release_key_for_lane(release.LANE_PRE), None)
    job.question_inventory = inventory
    flag_modified(job, "question_inventory")
    db.commit()
    db.refresh(job)
    assert release.release_payload(job, lane=release.LANE_PRE) is None

    called = []

    def spy(name, result):
        def call(_db, job_id, **kwargs):
            called.append(name)
            if isinstance(result, BaseException):
                raise result
            return result(_db, job_id, **kwargs)
        return call

    monkeypatch.setattr(
        run, "run_pre_release_for_job",
        spy("pre", run.ReleaseRunError("no staged Pre concept release")))
    monkeypatch.setattr(
        run, "run_release_for_job",
        spy("post", _stub_release(db, chapter, release.LANE_POST)))

    built = release_contract._build_master_siblings(
        db, job.id, chapter.id, owner_sub=OWNER)
    db.refresh(job)

    assert called == ["post"], "the empty lane is not attempted at all"
    assert built[release.LANE_PRE] is None
    assert built[release.LANE_POST] is not None
    # The Post lane is untouched by the Pre lane's absence.
    assert release.assessment_lane_issue(
        release.release_payload(job)) is None
    entry = release_files.master_entry(job, lane=release.LANE_PRE)
    assert entry["kind"] == "pre_release_master"
    assert entry["disabled"] is True
    assert entry["disabled_reason"]


def test_no_concept_payload_can_take_the_assessment_branch(db):
    """The disjointness ``structural_defects`` silently depends on.

    ``_is_assessment_payload`` is a one-key shape sniff, and it returns
    EARLY out of the concept branch — including out of the concept lane's
    structural-defects refusal in ``upload_release_to_database`` (and,
    since S10, the seal, row-defect and duplicate-identity gates beside
    it). [measured] neither
    staging site emits ``candidates`` or ``groups`` today, so the branch
    is unreachable from a concept payload. Nothing asserted that, and the
    bypass condition is "a future slice adds a key named ``groups``".
    """

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    for lane in release.RELEASE_LANES:
        payload = release.release_payload(job, lane=lane)
        assert payload is not None, lane
        assert not (
            set(release._ASSESSMENT_PAYLOAD_KEYS) & set(payload)
        ), lane
        assert not release._is_assessment_payload(payload), lane


def test_a_migrated_database_gets_the_lane_index_a_fresh_one_gets(tmp_path):
    """The deployed database is the MIGRATED one.

    ``models.py`` declares ``lane`` ``index=True``, which only reaches
    databases ``create_all`` mints; ``_ensure_columns`` emits ``ADD
    COLUMN``, which carries no index. Without the explicit ``CREATE INDEX
    IF NOT EXISTS`` the deployed schema differs permanently from the
    declared model — and ``db.py`` already carries exactly that line for
    the two ``owner_sub`` indexes, so this is the house pattern, not a
    new one.
    """

    import sqlalchemy

    def indexes(url: str) -> set[str]:
        engine = sqlalchemy.create_engine(url)
        with engine.connect() as conn:
            rows = conn.exec_driver_sql(
                "SELECT name FROM sqlite_master WHERE type = 'index' "
                "AND tbl_name = 'assessment_releases'"
            ).fetchall()
        engine.dispose()
        return {row[0] for row in rows}

    fresh = tmp_path / "fresh.db"
    migrated = tmp_path / "migrated.db"
    for path in (fresh, migrated):
        engine = sqlalchemy.create_engine(f"sqlite:///{path}")
        models.Base.metadata.create_all(engine)
        engine.dispose()

    # Emulate a pre-S6 database: the two columns and the index absent.
    engine = sqlalchemy.create_engine(f"sqlite:///{migrated}")
    with engine.connect() as conn:
        conn.exec_driver_sql(
            "DROP INDEX IF EXISTS ix_assessment_releases_lane")
        conn.exec_driver_sql(
            "ALTER TABLE assessment_releases DROP COLUMN lane")
        conn.exec_driver_sql(
            "ALTER TABLE assessment_releases DROP COLUMN layout_id")
        conn.commit()
    engine.dispose()

    from app import db as db_module

    original_url, original_engine = db_module.DB_URL, db_module.engine
    try:
        db_module.DB_URL = f"sqlite:///{migrated}"
        db_module.engine = sqlalchemy.create_engine(db_module.DB_URL)
        db_module._ensure_columns()
        db_module.engine.dispose()
    finally:
        db_module.DB_URL, db_module.engine = original_url, original_engine

    assert "ix_assessment_releases_lane" in indexes(f"sqlite:///{fresh}")
    assert indexes(f"sqlite:///{migrated}") == indexes(f"sqlite:///{fresh}")


# --------------------------------------------------------------------------- #
# 7. S8's headline regression — the R4 hole, closed end to end
# --------------------------------------------------------------------------- #

def test_a_candidate_with_no_resolved_home_is_a_named_defect_not_a_silent_warning(
    db, client,
):
    """The whole of spec-step8's first atomic, proved as an END STATE.

    A candidate whose ``concept_key`` does not resolve:

    * yields a NAMED ``unresolved_question_home`` entry in
      ``diagnostics["payload_errors"]`` — decided at STAGING, by a pure
      function of the frozen snapshot, not discovered at projection time;
    * leaves ``_readiness`` BLOCKED and the database write refused;
    * leaves ALL FOUR output downloads returning 200;
    * and its ``question_label`` appears in NO DATA ROW OF ANY OF THE FOUR
      OUTPUTS. That last fact is what makes it a DEFECT rather than a flag:
      §4's "flags never block" governs semantic doubt on a row that SHIPS,
      and this row does not ship at all.
    """

    import io

    from openpyxl import load_workbook

    chapter = _chapter_with_concepts(db)
    job = _both_lanes_job(db, chapter)
    _pre, post = _run_both_lanes(db, job, chapter)

    orphaned = copy.deepcopy(dict(post.payload))
    label = str(orphaned["candidates"][0]["question_label"])
    orphaned["candidates"][0]["group_key"] = "(no-such-concept) BG01"
    orphaned["candidates"][0]["concept_key"] = "no-such-concept"

    broken = svc.create_release(
        db,
        chapter_id=chapter.id,
        payload=orphaned,
        job_id=job.id,
        owner_sub=OWNER,
        supersedes=post,
        lane=post.lane,
        layout_id=post.layout_id,
        provider_identity=dict(post.provider_identity or {}),
    )

    # (1) the named verdict, at staging.
    named = [
        error for error in broken.diagnostics["payload_errors"]
        if error.startswith(rel.UNRESOLVED_QUESTION_HOME)
    ]
    assert len(named) == 1, broken.diagnostics["payload_errors"]
    assert label in named[0]

    published = svc.publish_release(db, broken)

    # (2) readiness BLOCKED and the database write refused.
    assert published.diagnostics["readiness"] == svc.BLOCKED
    with pytest.raises(svc.UploadRefused, match="blocked"):
        svc.upload_master_to_database(db, published, owner_sub=OWNER)

    # (3) all four downloads 200.
    db.refresh(job)
    downloads: list[tuple[str, bytes]] = []
    for lane_query in ("", "?lane=pre"):
        response = client.get(
            f"/build-concepts/uploads/{job.id}"
            f"/release-bulk-import.xlsx{lane_query}")
        assert response.status_code == 200, lane_query
        downloads.append((lane_query or "post", response.content))
    for row_release in (_pre, published):
        response = client.get(
            f"/build-assessments/releases/{row_release.id}/master.xlsx")
        assert response.status_code == 200, row_release.id
        downloads.append((f"master-{row_release.id}", response.content))
    assert len(downloads) == 4

    # (4) and the label reaches no data row of any of them.
    for name, content in downloads:
        book = load_workbook(io.BytesIO(content), read_only=True)
        for sheet in book.worksheets:
            for row in sheet.iter_rows(values_only=True):
                assert label not in [
                    None if cell is None else str(cell) for cell in row
                ], (name, sheet.title)
        book.close()


def test_terminal_diagnosis_names_the_failing_predicate():
    """The checkpoint diagnosis helper must say WHY, per predicate.

    Restructure A (2026-08-29) removed this from Master eligibility —
    the gate now reads the verdict recorded at staging instead of
    re-deriving from the live checkpoint — but the helper remains the
    named-predicate diagnostic for checkpoint troubleshooting, and its
    per-predicate wording stays pinned."""

    from app.services import generation

    assert "no checkpoint entries" in (
        generation.concept_checkpoint_terminal_diagnosis({})
    )
    early_only = {"stage": "question_inventory"}
    assert "no terminal-stage entry" in (
        generation.concept_checkpoint_terminal_diagnosis(early_only)
    )
    broken_terminal = {
        "schema_version": generation._CONCEPT_CHECKPOINT_SCHEMA,
        "stage": "final_content_ready",
        "stage_schema_version": 1,
    }
    diagnosis = generation.concept_checkpoint_terminal_diagnosis(
        broken_terminal
    )
    assert "rejected" in diagnosis
    assert "stage_schema_version" in diagnosis or "bundle" in diagnosis
