"""Many-to-many tagging: repeated-row export, placement-aware dedupe, preview."""
import io

import openpyxl

from app import models
from app.bulk_import import SHEET_OBJECTIVE, strip_title_tag, writer
from app.services import tagging


def _two_concepts(client, chapter):
    detail = client.get(f"/directory/chapters/{chapter['id']}").json()
    concepts = [c for t in detail["topics"] for c in t["concepts"]]
    return concepts


def _generate_one_question(client, concept_id):
    s = client.post("/build-assessments/sessions", json={
        "scope_type": "concept", "scope_ids": [concept_id],
    }).json()
    client.post(f"/build-assessments/sessions/{s['id']}/batches", json={
        "cognitive_skills": ["Understanding"], "difficulty_levels": ["Moderate"],
        "categories": ["Multiple Choice Question"], "question_type": "objective",
        "num_questions": 1,
    })
    client.post(f"/build-assessments/sessions/{s['id']}/generate")
    return client.get(f"/build-assessments/sessions/{s['id']}").json()["generated_question_ids"][0]


def _labels_to_concepts(data: bytes) -> dict[str, set[str]]:
    """Map question_label -> set of concept_titles it appears under (objective sheet)."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb[SHEET_OBJECTIVE]
    qs = writer._q_start("objective")
    out: dict[str, set[str]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        label = writer._cell_str(row, qs)
        if not label:
            continue
        out.setdefault(label, set()).add(
            strip_title_tag(writer._cell_str(row, writer._IDX_CONCEPT_TITLE)))
    return out


def test_tag_question_emits_repeated_row_same_label(client, db, first_chapter):
    concepts = _two_concepts(client, first_chapter)
    assert len(concepts) >= 2, "seed chapter needs >= 2 concepts"
    c1, c2 = concepts[0], concepts[1]
    qid = _generate_one_question(client, c1["id"])

    r = client.post(f"/tagging/questions/{qid}/tag-to-concept",
                    json={"concept_id": c2["id"]}).json()
    assert r["status"] == "tagged"

    db.expire_all()
    data = writer.write_workbook(db, question_ids=[qid])
    label_map = _labels_to_concepts(data)
    label = db.get(models.Question, qid).question_label
    # Same question_label appears twice, under two distinct concept_titles.
    assert label in label_map
    assert len(label_map[label]) == 2
    assert {c1["concept_title"], c2["concept_title"]} == label_map[label]


def test_placement_aware_dedupe_skips_exact_repeat(client, db, first_chapter, tmp_path):
    concepts = _two_concepts(client, first_chapter)
    c1, c2 = concepts[0], concepts[1]
    qid = _generate_one_question(client, c1["id"])
    client.post(f"/tagging/questions/{qid}/tag-to-concept", json={"concept_id": c2["id"]})
    db.expire_all()

    path = tmp_path / "wb.xlsx"
    first = writer.append_questions(db, path, [qid])
    # Two placements written (home + tag); exactly one is a tag.
    assert first["objective"] == 2
    assert first["tagged"] == 1
    assert first["skipped"] == 0

    second = writer.append_questions(db, path, [qid])
    # Re-running adds nothing — both placements already present.
    assert second["objective"] == 0
    assert second["skipped"] == 2


def test_concept_tag_emits_repeated_concept_row(client, db, first_chapter, tmp_path):
    detail = client.get(f"/directory/chapters/{first_chapter['id']}").json()
    topics = detail["topics"]
    assert len(topics) >= 2, "seed chapter needs >= 2 topics"
    concept = topics[0]["concepts"][0]
    other_topic = topics[1]

    r = client.post(f"/tagging/concepts/{concept['id']}/tag-to-topic",
                    json={"topic_id": other_topic["id"]}).json()
    assert r["status"] == "tagged"

    db.expire_all()
    path = tmp_path / "concepts.xlsx"
    written = writer.append_concepts(db, path, [concept["id"]])
    assert written["written"] == 2  # home topic + tagged topic

    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_OBJECTIVE]
    topic_titles = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if strip_title_tag(writer._cell_str(row, writer._IDX_CONCEPT_TITLE)) == concept["concept_title"]:
            topic_titles.add(writer._cell_str(row, writer._IDX_TOPIC_TITLE))
    assert len(topic_titles) == 2


def test_preview_classifies_add_tag_skip(client, db, first_chapter, tmp_path):
    concepts = _two_concepts(client, first_chapter)
    c1, c2 = concepts[0], concepts[1]
    qid = _generate_one_question(client, c1["id"])
    client.post(f"/tagging/questions/{qid}/tag-to-concept", json={"concept_id": c2["id"]})
    db.expire_all()

    path = tmp_path / "preview.xlsx"

    # Nothing written yet: home = ADD, tag = TAG.
    p1 = tagging.preview(db, question_ids=[qid], path=path)
    assert p1["summary"]["ADD"] == 1
    assert p1["summary"]["TAG"] == 1
    assert p1["summary"]["SKIP"] == 0

    # Write both placements, then preview again: everything already present = SKIP.
    writer.append_questions(db, path, [qid])
    p2 = tagging.preview(db, question_ids=[qid], path=path)
    assert p2["summary"]["SKIP"] == 2
    assert p2["summary"]["ADD"] == 0
    assert p2["summary"]["TAG"] == 0


def test_tag_to_home_placement_is_noop(client, db, first_chapter):
    concepts = _two_concepts(client, first_chapter)
    qid = _generate_one_question(client, concepts[0]["id"])
    q = db.get(models.Question, qid)
    r = client.post(f"/tagging/questions/{qid}/tag-to-group",
                    json={"group_id": q.group_id}).json()
    assert r["status"] == "noop"
