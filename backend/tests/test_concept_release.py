from __future__ import annotations

import copy
from contextlib import nullcontext
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app import config, models
from app.main import app
from app.services import build_concepts
from app.services import concept_release as release


def _record(title: str = "Common Difference") -> dict:
    return {
        "topic": "Arithmetic Progressions",
        "parent_concept": "Sequences",
        "concept_title": title,
        "concept_details": (
            "Description: The common difference is the fixed amount added "
            "between consecutive terms.\n"
            "Achieving Mastery: Identify the common difference. // "
            "Types: Type 01: Applying a fixed-difference relationship "
            "Case 01: Determine the common difference from consecutive terms "
            "Example 01: Find the common difference in 3, 7, 11, 15. // "
            "Misconception/ Error Analysis: Misconceptions: Learners may "
            "believe the terms are multiplied.; Error Analysis: Learners may "
            "subtract in the wrong order."
        ),
        "keywords": "arithmetic progression, common difference",
        "_source_block_ids": ["BLK-0001"],
    }


def _split_type() -> tuple[dict, dict]:
    inventory = {
        "items": [
            {"qid": "QINV-0001", "raw_task": "Find the nth term."},
            {"qid": "QINV-0002", "raw_task": "Find the finite sum."},
        ]
    }
    mined = {
        "types": [{
            "type_id": "TYPE-0001",
            "type_title": "Using an arithmetic progression relationship",
            "type_description": (
                "Select the relationship required by the requested quantity."
            ),
            "case_prompts": [
                {
                    "case_id": "CASE-0001",
                    "case_title": "Nth-term request",
                    "case_signature": "The requested quantity is a_n.",
                    "examples": [{
                        "source_question_id": "QINV-0001",
                        "example_prompt": "Find the nth term.",
                        "_type_case_placement_contract": {
                            "owner_topic_id": "TOPIC-NTH",
                            "owner_topic_title": "The nth Term",
                        },
                    }],
                },
                {
                    "case_id": "CASE-0002",
                    "case_title": "Finite-sum request",
                    "case_signature": "The requested quantity is S_n.",
                    "examples": [{
                        "source_question_id": "QINV-0002",
                        "example_prompt": "Find the finite sum.",
                        "_type_case_placement_contract": {
                            "owner_topic_id": "TOPIC-SUM",
                            "owner_topic_title": "Sum of First n Terms",
                        },
                    }],
                },
            ],
        }]
    }
    return mined, inventory


def test_release_router_precedes_legacy_generation_router():
    path = "/build-concepts/post-learning/uploads/{job_id}/generate"
    matches = [route for route in app.routes if getattr(route, "path", "") == path]
    assert len(matches) >= 2
    assert matches[0].endpoint.__module__ == "app.api.concept_releases"


def test_one_reusable_type_can_route_cases_to_different_topics_exactly_once():
    mined, inventory = _split_type()

    routes, issues = release._type_case_routes(mined, inventory)

    assert [(row["type_id"], row["case_id"], row["qid"]) for row in routes] == [
        ("TYPE-0001", "CASE-0001", "QINV-0001"),
        ("TYPE-0001", "CASE-0002", "QINV-0002"),
    ]
    assert [row["owner_topic_id"] for row in routes] == [
        "TOPIC-NTH", "TOPIC-SUM",
    ]
    assert not [issue for issue in issues if issue["severity"] == "error"]


def test_duplicate_qid_across_cases_is_a_release_error():
    mined, inventory = _split_type()
    mined["types"][0]["case_prompts"][1]["examples"][0][
        "source_question_id"
    ] = "QINV-0001"

    _routes, issues = release._type_case_routes(mined, inventory)

    assert any(issue["code"] == "duplicate_qid_assignment" for issue in issues)
    assert any(issue["code"] == "qid_not_assigned_to_type_case" for issue in issues)


def test_review_workbook_highlights_error_rows_and_has_context_sheets(tmp_path):
    workbook_path = tmp_path / "review.xlsx"
    issue = release._issue(
        "test_error",
        "This row needs review.",
        row_index=0,
        concept="Common Difference",
    )

    release._concept_workbook(
        workbook_path,
        records=[_record()],
        issues=[issue],
        routes=[],
        blocks=[{
            "block_id": "BLK-0001",
            "topic_id": "TOPIC-1",
            "kind": "paragraph",
            "text": "Canonical source evidence.",
        }],
        events=[{"type": "log", "level": "warning", "message": "Test log"}],
        manifest_summary={"database_upload_allowed": False},
    )

    workbook = load_workbook(workbook_path)
    assert {
        "Concepts Review", "Issues", "QID Type Case Routes",
        "Evidence BLKs", "Run Log", "Release Manifest",
    } <= set(workbook.sheetnames)
    sheet = workbook["Concepts Review"]
    assert sheet["B2"].value == "ERROR - REVIEW REQUIRED"
    assert sheet["A2"].fill.fgColor.rgb.endswith("F4CCCC")
    assert sheet["F2"].value == "This row needs review."


def test_intercepted_deposit_captures_candidate_without_database_write():
    context = release._ReleaseContext(
        db=SimpleNamespace(),
        job_id=41,
        target_chapter_id=7,
        learning_kind="post",
        owner_sub="local:default",
    )
    job = SimpleNamespace(
        generation_checkpoint={"stage": "final"},
        question_inventory={"items": []},
    )
    token = release._RELEASE_CONTEXT.set(context)
    try:
        created, merged, written = release._intercept_deposit_and_publish(
            SimpleNamespace(),
            chapter_id=7,
            records=[_record()],
            pre_post="Post",
            source_book="NCERT",
            inventory={"items": []},
            mined_types={"types": []},
            source_text="source",
            final_grounding_certificate={"certificate": "ok"},
            grounding_audit_job=job,
        )
    finally:
        release._RELEASE_CONTEXT.reset(token)

    assert created == []
    assert merged == []
    assert written["release_intercepted"] is True
    assert context.captured["records"] == [_record()]
    assert context.captured["final_grounding_certificate"] == {
        "certificate": "ok"
    }


def test_complete_context_release_is_created_even_with_errors(
    tmp_path,
    monkeypatch,
):
    monkeypatch.setattr(config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(release, "_RELEASE_ROOT", tmp_path / "concept-releases")
    monkeypatch.setattr(release, "_copy_original_source", lambda *_args: "")
    monkeypatch.setattr(release, "_copy_rules", lambda *_args: "")
    job = SimpleNamespace(
        id=44,
        owner_sub="local:default",
        filename="source.pdf",
        source_book="NCERT",
        status="converted",
        learning_kind="post",
        mmd_text="# Source",
        generation_log=[{
            "type": "log", "level": "error", "message": "Semantic conflict"
        }],
        generation_checkpoint={"stage": "pre_type_assignment"},
        question_inventory={},
    )

    metadata = release.create_release(
        None,
        job,
        target_chapter_id=7,
        learning_kind="post",
        captured={
            "records": [_record()],
            "inventory": {"items": []},
            "mined_types": {"types": []},
            "source_text": "# Source",
            "source_context": {
                "graph": {
                    "blocks": [{
                        "block_id": "BLK-0001",
                        "topic_id": "TOPIC-1",
                        "kind": "paragraph",
                    }]
                },
                "canonical": {
                    "blocks": [{
                        "block_id": "BLK-0001",
                        "display_text": "Canonical source evidence.",
                    }]
                },
            },
        },
        terminal_failure=RuntimeError("terminal semantic conflict"),
    )

    release_dir = tmp_path / metadata["release_dir"]
    assert (release_dir / "concepts_review.xlsx").is_file()
    assert (release_dir / "context_bundle.zip").is_file()
    assert (release_dir / "generation_log.json").is_file()
    assert (release_dir / "source_blocks.json").is_file()
    assert (release_dir / "source_evidence.json").is_file()
    assert metadata["records_released"] == 1
    assert metadata["error_count"] >= 1
    assert metadata["database_upload_allowed"] is False


def test_generation_exception_is_converted_to_release_not_manual_pause(
    db,
    monkeypatch,
):
    job = models.UploadJob(
        owner_sub="local:default",
        module="build_concepts",
        upload_type="document",
        learning_kind="post",
        filename="chapter.mmd",
        mmd_text="# Chapter",
        status="converted",
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    captured: dict = {}

    def fails(*_args, **_kwargs):
        raise RuntimeError("phase32 semantic disagreement")

    def fake_release(_db, _job, **kwargs):
        captured.update(kwargs)
        return {
            "version": release.RELEASE_VERSION,
            "release_id": "release-test",
            "created_at": "2026-08-04T00:00:00+00:00",
            "job_id": job.id,
            "records_released": 3,
            "inventory_items": 2,
            "database_upload_allowed": False,
            "issue_count": 1,
            "error_count": 1,
            "warning_count": 0,
            "error_row_count": 1,
            "warning_row_count": 0,
            "ready_row_count": 2,
            "terminal_failure": {
                "exception_type": "RuntimeError",
                "reason": "phase32 semantic disagreement",
            },
            "selected_checkpoint_stage": "pre_type_assignment",
        }

    monkeypatch.setattr(build_concepts, "generate_post_learning", fails)
    monkeypatch.setattr(release, "create_release", fake_release)
    monkeypatch.setattr(
        release.uploads,
        "persist_current_generation_log",
        lambda *_args, **_kwargs: [],
    )
    monkeypatch.setattr(
        release.drive_checkpoints,
        "schedule_checkpoint_backup",
        lambda *_args, **_kwargs: None,
    )

    result = release.generate_post_learning_release(
        db,
        job.id,
        7,
        owner_sub="local:default",
    )

    db.refresh(job)
    assert result["status"] == "released"
    assert result["output_released"] is True
    assert result["database_written"] is False
    assert job.status == "released"
    assert job.pending_decision is None
    assert isinstance(captured["terminal_failure"], RuntimeError)


def test_dirty_release_cannot_be_uploaded_to_database(db, monkeypatch):
    job = models.UploadJob(
        owner_sub="local:default",
        module="build_concepts",
        upload_type="document",
        learning_kind="post",
        filename="chapter.mmd",
        mmd_text="# Chapter",
        status="released",
        question_inventory={
            release.RELEASE_METADATA_KEY: {
                "release_id": "dirty-release",
                "database_upload_allowed": False,
                "records_released": 4,
            }
        },
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    monkeypatch.setattr(
        build_concepts,
        "_deposit_concepts",
        lambda *_args, **_kwargs: pytest.fail(
            "dirty release reached database deposit"
        ),
    )

    with pytest.raises(
        build_concepts.DepositValidationError,
        match="not database-ready",
    ):
        release.publish_release(
            db,
            job.id,
            owner_sub="local:default",
        )
