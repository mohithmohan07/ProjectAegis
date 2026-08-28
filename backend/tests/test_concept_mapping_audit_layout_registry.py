"""Reader registry coverage for the normalized 2026-08-27 audit Masters."""
from __future__ import annotations

import io

import openpyxl
import pytest

from app import models
from app.bulk_import import assessment_workbook, layouts, reader
from app.services import assessment_profile


@pytest.fixture()
def import_db(db):
    """DB session whose importer-created rows are removed after the test.

    The suite's database is session-populated and function sessions do not
    roll back ``reader.import_workbook`` because the reader commits.  Record
    the shared baseline IDs and delete only rows created by this test, in
    foreign-key order; baseline chapters and their descendants are untouched.
    """

    cleanup_order = (
        models.QuestionTag,
        models.ConceptTag,
        models.Question,
        models.Group,
        models.Concept,
        models.Topic,
        models.Chapter,
    )
    baseline_ids = {
        model: {row_id for (row_id,) in db.query(model.id).all()}
        for model in cleanup_order
    }
    try:
        yield db
    finally:
        db.rollback()
        for model in cleanup_order:
            query = db.query(model)
            if baseline_ids[model]:
                query = query.filter(~model.id.in_(baseline_ids[model]))
            query.delete(synchronize_session=False)
        db.commit()


def _profile(subject: str) -> dict:
    return assessment_profile.resolve_for_metadata(None, {
        "board": "MSBSHSE",
        "grade": "06",
        "subject": subject,
    })


def _snapshot(subject: str, phase: str, marker: str) -> dict:
    chapter_tag = f"06_{subject}_MSBSHSE_Balbharati"
    concept_key = f"concept-{marker}"
    concept_id = f"06MS{marker}_T01_C01"
    return {
        "chapter": {
            "chapter_title": f"Audit {marker} ({chapter_tag})",
            "chapter_display_name": f"Audit {marker}",
            "chapter_duration": "362",
            "pre_topics": "",
            "post_topics": "",
            "chapter_description": "Normalized audit layout test.",
        },
        "topics": [{
            "topic_title": f"Topic 01: Audit {marker}",
            "topic_display_name": f"Audit {marker}",
            "pre_post_learning": phase,
            "topic_concept_labels": f"Audit {marker} ({concept_id})",
            "related_topics": "",
            "topic_description": "Normalized audit topic.",
            "concepts": [{
                "concept_key": concept_key,
                "concept_machine_id": concept_id,
                "concept_title": f"Audit {marker}",
                "concept_display_name": f"Audit {marker}",
                "concept_details": (
                    "Description: Normalized audit concept.\n"
                    "Achieving Mastery: Explain it."
                ),
                "keywords": "audit",
                "digicards": "",
                "related_concepts": "",
                "concept_source": "Balbharati",
            }],
        }],
        "groups": [],
        "candidates": [],
    }


def _headers(data: bytes) -> dict[str, tuple]:
    workbook = openpyxl.load_workbook(
        io.BytesIO(data), data_only=True, read_only=True,
    )
    try:
        return {
            name: next(
                workbook[name].iter_rows(
                    min_row=2, max_row=2, values_only=True,
                ),
                (),
            )
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("subject", "phase", "layout_id", "counts"),
    [
        (
            "Mathematics",
            "Post",
            layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
            (72, 380, 149),
        ),
        (
            "English",
            "Post",
            layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
            (72, 440, 149),
        ),
    ],
)
def test_rendered_normalized_master_identifies_and_reads(
    import_db, tmp_path, subject, phase, layout_id, counts,
):
    db = import_db
    snapshot = _snapshot(subject, phase, layout_id[-8:])
    data, _issues = assessment_workbook.render_master_file(
        snapshot, _profile(subject),
    )

    identity = layouts.identify_workbook(_headers(data))
    assert identity.layout_id == layout_id
    assert tuple(found.sheet_name for found in identity.sheets) == (
        "Objective", "Descriptive", "Subjective",
    )
    assert tuple(
        len(identity.layout.sheet(kind).fields)
        for kind in ("objective", "descriptive", "subjective")
    ) == counts

    path = tmp_path / f"{layout_id}.xlsx"
    path.write_bytes(data)
    imported = reader.import_workbook(db, path)
    assert imported["layout_id"] == layout_id
    assert imported["chapters"] == 1
    assert imported["topics"] == 1
    assert imported["concepts"] == 1


def test_registry_schema_is_the_renderer_schema_field_for_field() -> None:
    fixtures = (
        (
            _profile("Mathematics"),
            _snapshot("Mathematics", "Post", "MathRegistry"),
            layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
        ),
        (
            _profile("English"),
            _snapshot("English", "Post", "EnglishRegistry"),
            layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
        ),
    )
    for profile, snapshot, layout_id in fixtures:
        schema = assessment_workbook.output_schema(
            "master", profile, snapshot,
        )
        registered = layouts.layout(layout_id)
        for kind, sheet_name in (
            ("objective", "Objective"),
            ("descriptive", "Descriptive"),
            ("subjective", "Subjective"),
        ):
            sheet = registered.sheet(kind)
            assert list(sheet.fields) == schema["fields"][sheet_name]
            assert [band.as_dict() for band in sheet.bands] == (
                schema["bands"][sheet_name]
            )
            assert len(sheet.fields) == len(set(sheet.fields))


def test_english_post_reader_preserves_all_thirty_descriptive_answers(
    import_db, tmp_path,
) -> None:
    db = import_db
    marker = "EnglishThirty"
    snapshot = _snapshot("English", "Post", marker)
    concept = snapshot["topics"][0]["concepts"][0]
    concept_key = concept["concept_key"]
    group_key = f"{concept['concept_machine_id']} BG01"
    snapshot["groups"] = [{
        "group_key": group_key,
        "concept_key": concept_key,
        "group_name": group_key,
        "group_display_name": group_key,
        "semantic_description": "Thirty-part rubric.",
        "group_status": "Active",
        "group_type": "Basic",
    }]
    question_label = f"{concept['concept_machine_id']} Q01"
    snapshot["candidates"] = [{
        "candidate_id": "english-thirty-answer-candidate",
        "question_label": question_label,
        "sheet_kind": "descriptive",
        "question_category": "Composition Writing",
        "cognitive_skill": "Apply",
        "question_source": "UpSchool DB",
        "question_disclaimer": "",
        "question_duration": 30,
        "question_appears_in": "Pre/Post-Worksheet/Test",
        "answer_restriction": "Specific",
        "difficulty": "High",
        "question": "Write a response covering all thirty criteria.",
        "question_text": "Write a response covering all thirty criteria.",
        "marks": 30,
        "math_keyboard": "No",
        "display_answer": "A complete response covers every criterion.",
        "answer_explanation": "One mark per criterion.",
        "answers": [
            {
                "answer_type": "Phrases",
                "answer_content": f"[content]: criterion {number}",
                "answer_weightage": 1,
            }
            for number in range(1, 31)
        ],
        "sub_questions": [],
        "concept_key": concept_key,
        "group_key": group_key,
        "flags": [],
    }]

    data, issues = assessment_workbook.render_master_file(
        snapshot, _profile("English"),
    )
    assert issues["truncated_rows"] == []
    assert layouts.identify_workbook(_headers(data)).layout_id == (
        layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID
    )

    path = tmp_path / "english-post-thirty-answers.xlsx"
    path.write_bytes(data)
    imported = reader.import_workbook(db, path)
    assert imported["questions"] == 1
    question = db.query(models.Question).filter_by(
        question_label=question_label,
    ).one()
    assert len(question.answers) == 30
    assert question.answers[-1]["answer_content"] == "[content]: criterion 30"


def test_normalized_master_refuses_duplicate_headers_and_wrong_sheet_order() -> None:
    layout = layouts.layout(layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID)
    canonical = {
        sheet.sheet_name: sheet.fields
        for sheet in layout.sheets.values()
    }

    duplicate = dict(canonical)
    objective = list(duplicate["Objective"])
    objective.insert(objective.index("chapter_display_name"), "chapter_title")
    duplicate["Objective"] = tuple(objective)
    with pytest.raises(layouts.WorkbookLayoutError):
        layouts.identify_workbook(duplicate)

    reordered = {
        "Descriptive": canonical["Descriptive"],
        "Objective": canonical["Objective"],
        "Subjective": canonical["Subjective"],
    }
    with pytest.raises(layouts.WorkbookLayoutError) as exc:
        layouts.identify_workbook(reordered)
    assert "canonical order" in str(exc.value)

    incomplete = {
        "Objective": canonical["Objective"],
        "Descriptive": canonical["Descriptive"],
    }
    with pytest.raises(layouts.WorkbookLayoutError) as exc:
        layouts.identify_workbook(incomplete)
    assert "must be complete" in str(exc.value)


@pytest.mark.parametrize("layout_id", [
    layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
    layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
])
@pytest.mark.parametrize("suffix", [" ", "\n"])
def test_normalized_master_refuses_whitespace_mutated_field_names(
    layout_id: str,
    suffix: str,
) -> None:
    layout = layouts.layout(layout_id)
    headers = {
        sheet.sheet_name: sheet.fields
        for sheet in layout.sheets.values()
    }
    descriptive = list(headers["Descriptive"])
    index = descriptive.index("question_label")
    descriptive[index] = descriptive[index] + suffix
    headers["Descriptive"] = tuple(descriptive)

    assert layouts.identify_sheet(
        "Descriptive", headers["Descriptive"],
    ) is None
    with pytest.raises(layouts.WorkbookLayoutError):
        layouts.identify_workbook(headers)


@pytest.mark.parametrize("layout_id", [
    layouts.MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
    layouts.MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
])
def test_normalized_master_ignores_only_absent_trailing_header_cells(
    layout_id: str,
) -> None:
    layout = layouts.layout(layout_id)
    headers = {
        sheet.sheet_name: (*sheet.fields, None, None)
        for sheet in layout.sheets.values()
    }

    assert layouts.identify_workbook(headers).layout_id == layout_id
