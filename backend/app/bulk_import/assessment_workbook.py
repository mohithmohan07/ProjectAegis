"""Assessment workbook: dual projections of one immutable release snapshot.

The positional authority for everything here is the Grade-6 MES Bulk Import
FINAL template set (spec §0 item 2, §23), captured verbatim in
``assessment_workbook_template.json``: three sheets (Objective, Descriptive,
Subjective — no Doc Link, no trailing spaces in sheet names), the exact
two-row headers, the row-1 merged bands with their sheet-specific labels,
``answer_restriction`` between ``question_appears_in`` and
``level_of_difficulty``, ``question_text`` directly after ``question``, and
``answer_explanation`` closing the Objective/Subjective bands.

Two renderers project ONE snapshot (spec §1, §13.1):

* Concept File — hierarchy catalogue only: chapter/topic/concept bands in
  source order, Group and Question bands blank, Descriptive and Subjective
  header-only.
* Master File — the complete import: every concept including questionless
  ones, every occupied group, and every Objective, Descriptive, and Subjective
  assessment with answers, rubrics, subquestions, restriction, the exact MES
  appears-in wire value, and complete label aggregates. Required but unused
  group shells remain in the release snapshot and are named in the issues
  manifest instead of becoming empty assessment rows.

Deterministic mechanics only: positional serialization, formula-injection
escaping, cell limits, read-back parsing, and count/arithmetic validation.
Nothing here decides content — flags and unplaced candidates ride into the
returned manifest, never disappear.
"""
from __future__ import annotations

import hashlib
import io
import json
import math
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font

from . import ANSWER_TYPES
from ..services import assessment_profile
from ..services import assessment_release as rel
from ..services import identity
from ..services import katex_rules
from . import layouts

CELL_LIMIT = 32_767

_MANIFEST_PATH = Path(__file__).with_name("assessment_workbook_template.json")
MANIFEST: dict = json.loads(_MANIFEST_PATH.read_text(encoding="utf-8"))
SHEET_ORDER: list[str] = list(MANIFEST["sheet_order"])
FIELDS: dict[str, list[str]] = {
    name: list(sheet["fields"]) for name, sheet in MANIFEST["sheets"].items()
}
BANDS: dict[str, list[dict]] = {
    name: list(sheet["bands"]) for name, sheet in MANIFEST["sheets"].items()
}
_INDEX: dict[str, dict[str, int]] = {
    name: {field: i for i, field in enumerate(fields)}
    for name, fields in FIELDS.items()
}

# Which sheet kinds ``_question_record`` can write TODAY — a property of
# THIS CODE, not of any school's profile and not of the layout.  It has
# explicit branches for all three canonical answer-style sheets.
#
# ``sheet_for_kind`` is derived from THIS tuple × the layout's own sheet-name
# map, deliberately NOT from the profile: a profile-derived map has no
# ``None`` branch to take the moment a profile widens, so the defect could
# never fire and ``_question_record`` would raise instead.
RENDERABLE_SHEET_KINDS = ("objective", "descriptive", "subjective")


def sheet_for_kind() -> dict[str, str]:
    """``RENDERABLE_SHEET_KINDS`` × the layout's sheet-name map."""

    names = layouts.layout(layouts.REFERENCE_LAYOUT_ID).sheet_name_by_kind()
    return {
        kind: names[kind] for kind in RENDERABLE_SHEET_KINDS if kind in names
    }


MAX_OBJECTIVE_OPTIONS = 6
MAX_SUBJECTIVE_ANSWERS = 20
MAX_DESCRIPTIVE_ANSWERS = 10
MAX_SUBQUESTIONS = 15
MAX_SUBQUESTION_KEYWORDS = 6

_UPDATE_FIELD_AFTER = {
    "is_update_chapter": "chapter_title",
    "is_update_topic": "topic_title",
    "is_update_concept": "concept_title",
    "is_update_question": "question_label",
}

# One owner for both render and read-back of the update-marker contract.  The
# identity probes deliberately exclude the marker itself and non-identifying
# decoration: a concept-only tail has Chapter/Topic/Concept markers, but no
# Group/Question markers.
_UPDATE_FIELD_PRESENCE = {
    "is_update_chapter": ("chapter_title",),
    "is_update_topic": ("topic_title",),
    "is_update_concept": ("concept_title",),
    "is_update_group": (
        "group_name", "group_display_name", "group_type",
    ),
    "is_update_question": ("question_label", "question"),
}


def _snapshot_learning_phase(snapshot: Mapping | None) -> str:
    """One explicit topic lane, or blank when a snapshot is mixed/unknown."""

    phases = {
        str(topic.get("pre_post_learning") or "").strip().casefold()
        for topic in (snapshot or {}).get("topics") or []
        if isinstance(topic, Mapping)
        and str(topic.get("pre_post_learning") or "").strip()
    }
    return next(iter(phases)) if len(phases) == 1 else ""


def _insert_after(fields: list[str], anchor: str, value: str) -> None:
    if value in fields:
        return
    fields.insert(fields.index(anchor) + 1, value)


def _master_fields(contract: Mapping[str, Any]) -> dict[str, list[str]]:
    """Build a clean Master schema from the committed Concept schema.

    This is positional mechanics only.  It removes the possibility of the
    duplicate headers and styled trailing columns present in the hand-edited
    audit files: every field is inserted once into a unique base list.
    """

    fields_by_sheet = {name: list(fields) for name, fields in FIELDS.items()}
    include_updates = bool(contract.get("include_update_fields"))
    include_descriptive_source = bool(
        contract.get("include_descriptive_concept_source")
    )
    try:
        descriptive_slots = max(
            MAX_DESCRIPTIVE_ANSWERS,
            int(contract.get(
                "descriptive_answer_slots", MAX_DESCRIPTIVE_ANSWERS,
            )),
        )
    except (TypeError, ValueError):
        descriptive_slots = MAX_DESCRIPTIVE_ANSWERS

    for sheet, fields in fields_by_sheet.items():
        if include_updates:
            for update_field, anchor in _UPDATE_FIELD_AFTER.items():
                _insert_after(fields, anchor, update_field)
            group_anchor = (
                "group_display_name"
                if sheet == "Descriptive"
                else "group_name"
            )
            _insert_after(fields, group_anchor, "is_update_group")
        if (
            sheet == "Descriptive"
            and include_descriptive_source
            and "concept_source" not in fields
        ):
            _insert_after(fields, "concept_question_labels", "concept_source")

    descriptive = fields_by_sheet["Descriptive"]
    answer_explanation = descriptive.index("answer_explanation")
    expanded_answers = [
        field
        for number in range(
            MAX_DESCRIPTIVE_ANSWERS + 1, descriptive_slots + 1,
        )
        for field in (
            f"answer_type_{number}",
            f"answer_weightage_{number}",
            f"answer_content_{number}",
        )
    ]
    descriptive[answer_explanation:answer_explanation] = expanded_answers
    return fields_by_sheet


def _master_bands(
    fields_by_sheet: Mapping[str, list[str]],
    contract: Mapping[str, Any],
) -> dict[str, list[dict]]:
    """Rebase row-1 bands by their authoritative boundary field names."""

    bands_by_sheet: dict[str, list[dict]] = {}
    for sheet, base_bands in BANDS.items():
        base_fields = FIELDS[sheet]
        final_fields = fields_by_sheet[sheet]
        rebased = []
        for band in base_bands:
            start_field = base_fields[band["start"] - 1]
            end_field = base_fields[band["end"] - 1]
            end = final_fields.index(end_field) + 1
            if (
                sheet == "Descriptive"
                and band["label"].strip() == "Concept"
                and contract.get("include_descriptive_concept_source")
            ):
                end = final_fields.index("concept_source") + 1
            rebased.append({
                "label": band["label"],
                "start": final_fields.index(start_field) + 1,
                "end": end,
            })
        bands_by_sheet[sheet] = rebased
    return bands_by_sheet


def output_schema(
    role: str,
    profile: Mapping | str | None = None,
    snapshot: Mapping | None = None,
) -> dict[str, Any]:
    """Return exact fields/bands/capacity for one deterministic output role.

    Concept files always use the committed 67/374/144 schema. Master files
    read the resolved run profile and the snapshot's explicit Pre/Post lane.
    The returned structures are copies so callers cannot mutate module state.
    """

    normalized_role = str(role or "").strip().casefold()
    if normalized_role == "concept":
        return {
            "role": "concept",
            "contract_id": "concept-reference-1",
            "fields": {name: list(fields) for name, fields in FIELDS.items()},
            "bands": {
                name: [dict(band) for band in bands]
                for name, bands in BANDS.items()
            },
            "descriptive_answer_slots": MAX_DESCRIPTIVE_ANSWERS,
        }
    if normalized_role != "master":
        raise ValueError(f"unknown workbook output role {role!r}")
    contract = assessment_profile.master_workbook_contract(
        profile,
        learning_phase=_snapshot_learning_phase(snapshot),
    )
    fields_by_sheet = _master_fields(contract)
    return {
        "role": "master",
        "contract_id": str(contract.get("contract_id") or ""),
        "fields": fields_by_sheet,
        "bands": _master_bands(fields_by_sheet, contract),
        "descriptive_answer_slots": int(
            contract.get("descriptive_answer_slots")
            or MAX_DESCRIPTIVE_ANSWERS
        ),
        "natural_label_aggregates": bool(
            contract.get("natural_label_aggregates", False)
        ),
        "aggregate_rendered_questions_only": bool(
            contract.get("aggregate_rendered_questions_only", False)
        ),
    }


def _schema_identity(schema: Mapping[str, Any]) -> dict[str, str]:
    """Join one exact output schema to the registered reader layout."""

    headers = {
        sheet: tuple(schema["fields"][sheet])
        for sheet in SHEET_ORDER
    }
    identified = layouts.identify_workbook(headers)
    return {
        "layout_id": identified.layout_id,
        "contract_id": str(schema.get("contract_id") or ""),
    }


def output_identities(
    profile: Mapping | str | None = None,
    snapshot: Mapping | None = None,
) -> dict[str, dict[str, str]]:
    """Registered layout + schema contract identity for both projections."""

    return {
        "concepts_xlsx": _schema_identity(
            output_schema("concept", profile, snapshot)
        ),
        "master_xlsx": _schema_identity(
            output_schema("master", profile, snapshot)
        ),
    }


def _natural_label_key(value: str) -> tuple:
    """Stable human ordering (``Q2`` before ``Q10``), without mutation."""

    return tuple(
        (0, int(token)) if token.isdigit() else (1, token.casefold())
        for token in re.split(r"(\d+)", str(value))
    )


class WorkbookRenderError(ValueError):
    """A workbook cannot be rendered without violating a mechanical rule.

    Since spec-step8 S9 NOTHING in the Master or Concept lane raises it.
    ``render_master_file``, ``_question_record`` and ``_cell_value`` all
    record instead, because a raise at projection time costs every row on
    every sheet of all four outputs while the same verdict is available at
    staging (T7.5/B4, and S9 for the two cell-shape cases).  The class is
    kept — it is public API and a caller outside this package may still be
    catching it — but this module no longer raises it.
    """


# THE TWO SHAPES NO XLSX CELL CAN HOLD, in ONE place (T1: one
# implementation, two callers) — and ONE cell can carry BOTH at once, which
# is why the predicate below answers with a list.  ``_cell_value`` repairs
# every shape a cell trips, not the first one;
# ``assessment_release.unresolved_question_homes`` calls the SAME predicate
# at staging and names one ``render_shape_overflow`` finding per shape,
# which refuses the database write.  Two inline copies of one rule is
# exactly the drift that put the staged verdict and the renderer out of
# step over renderable sheet kinds.
#
# Neither is a judgment about content: 32 767 is the XLSX format's own
# per-cell character cap, and the illegal-character set is openpyxl's own
# ``ILLEGAL_CHARACTERS_RE`` — the code points XML 1.0 forbids in character
# data.  Both are read from the format, not authored here.

CELL_TEXT_TOO_LONG = "cell_text_too_long"
CELL_TEXT_ILLEGAL_CHARACTER = "cell_text_illegal_character"


def cell_text_defects(value: Any) -> list[dict[str, Any]]:
    """EVERY reason this value cannot be written into one XLSX cell.

    A LIST, and that is the whole point of the plural.  [measured] while
    this returned the FIRST reason only, a value that was both over the cap
    and carrying an XML-illegal code point reported ``cell_text_too_long``
    alone; ``_cell_value`` dispatched on that one reason, truncated, never
    stripped, and openpyxl raised ``IllegalCharacterError`` out of
    ``wb.save`` — the exact all-four-outputs loss S9 exists to remove, and
    invisible at staging because the finding named the other half.  A value
    can be wrong in two ways at once and the record has to say both.

    Ordered ``too long`` first, then ``illegal characters``, so the primary
    entry stays the one earlier readers saw for a singly-defective value.
    """

    if value is None or (
        isinstance(value, (int, float)) and not isinstance(value, bool)
    ):
        return []
    text = str(value)
    defects: list[dict[str, Any]] = []
    if len(text) > CELL_LIMIT:
        defects.append({
            "reason": CELL_TEXT_TOO_LONG,
            "cap": CELL_LIMIT,
            "actual": len(text),
        })
    illegal = ILLEGAL_CHARACTERS_RE.findall(text)
    if illegal:
        defects.append({
            "reason": CELL_TEXT_ILLEGAL_CHARACTER,
            "cap": 0,
            "actual": len(illegal),
            "code_points": sorted({f"U+{ord(ch):04X}" for ch in illegal}),
        })
    return defects


# What a truncated cell says in the workbook itself, so a reviewer reading
# the file — not only ``manifest.json`` — can see that the tail is
# elsewhere.  R4: the silence is the defect.
_TRUNCATION_MARK = (
    "\n[Aegis: this cell holds the first {kept} of {actual} characters. The "
    "complete value is recorded in this release's issues manifest under "
    "'oversized_cells' and nothing has been lost.]"
)


def _cell_value(
    value: Any, *, context: str, oversized: list[dict] | None = None,
) -> Any:
    """One cell's value, repaired rather than refused (spec-step8 S9).

    [measured at 76c84fb] a 40 000-character question raised
    ``WorkbookRenderError`` out of ``render_master_file`` and took all four
    outputs with it, and a single ``\\x01`` reached openpyxl and raised
    ``IllegalCharacterError`` from ``wb.save`` — while the staged verdict
    returned ``[]`` for both, so nothing refused the database write either.
    Both are now repaired here, named at staging as
    ``render_shape_overflow``, and recorded WHOLE in ``oversized`` — the
    ledger that becomes the issues manifest's ``oversized_cells``, beside
    S8's ``truncated_rows``.

    THE TRUNCATION QUESTION, answered rather than defaulted: the cell keeps
    as much as the format allows and says in the cell that it was cut, and
    the FULL value rides the manifest.  Dropping the tail silently is the
    one thing R4 forbids; refusing the file is the one thing Rule E
    forbids; and no reviewer can read a 40 000-character answer inside a
    spreadsheet cell anyway.

    EVERY REPAIR IS APPLIED, IN THIS ORDER, and the order is load-bearing.
    [measured] a value that was BOTH over the cap and carrying ``\\x01``
    reported one reason, took the branch for that reason, and shipped the
    illegal code point into openpyxl — which raised out of ``wb.save`` and
    cost all four outputs, the very failure this function was written to
    end.  So: strip the code points no cell may carry, decide the
    formula-like text on what SURVIVES that strip, then truncate. The Cell
    writer explicitly marks that text as a string; mutating its payload with
    an apostrophe corrupts legitimate negative Equation answers such as
    ``-3``.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)
    defects = cell_text_defects(text)
    if defects and oversized is not None:
        for defect in defects:
            oversized.append({
                "context": str(context),
                "reason": defect["reason"],
                "cap": defect["cap"],
                "actual": defect["actual"],
                "code_points": defect.get("code_points", []),
                # The whole value, unabridged. This is the record that
                # makes truncation a repair rather than a loss.
                "full_value": text,
            })
    if any(
        defect["reason"] == CELL_TEXT_ILLEGAL_CHARACTER for defect in defects
    ):
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
    budget = CELL_LIMIT
    if len(text) > budget:
        actual = len(text)
        # Two passes, because the note's own length depends on the number
        # it states.  The first pass sizes the note with a count that is
        # never narrower than the real one, so ``kept`` comes out
        # conservative; the second states the truth.  [measured] the single
        # pass this replaces formatted the TEMPLATE and cut by the
        # FORMATTED mark, so a 40 000-character value shipped a cell that
        # said it held 32 585 characters while holding 32 589 — a false
        # sentence inside a repair whose justification is that the record
        # stays true.
        kept = budget - len(
            _TRUNCATION_MARK.format(kept=budget, actual=actual))
        mark = _TRUNCATION_MARK.format(kept=kept, actual=actual)
        kept = min(kept, budget - len(mark))
        text = text[:kept] + mark
    return text


def _row_values(
    sheet: str, record: Mapping[str, Any],
    forced_blank: tuple[str, ...] = (),
    *, oversized: list[dict] | None = None,
    schema: Mapping[str, Any] | None = None,
) -> list:
    """Serialize one record positionally.

    ``forced_blank`` is the profile's ``forced_blank_fields`` (spec-step8
    T12/M4) — which fields ship blank whatever the row carries is a school's
    fill practice, not a fact about the format, and the same key is read by
    ``validate_concept_file``'s must-keep-blank gate so the renderer and the
    read-back can never disagree.
    """
    blank = frozenset(forced_blank)
    fields = (schema or output_schema("concept"))["fields"][sheet]
    materialized = dict(record)
    for update_field, identity_fields in _UPDATE_FIELD_PRESENCE.items():
        if update_field not in fields:
            continue
        populated = any(
            value is not None and str(value).strip()
            for value in (materialized.get(name) for name in identity_fields)
        )
        materialized[update_field] = "No" if populated else ""
    row = []
    for field in fields:
        value = "" if field in blank else materialized.get(field, "")
        row.append(_cell_value(
            value, context=f"{sheet}:{field}", oversized=oversized))
    return row


# A11 (owner audit, 2026-08-29): the numeric fields whose cells display
# one decimal place. chapter_duration is deliberately excluded — the
# corrected files carry it plain.
_ONE_DECIMAL_FIELD_RE = re.compile(
    r"^(?:marks|question_duration"
    r"|answer_weightage_\d+"
    r"|weightage_\d+"
    r"|sub_question_marks_\d+"
    r"|sq\d+_weightage_\d+)$"
)


def _append_record(
    ws, sheet: str, record: Mapping[str, Any],
    forced_blank: tuple[str, ...] = (),
    *, oversized: list[dict] | None = None,
    schema: Mapping[str, Any] | None = None,
) -> None:
    """Append exact content and literalize every formula-like string.

    Assessment workbooks carry learner content, never executable formulas.
    Explicit string typing blocks formula injection without adding an
    apostrophe to the stored payload, so negative mathematical answers remain
    byte-exact on read-back.
    """

    values = _row_values(
        sheet, record, forced_blank, oversized=oversized, schema=schema,
    )
    ws.append(values)
    row_number = ws.max_row
    active_fields = (schema or output_schema("concept"))["fields"][sheet]
    for column, value in enumerate(values, start=1):
        if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
            ws.cell(row=row_number, column=column).data_type = "s"
        elif (
            isinstance(value, (int, float))
            and not isinstance(value, bool)
            and column <= len(active_fields)
            and _ONE_DECIMAL_FIELD_RE.match(active_fields[column - 1])
        ):
            # A11 (owner audit): marks, durations, and weightages display
            # with one decimal ("1.0"), matching the corrected files. The
            # stored value stays numeric — this is presentation only.
            ws.cell(row=row_number, column=column).number_format = "0.0"


def _write_headers(
    ws, sheet: str, schema: Mapping[str, Any] | None = None,
) -> None:
    active = schema or output_schema("concept")
    for band in active["bands"][sheet]:
        cell = ws.cell(row=1, column=band["start"])
        cell.value = band["label"]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if band["end"] > band["start"]:
            ws.merge_cells(
                start_row=1, start_column=band["start"],
                end_row=1, end_column=band["end"])
    for i, field in enumerate(active["fields"][sheet], start=1):
        cell = ws.cell(row=2, column=i)
        cell.value = field
        cell.font = Font(bold=True, size=9)
    ws.freeze_panes = "A3"


def _new_workbook(
    schema: Mapping[str, Any] | None = None,
) -> openpyxl.Workbook:
    active = schema or output_schema("concept")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        _write_headers(wb.create_sheet(name), name, active)
    return wb


def _workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Snapshot access
# --------------------------------------------------------------------------- #
# The snapshot is one immutable dict (the release projection):
#   chapter: {chapter_title, chapter_display_name, pre_topics, post_topics,
#             chapter_description}
#   topics:  ordered [{topic fields..., "concepts": ordered [concept fields
#            ..., "concept_key"]}]
#   groups:  [group records (assessment_grouping.group_record) each with
#            "concept_key"]
#   candidates: [materialized candidates with "question_label" and
#            "concept_key"/"group_key" filled from their placement]
# ``concept_key`` is any stable join key (machine id); titles are display
# values and never used for joining.


def _concept_rows(snapshot: Mapping) -> list[dict]:
    rows = []
    chapter = dict(snapshot.get("chapter") or {})
    lane_positions: dict[str, int] = {}
    for topic in snapshot.get("topics") or []:
        lane = identity.lane_token(str(topic.get("pre_post_learning") or ""))
        topic_ordinal = lane_positions.get(lane, 0) + 1
        lane_positions[lane] = topic_ordinal
        topic_fields = {
            k: v for k, v in topic.items() if k != "concepts"
        }
        for concept in topic.get("concepts") or []:
            rows.append({
                "chapter": chapter,
                "topic": topic_fields,
                "topic_ordinal": topic_ordinal,
                "concept": dict(concept),
            })
    return rows


# The concept identity trio: join keys and provenance, never workbook cells.
# ``_row_values`` iterates template fields, so a record key with no column
# vanishes with no flag — stripping them HERE is what keeps T3.6's
# manifest-union gate from reporting two identity fields as structural
# defects on every release.
_IDENTITY_TRIO = ("concept_key", "concept_machine_id", "release_row_identity")
# The topic's identity key, stripped from cells for the same reason.
_TOPIC_IDENTITY = ("topic_machine_id",)


def _titled_topic(topic: Mapping, ordinal: int = 1) -> str:
    """The shared topic-title cell, numbered within its snapshot lane."""
    title = str(topic.get("topic_title") or "")
    machine_id = str(topic.get("topic_machine_id") or "").strip()
    return identity.topic_title_cell(title, machine_id, ordinal)


def _identity_tag_replacement(kind: str, mapping: Mapping) -> dict | None:
    """A record for a baked minted-grammar tag replaced by a DIFFERENT id.

    ``identity.titled`` replaces a minted-shaped carried tag with the
    persisted id; when the two disagree the loser must not vanish silently
    (R4 — the reader records the mirror case as ``machine_id_conflict``).
    Detection only; the replacement itself stays with ``identity.titled``.
    """
    title_key, id_key = (
        ("topic_title", "topic_machine_id") if kind == "topic"
        else ("concept_title", "concept_machine_id")
    )
    raw = str(mapping.get(title_key) or "")
    machine_id = str(mapping.get(id_key) or "").strip()
    if not machine_id:
        return None
    carried = identity.title_tag(raw)
    if not carried or carried == machine_id:
        return None
    minted_shape = (
        identity.minted_topic_ordinal(carried) is not None
        or identity.minted_concept_ordinal(carried) is not None
    )
    if not minted_shape:
        return None
    return {
        "kind": kind,
        "title": raw,
        "carried_tag": carried,
        "applied_machine_id": machine_id,
    }


def _titled_concept(concept: Mapping) -> str:
    """The concept title CELL: the shared composed identity (T14/B4).

    ``identity.titled`` over the persisted ``concept_machine_id``, so the
    cell reads exactly as ``topic_concept_labels`` names it and a re-import
    restores the persisted id instead of re-minting. A concept with no id
    keeps its title verbatim — the gold reference snapshots never mint, and
    ``titled(name, "")`` would strip a tag baked into a gold title.
    """
    title = str(concept.get("concept_title") or "")
    machine_id = str(concept.get("concept_machine_id") or "").strip()
    return identity.titled(title, machine_id) if machine_id else title


def _bands_record(entry: Mapping) -> dict:
    record: dict = {}
    record.update(entry["chapter"])
    record.update({
        k: v for k, v in entry["topic"].items() if k not in _TOPIC_IDENTITY
    })
    record["topic_title"] = _titled_topic(
        entry["topic"], int(entry.get("topic_ordinal") or 1),
    )
    record.update({
        k: v for k, v in entry["concept"].items() if k not in _IDENTITY_TRIO
    })
    # The one composition point for the title cell of every row this
    # concept owns, in both renderers (spec-step8 T14/B4): before this,
    # Outputs 02/04 shipped the bare title while the roster carried the
    # tagged one, and a re-import re-minted every identity.
    record["concept_title"] = _titled_concept(entry["concept"])
    return record


def snapshot_sha256(snapshot: Mapping) -> str:
    return rel.sha256_json(snapshot)


# --------------------------------------------------------------------------- #
# Output A — Concept File (spec §1, §9)
# --------------------------------------------------------------------------- #

def render_concept_file(
    snapshot: Mapping, profile: Mapping | str | None = None,
    *, oversized: list[dict] | None = None,
) -> bytes:
    """Output 01/03 — the Concept File.

    ``oversized`` is the optional cell-shape ledger (spec-step8 S9). This
    renderer returns bytes rather than (bytes, issues), so the caller that
    wants the record passes a list in; passing none loses nothing a
    reviewer sees, because the SAME defect is named at staging as
    ``render_shape_overflow`` on the concept row it belongs to.
    """
    forced_blank = assessment_profile.forced_blank_fields(profile)
    schema = output_schema("concept", profile, snapshot)
    wb = _new_workbook(schema)
    ws = wb["Objective"]
    for entry in _concept_rows(snapshot):
        record = _bands_record(entry)
        # A clean catalogue: Group and Question bands stay blank, and the
        # concept-band group labels stay blank too — groups do not exist in
        # Output A (spec §1: "no populated Group fields").
        for field in ("basic_groups", "intermediate_groups",
                      "advanced_groups", "concept_question_labels"):
            record[field] = ""
        _append_record(
            ws, "Objective", record, forced_blank, oversized=oversized,
            schema=schema,
        )
    return _workbook_bytes(wb)


# --------------------------------------------------------------------------- #
# Output B — Master File (spec §1, §10)
# --------------------------------------------------------------------------- #

def _question_record(
    candidate: Mapping, sheet: str, profile: Mapping | str | None = None,
    *, truncated: list[dict] | None = None,
    descriptive_answer_slots: int = MAX_DESCRIPTIVE_ANSWERS,
) -> dict:
    """Project one candidate onto its sheet's question columns.

    THIS FUNCTION RAISES NOTHING (spec-step8 T7.5/B4).  Its seven former
    raises are gone: ``question_duration``, objective ``math_keyboard`` and
    descriptive ``math_keyboard`` already have a staging twin in
    ``assessment_release.validate_candidate``, and the four count caps are
    named at staging as ``render_shape_overflow``.  What it does with a row
    that overflows the layout is write what the layout HAS slots for — the
    row ships and is visible, the overflow is named, and the database write
    is refused — instead of costing every row on every sheet of all four
    outputs.

    ``truncated`` is the renderer's own ledger.  Every one of the four caps
    appends ``{question_label, field, cap, actual}`` to it, so the loss is
    recorded in the issues manifest that lands on disk and not only in the
    staged verdict — a caller outside ``create_release`` used to truncate
    with no record anywhere (R4: the silence is the defect).
    """
    def _record_shape(field: str, value: Any) -> None:
        if truncated is not None:
            truncated.append({
                "candidate_id": str(candidate.get("candidate_id") or ""),
                "question_label": str(candidate.get("question_label") or ""),
                "field": field,
                "reason": "invalid_container_shape",
                "expected": "array of objects",
                "actual_type": type(value).__name__,
            })

    def _mapping_array(field: str, value: Any) -> list[Mapping]:
        """Project only object members while recording every shape repair."""

        if value is None:
            return []
        if not isinstance(value, list):
            _record_shape(field, value)
            return []
        items: list[Mapping] = []
        for position, item in enumerate(value, start=1):
            if isinstance(item, Mapping):
                items.append(item)
            else:
                _record_shape(f"{field}[{position}]", item)
        return items

    def _cap(field: str, items: list, cap: int) -> list:
        if len(items) > cap and truncated is not None:
            truncated.append({
                "candidate_id": str(candidate.get("candidate_id") or ""),
                "question_label": str(candidate.get("question_label") or ""),
                "field": field,
                "cap": cap,
                "actual": len(items),
            })
        return items[:cap]

    record = {
        "question_label": candidate.get("question_label", ""),
        "question_category": candidate.get("question_category", ""),
        "cognitive_skills": candidate.get("cognitive_skill", ""),
        "question_source": candidate.get(
            "question_source", assessment_profile.question_source(profile)),
        "question_disclaimer": candidate.get("question_disclaimer", ""),
        "question_duration": candidate.get("question_duration", ""),
        "question_appears_in": candidate.get("question_appears_in", ""),
        "answer_restriction": candidate.get("answer_restriction", ""),
        "level_of_difficulty": candidate.get("difficulty", ""),
        "question": candidate.get("question", ""),
        "question_text": candidate.get("question_text", ""),
        "marks": candidate.get("marks", ""),
        "answer_explanation": candidate.get("answer_explanation", ""),
    }
    answers = _mapping_array("answers", candidate.get("answers"))
    if sheet == "Objective":
        # The layout has exactly this many slots; an overflow is named at
        # staging as ``render_shape_overflow`` and refuses the DB write.
        capped_answers = _cap("answers", answers, MAX_OBJECTIVE_OPTIONS)
        for n, answer in enumerate(capped_answers, start=1):
            record[f"answer_type_{n}"] = answer.get("answer_type", "")
            # The declared medium controls the whole CMS cell.  Equation is
            # raw LaTeX with no [Katex] wrapper; Phrases stays plain text.
            record[f"answer_content_{n}"] = katex_rules.raw_answer_cell(
                answer.get("answer_type", ""),
                answer.get("answer_content", ""),
            )
            # The CMS import (and the reference workbooks) mark the
            # correct option "Yes" / wrong options "No"; the pipeline's
            # internal wire uses "1"/"0". Normalize at the render seam
            # against the one shared predicate — the read-back and the
            # reader accept both spellings, so the round-trip holds.
            record[f"correct_answer_{n}"] = (
                "Yes" if rel.is_correct_option(
                    answer.get("correct_answer")) else "No"
            ) if str(answer.get("correct_answer", "")).strip() != "" else ""
            record[f"answer_weightage_{n}"] = answer.get(
                "answer_weightage", "")
        # ``question_text`` is the CMS's one-cell rendering of the whole
        # item: for an objective question that is the stem PLUS the
        # ordered options (the owner's ruling; the reference workbooks
        # carry the options in the text). The stem-equality gates
        # upstream stay untouched — this is a render-time composition of
        # already-decided content, options lettered in answer order.
        option_lines = [
            f"{chr(ord('a') + index)}) "
            + katex_rules.rich_answer_display(
                a.get("answer_type", ""),
                str(a.get("answer_content") or ""),
            )
            for index, a in enumerate(capped_answers)
            if str(a.get("answer_content") or "").strip()
        ]
        if option_lines:
            record["question_text"] = (
                str(record.get("question_text") or "").rstrip()
                + "\n" + "\n".join(option_lines)
            ).strip()
    elif sheet == "Subjective":
        record["math_keyboard"] = candidate.get("math_keyboard", "")
        for n, answer in enumerate(
            _cap("answers", answers, MAX_SUBJECTIVE_ANSWERS), start=1
        ):
            answer_type = answer.get("answer_type", "")
            content = answer.get("answer_content", "")
            record[f"answer_type_{n}"] = answer_type
            record[f"answer_{n}"] = katex_rules.raw_answer_cell(
                answer_type, content,
            )
            record[f"answer_display_{n}"] = (
                answer.get("answer_display")
                or katex_rules.rich_answer_display(answer_type, content)
            )
            record[f"weightage_{n}"] = answer.get(
                "answer_weightage", ""
            )
            record[f"placeholder_{n}"] = answer.get("placeholder", "")
    else:  # Descriptive
        record["math_keyboard"] = candidate.get("math_keyboard", "")
        record["display_answer"] = candidate.get("display_answer", "")
        # The newer source-audit contract keeps part text only in the
        # dedicated sub_question_N columns. Re-appending it to question_text
        # duplicates both the visible task and, downstream, its scoring.
        for n, answer in enumerate(
            _cap("answers", answers, descriptive_answer_slots), start=1
        ):
            record[f"answer_type_{n}"] = answer.get("answer_type", "")
            record[f"answer_weightage_{n}"] = answer.get(
                "answer_weightage", "")
            record[f"answer_content_{n}"] = katex_rules.raw_answer_cell(
                answer.get("answer_type", ""),
                answer.get("answer_content", ""),
            )
        sub_questions = _mapping_array(
            "sub_questions", candidate.get("sub_questions"),
        )
        for n, sub in enumerate(
            _cap("sub_questions", sub_questions, MAX_SUBQUESTIONS), start=1
        ):
            record[f"sub_question_{n}"] = sub.get("text", "")
            record[f"sub_question_marks_{n}"] = sub.get("marks", "")
            keywords = _cap(
                f"sub_questions[{n}].keywords",
                _mapping_array(
                    f"sub_questions[{n}].keywords", sub.get("keywords"),
                ),
                MAX_SUBQUESTION_KEYWORDS)
            for m, keyword in enumerate(keywords, start=1):
                record[f"sq{n}_answer_type_{m}"] = keyword.get(
                    "answer_type", "")
                record[f"sq{n}_weightage_{m}"] = keyword.get("weightage", "")
                # SOP §4.3: keyword cells are raw for Equation/Image too.
                record[f"sq{n}_keyword_{m}"] = katex_rules.raw_answer_cell(
                    keyword.get("answer_type", ""),
                    keyword.get("keyword", ""),
                )
    # Release freeze refuses unsupported table dialects for fresh decisions.
    # Keep the renderer itself total for accepted legacy/direct callers too:
    # mechanically project every untyped rich-text table to ordered
    # row/column labels.  This is the last serializer seam, not a semantic
    # reconstruction; already-authored source images remain untouched.
    for field in (
        "question", "question_text", "display_answer", "answer_explanation",
    ):
        if field in record:
            record[field] = katex_rules.replace_unsupported_tables(
                str(record.get(field) or "")
            )
    for n in range(1, MAX_SUBQUESTIONS + 1):
        field = f"sub_question_{n}"
        if field in record:
            record[field] = katex_rules.replace_unsupported_tables(
                str(record.get(field) or "")
            )
    return record


def _group_record_fields(
    group: Mapping, group_labels: list[str],
    profile: Mapping | str | None = None,
) -> dict:
    return {
        "group_name": group.get("group_name", ""),
        "group_display_name": group.get("group_display_name", ""),
        "group_description": group.get("semantic_description", ""),
        "group_status": group.get(
            "group_status", assessment_profile.group_status(profile)),
        "group_type": group.get("group_type", ""),
        "group_question_labels": ", ".join(group_labels),
        "related_digicards": "",
    }


def render_master_file(
    snapshot: Mapping, profile: Mapping | str | None = None,
) -> tuple[bytes, dict]:
    """Render the Master File; returns (bytes, issues manifest).

    THIS FUNCTION RAISES NOTHING (spec-step8 T7.5/B4).  Its seven former
    raises are gone.  Three of them (blank ``group_key``, duplicate
    ``group_key``, invalid ``group_type``) already have a staging twin in
    ``assessment_release.validate_group`` / ``duplicate_group_keys``; the
    first two ALSO have a read-back twin in ``validate_master_file``, but
    invalid ``group_type`` does NOT — the read-back compares ``group_type``
    only against the snapshot's own value, never against the valid set, so
    an invalid value that round-trips consistently passes it; the one
    validity twin is ``validate_group`` at freeze (the S8 commit measured
    this; an earlier version of this docstring claimed a read-back twin
    for all three).  The other four are
    named at staging by ``assessment_release.unresolved_question_homes`` as
    ``group_concept_home_unknown``, ``group_home_unnamed``,
    ``group_visible_name_mismatch`` and ``group_home_disagreement``.  A raise
    here would cost NO WORKBOOK FOR ANY OF THE FOUR OUTPUTS while the same
    verdict is already recorded on the release and already refuses the
    database write.

    Nothing disappears: a candidate whose placement never resolved to a
    concept/group cannot be positioned in the hierarchy, so it is recorded
    in the issues manifest (spec §13.3 keeps the full ledger in the release
    manifest, never an unrecognized extra sheet); a group the renderer
    cannot place is recorded in ``issues["group_defects"]`` for the same
    reason.
    """
    forced_blank = assessment_profile.forced_blank_fields(profile)
    schema = output_schema("master", profile, snapshot)
    descriptive_answer_slots = int(schema["descriptive_answer_slots"])
    concept_entries = {
        str(entry["concept"].get("concept_key") or ""): entry
        for entry in _concept_rows(snapshot)
    }
    concept_order = [
        str(entry["concept"].get("concept_key") or "")
        for entry in _concept_rows(snapshot)
    ]
    groups = [dict(g) for g in snapshot.get("groups") or []]
    group_defects: list[dict] = []

    def _group_defect(group_key: str, code: str, message: str) -> None:
        group_defects.append(
            {"group_key": group_key, "code": code, "message": message})

    groups_by_key: dict[str, dict] = {}
    renderable_groups: list[dict] = []
    for group in groups:
        group_key = str(group.get("group_key") or "")
        if not group_key:
            _group_defect(
                "", rel.GROUP_KEY_BLANK,
                "group_key must not be blank")
            continue
        if group_key in groups_by_key:
            _group_defect(
                group_key, rel.GROUP_KEY_DUPLICATE,
                f"duplicate group_key {group_key!r}")
            continue
        groups_by_key[group_key] = group
        concept_key = str(group.get("concept_key") or "")
        concept_entry = concept_entries.get(concept_key)
        if concept_entry is None:
            # No hierarchy row to hang a catalogue row off. Named at
            # staging as ``group_concept_home_unknown``; recorded here.
            _group_defect(
                group_key, rel.GROUP_CONCEPT_HOME_UNKNOWN,
                f"group {group_key!r} has unknown concept home "
                f"{concept_key!r}")
            continue
        renderable_groups.append(group)
    candidates = [dict(c) for c in snapshot.get("candidates") or []]

    sheets_by_kind = sheet_for_kind()
    renderable = tuple(sheets_by_kind)
    profile_name = assessment_profile.name(profile)

    # Complete ordered label aggregates (spec §8.4).
    #
    # The three "can this candidate be written to a data row" branches are
    # NOT re-discovered here: ``assessment_release.candidate_home_finding``
    # is the one implementation and the staged verdict calls the same
    # function on the same four inputs (T1). Two inline copies had already
    # drifted on the renderable set — see that function's docstring.
    concept_labels: dict[str, list[str]] = {}
    group_labels: dict[str, list[str]] = {}
    unplaced: list[dict] = []
    placed: list[tuple[dict, dict | None]] = []

    def _record_unplaced(candidate: Mapping, finding: Mapping) -> None:
        unplaced.append({
            "candidate_id": finding["candidate_id"],
            "question_label": finding["question_label"],
            "code": finding["code"],
            "reason": finding["message"],
            "flags": list(candidate.get("flags") or []),
        })

    for candidate in candidates:
        concept_key = str(candidate.get("concept_key") or "")
        group_key = str(candidate.get("group_key") or "")
        label = str(candidate.get("question_label") or "")
        home = rel.candidate_home_finding(
            candidate, concept_entries, groups_by_key, renderable,
            profile_name)
        if home is not None and home["code"] != rel.SHEET_KIND_NOT_RENDERABLE:
            _record_unplaced(candidate, home)
            continue
        # A candidate whose HOME resolved still contributes its label to
        # the ordered aggregates even when this step cannot render its
        # sheet kind — the label is a fact about the concept and the group,
        # not about which sheet the row lands on, and
        # ``validate_master_file``'s read-back expects exactly this set.
        if (
            home is None
            or not schema.get("aggregate_rendered_questions_only")
        ):
            concept_labels.setdefault(concept_key, []).append(label)
            group_labels.setdefault(group_key, []).append(label)
        placed.append((candidate, home))

    if schema.get("natural_label_aggregates"):
        for aggregates in (concept_labels, group_labels):
            for labels in aggregates.values():
                labels.sort(key=_natural_label_key)

    wb = _new_workbook(schema)
    group_provenance: list[dict] = []

    def _append_group_row(
        sheet: str, group_key: str, record: Mapping[str, Any],
    ) -> None:
        ws = wb[sheet]
        _append_record(
            ws, sheet, record, forced_blank, oversized=oversized,
            schema=schema,
        )
        group_provenance.append({
            "sheet": sheet,
            "row": ws.max_row,
            "group_key": group_key,
        })

    truncated: list[dict] = []
    # S9's cell-shape ledger. Every entry carries the FULL value, so a cell
    # the format cannot hold is repaired in the workbook and recorded whole
    # here — the ``truncated_rows`` pattern one level down, at the cell.
    oversized: list[dict] = []

    rollup_field = {
        "Basic": "basic_groups",
        "Intermediate": "intermediate_groups",
        "Advanced": "advanced_groups",
    }
    rollups_by_concept: dict[str, dict[str, list[str]]] = {}
    omitted_empty_group_shells: list[str] = []
    for group in renderable_groups:
        group_key = str(group.get("group_key") or "")
        if not group_labels.get(group_key):
            omitted_empty_group_shells.append(group_key)
            continue
        concept_key = str(group.get("concept_key") or "")
        field = rollup_field.get(str(group.get("group_type") or ""))
        if not field:
            continue
        rollups_by_concept.setdefault(concept_key, {}).setdefault(
            field, []
        ).append(str(group.get("group_name") or group_key))

    def _apply_rollups(record: dict, concept_key: str) -> None:
        values = rollups_by_concept.get(concept_key, {})
        for field in rollup_field.values():
            record[field] = ", ".join(values.get(field, []))

    def _full_record(candidate: Mapping, sheet: str) -> dict:
        concept_key = str(candidate.get("concept_key") or "")
        group_key = str(candidate.get("group_key") or "")
        entry = concept_entries[concept_key]
        record = _bands_record(entry)
        record["concept_question_labels"] = ", ".join(
            concept_labels.get(concept_key, []))
        _apply_rollups(record, concept_key)
        record.update(_group_record_fields(
            groups_by_key[group_key], group_labels.get(group_key, []),
            profile))
        record.update(
            _question_record(
                candidate, sheet, profile, truncated=truncated,
                descriptive_answer_slots=descriptive_answer_slots,
            )
        )
        return record

    question_rows = 0
    placed_by_concept: dict[str, int] = {}
    for candidate, home in placed:
        if home is not None:
            # A kind this step cannot render. Named at staging as
            # ``sheet_kind_not_renderable``, carrying the kind, the profile
            # name and the renderable set — never a school's name. The
            # verdict is ``candidate_home_finding``'s, not a second copy.
            _record_unplaced(candidate, home)
            continue
        sheet = sheets_by_kind[str(candidate.get("sheet_kind") or "")]
        group_key = str(candidate.get("group_key") or "")
        _append_group_row(
            sheet, group_key, _full_record(candidate, sheet))
        question_rows += 1
        concept_key = str(candidate.get("concept_key") or "")
        placed_by_concept[concept_key] = (
            placed_by_concept.get(concept_key, 0) + 1)

    # Empty BG/IG/AG shells remain in the immutable snapshot as the editing
    # surface, but they are not assessment rows. Emitting them produced the
    # audit's questionless group scaffolding. A genuinely questionless
    # concept still receives its one concept-only tail row below (OD5).

    # OWNER RULING OD5: "the Master should contain everything (including the
    # concepts that have no questions … so they should appear in the end
    # till concepts columns)". Exactly ONE row per questionless concept, no
    # Group-band update at all, appended LAST in the snapshot's own concept
    # order. The payload keeps its BG01/IG01/AG01 shells — they are group
    # structure the grouping lane and step 9's edit surface read — so the
    # payload and the workbook differ, and that difference is RECORDED here
    # rather than being silent (R4).
    questionless: list[dict] = []
    shells_by_concept: dict[str, list[str]] = {}
    for group in renderable_groups:
        shells_by_concept.setdefault(
            str(group.get("concept_key") or ""), []
        ).append(str(group.get("group_key") or ""))
    for concept_key in concept_order:
        if placed_by_concept.get(concept_key):
            continue
        entry = concept_entries[concept_key]
        record = _bands_record(entry)
        record["concept_question_labels"] = ""
        for field in rollup_field.values():
            record[field] = ""
        _append_record(
            wb["Objective"], "Objective", record, forced_blank,
            oversized=oversized, schema=schema,
        )
        questionless.append({
            "concept_key": concept_key,
            "concept_machine_id": str(
                entry["concept"].get("concept_machine_id") or ""),
            "concept_title": str(
                entry["concept"].get("concept_title") or ""),
            # The string the tail row's cell ACTUALLY reads (composed);
            # the bare title above names no cell once an id exists.
            "rendered_concept_title": _titled_concept(entry["concept"]),
            "shell_group_keys": shells_by_concept.get(concept_key, []),
        })

    tag_replacements = []
    seen_replacement_keys: set[tuple] = set()
    for entry in _concept_rows(snapshot):
        for kind, mapping in (("topic", entry["topic"]),
                              ("concept", entry["concept"])):
            record = _identity_tag_replacement(kind, mapping)
            if record is None:
                continue
            key = (record["kind"], record["title"],
                   record["applied_machine_id"])
            if key in seen_replacement_keys:
                continue
            seen_replacement_keys.add(key)
            tag_replacements.append(record)

    issues = {
        "unplaced": unplaced,
        "placed_questions": question_rows,
        # A baked tag that disagreed with the persisted id was REPLACED in
        # the cell; the loser is recorded here, never dropped silently (R4).
        "identity_tag_replacements": tag_replacements,
        "groups": len(groups),
        "group_provenance": group_provenance,
        "questionless_concepts": questionless,
        "omitted_empty_group_shells": omitted_empty_group_shells,
        "group_defects": group_defects,
        # What the layout had no slot for. The row SHIPS truncated (that is
        # what stops a raise costing all four outputs) and the loss is named
        # at staging as ``render_shape_overflow``, which refuses the
        # database write — but a reviewer reading only ``manifest.json`` saw
        # a complete-looking six-option row with no sign an option was
        # dropped. R4: the silence was the defect.
        "truncated_rows": truncated,
        # What no XLSX cell can hold, with the whole value beside it
        # (spec-step8 S9). Refused at staging as ``render_shape_overflow``.
        "oversized_cells": oversized,
    }
    return _workbook_bytes(wb), issues


# --------------------------------------------------------------------------- #
# Read-back parsing and validation (spec §13.2 steps 2–4)
# --------------------------------------------------------------------------- #

def parse_workbook(data: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                data_only=True)
    parsed: dict = {"sheet_order": list(wb.sheetnames), "sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        next(rows, ())  # band row
        header = [h for h in next(rows, ()) if h is not None]
        records = []
        row_numbers = []
        for row_number, row in enumerate(rows, start=3):
            if row is None or not any(row):
                continue
            records.append({
                field: ("" if i >= len(row) or row[i] is None else row[i])
                for i, field in enumerate(header)
            })
            row_numbers.append(row_number)
        parsed["sheets"][name] = {
            "fields": header,
            "rows": records,
            "row_numbers": row_numbers,
        }
    wb.close()
    return parsed


def _header_errors(
    parsed: Mapping, schema: Mapping[str, Any] | None = None,
) -> list[str]:
    active = schema or output_schema("concept")
    errors = []
    if parsed.get("sheet_order") != SHEET_ORDER:
        errors.append(
            f"sheet order {parsed.get('sheet_order')} != {SHEET_ORDER}")
    for name in SHEET_ORDER:
        sheet = (parsed.get("sheets") or {}).get(name)
        if sheet is None:
            errors.append(f"missing sheet {name!r}")
            continue
        if sheet["fields"] != active["fields"][name]:
            errors.append(f"{name}: header row differs from the template")
    return errors


def _readback_decimal(value: Any) -> Decimal | None:
    """Parse one exact finite workbook number without truthy-zero tricks."""

    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.startswith(("+", "-")):
            return None
    elif isinstance(value, (int, float, Decimal)):
        text = str(value)
    else:
        return None
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    if not number.is_finite():
        return None
    try:
        wire_number = float(number)
    except (OverflowError, ValueError):
        return None
    if not math.isfinite(wire_number):
        return None
    if number != 0 and wire_number == 0:
        return None
    if Decimal(str(wire_number)) != number:
        return None
    return number


def _populated(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _objective_marking_errors(
    row: Mapping[str, Any], *, label: str, marks: Decimal | None,
) -> list[str]:
    errors: list[str] = []
    correct_count = 0
    weights: list[Decimal] = []
    populated_options = 0
    populated_option_numbers: list[int] = []
    for n in range(1, MAX_OBJECTIVE_OPTIONS + 1):
        fields = (
            f"answer_type_{n}",
            f"answer_content_{n}",
            f"correct_answer_{n}",
            f"answer_weightage_{n}",
        )
        if not any(_populated(row.get(field)) for field in fields):
            continue
        populated_options += 1
        populated_option_numbers.append(n)
        answer_type = str(row.get(f"answer_type_{n}") or "")
        answer_content = str(row.get(f"answer_content_{n}") or "")
        if answer_type not in ANSWER_TYPES:
            errors.append(
                f"{label}: option {n} has unsupported answer_type "
                f"{answer_type!r}"
            )
        if not answer_content.strip():
            errors.append(f"{label}: option {n} has no answer content")
        if rel.option_content_has_label(answer_content):
            errors.append(
                f"{label}: option {n} repeats its letter label"
            )
        for issue in katex_rules.answer_cell_issues(
            answer_type, answer_content,
        ):
            errors.append(
                f"{label}: option {n} violates declared medium: {issue}"
            )
        is_correct = rel.is_correct_option(row.get(f"correct_answer_{n}"))
        if is_correct:
            correct_count += 1
        weight = _readback_decimal(row.get(f"answer_weightage_{n}"))
        if weight is None:
            errors.append(
                f"{label}: option {n} weight must be finite and numeric"
            )
            continue
        weights.append(weight)
        if is_correct:
            if weight <= 0:
                errors.append(
                    f"{label}: correct option {n} weight must be positive"
                )
            if marks is not None and weight != marks:
                errors.append(
                    f"{label}: correct weightage {weight} != marks {marks}"
                )
        elif weight != 0:
            errors.append(
                f"{label}: wrong option {n} weight must be exact zero"
            )
    if populated_options == 0:
        errors.append(f"{label}: objective has no option/rubric blocks")
    elif populated_option_numbers != list(range(1, populated_options + 1)):
        errors.append(
            f"{label}: objective option blocks must be contiguous from "
            "slot 1"
        )
    uppercase_labels = katex_rules.uppercase_objective_option_labels(
        str(row.get("question_text") or ""), MAX_OBJECTIVE_OPTIONS,
    )
    if uppercase_labels:
        errors.append(
            f"{label}: question_text uses uppercase objective option "
            f"label(s) {', '.join(uppercase_labels)}; use lowercase labels"
        )
    if correct_count != 1:
        errors.append(f"{label}: {correct_count} correct options")
    if (
        marks is not None
        and len(weights) == populated_options
        and sum(weights, Decimal(0)) != marks
    ):
        errors.append(
            f"{label}: option weights must sum exactly to marks {marks}"
        )
    return errors


def _subjective_marking_errors(
    row: Mapping[str, Any], *, label: str, marks: Decimal | None,
) -> list[str]:
    errors: list[str] = []
    weights: list[Decimal] = []
    populated = 0
    populated_numbers: list[int] = []
    question = str(row.get("question") or "")
    for n in range(1, MAX_SUBJECTIVE_ANSWERS + 1):
        fields = (
            f"answer_type_{n}", f"answer_{n}", f"answer_display_{n}",
            f"weightage_{n}", f"placeholder_{n}",
        )
        if not any(_populated(row.get(field)) for field in fields):
            continue
        populated += 1
        populated_numbers.append(n)
        answer_type = str(row.get(f"answer_type_{n}") or "")
        answer_content = str(row.get(f"answer_{n}") or "")
        if answer_type not in ANSWER_TYPES:
            errors.append(
                f"{label}: subjective answer {n} has unsupported "
                f"answer_type {answer_type!r}"
            )
        if not answer_content.strip():
            errors.append(f"{label}: subjective answer {n} has no content")
        for issue in katex_rules.answer_cell_issues(
            answer_type, answer_content,
        ):
            errors.append(
                f"{label}: subjective answer {n} violates declared medium: "
                f"{issue}"
            )
        answer_display = str(row.get(f"answer_display_{n}") or "")
        if not answer_display.strip():
            errors.append(
                f"{label}: subjective answer {n} has no answer_display"
            )
        for issue in katex_rules.rich_text_issues(answer_display):
            errors.append(
                f"{label}: subjective answer {n} answer_display rich-text: "
                f"{issue}"
            )
        expected_placeholder = chr(ord("a") + n - 1)
        placeholder = str(row.get(f"placeholder_{n}") or "")
        if placeholder != expected_placeholder:
            errors.append(
                f"{label}: subjective placeholder {n} {placeholder!r} != "
                f"{expected_placeholder!r}"
            )
        token = f"$${expected_placeholder}$$"
        if question.count(token) != 1:
            errors.append(
                f"{label}: subjective question must contain {token!r} "
                "exactly once"
            )
        weight = _readback_decimal(row.get(f"weightage_{n}"))
        if weight is None or weight <= 0:
            errors.append(
                f"{label}: subjective weight {n} must be finite and positive"
            )
        else:
            weights.append(weight)
    if populated == 0:
        errors.append(f"{label}: subjective has no answer blocks")
    elif populated_numbers != list(range(1, populated + 1)):
        errors.append(
            f"{label}: subjective answer blocks must be contiguous from "
            "slot 1"
        )
    if (
        marks is not None
        and len(weights) == populated
        and sum(weights, Decimal(0)) != marks
    ):
        errors.append(
            f"{label}: subjective weights must sum exactly to marks {marks}"
        )
    return errors


def _descriptive_marking_errors(
    row: Mapping[str, Any], *, label: str, marks: Decimal | None,
    answer_slots: int = MAX_DESCRIPTIVE_ANSWERS,
) -> list[str]:
    errors: list[str] = []
    answer_weights: list[Decimal] = []
    populated_answers = 0
    complete_answers = 0
    populated_answer_numbers: list[int] = []
    for n in range(1, answer_slots + 1):
        fields = (
            f"answer_type_{n}",
            f"answer_content_{n}",
            f"answer_weightage_{n}",
        )
        if not any(_populated(row.get(field)) for field in fields):
            continue
        populated_answers += 1
        populated_answer_numbers.append(n)
        answer_type = str(row.get(f"answer_type_{n}") or "")
        answer_content = str(row.get(f"answer_content_{n}") or "")
        if answer_type not in ANSWER_TYPES:
            errors.append(
                f"{label}: answer/rubric block {n} has unsupported "
                f"answer_type {answer_type!r}"
            )
        if not answer_content.strip():
            errors.append(
                f"{label}: answer/rubric block {n} has no content"
            )
        malformed_tag = rel.malformed_rubric_tag(answer_content, answer_type)
        if malformed_tag:
            errors.append(
                f"{label}: answer/rubric block {n} does not start with an "
                "allowed functional tag or is without its required colon"
            )
        medium_issues = katex_rules.answer_cell_issues(
            answer_type, answer_content,
        )
        for issue in medium_issues:
            errors.append(
                f"{label}: answer/rubric block {n} violates declared "
                f"medium: {issue}"
            )
        weight = _readback_decimal(row.get(f"answer_weightage_{n}"))
        if weight is None or weight <= 0:
            errors.append(
                f"{label}: answer/rubric weight {n} must be finite and positive"
            )
            continue
        answer_weights.append(weight)
        if (
            answer_type in ANSWER_TYPES
            and answer_content.strip()
            and not malformed_tag
            and not medium_issues
        ):
            complete_answers += 1
    sub_marks: list[Decimal] = []
    populated_subquestions = 0
    populated_subquestion_numbers: list[int] = []
    for n in range(1, MAX_SUBQUESTIONS + 1):
        keyword_fields = [
            field
            for m in range(1, MAX_SUBQUESTION_KEYWORDS + 1)
            for field in (
                f"sq{n}_answer_type_{m}",
                f"sq{n}_weightage_{m}",
                f"sq{n}_keyword_{m}",
            )
        ]
        fields = [f"sub_question_{n}", f"sub_question_marks_{n}"]
        if not any(
            _populated(row.get(field)) for field in [*fields, *keyword_fields]
        ):
            continue
        populated_subquestions += 1
        populated_subquestion_numbers.append(n)
        subquestion_text = str(row.get(f"sub_question_{n}") or "").strip()
        if not subquestion_text:
            errors.append(f"{label}: subquestion {n} has no text")
        elif any(
            subquestion_text in str(row.get(field) or "")
            for field in ("question", "question_text")
        ):
            errors.append(
                f"{label}: subquestion {n} text is duplicated in the main "
                "question/question_text"
            )
        sub_mark = _readback_decimal(row.get(f"sub_question_marks_{n}"))
        if sub_mark is None or sub_mark <= 0:
            errors.append(
                f"{label}: subquestion {n} marks must be finite and positive"
            )
        else:
            sub_marks.append(sub_mark)

        keyword_weights: list[Decimal] = []
        populated_keywords = 0
        populated_keyword_numbers: list[int] = []
        for m in range(1, MAX_SUBQUESTION_KEYWORDS + 1):
            fields = (
                f"sq{n}_answer_type_{m}",
                f"sq{n}_weightage_{m}",
                f"sq{n}_keyword_{m}",
            )
            if not any(_populated(row.get(field)) for field in fields):
                continue
            populated_keywords += 1
            populated_keyword_numbers.append(m)
            keyword_type = str(row.get(f"sq{n}_answer_type_{m}") or "")
            keyword_content = str(row.get(f"sq{n}_keyword_{m}") or "")
            if keyword_type not in ANSWER_TYPES:
                errors.append(
                    f"{label}: subquestion {n} keyword {m} has unsupported "
                    f"answer_type {keyword_type!r}"
                )
            if not keyword_content.strip():
                errors.append(
                    f"{label}: subquestion {n} keyword {m} has no text"
                )
            for issue in katex_rules.answer_cell_issues(
                keyword_type, keyword_content,
            ):
                errors.append(
                    f"{label}: subquestion {n} keyword {m} violates "
                    f"declared medium: {issue}"
                )
            if rel.malformed_rubric_tag(
                keyword_content, keyword_type,
            ):
                errors.append(
                    f"{label}: subquestion {n} keyword {m} does not start "
                    "with an allowed functional tag or is without its "
                    "required colon"
                )
            weight = _readback_decimal(row.get(f"sq{n}_weightage_{m}"))
            if weight is None or weight <= 0:
                errors.append(
                    f"{label}: subquestion {n} keyword {m} weight must be "
                    "finite and positive"
                )
                continue
            keyword_weights.append(weight)
        if (
            populated_keywords
            and sub_mark is not None
            and sub_mark > 0
            and len(keyword_weights) == populated_keywords
            and sum(keyword_weights, Decimal(0)) != sub_mark
        ):
            errors.append(
                f"{label}: subquestion {n} keyword weights must sum exactly "
                "to its marks"
            )
        if populated_keywords == 0:
            errors.append(
                f"{label}: subquestion {n} needs at least one keyword rubric"
            )
        elif populated_keyword_numbers != list(
            range(1, populated_keywords + 1)
        ):
            errors.append(
                f"{label}: subquestion {n} keyword blocks must be "
                "contiguous from slot 1"
            )
    if populated_answers == 0 and populated_subquestions == 0:
        errors.append(f"{label}: descriptive has no answer/rubric blocks")
    if populated_answers and populated_subquestions:
        errors.append(
            f"{label}: multipart descriptive duplicates scoring in main "
            "answer/rubric blocks"
        )
    if populated_answers and populated_answer_numbers != list(
        range(1, populated_answers + 1)
    ):
        errors.append(
            f"{label}: Descriptive main rubric blocks must be contiguous "
            "from slot 1"
        )
    if populated_subquestions and populated_subquestion_numbers != list(
        range(1, populated_subquestions + 1)
    ):
        errors.append(
            f"{label}: Descriptive subquestion blocks must be contiguous "
            "from slot 1"
        )
    if (
        marks == Decimal(4)
        and populated_subquestions == 0
        and complete_answers < 2
    ):
        errors.append(
            f"{label}: 4-mark single-part descriptive requires at least two "
            "answer/rubric blocks"
        )
    if (
        populated_subquestions == 0
        and marks is not None
        and len(answer_weights) == populated_answers
        and sum(answer_weights, Decimal(0)) != marks
    ):
        errors.append(
            f"{label}: answer/rubric weights must sum exactly to marks {marks}"
        )
    if (
        populated_subquestions
        and marks is not None
        and len(sub_marks) == populated_subquestions
        and sum(sub_marks, Decimal(0)) != marks
    ):
        errors.append(
            f"{label}: subquestion marks must sum exactly to marks {marks}"
        )
    return errors


def validate_concept_file(
    parsed: Mapping, snapshot: Mapping,
    profile: Mapping | str | None = None,
) -> list[str]:
    errors = _header_errors(parsed, output_schema("concept", profile, snapshot))
    if errors:
        return errors
    rows = parsed["sheets"]["Objective"]["rows"]
    expected = [
        _titled_concept(e["concept"]) for e in _concept_rows(snapshot)
    ]
    actual = [str(r.get("concept_title") or "") for r in rows]
    if actual != expected:
        errors.append(
            f"concept rows differ: expected {len(expected)}, "
            f"got {len(actual)} (or out of source order)")
    # The topic cell's read-back twin (B4 repair round: the composed topic
    # cell shipped with no verification at all while the concept cell had
    # three).
    expected_topics = [
        _titled_topic(e["topic"], int(e.get("topic_ordinal") or 1))
        for e in _concept_rows(snapshot)
    ]
    actual_topics = [str(r.get("topic_title") or "") for r in rows]
    if actual_topics != expected_topics:
        errors.append("topic title cells differ from the snapshot's topics")
    # The must-keep-blank list reads the SAME profile key the renderer's
    # ``_row_values`` reads (spec-step8 T12/M4), so the read-back gate and
    # the renderer can never disagree about what ships blank.
    must_be_blank = (
        "group_name", "group_type", "question_label", "question",
        "basic_groups", "intermediate_groups", "advanced_groups",
        "concept_question_labels",
        *assessment_profile.forced_blank_fields(profile),
    )
    for i, row in enumerate(rows, start=3):
        populated = [
            field for field in must_be_blank
            if str(row.get(field) or "").strip()
        ]
        if populated:
            errors.append(
                f"Objective row {i}: Concept File must keep {populated} "
                "blank")
    for name in ("Descriptive", "Subjective"):
        if parsed["sheets"][name]["rows"]:
            errors.append(f"{name}: Concept File must be header-only")
    return errors


def validate_master_file(
    parsed: Mapping, snapshot: Mapping, profile: Mapping | str | None = None,
    *, group_provenance: list[Mapping[str, Any]] | None = None,
) -> list[str]:
    profile = assessment_profile.resolve(profile)
    forced_blank = assessment_profile.forced_blank_fields(profile)
    schema = output_schema("master", profile, snapshot)
    descriptive_answer_slots = int(schema["descriptive_answer_slots"])
    errors = _header_errors(parsed, schema)
    if errors:
        return errors
    if "subjective" not in assessment_profile.sheet_kinds(profile) and (
        parsed["sheets"]["Subjective"]["rows"]
    ):
        errors.append(
            "Subjective: profile "
            f"{assessment_profile.name(profile)!r} allows no data rows")

    groups = [dict(g) for g in snapshot.get("groups") or []]
    groups_by_key: dict[str, dict] = {}
    for group in groups:
        group_key = str(group.get("group_key") or "")
        if not group_key:
            errors.append("snapshot contains a blank group_key")
        elif group_key in groups_by_key:
            errors.append(f"snapshot contains duplicate group_key {group_key!r}")
        else:
            groups_by_key[group_key] = group
    keys_by_visible_name: dict[str, list[str]] = {}
    for group_key, group in groups_by_key.items():
        group_name = str(group.get("group_name") or "")
        keys_by_visible_name.setdefault(group_name, []).append(group_key)

    provenance_by_row: dict[tuple[str, int], str] = {}
    if group_provenance is not None:
        for n, item in enumerate(group_provenance, start=1):
            sheet = str(item.get("sheet") or "")
            try:
                row_number = int(item.get("row"))
            except (TypeError, ValueError):
                errors.append(
                    f"group provenance entry {n}: row must be an integer")
                continue
            group_key = str(item.get("group_key") or "")
            coordinate = (sheet, row_number)
            if sheet not in {
                "Objective", "Subjective", "Descriptive",
            } or row_number < 3:
                errors.append(
                    f"group provenance entry {n}: invalid coordinate "
                    f"{coordinate!r}")
                continue
            if coordinate in provenance_by_row:
                errors.append(
                    f"group provenance repeats {sheet} row {row_number}")
                continue
            if group_key not in groups_by_key:
                errors.append(
                    f"group provenance {sheet} row {row_number}: unknown "
                    f"group_key {group_key!r}")
                continue
            provenance_by_row[coordinate] = group_key

    concept_entries = {
        str(entry["concept"].get("concept_key") or ""): entry
        for entry in _concept_rows(snapshot)
    }
    concept_keys = set(concept_entries)
    expected_group_labels: dict[str, list[str]] = {}
    expected_concept_labels: dict[str, list[str]] = {}
    expected_question_labels: dict[str, list[str]] = {
        name: [] for name in ("Objective", "Subjective", "Descriptive")
    }
    # The same count the renderer used: how many question rows each concept
    # actually places. A concept with none gets ONE tail row stopping at the
    # concept columns (OD5/T16) and no Group row at all, so demanding a Group
    # row for it would turn every questionless concept into an error.
    placed_by_concept: dict[str, int] = {}
    renderable_sheets = sheet_for_kind()
    profile_name = assessment_profile.name(profile)
    for candidate in snapshot.get("candidates") or []:
        concept_key = str(candidate.get("concept_key") or "")
        group_key = str(candidate.get("group_key") or "")
        label = str(candidate.get("question_label") or "")
        home = rel.candidate_home_finding(
            candidate, concept_entries, groups_by_key,
            tuple(renderable_sheets), profile_name,
        )
        if (
            concept_key in concept_keys
            and group_key in groups_by_key
            and label
        ):
            group_concept_key = str(
                groups_by_key[group_key].get("concept_key") or "")
            if group_concept_key != concept_key:
                errors.append(
                    f"{label}: candidate concept_key {concept_key!r} does "
                    f"not match group {group_key!r} home "
                    f"{group_concept_key!r}")
                continue
            if (
                home is None
                or not schema.get("aggregate_rendered_questions_only")
            ):
                expected_group_labels.setdefault(group_key, []).append(label)
                expected_concept_labels.setdefault(
                    concept_key, [],
                ).append(label)
            if home is None:
                placed_by_concept[concept_key] = (
                    placed_by_concept.get(concept_key, 0) + 1)
        if home is None:
            sheet_name = renderable_sheets[
                str(candidate.get("sheet_kind") or "")
            ]
            expected_question_labels[sheet_name].append(label)
    if schema.get("natural_label_aggregates"):
        for aggregates in (expected_group_labels, expected_concept_labels):
            for labels in aggregates.values():
                labels.sort(key=_natural_label_key)
    expected_group_keys = {
        group_key for group_key, labels in expected_group_labels.items()
        if labels
    }
    rollup_field = {
        "Basic": "basic_groups",
        "Intermediate": "intermediate_groups",
        "Advanced": "advanced_groups",
    }
    expected_rollups: dict[str, dict[str, list[str]]] = {}
    for group_key, group in groups_by_key.items():
        if group_key not in expected_group_keys:
            continue
        concept_key = str(group.get("concept_key") or "")
        field = rollup_field.get(str(group.get("group_type") or ""))
        if not field:
            continue
        expected_rollups.setdefault(concept_key, {}).setdefault(
            field, []
        ).append(str(group.get("group_name") or group_key))
    expected_questionless_concepts = [
        _titled_concept(entry["concept"])
        for concept_key, entry in concept_entries.items()
        if not placed_by_concept.get(concept_key)
    ]

    seen_group_keys: set[str] = set()
    seen_concepts: set[str] = set()
    actual_questionless_concepts: list[str] = []
    actual_question_labels: dict[str, list[str]] = {
        name: [] for name in ("Objective", "Subjective", "Descriptive")
    }
    aggregate_by_group: dict[str, set[str]] = {}
    visited_provenance: set[tuple[str, int]] = set()
    # The topic cell's read-back twin (B4 repair round): keyed off the
    # COMPOSED concept title, which embeds the machine id, so two
    # same-titled concepts under two topics map to their own topics.
    #
    # Round 7: a LEGACY snapshot (``snapshot_from_chapter``, no machine
    # ids) composes the bare title, so two same-titled concepts collapse
    # onto one key. Last-wins there turned this twin into a false gate —
    # [measured] it blocked a valid legacy Master with "topic_title does
    # not match" over rows nobody mis-filed, permanently, the moment a
    # chapter legitimately held two same-titled concepts (a state S10's
    # publication supports). A key two rows share with two different
    # topics proves nothing about either row, so the twin DECLINES to
    # judge those (``None`` skips the check below) instead of judging by
    # collapse order.
    expected_topic_by_concept: dict[str, str | None] = {}
    expected_source_by_path: dict[tuple[str, str], str | None] = {}
    for e in _concept_rows(snapshot):
        composed_key = _titled_concept(e["concept"])
        expected_value = _titled_topic(
            e["topic"], int(e.get("topic_ordinal") or 1),
        )
        if composed_key in expected_topic_by_concept:
            if expected_topic_by_concept[composed_key] != expected_value:
                expected_topic_by_concept[composed_key] = None
        else:
            expected_topic_by_concept[composed_key] = expected_value
        source_key = (expected_value, composed_key)
        source_value = str(e["concept"].get("concept_source") or "")
        if source_key in expected_source_by_path:
            if expected_source_by_path[source_key] != source_value:
                # A duplicated snapshot path that disagrees about its source
                # is not something read-back may silently resolve by order.
                expected_source_by_path[source_key] = None
        else:
            expected_source_by_path[source_key] = source_value
    for name in ("Objective", "Subjective", "Descriptive"):
        sheet_rows = parsed["sheets"][name]["rows"]
        row_numbers = parsed["sheets"][name].get("row_numbers")
        if row_numbers is None:
            row_numbers = list(range(3, 3 + len(sheet_rows)))
        elif len(row_numbers) != len(sheet_rows):
            errors.append(
                f"{name}: read-back row number ledger differs from rows")
            row_numbers = list(range(3, 3 + len(sheet_rows)))
        for i, row in zip(row_numbers, sheet_rows):
            row_concept_title = str(row.get("concept_title") or "")
            seen_concepts.add(row_concept_title)
            sheet_fields = schema["fields"][name]
            for update_field, identity_fields in _UPDATE_FIELD_PRESENCE.items():
                if update_field not in sheet_fields:
                    continue
                entity_populated = any(
                    _populated(row.get(field)) for field in identity_fields
                )
                expected_update = "No" if entity_populated else ""
                raw_update = row.get(update_field)
                actual_update = (
                    "" if raw_update is None else str(raw_update)
                )
                if actual_update != expected_update:
                    errors.append(
                        f"{name} row {i}: {update_field} "
                        f"{actual_update!r} != {expected_update!r} for "
                        f"the {'populated' if entity_populated else 'blank'} "
                        "entity band"
                    )

            if "concept_source" in sheet_fields:
                raw_source = row.get("concept_source")
                actual_source = (
                    "" if raw_source is None else str(raw_source)
                )
                if not row_concept_title:
                    expected_source = ""
                else:
                    source_key = (
                        str(row.get("topic_title") or ""),
                        row_concept_title,
                    )
                    if source_key not in expected_source_by_path:
                        expected_source = None
                        errors.append(
                            f"{name} row {i}: concept_source cannot be "
                            "matched to a snapshot concept"
                        )
                    else:
                        expected_source = expected_source_by_path[source_key]
                        if expected_source is None:
                            errors.append(
                                f"{name} row {i}: snapshot concept_source "
                                "is ambiguous for this topic/concept path"
                            )
                if (
                    expected_source is not None
                    and actual_source != expected_source
                ):
                    errors.append(
                        f"{name} row {i}: concept_source "
                        f"{actual_source!r} != snapshot value "
                        f"{expected_source!r}"
                    )
            question_fields = sheet_fields[
                sheet_fields.index("question_label"):
            ]
            has_question_band = any(
                row.get(field) is not None
                and str(row.get(field)).strip()
                for field in question_fields
            )
            group_fields = (
                "group_name", "group_display_name", "group_description",
                "group_status", "group_type", "group_question_labels",
                "related_digicards",
            )
            has_group_band = any(
                row.get(field) is not None
                and str(row.get(field)).strip()
                for field in group_fields
            )
            expected_topic = expected_topic_by_concept.get(row_concept_title)
            if expected_topic is not None and (
                str(row.get("topic_title") or "") != expected_topic
            ):
                errors.append(
                    f"{name} row {i}: topic_title does not match the "
                    "snapshot's topic for this concept")
            group_name = str(row.get("group_name") or "")
            coordinate = (name, i)
            group_key = provenance_by_row.get(coordinate, "")
            if group_key:
                visited_provenance.add(coordinate)
            elif group_name and group_provenance is not None:
                errors.append(
                    f"{name} row {i}: group provenance is missing")
            elif group_name:
                matching_keys = keys_by_visible_name.get(group_name, [])
                if len(matching_keys) == 1:
                    group_key = matching_keys[0]
                elif len(matching_keys) > 1:
                    errors.append(
                        f"{name} row {i}: group provenance required for "
                        f"non-unique visible group name {group_name!r}")
                else:
                    errors.append(
                        f"{name} row {i}: unknown group name {group_name!r}")
            if group_key:
                seen_group_keys.add(group_key)
                expected_group = groups_by_key[group_key]
                expected_name = str(expected_group.get("group_name") or "")
                expected_display = str(
                    expected_group.get("group_display_name") or "")
                expected_type = str(expected_group.get("group_type") or "")
                group_concept_key = str(
                    expected_group.get("concept_key") or "")
                concept_entry = concept_entries.get(group_concept_key)
                if group_name != expected_name:
                    errors.append(
                        f"{name} row {i}: group_name {group_name!r} does "
                        f"not match {group_key!r}")
                if str(row.get("group_display_name") or "") != expected_display:
                    errors.append(
                        f"{name} row {i}: group_display_name does not "
                        f"match {group_key!r}")
                if str(row.get("group_type") or "") != expected_type:
                    errors.append(
                        f"{name} row {i}: group_type does not match "
                        f"{group_key!r}")
                if concept_entry is None:
                    errors.append(
                        f"{name} row {i}: {group_key!r} has unknown "
                        f"concept home {group_concept_key!r}")
                else:
                    expected_concept_title = _titled_concept(
                        concept_entry["concept"])
                    if str(row.get("concept_title") or "") != (
                        expected_concept_title
                    ):
                        errors.append(
                            f"{name} row {i}: concept home does not match "
                            f"{group_key!r}")
                    expected_concept_aggregate = ", ".join(
                        expected_concept_labels.get(group_concept_key, [])
                    )
                    if str(row.get("concept_question_labels") or "") != (
                        expected_concept_aggregate
                    ):
                        errors.append(
                            f"{name} row {i}: concept_question_labels "
                            "does not match surviving concept members"
                        )
                    expected_concept_rollups = expected_rollups.get(
                        group_concept_key, {}
                    )
                    for field in rollup_field.values():
                        expected_value = ", ".join(
                            expected_concept_rollups.get(field, [])
                        )
                        actual_value = str(row.get(field) or "")
                        if actual_value != expected_value:
                            errors.append(
                                f"{name} row {i}: {field} "
                                f"{actual_value!r} != {expected_value!r}"
                            )
                aggregate_by_group.setdefault(group_key, set()).add(
                    str(row.get("group_question_labels") or ""))
            # The fourth C10 site (spec-step8): the read-back blanks by the
            # SAME profile key the renderer blanked by, so a profile that
            # un-blanks a field cannot fail its own read-back on the value
            # it just rendered.
            for field in forced_blank:
                if str(row.get(field) or "").strip():
                    errors.append(f"{name} row {i}: {field} must be blank")
            if not has_question_band:
                if has_group_band:
                    errors.append(
                        f"{name} row {i}: a row without a Question band "
                        "must also keep the Group band blank"
                    )
                if name != "Objective":
                    errors.append(
                        f"{name} row {i}: concept-only tail rows are allowed "
                        "only on Objective"
                    )
                else:
                    actual_questionless_concepts.append(row_concept_title)
                continue
            raw_label = str(row.get("question_label") or "").strip()
            actual_question_labels[name].append(raw_label)
            if not raw_label:
                errors.append(
                    f"{name} row {i}: populated Question band requires a "
                    "non-blank question_label"
                )
            label = raw_label or f"{name} row {i}"
            if not str(row.get("question") or "").strip():
                errors.append(f"{label}: question must not be blank")
            appears = str(row.get("question_appears_in") or "")
            wire = assessment_profile.appears_in(profile)
            if appears != wire:
                errors.append(
                    f"{label}: question_appears_in {appears!r} is not the "
                    f"profile wire value {wire!r}")
            restriction = str(row.get("answer_restriction") or "")
            if restriction not in rel.ANSWER_RESTRICTIONS:
                errors.append(
                    f"{label}: answer_restriction {restriction!r} invalid")
            duration = _readback_decimal(row.get("question_duration"))
            if duration is None or duration <= 0:
                errors.append(
                    f"{label}: question_duration must be finite and positive"
                )
            if name in {"Subjective", "Descriptive"} and row.get(
                "math_keyboard"
            ) not in {
                "Yes", "No",
            }:
                errors.append(
                    f"{label}: {name.lower()} math_keyboard must be exactly "
                    "Yes or No"
                )
            marks = _readback_decimal(row.get("marks"))
            if marks is None or marks <= 0:
                errors.append(f"{label}: marks must be finite and positive")
            for field in (
                "question", "question_text", "display_answer",
                "answer_explanation",
            ):
                rich_value = str(row.get(field) or "")
                if name == "Subjective" and field in {
                    "question", "question_text",
                }:
                    for n in range(1, MAX_SUBJECTIVE_ANSWERS + 1):
                        placeholder = str(
                            row.get(f"placeholder_{n}") or ""
                        )
                        if len(placeholder) == 1 and "a" <= placeholder <= "t":
                            rich_value = rich_value.replace(
                                f"$${placeholder}$$", ""
                            )
                for issue in katex_rules.rich_text_issues(rich_value):
                    # Full rich-text validation happens before freeze.  The
                    # workbook renderer may add mechanical truncation notes
                    # whose field-like underscores are not authored LaTeX;
                    # read-back owns only Q21's table-dialect refusal here.
                    if issue == "unsupported_table":
                        errors.append(
                            f"{label}: {field} rich-text: {issue}"
                        )
            for n in range(1, MAX_SUBQUESTIONS + 1):
                for issue in katex_rules.rich_text_issues(
                    str(row.get(f"sub_question_{n}") or "")
                ):
                    if issue == "unsupported_table":
                        errors.append(
                            f"{label}: sub_question_{n} rich-text: {issue}"
                        )
            if name == "Objective":
                errors.extend(_objective_marking_errors(
                    row, label=label, marks=marks,
                ))
            elif name == "Subjective":
                errors.extend(_subjective_marking_errors(
                    row, label=label, marks=marks,
                ))
            else:
                errors.extend(_descriptive_marking_errors(
                    row, label=label, marks=marks,
                    answer_slots=descriptive_answer_slots,
                ))

    # Every concept (questionless included) and every created group appears.
    # Keyed on the COMPOSED title, which embeds the machine id: two
    # same-titled concepts under two topics are two distinct expected
    # members, so dropping one of two same-titled tail rows is visible
    # (spec-step8 B4; before this the set collapsed them).
    expected_concepts = {
        _titled_concept(e["concept"]) for e in _concept_rows(snapshot)
    }
    for concept_title in sorted(expected_concepts - seen_concepts):
        errors.append(f"concept missing from Master: {concept_title!r}")
    for group_key in sorted(expected_group_keys - seen_group_keys):
        errors.append(f"group missing from Master: {group_key!r}")
    for sheet, row_number in sorted(
        set(provenance_by_row) - visited_provenance
    ):
        errors.append(
            f"group provenance {sheet} row {row_number}: no read-back row")
    for group_key, aggregates in aggregate_by_group.items():
        if len(aggregates) > 1:
            errors.append(
                f"{group_key}: group_question_labels differs across "
                "repeated rows")
            continue
        actual_aggregate = next(iter(aggregates), "")
        expected_aggregate = ", ".join(
            expected_group_labels.get(group_key, []))
        if actual_aggregate != expected_aggregate:
            errors.append(
                f"{group_key}: group_question_labels {actual_aggregate!r} "
                f"!= {expected_aggregate!r}")
    for name in ("Objective", "Subjective", "Descriptive"):
        expected_labels = expected_question_labels[name]
        actual_labels = actual_question_labels[name]
        if actual_labels != expected_labels:
            errors.append(
                f"{name}: question label row set/order differs from the "
                f"snapshot (expected {len(expected_labels)} labels, got "
                f"{len(actual_labels)}; expected {expected_labels!r}, got "
                f"{actual_labels!r})"
            )
    if actual_questionless_concepts != expected_questionless_concepts:
        errors.append(
            "Objective: concept-only tail row set/order differs from the "
            f"snapshot (expected {expected_questionless_concepts!r}, got "
            f"{actual_questionless_concepts!r})"
        )
    return errors


# --------------------------------------------------------------------------- #
# Dual projection (spec §13.1–§13.2)
# --------------------------------------------------------------------------- #

def build_dual_output(
    snapshot: Mapping, profile: Mapping | str | None = None,
) -> dict:
    """Render, read back, validate, and hash both projections of one
    snapshot. Returns concepts/master bytes plus the shared manifest."""
    profile = assessment_profile.resolve(profile)
    snapshot_hash = snapshot_sha256(snapshot)
    concept_schema = output_schema("concept", profile, snapshot)
    master_schema = output_schema("master", profile, snapshot)
    concept_identity = _schema_identity(concept_schema)
    master_identity = _schema_identity(master_schema)
    concepts_bytes = render_concept_file(snapshot, profile)
    master_bytes, issues = render_master_file(snapshot, profile)
    concept_errors = validate_concept_file(
        parse_workbook(concepts_bytes), snapshot, profile)
    master_errors = validate_master_file(
        parse_workbook(master_bytes), snapshot, profile,
        group_provenance=issues["group_provenance"])
    manifest = {
        "profile": assessment_profile.name(profile),
        "template_source": MANIFEST.get("source", ""),
        "workbook_contracts": {
            "concepts_xlsx": {
                **concept_identity,
                "field_counts": {
                    sheet: len(concept_schema["fields"][sheet])
                    for sheet in SHEET_ORDER
                },
            },
            "master_xlsx": {
                **master_identity,
                "field_counts": {
                    sheet: len(master_schema["fields"][sheet])
                    for sheet in SHEET_ORDER
                },
                "descriptive_answer_slots": master_schema[
                    "descriptive_answer_slots"
                ],
            },
        },
        "concept_snapshot_sha256": snapshot_hash,
        "workbook_sha256s": {
            "concepts_xlsx": hashlib.sha256(concepts_bytes).hexdigest(),
            "master_xlsx": hashlib.sha256(master_bytes).hexdigest(),
        },
        "read_back": {
            "concepts_errors": concept_errors,
            "master_errors": master_errors,
        },
        "issues": issues,
    }
    return {
        "concepts_xlsx": concepts_bytes,
        "master_xlsx": master_bytes,
        "manifest": manifest,
        "valid": not concept_errors and not master_errors,
    }
