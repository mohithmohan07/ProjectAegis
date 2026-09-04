"""Registry of the Bulk Import workbook layouts this tool can read.

A workbook is *identified* by its header row (row 2) and its sheet name, by
exact equality against a registered layout. Nothing is guessed: a header that
matches no registered layout is refused, and every column is then addressed
**by name** through the identified layout instead of by a position derived
from today's canonical constants.

Why this module exists (spec-step8 T6, slice S2). ``reader.import_workbook``
used to address columns positionally against ``bulk_import.FIELDS_BY_KIND``
and to skip any sheet whose *name* it did not recognise, with no issue
recorded. On the accepted reference workbooks that silently dropped the whole
Objective sheet (``SHEET_OBJECTIVE`` is ``"Objective "`` — with a trailing
space — and the reference sheet is ``"Objective"``) and mis-banded the other
two (342 of 344 Descriptive and 63 of 63 Subjective question-band positions
read the wrong column). Identification + name addressing is pure mechanics: it
makes no judgment about what any cell MEANS, it only declines to guess which
column a value came from.

Two kinds of entry live here, and the difference is load-bearing:

* ``sop-mes-1`` is read AT IMPORT TIME from the committed format workbook
  ``backend/data/Testing/reference_bulk_import/bulk_import_format.xlsx`` — the
  owner-supplied layout authority (T3.1b). ``assessment_workbook_template.json``
  is compared against it and is a derived convenience cache: **on any
  disagreement the workbook wins** and a ``layout_manifest_drift`` defect is
  recorded (never a halt — Q13).
* the three ``canonical-*`` entries are **FROZEN LITERAL transcriptions** of
  the layouts this repo has historically written. They are deliberately NOT
  derived from ``bulk_import.FIELDS_BY_KIND``: slice S7 redefines those
  constants to the reference layout, and an entry derived from them would
  silently BECOME the reference layout, un-registering the canonical one and
  turning every older workbook into a 422. S7 must not touch the literals in
  this file.
"""
from __future__ import annotations

import hashlib
import json
import logging
import re
import shutil
from dataclasses import dataclass, field as dataclass_field
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import openpyxl
from openpyxl.styles import Alignment, Font

# --------------------------------------------------------------------------- #
# Errors and recorded defects
# --------------------------------------------------------------------------- #


class WorkbookLayoutError(ValueError):
    """A workbook's header row matches no registered layout.

    Mechanical gate on an artifact (CLAUDE.md "gates that refuse to accept a
    broken artifact"): it judges nothing about content, it declines to read
    columns whose meaning it cannot establish. ``import_workbook`` raises it
    before any ``db.add``, so an unidentified workbook imports nothing at all
    rather than importing part of itself wrongly.
    """

    def __init__(self, message: str, *, detail: Mapping | None = None) -> None:
        super().__init__(message)
        self.detail = dict(detail or {})


_REGISTRY_DEFECTS: list[dict] = []


def registry_defects() -> list[dict]:
    """Defects recorded while BUILDING the registry — recorded, not decided.

    Building the registry never halts on its own: it records what it found and
    returns whatever it could register. Today that is ``layout_manifest_drift``
    (the committed format workbook and the template JSON disagree; the workbook
    won, and the drift is a recorded defect) and ``layout_source_missing`` (the
    committed workbook is not on disk, so the reference layout is not
    registered in this process).

    What the CALLER does with a defect is the caller's contract, and the two
    differ deliberately. ``layout_manifest_drift`` has an answer — the workbook
    — so it stays a record. ``layout_source_missing`` leaves this package with
    no column geometry at all, so ``bulk_import/__init__.py`` turns it into an
    ``ImportError`` at import time: a pre-spend gate on a broken artifact, and
    the opposite of the mid-run halt Q13 retired, which is a refusal to finish
    work already paid for.
    """
    return [dict(defect) for defect in _REGISTRY_DEFECTS]


# --------------------------------------------------------------------------- #
# Layout objects
# --------------------------------------------------------------------------- #

_ANSWER_BLOCK_RE = re.compile(r"^answer_type_(\d+)$")
_SUBQUESTION_RE = re.compile(r"^sub_question_(\d+)$")

@dataclass(frozen=True)
class Band:
    """One row-1 section band, 1-based inclusive column range."""

    label: str
    start: int
    end: int

    def as_dict(self) -> dict:
        return {"label": self.label, "start": self.start, "end": self.end}


class SheetLayout:
    """One sheet of one layout: its exact field names and their addresses."""

    __slots__ = ("layout_id", "kind", "sheet_name", "fields", "bands",
                 "blocks", "_index")

    def __init__(self, layout_id: str, kind: str, sheet_name: str,
                 blocks: Sequence[tuple[str, Sequence[str]]],
                 bands: Sequence[Band]) -> None:
        self.layout_id = layout_id
        self.kind = kind
        self.sheet_name = sheet_name
        self.blocks: tuple[tuple[str, tuple[str, ...]], ...] = tuple(
            (name, tuple(names)) for name, names in blocks
        )
        self.fields: tuple[str, ...] = tuple(
            name for _block, names in self.blocks for name in names
        )
        self.bands: tuple[Band, ...] = tuple(bands)
        index: dict[tuple[str, str], int] = {}
        position = 0
        for block, names in self.blocks:
            for name in names:
                # Band-qualified so a layout that repeats a field name across
                # bands (canonical Descriptive carries ``question_label`` in
                # both the Group and the Question band) resolves explicitly
                # instead of by scanning.
                index.setdefault((block, name), position)
                position += 1
        self._index = index

    # -- addressing ------------------------------------------------------- #

    def block_fields(self, block: str) -> tuple[str, ...]:
        for name, names in self.blocks:
            if name == block:
                return names
        return ()

    def block_start(self, block: str) -> int:
        position = 0
        for name, names in self.blocks:
            if name == block:
                return position
            position += len(names)
        return position

    def column(self, block: str, field: str) -> int | None:
        """0-based column index of one band-qualified field, or None."""
        return self._index.get((block, field))

    def block_values(self, row: Sequence, block: str) -> dict[str, str]:
        """Every field of one band, read BY NAME, stripped."""
        out: dict[str, str] = {}
        for field in self.block_fields(block):
            index = self._index[(block, field)]
            value = row[index] if index < len(row) else None
            out.setdefault(
                field, "" if value is None else str(value).strip())
        return out

    # -- derived, mechanical counts --------------------------------------- #

    @property
    def question_scalar_fields(self) -> tuple[str, ...]:
        """Question-band fields before the first answer block.

        Derived from the layout's own field list, never hard-coded: the
        canonical sheets yield 10/11/12 and the reference sheets 12/13/14.
        """
        out: list[str] = []
        for field in self.block_fields("question"):
            if _ANSWER_BLOCK_RE.match(field):
                break
            out.append(field)
        return tuple(out)

    @property
    def answer_block_numbers(self) -> tuple[int, ...]:
        return tuple(sorted(
            int(match.group(1))
            for match in (
                _ANSWER_BLOCK_RE.match(field)
                for field in self.block_fields("question")
            )
            if match
        ))

    @property
    def sub_question_numbers(self) -> tuple[int, ...]:
        return tuple(sorted(
            int(match.group(1))
            for match in (
                _SUBQUESTION_RE.match(field)
                for field in self.block_fields("question")
            )
            if match
        ))

    def sub_question_keyword_numbers(self, number: int) -> tuple[int, ...]:
        prefix = re.compile(rf"^sq{number}_answer_type_(\d+)$")
        return tuple(sorted(
            int(match.group(1))
            for match in (
                prefix.match(field) for field in self.block_fields("question")
            )
            if match
        ))


class Layout:
    """One registered workbook layout: its sheets and its ignored sheets."""

    __slots__ = ("id", "sheets", "ignored_sheets", "source")

    def __init__(self, layout_id: str, sheets: Mapping[str, SheetLayout],
                 ignored_sheets: Sequence[str] = (), source: str = "") -> None:
        self.id = layout_id
        self.sheets = dict(sheets)
        self.ignored_sheets = tuple(ignored_sheets)
        self.source = source

    def sheet(self, kind: str) -> SheetLayout:
        return self.sheets[kind]

    def sheet_name_by_kind(self) -> dict[str, str]:
        return {kind: sheet.sheet_name for kind, sheet in self.sheets.items()}

    def fields_by_kind(self) -> dict[str, tuple[str, ...]]:
        return {kind: sheet.fields for kind, sheet in self.sheets.items()}

    def bands_by_kind(self) -> dict[str, list[dict]]:
        return {
            kind: [band.as_dict() for band in sheet.bands]
            for kind, sheet in self.sheets.items()
        }


# --------------------------------------------------------------------------- #
# FROZEN LITERAL canonical transcriptions
#
# Every tuple below is written out here on purpose. It is a transcription of
# the layout this repo wrote before slice S7, and it must NEVER be re-derived
# from ``bulk_import.FIELDS_BY_KIND`` / ``SECTION_BANDS`` / ``SHEET_BY_KIND``:
# S7 redefines those to the reference layout, and a derived entry would follow
# it, silently un-registering the canonical layout that older workbooks use.
# The regular answer / sub-question blocks are expanded from FROZEN LITERAL
# counts below, the way this package has always described them ("the answer /
# sub-question blocks are regular, so they are generated rather than
# transcribed" — bulk_import/__init__.py), so the expansion depends on nothing
# outside this module.
# --------------------------------------------------------------------------- #

# Sheet names, frozen — note the trailing space on Objective. It is the exact
# byte sequence historical workbooks carry, and getting it wrong is the bug
# this module exists to make impossible.
_FROZEN_SHEET_NAMES = {
    "objective": "Objective ",
    "subjective": "Subjective",
    "descriptive": "Descriptive",
}
_FROZEN_DOC_LINK_SHEET = "Doc Link <> Each fields "

_FROZEN_CHAPTER_FIELDS = (
    "chapter_title", "chapter_display_name", "chapter_duration",
    "pre_topics", "post_topics", "chapter_description",
)
_FROZEN_TOPIC_FIELDS = (
    "topic_title", "topic_display_name", "pre_post_learning",
    "topic_concept_labels", "related_topics", "topic_description",
)
_FROZEN_CONCEPT_FIELDS = (
    "concept_title", "concept_display_name", "parent_concept",
    "concept_details", "digicards",
    "basic_groups", "intermediate_groups", "advanced_groups",
    "concept_source",
)
_FROZEN_LEGACY_CONCEPT_FIELDS = (
    "concept_title", "concept_display_name", "concept_details",
    "keywords", "digicards", "related_concepts",
    "basic_groups", "intermediate_groups", "advanced_groups",
)
_FROZEN_OBJECTIVE_GROUP_FIELDS = (
    "concept_question_labels",
    "group_name", "group_display_name", "group_description",
    "group_status", "group_type",
    "group_question_labels", "related_digicards",
)
_FROZEN_SUBJECTIVE_GROUP_FIELDS = _FROZEN_OBJECTIVE_GROUP_FIELDS
_FROZEN_DESCRIPTIVE_GROUP_FIELDS = (
    "concept_question_labels", "group_display_name", "group_description",
    "group_name", "group_status", "group_type",
    "question_label", "group_question_labels", "related_digicards",
)

# The 10 / 11 / 12 scalar question fields, frozen literally. The reader used
# to hard-code these three counts (reader.py:85/96/108); they are properties
# of a LAYOUT, and every consumer now derives them from the field list above
# (`SheetLayout.question_scalar_fields`).
_FROZEN_OBJECTIVE_SCALARS = (
    "question_label", "question_category", "cognitive_skills",
    "question_source", "question_disclaimer", "question_duration",
    "question_appears_in", "level_of_difficulty", "question", "marks",
)
_FROZEN_SUBJECTIVE_SCALARS = (
    "question_label", "question_category", "cognitive_skills",
    "question_source", "question_disclaimer", "question_duration",
    "math_keyboard", "question_appears_in", "level_of_difficulty",
    "question", "marks",
)
_FROZEN_DESCRIPTIVE_SCALARS = (
    "question_label", "question_category", "cognitive_skills",
    "question_source", "question_disclaimer", "question_duration",
    "math_keyboard", "question_appears_in", "level_of_difficulty",
    "question", "marks", "display_answer",
)

# Frozen block counts of the canonical layout.
_FROZEN_OBJECTIVE_ANSWER_BLOCKS = 6      # answer_type_1 .. answer_type_6
_FROZEN_SUBJECTIVE_ANSWER_BLOCKS = 10    # the reader's historical range(10)
_FROZEN_DESCRIPTIVE_ANSWER_BLOCKS = 10
_FROZEN_DESCRIPTIVE_SUBQUESTIONS = 15
_FROZEN_DESCRIPTIVE_SUBQUESTION_KEYWORDS = 6


def _frozen_objective_question_fields(*, question_text: bool) -> tuple[str, ...]:
    fields = list(_FROZEN_OBJECTIVE_SCALARS)
    for n in range(1, _FROZEN_OBJECTIVE_ANSWER_BLOCKS + 1):
        for prefix in ("answer_type", "answer_content", "correct_answer",
                       "answer_weightage"):
            fields.append(f"{prefix}_{n}")
    fields.append("answer_explanation")
    if question_text:
        fields.append("question_text")
    return tuple(fields)


def _frozen_subjective_question_fields(*, question_text: bool) -> tuple[str, ...]:
    fields = list(_FROZEN_SUBJECTIVE_SCALARS)
    for n in range(1, _FROZEN_SUBJECTIVE_ANSWER_BLOCKS + 1):
        for prefix in ("answer_type", "answer", "answer_display", "weightage",
                       "placeholder"):
            fields.append(f"{prefix}_{n}")
    fields.append("answer_explanation")
    if question_text:
        fields.append("question_text")
    return tuple(fields)


def _frozen_descriptive_question_fields(*, question_text: bool) -> tuple[str, ...]:
    fields = list(_FROZEN_DESCRIPTIVE_SCALARS)
    for n in range(1, _FROZEN_DESCRIPTIVE_ANSWER_BLOCKS + 1):
        for prefix in ("answer_type", "answer_weightage", "answer_content"):
            fields.append(f"{prefix}_{n}")
    fields.append("answer_explanation")
    for n in range(1, _FROZEN_DESCRIPTIVE_SUBQUESTIONS + 1):
        fields.append(f"sub_question_{n}")
        fields.append(f"sub_question_marks_{n}")
        for m in range(1, _FROZEN_DESCRIPTIVE_SUBQUESTION_KEYWORDS + 1):
            for prefix in ("answer_type", "weightage", "keyword"):
                fields.append(f"sq{n}_{prefix}_{m}")
    if question_text:
        fields.append("question_text")
    return tuple(fields)


def _bands_from_spans(spans: Sequence[tuple[str, int]]) -> tuple[Band, ...]:
    bands: list[Band] = []
    column = 1
    for label, span in spans:
        bands.append(Band(label, column, column + span - 1))
        column += span
    return tuple(bands)


def _canonical_layout(layout_id: str, *, concept_fields: Sequence[str],
                      question_text: bool) -> Layout:
    """Build one frozen canonical entry from the literals above."""
    question_by_kind = {
        "objective": _frozen_objective_question_fields(
            question_text=question_text),
        "subjective": _frozen_subjective_question_fields(
            question_text=question_text),
        "descriptive": _frozen_descriptive_question_fields(
            question_text=question_text),
    }
    group_by_kind = {
        "objective": _FROZEN_OBJECTIVE_GROUP_FIELDS,
        "subjective": _FROZEN_SUBJECTIVE_GROUP_FIELDS,
        "descriptive": _FROZEN_DESCRIPTIVE_GROUP_FIELDS,
    }
    sheets: dict[str, SheetLayout] = {}
    for kind, question_fields in question_by_kind.items():
        blocks = (
            ("chapter", _FROZEN_CHAPTER_FIELDS),
            ("topic", _FROZEN_TOPIC_FIELDS),
            ("concept", tuple(concept_fields)),
            ("group", group_by_kind[kind]),
            ("question", question_fields),
        )
        # Row-1 bands of the canonical sheets: the Concept band visually owns
        # the trailing linkage column and the Group band is one shorter, which
        # is why ``bands`` and ``blocks`` are separate things here.
        spans = (
            ("Chapter", len(_FROZEN_CHAPTER_FIELDS)),
            ("Topic", len(_FROZEN_TOPIC_FIELDS)),
            ("Concept", len(concept_fields) + 1),
            ("Group", len(group_by_kind[kind]) - 1),
            ("Question", len(question_fields)),
        )
        sheets[kind] = SheetLayout(
            layout_id, kind, _FROZEN_SHEET_NAMES[kind], blocks,
            _bands_from_spans(spans),
        )
    return Layout(layout_id, sheets,
                  ignored_sheets=(_FROZEN_DOC_LINK_SHEET,),
                  source="frozen literal transcription (spec-step8 T6)")


# --------------------------------------------------------------------------- #
# sop-mes-1 — read from the COMMITTED FORMAT WORKBOOK (T3.1b)
# --------------------------------------------------------------------------- #

REFERENCE_LAYOUT_ID = "sop-mes-1"
# The update-aware layouts of Master Governing Contract v2.0 §14: the
# committed reference schema plus the five ``is_update_*`` columns and the
# Descriptive ``concept_source`` column — 72 / 380 / 149 columns — and the
# expanded-Descriptive variant with 30 rubric slots (440 columns). They were
# first evidenced by the 2026-08-27 Grade-6 audit workbooks, but they are
# NOT board layouts: every generated output of every board uses the
# update-aware geometry, and only an explicitly frozen profile widens the
# Descriptive slots. The historical names below remain as aliases.
UPDATE_AWARE_MASTER_LAYOUT_ID = "update-aware-master-1"
UPDATE_AWARE_EXPANDED_DESCRIPTIVE_LAYOUT_ID = (
    "update-aware-master-expanded-descriptive-1"
)
MSBSHSE_GRADE_6_MASTER_LAYOUT_ID = UPDATE_AWARE_MASTER_LAYOUT_ID
MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID = (
    UPDATE_AWARE_EXPANDED_DESCRIPTIVE_LAYOUT_ID
)
REFERENCE_WORKBOOK_PATH = (
    Path(__file__).resolve().parents[2]
    / "data" / "Testing" / "reference_bulk_import" / "bulk_import_format.xlsx"
)
# The template JSON is a DERIVED convenience cache of the workbook above, not
# an independent authority. It is compared, never trusted over the workbook.
MANIFEST_PATH = Path(__file__).with_name("assessment_workbook_template.json")

_REFERENCE_KIND_BY_SHEET = {
    "Objective": "objective",
    "Descriptive": "descriptive",
    "Subjective": "subjective",
}


def _header_names(values: Iterable) -> tuple[str, ...]:
    """Return header cell text byte-exactly, dropping only absent tail cells.

    Whitespace is data in a field name.  Stripping here made a workbook with
    ``"question_label "`` or ``"question_label\n"`` identify as canonical
    even though every later lookup addresses ``"question_label"``.  A
    trailing ``None`` is different: openpyxl may expose styled but genuinely
    absent cells after the final header, and those carry no field name.
    Interior ``None`` values become an explicit empty name and therefore fail
    exact tuple equality as they should.
    """

    raw = list(values)
    while raw and raw[-1] is None:
        raw.pop()
    return tuple("" if value is None else str(value) for value in raw)


def _blocks_from_bands(fields: Sequence[str],
                       bands: Sequence[Band]) -> tuple[tuple[str, tuple[str, ...]], ...]:
    """Partition a sheet's columns using its own row-1 bands.

    A column that carries no band of its own belongs to the nearest band that
    starts at or before it — pure column arithmetic (the reference Objective
    sheet leaves ``concept_source`` unbanded between Concept and Group, and
    every column after the one-cell ``Question`` band is question content).
    """
    ordered = sorted(bands, key=lambda band: band.start)
    blocks: list[tuple[str, list[str]]] = []
    for band in ordered:
        blocks.append((band.label.strip().casefold(), []))
    for position, field in enumerate(fields, start=1):
        target = 0
        for i, band in enumerate(ordered):
            if band.start <= position:
                target = i
        blocks[target][1].append(field)
    return tuple((name, tuple(names)) for name, names in blocks)


def _sheet_bands(worksheet, field_count: int) -> tuple[Band, ...]:
    merged = {
        cell_range.min_col: cell_range.max_col
        for cell_range in worksheet.merged_cells.ranges
        if cell_range.min_row == 1 and cell_range.max_row == 1
    }
    bands: list[Band] = []
    for column in range(1, field_count + 1):
        value = worksheet.cell(row=1, column=column).value
        label = "" if value is None else str(value)
        if not label.strip():
            continue
        bands.append(Band(label, column, merged.get(column, column)))
    return tuple(bands)


def build_reference_layout(
    workbook_path: Path | None = None,
    manifest: Mapping | None = None,
) -> tuple[Layout | None, list[dict]]:
    """Build ``sop-mes-1`` from the committed workbook; compare the JSON.

    Returns ``(layout, defects)``. The WORKBOOK WINS on any disagreement: the
    returned layout is always the workbook's, and a ``layout_manifest_drift``
    defect names the sheet, the first divergent column index and both values.
    A drift never halts anything (Q13) — the run completes on the workbook's
    layout.
    """
    path = Path(workbook_path or REFERENCE_WORKBOOK_PATH)
    defects: list[dict] = []
    if not path.exists():
        defects.append({
            "code": "layout_source_missing",
            "path": str(path),
            "detail": (
                "the committed Bulk Import format workbook is not on disk, so "
                f"{REFERENCE_LAYOUT_ID} is not registered in this process"
            ),
        })
        return None, defects

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        sheets: dict[str, SheetLayout] = {}
        sheet_order: list[str] = []
        for sheet_name in workbook.sheetnames:
            kind = _REFERENCE_KIND_BY_SHEET.get(sheet_name)
            if kind is None:
                defects.append({
                    "code": "layout_sheet_unknown",
                    "sheet": sheet_name,
                    "detail": "no content kind is declared for this sheet",
                })
                continue
            worksheet = workbook[sheet_name]
            fields = _header_names(
                next(worksheet.iter_rows(min_row=2, max_row=2,
                                         values_only=True), ())
            )
            bands = _sheet_bands(worksheet, len(fields))
            sheets[kind] = SheetLayout(
                REFERENCE_LAYOUT_ID, kind, sheet_name,
                _blocks_from_bands(fields, bands), bands,
            )
            sheet_order.append(sheet_name)
    finally:
        workbook.close()

    if not sheets:
        defects.append({
            "code": "layout_source_missing",
            "path": str(path),
            "detail": "the format workbook carries no recognised content sheet",
        })
        return None, defects

    cache = manifest
    if cache is None:
        try:
            cache = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            cache = None
    if cache is not None:
        defects.extend(_manifest_drift(sheets, sheet_order, cache, path))

    return Layout(REFERENCE_LAYOUT_ID, sheets, ignored_sheets=(),
                  source=f"committed format workbook {path.name}"), defects


def _manifest_drift(sheets: Mapping[str, SheetLayout], sheet_order: Sequence[str],
                    manifest: Mapping, path: Path) -> list[dict]:
    """Tuple inequality between the workbook and its derived JSON cache."""
    defects: list[dict] = []
    cached_order = list(manifest.get("sheet_order") or [])
    if cached_order != list(sheet_order):
        defects.append({
            "code": "layout_manifest_drift",
            "layout_id": REFERENCE_LAYOUT_ID,
            "sheet": "",
            "column_index": None,
            "workbook_value": list(sheet_order),
            "manifest_value": cached_order,
            "detail": "sheet order differs; the workbook's order is used",
            "source": str(path),
        })
    cached_sheets = manifest.get("sheets") or {}
    for kind, sheet in sheets.items():
        cached = (cached_sheets.get(sheet.sheet_name) or {}).get("fields")
        if cached is None:
            defects.append({
                "code": "layout_manifest_drift",
                "layout_id": REFERENCE_LAYOUT_ID,
                "sheet": sheet.sheet_name,
                "column_index": None,
                "workbook_value": len(sheet.fields),
                "manifest_value": None,
                "detail": "the JSON cache carries no field list for this sheet",
                "source": str(path),
            })
            continue
        cached_fields = list(cached)
        if cached_fields == list(sheet.fields):
            continue
        index = next(
            (
                i for i in range(max(len(cached_fields), len(sheet.fields)))
                if (sheet.fields[i] if i < len(sheet.fields) else None)
                != (cached_fields[i] if i < len(cached_fields) else None)
            ),
            None,
        )
        defects.append({
            "code": "layout_manifest_drift",
            "layout_id": REFERENCE_LAYOUT_ID,
            "sheet": sheet.sheet_name,
            "column_index": index,
            "workbook_value": (
                sheet.fields[index] if index is not None
                and index < len(sheet.fields) else None
            ),
            "manifest_value": (
                cached_fields[index] if index is not None
                and index < len(cached_fields) else None
            ),
            "detail": (
                "the committed format workbook and its JSON cache disagree; "
                "the workbook wins"
            ),
            "source": str(path),
        })
    return defects


# --------------------------------------------------------------------------- #
# Normalized 2026-08-27 audit Master layouts
#
# These are output-role layouts, not replacements for ``sop-mes-1``.  They
# are derived only from the committed reference workbook's registered fields
# and bands.  In particular, they do not import ``assessment_workbook`` (which
# already imports this registry through ``bulk_import``) and therefore cannot
# create a renderer/reader import cycle.
#
# The hand-edited audit workbooks are evidence, not layout authorities: some
# carry duplicate headers and styled trailing columns.  The registered target
# is instead the unique insertion contract below.  Exact header equality then
# makes those raw defects fail closed while accepting the clean renderer.
# --------------------------------------------------------------------------- #

_AUDIT_UPDATE_FIELD_AFTER = (
    ("is_update_chapter", "chapter_title"),
    ("is_update_topic", "topic_title"),
    ("is_update_concept", "concept_title"),
    ("is_update_question", "question_label"),
)
_AUDIT_BASE_DESCRIPTIVE_ANSWER_SLOTS = 10
_AUDIT_ENGLISH_POST_DESCRIPTIVE_ANSWER_SLOTS = 30


def _insert_unique_after(fields: list[str], anchor: str, field: str) -> None:
    """Insert one new field after one unambiguous reference-field anchor."""

    if fields.count(anchor) != 1:
        raise ValueError(
            f"audit Master anchor {anchor!r} is not unique in the reference "
            "layout"
        )
    if field in fields:
        raise ValueError(
            f"audit Master field {field!r} already exists in the reference "
            "layout"
        )
    fields.insert(fields.index(anchor) + 1, field)


def _audit_master_fields(
    reference: SheetLayout,
    *,
    descriptive_answer_slots: int,
) -> tuple[str, ...]:
    """Reference fields plus the audit's documented, unique insertions."""

    fields = list(reference.fields)
    for field, anchor in _AUDIT_UPDATE_FIELD_AFTER:
        _insert_unique_after(fields, anchor, field)
    _insert_unique_after(
        fields,
        "group_display_name" if reference.kind == "descriptive" else "group_name",
        "is_update_group",
    )

    if reference.kind == "descriptive":
        _insert_unique_after(
            fields, "concept_question_labels", "concept_source",
        )
        if descriptive_answer_slots < _AUDIT_BASE_DESCRIPTIVE_ANSWER_SLOTS:
            raise ValueError(
                "an audit Master cannot remove committed descriptive answer "
                "slots"
            )
        extra = [
            f"{prefix}_{number}"
            for number in range(
                _AUDIT_BASE_DESCRIPTIVE_ANSWER_SLOTS + 1,
                descriptive_answer_slots + 1,
            )
            for prefix in (
                "answer_type", "answer_weightage", "answer_content",
            )
        ]
        explanation = fields.index("answer_explanation")
        fields[explanation:explanation] = extra

    if len(fields) != len(set(fields)):
        duplicates = sorted({name for name in fields if fields.count(name) > 1})
        raise ValueError(
            "normalized audit Master fields must be unique; duplicates: "
            + ", ".join(duplicates)
        )
    return tuple(fields)


def _audit_master_bands(
    reference: SheetLayout,
    fields: Sequence[str],
) -> tuple[Band, ...]:
    """Rebase reference bands by their boundary field names.

    This preserves the committed workbook's intentional band gaps.  The only
    new ownership statement is Descriptive ``concept_source``: the audit log
    places it at the end of the Concept band.
    """

    rebased: list[Band] = []
    for band in reference.bands:
        start_field = reference.fields[band.start - 1]
        end_field = reference.fields[band.end - 1]
        end = fields.index(end_field) + 1
        if (
            reference.kind == "descriptive"
            and band.label.strip() == "Concept"
        ):
            end = fields.index("concept_source") + 1
        rebased.append(Band(
            band.label,
            fields.index(start_field) + 1,
            end,
        ))
    return tuple(rebased)


def _audit_master_layout(
    reference: Layout,
    layout_id: str,
    *,
    descriptive_answer_slots: int,
) -> Layout:
    """Build one normalized Master reader layout from ``sop-mes-1``."""

    sheets: dict[str, SheetLayout] = {}
    for kind, reference_sheet in reference.sheets.items():
        fields = _audit_master_fields(
            reference_sheet,
            descriptive_answer_slots=descriptive_answer_slots,
        )
        bands = _audit_master_bands(reference_sheet, fields)
        sheets[kind] = SheetLayout(
            layout_id,
            kind,
            reference_sheet.sheet_name,
            _blocks_from_bands(fields, bands),
            bands,
        )
    return Layout(
        layout_id,
        sheets,
        ignored_sheets=(),
        source=(
            "update-aware layout (contract v2.0 §14) derived from "
            f"{reference.id}"
        ),
    )


# --------------------------------------------------------------------------- #
# The registry
# --------------------------------------------------------------------------- #

CANONICAL_CURRENT = _canonical_layout(
    "canonical-current",
    concept_fields=_FROZEN_CONCEPT_FIELDS, question_text=True,
)
# The 64-column Objective variant (and its siblings): templates predating the
# trailing ``question_text`` column. These really do reach the reader.
CANONICAL_NO_QUESTION_TEXT = _canonical_layout(
    "canonical-no-question-text",
    concept_fields=_FROZEN_CONCEPT_FIELDS, question_text=False,
)
# The concept band of the OLD layout (no parent_concept / concept_source, with
# keywords / related_concepts).
CANONICAL_LEGACY_CONCEPT_BAND = _canonical_layout(
    "canonical-legacy-concept-band",
    concept_fields=_FROZEN_LEGACY_CONCEPT_FIELDS, question_text=True,
)

LAYOUTS: dict[str, Layout] = {}


def register(layout: Layout) -> None:
    LAYOUTS[layout.id] = layout


_reference_layout, _reference_defects = build_reference_layout()
_REGISTRY_DEFECTS.extend(_reference_defects)
for _defect in _REGISTRY_DEFECTS:
    # Recorded and visible, never a halt (Q13). Surfacing these on the release
    # audit is the audit seam's job and arrives with it; nothing here guesses.
    logging.getLogger(__name__).warning(
        "bulk-import layout registry: %s (%s)",
        _defect.get("code"), _defect.get("detail"),
    )
if _reference_layout is not None:
    register(_reference_layout)
    register(_audit_master_layout(
        _reference_layout,
        MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
        descriptive_answer_slots=_AUDIT_BASE_DESCRIPTIVE_ANSWER_SLOTS,
    ))
    register(_audit_master_layout(
        _reference_layout,
        MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
        descriptive_answer_slots=(
            _AUDIT_ENGLISH_POST_DESCRIPTIVE_ANSWER_SLOTS
        ),
    ))
register(CANONICAL_CURRENT)
register(CANONICAL_NO_QUESTION_TEXT)
register(CANONICAL_LEGACY_CONCEPT_BAND)

# These layouts are current, deliberate output roles.  They are not the
# historical inputs that must have a migration home in ``sop-mes-1``.  Keeping
# that classification explicit prevents a future registry-wide migration
# invariant from treating 20 legitimate English descriptive answer slots as
# disposable legacy columns.
CURRENT_TARGET_LAYOUT_IDS = frozenset({
    REFERENCE_LAYOUT_ID,
    MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
    MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
})

# A generated Master is one three-tab artifact, in this exact order.  Older
# registered inputs retain their historical partial-workbook tolerance.
STRICT_COMPLETE_LAYOUT_IDS = frozenset({
    MSBSHSE_GRADE_6_MASTER_LAYOUT_ID,
    MSBSHSE_GRADE_6_ENGLISH_POST_MASTER_LAYOUT_ID,
})


def layout(layout_id: str) -> Layout:
    return LAYOUTS[layout_id]


def sheet(layout_id: str, kind: str) -> SheetLayout:
    return LAYOUTS[layout_id].sheet(kind)


@dataclass(frozen=True)
class SheetIdentity:
    layout_id: str
    kind: str
    sheet_name: str

    @property
    def sheet(self) -> SheetLayout:
        return LAYOUTS[self.layout_id].sheet(self.kind)


def identify_sheet(sheet_name: str, header_names: Iterable) -> SheetIdentity | None:
    """Identify one sheet by EXACT sheet name + header tuple equality.

    No prefix matching, no fuzzy matching, no scanning. A trailing space in a
    sheet name makes a different layout, not a skip.
    """
    names = _header_names(header_names)
    if not names:
        return None
    for entry in LAYOUTS.values():
        for kind, candidate in entry.sheets.items():
            if candidate.sheet_name == sheet_name and candidate.fields == names:
                return SheetIdentity(entry.id, kind, sheet_name)
    return None


@dataclass(frozen=True)
class WorkbookIdentity:
    layout_id: str
    sheets: tuple[SheetIdentity, ...]
    ignored_sheets: tuple[str, ...]

    @property
    def layout(self) -> Layout:
        return LAYOUTS[self.layout_id]


def _closest(sheet_name: str, names: Sequence[str]) -> tuple[SheetLayout | None, int | None]:
    """The registered sheet a header most nearly is, by common prefix."""
    best: SheetLayout | None = None
    best_score = -1
    for entry in LAYOUTS.values():
        for candidate in entry.sheets.values():
            shared = 0
            for a, b in zip(candidate.fields, names):
                if a != b:
                    break
                shared += 1
            score = shared + (1000 if candidate.sheet_name == sheet_name else 0)
            if score > best_score:
                best, best_score = candidate, score
    if best is None:
        return None, None
    divergent = next(
        (
            i for i in range(max(len(best.fields), len(names)))
            if (best.fields[i] if i < len(best.fields) else None)
            != (names[i] if i < len(names) else None)
        ),
        None,
    )
    return best, divergent


def _refuse(sheet_name: str, names: Sequence[str], reason: str) -> WorkbookLayoutError:
    closest, divergent = _closest(sheet_name, names)
    detail = {
        "sheet": sheet_name,
        "column_count": len(names),
        "reason": reason,
        "closest_layout_id": closest.layout_id if closest else None,
        "closest_kind": closest.kind if closest else None,
        "closest_sheet_name": closest.sheet_name if closest else None,
        "closest_column_count": len(closest.fields) if closest else None,
        "first_divergent_index": divergent,
        "found": (
            names[divergent]
            if divergent is not None and divergent < len(names) else None
        ),
        "expected": (
            closest.fields[divergent]
            if closest and divergent is not None
            and divergent < len(closest.fields) else None
        ),
    }
    message = (
        f"unrecognised Bulk Import layout: sheet {sheet_name!r} carries "
        f"{len(names)} columns and matches no registered layout ({reason})."
    )
    if closest is not None:
        message += (
            f" Closest registered layout: {closest.layout_id!r} sheet "
            f"{closest.sheet_name!r} ({len(closest.fields)} columns)"
        )
        if divergent is not None:
            message += (
                f"; first difference at column {divergent + 1}: found "
                f"{detail['found']!r}, expected {detail['expected']!r}"
            )
        message += "."
    return WorkbookLayoutError(message, detail=detail)


def _workbook_claim(
    entry: Layout,
    headers: Mapping[str, Sequence],
) -> WorkbookIdentity | None:
    """Return ``entry``'s whole-workbook claim, if every tab belongs to it.

    Two normalized Master layouts intentionally share their Objective and
    Subjective headers; only English Post expands Descriptive.  Choosing a
    layout one sheet at a time would therefore produce a false mixed-layout
    refusal.  Whole-workbook matching resolves the discriminator without any
    fuzzy/header-prefix logic.
    """

    matched: list[SheetIdentity] = []
    ignored: list[str] = []
    for sheet_name, values in headers.items():
        names = _header_names(values)
        candidate = next(
            (
                (kind, sheet_layout)
                for kind, sheet_layout in entry.sheets.items()
                if sheet_layout.sheet_name == sheet_name
                and sheet_layout.fields == names
            ),
            None,
        )
        if candidate is not None:
            kind, _sheet_layout = candidate
            matched.append(SheetIdentity(entry.id, kind, sheet_name))
            continue
        if sheet_name in entry.ignored_sheets:
            ignored.append(sheet_name)
            continue
        return None
    if not matched:
        return None
    return WorkbookIdentity(entry.id, tuple(matched), tuple(ignored))


def _require_strict_workbook_shape(
    identity: WorkbookIdentity,
    headers: Mapping[str, Sequence],
) -> None:
    """Refuse an incomplete/reordered normalized Master output."""

    if identity.layout_id not in STRICT_COMPLETE_LAYOUT_IDS:
        return
    expected_order = tuple(
        sheet_layout.sheet_name
        for sheet_layout in identity.layout.sheets.values()
    )
    found_order = tuple(headers)
    if found_order == expected_order:
        return
    found = identity.sheets[0]
    raise _refuse(
        found.sheet_name,
        found.sheet.fields,
        "normalized Master sheets must be complete and in canonical "
        f"order {expected_order!r}; found {found_order!r}",
    )


def identify_workbook(headers: Mapping[str, Sequence]) -> WorkbookIdentity:
    """Identify a whole workbook, or refuse it.

    ``headers`` maps every sheet name in the file to its row-2 values, in
    workbook order. Exactly one registered layout must claim every content
    sheet; a sheet that no layout claims and that the claimed layout does not
    declare as a non-content sheet refuses the WHOLE workbook — reading part
    of a file whose geometry is unknown is the corruption this gate exists to
    prevent.
    """
    # Prefer one layout that claims the artifact as a whole.  This is still
    # exact equality, and is necessary because the 10-slot and 30-slot Master
    # roles deliberately share two of their three sheet schemas.
    claims = [
        claim
        for entry in LAYOUTS.values()
        if (claim := _workbook_claim(entry, headers)) is not None
    ]
    if len(claims) == 1:
        _require_strict_workbook_shape(claims[0], headers)
        return claims[0]

    identified: list[SheetIdentity] = []
    unidentified: list[tuple[str, tuple[str, ...]]] = []
    for sheet_name, values in headers.items():
        names = _header_names(values)
        found = identify_sheet(sheet_name, names)
        if found is None:
            unidentified.append((sheet_name, names))
        else:
            identified.append(found)

    if not identified:
        sheet_name, names = next(
            ((name, values) for name, values in unidentified if values),
            (next(iter(headers), ""), ()),
        )
        raise _refuse(sheet_name, names, "no sheet matched a registered layout")

    layout_ids = {found.layout_id for found in identified}
    if len(layout_ids) > 1:
        first = identified[0]
        raise _refuse(
            first.sheet_name, first.sheet.fields,
            "the sheets of this workbook match "
            f"{len(layout_ids)} different layouts ({', '.join(sorted(layout_ids))})",
        )

    layout_id = layout_ids.pop()
    entry = LAYOUTS[layout_id]
    _require_strict_workbook_shape(
        WorkbookIdentity(layout_id, tuple(identified), ()), headers,
    )
    ignored: list[str] = []
    for sheet_name, names in unidentified:
        if sheet_name in entry.ignored_sheets:
            ignored.append(sheet_name)
            continue
        raise _refuse(
            sheet_name, names,
            f"the rest of this workbook is layout {layout_id!r}",
        )
    return WorkbookIdentity(layout_id, tuple(identified), tuple(ignored))


# --------------------------------------------------------------------------- #
# Migrating a workbook from one registered layout to another (Q16 / T7.6)
#
# The SUBJECT is ``config.BULK_IMPORT_OUTPUT`` — the append-only generation
# accumulator, gitignored runtime state on the deployed volume. It is NOT
# ``backend/data/bulk_import_database.xlsx``, which nothing appends to and
# which CI regenerates through this package's writer before every pytest run
# (OWNER RULING OD2 / T7.6a). There is no fallback path, no second file and no
# sequence-of-paths plumbing.
#
# This function NEVER RAISES on row content. ``writer.append_concepts`` is
# reached from ``build_concepts.py`` inside a generation run that has already
# spent its entire model budget, so a refusal here is the halt Q13 retired.
# An unmappable non-blank value lands in the receipt's ``unmappable`` list and
# becomes one recorded Fixer decision plus a release issue; the pre-migration
# workbook is retained beside the target; the run completes.
# --------------------------------------------------------------------------- #

MIGRATION_UNMAPPABLE_VALUE = "layout_migration_unmappable_value"
MIGRATION_DROPPED_SHEET = "layout_migration_dropped_sheet"


@dataclass(frozen=True)
class LayoutAlias:
    """One RECORDED column alias between two registered layouts.

    An alias is never inferred. [measured] canonical Descriptive carries
    ``question_label`` twice — at index 27 inside the Group band and at index
    30 inside the Question band — and on this repo's own artifact both are
    non-blank on 56 of 56 rows with IDENTICAL values on all 56, because
    ``writer._question_to_row`` wrote ``q.question_label`` into both. The
    target carries the column once. Under the band-qualified logical key that
    makes name-addressing possible, ``group.question_label`` therefore has no
    destination and would be reported unmappable on every real workbook. The
    collapse is provably safe, but it is written down here as an explicit
    entry rather than derived by a "same name, must be the same thing" rule —
    which is precisely the kind of shape-matching that reads a workbook wrong.
    """

    source_layout_id: str
    target_layout_id: str
    kind: str
    source: tuple[str, str]      # (band, field) in the source layout
    target: tuple[str, str]      # (band, field) in the target layout
    reason: str

    @property
    def name(self) -> str:
        return (
            f"{self.source_layout_id}:{self.kind}:"
            f"{self.source[0]}.{self.source[1]} -> "
            f"{self.target_layout_id}:{self.target[0]}.{self.target[1]}"
        )


_DUPLICATE_QUESTION_LABEL_REASON = (
    "canonical Descriptive carries question_label in BOTH the Group band "
    "(idx 27) and the Question band (idx 30); the writer wrote "
    "q.question_label into both, and [measured] the two agree on 56 of 56 "
    "rows of this repo's own artifact. The target carries the column once, "
    "in the Question band, so the Group-band copy collapses onto it."
)

_CONCEPT_QUESTION_LABELS_REASON = (
    "the canonical layouts open the Group band with concept_question_labels "
    "(the label list linking a concept to its questions); the target keeps "
    "the same column with the same name at the END of the Concept band. Same "
    "column, same meaning, different band — so it is an alias, and without it "
    "every migrated question row would report the concept->question linkage "
    "unmappable and lose it."
)

_CANONICAL_LAYOUT_IDS = (
    "canonical-current",
    "canonical-no-question-text",
    "canonical-legacy-concept-band",
)

MIGRATION_ALIASES: tuple[LayoutAlias, ...] = tuple(
    LayoutAlias(
        source_layout_id=source_id,
        target_layout_id=REFERENCE_LAYOUT_ID,
        kind="descriptive",
        source=("group", "question_label"),
        target=("question", "question_label"),
        reason=_DUPLICATE_QUESTION_LABEL_REASON,
    )
    for source_id in _CANONICAL_LAYOUT_IDS
) + tuple(
    LayoutAlias(
        source_layout_id=source_id,
        target_layout_id=REFERENCE_LAYOUT_ID,
        kind=kind,
        source=("group", "concept_question_labels"),
        target=("concept", "concept_question_labels"),
        reason=_CONCEPT_QUESTION_LABELS_REASON,
    )
    for source_id in _CANONICAL_LAYOUT_IDS
    for kind in ("objective", "descriptive", "subjective")
)


def aliases_for(source_layout_id: str, target_layout_id: str,
                kind: str) -> tuple[LayoutAlias, ...]:
    return tuple(
        alias for alias in MIGRATION_ALIASES
        if alias.source_layout_id == source_layout_id
        and alias.target_layout_id == target_layout_id
        and alias.kind == kind
    )


@dataclass
class MigrationReceipt:
    """Pure mechanics of one layout migration — Q16's recorded receipt."""

    source_layout_id: str
    target_layout_id: str
    sha256_before: str
    sha256_after: str
    rows_by_sheet_before: dict[str, int]
    rows_by_sheet_after: dict[str, int]
    sibling_path: str
    alias_entries_applied: list[str] = dataclass_field(default_factory=list)
    unmappable: list[dict] = dataclass_field(default_factory=list)
    dropped_sheets: list[str] = dataclass_field(default_factory=list)
    # Non-blank data cells (row 3 onward) each dropped sheet was carrying when
    # it was dropped. A sheet the source layout declares ``ignored`` is not
    # part of either layout, so its columns have no logical names to map — but
    # a reviewer who pasted a link into that tab put content there, and a
    # receipt that records only the sheet's NAME cannot say whether anything
    # was in it. The values themselves are retained verbatim in the
    # pre-migration sibling; this is the magnitude that makes the drop
    # escalatable into one recorded decision instead of a silent tidy-up.
    dropped_sheet_cells: dict[str, int] = dataclass_field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            "source_layout_id": self.source_layout_id,
            "target_layout_id": self.target_layout_id,
            "sha256_before": self.sha256_before,
            "sha256_after": self.sha256_after,
            "rows_by_sheet_before": dict(self.rows_by_sheet_before),
            "rows_by_sheet_after": dict(self.rows_by_sheet_after),
            "sibling_path": self.sibling_path,
            "alias_entries_applied": list(self.alias_entries_applied),
            "unmappable": [dict(entry) for entry in self.unmappable],
            "dropped_sheets": list(self.dropped_sheets),
            "dropped_sheet_cells": dict(self.dropped_sheet_cells),
        }


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _write_layout_headers(worksheet, sheet_layout: "SheetLayout") -> None:
    """The two header rows of one sheet of one layout, byte-exactly.

    Band labels keep their per-sheet trailing whitespace and the bands keep
    their gaps: the reference Objective sheet leaves column 23 unbanded.
    """
    for band in sheet_layout.bands:
        cell = worksheet.cell(row=1, column=band.start)
        cell.value = band.label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if band.end > band.start:
            worksheet.merge_cells(
                start_row=1, start_column=band.start,
                end_row=1, end_column=band.end,
            )
    for index, name in enumerate(sheet_layout.fields, start=1):
        cell = worksheet.cell(row=2, column=index)
        cell.value = name
        cell.font = Font(bold=True, size=9)
    worksheet.freeze_panes = "A3"


def _is_blank(value) -> bool:
    """Whether a source cell is EMPTY. Mechanics, and deliberately narrow.

    ``0`` and ``False`` are values a workbook can legitimately carry (a mark, a
    count), and ``str(value or "").strip()`` calls both of them blank — which
    would drop them from a migration and record nothing. Only ``None`` and
    whitespace-only text are empty.
    """
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _collapse_placements(
    candidates: list[dict],
) -> tuple[dict, list[dict]]:
    """Which candidate lands in a shared target column, and what is recorded.

    Two source cells reach ONE target column only through a recorded alias
    (canonical Descriptive carries ``question_label`` in both the Group and
    the Question band; the target carries it once). Deciding between them is
    mechanics, not judgment about content, and the rule is fixed:

    * a blank candidate never displaces a non-blank one — so a whitespace-only
      cell in the target's own band cannot erase a real aliased value;
    * among the non-blank candidates the one whose OWN band carries this
      logical name in both layouts (the un-aliased one) is kept;
    * every other non-blank candidate that DIFFERS from the keeper is returned
      as a conflict, for the caller to record in the receipt's ``unmappable``
      list. It is never dropped in silence, and the alias is not reported
      applied for that row.

    Identical copies collapse losslessly, which is the case this repo's own
    artifact is in on 56 of 56 rows.
    """
    if len(candidates) == 1:
        return candidates[0], []
    filled = [
        candidate for candidate in candidates
        if not _is_blank(candidate["value"])
    ]
    if not filled:
        return candidates[0], []
    keeper = next(
        (candidate for candidate in filled if candidate["alias"] is None),
        filled[0],
    )
    conflicts = [
        candidate for candidate in filled
        if candidate is not keeper
        and str(candidate["value"]) != str(keeper["value"])
    ]
    return keeper, conflicts


def _non_blank_data_cells(worksheet) -> int:
    """Non-blank cells from row 3 down — a sheet's DATA, not its banner rows."""
    return sum(
        1
        for row in worksheet.iter_rows(min_row=3, values_only=True)
        for value in row
        if not _is_blank(value)
    )


def _workbook_headers(path: Path) -> dict[str, tuple]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return {
            name: next(
                workbook[name].iter_rows(
                    min_row=2, max_row=2, values_only=True),
                (),
            )
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()


def migrate_workbook_layout(
    path: Path, *, target_layout_id: str,
    sibling_for: Path | None = None,
) -> MigrationReceipt | None:
    """Rewrite ``path`` from its registered layout onto ``target_layout_id``.

    Returns ``None`` when there is nothing to do — the file does not exist, it
    is already on the target layout, or it matches no registered layout (which
    ``writer.scan_workbook`` records as a mismatch on the same pass). Returns a
    ``MigrationReceipt`` otherwise. It never raises on row content.

    Every value moves by its BAND-QUALIFIED logical name plus the explicit
    aliases above. A non-blank source value whose logical name has no home in
    the target is left in the receipt's ``unmappable`` list; the caller turns
    it into one recorded Fixer decision and a release issue.

    ``sibling_for`` names the retained pre-migration copy after the workbook
    the OPERATOR knows rather than after ``path``. The generation run hands
    this function the staged temp sibling the transactional outbox provides
    (``build_concepts._stage_concept_workbook``), whose stem is
    ``tempfile.mkstemp``'s dot-hidden random token — so without this the
    recorded recovery path points at a hidden file whose name says nothing
    about which workbook it came from, and every failed run leaves another
    one. Named for the target, the copy is beside the target, says what it is,
    and a re-run overwrites it instead of accumulating orphans.
    """
    from . import workbook_sync
    from .writer import _safe_cell

    path = Path(path)
    if not path.exists():
        return None
    target = LAYOUTS.get(target_layout_id)
    if target is None:
        return None

    with workbook_sync.output_workbook_lock():
        try:
            identity = identify_workbook(_workbook_headers(path))
        except WorkbookLayoutError:
            # Unidentifiable: guessing its geometry is the corruption the
            # registry exists to prevent, and refusing mid-run is the halt
            # Q13 retired. Recorded by the caller's scan; nothing is touched.
            return None
        if identity.layout_id == target_layout_id:
            return None

        source = identity.layout
        sha256_before = _sha256_file(path)
        named_for = Path(sibling_for) if sibling_for is not None else path
        sibling = named_for.with_name(
            f"{named_for.stem}.pre-{identity.layout_id}{named_for.suffix}")
        shutil.copy2(path, sibling)

        rows_before: dict[str, int] = {}
        rows_after: dict[str, int] = {}
        applied: list[str] = []
        unmappable: list[dict] = []
        dropped = [name for name in identity.ignored_sheets]

        old = openpyxl.load_workbook(path, data_only=True)
        new = openpyxl.Workbook()
        new.remove(new.active)
        dropped_cells: dict[str, int] = {}
        try:
            dropped_cells = {
                name: _non_blank_data_cells(old[name])
                for name in dropped if name in old.sheetnames
            }
            for kind, sheet_layout in target.sheets.items():
                worksheet = new.create_sheet(sheet_layout.sheet_name)
                _write_layout_headers(worksheet, sheet_layout)
                source_sheet = source.sheets.get(kind)
                rows_before[sheet_layout.sheet_name] = 0
                rows_after[sheet_layout.sheet_name] = 0
                if source_sheet is None or source_sheet.sheet_name not in old.sheetnames:
                    continue
                alias_by_source = {
                    alias.source: alias
                    for alias in aliases_for(
                        identity.layout_id, target_layout_id, kind)
                }
                out_row = 3
                for row_index, values in enumerate(
                    old[source_sheet.sheet_name].iter_rows(
                        min_row=3, values_only=True),
                    start=3,
                ):
                    if not values or all(
                        _is_blank(value) for value in values
                    ):
                        continue
                    rows_before[sheet_layout.sheet_name] += 1
                    # RESOLVE the whole row before writing any of it. Writing
                    # inside the walk made the alias's own conflict guard
                    # unreachable: source bands are visited in layout order, so
                    # ``group.question_label`` always landed in a still-blank
                    # cell and the Question band then overwrote it with no
                    # comparison at all — a divergent duplicate label left with
                    # no ``unmappable`` entry while the receipt still reported
                    # the alias applied. Resolving first makes the comparison
                    # symmetric and order-independent.
                    placements: dict[int, list[dict]] = {}
                    for block, names in source_sheet.blocks:
                        for name in names:
                            index = source_sheet.column(block, name)
                            if index is None or index >= len(values):
                                continue
                            value = values[index]
                            alias = alias_by_source.get((block, name))
                            key = alias.target if alias is not None else (
                                block, name)
                            column = sheet_layout.column(*key)
                            if column is None:
                                if not _is_blank(value):
                                    unmappable.append({
                                        "sheet": source_sheet.sheet_name,
                                        "kind": kind,
                                        "row": row_index,
                                        "band": block,
                                        "field": name,
                                        "value": str(value),
                                        "source_layout_id": identity.layout_id,
                                        "target_layout_id": target_layout_id,
                                    })
                                continue
                            placements.setdefault(column, []).append({
                                "value": value,
                                "band": block,
                                "field": name,
                                "alias": alias,
                            })
                    wrote_any = False
                    for column, candidates in placements.items():
                        keeper, conflicts = _collapse_placements(candidates)
                        for loser in conflicts:
                            unmappable.append({
                                "sheet": source_sheet.sheet_name,
                                "kind": kind,
                                "row": row_index,
                                "band": loser["band"],
                                "field": loser["field"],
                                "value": str(loser["value"]),
                                "conflicts_with": str(keeper["value"]),
                                "alias": (
                                    loser["alias"].name
                                    if loser["alias"] is not None else None
                                ),
                                "source_layout_id": identity.layout_id,
                                "target_layout_id": target_layout_id,
                            })
                        for candidate in candidates:
                            alias = candidate["alias"]
                            if alias is None or alias.name in applied:
                                continue
                            if any(loser is candidate for loser in conflicts):
                                continue
                            if _is_blank(candidate["value"]):
                                continue
                            applied.append(alias.name)
                        if _is_blank(keeper["value"]):
                            # Nothing to carry: the target cell is already
                            # empty in this freshly built sheet.
                            continue
                        worksheet.cell(
                            row=out_row, column=column + 1,
                            # openpyxl refuses to write a control character;
                            # a raise here is the halt Q13 retired, so the
                            # writer's own sanitiser guards the save.
                            value=_safe_cell(keeper["value"]),
                        )
                        wrote_any = True
                    # ``rows_after`` counts rows that actually LANDED. It used
                    # to be incremented in the same breath as ``rows_before``
                    # with no branch between them, so the receipt's "no rows
                    # lost" field was true by construction and could not be
                    # evidence of anything. A source row whose every non-blank
                    # value is homeless in the target lands nothing — each of
                    # those values is already an ``unmappable`` record and the
                    # sibling still holds the row — and the counts now say so
                    # instead of writing an empty row and calling it survival.
                    if not wrote_any:
                        continue
                    out_row += 1
                    rows_after[sheet_layout.sheet_name] += 1
            # Atomic, like every other write in this package: the subject is
            # runtime state on a deployed volume, and a crash mid-save must
            # not leave it truncated.
            workbook_sync.atomic_save_workbook(new, path)
        finally:
            old.close()
            new.close()

        return MigrationReceipt(
            source_layout_id=identity.layout_id,
            target_layout_id=target_layout_id,
            sha256_before=sha256_before,
            sha256_after=_sha256_file(path),
            rows_by_sheet_before=rows_before,
            rows_by_sheet_after=rows_after,
            sibling_path=str(sibling),
            alias_entries_applied=applied,
            unmappable=unmappable,
            dropped_sheets=dropped,
            dropped_sheet_cells=dropped_cells,
        )
