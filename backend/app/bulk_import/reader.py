"""Read a Bulk Import workbook and normalize it into the DB hierarchy.

The workbook's layout is IDENTIFIED first, by exact sheet name + header-row
equality against ``bulk_import.layouts``; every column is then addressed **by
name** through that layout. A workbook whose header matches no registered
layout is refused whole (``WorkbookLayoutError`` -> 422) before anything is
written: reading part of a file whose column geometry is unknown is how 342 of
344 question-band positions were silently mis-read (spec-step8 T6).

Rows that only carry chapter/topic/concept/group context (no question) still
create the hierarchy nodes; a Question is created only when the Question band
carries text or a label.
"""
from __future__ import annotations

import math
import re
from collections.abc import Mapping
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from . import (
    ANSWER_TYPES, APPEARS_IN_ALL, COGNITIVE_SKILLS, DIFFICULTY_LEVELS,
    merge_sources, normalize_answer_type, normalize_appears_in,
    normalize_cognitive_skills, normalize_difficulty, normalize_question_text,
    split_multi, strip_title_tag, strip_topic_title, to_plain_text,
)
from . import layouts
from . import assessment_workbook as workbook_contract
from .layouts import WorkbookLayoutError  # noqa: F401  (re-exported for api)
from .. import models
from ..services import assessment_profile
from ..services import assessment_release as release_contract
from ..services import directory, identity, katex_rules


class WorkbookContentError(ValueError):
    """A recognized workbook contains content unsafe to persist."""


def _parse_answers(
    qd: dict, kind: str, sheet_layout: layouts.SheetLayout,
) -> tuple[list[dict], list[dict]]:
    """Return (answers, sub_questions) read BY NAME from the question band.

    The block COUNT comes from the identified layout (how many
    ``answer_type_N`` / ``sub_question_N`` columns it declares), never from a
    literal: the canonical Subjective sheet carries 10 answer blocks and the
    reference one carries 20, and the old ``range(10)`` made blocks 11-20
    unreadable by construction.
    """
    answers: list[dict] = []
    sub_questions: list[dict] = []

    def cell(name: str) -> str:
        return qd.get(name, "")

    if kind == "objective":
        for n in sheet_layout.answer_block_numbers:
            atype, content = cell(f"answer_type_{n}"), cell(f"answer_content_{n}")
            if not (atype or content):
                continue
            answers.append({
                "answer_type": atype, "answer_content": content,
                "correct_answer": cell(f"correct_answer_{n}"),
                "answer_weightage": cell(f"answer_weightage_{n}"),
            })
    elif kind == "subjective":
        for n in sheet_layout.answer_block_numbers:
            atype, ans = cell(f"answer_type_{n}"), cell(f"answer_{n}")
            if not (atype or ans):
                continue
            answers.append({
                "answer_type": atype, "answer": ans,
                "answer_display": cell(f"answer_display_{n}"),
                "weightage": cell(f"weightage_{n}"),
                "placeholder": cell(f"placeholder_{n}"),
            })
    else:  # descriptive
        for n in sheet_layout.answer_block_numbers:
            atype, content = cell(f"answer_type_{n}"), cell(f"answer_content_{n}")
            if not (atype or content):
                continue
            answers.append({
                "answer_type": atype,
                "answer_weightage": cell(f"answer_weightage_{n}"),
                "answer_content": content,
            })
        for n in sheet_layout.sub_question_numbers:
            text = cell(f"sub_question_{n}")
            if not text:
                continue
            keywords = []
            for m in sheet_layout.sub_question_keyword_numbers(n):
                atype = cell(f"sq{n}_answer_type_{m}")
                weight = cell(f"sq{n}_weightage_{m}")
                kw = cell(f"sq{n}_keyword_{m}")
                if not (atype or kw):
                    continue
                keywords.append({"answer_type": atype, "weightage": weight, "keyword": kw})
            sub_questions.append({
                "text": text, "marks": cell(f"sub_question_marks_{n}"),
                "keywords": keywords,
            })
    return answers, sub_questions


_MAX_ISSUES = 200
# Identity notes get a slice of the ledger, never all of it: a wholly
# tagless legacy book raises up to two per distinct topic/concept name, and
# an uncapped stream of them would push out the placement and mangled-row
# issues raised later in the same row loop.
_MAX_IDENTITY_NOTES = 50

# A RESTORED id is bounded by the column that has to carry it downstream.
# ``Topic/Concept.machine_id`` is ``String(255)`` but
# ``Question.question_label`` is ``String(128)`` and is
# ``f"{machine_id} Q##"`` — so an id longer than this cannot become a label.
# Read off the column itself rather than written as a literal.
_MAX_RESTORED_MACHINE_ID = (
    models.Question.question_label.type.length - len(" Q00")
)

# Group-band fields that carry meaningful Group identity. The linkage columns
# (``concept_question_labels`` and Descriptive's linkage ``question_label``)
# belong to neighbouring bands and never establish a group by themselves.
_GROUP_IDENTITY_FIELDS = (
    "group_name", "group_display_name", "group_type",
    "group_description", "group_question_labels",
)

_GROUP_SEQUENCE_RE = re.compile(r"(\d+)\s*\)?\s*$")


def _group_sequence(machine_name: str) -> int:
    """Tier sequence from a machine Group ID tail (``... BG02`` -> 2).

    Mechanical ID parsing only — it never decides which group a question
    belongs to.
    """
    match = _GROUP_SEQUENCE_RE.search(str(machine_name or "").strip())
    return int(match.group(1)) if match else 0


_STRICT_RICH_TEXT_CODES = frozenset({
    "unsupported_table",
    "unsupported_katex_command",
    "raw_math_delimiter",
    "literal_newline_escape",
    "katex_row_spacing",
})


def _format_issues(
    label: str,
    *texts: str,
    allowed_codes: frozenset[str] | None = None,
) -> list[str]:
    """Content-format validation: katex/img/link rules (allowed CMS formats)."""
    issues: list[str] = []
    blob = "\n".join(t for t in texts if t)
    messages = {
        "unbalanced_katex": "unbalanced [Katex] tag",
        "nested_katex": "nested [Katex] tag",
        "malformed_katex": "malformed [Katex] tag",
        "empty_katex": "empty [Katex] tag",
        "markdown_image": "Markdown image found — use [img src=\"...\" alt=\"...\"]",
        "raw_math_delimiter": "raw math delimiters found — use [Katex]...[/Katex]",
        "raw_math_expression": (
            "raw equation found — use [Katex]...[/Katex]"),
        "raw_latex": "raw LaTeX found outside a [Katex] tag",
        "unsupported_table": (
            "tabular/Markdown or noncanonical array markup is unsupported "
            "(unsupported_table); "
            "use a canonical [Katex] array, source image, or row/column-"
            "labelled plain text"
        ),
        "unsupported_katex_command": (
            "unsupported KaTeX command found; do not use \\mathrm, \\hspace, "
            "\\phantom, or \\boxed"
        ),
        "katex_row_spacing": (
            "KaTeX row breaks must not include an optional spacing argument"
        ),
        "literal_newline_escape": (
            "literal \\n found before a list/option label — use a real line "
            "break"
        ),
        "unbalanced_image": "unclosed [img] tag",
        "invalid_image_src": "[img] without a full HTTPS src URL",
        "missing_image_alt": "[img] missing alt text",
        "noncanonical_image": (
            "[img] must contain only ordered src and alt attributes"),
    }
    # Import accepts legacy lower-case [katex] and canonicalizes it before
    # persistence; all other syntax rules stay strict.
    for code in katex_rules.rich_text_issues(
        blob, require_canonical_case=False
    ):
        if allowed_codes is not None and code not in allowed_codes:
            continue
        message = messages.get(code)
        if message:
            issues.append(f"{label}: {message}")
    return issues


def _rich_text_for_validation(
    kind: str, answers: list[dict], *texts: str,
) -> tuple[str, ...]:
    """Return learner-facing text with Subjective placeholders masked.

    ``$$a$$`` through ``$$t$$`` are canonical Subjective answer-slot tokens,
    not raw-math delimiters.  Mask only tokens declared by an ordered answer
    block; arbitrary ``$$...$$`` remains visible to the rich-text validator.
    """

    values = tuple(str(text or "") for text in texts)
    if kind != "subjective":
        return values
    placeholders = {
        str(answer.get("placeholder") or "")
        for answer in answers
        if isinstance(answer, dict)
    }
    tokens = tuple(
        f"$${placeholder}$$"
        for placeholder in sorted(placeholders)
        if len(placeholder) == 1 and "a" <= placeholder <= "t"
    )
    if not tokens:
        return values
    masked: list[str] = []
    for value in values:
        for token in tokens:
            value = value.replace(token, "")
        masked.append(value)
    return tuple(masked)


def _answer_format_issues(
    label: str, answers: list[dict], sub_questions: list[dict],
) -> list[str]:
    """Validate typed answer_content cells independently of rich text."""

    messages = {
        "unsupported_table": (
            "contains unsupported tabular/Markdown or noncanonical array "
            "markup"
        ),
        "equation_unsupported_command": (
            "Equation content must not use \\mathrm, \\hspace, \\phantom, "
            "or \\boxed"
        ),
        "equation_row_spacing": (
            "Equation array rows must not include a spacing argument"
        ),
        "equation_katex_wrapper": (
            "Equation content must be raw LaTeX without [Katex]"
        ),
        "equation_math_delimiter": (
            "Equation content must not include math delimiters"
        ),
        "equation_non_latex_markup": (
            "Equation content must not include image/link markup"
        ),
        "equation_plain_text": (
            "Equation prose must be encoded inside a TeX text atom"
        ),
        "phrases_katex": "Phrases content must not include [Katex]",
        "phrases_latex": "Phrases content must not include LaTeX",
        "phrases_math_delimiter": (
            "Phrases content must not include math delimiters"
        ),
        "phrases_markup": (
            "Phrases content must not include image/link markup"
        ),
        "image_missing_source": (
            "Image content must carry an image source (canonical [img] tag "
            "or its URL)"
        ),
        "image_katex_markup": (
            "Image content must not include [Katex]/LaTeX markup"
        ),
    }
    issues: list[str] = []
    for position, answer in enumerate(answers, start=1):
        content = (
            answer.get("answer_content")
            if "answer_content" in answer
            else answer.get("answer")
        )
        if content is None:
            continue
        for code in katex_rules.answer_cell_issues(
            str(answer.get("answer_type") or ""), str(content or "")
        ):
            issues.append(
                f"{label}: answer/rubric block {position}: "
                f"{messages.get(code, code)}"
            )
    for sub_position, subquestion in enumerate(sub_questions, start=1):
        for keyword_position, keyword in enumerate(
            subquestion.get("keywords") or [], start=1
        ):
            for code in katex_rules.answer_cell_issues(
                str(keyword.get("answer_type") or ""),
                str(keyword.get("keyword") or ""),
            ):
                issues.append(
                    f"{label}: subquestion {sub_position} keyword "
                    f"{keyword_position}: {messages.get(code, code)}"
                )
    return issues


def _weightage_sum(answers: list[dict], kind: str) -> float | None:
    key = "weightage" if kind == "subjective" else "answer_weightage"
    total = 0.0
    found = False
    for a in answers:
        raw = str(a.get(key, "") or "").strip()
        if not raw:
            continue
        try:
            total += float(raw)
            found = True
        except ValueError:
            return None
    return total if found else None


def _sheet_headers(wb) -> dict[str, tuple]:
    """Row 2 of every sheet, in workbook order."""
    headers: dict[str, tuple] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers[sheet_name] = next(
            ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    return headers


def _row_chapter_meta(
    sheet_layout: layouts.SheetLayout, row: tuple,
) -> dict:
    """Derive declared board/grade/subject from one tag-bearing row.

    This deliberately uses the same probes as the mutating import path. The
    result selects a declarative profile policy; it never guesses a category
    or changes a row's authored values.
    """

    chapter = sheet_layout.block_values(row, "chapter")
    topic = sheet_layout.block_values(row, "topic")
    concept = sheet_layout.block_values(row, "concept")
    return directory.derive_chapter_meta(
        str(chapter.get("chapter_title") or ""),
        str(chapter.get("chapter_display_name") or ""),
        str(topic.get("topic_title") or ""),
        str(topic.get("topic_display_name") or ""),
        str(topic.get("topic_concept_labels") or ""),
        str(topic.get("related_topics") or ""),
        str(concept.get("concept_title") or ""),
        str(concept.get("concept_display_name") or ""),
        str(chapter.get("post_topics") or ""),
        str(chapter.get("pre_topics") or ""),
    )


def _policy_decimal(value) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, TypeError, ValueError):
        return None
    return number if number.is_finite() else None


def _strict_assessment_policy_issues(
    *,
    sheet_layout: layouts.SheetLayout,
    row: tuple,
    kind: str,
    question: Mapping,
    answers: list[dict],
    row_label: str,
) -> list[str]:
    """Validate a non-generic row policy before the first database write.

    The row already declares its sheet, category, marks, difficulty, duration,
    and answers. Policy validation only compares those values. For a
    per-subpoint contract the represented count is derived exactly from
    ``marks / marks_per_subpoint``; no missing basis is synthesized.
    """

    metadata = _row_chapter_meta(sheet_layout, row)
    policy = assessment_profile.assessment_format_policy(metadata=metadata)
    policy_id = str(policy.get("policy_id") or "")
    if not policy_id or policy_id == "generic-cms":
        return []
    formats = policy.get("formats_by_sheet")
    if not isinstance(formats, Mapping):
        return [
            f"{row_label}: assessment policy {policy_id!r} has no "
            "formats_by_sheet mapping"
        ]
    category = str(question.get("question_category") or "").strip()
    sheet_formats = formats.get(kind)
    rule = (
        sheet_formats.get(category)
        if isinstance(sheet_formats, Mapping)
        else None
    )
    if not isinstance(rule, Mapping):
        return [
            f"{row_label}: question_category {category!r} is not permitted "
            f"for sheet_kind {kind!r} by assessment policy {policy_id!r}"
        ]

    issues: list[str] = []
    marks = _policy_decimal(question.get("marks"))
    represented_count: int | None = None
    marks_rule = rule.get("marks")
    if "marks" in rule and not isinstance(marks_rule, Mapping):
        issues.append(
            f"{row_label}: assessment policy {policy_id!r} has a non-object "
            f"marks rule for {category!r}"
        )
    if marks is not None and marks > 0 and isinstance(marks_rule, Mapping):
        marks_mode = str(marks_rule.get("mode") or "")
        if not marks_mode:
            issues.append(
                f"{row_label}: assessment policy {policy_id!r} has a marks "
                f"rule without a mode for {category!r}"
            )
        elif marks_mode == "fixed":
            allowed = tuple(
                value
                for raw in marks_rule.get("allowed") or ()
                if (value := _policy_decimal(raw)) is not None and value > 0
            )
            if not allowed:
                issues.append(
                    f"{row_label}: assessment policy {policy_id!r} has no "
                    f"positive fixed marks for {category!r}"
                )
            elif marks not in set(allowed):
                issues.append(
                    f"{row_label}: marks {marks:g} must be one of "
                    f"{tuple(str(value) for value in allowed)} for "
                    f"question_category {category!r}"
                )
        elif marks_mode == "per_subpoint":
            unit = _policy_decimal(marks_rule.get("marks_per_subpoint"))
            raw_maximum = marks_rule.get("max_subpoints")
            maximum = _policy_decimal(raw_maximum)
            if "max_subpoints" in marks_rule and (
                isinstance(raw_maximum, bool)
                or not isinstance(raw_maximum, (int, float, Decimal))
                or maximum is None
                or maximum <= 0
                or maximum != maximum.to_integral_value()
            ):
                issues.append(
                    f"{row_label}: assessment policy {policy_id!r} has an "
                    f"invalid max_subpoints for {category!r}; it must be a "
                    "positive integer"
                )
            if unit is None or unit <= 0:
                issues.append(
                    f"{row_label}: assessment policy {policy_id!r} has an "
                    f"invalid marks_per_subpoint for {category!r}"
                )
            else:
                quotient = marks / unit
                if quotient <= 0 or quotient != quotient.to_integral_value():
                    issues.append(
                        f"{row_label}: marks {marks:g} must represent a "
                        "positive whole number of policy subpoints at "
                        f"{unit:g} mark(s) each"
                    )
                else:
                    represented_count = int(quotient)
                    if (
                        maximum is not None
                        and quotient > maximum
                    ):
                        issues.append(
                            f"{row_label}: represented subpoint count "
                            f"{quotient:g} exceeds the wire maximum "
                            f"{maximum:g} for {category!r}"
                        )
                    if kind == "subjective" and len(answers) != represented_count:
                        issues.append(
                            f"{row_label}: Subjective answer count "
                            f"{len(answers)} does not match represented "
                            f"subpoint count {represented_count}"
                        )
                    if kind == "subjective":
                        for position, answer in enumerate(answers, start=1):
                            weight = _policy_decimal(
                                answer.get("weightage")
                            )
                            if weight != unit:
                                issues.append(
                                    f"{row_label}: Subjective answer "
                                    f"{position} weight {weight!s} does not "
                                    "equal policy marks_per_subpoint "
                                    f"{unit:g}"
                                )
        elif marks_mode:
            issues.append(
                f"{row_label}: assessment policy {policy_id!r} has unknown "
                f"marks mode {marks_mode!r} for {category!r}"
            )

    duration = _policy_decimal(question.get("question_duration"))
    duration_rule = rule.get("duration")
    if (
        duration is None
        or duration <= 0
        or not isinstance(duration_rule, Mapping)
        or not duration_rule
    ):
        # The general strict numeric gate reports an absent/non-positive
        # duration. A category with no duration rule retains that gate only.
        return issues
    duration_mode = str(duration_rule.get("mode") or "")
    expected_duration: Decimal | None = None
    if duration_mode == "matrix":
        difficulty = normalize_difficulty(
            str(question.get("level_of_difficulty") or "")
        )
        minutes = duration_rule.get("minutes_by_difficulty")
        expected_duration = (
            _policy_decimal(minutes.get(difficulty))
            if isinstance(minutes, Mapping)
            else None
        )
        if expected_duration is None or expected_duration <= 0:
            issues.append(
                f"{row_label}: assessment policy {policy_id!r} has no "
                f"positive matrix duration for difficulty {difficulty!r}"
            )
            expected_duration = None
    elif duration_mode == "per_subpoint":
        minutes_per_subpoint = _policy_decimal(
            duration_rule.get("minutes_per_subpoint")
        )
        if minutes_per_subpoint is None or minutes_per_subpoint <= 0:
            issues.append(
                f"{row_label}: assessment policy {policy_id!r} has an "
                f"invalid minutes_per_subpoint for {category!r}"
            )
        elif represented_count is not None:
            expected_duration = minutes_per_subpoint * represented_count
        # If marks could not establish a count, the marks defect above is the
        # conclusive failure. Do not invent a duration basis for another one.
    elif duration_mode:
        issues.append(
            f"{row_label}: assessment policy {policy_id!r} has unknown "
            f"duration mode {duration_mode!r} for {category!r}"
        )
    if expected_duration is not None and duration != expected_duration:
        issues.append(
            f"{row_label}: question_duration {duration:g} does not match "
            f"policy duration {expected_duration:g} for {category!r}"
        )
    return issues


def _blocking_content_issues(wb, identified) -> list[str]:
    """Find mechanical content defects before the import mutates the DB."""

    issues: list[str] = []

    def flag(message: str) -> None:
        if len(issues) < _MAX_ISSUES:
            issues.append(message)

    for found in identified.sheets:
        kind = found.kind
        sheet_name = found.sheet_name
        sheet_layout = found.sheet
        worksheet = wb[sheet_name]
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3,
        ):
            if row is None or not any(row):
                continue
            question = sheet_layout.block_values(row, "question")
            # A concept-catalogue tail row deliberately stops before the
            # Question band. It is not a malformed zero-mark question, so
            # question-only medium, rubric, and numeric checks do not apply.
            if not any(
                value is not None and str(value).strip()
                for value in question.values()
            ):
                continue
            label = str(question.get("question_label") or "").strip()
            row_label = label or f"{sheet_name!r} row {row_number}"
            if not label:
                flag(
                    f"{row_label}: populated Question band requires a "
                    "non-blank question_label"
                )
            if not str(question.get("question") or "").strip():
                flag(f"{row_label}: question must not be blank")
            answers, sub_questions = _parse_answers(
                question, kind, sheet_layout,
            )
            normalized_answers = [dict(answer) for answer in answers]
            normalized_sub_questions = [
                {
                    **dict(subquestion),
                    "keywords": [
                        dict(keyword)
                        for keyword in subquestion.get("keywords") or []
                    ],
                }
                for subquestion in sub_questions
            ]
            for issue in _strict_assessment_policy_issues(
                sheet_layout=sheet_layout,
                row=row,
                kind=kind,
                question=question,
                answers=normalized_answers,
                row_label=row_label,
            ):
                flag(issue)
            for position, answer in enumerate(normalized_answers, start=1):
                answer["answer_type"] = normalize_answer_type(
                    str(answer.get("answer_type") or "")
                )
                if answer["answer_type"] not in ANSWER_TYPES:
                    flag(
                        f"{row_label}: answer/rubric block {position} has "
                        f"unsupported or blank answer_type "
                        f"{answer['answer_type']!r}"
                    )
                content_field = (
                    "answer_content"
                    if "answer_content" in answer
                    else "answer" if "answer" in answer else ""
                )
                if content_field:
                    answer[content_field] = katex_rules.raw_answer_cell(
                        str(answer.get("answer_type") or ""),
                        str(answer.get(content_field) or ""),
                    )
                    content = answer.get(content_field)
                    if (
                        kind == "objective"
                        and release_contract.option_content_has_label(content)
                    ):
                        flag(
                            f"{row_label}: objective option {position} "
                            "repeats its letter label"
                        )
                    if (
                        kind == "descriptive"
                        and release_contract.malformed_rubric_tag(
                            content, answer.get("answer_type"),
                        )
                    ):
                        flag(
                            f"{row_label}: descriptive rubric {position} "
                            "does not start with an allowed functional tag or "
                            "is without its required colon"
                        )
            for subquestion in normalized_sub_questions:
                for keyword_position, keyword in enumerate(
                    subquestion.get("keywords") or [], start=1,
                ):
                    keyword["answer_type"] = normalize_answer_type(
                        str(keyword.get("answer_type") or "")
                    )
                    if keyword["answer_type"] not in ANSWER_TYPES:
                        flag(
                            f"{row_label}: keyword block {keyword_position} "
                            "has unsupported or blank answer_type "
                            f"{keyword['answer_type']!r}"
                        )
                    keyword["keyword"] = katex_rules.raw_answer_cell(
                        str(keyword.get("answer_type") or ""),
                        str(keyword.get("keyword") or ""),
                    )
                    if release_contract.malformed_rubric_tag(
                        keyword.get("keyword"), keyword.get("answer_type"),
                    ):
                        flag(
                            f"{row_label}: subquestion keyword "
                            f"{keyword_position} does not start with an "
                            "allowed functional tag or is without its "
                            "required colon"
                        )
            for field in (
                "question", "question_text", "display_answer",
                "answer_explanation",
            ):
                validation_text, = _rich_text_for_validation(
                    kind, normalized_answers, str(question.get(field) or ""),
                )
                for issue in _format_issues(
                    f"{row_label}: {field}", validation_text,
                    allowed_codes=_STRICT_RICH_TEXT_CODES,
                ):
                    flag(issue)
            for position, subquestion in enumerate(
                sub_questions, start=1,
            ):
                for issue in _format_issues(
                    f"{row_label}: sub_question_{position}",
                    str(subquestion.get("text") or ""),
                    allowed_codes=_STRICT_RICH_TEXT_CODES,
                ):
                    flag(issue)
            for position, answer in enumerate(answers, start=1):
                for field in ("answer_display",):
                    for issue in _format_issues(
                        f"{row_label}: {field}_{position}",
                        str(answer.get(field) or ""),
                        allowed_codes=_STRICT_RICH_TEXT_CODES,
                    ):
                        flag(issue)
            for issue in _answer_format_issues(
                row_label,
                normalized_answers,
                normalized_sub_questions,
            ):
                flag(issue)
            marks = workbook_contract._readback_decimal(
                question.get("marks")
            )
            marking_validator = {
                "objective": workbook_contract._objective_marking_errors,
                "subjective": workbook_contract._subjective_marking_errors,
                "descriptive": workbook_contract._descriptive_marking_errors,
            }[kind]
            # Preserve the workbook's slot positions for contiguity checks,
            # while applying the same documented legacy answer-type aliases
            # that import will persist.  Passing the compact parsed lists
            # would hide gaps; passing the entirely raw row would make an
            # accepted alias such as ``Words`` fail the canonical enum gate.
            marking_row = dict(question)
            for number in sheet_layout.answer_block_numbers:
                field = f"answer_type_{number}"
                if field in marking_row:
                    marking_row[field] = normalize_answer_type(
                        str(marking_row.get(field) or "")
                    )
            for sub_number in sheet_layout.sub_question_numbers:
                for keyword_number in (
                    sheet_layout.sub_question_keyword_numbers(sub_number)
                ):
                    field = (
                        f"sq{sub_number}_answer_type_{keyword_number}"
                    )
                    if field in marking_row:
                        marking_row[field] = normalize_answer_type(
                            str(marking_row.get(field) or "")
                        )
            for issue in marking_validator(
                marking_row, label=row_label, marks=marks,
            ):
                flag(issue)
            keyboard = str(question.get("math_keyboard") or "")
            if kind == "objective" and keyboard:
                flag(
                    f"{row_label}: objective math_keyboard must be blank"
                )
            elif kind in {"subjective", "descriptive"} and keyboard not in {
                "Yes", "No",
            }:
                flag(
                    f"{row_label}: {kind} math_keyboard must be exactly "
                    "Yes or No"
                )
            for numeric_field in ("marks", "question_duration"):
                raw_numeric = question.get(numeric_field)
                numeric = workbook_contract._readback_decimal(raw_numeric)
                if numeric is None or numeric <= 0:
                    flag(
                        f"{row_label}: {numeric_field} must be finite and "
                        "positive"
                    )
            if kind == "objective":
                option_text = str(
                    question.get("question_text")
                    or question.get("question")
                    or ""
                )
                uppercase_labels = (
                    katex_rules.uppercase_objective_option_labels(
                        option_text,
                        len(sheet_layout.answer_block_numbers),
                    )
                )
                if uppercase_labels:
                    flag(
                        f"{row_label}: question_text uses uppercase "
                        "objective option label(s) "
                        f"{', '.join(uppercase_labels)}; use lowercase "
                        "labels (uppercase_objective_option_label)"
                    )
    return issues


def import_workbook(
    db: Session,
    path: Path,
    *,
    strict_content: bool = False,
) -> dict:
    """Import every content sheet; returns counts of created nodes + issues.

    Raises ``WorkbookLayoutError`` (mapped to 422 by ``api/data.py``) when the
    workbook matches no registered layout. The refusal happens before any
    ``db.add``, so an unidentified file imports nothing at all.

    ``strict_content=True`` is the public-upload contract: current mechanical
    Q21 defects raise ``WorkbookContentError`` before any query/mutation.
    Bootstrap and explicit migration callers keep the default so frozen
    legacy workbooks import with their existing review flags.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        identified = layouts.identify_workbook(_sheet_headers(wb))
    except WorkbookLayoutError:
        wb.close()
        raise
    # Public uploads use the current owner contract and fail closed before a
    # write. Internal bootstrap/migration callers deliberately keep the
    # historic flag-and-import mode so the frozen legacy database workbook
    # can be loaded and reviewed rather than making application startup
    # impossible. The API is the only untrusted persistence boundary.
    blocking_issues = (
        _blocking_content_issues(wb, identified)
        if strict_content else []
    )
    if blocking_issues:
        wb.close()
        raise WorkbookContentError(
            "workbook content validation failed before import: "
            + "; ".join(blocking_issues)
        )
    counts: dict = {"chapters": 0, "topics": 0, "concepts": 0, "groups": 0,
                    "questions": 0, "question_tags": 0, "issues": [],
                    "layout_id": identified.layout_id}

    def _flag(msg: str) -> None:
        if len(counts["issues"]) < _MAX_ISSUES:
            counts["issues"].append(msg)

    identity_notes: set[tuple[str, str, str]] = set()

    def _identity_flag(kind: str, name: str, code: str, msg: str) -> None:
        """One note per row per code, and a BUDGET of its own.

        A workbook repeats its topic on every line, so an un-deduped note
        would fill ``issues`` with one defect. The separate budget matters
        too: a wholly tagless legacy book raises two notes per distinct
        topic/concept name, and without a cap those notes would consume
        ``_MAX_ISSUES`` and silently truncate the placement and mangled-row
        issues raised later in the same row loop — the ledger S2 built.
        """
        if (kind, name, code) in identity_notes:
            return
        identity_notes.add((kind, name, code))
        if len(identity_notes) > _MAX_IDENTITY_NOTES:
            if len(identity_notes) == _MAX_IDENTITY_NOTES + 1:
                _flag(
                    f"more than {_MAX_IDENTITY_NOTES} rows carry no usable "
                    "machine id; the remaining identity notes are omitted so "
                    "the rest of the ledger survives "
                    "(identity_notes_truncated)")
            return
        _flag(msg)

    # Every machine id the database already holds, read ONCE. An id restored
    # out of a workbook cell must be unique or it is not an identity: two rows
    # sharing one id share one ``question_label``, and
    # ``assessment_release_service`` skips a repeated label with no flag (R4).
    claimed: dict[str, set[str]] = {
        "topic": {
            str(value or "").strip()
            for (value,) in db.query(models.Topic.machine_id).all()
            if str(value or "").strip()
        },
        "concept": {
            str(value or "").strip()
            for (value,) in db.query(models.Concept.machine_id).all()
            if str(value or "").strip()
        },
    }
    _ORDINAL_OF = {
        "topic": identity.minted_topic_ordinal,
        "concept": identity.minted_concept_ordinal,
    }

    def _usable_tag(tag: str, kind: str, name: str, stored: str) -> str:
        """The tag if it is an id THIS minter could have issued and is free.

        Two gates, both mechanics, both learned from a measurement.

        SHAPE. ``identity.title_tag`` only proves the tag round-trips
        ``strip_title_tag``; it does not prove the tag is an identity. The
        pre-S4 writer stamped ``directory.topic_tag``, which is CHAPTER-level
        — one tag on every topic of the chapter — and that is what the
        committed reference workbooks carry. [measured] restoring it unparsed
        gave the owner's own eight-topic ``grade6_science.xlsx`` six duplicate
        topic identities, with no issue recorded.

        UNIQUENESS. A minted-shape tag that another row already holds is not
        this row's identity either — importing one file against two chapters
        would otherwise hand both the same string.

        A rejected tag is not a refusal: the row imports, the note is
        recorded, and the id is minted on first use (T4-8, R4).
        """
        if not tag:
            return ""
        if len(tag) > _MAX_RESTORED_MACHINE_ID:
            _identity_flag(
                kind, name, "imported_without_machine_id",
                f"{kind} {name!r}: the title cell's tag is "
                f"{len(tag)} characters, more than a "
                f"{models.Question.question_label.type.length}-character "
                "question label can carry; the id will be minted on first "
                "use (imported_without_machine_id)")
            return ""
        if _ORDINAL_OF[kind](tag) is None:
            _identity_flag(
                kind, name, "imported_without_machine_id",
                f"{kind} {name!r}: the title cell's tag {tag!r} is not a "
                "machine id this pipeline mints; the id will be minted on "
                "first use (imported_without_machine_id)")
            return ""
        if tag != stored and tag in claimed[kind]:
            _identity_flag(
                kind, name, "machine_id_already_claimed",
                f"{kind} {name!r}: machine id {tag!r} is already held by "
                "another row; this row keeps its own and the id is minted on "
                "first use (machine_id_already_claimed)")
            return ""
        return tag

    def _claim(kind: str, tag: str) -> str:
        if tag:
            claimed[kind].add(tag)
        return tag

    def _restore_machine_id(row, tag: str, kind: str, name: str) -> None:
        """Keep the identity the writer exported instead of erasing it.

        T4-8. This endpoint STRIPPED the tag and threw it away, so the one
        authenticated POST S2 hardened was also the one that erased the
        identity S4 mints. ``identity.title_tag`` returns a tag only when it
        is exactly what ``bi.strip_title_tag`` removes, so a value restored
        here is one the writer could have written; anything else reads as
        tagless.

        A blank column is filled — that is the restore, and it overwrites
        nothing. A DIFFERENT persisted id is NEVER overwritten: §6:523's
        "stable forever" is a property of storage (P-C1), so an uploaded file
        may not re-key a published row. The disagreement is RECORDED, naming
        both, and the row still imports: this is an import, not an identity
        corruption, and refusing it would lose the file's content (R4).
        """
        stored = str(getattr(row, "machine_id", "") or "").strip()
        if stored:
            if tag and tag != stored:
                _identity_flag(
                    kind, name, "machine_id_conflict",
                    f"{kind} {name!r}: the workbook carries machine id "
                    f"{tag!r} but the stored row is {stored!r}; the stored id "
                    "stands (machine_id_conflict)")
            return
        usable = _usable_tag(tag, kind, name, stored)
        if usable:
            row.machine_id = _claim(kind, usable)
            return
        if not tag:
            _identity_flag(
                kind, name, "imported_without_machine_id",
                f"{kind} {name!r}: no machine id in the title cell; it "
                "will be minted on first use "
                "(imported_without_machine_id)")

    # A content sheet the layout declares but the file does not carry, and a
    # tab the layout declares non-content, are both RECORDED. Neither loses a
    # row -- an absent sheet has none, and an ignored tab carries none -- but
    # the defect this slice replaced was never the drop, it was the SILENCE:
    # the old reader skipped an unmatched sheet with a bare ``continue`` and
    # no flag, so a whole missing Objective sheet read as a clean import. A
    # reviewer must be able to tell "this file has no Subjective questions"
    # from "this file's Subjective sheet went missing" (R4).
    for missing in sorted(
        set(identified.layout.sheets) - {f.kind for f in identified.sheets}
    ):
        _flag(f"{identified.layout.sheet(missing).sheet_name}: sheet not present "
              f"in this workbook; nothing was imported for it")
    for ignored in identified.ignored_sheets:
        _flag(f"{ignored}: tab is not a content sheet in layout "
              f"{identified.layout_id!r}; it was read past, not imported")

    # Caches keyed by natural keys for de-duplication within one import.
    chapters: dict[str, models.Chapter] = {}
    topics: dict[tuple, models.Topic] = {}
    concepts: dict[tuple, models.Concept] = {}
    groups: dict[tuple, models.Group] = {}
    seen_labels: set[str] = {
        q.question_label for q in db.query(models.Question).all() if q.question_label
    }
    # Questions created or matched during THIS import, by label, so repeated
    # rows reconstruct placements against the right entity.
    label_questions: dict[str, models.Question] = {}

    # Cache of existing question texts per chapter, for cross-book dedupe.
    qtext_cache: dict[int, dict[str, models.Question]] = {}

    def _chapter_qtexts(chapter_id: int) -> dict[str, models.Question]:
        if chapter_id not in qtext_cache:
            qtext_cache[chapter_id] = {
                normalize_question_text(qq.question): qq
                for qq in (
                    db.query(models.Question)
                    .join(models.Group).join(models.Concept).join(models.Topic)
                    .filter(models.Topic.chapter_id == chapter_id)
                )
                if qq.question
            }
        return qtext_cache[chapter_id]

    for found in identified.sheets:
        kind = found.kind
        sheet_name = found.sheet_name
        sheet_layout = found.sheet
        ws = wb[sheet_name]

        for row_i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if row is None or not any(row):
                continue
            chap = sheet_layout.block_values(row, "chapter")
            top = sheet_layout.block_values(row, "topic")
            con = sheet_layout.block_values(row, "concept")
            grp = sheet_layout.block_values(row, "group")

            if not chap.get("chapter_title"):
                _flag(f"{sheet_name!r} row {row_i}: skipped — missing chapter_title")
                continue

            # ---- Chapter ----
            # Derive board/grade/subject/code from the RAW (tag-bearing) cells
            # first; the topic/concept tags carry the ID code prefix.
            meta = directory.derive_chapter_meta(
                chap["chapter_title"], chap.get("chapter_display_name", ""),
                top.get("topic_title", ""), top.get("topic_display_name", ""),
                top.get("topic_concept_labels", ""), top.get("related_topics", ""),
                con.get("concept_title", ""), con.get("concept_display_name", ""),
                chap.get("post_topics", ""), chap.get("pre_topics", ""),
            )
            chapter_code = meta["chapter_code"]
            # Recover CLEAN titles (strip embedded tags / "Topic NN:" prefix).
            chap_title = strip_title_tag(chap["chapter_title"])
            t_title_clean = strip_topic_title(top.get("topic_title", ""))
            c_title_clean = strip_title_tag(con.get("concept_title", ""))
            # T6.2: the chapter code truncates ('The Rise of Nationalism in
            # Europe' and '... in Asia' both derive 10CBSS_TheRiseOfNat), so
            # the code alone silently MERGED two chapters. The identity
            # carries the title as well; a code collision with a different
            # title is recorded and the rows are not merged.
            chap_title_key = normalize_question_text(
                chap_title or chap["chapter_title"])
            ch_key = (chapter_code, chap_title_key)
            chapter = chapters.get(ch_key)
            if chapter is None:
                same_code = db.query(models.Chapter).filter_by(
                    chapter_code=chapter_code).all()
                chapter = next(
                    (
                        row_chapter for row_chapter in same_code
                        if normalize_question_text(strip_title_tag(
                            row_chapter.chapter_title)) == chap_title_key
                    ),
                    None,
                )
                if chapter is None and same_code:
                    _flag(
                        f"chapter code {chapter_code!r} is already used by "
                        f"{same_code[0].chapter_title!r}; "
                        f"{(chap_title or chap['chapter_title'])!r} imports as "
                        "a separate chapter (chapter_code_collision)"
                    )
            if chapter is None:
                chapter = models.Chapter(
                    chapter_code=chapter_code,
                    board=meta["board"], grade=meta["grade"],
                    subject=meta["subject"], unit=meta["unit"],
                    chapter_title=chap_title or chap["chapter_title"],
                    chapter_display_name=chap.get("chapter_display_name", ""),
                    chapter_duration=chap.get("chapter_duration", ""),
                    pre_topics=chap.get("pre_topics", ""),
                    post_topics=chap.get("post_topics", ""),
                    chapter_description=chap.get("chapter_description", ""),
                )
                db.add(chapter)
                db.flush()
                counts["chapters"] += 1
            chapters[ch_key] = chapter

            # ---- Topic ----
            # T6.4: the identity carries the LANE and goes through the one
            # shared normaliser. Without the lane a Pre topic and a Post topic
            # sharing a title merged into whichever row was created first and
            # the other lane's concepts were re-parented under it, silently.
            # The stored ``topic_title`` is NEVER rewritten to the incoming
            # casing (T4-6: exact-match-plus-lenient-write is what created the
            # UploadRefused class).
            t_title = t_title_clean or "Topic 01"
            t_lane = top.get("pre_post_learning", "") or "Post"
            t_tag = identity.title_tag(top.get("topic_title", ""))
            t_ident = identity.topic_identity(t_title)
            t_key = (chapter.id, t_ident, t_lane)
            topic = topics.get(t_key)
            if topic is None:
                topic = next(
                    (
                        row_topic for row_topic in db.query(models.Topic)
                        .filter_by(chapter_id=chapter.id,
                                   pre_post_learning=t_lane).all()
                        if identity.topic_identity(row_topic.topic_title)
                        == t_ident
                    ),
                    None,
                )
            if topic is None:
                other_lane = next(
                    (
                        row_topic for row_topic in db.query(models.Topic)
                        .filter_by(chapter_id=chapter.id).all()
                        if identity.topic_identity(row_topic.topic_title)
                        == t_ident and row_topic.pre_post_learning != t_lane
                    ),
                    None,
                )
                if other_lane is not None:
                    _flag(
                        f"{t_title!r}: a {other_lane.pre_post_learning!r} topic "
                        f"already carries this title; the {t_lane!r} row "
                        "imports as a separate topic (topic_lane_conflict)"
                    )
                topic = models.Topic(
                    chapter_id=chapter.id, topic_title=t_title,
                    topic_display_name=top.get("topic_display_name", ""),
                    pre_post_learning=t_lane,
                    related_topics=top.get("related_topics", ""),
                    topic_description=top.get("topic_description", ""),
                    machine_id=_claim(
                        "topic", _usable_tag(t_tag, "topic", t_title, "")),
                )
                db.add(topic)
                db.flush()
                counts["topics"] += 1
                if not t_tag:
                    _identity_flag(
                        "topic", t_title, "imported_without_machine_id",
                        f"topic {t_title!r}: no machine id in the title cell; "
                        "it will be minted on first use "
                        "(imported_without_machine_id)")
            else:
                _restore_machine_id(topic, t_tag, "topic", t_title)
            topics[t_key] = topic

            # ---- Concept ----
            c_title = c_title_clean or "Concept"
            c_tag = identity.title_tag(con.get("concept_title", ""))
            c_source = con.get("concept_source", "")
            c_key = (topic.id, c_title)
            concept = concepts.get(c_key)
            if concept is None:
                concept = db.query(models.Concept).filter_by(
                    topic_id=topic.id, concept_title=c_title).first()
            if concept is None:
                concept_details = katex_rules.canonicalize_rich_text(
                    con.get("concept_details", ""))
                for issue in _format_issues(
                    f"{sheet_name!r} row {row_i} concept_details",
                    concept_details,
                ):
                    _flag(issue)
                concept = models.Concept(
                    topic_id=topic.id, concept_title=c_title,
                    concept_display_name=con.get("concept_display_name", ""),
                    parent_concept=con.get("parent_concept", ""),
                    concept_details=concept_details,
                    keywords=con.get("keywords", ""),
                    digicards=con.get("digicards", ""),
                    related_concepts=con.get("related_concepts", ""),
                    sources=c_source,
                    machine_id=_claim(
                        "concept", _usable_tag(c_tag, "concept", c_title, "")),
                )
                db.add(concept)
                db.flush()
                counts["concepts"] += 1
                if not c_tag:
                    _identity_flag(
                        "concept", c_title, "imported_without_machine_id",
                        f"concept {c_title!r}: no machine id in the title "
                        "cell; it will be minted on first use "
                        "(imported_without_machine_id)")
            else:
                _restore_machine_id(concept, c_tag, "concept", c_title)
                if c_source:
                    # Same concept from another book: accumulate sources.
                    concept.sources = merge_sources(concept.sources, c_source)
            if con.get("parent_concept") and not concept.parent_concept:
                concept.parent_concept = con.get("parent_concept", "")
            concepts[c_key] = concept

            # ---- Question band (parsed before the Group so a pure Concept
            # row can be recognized without minting a default group) ----
            # Names AND values now come from the same identified layout. They
            # used to decouple by construction: the names were taken from the
            # canonical FIELDS_BY_KIND while the values were sliced from the
            # detected geometry.
            qd = sheet_layout.block_values(row, "question")
            label = qd.get("question_label", "")
            has_question = bool(label or qd.get("question"))

            # ---- Group ----
            # A Group exists only when the row carries meaningful Group
            # identity or a populated Question band (spec §9): a pure Concept
            # row must not mint a default Basic group.
            has_group_identity = any(
                grp.get(field, "").strip()
                for field in _GROUP_IDENTITY_FIELDS
            )
            if not (has_group_identity or has_question):
                continue  # hierarchy-only row: chapter/topic/concept created
            g_type = grp.get("group_type") or "Basic"
            g_name = grp.get("group_name") or grp.get("group_display_name") or f"{g_type} Group"
            g_key = (concept.id, g_type, g_name)
            group = groups.get(g_key)
            if group is None:
                group = db.query(models.Group).filter_by(
                    concept_id=concept.id, group_type=g_type, group_name=g_name).first()
            if group is None:
                group = models.Group(
                    concept_id=concept.id, group_type=g_type, group_name=g_name,
                    group_display_name=grp.get("group_display_name", ""),
                    group_description=grp.get("group_description", ""),
                    group_status=grp.get("group_status", "Active"),
                    related_digicards=grp.get("related_digicards", ""),
                    group_key=g_name,
                    group_sequence=_group_sequence(g_name),
                )
                db.add(group)
                db.flush()
                counts["groups"] += 1
            elif not group.group_key:
                group.group_key = g_name
                group.group_sequence = _group_sequence(g_name)
            groups[g_key] = group

            # ---- Question ----
            if not has_question:
                continue
            if label and label in seen_labels:
                # One label = one Question, many placements. A repeated label
                # reconstructs a placement edge instead of being dropped, so
                # an export re-imported into an empty DB rebuilds every
                # QuestionTag (spec §7.5).
                existing_q = label_questions.get(label)
                if existing_q is None:
                    existing_q = db.query(models.Question).filter_by(
                        question_label=label).first()
                if existing_q is None:
                    _flag(f"{label}: repeated label has no importable "
                          "original question — row skipped")
                    continue
                repeat_norm = normalize_question_text(qd.get("question", ""))
                if repeat_norm and repeat_norm != normalize_question_text(
                        existing_q.question):
                    _flag(f"{label}: conflicting question content under one "
                          "label — row rejected")
                    continue
                if existing_q.group_id == group.id:
                    continue  # exact duplicate of the home placement
                tag_exists = db.query(models.QuestionTag).filter_by(
                    question_id=existing_q.id, group_id=group.id).first()
                if tag_exists is None:
                    db.add(models.QuestionTag(
                        question_id=existing_q.id, group_id=group.id))
                    counts["question_tags"] += 1
                continue
            # Cross-book duplicate check: same question text under the same
            # chapter (any label) is not re-added — its sources merge instead.
            norm = normalize_question_text(qd.get("question", ""))
            if norm:
                existing_q = _chapter_qtexts(chapter.id).get(norm)
                if existing_q is not None:
                    existing_q.question_source = merge_sources(
                        existing_q.question_source, qd.get("question_source", ""))
                    counts["question_sources_merged"] = counts.get(
                        "question_sources_merged", 0) + 1
                    seen_labels.add(label)
                    if label:
                        label_questions[label] = existing_q
                    continue
            seen_labels.add(label)

            answers, sub_questions = _parse_answers(qd, kind, sheet_layout)
            try:
                marks = float(qd.get("marks") or 0)
            except ValueError:
                marks = 0.0
                _flag(f"{label or 'row ' + str(row_i)}: marks not numeric "
                      f"({qd.get('marks')!r})")
            try:
                duration = float(qd.get("question_duration"))
                if not math.isfinite(duration) or duration <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                duration = 0.0
                _flag(
                    f"{label or 'row ' + str(row_i)}: question_duration "
                    "must be finite and positive"
                )

            # ---- Normalization to standard values ----
            skills = normalize_cognitive_skills(qd.get("cognitive_skills", ""))
            for part in split_multi(skills):
                if part not in COGNITIVE_SKILLS:
                    _flag(f"{label}: unknown cognitive skill {part!r}")
            difficulty = normalize_difficulty(qd.get("level_of_difficulty", ""))
            if difficulty and difficulty not in DIFFICULTY_LEVELS:
                _flag(f"{label}: unknown level_of_difficulty {difficulty!r}")
            appears = normalize_appears_in(qd.get("question_appears_in", ""))
            for a in answers:
                a["answer_type"] = normalize_answer_type(a.get("answer_type", ""))
                if a["answer_type"] and a["answer_type"] not in ANSWER_TYPES:
                    _flag(f"{label}: unknown answer_type {a['answer_type']!r}")
                content_field = (
                    "answer_content"
                    if "answer_content" in a
                    else "answer" if "answer" in a else ""
                )
                if content_field:
                    a[content_field] = katex_rules.raw_answer_cell(
                        a.get("answer_type", ""),
                        a.get(content_field, ""),
                    )
                    if (
                        kind == "objective"
                        and release_contract.option_content_has_label(
                            a.get(content_field)
                        )
                    ):
                        _flag(
                            f"{label}: objective option repeats its letter "
                            "label"
                        )
                    if (
                        kind == "descriptive"
                        and release_contract.malformed_rubric_tag(
                            a.get(content_field), a.get("answer_type"),
                        )
                    ):
                        _flag(
                            f"{label}: descriptive rubric does not start "
                            "with an allowed functional tag or is without "
                            "its required colon"
                        )
            for subquestion in sub_questions:
                for keyword in subquestion.get("keywords") or []:
                    keyword["answer_type"] = normalize_answer_type(
                        keyword.get("answer_type", "")
                    )
                    if (
                        keyword["answer_type"]
                        and keyword["answer_type"] not in ANSWER_TYPES
                    ):
                        _flag(
                            f"{label}: unknown keyword answer_type "
                            f"{keyword['answer_type']!r}"
                        )
                    keyword["keyword"] = katex_rules.raw_answer_cell(
                        keyword.get("answer_type", ""),
                        keyword.get("keyword", ""),
                    )
                    if (
                        kind == "descriptive"
                        and release_contract.malformed_rubric_tag(
                            keyword.get("keyword"),
                            keyword.get("answer_type"),
                        )
                    ):
                        _flag(
                            f"{label}: descriptive subquestion keyword does "
                            "not start with an allowed functional tag or is "
                            "without its required colon"
                        )

            # ---- Validation: weightage sum vs marks; content formats ----
            if (
                kind in {"subjective", "descriptive"}
                and marks
                and not (kind == "descriptive" and sub_questions)
            ):
                total = _weightage_sum(answers, kind)
                if total is not None and abs(total - marks) > 0.01:
                    _flag(f"{label}: answer weightage sum {total:g} != marks {marks:g}")
            if kind == "descriptive" and answers and sub_questions:
                _flag(
                    f"{label}: multipart descriptive duplicates scoring in "
                    "main answer/rubric blocks"
                )
            if (
                kind == "descriptive"
                and marks == 4
                and not sub_questions
                and len(answers) < 2
            ):
                _flag(
                    f"{label}: 4-mark descriptive requires at least two "
                    "answer/rubric blocks"
                )
            rich_text_values = _rich_text_for_validation(
                kind,
                answers,
                qd.get("question", ""),
                qd.get("question_text", ""),
                qd.get("display_answer", ""),
                qd.get("answer_explanation", ""),
                *(str(subquestion.get("text") or "")
                  for subquestion in sub_questions),
            )
            for issue in _format_issues(
                label or f"row {row_i}",
                *rich_text_values,
            ):
                _flag(issue)
            for issue in _answer_format_issues(
                label or f"row {row_i}", answers, sub_questions,
            ):
                _flag(issue)

            # ---- question_text: parse if present, else backfill (plain text) ----
            question_text = qd.get("question_text", "").strip()
            if not question_text and qd.get("question"):
                question_text = to_plain_text(qd.get("question", ""))

            new_q = models.Question(
                group_id=group.id, sheet_kind=kind, question_label=label,
                question_category=qd.get("question_category", ""),
                cognitive_skills=skills,
                question_source=qd.get("question_source", ""),
                question_disclaimer=qd.get("question_disclaimer", ""),
                question_duration=duration,
                math_keyboard=qd.get("math_keyboard", ""),
                # T6.3: the READER's default is the one its own layout
                # implies, never the active school profile's wire value —
                # sourcing it from the profile would stamp one school's
                # convention onto every other school's imported rows.
                question_appears_in=appears or APPEARS_IN_ALL,
                answer_restriction=qd.get("answer_restriction", ""),
                level_of_difficulty=difficulty,
                question=qd.get("question", ""),
                question_text=question_text,
                marks=marks,
                display_answer=qd.get("display_answer", ""),
                answer_explanation=qd.get("answer_explanation", ""),
                answers=answers, sub_questions=sub_questions, origin="seed",
            )
            db.add(new_q)
            db.flush()
            if norm:
                _chapter_qtexts(chapter.id)[norm] = new_q
            if label:
                label_questions[label] = new_q
            counts["questions"] += 1

    db.commit()
    wb.close()
    return counts
