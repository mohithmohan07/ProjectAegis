"""Write normalized questions back to the canonical Bulk Import workbook.

Two header rows are emitted per content sheet (section bands + field names).
Writes are **append-only**: ``append_questions`` reads existing
``question_label`` values across all tabs and skips anything already present,
so re-running a generation never overwrites or deletes prior rows.
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from . import (
    CHAPTER_FIELDS, TOPIC_FIELDS, FIELDS_BY_KIND, SHEET_BY_KIND,
    SECTION_BANDS, GROUP_FIELDS_BY_KIND,
    merge_sources, normalize_question_text, strip_title_tag, strip_topic_title,
)
from . import layouts
from . import workbook_sync
from .. import models
from ..services import directory, identity

_BAND_FILL = {
    "Chapter": "FCE4D6", "Topic": "FFF2CC", "Concept": "D9EAD3",
    "Group": "D0E0E3", "Question": "CFE2F3",
}


def _target_sheet(kind: str) -> layouts.SheetLayout:
    """The sheet of the TARGET layout this writer emits (spec-step8 S7)."""
    return layouts.sheet(layouts.REFERENCE_LAYOUT_ID, kind)


def _group_fields(kind: str) -> list[str]:
    return list(GROUP_FIELDS_BY_KIND[kind])


# Positional indices shared by every content sheet (the Chapter and Topic
# bands are identical on all three sheets of every registered layout; the
# schema gate in tests/test_bulk_import_schema.py pins that).
_IDX_CHAPTER_TITLE = 0
_IDX_TOPIC_TITLE = len(CHAPTER_FIELDS)
_IDX_TOPIC_PRE_POST = _IDX_TOPIC_TITLE + TOPIC_FIELDS.index(
    "pre_post_learning")
_IDX_CONCEPT_TITLE = len(CHAPTER_FIELDS) + len(TOPIC_FIELDS)


def _q_start(kind: str) -> int:
    """Column index where the Question band's first ``question_label`` lives.

    Read off the layout, never summed from constants: the Concept band is per
    sheet now (Descriptive carries no ``concept_source``) and the reference
    Objective sheet leaves column 23 unbanded, so a sum of shared spans lands
    on the wrong column.
    """
    return _target_sheet(kind).block_start("question")


def _identify(sheet_name: str, header_row: tuple) -> layouts.SheetIdentity | None:
    """Identify one sheet through the shared registry (spec-step8 T6).

    The writer's own header-shape detection is gone: it was the twin of
    ``reader._concept_fields``, two functions deciding the same column
    geometry of the same file under the same ``output_workbook_lock()``.
    """
    return layouts.identify_sheet(sheet_name, header_row)


def _identified_sheet(
    sheet_name: str, header_row: tuple, kind: str = "objective", *,
    mismatches: list[dict] | None = None,
) -> layouts.SheetLayout:
    """The identified layout of one sheet, or the target layout's own.

    An unidentified sheet is a RECORDED mismatch, never a raise: the writer is
    reached mid-run with the model budget already spent, so T6's fail-closed
    asymmetry stops at the reader. The fallback is the layout this writer
    itself emits.
    """
    found = _identify(sheet_name, header_row)
    if found is not None:
        return found.sheet
    if mismatches is not None:
        mismatches.append({
            "code": "sheet_layout_unidentified",
            "sheet": sheet_name,
            "column_count": len([h for h in header_row if h is not None]),
            "detail": (
                "no registered layout matches this sheet's header row; the "
                "writer's own layout was assumed"
            ),
        })
    return _target_sheet(kind)


def _sheet_name_for(wb, index: "WorkbookIndex", kind: str) -> str:
    """The sheet in ``wb`` that carries ``kind``, by IDENTITY not by name.

    Mechanics for reading older files, not a second layout: the canonical
    layouts spell the Objective sheet ``'Objective '`` (trailing space) and the
    authority spells it ``'Objective'``. Falling back to the target name is
    right for a workbook this writer just created.
    """
    for sheet_name, meta in index.sheet_meta.items():
        if meta.get("kind") == kind and sheet_name in wb.sheetnames:
            return sheet_name
    target = SHEET_BY_KIND[kind]
    if target in wb.sheetnames:
        return target
    for entry in layouts.LAYOUTS.values():
        candidate = entry.sheets.get(kind)
        if candidate is not None and candidate.sheet_name in wb.sheetnames:
            return candidate.sheet_name
    return target


# --------------------------------------------------------------------------- #
# T3.8 — a row that does not fit its sheet is RECORDED, never silently repaired
#
# ``_question_to_row`` and ``_concept_to_row`` used to end
# ``row += [""] * (expected - len(row))`` / ``return row[:expected]``: a row
# wider than its sheet lost its tail with no record and a row narrower gained
# blanks with no record, on the export path R5 calls the database of record.
# Since S7 moves ``expected`` on every sheet, that silence is exactly what
# would make a half-migrated file look green.
#
# The width comparison itself is mechanics — ``expected`` is the layout's own
# column count, not a judgment about content. Its CONSEQUENCE follows Q13: one
# recorded, flagged, content-addressed best-judgment decision naming the sheet
# kind, the expected width, the actual width, the row's identity and every
# value that had no column to land in, and the run completes. Never a raise
# (``append_concepts`` is reached mid-run with the model budget already spent),
# never a silent slice.
# --------------------------------------------------------------------------- #

ROW_WIDTH_MISMATCH = "bulk_import_row_width_mismatch"


def _fixer_decision(kind: str, *, detail: str, context: dict) -> dict:
    """One recorded, flagged, content-addressed decision (CLAUDE.md Q13)."""
    from ..services.phase3.fixer import FIXER_POLICY_VERSION

    body = json.dumps(
        {"kind": kind, "detail": detail, "context": context},
        sort_keys=True, ensure_ascii=False, default=str,
    )
    decision = {
        "code": kind,
        "policy_version": FIXER_POLICY_VERSION,
        "decision_sha256": hashlib.sha256(body.encode("utf-8")).hexdigest(),
        "review_flag": True,
        "detail": detail,
    }
    decision.update(context)
    return decision


def _fit_row_to_sheet(
    row: list, expected: int, *, kind: str, row_identity: str,
    decisions: list[dict] | None = None,
) -> list:
    """Return ``row`` at the sheet's width, recording any mismatch."""
    if len(row) == expected:
        return row
    overflow = [
        "" if value is None else str(value)
        for value in row[expected:]
        if str(value or "").strip()
    ]
    decision = _fixer_decision(
        ROW_WIDTH_MISMATCH,
        detail=(
            f"a {kind} row was composed {len(row)} cells wide for a sheet "
            f"{expected} cells wide; the row ships at the sheet's width and "
            "every value with no column is recorded here"
        ),
        context={
            "sheet_kind": kind,
            "expected_width": expected,
            "actual_width": len(row),
            "row_identity": row_identity,
            "overflow_values": overflow,
        },
    )
    if decisions is not None:
        decisions.append(decision)
    # The whole decision, not just its prose: the export paths that build a
    # fresh workbook return bytes and have no decisions list to collect into,
    # and on those a partial log line would be the ONLY record — losing the
    # row identity and the overflow values, which is the half of T3.8 that
    # matters. Carrying the full record here keeps the log a complete one.
    logging.getLogger(__name__).warning(
        "bulk-import %s: %s", ROW_WIDTH_MISMATCH,
        json.dumps(decision, sort_keys=True, ensure_ascii=False, default=str),
    )
    if len(row) < expected:
        return row + [""] * (expected - len(row))
    return row[:expected]


def _cell_str(row: tuple, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


# Control characters (except tab/newline) are illegal in xlsx cells; OCR'd
# content occasionally smuggles one in (e.g. a mangled degree sign). openpyxl
# raises IllegalCharacterError on write, so every outgoing value is sanitized.
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
EXCEL_CELL_CHARACTER_LIMIT = 32_767


class ExcelCellLimitError(ValueError):
    """A workbook value cannot be exported losslessly in one Excel cell."""


def _safe_cell(value):
    if isinstance(value, str):
        return _ILLEGAL_XLSX_RE.sub("", value)
    return value


def _validate_cell_length(cell, value) -> None:
    """Reject text that OpenPyXL would otherwise silently truncate."""
    if not isinstance(value, str) or len(value) <= EXCEL_CELL_CHARACTER_LIMIT:
        return
    excess = len(value) - EXCEL_CELL_CHARACTER_LIMIT
    worksheet = getattr(cell, "parent", None)
    sheet_name = getattr(worksheet, "title", "unknown")
    coordinate = getattr(cell, "coordinate", "unknown")
    field_name = ""
    if worksheet is not None and getattr(cell, "row", 0) > 2:
        header = worksheet.cell(row=2, column=cell.column).value
        if header:
            field_name = f" (field {str(header)!r})"
    raise ExcelCellLimitError(
        "Excel export blocked to prevent data loss: "
        f"cell {sheet_name!r}!{coordinate}{field_name} contains "
        f"{len(value):,} characters, "
        f"exceeding Excel's {EXCEL_CELL_CHARACTER_LIMIT:,}-character cell "
        f"limit by {excess:,}. Shorten this value or split it across multiple "
        "cells or records, then export again."
    )


def _set_cell_value(cell, value) -> None:
    """Write untrusted text without allowing Excel formula interpretation.

    Setting ``data_type`` explicitly preserves the exact displayed and
    round-tripped value. Prefixing an apostrophe is intentionally avoided for
    XLSX because it would change the stored content.
    """
    safe = _safe_cell(value)
    _validate_cell_length(cell, safe)
    cell.value = safe
    if isinstance(safe, str) and safe.startswith(_FORMULA_PREFIXES):
        cell.data_type = "s"


def _write_cell(ws, *, row: int, column: int, value) -> None:
    _set_cell_value(ws.cell(row=row, column=column), value)


def question_placement_key(label: str, group: models.Group) -> tuple:
    """Identity + ancestor path for one assessment placement.

    Matches the CMS dedupe unit: a repeat of this exact tuple is a duplicate
    (skip), the same ``label`` under a different tuple is a tag. Stable group
    identity is part of the key (spec §7.5): without it, two groups of the
    same tier under one concept collapse into one placement.
    """
    concept = group.concept
    topic = concept.topic
    chapter = topic.chapter
    return (label, chapter.chapter_title, topic.topic_title,
            concept.concept_title, group.group_type,
            group.group_key or group.group_name)


def visible_question_placement_key(
    label: str, group: models.Group,
) -> tuple:
    """Legacy workbook identity when no internal-key column is available.

    The canonical workbook has only Q12's friendly ``group_name`` slot.  The
    legacy Build Assessments append lane therefore needs this read-back alias
    after its internal ``group_key`` and visible name are separated.  MES
    releases keep using their snapshot ledger and never depend on this alias.
    """
    return (*question_placement_key(label, group)[:-1], group.group_name)


def concept_placement_key(concept: models.Concept, topic: models.Topic) -> tuple:
    """Normalized identity + ancestor path for one concept placement.

    Concept generation may improve title capitalization or whitespace on a
    later pass. Those presentation-only changes must refresh the original row,
    not create a second CMS placement. Learning kind is part of the path
    because one chapter may intentionally teach the same normalized concept in
    distinct Pre and Post topic bands.
    """
    chapter = topic.chapter
    return (
        normalize_question_text(strip_title_tag(concept.concept_title)),
        normalize_question_text(strip_title_tag(chapter.chapter_title)),
        normalize_question_text(strip_topic_title(topic.topic_title)),
        normalize_question_text(topic.pre_post_learning),
    )


def _row_question_placement_key(
    row: tuple, sheet_layout: layouts.SheetLayout,
) -> tuple | None:
    """Placement identity of one workbook row, addressed BY NAME."""
    qs = sheet_layout.column("question", "question_label")
    g_type = sheet_layout.column("group", "group_type")
    g_name = sheet_layout.column("group", "group_name")
    if qs is None or g_type is None or g_name is None:
        return None
    label = _cell_str(row, qs)
    if not label:
        return None
    # Strip the title-column tags so keys match the clean DB-derived keys.
    return (label,
            strip_title_tag(_cell_str(row, _IDX_CHAPTER_TITLE)),
            strip_topic_title(_cell_str(row, _IDX_TOPIC_TITLE)),
            strip_title_tag(_cell_str(row, _IDX_CONCEPT_TITLE)),
            _cell_str(row, g_type),
            _cell_str(row, g_name))


def _row_concept_placement_key(row: tuple) -> tuple | None:
    title = strip_title_tag(_cell_str(row, _IDX_CONCEPT_TITLE))
    if not title:
        return None
    return (
        normalize_question_text(title),
        normalize_question_text(
            strip_title_tag(_cell_str(row, _IDX_CHAPTER_TITLE))),
        normalize_question_text(
            strip_topic_title(_cell_str(row, _IDX_TOPIC_TITLE))),
        normalize_question_text(_cell_str(row, _IDX_TOPIC_PRE_POST)),
    )


class WorkbookIndex:
    """A scan of what already exists in a workbook, for placement-aware writes.

    - ``q_placements`` / ``c_placements``: exact (identity, placement) tuples present.
    - ``labels`` / ``concept_titles``: entity identities present anywhere (used to
      classify a new placement as a *tag* vs a brand-new *add*).
    - ``q_rows``: placement key -> (sheet, row) for in-place source merges.
    - ``concept_rows``: normalized placement key -> [(sheet, row), ...] —
      exact-placement rows carrying that concept's band, for full refreshes.
    - ``sheet_meta``: per-sheet column geometry (legacy vs current layout).
    """

    __slots__ = ("q_placements", "labels", "c_placements", "concept_titles",
                 "q_rows", "concept_rows", "sheet_meta", "mismatches")

    def __init__(self) -> None:
        self.q_placements: set[tuple] = set()
        self.labels: set[str] = set()
        self.c_placements: set[tuple] = set()
        self.concept_titles: set[str] = set()
        self.q_rows: dict[tuple, tuple] = {}
        self.concept_rows: dict[tuple, list[tuple]] = {}
        self.sheet_meta: dict[str, dict] = {}
        # Sheets whose header row matched no registered layout. Recorded, not
        # skipped in silence: the unflagged ``continue`` here was the writer's
        # twin of the reader hole S2 closes.
        self.mismatches: list[dict] = []


def scan_workbook(path: Path) -> WorkbookIndex:
    idx = WorkbookIndex()
    if not path.exists():
        return idx
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
        found = _identify(sheet_name, header)
        if found is None:
            # A sheet with no header row at all carries no columns to place
            # anything in (the Doc Link tab). A sheet that HAS a header and
            # still matches nothing is the recorded mismatch.
            if any(cell is not None for cell in header):
                idx.mismatches.append({
                    "code": "sheet_layout_unidentified",
                    "sheet": sheet_name,
                    "column_count": len(
                        [cell for cell in header if cell is not None]),
                    "detail": (
                        "no registered layout matches this sheet's header "
                        "row; it was not scanned for existing placements"
                    ),
                })
            continue
        sheet_layout = found.sheet
        concept_fields = list(sheet_layout.block_fields("concept"))
        concept_start = sheet_layout.block_start("concept")
        idx.sheet_meta[sheet_name] = {
            "layout_id": found.layout_id,
            "kind": found.kind,
            "concept_len": len(concept_fields),
            "concept_fields": concept_fields,
            "q_start": sheet_layout.block_start("question"),
            "q_src_col": sheet_layout.column("question", "question_source"),
            # concept_source only exists in some layouts.
            "c_src_col": sheet_layout.column("concept", "concept_source"),
            "parent_col": sheet_layout.column("concept", "parent_concept"),
        }
        idx.sheet_meta[sheet_name]["concept_start"] = concept_start
        for row_i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not any(row):
                continue
            qk = _row_question_placement_key(row, sheet_layout)
            if qk:
                idx.q_placements.add(qk)
                idx.labels.add(qk[0])
                idx.q_rows[qk] = (sheet_name, row_i)
            ck = _row_concept_placement_key(row)
            if ck:
                idx.c_placements.add(ck)
                idx.concept_titles.add(ck[0])
                idx.concept_rows.setdefault(ck, []).append((sheet_name, row_i))
    wb.close()
    return idx


def _question_placements(q: models.Question) -> list[models.Group]:
    """Authoring home + every tagged group, de-duplicated, order-stable."""
    groups = [q.group] + [t.group for t in q.tags]
    seen: set[int] = set()
    out: list[models.Group] = []
    for g in groups:
        if g is not None and g.id not in seen:
            seen.add(g.id)
            out.append(g)
    return out


def _concept_placements(c: models.Concept) -> list[models.Topic]:
    """Authoring home topic + every tagged topic, de-duplicated, order-stable."""
    topics = [c.topic] + [t.topic for t in c.tags]
    seen: set[int] = set()
    out: list[models.Topic] = []
    for t in topics:
        if t is not None and t.id not in seen:
            seen.add(t.id)
            out.append(t)
    return out


class ConceptExportScope:
    """The exact topology selected for one concept workbook/export.

    A scoped Build Concepts download must not derive labels, topic numbers, or
    chapter topic lists from unrelated historical rows still present in the
    database. The accepted concept ids are the closed-world export topology.
    """

    def __init__(self, concepts: list[models.Concept]) -> None:
        by_topic: dict[int, dict[int, models.Concept]] = defaultdict(dict)
        topics: dict[int, models.Topic] = {}
        for concept in concepts:
            for topic in _concept_placements(concept):
                topics[topic.id] = topic
                by_topic[topic.id][concept.id] = concept

        self.concepts_by_topic = {
            topic_id: sorted(values.values(), key=identity.source_order_key)
            for topic_id, values in by_topic.items()
        }
        grouped_topics: dict[tuple[int, str], list[models.Topic]] = defaultdict(list)
        for topic in topics.values():
            grouped_topics[
                (topic.chapter_id, (topic.pre_post_learning or "").casefold())
            ].append(topic)
        self.topics_by_chapter_kind = {
            key: sorted(values, key=identity.source_order_key)
            for key, values in grouped_topics.items()
        }
        self.topic_numbers = {
            topic.id: position
            for topic_group in self.topics_by_chapter_kind.values()
            for position, topic in enumerate(topic_group, start=1)
        }

    def concepts_for(self, topic: models.Topic) -> list[models.Concept]:
        return list(self.concepts_by_topic.get(topic.id, ()))

    def topics_for(
        self, chapter: models.Chapter, learning_kind: str,
    ) -> list[models.Topic]:
        return list(self.topics_by_chapter_kind.get(
            (chapter.id, (learning_kind or "").casefold()), ()))


# T10-4 (S11): the read-back's finding is RECORDED, never a raise. The old
# ``ConceptWorkbookValidationError`` had exactly two repo hits — its
# definition and its raise — nothing caught it and no test pinned it, while
# it gated both the mid-run deposit (after the model budget was spent) and
# the explicit publication. The identity comparison below is a gate on an
# artifact and stays exactly where it is (R5: the workbook is the database);
# what changed is the CONSEQUENCE — one recorded, flagged, content-addressed
# Fixer decision per validation pass, and the run completes (Q13). The
# writer's two content judgments (the Activity-hub marker regex and the
# culmination keyword/count/length block) and the dead split-Type branch
# behind the constant-True ``_api_question_placement_active`` are deleted
# outright: keyword vocabularies and thresholds judging model-authored
# prose, Rule 1's forbidden bullets, with [measured] no test anywhere
# pinning any of them firing.
READBACK_TOPOLOGY_MISMATCH = "bulk_import_readback_topology_mismatch"


def _validate_concepts_workbook_bytes(
    data: bytes,
    concepts: list[models.Concept],
    export_scope: ConceptExportScope,
    *,
    exact_rows: bool,
) -> list[dict]:
    """Read the serialized XLSX back and verify final delivery identity.

    Returns the recorded Fixer decisions ([] when the artifact agrees with
    its accepted DB topology). Each finding batch is one content-addressed
    decision, logged in full here because the fresh-workbook export paths
    return bytes and have no decisions list to collect into.
    """
    expected: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for concept in concepts:
        for topic in _concept_placements(concept):
            key = (
                normalize_question_text(topic.chapter.chapter_title),
                normalize_question_text(concept.concept_title),
                normalize_question_text(
                    strip_topic_title(topic.topic_title) or topic.topic_title
                ),
                normalize_question_text(topic.pre_post_learning),
            )
            front = _front_bands(
                concept,
                topic,
                include_group_columns=False,
                export_scope=export_scope,
            )
            expected[key] = {
                "topic_title": str(front[_IDX_TOPIC_TITLE] or ""),
                "concept_labels": str(front[
                    len(CHAPTER_FIELDS)
                    + TOPIC_FIELDS.index("topic_concept_labels")
                ] or ""),
                "topic_description": str(front[
                    len(CHAPTER_FIELDS)
                    + TOPIC_FIELDS.index("topic_description")
                ] or ""),
            }

    workbook = openpyxl.load_workbook(
        io.BytesIO(data), data_only=True, read_only=True)
    try:
        sheet_name = SHEET_BY_KIND["objective"]
        ws = workbook[sheet_name]
        seen: dict[tuple[str, str, str, str], int] = {}
        issues: list[str] = []

        for row in ws.iter_rows(min_row=3, values_only=True):
            chapter_title = strip_title_tag(
                _cell_str(row, _IDX_CHAPTER_TITLE))
            concept_title = strip_title_tag(
                _cell_str(row, _IDX_CONCEPT_TITLE))
            topic_title = strip_topic_title(
                _cell_str(row, _IDX_TOPIC_TITLE))
            learning_kind = _cell_str(row, _IDX_TOPIC_PRE_POST)
            key = (
                normalize_question_text(chapter_title),
                normalize_question_text(concept_title),
                normalize_question_text(topic_title),
                normalize_question_text(learning_kind),
            )
            if key not in expected:
                continue
            seen[key] = seen.get(key, 0) + 1
            contract = expected[key]
            if _cell_str(row, _IDX_TOPIC_TITLE) != contract["topic_title"]:
                issues.append(
                    f"{concept_title}: noncanonical topic number/title")
            labels_index = (
                len(CHAPTER_FIELDS)
                + TOPIC_FIELDS.index("topic_concept_labels"))
            if _cell_str(row, labels_index) != contract["concept_labels"]:
                issues.append(
                    f"{topic_title}: topic_concept_labels do not match "
                    "the selected topology")
            description_index = (
                len(CHAPTER_FIELDS)
                + TOPIC_FIELDS.index("topic_description"))
            if _cell_str(
                row, description_index
            ) != contract["topic_description"]:
                issues.append(
                    f"{topic_title}: stale topic_description was serialized")

        missing = sorted(set(expected) - set(seen))
        if missing:
            issues.append(
                f"{len(missing)} selected concept placement(s) are missing")
        if exact_rows and any(count != 1 for count in seen.values()):
            issues.append(
                "fresh concept export contains duplicate selected placements")

        if not issues:
            return []
        findings = list(dict.fromkeys(issues))
        decision = _fixer_decision(
            READBACK_TOPOLOGY_MISMATCH,
            detail=(
                "the serialized concept workbook disagrees with its "
                "accepted DB topology; the artifact ships with this "
                "recorded decision instead of being withheld — every "
                "finding is named here for the reviewer (T10-4, Q13)"
            ),
            context={
                "findings": findings,
                "exact_rows": bool(exact_rows),
            },
        )
        logging.getLogger(__name__).warning(
            "bulk-import %s: %s", READBACK_TOPOLOGY_MISMATCH,
            json.dumps(
                decision, sort_keys=True, ensure_ascii=False, default=str),
        )
        return [decision]
    finally:
        workbook.close()


def _topic_number(
    topic: models.Topic, export_scope: ConceptExportScope | None = None,
) -> int:
    """1-based position of the topic within its chapter (textbook order)."""
    if export_scope is not None:
        scoped = export_scope.topic_numbers.get(topic.id)
        if scoped is not None:
            return scoped
    siblings = sorted(
        (
            sibling for sibling in topic.chapter.topics
            if sibling.pre_post_learning == topic.pre_post_learning
        ),
        key=identity.source_order_key,
    )
    try:
        return siblings.index(topic) + 1
    except ValueError:
        return 1


def composed_topic_title(
    topic: models.Topic, export_scope: ConceptExportScope | None = None,
) -> str:
    """Tagged topic title cell, e.g. 'Topic 01: <Title> (<machine id>)'.

    ``strip_topic_title`` normalizes the stored title first so an already-tagged
    value never gets a second 'Topic NN:'/code prefix.

    The tag is the topic's PERSISTED ``machine_id``, read through
    ``identity.machine_id_for_topic``, not ``directory.topic_tag``: that
    function stamps ONE chapter-level code on every topic of the chapter, so
    the cell could not tell two topics apart, and it re-derives from the
    chapter title, so a rename re-keyed every row. The composition itself
    lives in ``identity.titled`` — shared with the Master renderer, stripped
    from neither (T14).
    """
    clean = strip_topic_title(topic.topic_title) or topic.topic_title
    return (
        f"Topic {_topic_number(topic, export_scope):02d}: "
        f"{identity.titled(clean, identity.machine_id_for_topic(topic))}"
    )


def composed_topic_display(topic: models.Topic) -> str:
    """Clean topic display name, e.g. '<Title>' (no topic number or tag/code)."""
    clean = strip_topic_title(topic.topic_title) or topic.topic_title
    return clean


def _groups_by_type(concept: models.Concept) -> dict[str, str]:
    """All groups of each type, comma-separated (S/T/U columns)."""
    buckets: dict[str, list[str]] = {"Basic": [], "Intermediate": [], "Advanced": []}
    for g in sorted(concept.groups, key=lambda g: g.id):
        if g.group_type in buckets:
            buckets[g.group_type].append(g.group_display_name or g.group_name)
    return {k: ", ".join(v) for k, v in buckets.items()}


def _concept_field_value(
    concept: models.Concept, topic: models.Topic, field: str, *,
    include_group_columns: bool,
) -> str:
    # Parent Concept ships empty by team decision: concepts sit flat under
    # their topic. The target layout (spec-step8 Q5) has no parent_concept
    # column at all; older workbooks that still carry one keep it blank.
    parent = ""
    if field == "concept_title":
        # The concept's own persisted id, not the topic-level
        # ``directory.concept_tag`` every concept under the topic used to
        # share (and which collapsed to ``..._X_PL_X`` for any non-Latin
        # source, so two different Marathi chapters minted the same tag).
        return identity.titled(
            concept.concept_title, identity.machine_id_for_concept(concept))
    if field == "concept_display_name":
        return concept.concept_title
    if field == "parent_concept":
        return parent
    if field == "concept_details":
        return concept.concept_details
    if field == "keywords":
        return concept.keywords
    if field == "digicards":
        return concept.digicards
    if field == "related_concepts":
        # The "parent: X" marker this used to smuggle in is gone with the
        # parent column: ``parent`` is unconditionally "" above, so the branch
        # was dead code claiming a fallback that could never fire.
        return concept.related_concepts or ""
    if field in {"basic_groups", "intermediate_groups", "advanced_groups"}:
        if not include_group_columns:
            return ""
        by_type = _groups_by_type(concept)
        return {
            "basic_groups": by_type["Basic"],
            "intermediate_groups": by_type["Intermediate"],
            "advanced_groups": by_type["Advanced"],
        }[field]
    if field == "concept_source":
        return concept.sources
    return ""


def _chapter_book_source(chapter: models.Chapter, concept: models.Concept) -> str:
    """Best book/publication token for the chapter tag (Fullmarks, NCERT, …)."""
    book = directory.primary_book_source(concept.sources)
    if book:
        return book
    for topic in chapter.topics:
        for sibling in topic.concepts:
            book = directory.primary_book_source(sibling.sources)
            if book:
                return book
    return ""


def _front_bands(concept: models.Concept, topic: models.Topic, *,
                 include_group_columns: bool = True,
                 concept_fields: list[str] | None = None,
                 concept_question_labels: str = "",
                 export_scope: ConceptExportScope | None = None) -> list:
    """Chapter + Topic + Concept bands, with tags in the title columns.

    The title columns carry a human-readable tag; the display columns stay
    clean (the reader strips the tags back to the clean model values).

    ``include_group_columns`` is False for concept-catalog rows — group columns
    are filled later when assessments are built, not at concept generation.

    ``concept_question_labels`` is the row's question linkage, not a property
    of the concept: on the TARGET layout that column sits at the end of the
    Concept band (it used to head the Group band), so it is passed in per row
    and is "" for a concept-catalog row (T17 point 3).
    """
    chapter = topic.chapter
    book = _chapter_book_source(chapter, concept)
    c_tag = directory.chapter_tag(
        chapter.board, chapter.grade, chapter.subject, book=book)
    if concept_fields is None:
        concept_fields = list(_target_sheet("objective").block_fields("concept"))
    # Column J lists each concept exactly as its concept_title column reads —
    # the tagged "Name (machine id)" form — so the importer links them
    # (reviewers: "labels should be concept title, not concept display
    # name"). One roster, shared with the Master renderer (T14); the export
    # scope is PASSED, never imported, so neither renderer imports the other.
    concept_labels = identity.topic_concept_roster(topic, export_scope)
    if export_scope is not None:
        pre_topics = ", ".join(
            composed_topic_title(t, export_scope)
            for t in export_scope.topics_for(chapter, "Pre")
        )
        post_topics = ", ".join(
            composed_topic_title(t, export_scope)
            for t in export_scope.topics_for(chapter, "Post")
        )
    else:
        pre_topics = chapter.pre_topics
        post_topics = chapter.post_topics
    return [
        # ---- Chapter band (tag in title, clean display) ----
        f"{chapter.chapter_title} ({c_tag})", chapter.chapter_title,
        chapter.chapter_duration, pre_topics, post_topics,
        chapter.chapter_description,
        # ---- Topic band ("Topic NN: <title> (<tag>)", display "Topic NN: <title>") ----
        composed_topic_title(topic, export_scope),
        composed_topic_display(topic), topic.pre_post_learning, concept_labels,
        topic.related_topics, topic.topic_description,
    ] + [
        concept_question_labels if field == "concept_question_labels"
        else _concept_field_value(
            concept, topic, field,
            include_group_columns=include_group_columns,
        )
        for field in concept_fields
    ]


def _group_band_values(q: models.Question, group: models.Group) -> dict:
    """Group-band cells of one question row, BY NAME.

    Named rather than positional because the Group band is where the layouts
    disagree most: Descriptive orders its fields differently from Objective,
    the canonical layouts carry ``concept_question_labels`` at its head and a
    second ``question_label`` inside it, and the target carries neither.
    """
    return {
        "group_name": group.group_name,
        "group_display_name": group.group_display_name,
        "group_description": group.group_description,
        "group_status": group.group_status,
        "group_type": group.group_type,
        "group_question_labels": q.question_label,
        "related_digicards": group.related_digicards,
        # Canonical-layout columns only; the target carries these elsewhere or
        # not at all. Kept so appending to an older workbook still fills the
        # columns that workbook actually has.
        "concept_question_labels": q.question_label,
        "question_label": q.question_label,
    }


def _question_band_values(
    q: models.Question, sheet_layout: layouts.SheetLayout,
) -> dict:
    """Question-band cells of one question row, BY NAME.

    The answer-block and sub-question COUNTS come from the identified layout,
    never from a literal: the canonical Subjective sheet carries 10 answer
    blocks and the target carries 20, so the old ``range(10)`` left blocks
    11-20 unwritten by construction (T6).
    """
    values: dict = {
        "question_label": q.question_label,
        "question_category": q.question_category,
        "cognitive_skills": q.cognitive_skills,
        "question_source": q.question_source,
        "question_disclaimer": q.question_disclaimer,
        "question_duration": q.question_duration,
        "math_keyboard": q.math_keyboard,
        "question_appears_in": q.question_appears_in,
        "level_of_difficulty": q.level_of_difficulty,
        "question": q.question,
        "question_text": q.question_text,
        "marks": q.marks,
        "display_answer": q.display_answer,
        "answer_explanation": q.answer_explanation,
    }
    answers = q.answers or []
    for n in sheet_layout.answer_block_numbers:
        answer = answers[n - 1] if n - 1 < len(answers) else {}
        # The answer dict's keys ARE the column prefixes on every sheet
        # (objective answer_content_N, subjective answer_N, descriptive
        # answer_weightage_N), so no per-kind translation table is needed.
        for key, value in (answer or {}).items():
            values[f"{key}_{n}"] = value
    sub_questions = q.sub_questions or []
    for n in sheet_layout.sub_question_numbers:
        sub = sub_questions[n - 1] if n - 1 < len(sub_questions) else {}
        sub = sub or {}
        values[f"sub_question_{n}"] = sub.get("text", "")
        values[f"sub_question_marks_{n}"] = sub.get("marks", "")
        keywords = sub.get("keywords") or []
        for m in sheet_layout.sub_question_keyword_numbers(n):
            keyword = keywords[m - 1] if m - 1 < len(keywords) else {}
            keyword = keyword or {}
            values[f"sq{n}_answer_type_{m}"] = keyword.get("answer_type", "")
            values[f"sq{n}_weightage_{m}"] = keyword.get("weightage", "")
            values[f"sq{n}_keyword_{m}"] = keyword.get("keyword", "")
    return values


def _band_cells(
    sheet_layout: layouts.SheetLayout, block: str, values: dict,
) -> list:
    """One band emitted in the layout's own column order, addressed by name."""
    return [values.get(name, "") for name in sheet_layout.block_fields(block)]


def _question_to_row(q: models.Question, kind: str,
                     group: "models.Group | None" = None,
                     concept_fields: list[str] | None = None,
                     sheet_layout: layouts.SheetLayout | None = None,
                     decisions: list[dict] | None = None) -> list:
    """Build one flat row from a normalized Question, in a layout's order.

    ``group`` selects the *placement*: the question's authoring home
    (``q.group``) by default, or a tagged group when emitting a many-to-many
    tag row. The question content is identical across placements; only the
    Chapter/Topic/Concept/Group bands change.

    ``sheet_layout`` is the layout of the sheet being written — the identified
    layout of an existing workbook, or the target layout for a fresh one. The
    Group and Question bands are emitted BY NAME through it, so appending to an
    older workbook fills that workbook's columns instead of writing target
    values at canonical positions.
    """
    group = group or q.group
    concept = group.concept
    topic = concept.topic

    sheet_layout = sheet_layout or _target_sheet(kind)
    if concept_fields is None:
        concept_fields = list(sheet_layout.block_fields("concept"))
    row: list = list(_front_bands(
        concept, topic, concept_fields=concept_fields,
        concept_question_labels=q.question_label,
    ))
    row += _band_cells(sheet_layout, "group", _group_band_values(q, group))
    row += _band_cells(
        sheet_layout, "question", _question_band_values(q, sheet_layout))

    expected = len(sheet_layout.fields) + (
        len(concept_fields) - len(sheet_layout.block_fields("concept")))
    return _fit_row_to_sheet(
        row, expected, kind=kind,
        row_identity=str(q.question_label or ""), decisions=decisions,
    )


def _concept_to_row(concept: models.Concept, kind: str = "objective",
                    topic: "models.Topic | None" = None,
                    concept_fields: list[str] | None = None,
                    sheet_layout: layouts.SheetLayout | None = None,
                    export_scope: ConceptExportScope | None = None,
                    decisions: list[dict] | None = None) -> list:
    """Build a concept-catalog row (chapter/topic/concept filled, no question).

    ``topic`` selects the placement: the concept's authoring home
    (``concept.topic``) by default, or a tagged topic (possibly in another
    chapter) when emitting a many-to-many concept tag row.

    Every column from the Group band onward is blank (OWNER RULING OD6 /
    T17): a Concept File has no groups and no questions. That is now the
    explicit tail below rather than an accident of the row-width pad.

    Which is why the width check is on the FRONT bands, not on the whole row.
    Padding to the sheet width first and checking afterwards made T3.8's
    narrow-row record unreachable here: a genuinely short Chapter/Topic/
    Concept composition was absorbed by the very pad that writes the
    deliberate tail, so a real defect and the intended blank tail were
    indistinguishable. Checked against the front width they are distinct, and
    the tail is appended explicitly below exactly as this docstring claims.
    """
    topic = topic or concept.topic
    sheet_layout = sheet_layout or _target_sheet(kind)
    if concept_fields is None:
        concept_fields = list(sheet_layout.block_fields("concept"))
    row: list = list(_front_bands(
        concept,
        topic,
        include_group_columns=False,
        concept_fields=concept_fields,
        export_scope=export_scope,
    ))
    expected_front = (
        len(sheet_layout.block_fields("chapter"))
        + len(sheet_layout.block_fields("topic"))
        + len(concept_fields)
    )
    row = _fit_row_to_sheet(
        row, expected_front, kind=kind,
        row_identity=str(concept.concept_title or ""), decisions=decisions,
    )
    expected = len(sheet_layout.fields) + (
        len(concept_fields) - len(sheet_layout.block_fields("concept")))
    return row + [""] * (expected - len(row))


def _row_has_question(ws, row_i: int, q_start: int) -> bool:
    """Whether an existing workbook row carries any Question-band content."""
    return any(
        str(ws.cell(row=row_i, column=column).value or "").strip()
        for column in range(q_start + 1, ws.max_column + 1)
    )


def _refresh_concept_rows(
    wb,
    index: WorkbookIndex,
    concept: models.Concept,
    topic: models.Topic,
    locations: list[tuple],
    *,
    include_ancestors: bool,
    export_scope: ConceptExportScope | None = None,
) -> int:
    """Refresh selected rows from the DB concept, retaining row identity.

    Normal refreshes write only the Concept band. Placement reconciliation also
    writes the Chapter and Topic bands so a stale row moves to its current
    authoritative placement. Group and Question bands are never overwritten.
    Concept source history is merged rather than replaced.
    """
    sources_updated = 0
    for sheet_name, row_i in locations:
        meta = index.sheet_meta.get(sheet_name) or {}
        concept_fields = meta.get("concept_fields")
        q_start = meta.get("q_start")
        if not concept_fields or q_start is None:
            continue
        ws = wb[sheet_name]
        # Catalog-only rows deliberately leave basic/intermediate/advanced
        # summaries blank. Question rows carry the current group summaries.
        include_group_columns = _row_has_question(ws, row_i, q_start)
        front_values = _front_bands(
            concept,
            topic,
            include_group_columns=include_group_columns,
            concept_fields=concept_fields,
            export_scope=export_scope,
        )
        start = 0 if include_ancestors else _IDX_CONCEPT_TITLE
        for column_index, value in enumerate(front_values[start:], start=start):
            field_index = column_index - _IDX_CONCEPT_TITLE
            field = (
                concept_fields[field_index]
                if 0 <= field_index < len(concept_fields)
                else ""
            )
            if field == "concept_question_labels":
                # A question linkage, not a concept property. On the target
                # layout it sits inside the Concept band, so refreshing the
                # band would blank every question row's link — the silent loss
                # this refresh exists to prevent.
                continue
            cell = ws.cell(row=row_i, column=column_index + 1)
            if field == "concept_source":
                current = str(cell.value or "")
                value = merge_sources(current, str(value or ""))
                if value != current:
                    sources_updated += 1
            _set_cell_value(cell, value)
    return sources_updated


def _refresh_concept_band(
    wb,
    index: WorkbookIndex,
    concept: models.Concept,
    topic: models.Topic,
    export_scope: ConceptExportScope | None = None,
) -> int:
    """Refresh every row at one exact placement, including scoped metadata."""
    return _refresh_concept_rows(
        wb,
        index,
        concept,
        topic,
        list(index.concept_rows.get(concept_placement_key(concept, topic), [])),
        # Topic numbering, topic_concept_labels, descriptions, and chapter
        # pre/post lists belong to the accepted export topology too. Updating
        # only the Concept band is how stale 113-label metadata survived a
        # 42-row successful run.
        include_ancestors=True,
        export_scope=export_scope,
    )


def _move_indexed_concept_row(
    index: WorkbookIndex,
    old_key: tuple,
    new_key: tuple,
    location: tuple,
) -> None:
    """Reflect one in-place placement move in the current workbook index."""
    old_locations = index.concept_rows.get(old_key, [])
    if location in old_locations:
        old_locations.remove(location)
    if not old_locations:
        index.concept_rows.pop(old_key, None)
        index.c_placements.discard(old_key)
    new_locations = index.concept_rows.setdefault(new_key, [])
    if location not in new_locations:
        new_locations.append(location)
    index.c_placements.add(new_key)
    index.concept_titles.add(new_key[0])


def _reconcile_concept_placements(
    wb,
    index: WorkbookIndex,
    concept: models.Concept,
    desired_topics: list[models.Topic],
    export_scope: ConceptExportScope | None = None,
) -> int:
    """Move stale same-chapter rows to current DB placements conservatively.

    Question rows always belong to the concept's authoritative home topic.
    Catalog-only rows are reused only for desired placements that lack a
    catalog row. Rows in another chapter are never candidates, even when their
    normalized concept title is identical.
    """
    if not desired_topics or concept.topic is None:
        return 0
    home_key = concept_placement_key(concept, concept.topic)
    identity, home_chapter, _, home_learning_kind = home_key
    desired_by_key = {
        concept_placement_key(concept, topic): topic
        for topic in desired_topics
    }
    desired_same_chapter = {
        key: topic
        for key, topic in desired_by_key.items()
        if key[1] == home_chapter and key[3] == home_learning_kind
    }
    same_chapter_keys = [
        key
        for key in list(index.concept_rows)
        if (
            key[0] == identity
            and key[1] == home_chapter
            and key[3] == home_learning_kind
        )
    ]
    if not same_chapter_keys:
        return 0

    sources_updated = 0

    # A ConceptTag does not relocate the concept's assessments. Therefore any
    # question-bearing row under another topic is a stale former-home row,
    # including when that topic remains a legitimate catalog tag.
    for old_key in same_chapter_keys:
        if old_key == home_key:
            continue
        for location in list(index.concept_rows.get(old_key, [])):
            sheet_name, row_i = location
            meta = index.sheet_meta.get(sheet_name) or {}
            q_start = meta.get("q_start")
            if q_start is None or not _row_has_question(
                    wb[sheet_name], row_i, q_start):
                continue
            sources_updated += _refresh_concept_rows(
                wb,
                index,
                concept,
                concept.topic,
                [location],
                include_ancestors=True,
                export_scope=export_scope,
            )
            _move_indexed_concept_row(
                index, old_key, home_key, location)

    def _catalog_locations(key: tuple) -> list[tuple]:
        out: list[tuple] = []
        for location in index.concept_rows.get(key, []):
            sheet_name, row_i = location
            meta = index.sheet_meta.get(sheet_name) or {}
            q_start = meta.get("q_start")
            if q_start is not None and not _row_has_question(
                    wb[sheet_name], row_i, q_start):
                out.append(location)
        return out

    # Home comes first in ``desired_topics``. Reuse stale catalog rows for
    # missing desired catalog placements in that same stable order.
    missing_catalog = [
        (key, desired_by_key[key])
        for key in desired_by_key
        if key in desired_same_chapter and not _catalog_locations(key)
    ]
    stale_catalog: list[tuple[tuple, tuple]] = []
    for old_key in same_chapter_keys:
        if old_key in desired_same_chapter:
            continue
        stale_catalog.extend(
            (old_key, location)
            for location in _catalog_locations(old_key)
        )

    for (old_key, location), (new_key, topic) in zip(
            stale_catalog, missing_catalog):
        sources_updated += _refresh_concept_rows(
            wb,
            index,
            concept,
            topic,
            [location],
            include_ancestors=True,
            export_scope=export_scope,
        )
        _move_indexed_concept_row(index, old_key, new_key, location)

    return sources_updated


@workbook_sync.synchronized_output_workbook
def append_concepts(db: Session, path: Path, concept_ids: list[int],
                    *, sibling_for: Path | None = None) -> dict[str, int]:
    """Append concept-catalog rows (no questions) to the Objective sheet.

    One row per (concept, placement): the concept's home topic plus every
    tagged topic/chapter. Placements already present are never re-added —
    instead their complete Concept bands are refreshed in place from the
    current DB concept. ``concept_source`` is merged so a concept re-used from
    another book keeps all previously recorded sources.

    ``sibling_for`` is the published workbook ``path`` is a staged copy of, and
    it only names the retained pre-migration sibling (see
    ``layouts.migrate_workbook_layout``). Nothing is read from or written to
    it here.
    """
    # An existing workbook on an older registered layout is MIGRATED ONCE,
    # here, before it is opened for writing — otherwise this function would
    # write target-layout values at old-layout positions (Q16/T7.6). It runs
    # on the staged sibling the transactional outbox already provides, it
    # takes ``output_workbook_lock()`` itself (an RLock, so it composes with
    # this function's own decorator), and it NEVER raises: append_concepts is
    # reached mid-run with the model budget already spent, so an unmappable
    # value is a recorded Fixer decision plus a release issue and the run
    # completes (Q13).
    issues: list[dict] = []
    decisions: list[dict] = []
    receipt = layouts.migrate_workbook_layout(
        path, target_layout_id=layouts.REFERENCE_LAYOUT_ID,
        sibling_for=sibling_for)
    index = scan_workbook(path)
    wb = openpyxl.load_workbook(path) if path.exists() else _new_workbook()
    sheet_name = _sheet_name_for(wb, index, "objective")
    ws = wb[sheet_name]
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    objective_layout = _identified_sheet(
        sheet_name, header, "objective", mismatches=index.mismatches)
    concept_fields = list(objective_layout.block_fields("concept"))
    concepts = (
        db.query(models.Concept).filter(models.Concept.id.in_(concept_ids))
        .all()
    )
    concepts = sorted(concepts, key=lambda concept: (
        identity.source_order_key(concept.topic),
        identity.source_order_key(concept),
    ))
    export_scope = ConceptExportScope(concepts)
    result: dict = {
        "written": 0,
        "sources_updated": 0,
    }
    if receipt is not None:
        result["layout_migration"] = receipt.as_dict()
        for entry in receipt.unmappable:
            detail = (
                "migrating this workbook from layout "
                f"{receipt.source_layout_id!r} to {receipt.target_layout_id!r} "
                f"left a non-blank {entry['sheet']} value with no column: "
                f"{entry['band']}.{entry['field']} on row {entry['row']} "
                f"({entry['value']!r}). The pre-migration workbook is retained "
                f"at {receipt.sibling_path}."
            )
            decisions.append(_fixer_decision(
                layouts.MIGRATION_UNMAPPABLE_VALUE,
                detail=detail, context=dict(entry),
            ))
            issues.append({
                "code": layouts.MIGRATION_UNMAPPABLE_VALUE,
                "detail": detail,
                **entry,
            })
        # A sheet the source layout declares ``ignored`` is in NEITHER layout,
        # so it has no columns to map and the migrated workbook is rebuilt
        # without it. Recording only its name cannot say whether a reviewer had
        # put anything in it, so a dropped sheet that was carrying data rows
        # gets the same recorded-decision treatment an unmappable cell does —
        # once for the sheet, naming the cell count and where the values are
        # still readable, rather than once per cell.
        for sheet_name in receipt.dropped_sheets:
            cells = receipt.dropped_sheet_cells.get(sheet_name, 0)
            if not cells:
                continue
            detail = (
                f"migrating this workbook from layout "
                f"{receipt.source_layout_id!r} to {receipt.target_layout_id!r} "
                f"dropped the sheet {sheet_name!r}, which is in no registered "
                f"layout and was carrying {cells} non-blank data cell(s). "
                f"They are retained verbatim in the pre-migration workbook at "
                f"{receipt.sibling_path}."
            )
            context = {
                "sheet": sheet_name,
                "cells": cells,
                "source_layout_id": receipt.source_layout_id,
                "target_layout_id": receipt.target_layout_id,
                "sibling_path": receipt.sibling_path,
            }
            decisions.append(_fixer_decision(
                layouts.MIGRATION_DROPPED_SHEET,
                detail=detail, context=context,
            ))
            issues.append({
                "code": layouts.MIGRATION_DROPPED_SHEET,
                "detail": detail,
                **context,
            })
    for c in concepts:
        desired_topics = _concept_placements(c)
        result["sources_updated"] += _reconcile_concept_placements(
            wb, index, c, desired_topics, export_scope)
        for topic in desired_topics:
            key = concept_placement_key(c, topic)
            if key in index.c_placements:
                result["sources_updated"] += _refresh_concept_band(
                    wb, index, c, topic, export_scope)
                continue
            index.c_placements.add(key)
            target = ws.max_row + 1 if ws.max_row >= 2 else 3
            for i, value in enumerate(
                _concept_to_row(
                    c,
                    "objective",
                    topic,
                    concept_fields=concept_fields,
                    sheet_layout=objective_layout,
                    export_scope=export_scope,
                    decisions=decisions,
                ),
                start=1,
            ):
                _write_cell(ws, row=target, column=i, value=value)
            index.concept_rows.setdefault(key, []).append(
                (sheet_name, target))
            index.concept_titles.add(key[0])
            result["written"] += 1
    serialized = io.BytesIO()
    wb.save(serialized)
    # T10-4 (S11): a read-back disagreement is one recorded decision and
    # the workbook still lands — the raise that used to sit here gated the
    # mid-run deposit after the model budget was spent, and nothing ever
    # caught it.
    decisions.extend(_validate_concepts_workbook_bytes(
        serialized.getvalue(),
        concepts,
        export_scope,
        exact_rows=False,
    ))
    workbook_sync.atomic_save_workbook(wb, path)
    for decision in decisions:
        if decision["code"] in (
            ROW_WIDTH_MISMATCH, READBACK_TOPOLOGY_MISMATCH,
        ):
            issues.append({**decision})
    if issues:
        result["issues"] = issues
    if decisions:
        result["fixer_decisions"] = decisions
    if index.mismatches:
        result["layout_mismatches"] = list(index.mismatches)
    return result


def _write_headers(ws, kind: str) -> None:
    """Emit the two header rows of one sheet, exactly as the authority has them.

    Row 1's bands are ``{label, start, end}`` and they DO NOT tile the sheet:
    the reference Objective sheet leaves column 23 unbanded and bands nothing
    after column 31. Their labels carry per-sheet trailing whitespace
    (``'Chapter '`` vs ``'Chapter  '``) and it is written byte-exactly, never
    normalised — a normaliser here is what the trailing-space sheet-name
    constant was.
    """
    fields = FIELDS_BY_KIND[kind]
    # Row 1: section bands (merged, with gaps).
    for band in SECTION_BANDS[kind]:
        label, start, end = band["label"], band["start"], band["end"]
        cell = ws.cell(row=1, column=start)
        _set_cell_value(cell, label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            "solid", fgColor=_BAND_FILL.get(label.strip(), "EEEEEE"))
        if end > start:
            ws.merge_cells(
                start_row=1, start_column=start, end_row=1, end_column=end)
    # Row 2: field names.
    for i, name in enumerate(fields, start=1):
        c = ws.cell(row=2, column=i)
        _set_cell_value(c, name)
        c.font = Font(bold=True, size=9)
    ws.freeze_panes = "A3"
    ws.column_dimensions[get_column_letter(1)].width = 22


def _new_workbook() -> openpyxl.Workbook:
    """A fresh workbook on the target layout — the authority's THREE sheets.

    This function builds Outputs 01 and 03 (the two Concept Files) as well as
    the append-only output workbook, so OWNER RULING OD6's sheet half lands
    here (T12/M1, T17/B1). It used to emit four sheets — ``'Objective '``,
    ``'Subjective'``, ``'Descriptive'`` and ``'Doc Link <> Each fields '`` —
    against the committed format workbook's three, in the wrong order, with a
    Doc Link sheet that is in no registered layout and therefore made the file
    unidentifiable. The two data-less sheets stay: they are what makes the
    reviewer's four files one workbook family that a single ``identify_sheet``
    can read.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for kind, sheet_name in SHEET_BY_KIND.items():
        ws = wb.create_sheet(sheet_name)
        _write_headers(ws, kind)
    return wb


def _questions(db: Session, question_ids: list[int] | None) -> list[models.Question]:
    q = db.query(models.Question)
    if question_ids is not None:
        q = q.filter(models.Question.id.in_(question_ids))
    return q.order_by(models.Question.id).all()


def write_workbook(db: Session, dest: Path | None = None,
                   question_ids: list[int] | None = None) -> bytes:
    """Write a fresh canonical workbook with the selected questions.

    Emits one row per (question, placement): the authoring home plus every tag,
    so many-to-many associations survive a full export. Concepts that have no
    question rows still get concept-catalog rows on the Objective sheet —
    otherwise a database holding only generated concepts (no assessments yet)
    would export as an empty workbook.
    """
    wb = _new_workbook()
    next_row = {k: 3 for k in SHEET_BY_KIND}
    concepts_with_rows: set[int] = set()
    for q in _questions(db, question_ids):
        ws = wb[SHEET_BY_KIND[q.sheet_kind]]
        for group in _question_placements(q):
            for i, value in enumerate(_question_to_row(q, q.sheet_kind, group), start=1):
                _write_cell(
                    ws,
                    row=next_row[q.sheet_kind],
                    column=i,
                    value=value,
                )
            next_row[q.sheet_kind] += 1
            concepts_with_rows.add(group.concept_id)
    if question_ids is None:
        ws_obj = wb[SHEET_BY_KIND["objective"]]
        for concept in db.query(models.Concept).order_by(models.Concept.id):
            if concept.id in concepts_with_rows:
                continue
            for topic in _concept_placements(concept):
                for i, value in enumerate(
                    _concept_to_row(concept, "objective", topic), start=1
                ):
                    _write_cell(
                        ws_obj,
                        row=next_row["objective"],
                        column=i,
                        value=value,
                    )
                next_row["objective"] += 1
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if dest:
        dest.write_bytes(data)
    return data


def write_concepts_workbook(db: Session, concept_ids: list[int]) -> bytes:
    """Write a fresh canonical workbook holding only the given concepts.

    Concepts have no questions of their own here, so they are emitted as
    concept-catalog rows on the Objective sheet — one row per (concept,
    placement) (the authoring home topic plus every tagged topic/chapter) —
    exactly the shape ``append_concepts`` writes to the app-data output
    workbook. Used by the per-functionality "download Bulk Import Excel"
    export for the Build Concepts flows.
    """
    wb = _new_workbook()
    ws = wb[SHEET_BY_KIND["objective"]]
    concepts = (
        db.query(models.Concept).filter(models.Concept.id.in_(concept_ids))
        .all()
    )
    concepts = sorted(concepts, key=lambda concept: (
        identity.source_order_key(concept.topic),
        identity.source_order_key(concept),
    ))
    export_scope = ConceptExportScope(concepts)
    next_row = 3
    for c in concepts:
        for topic in sorted(_concept_placements(c), key=identity.source_order_key):
            for i, value in enumerate(
                _concept_to_row(
                    c,
                    "objective",
                    topic,
                    export_scope=export_scope,
                ),
                start=1,
            ):
                _write_cell(ws, row=next_row, column=i, value=value)
            next_row += 1
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    # T10-4 (S11): findings are recorded (the helper logs the full
    # decision — this path returns bytes and has no decisions list), and
    # the download ships. Withholding the file was the defect: nothing
    # ever caught the raise, and Rule E says no defect blocks a download.
    _validate_concepts_workbook_bytes(
        data,
        concepts,
        export_scope,
        exact_rows=True,
    )
    return data


def write_subject_workbook(
    db: Session, *, subject: str, board: str = "", grade: str = "",
    include_content: bool = True,
) -> bytes:
    """Create a canonical workbook scoped to one subject.

    ``include_content=False`` yields a blank authoring template (headers only,
    exact canonical layout). With content, every question placement and every
    concept placement falling inside the scoped chapters is emitted; concepts
    without questions still get concept-catalog rows on the Objective sheet so
    the full hierarchy is represented.
    """
    wb = _new_workbook()
    if include_content:
        q = db.query(models.Chapter).filter(models.Chapter.subject == subject)
        if board:
            q = q.filter(models.Chapter.board == board)
        if grade:
            q = q.filter(models.Chapter.grade == grade)
        chapter_ids = {c.id for c in q.all()}

        next_row = {k: 3 for k in SHEET_BY_KIND}
        concepts_with_rows: set[int] = set()
        for question in db.query(models.Question).order_by(models.Question.id):
            for group in _question_placements(question):
                if group.concept.topic.chapter_id not in chapter_ids:
                    continue
                ws = wb[SHEET_BY_KIND[question.sheet_kind]]
                for i, value in enumerate(
                    _question_to_row(question, question.sheet_kind, group), start=1
                ):
                    _write_cell(
                        ws,
                        row=next_row[question.sheet_kind],
                        column=i,
                        value=value,
                    )
                next_row[question.sheet_kind] += 1
                concepts_with_rows.add(group.concept_id)

        # Concept-catalog rows for in-scope concepts that have no question rows.
        # In-scope = home topic in a scoped chapter OR tagged into one.
        ws_obj = wb[SHEET_BY_KIND["objective"]]
        home = (
            db.query(models.Concept).join(models.Topic)
            .filter(models.Topic.chapter_id.in_(chapter_ids)).all()
        )
        tagged = (
            db.query(models.Concept).join(models.ConceptTag).join(
                models.Topic, models.ConceptTag.topic_id == models.Topic.id)
            .filter(models.Topic.chapter_id.in_(chapter_ids)).all()
        )
        in_scope: dict[int, models.Concept] = {c.id: c for c in home + tagged}
        for concept in sorted(in_scope.values(), key=lambda c: c.id):
            if concept.id in concepts_with_rows:
                continue
            for topic in _concept_placements(concept):
                if topic.chapter_id not in chapter_ids:
                    continue
                for i, value in enumerate(_concept_to_row(concept, "objective", topic), start=1):
                    _write_cell(
                        ws_obj,
                        row=next_row["objective"],
                        column=i,
                        value=value,
                    )
                next_row["objective"] += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@workbook_sync.synchronized_output_workbook
def append_questions(db: Session, path: Path, question_ids: list[int]) -> dict[str, int]:
    """Append-only write, placement-aware.

    Adds one row per (question, placement) — the authoring home plus every tag —
    skipping any (label, ancestor-path) already present. A repeated label under
    a *new* placement is therefore written as a tag rather than skipped.
    """
    index = scan_workbook(path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = _new_workbook()

    appended: dict = {"objective": 0, "subjective": 0, "descriptive": 0,
                      "tagged": 0, "skipped": 0, "sources_updated": 0}
    decisions: list[dict] = []
    for q in _questions(db, question_ids):
        for n, group in enumerate(_question_placements(q)):
            key = question_placement_key(q.question_label, group)
            existing_key = key
            if (
                q.origin == "concept_mapping"
                and key not in index.q_placements
            ):
                existing_key = visible_question_placement_key(
                    q.question_label, group)
            if q.question_label and existing_key in index.q_placements:
                appended["skipped"] += 1
                # Existing row: refresh its question_source in place so a
                # duplicate question arriving from another book accumulates
                # sources instead of duplicating the row.
                loc = index.q_rows.get(existing_key)
                col = (
                    (index.sheet_meta.get(loc[0]) or {}).get("q_src_col")
                    if loc else None
                )
                if loc and q.question_source and col is not None:
                    sheet_name, row_i = loc
                    cell = wb[sheet_name].cell(row=row_i, column=col + 1)
                    merged = merge_sources(str(cell.value or ""), q.question_source)
                    if merged != str(cell.value or ""):
                        _set_cell_value(cell, merged)
                        appended["sources_updated"] += 1
                continue
            is_tag = q.question_label in index.labels
            index.q_placements.add(key)
            index.labels.add(q.question_label)
            sheet_name = _sheet_name_for(wb, index, q.sheet_kind)
            ws = wb[sheet_name]
            header = next(
                ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
            sheet_layout = _identified_sheet(
                sheet_name, header, q.sheet_kind,
                mismatches=index.mismatches)
            concept_fields = list(sheet_layout.block_fields("concept"))
            target = ws.max_row + 1 if ws.max_row >= 2 else 3
            for i, value in enumerate(
                _question_to_row(
                    q, q.sheet_kind, group, concept_fields=concept_fields,
                    sheet_layout=sheet_layout, decisions=decisions),
                start=1,
            ):
                _write_cell(ws, row=target, column=i, value=value)
            appended[q.sheet_kind] += 1
            if is_tag:
                appended["tagged"] += 1

    workbook_sync.atomic_save_workbook(wb, path)
    if decisions:
        appended["fixer_decisions"] = decisions
    # Symmetric with ``append_concepts``: the same defect on the same row
    # composer reaches the release-issue channel from either entry point.
    issues = [
        {"code": ROW_WIDTH_MISMATCH, **decision}
        for decision in decisions
        if decision["code"] == ROW_WIDTH_MISMATCH
    ]
    if issues:
        appended["issues"] = issues
    if index.mismatches:
        appended["layout_mismatches"] = list(index.mismatches)
    return appended
