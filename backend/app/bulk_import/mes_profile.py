"""MES workbook profile: dual projections of one immutable release snapshot.

The positional authority for everything here is the Grade-6 MES Bulk Import
FINAL template set (spec §0 item 2, §23), captured verbatim in
``mes_template_manifest.json``: three sheets (Objective, Descriptive,
Subjective — no Doc Link, no trailing spaces in sheet names), the exact
two-row headers, the row-1 merged bands with their sheet-specific labels,
``answer_restriction`` between ``question_appears_in`` and
``level_of_difficulty``, ``question_text`` directly after ``question``, and
``answer_explanation`` closing the Objective/Subjective bands.

Two renderers project ONE snapshot (spec §1, §13.1):

* Concept File — hierarchy catalogue only: chapter/topic/concept bands in
  source order, Group and Question bands blank, Descriptive and Subjective
  header-only.
* Master File — the complete import: every concept including questionless
  ones, every group (required shells and occupied groups, with catalogue
  rows for groups no question row represents), every Objective and
  Descriptive assessment with answers, rubrics, subquestions, restriction,
  the exact MES appears-in wire value, and complete label aggregates; zero
  Subjective data rows.

Deterministic mechanics only: positional serialization, formula-injection
escaping, cell limits, read-back parsing, and count/arithmetic validation.
Nothing here decides content — flags and unplaced candidates ride into the
returned manifest, never disappear.
"""
from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path
from typing import Any, Mapping

import openpyxl
from openpyxl.styles import Alignment, Font

from ..services import assessment_release as rel

CELL_LIMIT = 32_767

_MANIFEST_PATH = Path(__file__).with_name("mes_template_manifest.json")
MANIFEST: dict = json.loads(_MANIFEST_PATH.read_text(encoding="utf-8"))
SHEET_ORDER: list[str] = list(MANIFEST["sheet_order"])
FIELDS: dict[str, list[str]] = {
    name: list(sheet["fields"]) for name, sheet in MANIFEST["sheets"].items()
}
BANDS: dict[str, list[dict]] = {
    name: list(sheet["bands"]) for name, sheet in MANIFEST["sheets"].items()
}
_INDEX: dict[str, dict[str, int]] = {
    name: {field: i for i, field in enumerate(fields)}
    for name, fields in FIELDS.items()
}

# MES §3.8: these fields ship blank regardless of what upstream rows carry.
_FORCED_BLANK = {"chapter_duration", "question_disclaimer"}

MAX_OBJECTIVE_OPTIONS = 6
MAX_DESCRIPTIVE_ANSWERS = 10
MAX_SUBQUESTIONS = 15
MAX_SUBQUESTION_KEYWORDS = 6


class MesRenderError(ValueError):
    """A workbook cannot be rendered without violating a mechanical rule."""


def _cell_value(value: Any, *, context: str) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)
    if len(text) > CELL_LIMIT:
        raise MesRenderError(
            f"{context}: cell exceeds {CELL_LIMIT} characters "
            f"({len(text)})")
    if text[:1] in {"=", "+", "-", "@"}:
        return "'" + text  # formula-injection guard (spec §11)
    return text


def _row_values(sheet: str, record: Mapping[str, Any]) -> list:
    fields = FIELDS[sheet]
    row = []
    for field in fields:
        value = "" if field in _FORCED_BLANK else record.get(field, "")
        row.append(_cell_value(
            value, context=f"{sheet}:{field}"))
    return row


def _write_headers(ws, sheet: str) -> None:
    for band in BANDS[sheet]:
        cell = ws.cell(row=1, column=band["start"])
        cell.value = band["label"]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if band["end"] > band["start"]:
            ws.merge_cells(
                start_row=1, start_column=band["start"],
                end_row=1, end_column=band["end"])
    for i, field in enumerate(FIELDS[sheet], start=1):
        cell = ws.cell(row=2, column=i)
        cell.value = field
        cell.font = Font(bold=True, size=9)
    ws.freeze_panes = "A3"


def _new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        _write_headers(wb.create_sheet(name), name)
    return wb


def _workbook_bytes(wb: openpyxl.Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Snapshot access
# --------------------------------------------------------------------------- #
# The snapshot is one immutable dict (the release projection):
#   chapter: {chapter_title, chapter_display_name, pre_topics, post_topics,
#             chapter_description}
#   topics:  ordered [{topic fields..., "concepts": ordered [concept fields
#            ..., "concept_key"]}]
#   groups:  [group records (assessment_grouping.group_record) each with
#            "concept_key"]
#   candidates: [materialized candidates with "question_label" and
#            "concept_key"/"group_key" filled from their placement]
# ``concept_key`` is any stable join key (machine id); titles are display
# values and never used for joining.


def _concept_rows(snapshot: Mapping) -> list[dict]:
    rows = []
    chapter = dict(snapshot.get("chapter") or {})
    for topic in snapshot.get("topics") or []:
        topic_fields = {
            k: v for k, v in topic.items() if k != "concepts"
        }
        for concept in topic.get("concepts") or []:
            rows.append({
                "chapter": chapter,
                "topic": topic_fields,
                "concept": dict(concept),
            })
    return rows


def _bands_record(entry: Mapping) -> dict:
    record: dict = {}
    record.update(entry["chapter"])
    record.update(entry["topic"])
    record.update({
        k: v for k, v in entry["concept"].items() if k != "concept_key"
    })
    return record


def snapshot_sha256(snapshot: Mapping) -> str:
    return rel.sha256_json(snapshot)


# --------------------------------------------------------------------------- #
# Output A — Concept File (spec §1, §9)
# --------------------------------------------------------------------------- #

def render_concept_file(snapshot: Mapping) -> bytes:
    wb = _new_workbook()
    ws = wb["Objective"]
    for entry in _concept_rows(snapshot):
        record = _bands_record(entry)
        # A clean catalogue: Group and Question bands stay blank, and the
        # concept-band group labels stay blank too — groups do not exist in
        # Output A (spec §1: "no populated Group fields").
        for field in ("basic_groups", "intermediate_groups",
                      "advanced_groups", "concept_question_labels"):
            record[field] = ""
        ws.append(_row_values("Objective", record))
    return _workbook_bytes(wb)


# --------------------------------------------------------------------------- #
# Output B — Master File (spec §1, §10)
# --------------------------------------------------------------------------- #

def _question_record(candidate: Mapping, sheet: str) -> dict:
    record = {
        "question_label": candidate.get("question_label", ""),
        "question_category": candidate.get("question_category", ""),
        "cognitive_skills": candidate.get("cognitive_skill", ""),
        "question_source": candidate.get("question_source", ""),
        "question_duration": candidate.get("question_duration", 1),
        "question_appears_in": candidate.get("question_appears_in", ""),
        "answer_restriction": candidate.get("answer_restriction", ""),
        "level_of_difficulty": candidate.get("difficulty", ""),
        "question": candidate.get("question", ""),
        "question_text": candidate.get("question_text", ""),
        "marks": candidate.get("marks", ""),
        "answer_explanation": candidate.get("answer_explanation", ""),
    }
    answers = [
        a for a in candidate.get("answers") or [] if isinstance(a, Mapping)
    ]
    if sheet == "Objective":
        if len(answers) > MAX_OBJECTIVE_OPTIONS:
            raise MesRenderError(
                f"{record['question_label']}: more than "
                f"{MAX_OBJECTIVE_OPTIONS} options")
        for n, answer in enumerate(answers, start=1):
            record[f"answer_type_{n}"] = answer.get("answer_type", "")
            record[f"answer_content_{n}"] = answer.get("answer_content", "")
            record[f"correct_answer_{n}"] = answer.get("correct_answer", "")
            record[f"answer_weightage_{n}"] = answer.get(
                "answer_weightage", "")
    else:  # Descriptive
        record["math_keyboard"] = candidate.get("math_keyboard", "")
        record["display_answer"] = candidate.get("display_answer", "")
        if len(answers) > MAX_DESCRIPTIVE_ANSWERS:
            raise MesRenderError(
                f"{record['question_label']}: more than "
                f"{MAX_DESCRIPTIVE_ANSWERS} answer blocks")
        for n, answer in enumerate(answers, start=1):
            record[f"answer_type_{n}"] = answer.get("answer_type", "")
            record[f"answer_weightage_{n}"] = answer.get(
                "answer_weightage", "")
            record[f"answer_content_{n}"] = answer.get("answer_content", "")
        sub_questions = [
            s for s in candidate.get("sub_questions") or []
            if isinstance(s, Mapping)
        ]
        if len(sub_questions) > MAX_SUBQUESTIONS:
            raise MesRenderError(
                f"{record['question_label']}: more than "
                f"{MAX_SUBQUESTIONS} subquestions")
        for n, sub in enumerate(sub_questions, start=1):
            record[f"sub_question_{n}"] = sub.get("text", "")
            record[f"sub_question_marks_{n}"] = sub.get("marks", "")
            keywords = [
                k for k in sub.get("keywords") or [] if isinstance(k, Mapping)
            ]
            if len(keywords) > MAX_SUBQUESTION_KEYWORDS:
                raise MesRenderError(
                    f"{record['question_label']}: subquestion {n} has more "
                    f"than {MAX_SUBQUESTION_KEYWORDS} keyword slots")
            for m, keyword in enumerate(keywords, start=1):
                record[f"sq{n}_answer_type_{m}"] = keyword.get(
                    "answer_type", "")
                record[f"sq{n}_weightage_{m}"] = keyword.get("weightage", "")
                record[f"sq{n}_keyword_{m}"] = keyword.get("keyword", "")
    return record


def _group_record_fields(group: Mapping, group_labels: list[str]) -> dict:
    return {
        "group_name": group.get("group_name", ""),
        "group_display_name": group.get("group_display_name", ""),
        "group_description": group.get("semantic_description", ""),
        "group_status": group.get("group_status", "Active"),
        "group_type": group.get("group_type", ""),
        "group_question_labels": ", ".join(group_labels),
        "related_digicards": "",
    }


def render_master_file(snapshot: Mapping) -> tuple[bytes, dict]:
    """Render the Master File; returns (bytes, issues manifest).

    Nothing disappears: a candidate whose placement never resolved to a
    concept/group cannot be positioned in the hierarchy, so it is recorded
    in the issues manifest (spec §13.3 keeps the full ledger in the release
    manifest, never an unrecognized extra sheet).
    """
    concept_entries = {
        str(entry["concept"].get("concept_key") or ""): entry
        for entry in _concept_rows(snapshot)
    }
    groups = [dict(g) for g in snapshot.get("groups") or []]
    groups_by_key = {str(g.get("group_key") or ""): g for g in groups}
    candidates = [dict(c) for c in snapshot.get("candidates") or []]

    # Complete ordered label aggregates (spec §8.4).
    concept_labels: dict[str, list[str]] = {}
    group_labels: dict[str, list[str]] = {}
    unplaced: list[dict] = []
    placed: list[dict] = []
    for candidate in candidates:
        concept_key = str(candidate.get("concept_key") or "")
        group_key = str(candidate.get("group_key") or "")
        label = str(candidate.get("question_label") or "")
        if (
            concept_key not in concept_entries
            or group_key not in groups_by_key
            or not label
        ):
            unplaced.append({
                "candidate_id": str(candidate.get("candidate_id") or ""),
                "question_label": label,
                "reason": "unresolved home concept/group placement",
                "flags": list(candidate.get("flags") or []),
            })
            continue
        concept_labels.setdefault(concept_key, []).append(label)
        group_labels.setdefault(group_key, []).append(label)
        placed.append(candidate)

    wb = _new_workbook()
    represented_groups: set[str] = set()

    def _full_record(candidate: Mapping, sheet: str) -> dict:
        concept_key = str(candidate.get("concept_key") or "")
        group_key = str(candidate.get("group_key") or "")
        entry = concept_entries[concept_key]
        record = _bands_record(entry)
        record["concept_question_labels"] = ", ".join(
            concept_labels.get(concept_key, []))
        record.update(_group_record_fields(
            groups_by_key[group_key], group_labels.get(group_key, [])))
        record.update(_question_record(candidate, sheet))
        return record

    sheet_for_kind = {"objective": "Objective", "descriptive": "Descriptive"}
    question_rows = 0
    for candidate in placed:
        kind = str(candidate.get("sheet_kind") or "")
        sheet = sheet_for_kind.get(kind)
        if sheet is None:
            # MES never emits Subjective data rows (spec §3.1).
            unplaced.append({
                "candidate_id": str(candidate.get("candidate_id") or ""),
                "question_label": str(candidate.get("question_label") or ""),
                "reason": f"sheet kind {kind!r} has no MES data sheet",
                "flags": list(candidate.get("flags") or []),
            })
            continue
        ws = wb[sheet]
        ws.append(_row_values(sheet, _full_record(candidate, sheet)))
        question_rows += 1
        represented_groups.add(str(candidate.get("group_key") or ""))

    # Group catalogue rows: every created group not otherwise represented by
    # a question row — required empty shells included — appears once on the
    # Objective sheet (the catalogue carrier), Question band blank. This is
    # also what keeps questionless concepts in the Master (spec §10).
    for group in groups:
        group_key = str(group.get("group_key") or "")
        if group_key in represented_groups:
            continue
        concept_key = str(group.get("concept_key") or "")
        entry = concept_entries.get(concept_key)
        if entry is None:
            unplaced.append({
                "candidate_id": "",
                "question_label": "",
                "reason": (
                    f"group {group_key} references unknown concept "
                    f"{concept_key!r}"),
                "flags": list(group.get("flags") or []),
            })
            continue
        record = _bands_record(entry)
        record["concept_question_labels"] = ", ".join(
            concept_labels.get(concept_key, []))
        record.update(_group_record_fields(
            group, group_labels.get(group_key, [])))
        wb["Objective"].append(_row_values("Objective", record))

    issues = {
        "unplaced": unplaced,
        "placed_questions": question_rows,
        "groups": len(groups),
    }
    return _workbook_bytes(wb), issues


# --------------------------------------------------------------------------- #
# Read-back parsing and validation (spec §13.2 steps 2–4)
# --------------------------------------------------------------------------- #

def parse_workbook(data: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                data_only=True)
    parsed: dict = {"sheet_order": list(wb.sheetnames), "sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        next(rows, ())  # band row
        header = [h for h in next(rows, ()) if h is not None]
        records = []
        for row in rows:
            if row is None or not any(row):
                continue
            records.append({
                field: ("" if i >= len(row) or row[i] is None else row[i])
                for i, field in enumerate(header)
            })
        parsed["sheets"][name] = {"fields": header, "rows": records}
    wb.close()
    return parsed


def _header_errors(parsed: Mapping) -> list[str]:
    errors = []
    if parsed.get("sheet_order") != SHEET_ORDER:
        errors.append(
            f"sheet order {parsed.get('sheet_order')} != {SHEET_ORDER}")
    for name in SHEET_ORDER:
        sheet = (parsed.get("sheets") or {}).get(name)
        if sheet is None:
            errors.append(f"missing sheet {name!r}")
            continue
        if sheet["fields"] != FIELDS[name]:
            errors.append(f"{name}: header row differs from the template")
    return errors


def validate_concept_file(parsed: Mapping, snapshot: Mapping) -> list[str]:
    errors = _header_errors(parsed)
    if errors:
        return errors
    rows = parsed["sheets"]["Objective"]["rows"]
    expected = [
        str(e["concept"].get("concept_title") or "")
        for e in _concept_rows(snapshot)
    ]
    actual = [str(r.get("concept_title") or "") for r in rows]
    if actual != expected:
        errors.append(
            f"concept rows differ: expected {len(expected)}, "
            f"got {len(actual)} (or out of source order)")
    for i, row in enumerate(rows, start=3):
        populated = [
            field for field in (
                "group_name", "group_type", "question_label", "question",
                "basic_groups", "intermediate_groups", "advanced_groups",
                "concept_question_labels", "chapter_duration",
            )
            if str(row.get(field) or "").strip()
        ]
        if populated:
            errors.append(
                f"Objective row {i}: Concept File must keep {populated} "
                "blank")
    for name in ("Descriptive", "Subjective"):
        if parsed["sheets"][name]["rows"]:
            errors.append(f"{name}: Concept File must be header-only")
    return errors


def validate_master_file(parsed: Mapping, snapshot: Mapping) -> list[str]:
    errors = _header_errors(parsed)
    if errors:
        return errors
    if parsed["sheets"]["Subjective"]["rows"]:
        errors.append("Subjective: MES must contain no data rows")

    groups = [dict(g) for g in snapshot.get("groups") or []]
    expected_group_keys = {str(g.get("group_key") or "") for g in groups}
    seen_group_names: set[str] = set()
    seen_concepts: set[str] = set()
    question_rows = 0
    aggregate_by_group: dict[str, set[str]] = {}
    for name in ("Objective", "Descriptive"):
        for i, row in enumerate(parsed["sheets"][name]["rows"], start=3):
            seen_concepts.add(str(row.get("concept_title") or ""))
            group_name = str(row.get("group_name") or "")
            if group_name:
                seen_group_names.add(group_name)
                aggregate_by_group.setdefault(group_name, set()).add(
                    str(row.get("group_question_labels") or ""))
            if str(row.get("chapter_duration") or "").strip():
                errors.append(f"{name} row {i}: chapter_duration must be blank")
            if str(row.get("question_disclaimer") or "").strip():
                errors.append(
                    f"{name} row {i}: question_disclaimer must be blank")
            label = str(row.get("question_label") or "")
            if not label:
                continue
            question_rows += 1
            appears = str(row.get("question_appears_in") or "")
            if appears != "Pre/Post-Worksheet/Test":
                errors.append(
                    f"{label}: question_appears_in {appears!r} is not the "
                    "MES wire value")
            restriction = str(row.get("answer_restriction") or "")
            if restriction not in rel.ANSWER_RESTRICTIONS:
                errors.append(
                    f"{label}: answer_restriction {restriction!r} invalid")
            if name == "Objective":
                marks = float(row.get("marks") or 0)
                correct_weight = 0.0
                correct_count = 0
                for n in range(1, MAX_OBJECTIVE_OPTIONS + 1):
                    if str(row.get(f"correct_answer_{n}") or "").strip() in {
                        "1", "true", "True",
                    }:
                        correct_count += 1
                        try:
                            correct_weight = float(
                                row.get(f"answer_weightage_{n}") or 0)
                        except (TypeError, ValueError):
                            correct_weight = -1.0
                if correct_count != 1:
                    errors.append(
                        f"{label}: {correct_count} correct options")
                elif abs(correct_weight - marks) > 0.01:
                    errors.append(
                        f"{label}: correct weightage {correct_weight:g} != "
                        f"marks {marks:g}")

    # Every concept (questionless included) and every created group appears.
    expected_concepts = {
        str(e["concept"].get("concept_title") or "")
        for e in _concept_rows(snapshot)
    }
    for concept_title in sorted(expected_concepts - seen_concepts):
        errors.append(f"concept missing from Master: {concept_title!r}")
    expected_group_names = {
        str(g.get("group_name") or "") for g in groups
    }
    for group_name in sorted(expected_group_names - seen_group_names):
        errors.append(f"group missing from Master: {group_name!r}")
    del expected_group_keys
    for group_name, aggregates in aggregate_by_group.items():
        if len(aggregates) > 1:
            errors.append(
                f"{group_name}: group_question_labels differs across "
                "repeated rows")
    return errors


# --------------------------------------------------------------------------- #
# Dual projection (spec §13.1–§13.2)
# --------------------------------------------------------------------------- #

def build_dual_output(snapshot: Mapping) -> dict:
    """Render, read back, validate, and hash both projections of one
    snapshot. Returns concepts/master bytes plus the shared manifest."""
    snapshot_hash = snapshot_sha256(snapshot)
    concepts_bytes = render_concept_file(snapshot)
    master_bytes, issues = render_master_file(snapshot)
    concept_errors = validate_concept_file(
        parse_workbook(concepts_bytes), snapshot)
    master_errors = validate_master_file(
        parse_workbook(master_bytes), snapshot)
    manifest = {
        "profile": "mes-1",
        "template_source": MANIFEST.get("source", ""),
        "concept_snapshot_sha256": snapshot_hash,
        "workbook_sha256s": {
            "concepts_xlsx": hashlib.sha256(concepts_bytes).hexdigest(),
            "master_xlsx": hashlib.sha256(master_bytes).hexdigest(),
        },
        "read_back": {
            "concepts_errors": concept_errors,
            "master_errors": master_errors,
        },
        "issues": issues,
    }
    return {
        "concepts_xlsx": concepts_bytes,
        "master_xlsx": master_bytes,
        "manifest": manifest,
        "valid": not concept_errors and not master_errors,
    }
