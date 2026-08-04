"""Release-first publication boundary for Build Concepts.

Generation is allowed to be ambitious, semantic, and expensive. Publication is
not allowed to be implicit. This module turns every uploaded-source concept run
into a portable review release before any concept reaches the database.

The release contains:

* a review workbook with row-level status/error columns and highlighted rows;
* the complete generation log and terminal exception context;
* the converted MMD and original uploaded source;
* every available canonical source block, source-evidence packet and BLK binding;
* the portable generation checkpoint, inventory, mined Types/Cases and QID routes;
* the semantic graph, canonical source bundle and governing placement rules; and
* a content-addressed manifest plus a ZIP containing the whole issue context.

A separate explicit action publishes an accepted release. That action re-runs
Aegis's ordinary deterministic deposit gates and performs no generation request.
"""
from __future__ import annotations

import copy
import hashlib
import json
import re
import shutil
import traceback
import uuid
import zipfile
from contextvars import ContextVar
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import config, models
from . import build_concepts
from . import canonical_source_phase2 as phase2
from . import canonical_source_phase3 as phase3
from . import concept_refiner
from . import concept_validator
from . import drive_checkpoints
from . import generation
from . import grounding_certificate
from . import progress
from . import uploads
from ..bulk_import import writer as bulk_writer
from ..bulk_import import workbook_sync


RELEASE_VERSION = "concept-review-release-1"
RELEASE_METADATA_KEY = "_concept_release"
DIAGNOSTIC_METADATA_KEY = "_concept_diagnostic_export"
_RELEASE_ROOT = config.DATA_DIR / "concept-releases"
_RELEASE_CONTEXT: ContextVar["_ReleaseContext | None"] = ContextVar(
    "aegis_concept_release_context", default=None
)
_ORIGINAL_DEPOSIT = build_concepts._deposit_and_publish_concepts
_INSTALLED = False

_SECRET_KEY_RE = re.compile(
    r"(?:api[_-]?key|app[_-]?key|authorization|password|session[_-]?token|"
    r"access[_-]?token|refresh[_-]?token|client[_-]?secret)$",
    re.IGNORECASE,
)
_BLOCK_KEY_RE = re.compile(r"(?:^|_)(?:block|blk)(?:_ids?|s)?$", re.IGNORECASE)
_EVIDENCE_KEY_RE = re.compile(r"evidence", re.IGNORECASE)
_QID_RE = re.compile(r"\bQINV-[A-Za-z0-9_.:-]+\b")
_EXCEL_CELL_LIMIT = 32_000

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_READY_FILL = PatternFill("solid", fgColor="D9EAD3")
_INFO_FILL = PatternFill("solid", fgColor="D9EAF7")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9E1F2"),
    right=Side(style="thin", color="D9E1F2"),
    top=Side(style="thin", color="D9E1F2"),
    bottom=Side(style="thin", color="D9E1F2"),
)


@dataclass
class _ReleaseContext:
    db: Session
    job_id: int
    target_chapter_id: int
    learning_kind: str
    owner_sub: str | None
    captured: dict[str, Any] = field(default_factory=dict)


def install() -> None:
    """Install the no-write deposit interceptor used only inside release runs."""

    global _INSTALLED
    if _INSTALLED:
        return
    build_concepts._deposit_and_publish_concepts = _intercept_deposit_and_publish
    _INSTALLED = True


def _json_safe(value: Any) -> Any:
    """Detach runtime objects and redact credentials before diagnostic export."""

    if isinstance(value, Mapping):
        result: dict[str, Any] = {}
        for raw_key, raw_value in value.items():
            key = str(raw_key)
            if _SECRET_KEY_RE.search(key):
                result[key] = "[REDACTED]"
            else:
                result[key] = _json_safe(raw_value)
        return result
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if isinstance(value, set):
        return sorted((_json_safe(item) for item in value), key=str)
    if isinstance(value, Path):
        return value.as_posix()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _write_json(path: Path, value: Any) -> None:
    path.write_text(
        json.dumps(_json_safe(value), ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _write_text(path: Path, value: object) -> None:
    path.write_text(str(value or ""), encoding="utf-8")


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _release_id(job_id: int, *, diagnostic: bool) -> str:
    prefix = "diagnostic" if diagnostic else "release"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return f"{prefix}-{job_id}-{stamp}-{uuid.uuid4().hex[:8]}"


def _release_directory(job_id: int, release_id: str) -> Path:
    path = _RELEASE_ROOT / str(job_id) / release_id
    path.mkdir(parents=True, exist_ok=False)
    return path


def _active_source_context() -> dict[str, Any]:
    graph: Any = None
    canonical: Any = None
    session: Any = None
    try:
        graph = phase3.active_graph()
        session = phase3.active_session()
        if isinstance(session, Mapping):
            canonical = session.get("canonical")
            graph = graph or session.get("graph")
    except Exception:
        graph = None
        canonical = None
        session = None
    if not isinstance(canonical, Mapping):
        try:
            canonical = phase2.active_canonical()
        except Exception:
            canonical = None
    return {
        "graph": _json_safe(graph) if isinstance(graph, Mapping) else {},
        "canonical": (
            _json_safe(canonical) if isinstance(canonical, Mapping) else {}
        ),
        "session": _json_safe(session) if isinstance(session, Mapping) else {},
    }


def _intercept_deposit_and_publish(
    db: Session,
    *,
    chapter_id: int,
    records: list[dict],
    pre_post: str,
    source_book: str,
    inventory: dict | None = None,
    mined_types: dict | None = None,
    source_text: str = "",
    final_grounding_certificate: dict | None = None,
    grounding_audit_job: models.UploadJob | None = None,
) -> tuple[list[int], list[int], dict]:
    """Capture a validated candidate and deliberately perform zero DB writes."""

    context = _RELEASE_CONTEXT.get()
    if context is None:
        return _ORIGINAL_DEPOSIT(
            db,
            chapter_id=chapter_id,
            records=records,
            pre_post=pre_post,
            source_book=source_book,
            inventory=inventory,
            mined_types=mined_types,
            source_text=source_text,
            final_grounding_certificate=final_grounding_certificate,
            grounding_audit_job=grounding_audit_job,
        )

    job = grounding_audit_job or db.get(models.UploadJob, context.job_id)
    context.captured = {
        "records": copy.deepcopy(records),
        "inventory": copy.deepcopy(inventory or {}),
        "mined_types": copy.deepcopy(mined_types or {}),
        "source_text": str(source_text or ""),
        "final_grounding_certificate": copy.deepcopy(
            final_grounding_certificate or {}
        ),
        "generation_checkpoint": copy.deepcopy(
            (job.generation_checkpoint if job is not None else {}) or {}
        ),
        "question_inventory": copy.deepcopy(
            (job.question_inventory if job is not None else {}) or {}
        ),
        "source_context": _active_source_context(),
        "chapter_id": int(chapter_id),
        "pre_post": str(pre_post),
        "source_book": str(source_book or ""),
    }
    progress.log(
        "Release boundary reached: the complete candidate was captured for "
        "review. No concept, topic, chapter or workbook row was written to the "
        "database.",
        level="success",
    )
    written: dict[str, Any] = {
        "written": len(records),
        "sources_updated": 0,
        "parent_column": True,
        "release_intercepted": True,
    }
    if isinstance(final_grounding_certificate, dict):
        written["grounding_certificate"] = copy.deepcopy(
            final_grounding_certificate
        )
    return [], [], written


def _checkpoint_entries(checkpoint: Mapping[str, Any] | None) -> list[dict]:
    try:
        return [
            copy.deepcopy(row)
            for row in generation._concept_checkpoint_entries(checkpoint or {})
            if isinstance(row, dict)
        ]
    except Exception:
        return []


def _latest_materialized_payload(
    checkpoint: Mapping[str, Any] | None,
    question_inventory: Mapping[str, Any] | None,
) -> dict[str, Any]:
    entries = _checkpoint_entries(checkpoint)
    indexed = sorted(
        enumerate(entries),
        key=lambda pair: (
            generation._checkpoint_order(str(pair[1].get("stage") or "")),
            pair[0],
        ),
        reverse=True,
    )
    selected: dict[str, Any] = {}
    for _index, entry in indexed:
        if isinstance(entry.get("records"), list) and entry.get("records"):
            selected = entry
            break
    inventory = copy.deepcopy((question_inventory or {}).get("items") or [])
    stored_mined = copy.deepcopy(
        (question_inventory or {}).get("mined_types") or []
    )
    return {
        "records": copy.deepcopy(selected.get("records") or []),
        "inventory": copy.deepcopy(
            selected.get("question_task_inventory")
            or ({"items": inventory} if inventory else {})
        ),
        "mined_types": copy.deepcopy(
            selected.get("mined_types")
            or ({"types": stored_mined} if stored_mined else {})
        ),
        "final_grounding_certificate": copy.deepcopy(
            selected.get(grounding_certificate.FINAL_CERTIFICATE_FIELD)
            or (question_inventory or {}).get(
                grounding_certificate.FINAL_CERTIFICATE_FIELD
            )
            or {}
        ),
        "selected_checkpoint": copy.deepcopy(selected),
        "selected_stage": str(selected.get("stage") or ""),
    }


def _pending_decision(checkpoint: Mapping[str, Any] | None) -> dict[str, Any]:
    try:
        value = build_concepts._pending_human_decision(dict(checkpoint or {}))
    except Exception:
        value = None
    return copy.deepcopy(value) if isinstance(value, dict) else {}


def _issue(
    code: str,
    message: str,
    *,
    severity: str = "error",
    row_index: int | None = None,
    concept: str = "",
    field_name: str = "",
    snippet: str = "",
    qid: str = "",
    type_id: str = "",
    case_id: str = "",
    block_ids: Iterable[str] = (),
    source: str = "release",
) -> dict[str, Any]:
    return {
        "severity": severity,
        "code": str(code or "release_issue"),
        "message": str(message or "")[:8000],
        "row_index": row_index,
        "concept": str(concept or ""),
        "field": str(field_name or ""),
        "snippet": str(snippet or "")[:2000],
        "qid": str(qid or ""),
        "type_id": str(type_id or ""),
        "case_id": str(case_id or ""),
        "block_ids": [str(value) for value in block_ids if str(value)],
        "source": str(source or "release"),
    }


def _normalize_report_issues(report: Mapping[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for raw in report.get("errors") or []:
        if not isinstance(raw, Mapping):
            continue
        row_index: int | None
        try:
            row_index = int(raw.get("row_index"))
        except (TypeError, ValueError):
            row_index = None
        out.append(_issue(
            str(raw.get("code") or "concept_validation"),
            str(raw.get("message") or raw),
            severity=str(raw.get("severity") or "warning"),
            row_index=row_index,
            concept=str(raw.get("concept") or raw.get("concept_title") or ""),
            field_name=str(raw.get("field") or ""),
            snippet=str(raw.get("snippet") or ""),
            source="concept_validator",
        ))
    return out


def _record_qids(record: Mapping[str, Any]) -> list[str]:
    values: set[str] = set()
    for key, raw in record.items():
        if "qid" not in str(key).casefold() and key not in {
            "concept_details", "source_evidence"
        }:
            continue
        if isinstance(raw, str):
            values.update(_QID_RE.findall(raw))
        elif isinstance(raw, Iterable) and not isinstance(raw, (str, bytes, Mapping)):
            for value in raw:
                values.update(_QID_RE.findall(str(value)))
    return sorted(values)


def _placement_contract(raw: object) -> dict[str, Any]:
    if not isinstance(raw, Mapping):
        return {}
    try:
        certified = generation._certified_type_case_placement_contract(raw)
    except Exception:
        certified = None
    return copy.deepcopy(certified or dict(raw))


def _case_examples(case: Mapping[str, Any]) -> list[dict[str, Any]]:
    try:
        return [
            copy.deepcopy(row)
            for row in generation._case_examples(dict(case))
            if isinstance(row, dict)
        ]
    except Exception:
        raw = case.get("examples") or []
        return [copy.deepcopy(row) for row in raw if isinstance(row, Mapping)]


def _type_case_routes(
    mined_types: Mapping[str, Any] | None,
    inventory: Mapping[str, Any] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    types = [
        row for row in (mined_types or {}).get("types") or []
        if isinstance(row, Mapping)
    ]
    inventory_by_qid = {
        str(row.get("qid") or "").strip(): row
        for row in (inventory or {}).get("items") or []
        if isinstance(row, Mapping) and str(row.get("qid") or "").strip()
    }
    routes: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    seen: dict[str, list[tuple[str, str]]] = {}

    for type_index, mtype in enumerate(types, start=1):
        type_id = str(mtype.get("type_id") or f"TYPE-{type_index:04d}")
        type_title = str(mtype.get("type_title") or "").strip()
        type_definition = str(
            mtype.get("type_description")
            or mtype.get("task_pattern")
            or ""
        ).strip()
        if not type_title or not type_definition:
            issues.append(_issue(
                "type_definition_missing",
                "Every reusable Type must have a title and a method/pattern "
                "definition before database upload.",
                type_id=type_id,
            ))

        cases = [
            row for row in mtype.get("case_prompts") or []
            if isinstance(row, Mapping)
        ]
        if not cases:
            issues.append(_issue(
                "type_has_no_cases",
                "Reusable Type has no defined Case.",
                type_id=type_id,
            ))
        for case_index, case in enumerate(cases, start=1):
            case_id = str(case.get("case_id") or f"CASE-{case_index:04d}")
            case_title = str(case.get("case_title") or "").strip()
            case_definition = str(
                case.get("case_signature")
                or case.get("case_description")
                or case.get("task_pattern")
                or ""
            ).strip()
            if not case_title or not case_definition:
                issues.append(_issue(
                    "case_definition_missing",
                    "Every Case must have a title and a distinguishing "
                    "definition/signature.",
                    type_id=type_id,
                    case_id=case_id,
                ))

            examples = _case_examples(case)
            qids: list[str] = []
            for example in examples:
                qid = str(example.get("source_question_id") or "").strip()
                if qid and qid not in qids:
                    qids.append(qid)
            for raw_qid in case.get("source_question_ids") or []:
                qid = str(raw_qid or "").strip()
                if qid and qid not in qids:
                    qids.append(qid)
            direct_qid = str(case.get("source_question_id") or "").strip()
            if direct_qid and direct_qid not in qids:
                qids.append(direct_qid)
            if not qids:
                qids = [""]

            example_by_qid = {
                str(row.get("source_question_id") or "").strip(): row
                for row in examples
            }
            for qid in qids:
                example = example_by_qid.get(qid) or (
                    examples[0] if len(examples) == 1 else {}
                )
                contract = _placement_contract(
                    example.get("_type_case_placement_contract")
                    or case.get("_type_case_placement_contract")
                    or mtype.get("_type_case_placement_contract")
                    or (inventory_by_qid.get(qid) or {}).get(
                        "_type_case_placement_contract"
                    )
                )
                owner_topic_id = str(
                    contract.get("owner_topic_id")
                    or case.get("owner_topic_id")
                    or example.get("owner_topic_id")
                    or ""
                )
                owner_topic_title = str(
                    contract.get("owner_topic_title")
                    or case.get("topic_match_hint")
                    or mtype.get("topic_match_hint")
                    or (inventory_by_qid.get(qid) or {}).get("topic_hint")
                    or ""
                )
                route = {
                    "type_id": type_id,
                    "type_title": type_title,
                    "type_definition": type_definition,
                    "case_id": case_id,
                    "case_title": case_title,
                    "case_definition": case_definition,
                    "qid": qid,
                    "owner_topic_id": owner_topic_id,
                    "owner_topic_title": owner_topic_title,
                    "source_location_topic_ids": list(
                        contract.get("source_location_topic_ids")
                        or case.get("source_location_topic_ids")
                        or []
                    ),
                    "concept_match_hint": str(
                        case.get("concept_match_hint")
                        or mtype.get("concept_match_hint")
                        or ""
                    ),
                    "parent_concept_match_hint": str(
                        case.get("parent_concept_match_hint")
                        or mtype.get("parent_concept_match_hint")
                        or ""
                    ),
                    "example": str(
                        example.get("example_prompt")
                        or (inventory_by_qid.get(qid) or {}).get("raw_task")
                        or ""
                    ),
                    "evidence_block_ids": sorted({
                        str(block_id)
                        for relationship in contract.get("topic_relationships") or []
                        if isinstance(relationship, Mapping)
                        for block_id in relationship.get("evidence_block_ids") or []
                        if str(block_id)
                    }),
                    "placement_certified": bool(contract.get("certified")),
                }
                routes.append(route)
                if qid:
                    seen.setdefault(qid, []).append((type_id, case_id))
                    if not route["example"]:
                        issues.append(_issue(
                            "case_example_missing",
                            "The source-owned Example must appear below its Case.",
                            qid=qid,
                            type_id=type_id,
                            case_id=case_id,
                        ))

    inventory_qids = {
        str(row.get("qid") or "").strip()
        for row in (inventory or {}).get("items") or []
        if isinstance(row, Mapping) and str(row.get("qid") or "").strip()
    }
    for qid in sorted(inventory_qids - set(seen)):
        issues.append(_issue(
            "qid_not_assigned_to_type_case",
            "Inventory QID is not assigned to any Type/Case.",
            qid=qid,
        ))
    for qid, destinations in sorted(seen.items()):
        if len(destinations) > 1:
            issues.append(_issue(
                "duplicate_qid_assignment",
                "QID is assigned more than once across Types/Cases: "
                + ", ".join(f"{tid}/{cid}" for tid, cid in destinations),
                qid=qid,
            ))
    return routes, issues


def _case_example_order_issues(records: Sequence[Mapping[str, Any]]) -> list[dict]:
    issues: list[dict[str, Any]] = []
    token_re = re.compile(
        r"\b(Type\s+\d+|Case\s+\d+|(?:Worked\s+)?Example(?:\s+\d+)?)\s*:",
        re.IGNORECASE,
    )
    for row_index, record in enumerate(records):
        details = str(
            record.get("concept_details")
            or record.get("concept_description")
            or ""
        )
        for label, content in concept_refiner.split_sections(details):
            if not str(label).strip().casefold().startswith("type"):
                continue
            matches = list(token_re.finditer(content))
            for index, match in enumerate(matches):
                if not match.group(1).casefold().startswith("case"):
                    continue
                boundary = (
                    matches[index + 1].start()
                    if index + 1 < len(matches)
                    else len(content)
                )
                segment = content[match.end():boundary]
                if not re.search(
                    r"\b(?:Worked\s+)?Example(?:\s+\d+)?\s*:",
                    segment,
                    re.IGNORECASE,
                ):
                    issues.append(_issue(
                        "case_example_order",
                        "Case is not immediately followed by at least one "
                        "Example before the next Case/Type.",
                        row_index=row_index,
                        concept=str(record.get("concept_title") or ""),
                        field_name="concept_details",
                        snippet=content[match.start():boundary],
                    ))
    return issues


def _validation_issues(
    records: list[dict[str, Any]],
    *,
    inventory: Mapping[str, Any] | None,
    mined_types: Mapping[str, Any] | None,
    source_text: str,
    learning_kind: str,
    terminal_failure: Exception | None,
    pending_decision: Mapping[str, Any] | None,
    final_grounding_certificate: Mapping[str, Any] | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    if not records:
        issues.append(_issue(
            "no_materialized_concept_rows",
            "No concept rows were materialized before the run ended. The "
            "context package still contains the source, log and checkpoint.",
        ))
    else:
        try:
            report = concept_validator.validate_concept_rows(
                records,
                allow_types=True,
                require_culmination=False,
                allow_culmination=True,
                allowed_source_examples=generation._inventory_source_examples(
                    inventory
                ),
                source_text=source_text,
            )
            issues.extend(_normalize_report_issues(report))
        except Exception as exc:
            issues.append(_issue(
                "release_validation_exception",
                f"Validation report could not be completed: {type(exc).__name__}: {exc}",
            ))

    routes, route_issues = _type_case_routes(mined_types, inventory)
    issues.extend(route_issues)
    issues.extend(_case_example_order_issues(records))

    if records and inventory and (inventory.get("items") or []):
        try:
            coverage = generation._rendered_inventory_coverage_defects(
                records, dict(inventory)
            )
            for qid in coverage.get("missing") or []:
                issues.append(_issue(
                    "qid_missing_from_rendered_output",
                    "Inventory QID is absent from the rendered concept output.",
                    qid=str(qid),
                ))
            for qid in coverage.get("duplicate") or []:
                issues.append(_issue(
                    "qid_duplicated_in_rendered_output",
                    "Inventory QID appears more than once in the rendered "
                    "concept output.",
                    qid=str(qid),
                ))
        except Exception as exc:
            issues.append(_issue(
                "inventory_coverage_audit_exception",
                f"Exact inventory coverage audit could not finish: {exc}",
            ))
        try:
            violations = generation._rendered_inventory_topic_violations(
                records,
                dict(inventory),
                dict(mined_types or {}),
            )
            for violation in violations:
                if not isinstance(violation, Mapping):
                    continue
                issues.append(_issue(
                    "qid_topic_placement_violation",
                    str(violation.get("reason") or violation),
                    qid=str(violation.get("qid") or ""),
                    concept=str(violation.get("concept") or ""),
                    block_ids=violation.get("evidence_block_ids") or [],
                ))
        except Exception as exc:
            issues.append(_issue(
                "topic_placement_audit_exception",
                f"QID topic-placement audit could not finish: {exc}",
            ))

    if terminal_failure is not None:
        issues.append(_issue(
            "terminal_generation_failure",
            f"{type(terminal_failure).__name__}: {terminal_failure}",
            snippet="\n".join(traceback.format_exception(
                type(terminal_failure),
                terminal_failure,
                terminal_failure.__traceback__,
            )[-8:]),
            source="generation",
        ))
    if pending_decision:
        item = pending_decision.get("item")
        qids = (
            item.get("qids") if isinstance(item, Mapping) else []
        ) or pending_decision.get("qids") or []
        issues.append(_issue(
            "autonomous_semantic_issue_released",
            str(
                pending_decision.get("diagnosis")
                or pending_decision.get("conflict")
                or pending_decision.get("decision_question")
                or "A semantic decision remained unresolved; the run was "
                "released instead of asking the user to select an action."
            ),
            qid=", ".join(str(value) for value in qids if str(value)),
            type_id=str(
                item.get("type_id") if isinstance(item, Mapping) else ""
            ),
            source="semantic_resolution",
        ))

    if (
        learning_kind.casefold() == "post"
        and config.use_live_generation()
        and not isinstance(final_grounding_certificate, Mapping)
    ):
        issues.append(_issue(
            "final_grounding_certificate_missing",
            "Live Post Learning output has no final payload/evidence "
            "grounding certificate. It is reviewable but not database-ready.",
        ))

    deduped: list[dict[str, Any]] = []
    seen_issue: set[str] = set()
    for issue in issues:
        key = hashlib.sha256(json.dumps(
            _json_safe(issue), sort_keys=True, separators=(",", ":")
        ).encode("utf-8")).hexdigest()
        if key not in seen_issue:
            seen_issue.add(key)
            deduped.append(issue)

    error_rows = {
        int(issue["row_index"])
        for issue in deduped
        if issue.get("severity") == "error"
        and isinstance(issue.get("row_index"), int)
    }
    warning_rows = {
        int(issue["row_index"])
        for issue in deduped
        if issue.get("severity") != "error"
        and isinstance(issue.get("row_index"), int)
    } - error_rows
    summary = {
        "issue_count": len(deduped),
        "error_count": sum(
            1 for issue in deduped if issue.get("severity") == "error"
        ),
        "warning_count": sum(
            1 for issue in deduped if issue.get("severity") != "error"
        ),
        "error_row_count": len(error_rows),
        "warning_row_count": len(warning_rows),
        "ready_row_count": max(0, len(records) - len(error_rows) - len(warning_rows)),
    }
    return deduped, routes, summary


def _find_block_ids(value: Any) -> set[str]:
    found: set[str] = set()

    def walk(node: Any, key: str = "") -> None:
        if isinstance(node, Mapping):
            for raw_key, raw_value in node.items():
                walk(raw_value, str(raw_key))
            return
        if isinstance(node, Iterable) and not isinstance(node, (str, bytes)):
            for item in node:
                walk(item, key)
            return
        if not isinstance(node, str):
            return
        if _BLOCK_KEY_RE.search(key) or _EVIDENCE_KEY_RE.search(key):
            for token in re.findall(
                r"\b(?:BLK|BLOCK|FIG|TASK)-[A-Za-z0-9_.:-]+\b", node
            ):
                found.add(token)
        for token in re.findall(r"\bBLK-[A-Za-z0-9_.:-]+\b", node):
            found.add(token)

    walk(value)
    return found


def _canonical_blocks(source_context: Mapping[str, Any]) -> list[dict[str, Any]]:
    canonical = source_context.get("canonical")
    graph = source_context.get("graph")
    canonical_by_id = {
        str(row.get("block_id") or ""): row
        for row in (canonical or {}).get("blocks") or []
        if isinstance(row, Mapping) and str(row.get("block_id") or "")
    } if isinstance(canonical, Mapping) else {}
    rows: list[dict[str, Any]] = []
    for raw in (graph or {}).get("blocks") or [] if isinstance(graph, Mapping) else []:
        if not isinstance(raw, Mapping):
            continue
        block_id = str(raw.get("block_id") or "")
        canonical_row = canonical_by_id.get(block_id, {})
        rows.append({
            **copy.deepcopy(dict(raw)),
            "text": str(
                canonical_row.get("display_text")
                or canonical_row.get("raw_text")
                or raw.get("text")
                or ""
            ),
        })
    known = {str(row.get("block_id") or "") for row in rows}
    for block_id, raw in canonical_by_id.items():
        if block_id not in known:
            rows.append({
                **copy.deepcopy(dict(raw)),
                "block_id": block_id,
                "text": str(
                    raw.get("display_text") or raw.get("raw_text") or ""
                ),
            })
    return rows


def _source_evidence(
    *,
    records: Sequence[Mapping[str, Any]],
    inventory: Mapping[str, Any] | None,
    mined_types: Mapping[str, Any] | None,
    checkpoint: Mapping[str, Any] | None,
    pending_decision: Mapping[str, Any] | None,
    source_context: Mapping[str, Any],
) -> dict[str, Any]:
    material = {
        "records": records,
        "inventory": inventory or {},
        "mined_types": mined_types or {},
        "checkpoint": checkpoint or {},
        "pending_decision": pending_decision or {},
    }
    requested = _find_block_ids(material)
    blocks = _canonical_blocks(source_context)
    exact = [
        row for row in blocks
        if str(row.get("block_id") or "") in requested
    ]
    return {
        "referenced_block_ids": sorted(requested),
        "resolved_referenced_blocks": exact,
        "unresolved_block_ids": sorted(
            requested - {str(row.get("block_id") or "") for row in exact}
        ),
        "pending_decision_evidence": copy.deepcopy(
            (pending_decision or {}).get("evidence") or []
        ),
        "pending_decision_candidates": copy.deepcopy(
            (pending_decision or {}).get("candidates") or []
        ),
    }


def _log_text(events: Sequence[Mapping[str, Any]]) -> str:
    lines: list[str] = []
    for event in events:
        kind = str(event.get("type") or "log")
        level = str(event.get("level") or "")
        label = str(event.get("label") or "")
        message = str(event.get("message") or "")
        progress_value = event.get("value")
        prefix = " ".join(value for value in (kind, level) if value).upper()
        payload = message or label
        if progress_value is not None:
            payload = f"{payload} [{progress_value}]".strip()
        lines.append(f"{prefix}: {payload}".strip())
    return "\n".join(lines)


def _excel_cell(value: Any) -> str | int | float | bool:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (Mapping, list, tuple, set)):
        text = json.dumps(_json_safe(value), ensure_ascii=False, sort_keys=True)
    else:
        text = str(value)
    if len(text) > _EXCEL_CELL_LIMIT:
        return text[:_EXCEL_CELL_LIMIT - 40] + " … [TRUNCATED; full value in JSON]"
    return text


def _style_sheet(sheet, *, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
    for index, column in enumerate(sheet.columns, start=1):
        width = 12
        for cell in list(column)[:200]:
            width = max(width, min(60, len(str(cell.value or "")) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _concept_workbook(
    path: Path,
    *,
    records: list[dict[str, Any]],
    issues: list[dict[str, Any]],
    routes: list[dict[str, Any]],
    blocks: list[dict[str, Any]],
    events: list[dict[str, Any]],
    manifest_summary: Mapping[str, Any],
) -> None:
    workbook = Workbook()
    concepts = workbook.active
    concepts.title = "Concepts Review"
    concept_headers = [
        "row_index", "release_status", "error_count", "warning_count",
        "error_codes", "error_messages", "topic", "parent_concept",
        "concept_title", "concept_details", "keywords", "source_block_ids",
        "qids", "database_upload_allowed",
    ]
    concepts.append(concept_headers)
    by_row: dict[int, list[dict[str, Any]]] = {}
    for issue in issues:
        if isinstance(issue.get("row_index"), int):
            by_row.setdefault(int(issue["row_index"]), []).append(issue)
    for row_index, record in enumerate(records):
        row_issues = by_row.get(row_index, [])
        errors = [issue for issue in row_issues if issue.get("severity") == "error"]
        warnings = [issue for issue in row_issues if issue.get("severity") != "error"]
        status = (
            "ERROR - REVIEW REQUIRED" if errors else
            "WARNING - REVIEW ADVISED" if warnings else
            "READY FOR DATABASE UPLOAD"
        )
        concepts.append([
            row_index,
            status,
            len(errors),
            len(warnings),
            ", ".join(sorted({str(issue.get("code") or "") for issue in errors})),
            "\n".join(str(issue.get("message") or "") for issue in row_issues),
            _excel_cell(record.get("topic")),
            _excel_cell(record.get("parent_concept")),
            _excel_cell(record.get("concept_title") or record.get("concept")),
            _excel_cell(record.get("concept_details") or record.get("concept_description")),
            _excel_cell(record.get("keywords")),
            _excel_cell(
                record.get("_source_block_ids")
                or record.get(grounding_certificate.EVIDENCE_SET_FIELD)
                or []
            ),
            ", ".join(_record_qids(record)),
            bool(manifest_summary.get("database_upload_allowed")),
        ])
        fill = _ERROR_FILL if errors else _WARNING_FILL if warnings else _READY_FILL
        for cell in concepts[concepts.max_row]:
            cell.fill = fill

    if not records:
        concepts.append([
            "", "ERROR - NO MATERIALIZED ROWS", 1, 0,
            "no_materialized_concept_rows",
            "Generation ended before any concept row was materialized. See Issues, Run Log and context ZIP.",
            "", "", "", "", "", "", "", False,
        ])
        for cell in concepts[concepts.max_row]:
            cell.fill = _ERROR_FILL
    _style_sheet(concepts)

    issue_sheet = workbook.create_sheet("Issues")
    issue_headers = [
        "severity", "code", "row_index", "concept", "field", "qid",
        "type_id", "case_id", "block_ids", "message", "snippet", "source",
    ]
    issue_sheet.append(issue_headers)
    for issue in issues:
        issue_sheet.append([_excel_cell(issue.get(header)) for header in issue_headers])
        fill = _ERROR_FILL if issue.get("severity") == "error" else _WARNING_FILL
        for cell in issue_sheet[issue_sheet.max_row]:
            cell.fill = fill
    if not issues:
        issue_sheet.append(["info", "no_issues", "", "", "", "", "", "", "", "No release issues were detected.", "", "release"])
        for cell in issue_sheet[issue_sheet.max_row]:
            cell.fill = _READY_FILL
    _style_sheet(issue_sheet)

    route_sheet = workbook.create_sheet("QID Type Case Routes")
    route_headers = [
        "type_id", "type_title", "type_definition", "case_id", "case_title",
        "case_definition", "qid", "owner_topic_id", "owner_topic_title",
        "source_location_topic_ids", "concept_match_hint",
        "parent_concept_match_hint", "example", "evidence_block_ids",
        "placement_certified",
    ]
    route_sheet.append(route_headers)
    for route in routes:
        route_sheet.append([_excel_cell(route.get(header)) for header in route_headers])
    if not routes:
        route_sheet.append(["", "No Type/Case routes materialized"] + [""] * (len(route_headers) - 2))
    _style_sheet(route_sheet)

    block_sheet = workbook.create_sheet("Evidence BLKs")
    block_headers = [
        "block_id", "topic_id", "subtopic_id", "kind", "order",
        "source_start", "source_end", "page_number", "figure_id", "text",
    ]
    block_sheet.append(block_headers)
    for block in blocks:
        block_sheet.append([_excel_cell(block.get(header)) for header in block_headers])
    if not blocks:
        block_sheet.append(["No canonical BLKs available"] + [""] * (len(block_headers) - 1))
    _style_sheet(block_sheet)

    log_sheet = workbook.create_sheet("Run Log")
    log_headers = ["index", "type", "level", "label", "message", "progress", "timestamp", "error"]
    log_sheet.append(log_headers)
    for index, event in enumerate(events, start=1):
        log_sheet.append([
            index,
            _excel_cell(event.get("type")),
            _excel_cell(event.get("level")),
            _excel_cell(event.get("label")),
            _excel_cell(event.get("message")),
            _excel_cell(event.get("value")),
            _excel_cell(event.get("ts")),
            _excel_cell(event.get("error")),
        ])
    _style_sheet(log_sheet)

    manifest_sheet = workbook.create_sheet("Release Manifest")
    manifest_sheet.append(["field", "value"])
    for key, value in manifest_summary.items():
        manifest_sheet.append([key, _excel_cell(value)])
    _style_sheet(manifest_sheet)
    workbook.save(path)


def _safe_filename(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(value or "").name).strip("._")
    return cleaned or fallback


def _copy_original_source(job: models.UploadJob, release_dir: Path) -> str:
    try:
        source_path = uploads.upload_file_path(job)
    except Exception:
        return ""
    if not source_path.exists() or not source_path.is_file():
        return ""
    destination = release_dir / (
        "original_source__" + _safe_filename(job.filename, "source.bin")
    )
    shutil.copy2(source_path, destination)
    return destination.name


def _copy_rules(release_dir: Path) -> str:
    source = config.ROOT.parent / "docs" / "concept-placement-rules.md"
    if not source.exists():
        return ""
    destination = release_dir / "concept-placement-rules.md"
    shutil.copy2(source, destination)
    return destination.name


def _release_metadata(job: models.UploadJob) -> dict[str, Any]:
    inventory = job.question_inventory if isinstance(job.question_inventory, dict) else {}
    value = inventory.get(RELEASE_METADATA_KEY)
    return copy.deepcopy(value) if isinstance(value, Mapping) else {}


def _release_path(metadata: Mapping[str, Any], key: str) -> Path:
    relative = str(metadata.get(key) or "")
    if not relative:
        raise ValueError(f"release has no {key}")
    root = config.DATA_DIR.resolve()
    path = (config.DATA_DIR / relative).resolve()
    if root != path and root not in path.parents:
        raise ValueError("release path escapes the configured data directory")
    if not path.is_file():
        raise ValueError(f"release file is missing: {key}")
    return path


def _relative_data_path(path: Path) -> str:
    return path.resolve().relative_to(config.DATA_DIR.resolve()).as_posix()


def _release_urls(job_id: int) -> dict[str, str]:
    base = f"/build-concepts/releases/{job_id}"
    return {
        "release_summary_url": base,
        "release_workbook_url": base + "/workbook",
        "context_bundle_url": base + "/context",
        "release_manifest_url": base + "/manifest",
        "upload_to_database_url": base + "/upload-to-database",
    }


def _summary_response(job: models.UploadJob, metadata: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "job_id": job.id,
        "status": "released" if job.status != "generated" else "generated",
        "output_released": True,
        "database_written": job.status == "generated",
        "database_upload_allowed": bool(metadata.get("database_upload_allowed")),
        "release_id": str(metadata.get("release_id") or ""),
        "records_released": int(metadata.get("records_released") or 0),
        "issue_count": int(metadata.get("issue_count") or 0),
        "error_count": int(metadata.get("error_count") or 0),
        "warning_count": int(metadata.get("warning_count") or 0),
        "error_row_count": int(metadata.get("error_row_count") or 0),
        "warning_row_count": int(metadata.get("warning_row_count") or 0),
        "ready_row_count": int(metadata.get("ready_row_count") or 0),
        "terminal_failure": copy.deepcopy(metadata.get("terminal_failure") or {}),
        "selected_checkpoint_stage": str(
            metadata.get("selected_checkpoint_stage") or ""
        ),
        "inventory_items": int(metadata.get("inventory_items") or 0),
        "concept_ids": list(job.result_ids or []),
        **_release_urls(job.id),
    }


def create_release(
    db: Session,
    job: models.UploadJob,
    *,
    target_chapter_id: int,
    learning_kind: str,
    captured: Mapping[str, Any] | None = None,
    terminal_failure: Exception | None = None,
    pending_decision: Mapping[str, Any] | None = None,
    diagnostic: bool = False,
) -> dict[str, Any]:
    """Create one complete, self-contained review release or diagnostic ZIP."""

    captured = copy.deepcopy(dict(captured or {}))
    checkpoint = copy.deepcopy(
        captured.get("generation_checkpoint")
        or job.generation_checkpoint
        or {}
    )
    question_inventory = copy.deepcopy(
        captured.get("question_inventory")
        or job.question_inventory
        or {}
    )
    fallback = _latest_materialized_payload(checkpoint, question_inventory)
    records = copy.deepcopy(captured.get("records") or fallback["records"] or [])
    inventory = copy.deepcopy(
        captured.get("inventory") or fallback["inventory"] or {}
    )
    mined_types = copy.deepcopy(
        captured.get("mined_types") or fallback["mined_types"] or {}
    )
    final_certificate = copy.deepcopy(
        captured.get("final_grounding_certificate")
        or fallback["final_grounding_certificate"]
        or {}
    )
    source_text = str(captured.get("source_text") or job.mmd_text or "")
    pending = copy.deepcopy(
        pending_decision or _pending_decision(checkpoint) or {}
    )
    source_context = copy.deepcopy(
        captured.get("source_context") or _active_source_context()
    )
    events = copy.deepcopy(job.generation_log or [])
    release_id = _release_id(job.id, diagnostic=diagnostic)
    release_dir = _release_directory(job.id, release_id)

    issues, routes, issue_summary = _validation_issues(
        records,
        inventory=inventory,
        mined_types=mined_types,
        source_text=source_text,
        learning_kind=learning_kind,
        terminal_failure=terminal_failure,
        pending_decision=pending,
        final_grounding_certificate=(
            final_certificate if isinstance(final_certificate, Mapping) else None
        ),
    )
    source_evidence = _source_evidence(
        records=records,
        inventory=inventory,
        mined_types=mined_types,
        checkpoint=checkpoint,
        pending_decision=pending,
        source_context=source_context,
    )
    blocks = _canonical_blocks(source_context)
    database_upload_allowed = bool(
        records
        and not any(issue.get("severity") == "error" for issue in issues)
        and (
            learning_kind.casefold() != "post"
            or not config.use_live_generation()
            or isinstance(final_certificate, Mapping) and bool(final_certificate)
        )
    )
    terminal = (
        {
            "exception_type": type(terminal_failure).__name__,
            "reason": str(terminal_failure),
        }
        if terminal_failure is not None else {}
    )
    manifest: dict[str, Any] = {
        "version": RELEASE_VERSION,
        "release_id": release_id,
        "created_at": _iso_now(),
        "diagnostic_only": diagnostic,
        "job_id": job.id,
        "owner_sub": job.owner_sub,
        "learning_kind": learning_kind,
        "target_chapter_id": int(target_chapter_id or 0),
        "filename": job.filename,
        "source_book": job.source_book,
        "job_status_at_release": job.status,
        "records_released": len(records),
        "inventory_items": len((inventory or {}).get("items") or []),
        "type_count": len((mined_types or {}).get("types") or []),
        "qid_route_count": len([route for route in routes if route.get("qid")]),
        "selected_checkpoint_stage": str(
            captured.get("selected_stage")
            or fallback.get("selected_stage")
            or checkpoint.get("stage")
            or ""
        ),
        "database_written": False,
        "database_upload_allowed": database_upload_allowed,
        "terminal_failure": terminal,
        **issue_summary,
    }

    _write_json(release_dir / "records.json", records)
    _write_json(release_dir / "issues.json", issues)
    _write_json(release_dir / "qid_type_case_routes.json", routes)
    _write_json(release_dir / "question_inventory.json", inventory)
    _write_json(release_dir / "mined_types.json", mined_types)
    _write_json(release_dir / "generation_checkpoint.json", checkpoint)
    _write_json(release_dir / "pending_semantic_issue.json", pending)
    _write_json(release_dir / "generation_log.json", events)
    _write_text(release_dir / "generation_log.txt", _log_text(events))
    _write_text(release_dir / "converted_source.mmd", source_text)
    _write_json(release_dir / "semantic_graph.json", source_context.get("graph") or {})
    _write_json(release_dir / "canonical_source.json", source_context.get("canonical") or {})
    _write_json(release_dir / "source_session.json", source_context.get("session") or {})
    _write_json(release_dir / "source_blocks.json", blocks)
    _write_json(release_dir / "source_evidence.json", source_evidence)
    _write_json(release_dir / "final_grounding_certificate.json", final_certificate)
    original_source_name = _copy_original_source(job, release_dir)
    rules_name = _copy_rules(release_dir)

    workbook_name = "concepts_review.xlsx"
    _concept_workbook(
        release_dir / workbook_name,
        records=records,
        issues=issues,
        routes=routes,
        blocks=blocks,
        events=events,
        manifest_summary=manifest,
    )
    readme = (
        "PROJECT AEGIS CONCEPT REVIEW RELEASE\n\n"
        "This package was released before database publication. Open "
        "concepts_review.xlsx first. Problem rows are highlighted and each row "
        "has separate release_status, error_codes and error_messages columns.\n\n"
        "The QID Type Case Routes sheet preserves reusable Type identity while "
        "showing each Case's independently certified owner topic. A Type may "
        "therefore have Case 01 under one topic and Case 02 under another. A QID "
        "must occur exactly once.\n\n"
        "The ZIP also includes the full run log, checkpoint, MMD, original "
        "source, semantic graph, canonical source, every available BLK, source "
        "evidence, inventory, Types/Cases and placement rules.\n\n"
        "Nothing in this release has been written to the concept database. "
        "Use the separate Upload to database action only after review. Aegis "
        "will re-run all deterministic deposit gates before accepting it.\n"
    )
    _write_text(release_dir / "README.txt", readme)

    manifest["original_source_file"] = original_source_name
    manifest["placement_rules_file"] = rules_name
    manifest["files"] = []
    manifest_path = release_dir / "release_manifest.json"
    _write_json(manifest_path, manifest)
    file_entries = []
    for path in sorted(release_dir.iterdir(), key=lambda item: item.name):
        if not path.is_file() or path.name in {"release_manifest.json", "context_bundle.zip"}:
            continue
        file_entries.append({
            "name": path.name,
            "size_bytes": path.stat().st_size,
            "sha256": _sha256_file(path),
        })
    manifest["files"] = file_entries
    _write_json(manifest_path, manifest)

    zip_path = release_dir / "context_bundle.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(release_dir.iterdir(), key=lambda item: item.name):
            if path.is_file() and path != zip_path:
                archive.write(path, arcname=path.name)

    metadata = {
        **manifest,
        "release_dir": _relative_data_path(release_dir),
        "workbook_path": _relative_data_path(release_dir / workbook_name),
        "context_bundle_path": _relative_data_path(zip_path),
        "manifest_path": _relative_data_path(manifest_path),
        "records_path": _relative_data_path(release_dir / "records.json"),
        "inventory_path": _relative_data_path(release_dir / "question_inventory.json"),
        "mined_types_path": _relative_data_path(release_dir / "mined_types.json"),
        "source_text_path": _relative_data_path(release_dir / "converted_source.mmd"),
        "certificate_path": _relative_data_path(
            release_dir / "final_grounding_certificate.json"
        ),
    }
    return metadata


def _store_release(
    db: Session,
    job: models.UploadJob,
    metadata: Mapping[str, Any],
    *,
    target_chapter_id: int,
) -> None:
    db.refresh(job)
    inventory = copy.deepcopy(
        job.question_inventory if isinstance(job.question_inventory, dict) else {}
    )
    inventory[RELEASE_METADATA_KEY] = copy.deepcopy(dict(metadata))
    inventory.pop(DIAGNOSTIC_METADATA_KEY, None)
    job.question_inventory = inventory
    job.status = "released"
    job.result_ids = []
    job.deposit_scope_type = "chapter"
    job.deposit_scope_ids = [int(target_chapter_id)] if target_chapter_id else []
    job.generation_checkpoint = {}
    job.detail = (
        f"Review release {metadata.get('release_id')} created with "
        f"{metadata.get('records_released', 0)} row(s), "
        f"{metadata.get('error_count', 0)} error(s) and "
        f"{metadata.get('warning_count', 0)} warning(s). No database rows "
        "were written."
    )
    db.commit()
    db.refresh(job)
    drive_checkpoints.schedule_checkpoint_backup(job.id)


def _ensure_generation_log(
    db: Session,
    job_id: int,
    *,
    owner_sub: str | None,
    error: Exception | None,
) -> models.UploadJob:
    try:
        uploads.persist_current_generation_log(
            db, job_id, owner_sub=owner_sub, error=error
        )
    except Exception:
        db.rollback()
    return uploads.get_job(
        db, job_id, owner_sub=owner_sub, module="build_concepts"
    )


def _existing_release_response(
    db: Session,
    job_id: int,
    *,
    owner_sub: str | None,
) -> dict[str, Any] | None:
    job = uploads.get_job(
        db, job_id, owner_sub=owner_sub, module="build_concepts"
    )
    metadata = _release_metadata(job)
    if job.status == "released" and metadata:
        return _summary_response(job, metadata)
    return None


def _generate_release(
    db: Session,
    job_id: int,
    target_chapter_id: int,
    *,
    learning_kind: str,
    owner_sub: str | None,
) -> dict[str, Any]:
    install()
    existing = _existing_release_response(
        db, job_id, owner_sub=owner_sub
    )
    if existing is not None:
        progress.log(
            "Reusing the already-created review release; no generation or "
            "database request was repeated.",
            level="success",
        )
        return existing

    job = uploads.get_job(
        db,
        job_id,
        owner_sub=owner_sub,
        module="build_concepts",
        learning_kind=learning_kind,
    )
    context = _ReleaseContext(
        db=db,
        job_id=job_id,
        target_chapter_id=target_chapter_id,
        learning_kind=learning_kind,
        owner_sub=owner_sub,
    )
    token = _RELEASE_CONTEXT.set(context)
    result: Any = None
    failure: Exception | None = None
    pending: dict[str, Any] = {}
    try:
        try:
            if learning_kind == "post":
                result = build_concepts.generate_post_learning(
                    db,
                    job_id,
                    target_chapter_id,
                    owner_sub=owner_sub,
                )
            else:
                result = build_concepts.generate_pre_learning_from_upload(
                    db,
                    job_id,
                    target_chapter_id,
                    owner_sub=owner_sub,
                )
            if isinstance(result, Mapping) and result.get("pending_decision"):
                pending = copy.deepcopy(dict(result["pending_decision"]))
        except Exception as exc:
            failure = exc
            db.rollback()
            try:
                db.refresh(job)
            except Exception:
                pass
    finally:
        _RELEASE_CONTEXT.reset(token)

    progress.set_progress(1.0, label="Review release ready")
    progress.log(
        "Generation has reached the release boundary. Any unresolved semantic "
        "issue is being carried into the workbook and context bundle instead "
        "of being shown as a manual selection screen.",
        level="warning" if failure or pending else "success",
    )
    job = _ensure_generation_log(
        db,
        job_id,
        owner_sub=owner_sub,
        error=failure,
    )
    captured = copy.deepcopy(context.captured)
    if not captured:
        fallback = _latest_materialized_payload(
            job.generation_checkpoint,
            job.question_inventory,
        )
        captured = {
            **fallback,
            "generation_checkpoint": copy.deepcopy(
                job.generation_checkpoint or {}
            ),
            "question_inventory": copy.deepcopy(job.question_inventory or {}),
            "source_text": job.mmd_text,
            "source_context": _active_source_context(),
        }
    if not pending:
        pending = _pending_decision(
            captured.get("generation_checkpoint")
            or job.generation_checkpoint
            or {}
        )
    metadata = create_release(
        db,
        job,
        target_chapter_id=target_chapter_id,
        learning_kind=learning_kind,
        captured=captured,
        terminal_failure=failure,
        pending_decision=pending,
        diagnostic=False,
    )
    _store_release(
        db, job, metadata, target_chapter_id=target_chapter_id
    )
    progress.log(
        "Review workbook and complete context ZIP released. Database upload "
        "remains a separate explicit action.",
        level="success",
    )
    return _summary_response(job, metadata)


def generate_post_learning_release(
    db: Session,
    job_id: int,
    target_chapter_id: int,
    *,
    owner_sub: str | None = None,
) -> dict[str, Any]:
    return _generate_release(
        db,
        job_id,
        target_chapter_id,
        learning_kind="post",
        owner_sub=owner_sub,
    )


def generate_pre_learning_release(
    db: Session,
    job_id: int,
    target_chapter_id: int,
    *,
    owner_sub: str | None = None,
) -> dict[str, Any]:
    return _generate_release(
        db,
        job_id,
        target_chapter_id,
        learning_kind="pre",
        owner_sub=owner_sub,
    )


def release_summary(
    db: Session,
    job_id: int,
    *,
    owner_sub: str | None = None,
) -> dict[str, Any]:
    job = uploads.get_job(
        db, job_id, owner_sub=owner_sub, module="build_concepts"
    )
    metadata = _release_metadata(job)
    if not metadata:
        raise ValueError("no review release exists for this job")
    return _summary_response(job, metadata)


def release_file(
    db: Session,
    job_id: int,
    key: str,
    *,
    owner_sub: str | None = None,
    create_diagnostic: bool = False,
) -> tuple[Path, str]:
    job = uploads.get_job(
        db, job_id, owner_sub=owner_sub, module="build_concepts"
    )
    metadata = _release_metadata(job)
    if not metadata and create_diagnostic:
        job = _ensure_generation_log(
            db, job_id, owner_sub=owner_sub, error=None
        )
        target_ids = list(job.deposit_scope_ids or [])
        checkpoint_target = (
            (job.generation_checkpoint or {}).get("target_chapter_id")
            if isinstance(job.generation_checkpoint, Mapping) else None
        )
        metadata = create_release(
            db,
            job,
            target_chapter_id=int(
                target_ids[0] if target_ids else checkpoint_target or 0
            ),
            learning_kind=job.learning_kind,
            captured={},
            terminal_failure=None,
            pending_decision=_pending_decision(job.generation_checkpoint),
            diagnostic=True,
        )
        inventory = copy.deepcopy(
            job.question_inventory if isinstance(job.question_inventory, dict) else {}
        )
        inventory[DIAGNOSTIC_METADATA_KEY] = metadata
        job.question_inventory = inventory
        db.commit()
        db.refresh(job)
    if not metadata:
        diagnostic = (
            job.question_inventory.get(DIAGNOSTIC_METADATA_KEY)
            if isinstance(job.question_inventory, Mapping) else None
        )
        metadata = copy.deepcopy(diagnostic) if isinstance(diagnostic, Mapping) else {}
    path = _release_path(metadata, key)
    return path, str(metadata.get("release_id") or f"job-{job_id}")


def _load_release_json(metadata: Mapping[str, Any], key: str) -> Any:
    path = _release_path(metadata, key)
    return json.loads(path.read_text(encoding="utf-8"))


def publish_release(
    db: Session,
    job_id: int,
    *,
    owner_sub: str | None = None,
) -> dict[str, Any]:
    """Explicitly publish one clean release after all deterministic gates pass."""

    install()
    with uploads.exclusive_job_operation(job_id):
        job = uploads.get_job(
            db, job_id, owner_sub=owner_sub, module="build_concepts"
        )
        db.refresh(job)
        if job.status == "generated":
            metadata = _release_metadata(job)
            return {
                "job_id": job.id,
                "status": "generated",
                "database_written": True,
                "concept_ids": list(job.result_ids or []),
                "records_released": int(metadata.get("records_released") or 0),
                **_release_urls(job.id),
            }
        if job.status != "released":
            raise ValueError("generate and release the review workbook first")
        metadata = _release_metadata(job)
        if not metadata:
            raise ValueError("release metadata is missing")
        if not bool(metadata.get("database_upload_allowed")):
            raise build_concepts.DepositValidationError(
                "this release contains errors and is not database-ready; "
                "review the highlighted rows and Issues sheet"
            )

        records = _load_release_json(metadata, "records_path")
        inventory = _load_release_json(metadata, "inventory_path")
        mined_types = _load_release_json(metadata, "mined_types_path")
        certificate = _load_release_json(metadata, "certificate_path")
        source_text = _release_path(metadata, "source_text_path").read_text(
            encoding="utf-8"
        )
        target_chapter_id = int(metadata.get("target_chapter_id") or 0)
        if not target_chapter_id:
            raise ValueError("release has no destination chapter")

        certificate_sink: dict[str, Any] = {}
        with workbook_sync.output_workbook_lock():
            db.expire_all()
            chapter = db.get(models.Chapter, target_chapter_id)
            if chapter is None:
                raise ValueError("target chapter not found")
            created_ids, merged_ids = build_concepts._deposit_concepts(
                db,
                chapter,
                records,
                "Post" if job.learning_kind == "post" else "Pre",
                build_concepts._job_source_label(job),
                inventory=inventory if job.learning_kind == "post" else None,
                mined_types=mined_types if job.learning_kind == "post" else None,
                source_text=source_text,
                final_grounding_certificate=(
                    certificate
                    if job.learning_kind == "post" and certificate else None
                ),
                grounding_certificate_sink=certificate_sink,
            )
            active_ids = set(created_ids + merged_ids)
            build_concepts._sync_chapter_topic_summary(
                chapter,
                {},
                active_concept_ids=active_ids,
                pre_post="Post" if job.learning_kind == "post" else "Pre",
            )
            written = build_concepts._commit_and_publish_concept_workbook(
                db,
                config.BULK_IMPORT_OUTPUT,
                created_ids + merged_ids,
            )

        inventory_audit = copy.deepcopy(
            job.question_inventory if isinstance(job.question_inventory, dict) else {}
        )
        metadata = copy.deepcopy(metadata)
        metadata["database_written"] = True
        metadata["database_uploaded_at"] = _iso_now()
        metadata["created_concept_ids"] = created_ids
        metadata["merged_concept_ids"] = merged_ids
        inventory_audit[RELEASE_METADATA_KEY] = metadata
        if isinstance(certificate_sink.get("certificate"), dict):
            inventory_audit[
                build_concepts._DEPOSITED_GROUNDING_CERTIFICATE_KEY
            ] = copy.deepcopy(certificate_sink["certificate"])
        job.question_inventory = inventory_audit
        job.status = "generated"
        job.deposit_scope_type = "chapter"
        job.deposit_scope_ids = [target_chapter_id]
        job.result_ids = created_ids + merged_ids
        job.detail = (
            f"Explicit release upload created {len(created_ids)} concept(s) "
            f"and merged {len(merged_ids)} existing concept(s)."
        )
        db.commit()
        db.refresh(job)
        progress.log(
            "Explicit Upload to database completed after deterministic release "
            "validation. No generation request was made.",
            level="success",
        )
        return {
            "job_id": job.id,
            "status": "generated",
            "database_written": True,
            "concepts_created": len(created_ids),
            "concepts_merged": len(merged_ids),
            "concept_ids": created_ids + merged_ids,
            "rows_appended": int(written.get("written") or 0),
            "sources_updated": int(written.get("sources_updated") or 0),
            "output_workbook": str(config.BULK_IMPORT_OUTPUT),
            **_release_urls(job.id),
        }


# Ensure service callers using the release router receive the interceptor.
install()
