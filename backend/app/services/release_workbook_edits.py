"""Upload an edited Concept workbook back into the staged release + CMS.

The reviewer downloads the canonical release-bulk-import workbook, edits
it locally in Excel, and uploads the file here (owner steer, 2026-08-20:
"upload the downloaded/edited excel to the CMS" replaces the separate
upload-to-database action). The upload is applied as ONE recorded review
round through ``release_review.apply_manual_edits`` — the same trail,
version rows, and field rules as the review page — and then published to
the CMS through the ONE publication writer. There is deliberately no
second import semantics: the previous path (``/data/import``) silently
NO-OPed content edits on existing concepts and forked a duplicate row on
a title edit, which is exactly the divergence this module closes.

Matching is mechanical and loud:

* primarily by the identity tag the downloaded workbook itself embeds in
  the title cells (rebuilt from the same staged payload through the same
  ``transient_release_hierarchy`` that rendered the download — one
  authority for the tag);
* fallback, for a row whose tag was lost in editing, exact
  (topic, concept title) match against the staged records;
* a workbook row that matches no staged record is REFUSED with its sheet
  and row named — never silently dropped, never invented as a new row
  (additions belong to the review page's instruction rounds).

Only the review page's ``EDITABLE_FIELDS`` travel; anything else in the
workbook (groups, questions, chapter metadata) is the release's own
render and is ignored here.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping

import openpyxl
from sqlalchemy.orm import Session

from .. import models
from ..bulk_import import layouts
from ..bulk_import import strip_title_tag, strip_topic_title
from ..bulk_import.layouts import WorkbookLayoutError
from . import build_concepts_release as bcr
from . import build_concepts_release_files as release_files
from . import identity
from . import generation_recovery
from . import release_review


class WorkbookEditError(ValueError):
    """The uploaded workbook cannot be applied mechanically."""


def _sheet_headers(wb) -> dict[str, tuple]:
    return {
        name: tuple(
            next(wb[name].iter_rows(min_row=2, max_row=2, values_only=True),
                 ()),
        )
        for name in wb.sheetnames
    }


def _record_titles(record: Mapping[str, Any]) -> tuple[str, str]:
    topic = str(record.get("topic") or "").strip()
    title = str(
        record.get("concept_title") or record.get("concept") or ""
    ).strip()
    return topic, title


def _staged_maps(
    db: Session, job: models.UploadJob, payload: Mapping[str, Any]
) -> tuple[dict[str, int], dict[tuple[str, str], list[int]]]:
    """(tag -> record_index, (topic, title) -> [record_index...]).

    The tag map is rebuilt through ``transient_release_hierarchy`` — the
    exact function that rendered the downloaded workbook — so the ids
    here are byte-identical to the tags the reviewer's file carries. The
    transient pass skips defective rows, so its (concept, record) pairs
    are re-anchored to payload record indexes by equality, consumed in
    order (duplicate-safe).
    """

    records = payload.get("records") or []
    by_title: dict[tuple[str, str], list[int]] = {}
    for index, record in enumerate(records):
        if isinstance(record, Mapping):
            by_title.setdefault(_record_titles(record), []).append(index)

    tag_map: dict[str, int] = {}
    try:
        _chapter, concepts, carried, _defects = (
            release_files.transient_release_hierarchy(
                db, job, payload=dict(payload)
            )
        )
    except Exception:
        # The tag join is an assist; title matching still works. A
        # payload the hierarchy refuses outright will be refused again,
        # loudly, by the publication step below.
        return {}, by_title
    consumed: set[int] = set()
    for concept, record in zip(concepts, carried):
        if not isinstance(record, Mapping):
            continue
        key = _record_titles(record)
        index = next(
            (
                candidate for candidate in by_title.get(key, [])
                if candidate not in consumed
                and records[candidate] == record
            ),
            None,
        )
        if index is None:
            index = next(
                (
                    candidate for candidate in by_title.get(key, [])
                    if candidate not in consumed
                ),
                None,
            )
        if index is None:
            continue
        consumed.add(index)
        tag = str(getattr(concept, "machine_id", "") or "").strip()
        if tag and tag not in tag_map:
            tag_map[tag] = index
    return tag_map, by_title


def _workbook_rows(path: Path) -> list[dict[str, str]]:
    """Every content row's topic/concept projection, layout-identified."""

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        identified = layouts.identify_workbook(_sheet_headers(wb))
        rows: list[dict[str, str]] = []
        for found in identified.sheets:
            ws = wb[found.sheet_name]
            for row_i, row in enumerate(
                ws.iter_rows(min_row=3, values_only=True), start=3
            ):
                if row is None or not any(row):
                    continue
                top = found.sheet.block_values(row, "topic")
                con = found.sheet.block_values(row, "concept")
                raw_title = str(con.get("concept_title") or "")
                if not raw_title.strip():
                    continue
                rows.append({
                    "sheet": found.sheet_name,
                    "row": str(row_i),
                    "tag": identity.title_tag(raw_title),
                    "topic": strip_topic_title(
                        str(top.get("topic_title") or "")
                    ).strip(),
                    "concept_title": strip_title_tag(raw_title).strip(),
                    "parent_concept": str(
                        con.get("parent_concept") or ""
                    ).strip(),
                    "concept_details": str(
                        con.get("concept_details") or ""
                    ),
                    "keywords": str(con.get("keywords") or "").strip(),
                })
        return rows
    finally:
        wb.close()


def apply_workbook_and_publish(
    db: Session,
    job: models.UploadJob,
    *,
    lane: str,
    workbook_path: Path,
    owner_sub: str = "",
) -> dict[str, Any]:
    """One recorded round of workbook edits, then the CMS publication."""

    db.refresh(job)
    generation_recovery.require_mutation_allowed(
        job, operation="upload and publish an edited workbook"
    )

    lane = bcr.normalize_lane(lane)
    payload = bcr.release_payload(job, lane=lane)
    if payload is None:
        raise WorkbookEditError(
            "this job has no staged release for this lane; generate and "
            "stage the release before uploading an edited workbook"
        )
    try:
        parsed = _workbook_rows(workbook_path)
    except WorkbookLayoutError as error:
        raise WorkbookEditError(
            "the uploaded file is not a canonical Concept workbook: "
            + str(error)
        ) from error
    if not parsed:
        raise WorkbookEditError(
            "the uploaded workbook carries no concept rows"
        )

    tag_map, by_title = _staged_maps(db, job, payload)
    records = payload.get("records") or []
    edits: list[dict[str, Any]] = []
    unmatched: list[str] = []
    matched_rows = 0
    consumed: set[int] = set()
    for row in parsed:
        index: int | None = None
        if row["tag"] and row["tag"] in tag_map:
            index = tag_map[row["tag"]]
        else:
            index = next(
                (
                    candidate for candidate in by_title.get(
                        (row["topic"], row["concept_title"]), []
                    )
                    if candidate not in consumed
                ),
                None,
            )
        if index is None or not isinstance(records[index], Mapping):
            unmatched.append(
                f"{row['sheet']!r} row {row['row']}: "
                f"{row['concept_title']!r} matches no staged concept"
            )
            continue
        consumed.add(index)
        matched_rows += 1
        record = records[index]
        for field in sorted(release_review.EDITABLE_FIELDS):
            after = str(row.get(field, "")).strip()
            if field == "concept_details":
                after = str(row.get(field, ""))
            current = str(
                record.get(field)
                or (record.get("concept") if field == "concept_title"
                    else "")
                or (record.get("concept_description")
                    if field == "concept_details" else "")
                or ""
            )
            if field in ("topic", "concept_title") and not after:
                # The workbook cell is blank (merged-cell rendering);
                # never treat that as "empty the field".
                continue
            if after == current or (not after and not current.strip()):
                continue
            edits.append({
                "record_index": index,
                "field": field,
                "before": current,
                "after": after,
            })
    if unmatched:
        raise WorkbookEditError(
            "refusing the upload — these workbook rows match no staged "
            "concept (a new concept is added on the review page, not by "
            "workbook row): " + "; ".join(unmatched[:10])
        )

    round_applied = False
    if edits:
        release_review.apply_manual_edits(
            db, job,
            lane=lane,
            staged_release_uid=str(
                payload.get(bcr.STAGED_RELEASE_UID_FIELD) or ""
            ),
            edits=edits,
            owner_sub=owner_sub,
        )
        round_applied = True

    from . import build_concepts_release_publication as publication

    published = publication.upload_release_to_database(
        db, int(job.id), owner_sub=owner_sub or None, lane=lane
    )
    return {
        "lane": lane,
        "workbook_rows": len(parsed),
        "matched_rows": matched_rows,
        "changed_fields": len(edits),
        "round_recorded": round_applied,
        "publication": published,
    }
