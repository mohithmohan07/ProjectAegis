"""Files exported by the unattended Build Concepts release contract."""
from __future__ import annotations

import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import models
from . import concept_run_report
from . import containers
from . import identity
from . import coverage_ledger
from . import uploads
from .build_concepts_release import (
    LANE_POST,
    LANE_PRE,
    PRE_LANE_ASSUMES_NOTHING,
    PRE_LANE_VERDICT_FIELD,
    PRE_ROW_RELATED_CONCEPTS_FIELD,
    assessment_lane_issue,
    RELEASE_ROW_BLOCKS_FIELD,
    RELEASE_ROW_ERRORS_FIELD,
    RELEASE_ROW_QIDS_FIELD,
    RELEASE_ROW_ROUTES_FIELD,
    RELEASE_ROW_STATUS_FIELD,
    normalize_lane,
    release_payload,
    release_state,
    row_projection_defect,
    structural_defects,
)


# The projection's own defect codes (spec-step8 D8.4/S9). Row-level
# findings reuse ``build_concepts_release.STAGED_ROW_UNUSABLE``, because
# the staging pass and this one call the SAME function; these three name
# what is wrong with the release as a whole rather than with one row.
RELEASE_TARGET_CHAPTER_MISSING = "release_target_chapter_missing"
RELEASE_DIRECTORY_METADATA_MISSING = "release_directory_metadata_missing"
STAGED_RECORDS_NOT_AN_ARRAY = "staged_records_not_an_array"


def pre_lane_empty_reason(payload: Mapping[str, Any] | None) -> str:
    """WHY a staged Pre release carries zero concept rows, or ``""``.

    An empty Pre pair still downloads (Rule E / D8.4: a defect blocks the
    database write, never a download) — but [owner report, 2026-08-21] a
    reviewer handed two empty files with the reason living only in the run
    log, the release JSON and the Issues sheet read it as silent loss,
    which is exactly the confusion R4 exists to prevent. This transcribes
    the run's own recorded decision — the fail-closed refusal, or the
    empty-capture verdict — onto the manifest entries the reviewer is
    actually looking at. Nothing is decided here: no record means the
    generic sentence, never a guess at one.
    """

    if payload is None or payload.get("records"):
        return ""
    refused = str(payload.get("refused") or "").strip()
    if refused:
        return (
            "This run REFUSED the Pre-Learning map, so the Pre files "
            "carry no concepts. Recorded reason: " + refused
            + " Fix the named leak and re-run generation to rebuild the "
            "Pre lane; the Post outputs are unaffected."
        )
    verdict_row = payload.get(PRE_LANE_VERDICT_FIELD)
    verdict = (
        str(verdict_row.get("verdict") or "").strip().lower()
        if isinstance(verdict_row, Mapping) else ""
    )
    rationale = (
        str(verdict_row.get("rationale") or "").strip()
        if isinstance(verdict_row, Mapping) else ""
    )
    if verdict == PRE_LANE_ASSUMES_NOTHING:
        return (
            "The run decided this chapter assumes no prior knowledge, so "
            "the Pre files are genuinely empty (recorded verdict "
            f"'{PRE_LANE_ASSUMES_NOTHING}'). "
            + (f"Rationale: {rationale}" if rationale else "")
        ).strip()
    if verdict:
        return (
            "The prerequisite capture was empty and the run decided "
            f"'{verdict}', so the Pre files carry no concepts. "
            + (f"Rationale: {rationale} " if rationale else "")
            + "Re-run generation once the capture issue is resolved."
        ).strip()
    return (
        "The staged Pre-Learning release contains no concept rows. The "
        "release JSON and the diagnostics export carry everything the "
        "run recorded about this lane."
    )


# ---------------------------------------------------------------------- #
# The owner's output numbering (OD4 / register entry D9-Q22), which is the
# order the reviewer's list must read in:
#
#   01 Pre-Learning Concept File   02 Pre-Learning Master File
#   03 Post-Learning Concept File  04 Post-Learning Master File
#
# then the evidence artifacts — the review workbook, the diagnostics zip
# and the release JSON — which are NOT outputs and are never numbered.
#
# No ``kind`` string carries a digit, deliberately (T14): a kind spelling
# an output NUMBER would have to be renamed by the next renumbering
# ruling, and renaming a manifest kind is a frontend break.
#
# This lives here, in the module that owns the eager manifest, and the
# lazy twin in ``build_concepts_release_manifest`` aliases it — the same
# "one owner, no twin" resolution ``_safe_filename`` already got, for the
# same reason: an ORDER that two implementations each define separately
# is an order the monkeypatch in ``install()`` can silently disagree with.
# ---------------------------------------------------------------------- #

OUTPUT_KIND_ORDER: tuple[str, ...] = (
    "pre_release_bulk_import",   # Output 01 · Pre-Learning Concept File
    "pre_release_master",        # Output 02 · Pre-Learning Master File
    "release_bulk_import",       # Output 03 · Post-Learning Concept File
    "release_master",            # Output 04 · Post-Learning Master File
)

_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# The two Master entries' labels, keyed by lane, so the twin manifests
# cannot drift on the text either.
_MASTER_LABELS = {
    LANE_PRE: ("pre_release_master", "Pre-Learning Master File"),
    LANE_POST: ("release_master", "Post-Learning Master File"),
}
# The two reasons that are NOT the generic placeholder, each stating a
# checked fact rather than a guess.
_MASTER_NOT_BUILT = (
    "This lane's Master File has not been built for this run yet. Build it "
    "from the release page; the Concept File for this lane is unaffected."
)
_MASTER_NOT_PUBLISHED = (
    "This lane's Master File exists but its release has not finished "
    "publishing, so nothing is served for it yet."
)
_MASTER_SUPERSEDED = (
    "This lane's Master File was superseded and no newer release has "
    "finished being built. The superseded release keeps its file and its "
    "receipt; the manifest does not point at a superseded release."
)


def master_entry(
    job: models.UploadJob, *, lane: object,
) -> dict[str, Any]:
    """Output 02 or Output 04's manifest row — PRESENT, always.

    The four outputs are enumerated in one place or map P16 survives at
    its root: a reviewer who cannot see four entries cannot check that
    four exist, and an ABSENT entry is indistinguishable from a lane that
    does not apply to this chapter. So this row is never omitted. When the
    lane has no servable Master it is present and ``disabled`` with a
    reason that says which of the three things happened:

    * the lane's Master was not built on this run;
    * it was built but its release has not finished publishing;
    * every release in the lane's chain has been superseded, so there is
      no live row to point at;
    * it is available, and the entry is enabled and points at it.

    In all three disabled cases, when the run RECORDED why
    (``assessment_lane_unavailable``, T15-2) the reason is that record's
    own message — naming the exception — rather than the generic
    sentence.

    Resolved through ``AssessmentRelease.job_id`` plus the ``lane`` column
    (spec-step8 T2/S6) — which is exactly why S3 could only leave a
    placeholder here: there was no column to ask, and the entry had to
    exist anyway so that the ORDER was fixed from that slice on.

    The session comes off the job itself. A manifest block is handed an
    ORM object and no session, and reaching for a new one here would open
    a second connection during a serialization; ``object_session`` is the
    one the caller already has.
    """

    from sqlalchemy.orm import object_session

    from . import release_core

    resolved = normalize_lane(lane)
    kind, label = _MASTER_LABELS[resolved]
    entry: dict[str, Any] = {
        "kind": kind,
        "label": label,
        "filename": "",
        "media_type": _XLSX_MEDIA_TYPE,
        "size_bytes": 0,
        "download_url": "",
        "action": "download",
    }
    # Read ONCE, above both disabled branches. The run that recorded a
    # named ``assessment_lane_unavailable`` failure is this run — staging
    # rebuilds ``payload["issues"]`` from scratch and ``_build_master_
    # siblings`` appends after it — so whenever this is present it is the
    # more specific of two true things, and the reviewer gets the named
    # exception rather than a generic sentence. It used to be read only in
    # the "no row at all" branch, which meant the realistic partial
    # failure — ``create_release`` commits, then ``publish_release``
    # raises — showed the generic reason and threw the recorded one away.
    lane_payload = release_payload(job, lane=resolved)
    recorded = assessment_lane_issue(lane_payload)
    # The Pre pair's zero-row reason (owner report: two empty files with
    # the recorded decision nowhere in sight). Stamped on the ENABLED
    # entry too — the empty Master still downloads (Rule E), it just says
    # why it is empty. Pre lane only: the refused/verdict record it
    # transcribes exists only on Pre payloads.
    empty_note = (
        pre_lane_empty_reason(lane_payload) if resolved == LANE_PRE else ""
    )
    release = None
    superseded = None
    try:
        session = object_session(job)
        release = release_core.latest_release_for_lane(
            session, job.id, resolved)
        if release is None:
            superseded = release_core.release_chain_head(
                session, job.id, resolved)
    except Exception:
        # A manifest block must never be the thing that fails a job
        # serialization. An unreadable release row leaves the entry
        # present and disabled, which is the shape this function
        # guarantees in every other branch too.
        release = None
        superseded = None
    if release is None:
        entry["disabled"] = True
        entry["disabled_reason"] = (
            str((recorded or {}).get("message") or "")
            or (_MASTER_SUPERSEDED if superseded is not None
                else _MASTER_NOT_BUILT)
        )
        return entry
    if not (release.publication or {}).get("manifest"):
        entry["disabled"] = True
        entry["disabled_reason"] = (
            str((recorded or {}).get("message") or "")
            or _MASTER_NOT_PUBLISHED
        )
        return entry
    entry["label"] = f"Download the {label}"
    entry["filename"] = (
        f"aegis_master_{release.release_uid}_v{release.version}.xlsx"
    )
    entry["download_url"] = (
        f"/build-assessments/releases/{release.id}/master.xlsx"
    )
    entry["release_state"] = release_core.release_state(release)
    if empty_note:
        entry["note"] = empty_note
    return entry


def in_owner_order(
    entries: Iterable[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    """Outputs 01-04 first in the owner's numbering, then everything else.

    Pure mechanics — a stable sort on a fixed key. It reads no content and
    decides nothing about meaning; it only fixes the sequence the reviewer
    sees so that the four outputs are enumerated in one place, in one
    order, in both lanes and in both manifest implementations.

    It is applied at every assembly point rather than relying on the
    literal order entries happen to be written in, so a later slice that
    APPENDS ``release_master`` / ``pre_release_master`` at the end of a
    block still lands them at positions 02 and 04. That is the whole
    point: the order is fixed from here on and cannot be quietly
    re-decided by whoever edits a list last.
    """

    rank = {kind: index for index, kind in enumerate(OUTPUT_KIND_ORDER)}
    listed = [dict(entry) for entry in entries]
    outputs = sorted(
        (entry for entry in listed if entry.get("kind") in rank),
        key=lambda entry: rank[entry["kind"]],
    )
    return outputs + [
        entry for entry in listed if entry.get("kind") not in rank
    ]


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_READY_FILL = PatternFill("solid", fgColor="D9EAD3")
_TYPE_FILL = PatternFill("solid", fgColor="D9EAF7")
_CASE_FILL = PatternFill("solid", fgColor="EADCF8")
_EXAMPLE_FILL = PatternFill("solid", fgColor="F3F3F3")
_BLOCK_ID_RE = re.compile(r"\bBLK-[A-Za-z0-9_-]+\b")


ISSUES_NOTE_LABEL = "Release issues"


def _write_issues_note(
    sheet, defects: list[dict[str, Any]], concepts: list[Any],
) -> None:
    """Row 1's note, past the last banded column. See the caller's docstring.

    Written whenever the projection recorded anything OR the release is
    empty, so the reviewer opening a zero-row Concept File is told why it
    is empty instead of being left to guess (R4: the silence is the
    defect). Never written when there is nothing to say, so a healthy
    workbook is byte-identical to the one this repo already ships.

    THE CUT SAYS THAT IT CUT. A cell holds ``CELL_LIMIT`` characters and
    no more; a note longer than that has to lose its tail. It was losing
    it to a bare ``[:32_000]`` — an unnamed number, four digits away from
    the format's real cap, cutting mid-sentence with no marker in a note
    that is the last thing a reviewer reads before opening the release
    JSON. The cap is read from the module that owns it, and what was cut
    is named.
    """
    from ..bulk_import import writer as bi_writer
    from ..bulk_import.assessment_workbook import CELL_LIMIT

    lines = [_cell_text(defect.get("message")) for defect in defects]
    lines = [line for line in lines if line]
    if not concepts:
        lines.append(
            "This release has no concept row to export. That is emptiness, "
            "not corruption: whether it is CORRECT is recorded as the "
            "release state, not guessed here."
        )
    if not lines:
        return
    column = len(bi_writer.FIELDS_BY_KIND["objective"]) + 2
    cell = sheet.cell(row=1, column=column)
    text = _cell_text(f"{ISSUES_NOTE_LABEL}: " + " | ".join(lines))
    if len(text) > CELL_LIMIT:
        mark = (
            f" […{len(lines)} findings were recorded and this note was cut "
            "to fit one spreadsheet cell.]"
        )
        text = text[:CELL_LIMIT - len(mark)] + mark
    cell.value = text
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left")


def build_release_bulk_import_workbook(
    db, job: models.UploadJob, *, lane: object = LANE_POST,
) -> bytes:
    """The released rows in the canonical Bulk Import workbook format.

    After the explicit database upload the released rows exist as real
    concepts, so the canonical by-id writer serves the exact export. Before
    that upload the same canonical row renderer runs over transient
    (never-persisted) chapter/topic/concept copies built from the release
    payload — creators get the Bulk Import file the moment a run completes,
    with the authored chapter/topic metadata applied, and the database stays
    untouched.

    ``lane`` names the staged slot: ``"post"`` projects Output 03, ``"pre"``
    Output 01. The published-concept shortcut reads the ids the lane's OWN
    publication recorded — ``job.result_ids`` belongs to the Post lane, so
    serving the Pre workbook from it would hand a reviewer the Post rows
    under a Pre filename.

    IT ALWAYS RETURNS A WORKBOOK (spec-step8 D8.4). [measured at 76c84fb]
    an empty Pre release and a Post release with one topic-less row both
    404'd this route while the other three downloads returned 200 — one
    bad row took the reviewer's Concept File away. Now the sound rows are
    written and the defects ride an issues note.

    WHERE THAT NOTE GOES, and it is not a free choice. This file is the
    canonical Bulk Import format and it is re-imported: an extra sheet is
    refused outright by ``layouts.identify_workbook`` ("a sheet that no
    layout claims … refuses the WHOLE workbook") unless it is registered
    in the layout's ``ignored_sheets``, which the reference layout leaves
    empty; and an extra DATA row would be read back as a concept by
    ``reader.import_workbook`` and by ``validate_concept_file``'s row
    comparison. So the note is written on ROW 1 — the band row, which no
    reader parses (``_header_names`` reads row 2, data starts at row 3) —
    past the last banded column, where the reference sheet is blank by
    construction. Format-safe, and visible above the frozen pane.

    WHAT THE NOTE DUPLICATES AND WHAT IT DOES NOT, stated exactly rather
    than generously. The ROW-level findings are decided at staging by
    ``build_concepts_release.staged_row_defects``, so they are already in
    ``payload["issues"]``, in the release workbook's Issues sheet, in the
    diagnostics zip and in the release state — this note repeats them.
    The two CHAPTER-level findings below are not: they are discovered
    here, at download time, from a live database read, and D8.5 forbids a
    projection-time discovery from mutating the frozen payload. Their
    consumer for the DATABASE WRITE exists and is not this note —
    ``upload_release_to_database`` re-checks the target chapter itself and
    refuses — but this note is their only place in an ARTEFACT. Recorded
    as a known asymmetry rather than papered over.
    """
    from ..bulk_import import writer as bi_writer
    resolved = normalize_lane(lane)
    payload = release_payload(job, lane=resolved)
    if payload is None:
        raise ValueError("this upload has no staged release")
    summary = payload.get("summary") or {}
    result_ids = [
        int(v)
        for v in (
            summary.get("concept_ids")
            if resolved == LANE_PRE
            else job.result_ids
        ) or []
        if v
    ]
    if bool(summary.get("database_uploaded")) and result_ids:
        return bi_writer.write_concepts_workbook(db, result_ids)

    chapter, concepts, _records, defects = transient_release_hierarchy(
        db, job, payload=payload
    )

    wb = bi_writer._new_workbook()
    ws = wb[bi_writer.SHEET_BY_KIND["objective"]]
    _write_issues_note(ws, defects, concepts)
    export_scope = bi_writer.ConceptExportScope(concepts)
    next_row = 3
    for concept in concepts:
        for topic in sorted(
            bi_writer._concept_placements(concept),
            key=identity.source_order_key,
        ):
            for column, value in enumerate(
                bi_writer._concept_to_row(
                    concept,
                    "objective",
                    topic,
                    export_scope=export_scope,
                ),
                start=1,
            ):
                bi_writer._write_cell(ws, row=next_row, column=column, value=value)
            next_row += 1
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def transient_release_hierarchy(
    db,
    job: models.UploadJob,
    *,
    payload: Mapping[str, Any] | None = None,
) -> tuple[
    models.Chapter,
    list[models.Concept],
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    """Build the one transient hierarchy a lane's Concept and Master share.

    RETURNS ``(chapter, concepts, records, defects)`` — spec-step8 D8.4,
    Rule E enforced for the first time. Seven of this function's eight
    raises are gone: a defect anywhere in the hierarchy used to cost the
    ARTIFACTS, 404-ing ``release-bulk-import.xlsx`` and 400-ing the two
    Master builds, which is precisely what "a defect blocks the database
    write, never a download" forbids. What each of the seven did instead
    is written at its own site below.

    The ONE raise that stays is ``release is None`` — "this upload has no
    staged release". That is genuine impossibility rather than a defect:
    there is no artefact to project, no frozen directory metadata to build
    a chapter row from, and fabricating an empty one would hand a reviewer
    an invented file. D8.4 names it explicitly ("keeps ValueError→404 only
    for 'this upload has no staged release'"), and ``api/build_concepts``
    keeps exactly that arm.

    The ROW-level verdict is not made here. ``build_concepts_release.
    row_projection_defect`` is the one implementation and the staged
    payload already carries its findings under ``staged_row_defects``
    (D8.5: this function runs at download time, so a discovery here could
    reach ``structural_defects`` only by mutating a frozen payload during
    a GET). Calling it here re-reads the same rows through the same
    function, so the skipped row and the reviewer's issue cannot drift.

    LANE-GENERIC, and named that way deliberately. ``payload`` is resolved
    by the caller — ``build_release_bulk_import_workbook:161`` passes the
    PRE payload when asked for the Pre lane — so this builds the Post pair
    (Outputs 03/04) or the Pre pair (Outputs 01/02) depending on what it is
    handed. Naming one pair here would be wrong for half the callers. This
    is T14's own remedy for `assessment_release_run.py:1459`, the one
    string OD4 rules must stop naming a lane rather than be renumbered.

    The staged release records are the concept authority.  The target chapter
    supplies directory metadata, and — since Round 7 — persisted Topic and
    Concept rows are consulted for IDENTITY ONLY (machine ids and the title
    join), never for content: the ids stamped on these transient rows must
    be the ids the publication persists, and [measured] stamping them
    without looking cost every Master question its home the moment the
    chapter held prior state.  Keeping this construction in one place
    prevents the Concept and Master projections from drifting before the
    explicit database upload.
    """

    from . import build_concepts

    release = dict(payload) if isinstance(payload, Mapping) else release_payload(job)
    if release is None:
        # THE ONE SURVIVING RAISE — see the docstring. Genuine
        # impossibility, not a defect: there is nothing staged to project.
        raise ValueError("this upload has no staged release")
    defects: list[dict[str, Any]] = []
    target_chapter_id = int(release.get("target_chapter_id") or 0)
    if db.get(models.Chapter, target_chapter_id) is None:
        # DE-RAISED. The persisted chapter row contributes NOTHING to this
        # projection — every chapter field below is read from the payload's
        # frozen ``directory_metadata`` — so its absence changed no cell and
        # cost all four outputs. Recorded because the publication DOES need
        # that row and will refuse without it.
        defects.append({
            "code": RELEASE_TARGET_CHAPTER_MISSING,
            "message": (
                "the release target chapter no longer exists, so these "
                "files project the chapter frozen into the release rather "
                "than a live one"
            ),
        })
    directory = release.get("directory_metadata")
    if not isinstance(directory, Mapping) or not directory:
        # DE-RAISED. Missing directory metadata blanks the chapter's
        # board/grade/subject cells; it does not stop a single concept row
        # from being written.
        defects.append({
            "code": RELEASE_DIRECTORY_METADATA_MISSING,
            "message": (
                "the staged release has no frozen chapter directory "
                "metadata, so the chapter's directory cells ship blank"
            ),
        })
        directory = {}
    raw_records = release.get("records") or []
    if not isinstance(raw_records, list):
        # DE-RAISED. A records value that is not an array yields no rows,
        # which is what an empty array yields; the difference is that this
        # one is a defect and is named as one.
        defects.append({
            "code": STAGED_RECORDS_NOT_AN_ARRAY,
            "message": (
                "the staged release concept records are not an array, so "
                "no concept row could be read from them"
            ),
        })
        raw_records = []
    records: list[dict[str, Any]] = []
    for position, row in enumerate(raw_records, start=1):
        # DE-RAISED, three at once (row not an object / no topic / no
        # concept title). ONE implementation, shared with staging.
        finding = row_projection_defect(row, position)
        if finding is not None:
            defects.append(dict(finding))
            continue
        records.append(dict(row))
    # THE EIGHTH RAISE IS SIMPLY GONE, with no defect in its place: "the
    # release contains no concept rows to export" was EMPTINESS, and
    # emptiness is ``nothing_to_publish``'s question, not a corruption
    # finding (D8.1). Recording it here would put the very conflation this
    # slice split back into the projection.

    pre_post = (
        "Pre"
        if str(release.get("learning_kind") or "").strip().lower() == "pre"
        else "Post"
    )
    chapter = models.Chapter(
        chapter_code=str(directory.get("chapter_code") or ""),
        board=str(directory.get("board") or ""),
        grade=str(directory.get("grade") or ""),
        subject=str(directory.get("subject") or ""),
        unit=str(directory.get("unit") or ""),
        chapter_title=str(directory.get("chapter_title") or ""),
        chapter_display_name=str(
            directory.get("chapter_display_name") or ""
        ),
        chapter_description="",
        chapter_duration=str(directory.get("chapter_duration") or ""),
        pre_topics="",
        post_topics="",
    )
    chapter.id = -1
    topics_by_key: dict[str, models.Topic] = {}
    concepts: list[models.Concept] = []
    chapter_key = identity.chapter_key(chapter, chapter_id=target_chapter_id)
    # T4-5: ``source_order`` is reconciled to PER-TOPIC here. It was the
    # chapter-wide record index in this export while
    # ``build_concepts_release_publication`` used the per-topic position, so
    # any ``C##`` read off it meant two different things on the two sides of
    # one publication.
    concept_positions: dict[str, int] = {}
    # Round 7: the ids stamped here must BE the ids the publication
    # persists. [measured] purely positional stamping diverged from the
    # publication's minting the moment the chapter held prior state — a
    # prepended topic, a blank-id legacy row — and every Master question
    # built on the stamped id then resolved against the wrong persisted row
    # or against nothing. So the persisted chapter is consulted for
    # IDENTITY ONLY (machine ids and the title join), mirroring the
    # publication's own resolution input for input; the staged records stay
    # the sole CONTENT authority, which is what the docstring's "no
    # persisted row is read" rule always meant to protect.
    from .. import bulk_import as bi_mod
    live_chapter = db.get(models.Chapter, target_chapter_id)
    lane = identity.lane_token(pre_post)
    persisted_topics: dict[str, models.Topic] = {}
    taken_topic_ids: set[str] = set()
    if live_chapter is not None:
        for row in live_chapter.topics:
            if identity.lane_token(row.pre_post_learning) == lane:
                mid = str(row.machine_id or "").strip()
                if mid:
                    taken_topic_ids.add(mid)
            if row.pre_post_learning == pre_post:
                persisted_topics.setdefault(
                    identity.topic_identity(row.topic_title), row)
    taken_concept_ids: dict[str, set[str]] = {}
    claimed_concept_ids: set[int] = set()
    for record in records:
        topic_title = str(record.get("topic") or "").strip()
        key = identity.topic_identity(topic_title)
        topic = topics_by_key.get(key)
        if topic is None:
            topic = models.Topic(
                topic_title=topic_title,
                pre_post_learning=pre_post,
            )
            topic.id = -(len(topics_by_key) + 1)
            topic.source_order = len(topics_by_key) + 1
            topic.chapter = chapter
            topic.chapter_id = chapter.id
            # These rows are NEVER persisted, so ``machine_id_for_topic``
            # refuses to mint on them (T4-7: a transient row must not invent
            # an identity a persisted row will later contradict). They are
            # stamped HERE instead, by the publication's own resolution: a
            # persisted topic's id is reused verbatim (P-C1 — a published
            # topic is never re-keyed), and a genuinely new topic takes the
            # first free slot from this release's position among the
            # persisted lane siblings — so the staged id and the published
            # id are the same string in every deterministic case, not only
            # on a fresh chapter.
            persisted = persisted_topics.get(key)
            persisted_mid = str(
                getattr(persisted, "machine_id", "") or "").strip()
            if persisted_mid:
                topic.machine_id = persisted_mid
            else:
                topic.machine_id = identity.free_slot(
                    lambda n: identity.compose_topic_machine_id(
                        chapter_key, lane, n),
                    topic.source_order,
                    taken_topic_ids,
                )
            taken_topic_ids.add(topic.machine_id)
            seeded = taken_concept_ids.setdefault(topic.machine_id, set())
            if persisted is not None:
                for row in persisted.concepts:
                    mid = str(row.machine_id or "").strip()
                    if mid:
                        seeded.add(mid)
            topics_by_key[key] = topic
        title = str(
            record.get("concept_title") or record.get("concept") or ""
        )
        concept = models.Concept(
            concept_title=title,
            concept_display_name=title,
            parent_concept=str(record.get("parent_concept") or ""),
            concept_details=str(
                record.get("concept_details")
                or record.get("concept_description")
                or ""
            ),
            keywords=str(record.get("keywords") or ""),
            # OWNER RULING OD3: ``keywords`` and ``related_concepts`` ship
            # FILLED. The reference workbook leaving them blank is that
            # school's fill practice, not a rule of the format — the
            # committed format workbook carries both columns and no data
            # rows at all, so the format asserts nothing about them. The
            # ``forced_blank_fields`` profile key stays as the one-line
            # lever if that ever changes.
            #
            # On a PRE row the value is the resolved marker T3.3 stamps at
            # staging (persisted Post ``machine_id``s, newline-joined —
            # concept titles legitimately contain commas). On a POST row
            # nothing authors a concept↔concept relation yet, so it is
            # whatever the record carries, which is blank.
            related_concepts=str(
                record.get(PRE_ROW_RELATED_CONCEPTS_FIELD)
                or record.get("related_concepts")
                or ""
            ),
            digicards=str(record.get("digicards") or ""),
            sources=str(
                release.get("source_book")
                or release.get("filename")
                or ""
            ),
        )
        concept.id = -len(concepts) - 1
        concept_positions[key] = concept_positions.get(key, 0) + 1
        concept.source_order = concept_positions[key]
        # The publication's per-record resolution, predicted with the same
        # inputs (Round 7): a carried ``machine_id`` verbatim; else the
        # exactly-one unclaimed persisted title match under this topic —
        # its persisted id, or, for a blank-id legacy row the publication
        # will adopt, the mint that adoption will produce; else a fresh
        # slot among this topic's taken ids.
        carried = str(record.get("machine_id") or "").strip()
        taken = taken_concept_ids.setdefault(topic.machine_id, set())
        if carried:
            concept.machine_id = carried
            taken.add(carried)
        else:
            match = None
            persisted = persisted_topics.get(key)
            if persisted is not None:
                norm = bi_mod.normalize_question_text(title)
                candidates = [
                    row for row in persisted.concepts
                    if row.id not in claimed_concept_ids
                    and bi_mod.normalize_question_text(row.concept_title)
                    == norm
                ]
                if len(candidates) == 1:
                    match = candidates[0]
            if match is not None:
                claimed_concept_ids.add(match.id)
            match_mid = str(
                getattr(match, "machine_id", "") or "").strip()
            if match_mid:
                concept.machine_id = match_mid
            else:
                concept.machine_id = identity.free_slot(
                    lambda m: identity.compose_concept_machine_id(
                        topic.machine_id, m),
                    concept.source_order,
                    taken,
                )
            taken.add(concept.machine_id)
        concept.topic = topic
        concepts.append(concept)

    build_concepts._sync_chapter_topic_summary(
        chapter,
        dict(release.get("chapter_meta") or {}),
        active_concept_ids=None,
        pre_post=pre_post,
    )
    return chapter, concepts, records, defects


def _json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        indent=2,
        default=str,
    ).encode("utf-8")


def _safe_filename(value: str, fallback: str) -> str:
    name = Path(str(value or "")).stem
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._")
    return name[:100] or fallback


def _cell_text(value: object) -> str:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    return str(value or "")


def _header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _fit_columns(sheet, maximum: int = 70) -> None:
    for column in range(1, sheet.max_column + 1):
        width = 10
        for row in range(1, min(sheet.max_row, 250) + 1):
            width = max(width, min(maximum, len(_cell_text(sheet.cell(row, column).value)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def _provenance_manifest_rows(payload: Mapping[str, Any]) -> dict[str, Any]:
    """Manifest lines that answer "was this chapter actually read by a model?".

    A reviewer holding only this workbook can otherwise not tell a chapter
    with genuinely few questions from one whose outline pass never ran.
    """
    verdict_rows: dict[str, Any] = {}
    verdict = payload.get("pre_lane_verdict")
    if isinstance(verdict, Mapping) and str(verdict.get("verdict") or ""):
        # S9 / D8.3. It belongs HERE and not in the Issues sheet: a
        # positive ``assumes_nothing`` is provenance, not doubt, and
        # counting it as an issue would name a correct empty release
        # "ready with flags". A ``capture_incomplete`` verdict IS an issue
        # and ``_pre_release_issues`` raises it as one; this row states
        # what was decided either way, which is exactly the question this
        # manifest exists to answer for a reviewer holding only the file.
        verdict_rows["Empty prerequisite capture"] = (
            f"{verdict.get('verdict')} — {verdict.get('rationale') or ''}"
        ).strip(" —")
        # Its own second pass, beside it (Q10). The critic advises and
        # never gates, so its dissent belongs where the verdict is read —
        # a reviewer holding only this workbook can otherwise not tell a
        # verdict the second pass agreed with from one it argued against.
        flags = [
            str(flag).strip()
            for flag in verdict.get("review_flags") or []
            if str(flag).strip()
        ]
        if flags:
            verdict_rows["Empty prerequisite capture — review"] = (
                " | ".join(flags)
            )
    provenance = payload.get("extraction_provenance")
    if not isinstance(provenance, Mapping) or not provenance:
        return verdict_rows
    applied = bool(provenance.get("chapter_outline_applied"))
    reader = str(provenance.get("source_reader") or "")
    flags = [str(flag) for flag in provenance.get("chapter_outline_review_flags") or []]
    rows: dict[str, Any] = {
        "Chapter outline": (
            f"applied ({provenance.get('chapter_outline_version') or 'unversioned'})"
            if applied
            else "NOT applied — topics and question boundaries fell back to "
                 "deterministic reading"
        ),
    }
    if reader:
        rows["Source reader"] = reader
    # A zero here is the diagnosis, not an absence — render the counts as text
    # so they survive the blank-for-falsy cell rendering.
    for label, key in (
        ("Outline topics decided", "chapter_outline_topics"),
        ("Outline question partitions", "chapter_outline_partitions"),
        ("Task blocks the outline never ruled on", "chapter_outline_unruled_tasks"),
        ("Source task blocks", "source_task_blocks"),
        ("Questions extracted", "inventory_items"),
        ("Questions from a model-decided split", "model_split_items"),
    ):
        rows[label] = str(int(provenance.get(key) or 0))
    if flags:
        rows["Outline normalization flags"] = "\n".join(flags)
    rows.update(verdict_rows)
    return rows


def build_release_workbook(
    job: models.UploadJob, *, lane: object = LANE_POST,
) -> bytes:
    payload = release_payload(job, lane=normalize_lane(lane))
    if payload is None:
        raise ValueError("this upload has no staged release")

    workbook = Workbook()
    concepts = workbook.active
    concepts.title = "Released Concepts"
    concept_headers = [
        "Release Row",
        "Release Status",
        "Errors / Warnings",
        "Topic",
        "Parent Concept",
        "Concept Title",
        "Concept Details",
        "Keywords",
        "Semantic Topic ID",
        "Source BLKs",
        "QIDs",
        "Type / Case Routes",
    ]
    _header(concepts, concept_headers)
    records = [row for row in payload.get("records") or [] if isinstance(row, Mapping)]
    for index, record in enumerate(records, start=1):
        status = str(record.get(RELEASE_ROW_STATUS_FIELD) or "ready")
        errors = record.get(RELEASE_ROW_ERRORS_FIELD) or []
        concepts.append([
            index,
            status,
            "\n".join(str(value) for value in errors),
            record.get("topic", ""),
            # Parent Concept ships empty, same as the Bulk Import deliverable.
            "",
            record.get("concept_title") or record.get("concept") or "",
            record.get("concept_details") or record.get("concept_description") or "",
            record.get("keywords", ""),
            record.get("_semantic_topic_id", ""),
            ", ".join(str(value) for value in record.get(RELEASE_ROW_BLOCKS_FIELD) or []),
            ", ".join(str(value) for value in record.get(RELEASE_ROW_QIDS_FIELD) or []),
            _cell_text(record.get(RELEASE_ROW_ROUTES_FIELD) or []),
        ])
        fill = _ERROR_FILL if status == "released_with_errors" else _READY_FILL
        for cell in concepts[concepts.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    routes = workbook.create_sheet("Type Case Routing")
    route_headers = [
        "Row Kind",
        "Type ID",
        "Type Title",
        "Type Definition",
        "Case ID",
        "Case Definition",
        "Owner Topic IDs",
        "QIDs",
        "Example QID",
        "Example Prompt",
        "Audit Status",
        "Error",
    ]
    _header(routes, route_headers)
    for row in payload.get("type_case_rows") or []:
        if not isinstance(row, Mapping):
            continue
        kind = str(row.get("row_kind") or "")
        routes.append([
            kind,
            row.get("type_id", ""),
            row.get("type_title", ""),
            row.get("type_definition", ""),
            row.get("case_id", ""),
            row.get("case_definition", ""),
            ", ".join(str(value) for value in row.get("owner_topic_ids") or []),
            ", ".join(str(value) for value in row.get("qids") or []),
            row.get("example_qid", ""),
            row.get("example_prompt", ""),
            row.get("audit_status", "ready"),
            row.get("error", ""),
        ])
        fill = _TYPE_FILL if kind == "type" else _CASE_FILL if kind == "case" else _EXAMPLE_FILL
        for cell in routes[routes.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    issues = workbook.create_sheet("Release Issues")
    issue_headers = [
        "Severity",
        "Code",
        "Phase",
        "Unit ID",
        "Topic",
        "QIDs",
        "BLKs",
        "Message",
        "Full Details",
    ]
    _header(issues, issue_headers)
    for issue in payload.get("issues") or []:
        if not isinstance(issue, Mapping):
            continue
        severity = str(issue.get("severity") or "error")
        issues.append([
            severity,
            issue.get("code", ""),
            issue.get("phase", ""),
            issue.get("unit_id", ""),
            issue.get("topic", ""),
            ", ".join(str(value) for value in issue.get("qids") or []),
            ", ".join(str(value) for value in issue.get("block_ids") or []),
            issue.get("message", ""),
            _cell_text(issue.get("details")),
        ])
        fill = _ERROR_FILL if severity == "error" else _WARNING_FILL
        for cell in issues[issues.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    refinements = payload.get("refinements")
    if isinstance(refinements, Mapping):
        # The Refiner's recorded diff (docs/aegis-restructure.md §8.3):
        # every refinement beside its row identity, plus the advisory and
        # availability flags. Present only when the release traversed the
        # Refiner seam.
        refined = workbook.create_sheet("Refinements")
        _header(refined, [
            "Kind",
            "Concept Key",
            "Field",
            "Before",
            "After",
            "Reason / Flag",
        ])
        for change in refinements.get("changes") or []:
            if not isinstance(change, Mapping):
                continue
            refined.append([
                "refinement",
                _cell_text(change.get("concept_key")),
                change.get("field", ""),
                change.get("before", ""),
                change.get("after", ""),
                change.get("reason", ""),
            ])
            for cell in refined[refined.max_row]:
                cell.fill = _READY_FILL
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for flag in refinements.get("review_flags") or []:
            refined.append(["flag", "", "", "", "", str(flag)])
            for cell in refined[refined.max_row]:
                cell.fill = _WARNING_FILL
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    manifest = workbook.create_sheet("Release Manifest")
    _header(manifest, ["Field", "Value"])
    manifest_rows = {
        "Release version": payload.get("version"),
        "Released at": payload.get("released_at"),
        "Release reason": payload.get("release_reason"),
        "Job ID": payload.get("job_id"),
        "Learning kind": payload.get("learning_kind"),
        "Source file": payload.get("filename"),
        "Source book": payload.get("source_book"),
        "Target chapter ID": payload.get("target_chapter_id"),
        "Checkpoint stage": payload.get("checkpoint_stage"),
        "Checkpoint progress": payload.get("checkpoint_progress"),
        **_provenance_manifest_rows(payload),
        **(
            {"Refinements": str(refinements.get("summary") or "")}
            if isinstance(refinements, Mapping)
            else {}
        ),
        "Summary": payload.get("summary"),
        "Database publication": (
            "uploaded" if (payload.get("summary") or {}).get("database_uploaded")
            else "not uploaded; use the separate Upload to Database action"
        ),
    }
    for key, value in manifest_rows.items():
        manifest.append([key, _cell_text(value)])
        manifest.cell(manifest.max_row, 1).font = Font(bold=True)
        manifest.cell(manifest.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        _fit_columns(sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def release_payload_bytes(
    job: models.UploadJob, *, lane: object = LANE_POST,
) -> bytes:
    payload = release_payload(job, lane=normalize_lane(lane))
    if payload is None:
        raise ValueError("this upload has no staged release")
    return _json_bytes(payload)


def _iter_blocks(value: object) -> Iterable[dict[str, Any]]:
    if isinstance(value, Mapping):
        block_id = str(value.get("block_id") or "")
        if block_id:
            yield dict(value)
        for raw in value.values():
            yield from _iter_blocks(raw)
    elif isinstance(value, (list, tuple)):
        for raw in value:
            yield from _iter_blocks(raw)


def _add_block_reference(
    blocks: dict[str, dict[str, Any]],
    block_id: object,
    *,
    reference: str,
) -> None:
    value = str(block_id or "").strip()
    if not value:
        return
    current = blocks.setdefault(value, {
        "block_id": value,
        "record_status": "referenced_id_only",
        "references": [],
    })
    references = current.setdefault("references", [])
    if isinstance(references, list) and reference not in references:
        references.append(reference)


def _index_string_block_references(
    blocks: dict[str, dict[str, Any]],
    payload: Mapping[str, Any],
) -> None:
    for index, record in enumerate(payload.get("records") or [], start=1):
        if not isinstance(record, Mapping):
            continue
        for block_id in record.get(RELEASE_ROW_BLOCKS_FIELD) or []:
            _add_block_reference(
                blocks,
                block_id,
                reference=f"released_record:{index}",
            )
    for index, issue in enumerate(payload.get("issues") or [], start=1):
        if not isinstance(issue, Mapping):
            continue
        for block_id in issue.get("block_ids") or []:
            _add_block_reference(
                blocks,
                block_id,
                reference=f"release_issue:{index}",
            )
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    for block_id in sorted(set(_BLOCK_ID_RE.findall(raw))):
        _add_block_reference(
            blocks,
            block_id,
            reference="release_payload_text",
        )


def _artifact_directory(job: models.UploadJob) -> Path | None:
    helper = getattr(uploads, "source_artifact_directory", None)
    if not callable(helper) or not getattr(job, "id", None):
        return None
    try:
        # The helper is keyed by job id, not by the job row.  Passing the row
        # raised inside this guard, so every export silently shipped without
        # the canonical artifacts the README promises -- including the Phase
        # 3.1/3.2 caches that explain a failure repeating across resumes.
        path = Path(helper(int(job.id))).resolve()
    except Exception:
        return None
    return path if path.is_dir() else None


def _original_source(job: models.UploadJob) -> Path | None:
    try:
        path = Path(uploads.upload_file_path(job)).resolve()
    except Exception:
        return None
    return path if path.is_file() else None


def build_diagnostics_zip(
    job: models.UploadJob, *, lane: object = LANE_POST,
) -> bytes:
    resolved = normalize_lane(lane)
    payload = release_payload(job, lane=resolved)
    if payload is None:
        raise ValueError("this upload has no staged release")
    release_workbook = build_release_workbook(job, lane=resolved)
    source_manifest = (
        uploads.source_artifact_manifest(job)
        if callable(getattr(uploads, "source_artifact_manifest", None))
        else {}
    )
    blocks: dict[str, dict[str, Any]] = {}
    for source in (
        payload,
        job.generation_checkpoint or {},
        job.question_inventory or {},
        source_manifest,
    ):
        for block in _iter_blocks(source):
            block_id = str(block.get("block_id") or "")
            if block_id:
                blocks.setdefault(block_id, block)
    _index_string_block_references(blocks, payload)

    # The persisted container projections and Phase 2.2 placement snapshot
    # (when the artifact directory holds them) let the ledger account every
    # canonical figure block and list both lanes' verbatim dropped
    # furniture uncapped; jobs from before these artifacts existed keep the
    # capped job-state accounting.
    artifact_dir = _artifact_directory(job)
    projections = containers.load_containers(artifact_dir)

    def _artifact_snapshot(name: str) -> dict[str, Any] | None:
        if artifact_dir is None:
            return None
        path = artifact_dir / name
        if not path.is_file():
            return None
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            return None
        return loaded if isinstance(loaded, dict) else None

    place_snapshot = _artifact_snapshot("source.phase3-place.json")
    # Q1: the chapter analysis inventory's recorded build + allotments
    # (phase3/analyse.py) — lets the ledger account every LA-item.
    analysis_snapshot = _artifact_snapshot("source.phase3-analysis.json")
    # Phase 03 (doc §4): the Pre-Learning map this run built and the
    # merged prerequisite capture it was built from. They let the ledger
    # account the PRE lane's own obligations — every prerequisite carried
    # into the map exactly once, every Pre inventory item allotted
    # exactly once, every needed-for link resolved — and let the run
    # report say what the Pre lane produced.
    pre_map_snapshot = _artifact_snapshot("source.phase3-prelearn-map.json")
    prelearn_snapshot = _artifact_snapshot(
        "source.phase3-prelearn-capture.json"
    )
    # Outputs 01/02 as SHIPPED: the coverage plan and generated questions
    # the run authored, and the staged Pre release that carries them. With
    # these the reviewer sees what the Pre lane produced AND what actually
    # shipped, in RUN_REPORT.txt, without opening any artifact JSON.
    pre_questions_snapshot = _artifact_snapshot(
        "source.phase3-prelearn-questions.json"
    )
    pre_release = release_payload(job, lane=LANE_PRE)

    coverage = coverage_ledger.build_coverage_ledger(
        question_inventory=job.question_inventory or {},
        records=[
            row for row in payload.get("records") or []
            if isinstance(row, Mapping)
        ],
        chapter_reading=(job.question_inventory or {}).get("chapter_reading"),
        container_projections=projections,
        place_snapshot=place_snapshot,
        analysis_snapshot=analysis_snapshot,
        pre_map_snapshot=pre_map_snapshot,
        prelearn_snapshot=prelearn_snapshot,
        pre_questions_snapshot=pre_questions_snapshot,
        pre_release_payload=pre_release,
    )
    run_report = concept_run_report.build_run_report(
        payload,
        generation_log=job.generation_log or [],
        generation_checkpoint=job.generation_checkpoint or {},
        dropped_furniture=coverage.get("dropped_furniture"),
        pre_map=pre_map_snapshot,
        coverage=coverage,
        release_lane=resolved,
    )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "README.txt",
            (
                "Project Aegis diagnostic context export\n\n"
                "Start with RUN_REPORT.txt: it states what stopped the run, "
                "why the orchestration boundary did or did not recover from "
                "it, which rows the failure implicates, and whether the same "
                "failure repeated across resumes. context/run_report.json "
                "carries the same facts with exact log indexes.\n\n"
                "Its COVERAGE section accounts for every question, hub, "
                "figure and question fragment (placed or unaccounted, with "
                "review flags) and names every released row missing its "
                "Achieving Mastery line or Misconception/ Error Analysis "
                "section. context/coverage_ledger.json carries the full "
                "per-item accounting.\n\n"
                "The rest of the archive keeps the released workbook beside "
                "the complete saved generation log, checkpoint, source "
                "evidence, BLK index, Question/Task Inventory, Type/Case "
                "routing, original upload, and canonical-source artifacts. "
                "The release is not proof that every highlighted row is "
                "error-free. See Release Issues.\n"
            ),
        )
        archive.writestr(
            "RUN_REPORT.txt",
            (
                concept_run_report.render_run_report(run_report)
                + coverage_ledger.render_coverage(coverage)
            ).encode("utf-8"),
        )
        archive.writestr("context/run_report.json", _json_bytes(run_report))
        archive.writestr("context/coverage_ledger.json", _json_bytes(coverage))
        if pre_map_snapshot is not None:
            archive.writestr(
                "context/pre_learning_map.json",
                _json_bytes(pre_map_snapshot),
            )
        if prelearn_snapshot is not None:
            archive.writestr(
                "context/pre_learning_capture.json",
                _json_bytes(prelearn_snapshot),
            )
        if pre_questions_snapshot is not None:
            archive.writestr(
                "context/pre_learning_questions.json",
                _json_bytes(pre_questions_snapshot),
            )
        if pre_release is not None:
            archive.writestr(
                "release/pre_release_payload.json",
                _json_bytes(pre_release),
            )
        archive.writestr("release/released_concepts.xlsx", release_workbook)
        archive.writestr("release/release_payload.json", _json_bytes(payload))
        archive.writestr(
            "context/generation_log.json",
            _json_bytes(job.generation_log or []),
        )
        archive.writestr(
            "context/generation_checkpoint.json",
            _json_bytes(job.generation_checkpoint or {}),
        )
        archive.writestr(
            "context/question_inventory.json",
            _json_bytes(job.question_inventory or {}),
        )
        archive.writestr(
            "context/source_artifact_manifest.json",
            _json_bytes(source_manifest),
        )
        archive.writestr("context/blks.json", _json_bytes({
            "count": len(blocks),
            "blocks": [blocks[key] for key in sorted(blocks)],
        }))
        archive.writestr(
            "context/source_evidence.json",
            _json_bytes({
                "pending_decision": payload.get("pending_decision_snapshot") or {},
                "issues": payload.get("issues") or [],
                "type_case_rows": payload.get("type_case_rows") or [],
            }),
        )
        archive.writestr(
            "source/converted_source.mmd",
            str(job.mmd_text or "").encode("utf-8"),
        )

        original = _original_source(job)
        if original is not None:
            archive.write(original, f"source/original/{original.name}")
        directory = _artifact_directory(job)
        if directory is not None:
            for path in sorted(directory.rglob("*")):
                if path.is_file():
                    archive.write(
                        path,
                        "source/canonical_artifacts/"
                        + str(path.relative_to(directory)).replace("\\", "/"),
                    )
    return buffer.getvalue()


def _pre_release_entries(
    job: models.UploadJob,
    *,
    sizes: bool,
) -> list[dict[str, Any]]:
    """Outputs 01/02's manifest rows, or none when the run built no Pre lane.

    (Numbered under the owner's ruling OD4 / register entry D9-Q22: the
    Pre lane is 01/02 and the Post lane 03/04. Earlier text in this repo
    called the same pair "Outputs 03/04"; it is superseded, not wrong.)

    NOTE FOR ANY LATER EDITOR: this function has a TWIN in
    ``build_concepts_release_manifest.py``, whose ``install()`` rebinds
    ``release_artifact_entries`` in THIS module for the whole process.
    Production therefore always serves the twin, and this eager version
    is reached only through ``eager_release_artifact_entries`` (defined
    below the function it aliases) — including by the pin in
    ``tests/test_pre_release_lane_wiring.py``. Editing one alone is a
    silent no-op; the pin exists so that fails loudly.
    """

    payload = release_payload(job, lane=LANE_PRE)
    if payload is None:
        return []
    stem = _safe_filename(job.filename, "concepts")
    query = f"?lane={LANE_PRE}"
    uploaded = bool((payload.get("summary") or {}).get("database_uploaded"))
    empty_note = pre_lane_empty_reason(payload)
    return in_owner_order([
        {
            # Output 01 · Pre-Learning Concept File.
            #
            # ``size_bytes`` is 0 even in this eager block: the builder
            # needs a database session and a manifest block has none.
            # Advertising the entry with an unmeasured size is what the
            # reviewer needs; a size is not.
            "kind": "pre_release_bulk_import",
            "label": "Download the Pre-Learning Concept File",
            "filename": f"{stem}_pre_bulk_import.xlsx",
            "media_type": _XLSX_MEDIA_TYPE,
            "size_bytes": 0,
            "download_url": (
                f"/build-concepts/uploads/{job.id}"
                f"/release-bulk-import.xlsx{query}"
            ),
            "action": "download",
            # The zero-row reason (the run's own recorded refusal or
            # verdict). The download stays enabled — Rule E — but a
            # reviewer sees WHY the file has no rows before opening it.
            **({"note": empty_note} if empty_note else {}),
        },
        # Output 02 · Pre-Learning Master File. Present and, when the lane
        # has no servable Master, disabled with a stated reason — never
        # absent, because an absent entry is the defect this manifest
        # exists to close and a reviewer who cannot see four outputs in
        # one place cannot check that four exist.
        master_entry(job, lane=LANE_PRE),
        {
            "kind": "released_pre_concepts",
            "label": "Download released Pre-Learning output",
            "filename": f"{stem}_pre_released.xlsx",
            "media_type": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            "size_bytes": (
                len(build_release_workbook(job, lane=LANE_PRE)) if sizes else 0
            ),
            "download_url": (
                f"/build-concepts/uploads/{job.id}/release.xlsx{query}"
            ),
            "action": "download",
        },
        {
            "kind": "pre_release_diagnostics",
            "label": "Export full Pre-Learning issue context",
            "filename": f"{stem}_pre_diagnostics.zip",
            "media_type": "application/zip",
            "size_bytes": (
                len(build_diagnostics_zip(job, lane=LANE_PRE)) if sizes else 0
            ),
            "download_url": (
                f"/build-concepts/uploads/{job.id}/diagnostics.zip{query}"
            ),
            "action": "download",
        },
        {
            "kind": "pre_release_payload",
            "label": "Download Pre-Learning release JSON",
            "filename": f"{stem}_pre_release.json",
            "media_type": "application/json",
            "size_bytes": (
                len(release_payload_bytes(job, lane=LANE_PRE)) if sizes else 0
            ),
            "download_url": (
                f"/build-concepts/uploads/{job.id}/release.json{query}"
            ),
            "action": "download",
        },
        {
            "kind": "pre_database_upload",
            "label": (
                "Already uploaded to database"
                if uploaded
                else "Upload released Pre-Learning output to database"
            ),
            "filename": "",
            "media_type": "application/json",
            "size_bytes": 0,
            "download_url": (
                f"/build-concepts/uploads/{job.id}/upload-release{query}"
            ),
            "action": "post",
            "disabled": uploaded,
            "requires_confirmation": True,
            # The lane's release state, and — when that state is
            # ``diagnostic_release`` — what made it one. Computed and
            # surfaced by no endpoint before this, so a reviewer facing a
            # publication the gate will refuse saw a live button and no
            # reason. Downloads are never blocked by either (Rule E).
            "release_state": release_state(payload),
            "structural_defects": structural_defects(payload),
        },
    ])


def release_artifact_entries(job: models.UploadJob) -> list[dict[str, Any]]:
    payload = release_payload(job)
    if payload is None:
        return []
    workbook = build_release_workbook(job)
    diagnostics = build_diagnostics_zip(job)
    raw_payload = release_payload_bytes(job)
    stem = _safe_filename(job.filename, "concepts")
    return in_owner_order([
        {
            # Output 03 · Post-Learning Concept File. Same builder as
            # Output 01, same route, no ``lane`` parameter — so no
            # already-published URL moves.
            "kind": "release_bulk_import",
            "label": "Download the Post-Learning Concept File",
            "filename": f"{stem}_bulk_import.xlsx",
            "media_type": _XLSX_MEDIA_TYPE,
            "size_bytes": 0,
            "download_url": (
                f"/build-concepts/uploads/{job.id}"
                f"/release-bulk-import.xlsx"
            ),
            "action": "download",
        },
        # Output 04 · Post-Learning Master File, resolved through this
        # job's Post ``AssessmentRelease`` row. Present in every branch.
        master_entry(job, lane=LANE_POST),
        {
            "kind": "released_concepts",
            "label": "Download released output",
            "filename": f"{_safe_filename(job.filename, 'concepts')}_released.xlsx",
            "media_type": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            "size_bytes": len(workbook),
            "download_url": f"/build-concepts/uploads/{job.id}/release.xlsx",
            "action": "download",
        },
        {
            "kind": "release_diagnostics",
            "label": "Export full issue context",
            "filename": f"{_safe_filename(job.filename, 'concepts')}_diagnostics.zip",
            "media_type": "application/zip",
            "size_bytes": len(diagnostics),
            "download_url": f"/build-concepts/uploads/{job.id}/diagnostics.zip",
            "action": "download",
        },
        {
            "kind": "release_payload",
            "label": "Download release JSON",
            "filename": f"{_safe_filename(job.filename, 'concepts')}_release.json",
            "media_type": "application/json",
            "size_bytes": len(raw_payload),
            "download_url": f"/build-concepts/uploads/{job.id}/release.json",
            "action": "download",
        },
        {
            "kind": "database_upload",
            "label": (
                "Already uploaded to database"
                if (payload.get("summary") or {}).get("database_uploaded")
                else "Upload released output to database"
            ),
            "filename": "",
            "media_type": "application/json",
            "size_bytes": 0,
            # The publish URL states its lane, and the four DOWNLOAD URLs
            # above do not. That asymmetry is the point: the publish
            # endpoint refuses a blank or absent lane (a defaulted
            # publication is an authenticated write nobody named), so a
            # lane-less publish URL here would be an affordance the
            # server's own gate rejects.
            "download_url": (
                f"/build-concepts/uploads/{job.id}"
                f"/upload-release?lane={LANE_POST}"
            ),
            "action": "post",
            "disabled": bool(
                (payload.get("summary") or {}).get("database_uploaded")
            ),
            "requires_confirmation": True,
            "release_state": release_state(payload),
            "structural_defects": structural_defects(payload),
        },
    ] + _pre_release_entries(job, sizes=True))


# The stable handle on the EAGER implementation.
#
# ``build_concepts_release_manifest.install()`` rebinds the module
# attribute ``release_artifact_entries`` above to its own lazy twin, for
# the whole process — ``app.main.bootstrap()`` does it at startup, and a
# test that calls ``install()`` does it for the rest of the session. So
# after install, ``release_files.release_artifact_entries`` IS the lazy
# one, and a pin that reaches the two implementations through their
# module attributes ends up comparing one function with itself and
# passes no matter how far the two have drifted.
#
# This name is never rebound. The twin pin in
# tests/test_pre_release_lane_wiring.py goes through it, which is what
# makes "an entry added to only one of them fails loudly" true rather
# than aspirational.
eager_release_artifact_entries = release_artifact_entries
