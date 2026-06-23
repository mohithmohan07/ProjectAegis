"""Import unit/chapter syllabus structure from Excel workbooks.

Loads chapter shells (no concepts or questions) so users can deposit concept
mapping into pre-defined units and chapters. English Language is replicated
across all boards because it is universal.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import config, models
from .. import bulk_import as bi
from . import directory
from .text_normalize import (
    normalize_board,
    normalize_chapter,
    normalize_grade,
    normalize_subject,
    normalize_unit,
)

# Boards that receive the shared English Language syllabus.
ALL_SYLLABUS_BOARDS = list(bi.BOARDS)

# Known syllabus files (place under ``data/syllabus/``).
SYLLABUS_FILES = {
    "english_language": "English Language Units and Chapters.xlsx",
    "karnataka": "Kstate Syllabus Grade 6-10.xlsx",
    "maharashtra": "Maharashtra Board Chapter List.xlsx",
    "cbse": "Unit-Chapter List_ CBSE.xlsx",
    "icse": "Unit-Chapter List_ ICSE.xlsx",
}

# Column header aliases (lowercase).
_COL_ALIASES: dict[str, tuple[str, ...]] = {
    "board": ("board", "syllabus", "examination board", "exam board"),
    "grade": ("grade", "class", "standard", "std", "class/grade"),
    "subject": ("subject", "subjects", "subject name"),
    "unit": ("unit", "units", "unit name", "unit title", "unit no", "unit number"),
    "chapter": (
        "chapter", "chapters", "chapter name", "chapter title",
        "lesson", "topic name", "topic title",
    ),
}


@dataclass(frozen=True, slots=True)
class SyllabusRow:
    board: str
    grade: str
    subject: str
    unit: str
    chapter: str


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9/ ]+", "", (value or "").lower()).strip()


def _match_columns(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        norm = _norm_header(header)
        if not norm:
            continue
        for field, aliases in _COL_ALIASES.items():
            if field in mapping:
                continue
            if norm in aliases or any(a in norm for a in aliases):
                mapping[field] = idx
    return mapping


def _parse_sheet_name(name: str) -> dict[str, str]:
    """Extract grade/subject hints from a worksheet tab name."""
    hints: dict[str, str] = {}
    text = name or ""
    grade = normalize_grade(text)
    if grade:
        hints["grade"] = grade
    # "Grade 8 Mathematics", "Class 10 - Science", "Mathematics G09"
    for token in re.split(r"[-_/]", text):
        token = token.strip()
        g = normalize_grade(token)
        if g and "grade" not in hints:
            hints["grade"] = g
        subj = normalize_subject(token)
        if subj and subj.lower() not in {"grade", "class", "standard"}:
            # Avoid treating "Grade" as a subject.
            if not re.fullmatch(r"(?i)(grade|class|standard|unit|chapter).*", token):
                hints.setdefault("subject", subj)
    return hints


def _forward_fill(values: list[str]) -> list[str]:
    last = ""
    out: list[str] = []
    for v in values:
        if v:
            last = v
        out.append(last)
    return out


def _detect_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    best_idx = -1
    best_map: dict[str, int] = {}
    best_score = 0
    for i, row in enumerate(rows[:8]):
        headers = [_cell_str(c) for c in row]
        mapping = _match_columns(headers)
        score = len(mapping)
        if "chapter" in mapping:
            score += 2
        if score > best_score:
            best_score = score
            best_idx = i
            best_map = mapping
    if best_idx < 0 or "chapter" not in best_map:
        return None
    return best_idx, best_map


def _rows_from_sheet(ws, *, default_board: str = "", default_subject: str = "",
                     default_grade: str = "") -> list[SyllabusRow]:
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    sheet_hints = _parse_sheet_name(ws.title)
    grade_default = default_grade or sheet_hints.get("grade", "")
    subject_default = default_subject or sheet_hints.get("subject", "")

    detected = _detect_header_row(raw_rows)
    if detected:
        header_idx, colmap = detected
        data_rows = raw_rows[header_idx + 1:]
    else:
        # Fallback: first four columns as grade, subject, unit, chapter.
        colmap = {"grade": 0, "subject": 1, "unit": 2, "chapter": 3}
        data_rows = raw_rows

    # Collect column values for forward-fill on sparse unit/subject/grade columns.
    grade_col = [_cell_str(r[colmap["grade"]]) if colmap.get("grade") is not None and len(r) > colmap["grade"] else ""
                 for r in data_rows]
    subject_col = [_cell_str(r[colmap["subject"]]) if colmap.get("subject") is not None and len(r) > colmap["subject"] else ""
                   for r in data_rows]
    unit_col = [_cell_str(r[colmap["unit"]]) if colmap.get("unit") is not None and len(r) > colmap["unit"] else ""
                for r in data_rows]

    grade_col = _forward_fill(grade_col)
    subject_col = _forward_fill(subject_col)
    unit_col = _forward_fill(unit_col)

    out: list[SyllabusRow] = []
    for i, row in enumerate(data_rows):
        chapter_idx = colmap.get("chapter")
        if chapter_idx is None or len(row) <= chapter_idx:
            continue
        chapter = normalize_chapter(_cell_str(row[chapter_idx]))
        if not chapter:
            continue
        # Skip header-like repeats and section labels.
        if chapter.lower() in {"chapter", "chapters", "unit", "subject", "grade", "class"}:
            continue

        board_raw = (
            _cell_str(row[colmap["board"]])
            if colmap.get("board") is not None and len(row) > colmap["board"]
            else default_board
        )
        board = normalize_board(board_raw, filename="")
        if not board and default_board:
            board = default_board

        grade = normalize_grade(grade_col[i]) or grade_default
        subject = normalize_subject(subject_col[i]) or subject_default
        unit = normalize_unit(unit_col[i]) if unit_col[i] else ""

        if not grade or not subject:
            continue
        if not unit:
            unit = "General"

        out.append(SyllabusRow(
            board=board, grade=grade, subject=subject, unit=unit, chapter=chapter,
        ))
    return out


def parse_workbook(
    path: Path,
    *,
    default_board: str = "",
    default_subject: str = "",
    universal_boards: list[str] | None = None,
) -> list[SyllabusRow]:
    """Parse one syllabus Excel file into normalized rows."""
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[SyllabusRow] = []
    boards = universal_boards or ([default_board] if default_board else [])

    for ws in wb.worksheets:
        sheet_rows = _rows_from_sheet(
            ws,
            default_board=default_board,
            default_subject=default_subject,
        )
        if universal_boards:
            expanded: list[SyllabusRow] = []
            for b in boards:
                for r in sheet_rows:
                    expanded.append(SyllabusRow(
                        board=b, grade=r.grade, subject=r.subject,
                        unit=r.unit, chapter=r.chapter,
                    ))
            rows.extend(expanded)
        else:
            rows.extend(sheet_rows)
    wb.close()
    return rows


def _chapter_key(row: SyllabusRow) -> tuple[str, str, str, str, str]:
    return (row.board, row.grade, row.subject, row.unit, row.chapter.lower())


def upsert_chapters(db: Session, rows: list[SyllabusRow]) -> dict[str, int]:
    """Insert chapter shells; skip duplicates already in the DB."""
    created = 0
    skipped = 0
    seen: set[tuple[str, str, str, str, str]] = set()

    existing_codes = {
        c.chapter_code
        for c in db.query(models.Chapter.chapter_code).all()
    }

    for row in rows:
        dedupe = _chapter_key(row)
        if dedupe in seen:
            skipped += 1
            continue
        seen.add(dedupe)

        code = directory.make_chapter_code(
            row.board, row.grade, row.subject, row.chapter,
        )
        if code in existing_codes:
            skipped += 1
            continue

        ch = models.Chapter(
            chapter_code=code,
            board=row.board,
            grade=row.grade,
            subject=row.subject,
            unit=row.unit,
            chapter_title=row.chapter,
            chapter_display_name=f"{row.chapter} ({code})",
        )
        db.add(ch)
        existing_codes.add(code)
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "total_rows": len(rows)}


def _resolve_file(key: str) -> Path | None:
    name = SYLLABUS_FILES.get(key, "")
    if not name:
        return None
    path = config.SYLLABUS_DIR / name
    return path if path.exists() else None


def load_all_syllabus_files(db: Session) -> dict:
    """Load every known syllabus workbook from ``data/syllabus/``."""
    all_rows: list[SyllabusRow] = []
    loaded: list[str] = []
    missing: list[str] = []

    # Board-specific files.
    board_keys = {
        "cbse": "CBSE",
        "icse": "ICSE",
        "maharashtra": "Maharashtra",
        "karnataka": "Karnataka",
    }
    for key, board in board_keys.items():
        path = _resolve_file(key)
        if not path:
            missing.append(SYLLABUS_FILES[key])
            continue
        all_rows.extend(parse_workbook(path, default_board=board))
        loaded.append(path.name)

    # English Language — universal across all boards.
    eng_path = _resolve_file("english_language")
    if eng_path:
        all_rows.extend(parse_workbook(
            eng_path,
            default_subject="English Language",
            universal_boards=ALL_SYLLABUS_BOARDS,
        ))
        loaded.append(eng_path.name)
    else:
        missing.append(SYLLABUS_FILES["english_language"])

    counts = upsert_chapters(db, all_rows) if all_rows else {
        "created": 0, "skipped": 0, "total_rows": 0,
    }
    return {
        "loaded_files": loaded,
        "missing_files": missing,
        **counts,
    }


def bootstrap_syllabus(db: Session) -> dict | None:
    """Load syllabus structure when the database has no chapters yet."""
    if db.query(models.Chapter).count() > 0:
        return None
    if not config.SYLLABUS_DIR.is_dir():
        return None
    if not any(config.SYLLABUS_DIR.glob("*.xlsx")):
        return None
    return load_all_syllabus_files(db)
